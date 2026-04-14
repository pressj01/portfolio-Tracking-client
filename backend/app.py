import os
import re
import sys
import time

# Ensure backend directory is on the Python path so sibling imports work
# regardless of the working directory (e.g. when launched from project root).
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import pandas as pd
from flask import Flask, request, jsonify, session, send_file
from flask_cors import CORS
from config import get_connection, FRED_API_KEY
from database import ensure_tables_exist
from import_data import (
    import_from_excel, import_from_upload,
    import_weekly_payouts, import_monthly_payouts,
    import_monthly_payout_tickers,
    import_multi_excel, import_multi_upload,
)
from normalize import (
    populate_holdings,
    populate_dividends,
    populate_income_tracking,
    populate_pillar_weights,
)
from transaction_import import PARSERS as TXN_PARSERS

app = Flask(__name__)
app.secret_key = "portfolio-tracking-client-secret-key"
CORS(app)


@app.errorhandler(500)
def handle_500(e):
    return jsonify({"error": "Internal server error", "detail": str(e)}), 500


@app.errorhandler(404)
def handle_404(e):
    return jsonify({"error": "Not found"}), 404

UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), "uploads")
os.makedirs(UPLOAD_FOLDER, exist_ok=True)


# ── Helpers ────────────────────────────────────────────────────────────────────

def get_profile_id():
    """Return a single profile_id (for write operations)."""
    return int(request.args.get("profile_id", session.get("profile_id", 1)))


def get_profile_filter():
    """Return (is_aggregate, profile_ids_list) for read operations.

    If ?aggregate=true, reads member IDs from aggregate_config.
    Otherwise returns (False, [single_profile_id]).
    """
    if request.args.get("aggregate") == "true":
        conn = get_connection()
        rows = conn.execute("SELECT member_profile_id FROM aggregate_config").fetchall()
        conn.close()
        ids = [r["member_profile_id"] if isinstance(r, dict) else r[0] for r in rows]
        return True, ids if ids else [1]
    pid = int(request.args.get("profile_id", session.get("profile_id", 1)))
    return False, [pid]


def _resolve_aggregate_profile(ticker, profile_ids):
    """For aggregate writes, find the profile with the largest position for a ticker."""
    conn = get_connection()
    placeholders = ",".join("?" * len(profile_ids))
    row = conn.execute(
        f"SELECT profile_id FROM all_account_info WHERE ticker = ? AND profile_id IN ({placeholders}) ORDER BY quantity DESC LIMIT 1",
        [ticker] + profile_ids,
    ).fetchone()
    conn.close()
    if row:
        return row["profile_id"] if isinstance(row, dict) else row[0]
    return profile_ids[0]


def _get_write_profile_id():
    """Get a single profile_id for write operations, resolving aggregate if needed."""
    is_agg, pids = get_profile_filter()
    if is_agg:
        return pids[0]  # default; caller should use _resolve_aggregate_profile for ticker-specific ops
    return pids[0]


def rows_to_dicts(rows):
    """Convert sqlite3.Row results to a list of dicts."""
    return [dict(r) for r in rows]


_ACCOUNT_MATCH_IGNORED_TOKENS = {
    "account", "acct", "portfolio",
    "etrade", "e", "trade",
    "schwab", "charles", "fidelity",
    "snowball", "analytics",
    "traditional",
}


def _normalize_account_tokens(label):
    text = re.sub(r"[^a-z0-9]+", " ", (label or "").strip().lower())
    return [tok for tok in text.split() if tok]


def _account_match_info(account_name, profile_name, source_format=None):
    """Return whether an import file account appears to match the target profile."""
    if not account_name:
        return {"matched": True, "reason": "missing_account_name"}

    account_tokens = _normalize_account_tokens(account_name)
    profile_tokens = _normalize_account_tokens(profile_name)
    source_format = (source_format or "").strip().lower()

    account_core = {
        tok for tok in account_tokens
        if not tok.isdigit() and tok not in _ACCOUNT_MATCH_IGNORED_TOKENS
    }
    profile_core = {
        tok for tok in profile_tokens
        if not tok.isdigit() and tok not in _ACCOUNT_MATCH_IGNORED_TOKENS
    }

    if not account_core:
        return {"matched": True, "reason": "no_core_tokens"}

    matched = account_core == profile_core
    reason = "exact_core_match" if matched else "core_token_mismatch"
    message = None
    if matched and source_format == "etrade" and "etrade" not in profile_tokens:
        matched = False
        reason = "broker_profile_mismatch"
        message = (
            f"E*TRADE files can only be imported into an E*TRADE portfolio. "
            f"'{profile_name}' is not marked as E*TRADE."
        )
    elif not matched:
        message = (
            f"File account '{account_name}' does not match selected portfolio "
            f"'{profile_name}'. Switch portfolios before importing."
        )

    return {
        "matched": matched,
        "reason": reason,
        "message": message,
    }


def _get_profile_name(profile_id):
    conn = get_connection()
    try:
        row = conn.execute("SELECT name FROM profiles WHERE id = ?", (profile_id,)).fetchone()
        if not row:
            return "Portfolio"
        return row["name"] if isinstance(row, dict) else row[0]
    finally:
        conn.close()


def _profile_is_positions_managed(profile_id, conn=None):
    close = False
    if conn is None:
        conn = get_connection()
        close = True
    try:
        row = conn.execute(
            "SELECT positions_managed FROM profiles WHERE id = ?",
            (profile_id,),
        ).fetchone()
        if not row:
            return False
        return bool(row["positions_managed"] if isinstance(row, dict) else row[0])
    finally:
        if close:
            conn.close()


def _set_profile_positions_managed(profile_id, managed, conn=None):
    close = False
    if conn is None:
        conn = get_connection()
        close = True
    try:
        conn.execute(
            "UPDATE profiles SET positions_managed = ? WHERE id = ?",
            (1 if managed else 0, profile_id),
        )
        if close:
            conn.commit()
    finally:
        if close:
            conn.close()


def _get_owner_source_profile_ids(conn):
    """Return profile ids that feed Owner reconciliation/DRIP sync."""
    rows = conn.execute(
        "SELECT id FROM profiles WHERE id != 1 AND include_in_owner = 1 ORDER BY id"
    ).fetchall()
    return [r["id"] if isinstance(r, dict) else r[0] for r in rows]


def _get_refresh_target_info(conn):
    """Resolve which profiles a holdings refresh should update."""
    if request.args.get("aggregate") == "true":
        rows = conn.execute(
            "SELECT member_profile_id FROM aggregate_config ORDER BY member_profile_id"
        ).fetchall()
        pids = [r["member_profile_id"] if isinstance(r, dict) else r[0] for r in rows]
        return {
            "scope": "aggregate",
            "source_profile_ids": pids or [1],
            "selected_profile_id": None,
        }

    profile_id = get_profile_id()
    if profile_id == 1:
        source_ids = _get_owner_source_profile_ids(conn)
        return {
            "scope": "owner",
            "source_profile_ids": source_ids or [1],
            "selected_profile_id": 1,
        }

    return {
        "scope": "profile",
        "source_profile_ids": [profile_id],
        "selected_profile_id": profile_id,
    }


def _estimate_month_income(holding, month):
    """Estimate income for a specific month (1-12) from a single holding.

    Uses div_frequency and ex_div_date to determine whether the holding
    pays out in the given calendar month and returns qty * div if so.
    Weekly payers use 4.33 payments per month (52 / 12).
    """
    import datetime

    qty = float(holding.get("quantity") or 0)
    div = float(holding.get("div") or 0)
    freq = (holding.get("div_frequency") or "").strip().upper()
    ex_raw = holding.get("ex_div_date") or ""

    if qty <= 0 or div <= 0 or not freq:
        return 0.0

    # Weekly: one payment per week, 52 per year ≈ 4.33 per month.
    # Consistent with the projected income chart (annual / 52 * 4.33).
    if freq in ("W", "52"):
        return round(qty * div * 4.33, 2)

    # Monthly: always pays
    if freq == "M":
        return round(qty * div, 2)

    # For Q / SA / A we need the ex_div_date to figure out the cycle
    ex_month = None
    for fmt in ("%m/%d/%y", "%Y-%m-%d", "%m/%d/%Y"):
        try:
            ex_month = datetime.datetime.strptime(ex_raw.strip(), fmt).month
            break
        except (ValueError, TypeError):
            continue

    if ex_month is None:
        return 0.0

    period = {"Q": 3, "SA": 6, "A": 12}.get(freq)
    if period is None:
        return 0.0

    if (month - ex_month) % period == 0:
        return round(qty * div, 2)

    return 0.0


def _estimate_current_month_income(holding):
    """Estimate income for the current month from a single holding."""
    import datetime
    return _estimate_month_income(holding, datetime.date.today().month)


def _estimate_ytd_income(holding):
    """Estimate year-to-date income from a single holding (Jan through last completed month).
    Excludes current month since it's incomplete — avoids overcounting."""
    import datetime
    cur_month = datetime.date.today().month
    return sum(_estimate_month_income(holding, m) for m in range(1, cur_month))


def _frequency_to_payments_per_year(freq):
    return {
        "W": 52,
        "52": 52,
        "M": 12,
        "Q": 4,
        "SA": 2,
        "A": 1,
    }.get((freq or "").strip().upper(), 0)


def _recompute_position_income_fields(conn, profile_id, ticker, broker_dividend_yield=None):
    row = conn.execute(
        """SELECT quantity, purchase_value, current_value, div, div_frequency, ex_div_date
           FROM all_account_info
           WHERE ticker = ? AND profile_id = ?""",
        (ticker, profile_id),
    ).fetchone()
    if not row:
        return

    quantity = float(row["quantity"] or 0)
    purchase_value = float(row["purchase_value"] or 0)
    current_value = float(row["current_value"] or 0)
    div = float(row["div"] or 0)
    div_frequency = row["div_frequency"] or ""
    payments_per_year = _frequency_to_payments_per_year(div_frequency)

    annual = 0.0
    if quantity > 0 and div > 0 and payments_per_year > 0:
        annual = quantity * div * payments_per_year
    elif quantity > 0 and current_value > 0 and broker_dividend_yield:
        annual = current_value * (float(broker_dividend_yield) / 100.0)

    monthly = annual / 12.0 if annual > 0 else 0.0
    yoc = (annual / purchase_value * 100.0) if annual > 0 and purchase_value > 0 else 0.0
    cur_yield = (annual / current_value * 100.0) if annual > 0 and current_value > 0 else 0.0
    current_month_income = _estimate_current_month_income({
        "quantity": quantity,
        "div": div,
        "div_frequency": div_frequency,
        "ex_div_date": row["ex_div_date"],
    }) if quantity > 0 else 0.0

    conn.execute(
        """UPDATE all_account_info
           SET estim_payment_per_year = ?, approx_monthly_income = ?,
               annual_yield_on_cost = ?, current_annual_yield = ?,
               current_month_income = ?
           WHERE ticker = ? AND profile_id = ?""",
        (
            round(annual, 2),
            round(monthly, 2),
            round(yoc, 4),
            round(cur_yield, 4),
            round(current_month_income, 2),
            ticker,
            profile_id,
        ),
    )


def _calc_dividend_growth_batch(tickers):
    """Calculate 3-year and 5-year dividend growth rates for a list of tickers.
    Returns dict of ticker -> {"div_growth_3y": float|None, "div_growth_5y": float|None}.
    Uses yfinance Ticker.dividends for accurate per-share dividend history.
    Compares annual per-share dividends (excluding current partial year).
    """
    import yfinance as yf
    import warnings
    import datetime
    warnings.filterwarnings("ignore")

    result = {t: {"div_growth_3y": None, "div_growth_5y": None} for t in tickers}
    if not tickers:
        return result

    current_year = datetime.date.today().year

    for t in tickers:
        try:
            divs = yf.Ticker(t).dividends
            if divs is None or len(divs) < 2:
                continue
            divs = divs[divs > 0]
            divs.index = pd.to_datetime(divs.index)
            # Exclude current partial year
            divs = divs[divs.index.year < current_year]
            if len(divs) < 2:
                continue

            yearly = divs.resample("YE").sum()
            yearly = yearly[yearly > 0]
            if len(yearly) < 2:
                continue

            recent = float(yearly.iloc[-1])

            if len(yearly) >= 4:
                past_3 = float(yearly.iloc[-4])
                if past_3 > 0:
                    result[t]["div_growth_3y"] = round(((recent / past_3) ** (1 / 3) - 1) * 100, 2)

            if len(yearly) >= 6:
                past_5 = float(yearly.iloc[-6])
                if past_5 > 0:
                    result[t]["div_growth_5y"] = round(((recent / past_5) ** (1 / 5) - 1) * 100, 2)
        except Exception:
            continue

    return result


def _auto_reconcile_owner():
    """Silently reconcile Owner (profile 1) from sub-profiles after any import.

    Only runs if the Owner import has been used (owner_import_used setting).
    Syncs quantities, prices, and income fields while preserving Owner-only
    data (ytd_divs, total_divs_received, paid_for_itself, current_month_income,
    estim_payment_per_year, approx_monthly_income).
    """
    from datetime import date as _date

    conn = get_connection()
    owner_id = 1

    # Only reconcile if Owner import was used
    _oiu = conn.execute(
        "SELECT value FROM settings WHERE key = 'owner_import_used'"
    ).fetchone()
    if not (_oiu and _oiu[0] == "true"):
        conn.close()
        return

    # Get all non-Owner profiles that are marked for inclusion
    rows = conn.execute("SELECT id FROM profiles WHERE id != ? AND include_in_owner = 1", (owner_id,)).fetchall()
    source_ids = [r["id"] if isinstance(r, dict) else r[0] for r in rows]
    if not source_ids:
        conn.close()
        return

    placeholders = ",".join("?" * len(source_ids))

    # Aggregate holdings from source portfolios
    agg_rows = conn.execute(f"""
        SELECT
            ticker,
            MAX(description) as description,
            MAX(classification_type) as classification_type,
            SUM(quantity) as quantity,
            CASE WHEN SUM(quantity) > 0 THEN SUM(purchase_value) / SUM(quantity) ELSE 0 END as price_paid,
            MAX(current_price) as current_price,
            SUM(purchase_value) as purchase_value,
            SUM(current_value) as current_value,
            SUM(gain_or_loss) as gain_or_loss,
            CASE WHEN SUM(purchase_value) > 0 THEN SUM(gain_or_loss) / SUM(purchase_value) ELSE 0 END as gain_or_loss_percentage,
            CASE WHEN SUM(purchase_value) > 0 THEN SUM(gain_or_loss) / SUM(purchase_value) ELSE 0 END as percent_change,
            MAX(div_frequency) as div_frequency,
            MAX(reinvest) as reinvest,
            MAX(ex_div_date) as ex_div_date,
            MAX(div_pay_date) as div_pay_date,
            MAX(div) as div,
            SUM(dividend_paid) as dividend_paid,
            SUM(estim_payment_per_year) as estim_payment_per_year,
            SUM(approx_monthly_income) as approx_monthly_income,
            SUM(withdraw_8pct_cost_annually) as withdraw_8pct_cost_annually,
            SUM(withdraw_8pct_per_month) as withdraw_8pct_per_month,
            SUM(cash_not_reinvested) as cash_not_reinvested,
            SUM(total_cash_reinvested) as total_cash_reinvested,
            SUM(shares_bought_from_dividend) as shares_bought_from_dividend,
            SUM(shares_bought_in_year) as shares_bought_in_year,
            SUM(shares_in_month) as shares_in_month,
            MIN(purchase_date) as purchase_date
        FROM all_account_info
        WHERE profile_id IN ({placeholders})
        GROUP BY ticker
    """, source_ids).fetchall()

    agg_map = {r["ticker"]: dict(r) for r in agg_rows}

    owner_rows = conn.execute(
        "SELECT ticker FROM all_account_info WHERE profile_id = ?", (owner_id,)
    ).fetchall()
    owner_tickers = {r["ticker"] for r in owner_rows}

    # Sync these fields from sub-profiles.  Owner-only payout history fields
    # (ytd_divs, total_divs_received, paid_for_itself, current_month_income)
    # are preserved since sub-profiles don't track those.
    sync_fields = [
        "description", "classification_type", "quantity", "price_paid",
        "current_price", "purchase_value", "current_value", "gain_or_loss",
        "gain_or_loss_percentage", "percent_change", "div_frequency", "reinvest",
        "ex_div_date", "div_pay_date", "div", "dividend_paid",
        "estim_payment_per_year", "approx_monthly_income",
        "withdraw_8pct_cost_annually", "withdraw_8pct_per_month",
        "cash_not_reinvested", "total_cash_reinvested",
        "shares_bought_from_dividend", "shares_bought_in_year", "shares_in_month",
        "purchase_date",
    ]

    for ticker, agg in agg_map.items():
        if ticker in owner_tickers:
            sets = [f"{f} = ?" for f in sync_fields] + ["import_date = ?"]
            vals = [agg.get(f) for f in sync_fields] + [_date.today().isoformat(), ticker, owner_id]
            conn.execute(
                f"UPDATE all_account_info SET {', '.join(sets)} WHERE ticker = ? AND profile_id = ?",
                vals,
            )
        else:
            cols = ["ticker", "profile_id", "import_date"] + sync_fields
            vals = [ticker, owner_id, _date.today().isoformat()] + [agg.get(f) for f in sync_fields]
            ph = ",".join("?" * len(cols))
            conn.execute(f"INSERT INTO all_account_info ({', '.join(cols)}) VALUES ({ph})", vals)

    # Remove tickers from Owner that no longer exist in any sub-portfolio
    for ticker in owner_tickers - set(agg_map.keys()):
        conn.execute(
            "DELETE FROM all_account_info WHERE ticker = ? AND profile_id = ?",
            (ticker, owner_id),
        )

    conn.commit()

    # Refresh derived tables for Owner
    populate_holdings(owner_id)
    populate_dividends(owner_id)
    populate_income_tracking(owner_id)

    conn.close()


# ── Startup ────────────────────────────────────────────────────────────────────

@app.before_request
def _ensure_db():
    """Create tables on first request if they don't exist."""
    if not getattr(app, '_db_initialized', False):
        conn = get_connection()
        ensure_tables_exist(conn)
        conn.close()
        app._db_initialized = True


# ── Profiles ───────────────────────────────────────────────────────────────────

@app.route("/api/profiles", methods=["GET"])
def list_profiles():
    conn = get_connection()
    rows = conn.execute("SELECT id, name, created_at FROM profiles ORDER BY id").fetchall()
    conn.close()
    return jsonify(rows_to_dicts(rows))


@app.route("/api/profiles", methods=["POST"])
def create_profile():
    data = request.get_json()
    name = data.get("name", "").strip()
    if not name:
        return jsonify({"error": "Name is required"}), 400
    conn = get_connection()
    cur = conn.execute(
        "INSERT INTO profiles (name, include_in_owner) VALUES (?, 1)", (name,)
    )
    pid = cur.lastrowid
    # Auto-add to Combined aggregate if one exists
    has_agg = conn.execute(
        "SELECT 1 FROM aggregate_config LIMIT 1"
    ).fetchone()
    if has_agg:
        conn.execute(
            "INSERT OR IGNORE INTO aggregate_config (member_profile_id) VALUES (?)",
            (pid,),
        )
    conn.commit()
    conn.close()
    return jsonify({"id": pid, "name": name}), 201


@app.route("/api/profiles/<int:pid>", methods=["PUT"])
def update_profile(pid):
    data = request.get_json()
    name = data.get("name", "").strip()
    if not name:
        return jsonify({"error": "Name is required"}), 400
    conn = get_connection()
    conn.execute("UPDATE profiles SET name = ? WHERE id = ?", (name, pid))
    conn.commit()
    conn.close()
    return jsonify({"id": pid, "name": name})


@app.route("/api/profiles/<int:pid>", methods=["DELETE"])
def delete_profile(pid):
    if pid == 1:
        return jsonify({"error": "Cannot delete the default profile"}), 400
    conn = get_connection()
    _clear_profile_data(conn, pid)
    conn.execute("DELETE FROM aggregate_config WHERE member_profile_id = ?", (pid,))
    conn.execute("DELETE FROM profiles WHERE id = ?", (pid,))
    conn.commit()
    conn.close()
    return jsonify({"deleted": pid})


def _clear_profile_data(conn, pid):
    """Remove all data for a profile from every profile-scoped table."""
    for table in [
        "all_account_info", "holdings", "dividends", "income_tracking",
        "weekly_payouts", "monthly_payouts", "weekly_payout_tickers",
        "monthly_payout_tickers", "drip_settings", "drip_contribution_targets",
        "drip_redirects", "swap_candidates", "ticker_categories",
    ]:
        conn.execute(f"DELETE FROM {table} WHERE profile_id = ?", (pid,))


@app.route("/api/profiles/<int:pid>/clear", methods=["POST"])
def clear_profile_data(pid):
    """Clear all data for a profile without deleting the profile itself."""
    conn = get_connection()
    _clear_profile_data(conn, pid)
    conn.commit()
    conn.close()
    return jsonify({"cleared": pid, "message": f"All data cleared for profile {pid}"})


@app.route("/api/profiles/<int:pid>/include-in-owner", methods=["PUT"])
def set_include_in_owner(pid):
    """Toggle whether a profile is included in Owner reconciliation."""
    if pid == 1:
        return jsonify({"error": "Owner is always included"}), 400
    data = request.get_json() or {}
    val = 1 if data.get("include") else 0
    conn = get_connection()
    conn.execute("UPDATE profiles SET include_in_owner = ? WHERE id = ?", (val, pid))
    conn.commit()
    conn.close()
    return jsonify({"id": pid, "include_in_owner": val})


@app.route("/api/profiles/summary", methods=["GET"])
def profiles_summary():
    """Return per-profile stats for the Manage Portfolios page."""
    conn = get_connection()
    rows = conn.execute("""
        SELECT p.id, p.name, p.created_at, p.include_in_owner,
               COUNT(a.ticker) as holdings_count,
               COALESCE(SUM(a.current_value), 0) as total_value
        FROM profiles p
        LEFT JOIN all_account_info a ON p.id = a.profile_id
        GROUP BY p.id
        ORDER BY p.id
    """).fetchall()
    flag = conn.execute("SELECT value FROM settings WHERE key = 'owner_import_used'").fetchone()
    conn.close()
    return jsonify({"profiles": rows_to_dicts(rows), "owner_import_used": bool(flag)})


@app.route("/api/aggregate-config", methods=["GET"])
def get_aggregate_config():
    conn = get_connection()
    rows = conn.execute("SELECT member_profile_id FROM aggregate_config").fetchall()
    name_row = conn.execute("SELECT value FROM settings WHERE key = 'aggregate_name'").fetchone()
    conn.close()
    ids = [r["member_profile_id"] if isinstance(r, dict) else r[0] for r in rows]
    name = (name_row[0] if name_row else "Aggregate")
    return jsonify({"member_ids": ids, "name": name})


@app.route("/api/aggregate-config", methods=["PUT"])
def set_aggregate_config():
    data = request.get_json()
    member_ids = data.get("member_ids", [])
    name = data.get("name", "Aggregate")
    conn = get_connection()
    conn.execute("DELETE FROM aggregate_config")
    for mid in member_ids:
        conn.execute("INSERT OR IGNORE INTO aggregate_config (member_profile_id) VALUES (?)", (mid,))
    conn.execute("INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)", ("aggregate_name", name))
    conn.commit()
    conn.close()
    return jsonify({"member_ids": member_ids, "name": name})


@app.route("/api/aggregate-config", methods=["DELETE"])
def delete_aggregate_config():
    conn = get_connection()
    conn.execute("DELETE FROM aggregate_config")
    conn.execute("DELETE FROM settings WHERE key = 'aggregate_name'")
    conn.commit()
    conn.close()
    return jsonify({"deleted": True})


@app.route("/api/profiles/reconcile-owner", methods=["POST"])
def reconcile_owner():
    """Compare Owner (profile 1) against the sum of the other portfolios and sync mismatches.

    For each ticker across all non-Owner profiles:
    - If missing from Owner: insert it
    - If quantity/value differs: update Owner
    - If ticker exists in Owner but not in any sub-portfolio: remove it
    """
    from datetime import date as _date

    data = request.get_json() or {}
    source_ids = data.get("source_ids", [])

    conn = get_connection()
    owner_id = 1

    if not source_ids:
        # Default: all profiles except Owner that are marked for inclusion
        rows = conn.execute("SELECT id FROM profiles WHERE id != ? AND include_in_owner = 1", (owner_id,)).fetchall()
        source_ids = [r["id"] if isinstance(r, dict) else r[0] for r in rows]

    if not source_ids:
        conn.close()
        return jsonify({"error": "No source portfolios to reconcile from"}), 400

    placeholders = ",".join("?" * len(source_ids))

    # Aggregate holdings from source portfolios
    agg_rows = conn.execute(f"""
        SELECT
            ticker,
            MAX(description) as description,
            MAX(classification_type) as classification_type,
            SUM(quantity) as quantity,
            CASE WHEN SUM(quantity) > 0 THEN SUM(purchase_value) / SUM(quantity) ELSE 0 END as price_paid,
            MAX(current_price) as current_price,
            SUM(purchase_value) as purchase_value,
            SUM(current_value) as current_value,
            SUM(gain_or_loss) as gain_or_loss,
            CASE WHEN SUM(purchase_value) > 0 THEN SUM(gain_or_loss) / SUM(purchase_value) ELSE 0 END as gain_or_loss_percentage,
            CASE WHEN SUM(purchase_value) > 0 THEN SUM(gain_or_loss) / SUM(purchase_value) ELSE 0 END as percent_change,
            MAX(div_frequency) as div_frequency,
            MAX(reinvest) as reinvest,
            MAX(ex_div_date) as ex_div_date,
            MAX(div_pay_date) as div_pay_date,
            MAX(div) as div,
            SUM(dividend_paid) as dividend_paid,
            SUM(estim_payment_per_year) as estim_payment_per_year,
            SUM(approx_monthly_income) as approx_monthly_income,
            SUM(withdraw_8pct_cost_annually) as withdraw_8pct_cost_annually,
            SUM(withdraw_8pct_per_month) as withdraw_8pct_per_month,
            SUM(cash_not_reinvested) as cash_not_reinvested,
            SUM(total_cash_reinvested) as total_cash_reinvested,
            SUM(shares_bought_from_dividend) as shares_bought_from_dividend,
            SUM(shares_bought_in_year) as shares_bought_in_year,
            SUM(shares_in_month) as shares_in_month,
            SUM(ytd_divs) as ytd_divs,
            SUM(total_divs_received) as total_divs_received,
            CASE WHEN SUM(purchase_value) > 0 THEN SUM(total_divs_received) / SUM(purchase_value) ELSE 0 END as paid_for_itself,
            MIN(purchase_date) as purchase_date,
            SUM(current_month_income) as current_month_income
        FROM all_account_info
        WHERE profile_id IN ({placeholders})
        GROUP BY ticker
    """, source_ids).fetchall()

    agg_map = {r["ticker"]: dict(r) for r in agg_rows}

    # Get current Owner holdings
    owner_rows = conn.execute(
        "SELECT ticker, quantity, current_value FROM all_account_info WHERE profile_id = ?",
        (owner_id,),
    ).fetchall()
    owner_tickers = {r["ticker"] for r in owner_rows}
    owner_qty = {r["ticker"]: r["quantity"] for r in owner_rows}

    inserted = 0
    updated = 0
    removed = 0

    # Fields to sync from sub-profiles to Owner.
    # Owner-only payout history fields are preserved: ytd_divs,
    # total_divs_received, paid_for_itself, current_month_income.
    update_fields = [
        "description", "classification_type", "quantity", "price_paid",
        "current_price", "purchase_value", "current_value", "gain_or_loss",
        "gain_or_loss_percentage", "percent_change", "div_frequency", "reinvest",
        "ex_div_date", "div_pay_date", "div", "dividend_paid",
        "estim_payment_per_year", "approx_monthly_income",
        "withdraw_8pct_cost_annually", "withdraw_8pct_per_month",
        "cash_not_reinvested", "total_cash_reinvested",
        "shares_bought_from_dividend", "shares_bought_in_year", "shares_in_month",
        "purchase_date",
    ]

    for ticker, agg in agg_map.items():
        if ticker in owner_tickers:
            # Sync quantity/income fields from sub-portfolios, preserve Owner-only fields
            sets = []
            vals = []
            for f in update_fields:
                v = agg.get(f)
                sets.append(f"{f} = ?")
                vals.append(v)
            sets.append("import_date = ?")
            vals.append(_date.today().isoformat())
            vals.extend([ticker, owner_id])
            conn.execute(
                f"UPDATE all_account_info SET {', '.join(sets)} WHERE ticker = ? AND profile_id = ?",
                vals,
            )
            updated += 1
        else:
            # Insert new ticker into Owner (includes all fields since Owner has no prior data)
            all_fields = update_fields + ["ytd_divs", "total_divs_received", "paid_for_itself", "current_month_income"]
            cols = ["ticker", "profile_id", "import_date"] + all_fields
            vals = [ticker, owner_id, _date.today().isoformat()] + [agg.get(f) for f in all_fields]
            ph = ",".join("?" * len(cols))
            conn.execute(f"INSERT INTO all_account_info ({', '.join(cols)}) VALUES ({ph})", vals)
            inserted += 1

    # Remove tickers from Owner that no longer exist in any sub-portfolio
    agg_tickers = set(agg_map.keys())
    for ticker in owner_tickers - agg_tickers:
        conn.execute(
            "DELETE FROM all_account_info WHERE ticker = ? AND profile_id = ?",
            (ticker, owner_id),
        )
        removed += 1

    conn.commit()

    # Refresh derived tables for Owner
    populate_holdings(owner_id)
    populate_dividends(owner_id)
    populate_income_tracking(owner_id)

    conn.close()

    parts = []
    if updated: parts.append(f"{updated} updated")
    if inserted: parts.append(f"{inserted} added")
    if removed: parts.append(f"{removed} removed")
    msg = f"Owner reconciled: {', '.join(parts)}." if parts else "Owner is already in sync — no changes needed."

    return jsonify({"updated": updated, "inserted": inserted, "removed": removed, "message": msg})


# ── Import endpoints ──────────────────────────────────────────────────────────

@app.route("/api/import/excel", methods=["POST"])
def api_import_excel():
    """Import the owner's Excel spreadsheet (All Accounts sheet)."""
    profile_id = get_profile_id()
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    f = request.files["file"]
    path = os.path.join(UPLOAD_FOLDER, f.filename)
    f.save(path)
    try:
        multi = request.form.get("multi_sheet", "false").lower() == "true"
        as_txns = request.form.get("as_transactions", "false").lower() == "true"
        if multi:
            # Snapshot all existing profiles before import
            pre_snaps = {}
            if as_txns:
                conn_snap = get_connection()
                all_profiles = conn_snap.execute("SELECT id FROM profiles").fetchall()
                conn_snap.close()
                for p in all_profiles:
                    pid = p["id"] if isinstance(p, dict) else p[0]
                    pre_snaps[pid] = _snapshot_positions(pid)
            results = import_multi_excel(path, default_profile_id=profile_id)
            # Populate derived tables for each imported profile
            for r in results:
                if r["rows"] > 0:
                    pid = r["profile_id"]
                    if as_txns:
                        _import_as_transactions(pid, pre_snaps.get(pid, {}))
                    populate_holdings(pid)
                    populate_dividends(pid)
                    populate_income_tracking(pid)
                    populate_pillar_weights(pid)
                    if not as_txns:
                        _rerollup_after_import(pid)
            total = sum(r["rows"] for r in results)
            # Mark owner import as used
            conn2 = get_connection()
            conn2.execute("INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)", ("owner_import_used", "true"))
            conn2.commit()
            conn2.close()
            # Auto-reconcile Owner quantities from sub-profiles
            _auto_reconcile_owner()
            return jsonify({"rows": total, "message": f"Imported {len(results)} sheets ({total} total holdings)", "details": results})
        else:
            sheet = request.form.get("sheet_name", "All Accounts")
            # Snapshot BEFORE import so we know the pre-import state
            pre_snap = _snapshot_positions(profile_id) if as_txns else None
            count, msg = import_from_excel(path, sheet_name=sheet, profile_id=profile_id)
            # Auto-populate derived tables
            if as_txns:
                _import_as_transactions(profile_id, pre_snap)
            populate_holdings(profile_id)
            populate_dividends(profile_id)
            populate_income_tracking(profile_id)
            populate_pillar_weights(profile_id)
            if not as_txns:
                _rerollup_after_import(profile_id)
            # Mark owner import as used
            conn2 = get_connection()
            conn2.execute("INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)", ("owner_import_used", "true"))
            conn2.commit()
            conn2.close()
            # Auto-reconcile Owner quantities from sub-profiles
            _auto_reconcile_owner()
            return jsonify({"rows": count, "message": msg})
    except Exception as e:
        return jsonify({"error": str(e)}), 400
    finally:
        try:
            if os.path.exists(path):
                os.remove(path)
        except OSError:
            pass


@app.route("/api/import/generic", methods=["POST"])
def api_import_generic():
    """Import a generic user spreadsheet (Ticker + Shares minimum)."""
    profile_id = get_profile_id()
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    f = request.files["file"]
    path = os.path.join(UPLOAD_FOLDER, f.filename)
    f.save(path)
    try:
        multi = request.form.get("multi_sheet", "false").lower() == "true"
        as_txns = request.form.get("as_transactions", "false").lower() == "true"
        if multi:
            # Snapshot all existing profiles before import
            pre_snaps = {}
            if as_txns:
                conn_snap = get_connection()
                all_profiles = conn_snap.execute("SELECT id FROM profiles").fetchall()
                conn_snap.close()
                for p in all_profiles:
                    pid = p["id"] if isinstance(p, dict) else p[0]
                    pre_snaps[pid] = _snapshot_positions(pid)
            results = import_multi_upload(path)
            for r in results:
                if r["rows"] > 0:
                    pid = r["profile_id"]
                    if as_txns:
                        _import_as_transactions(pid, pre_snaps.get(pid, {}))
                    populate_holdings(pid)
                    populate_dividends(pid)
                    populate_income_tracking(pid)
                    if not as_txns:
                        _rerollup_after_import(pid)
            total = sum(r["rows"] for r in results)
            # Auto-reconcile Owner quantities from sub-profiles
            _auto_reconcile_owner()
            return jsonify({"rows": total, "message": f"Imported {len(results)} portfolios ({total} total holdings)", "details": results})
        else:
            # Snapshot BEFORE import
            pre_snap = _snapshot_positions(profile_id) if as_txns else None
            df = pd.read_excel(path, engine="openpyxl")
            count, msg = import_from_upload(df, profile_id)
            if as_txns:
                _import_as_transactions(profile_id, pre_snap)
            populate_holdings(profile_id)
            populate_dividends(profile_id)
            populate_income_tracking(profile_id)
            if not as_txns:
                _rerollup_after_import(profile_id)
            # Auto-reconcile Owner if a sub-profile was imported
            if profile_id != 1:
                _auto_reconcile_owner()
            return jsonify({"rows": count, "message": msg})
    except Exception as e:
        return jsonify({"error": str(e)}), 400
    finally:
        try:
            if os.path.exists(path):
                os.remove(path)
        except OSError:
            pass


@app.route("/api/import/weekly-payouts", methods=["POST"])
def api_import_weekly():
    # Weekly payouts are portfolio-wide actuals — always store under Owner (id=1)
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    f = request.files["file"]
    import uuid
    path = os.path.join(UPLOAD_FOLDER, f"weekly_{uuid.uuid4().hex}_{f.filename}")
    try:
        f.save(path)
        count, msg = import_weekly_payouts(path, 1)
        return jsonify({"rows": count, "message": msg})
    except Exception as e:
        return jsonify({"error": str(e)}), 400
    finally:
        try:
            if os.path.exists(path):
                os.remove(path)
        except OSError:
            pass


@app.route("/api/import/monthly-payouts", methods=["POST"])
def api_import_monthly():
    # Monthly payouts are portfolio-wide actuals — always store under Owner (id=1)
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    f = request.files["file"]
    import uuid
    path = os.path.join(UPLOAD_FOLDER, f"monthly_{uuid.uuid4().hex}_{f.filename}")
    try:
        f.save(path)
        count, msg = import_monthly_payouts(path, 1)
        return jsonify({"rows": count, "message": msg})
    except Exception as e:
        return jsonify({"error": str(e)}), 400
    finally:
        try:
            if os.path.exists(path):
                os.remove(path)
        except OSError:
            pass


@app.route("/api/import/monthly-payout-tickers", methods=["POST"])
def api_import_monthly_tickers():
    # Monthly payout tickers are portfolio-wide — always store under Owner (id=1)
    profile_id = 1
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    f = request.files["file"]
    import uuid
    path = os.path.join(UPLOAD_FOLDER, f"tickers_{uuid.uuid4().hex}_{f.filename}")
    try:
        f.save(path)
        count, msg = import_monthly_payout_tickers(path, profile_id)
        return jsonify({"rows": count, "message": msg})
    except Exception as e:
        return jsonify({"error": str(e)}), 400
    finally:
        try:
            if os.path.exists(path):
                os.remove(path)
        except OSError:
            pass


# ── Transaction History Import ────────────────────────────────────────────────

@app.route("/api/import/transactions/preview", methods=["POST"])
def api_import_transactions_preview():
    """Parse a transaction history CSV and return a preview (no DB writes)."""
    profile_id = get_profile_id()
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    f = request.files["file"]
    fmt = request.form.get("format", "snowball")
    if fmt not in TXN_PARSERS:
        return jsonify({"error": f"Unknown format: {fmt}"}), 400

    import uuid
    path = os.path.join(UPLOAD_FOLDER, f"txn_{uuid.uuid4().hex}_{f.filename}")
    f.save(path)
    try:
        result = TXN_PARSERS[fmt](path, f.filename)
        if result.get("account_name"):
            result["target_profile_name"] = _get_profile_name(profile_id)
            result["account_match"] = _account_match_info(
                result.get("account_name"),
                result["target_profile_name"],
                fmt,
            )
        return jsonify(result)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        return jsonify({"error": f"Failed to parse file: {e}"}), 400
    finally:
        try:
            if os.path.exists(path):
                os.remove(path)
        except OSError:
            pass


def _import_positions(parsed, profile_id):
    """Import holdings from a positions-based file (e.g. Schwab Positions CSV).

    Writes directly to all_account_info — sets quantity, cost basis, price,
    and current value. Existing holdings for the profile are updated; new
    tickers are inserted. Holdings not in the import are zeroed out.
    """
    from datetime import date as _date
    conn = get_connection()

    positions = parsed["positions"]
    positions_by_ticker = {pos["ticker"]: pos for pos in positions}
    imported_tickers = set()
    updated = 0
    inserted = 0

    try:
        _set_profile_positions_managed(profile_id, True, conn)

        for pos in positions:
            ticker = pos["ticker"]
            imported_tickers.add(ticker)

            existing = conn.execute(
                "SELECT 1 FROM all_account_info WHERE ticker = ? AND profile_id = ?",
                (ticker, profile_id),
            ).fetchone()

            if existing:
                conn.execute(
                    """UPDATE all_account_info SET
                        quantity = ?, price_paid = ?, current_price = ?,
                        purchase_value = ?, current_value = ?, gain_or_loss = ?,
                        base_quantity = ?, import_date = ?,
                        description = CASE WHEN description IS NULL OR description = ''
                                       THEN ? ELSE description END
                    WHERE ticker = ? AND profile_id = ?""",
                    (
                        pos["quantity"], pos["cost_per_share"], pos["current_price"],
                        pos["purchase_value"], pos["current_value"], pos["gain_or_loss"],
                        pos["quantity"], _date.today().isoformat(),
                        pos["description"],
                        ticker, profile_id,
                    ),
                )
                updated += 1
            else:
                conn.execute(
                    """INSERT INTO all_account_info
                        (ticker, profile_id, quantity, base_quantity, price_paid, current_price,
                         purchase_value, current_value, gain_or_loss,
                         description, import_date)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                    (
                        ticker, profile_id, pos["quantity"], pos["quantity"],
                        pos["cost_per_share"], pos["current_price"],
                        pos["purchase_value"], pos["current_value"], pos["gain_or_loss"],
                        pos["description"], _date.today().isoformat(),
                    ),
                )
                inserted += 1

        # Zero out holdings not in the positions file
        zeroed = 0
        stale_rows = conn.execute(
            "SELECT ticker FROM all_account_info WHERE profile_id = ? AND quantity > 0",
            (profile_id,),
        ).fetchall()
        for row in stale_rows:
            if row["ticker"] not in imported_tickers:
                conn.execute(
                    """UPDATE all_account_info SET quantity = 0, current_value = 0,
                       purchase_value = 0, gain_or_loss = 0,
                       base_quantity = 0, import_date = ?
                    WHERE ticker = ? AND profile_id = ?""",
                    (_date.today().isoformat(), row["ticker"], profile_id),
                )
                zeroed += 1

        touched_tickers = set(imported_tickers)
        touched_tickers.update(
            row["ticker"] if isinstance(row, dict) else row[0]
            for row in stale_rows
            if (row["ticker"] if isinstance(row, dict) else row[0]) not in imported_tickers
        )
        for ticker in touched_tickers:
            broker_yield = None
            if ticker in positions_by_ticker:
                broker_yield = positions_by_ticker[ticker].get("dividend_yield")
            _recompute_position_income_fields(conn, profile_id, ticker, broker_yield)

        conn.commit()

        # Run normalize passes
        populate_holdings(profile_id)
        populate_dividends(profile_id)
        populate_income_tracking(profile_id)

        msg = f"Imported {updated + inserted} positions ({updated} updated, {inserted} new)."
        if zeroed:
            msg += f" {zeroed} stale holdings zeroed out."
        return jsonify({"message": msg, "updated": updated, "inserted": inserted, "zeroed": zeroed})

    except Exception as e:
        conn.rollback()
        return jsonify({"error": f"Import failed: {e}"}), 500
    finally:
        conn.close()


@app.route("/api/import/transactions", methods=["POST"])
def api_import_transactions():
    """Import transaction history from a CSV into the transactions + dividend_payments tables."""
    profile_id = get_profile_id()
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    f = request.files["file"]
    fmt = request.form.get("format", "snowball")
    if fmt not in TXN_PARSERS:
        return jsonify({"error": f"Unknown format: {fmt}"}), 400

    import uuid
    path = os.path.join(UPLOAD_FOLDER, f"txn_{uuid.uuid4().hex}_{f.filename}")
    f.save(path)
    try:
        parsed = TXN_PARSERS[fmt](path, f.filename)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        return jsonify({"error": f"Failed to parse file: {e}"}), 400
    finally:
        try:
            if os.path.exists(path):
                os.remove(path)
        except OSError:
            pass

    from datetime import date as _date, datetime as _dt

    if parsed.get("account_name"):
        profile_name = _get_profile_name(profile_id)
        account_match = _account_match_info(parsed.get("account_name"), profile_name, fmt)
        if not account_match["matched"]:
            return jsonify({"error": account_match["message"]}), 400

    # ── Positions-based import (e.g. Schwab) ─────────────────────────────────
    if parsed.get("format_type") == "positions":
        return _import_positions(parsed, profile_id)

    # ── Transaction-based import (e.g. Snowball) ─────────────────────────────
    conn = get_connection()
    preserve_positions = _profile_is_positions_managed(profile_id, conn)
    inserted_buys = 0
    inserted_sells = 0
    dividends_applied = 0
    duplicates_skipped = 0
    tickers_with_txns = set()
    tickers_with_divs = set()

    from collections import Counter

    # Pre-aggregate dividends by (ticker, date) so multiple distribution types
    # on the same day (Cash Div + Non-Qual Div, etc.) become a single row.
    from collections import defaultdict
    div_agg = defaultdict(lambda: {"amount": 0.0, "notes_parts": []})
    non_div_txns = []
    for txn in parsed["transactions"]:
        if txn["type"] == "DIVIDEND":
            key = (txn["ticker"], txn["date"])
            div_agg[key]["amount"] += txn["dividend_amount"] or 0
            if txn.get("notes"):
                div_agg[key]["notes_parts"].append(txn["notes"])
        else:
            non_div_txns.append(txn)

    try:
        # Insert aggregated dividends
        for (ticker, date_str), info in div_agg.items():
            dup = conn.execute(
                "SELECT 1 FROM dividend_payments WHERE ticker = ? AND profile_id = ? AND payment_date = ?",
                (ticker, profile_id, date_str),
            ).fetchone()
            if dup:
                duplicates_skipped += 1
                continue

            notes = "; ".join(dict.fromkeys(info["notes_parts"]))  # dedupe note parts
            conn.execute(
                "INSERT INTO dividend_payments (ticker, profile_id, payment_date, amount, source, notes) "
                "VALUES (?, ?, ?, ?, ?, ?)",
                (ticker, profile_id, date_str, round(info["amount"], 2), fmt, notes),
            )
            dividends_applied += 1
            tickers_with_divs.add(ticker)

        # Insert BUY/SELL transactions
        for txn in non_div_txns:
            ticker = txn["ticker"]
            # Count-based duplicate check so legitimate identical transactions
            # (same day/qty/price) are preserved while re-imports are deduped.
            # DB count includes uncommitted inserts from this batch (same conn).
            existing_count = conn.execute(
                "SELECT COUNT(*) FROM transactions WHERE ticker = ? AND profile_id = ? "
                "AND transaction_type = ? AND transaction_date = ? "
                "AND ABS(shares - ?) < 0.0001 AND price_per_share = ?",
                (ticker, profile_id, txn["type"], txn["date"],
                 txn["shares"], txn["price_per_share"] or 0),
            ).fetchone()[0]
            import_count = sum(1 for t in non_div_txns
                               if t["type"] == txn["type"] and t["ticker"] == ticker
                               and t["date"] == txn["date"]
                               and t["shares"] is not None
                               and abs(t["shares"] - txn["shares"]) < 0.0001
                               and (t["price_per_share"] or 0) == (txn["price_per_share"] or 0))
            if existing_count >= import_count:
                duplicates_skipped += 1
                continue

            # Ensure holding exists in all_account_info
            existing = conn.execute(
                "SELECT 1 FROM all_account_info WHERE ticker = ? AND profile_id = ?",
                (ticker, profile_id),
            ).fetchone()
            if not existing and not preserve_positions:
                conn.execute(
                    "INSERT INTO all_account_info (ticker, profile_id, description, import_date) "
                    "VALUES (?, ?, ?, ?)",
                    (ticker, profile_id, "", _date.today().isoformat()),
                )

            conn.execute(
                "INSERT INTO transactions (ticker, profile_id, transaction_type, transaction_date, "
                "shares, price_per_share, fees, notes) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                (ticker, profile_id, txn["type"], txn["date"],
                 txn["shares"], txn["price_per_share"], txn["fees"], txn.get("notes", "")),
            )

            if txn["type"] == "BUY":
                inserted_buys += 1
            else:
                inserted_sells += 1
            tickers_with_txns.add(ticker)

        conn.commit()

        # Rollup each ticker that had BUY/SELL inserts
        if not preserve_positions:
            for ticker in tickers_with_txns:
                _rollup_transactions(ticker, profile_id, conn)

        # Recompute dividend totals from dividend_payments for affected tickers
        current_year = str(_date.today().year)
        for ticker in tickers_with_divs:
            # Ensure holding exists (might only have dividend records, no buy/sell)
            existing = conn.execute(
                "SELECT 1 FROM all_account_info WHERE ticker = ? AND profile_id = ?",
                (ticker, profile_id),
            ).fetchone()
            if not existing:
                continue

            total = conn.execute(
                "SELECT COALESCE(SUM(amount), 0) FROM dividend_payments "
                "WHERE ticker = ? AND profile_id = ?",
                (ticker, profile_id),
            ).fetchone()[0]

            ytd = conn.execute(
                "SELECT COALESCE(SUM(amount), 0) FROM dividend_payments "
                "WHERE ticker = ? AND profile_id = ? AND payment_date LIKE ?",
                (ticker, profile_id, f"{current_year}%"),
            ).fetchone()[0]

            conn.execute(
                "UPDATE all_account_info SET total_divs_received = ?, ytd_divs = ?, "
                "paid_for_itself = CASE WHEN purchase_value > 0 THEN ? / purchase_value ELSE 0 END "
                "WHERE ticker = ? AND profile_id = ?",
                (total, ytd, total, ticker, profile_id),
            )

        conn.commit()

        # Run standard post-import chain
        populate_holdings(profile_id)
        populate_dividends(profile_id)
        populate_income_tracking(profile_id)

    except Exception as e:
        conn.close()
        return jsonify({"error": f"Import failed: {e}"}), 500
    finally:
        conn.close()

    return jsonify({
        "inserted_buys": inserted_buys,
        "inserted_sells": inserted_sells,
        "dividends_applied": dividends_applied,
        "duplicates_skipped": duplicates_skipped,
        "message": (
            f"Imported {inserted_buys} buys, {inserted_sells} sells, "
            f"{dividends_applied} dividends. {duplicates_skipped} duplicates skipped."
            + (
                " Current holdings were preserved from the broker positions import for this portfolio."
                if preserve_positions else ""
            )
        ),
    })


# ── Ticker Lookup (yfinance) ────────────────────────────────────────────────────

@app.route("/api/lookup/<ticker>", methods=["GET"])
def lookup_ticker(ticker):
    """Fetch current price, dividend info, and description from Yahoo Finance."""
    import yfinance as yf
    from datetime import datetime as _dt

    ticker = ticker.strip().upper()
    result = {
        "ticker": ticker,
        "description": ticker,
        "classification_type": "ETF",
        "current_price": 0,
        "div": 0,
        "div_frequency": "Q",
        "ex_div_date": None,
        "div_pay_date": None,
        "dividend_paid": 0,
        "ytd_divs": 0,
        "total_divs_received": 0,
        "paid_for_itself": 0,
    }

    try:
        tk = yf.Ticker(ticker)
        info = tk.info or {}

        # Detect ticker rename (e.g. TOPW → WPAY)
        actual_symbol = info.get("symbol", "").upper()
        if actual_symbol and actual_symbol != ticker:
            # yfinance resolved this to a different ticker — fetch data under the new symbol
            tk = yf.Ticker(actual_symbol)
            info = tk.info or {}
            result["renamed_to"] = actual_symbol

        result["description"] = (info.get("longName") or info.get("shortName") or ticker)[:200]
        result["classification_type"] = (info.get("quoteType") or "ETF")[:20]
        result["current_price"] = info.get("regularMarketPrice") or info.get("currentPrice") or 0

        # Infer frequency and div from dividend history
        freq_code = "Q"
        try:
            hist = tk.dividends
            if hist is not None and len(hist) > 0:
                # Match timezone awareness of the index
                if hist.index.tz is not None:
                    one_year_ago = pd.Timestamp.now(tz=hist.index.tz) - pd.Timedelta(days=365)
                else:
                    one_year_ago = pd.Timestamp.now() - pd.Timedelta(days=365)
                recent = hist[hist.index >= one_year_ago]
                n = len(recent[recent > 0])
                if n >= 45:
                    freq_code = "W"
                elif n >= 10:
                    freq_code = "M"
                elif n >= 3:
                    freq_code = "Q"
                elif n >= 2:
                    freq_code = "SA"
                elif n >= 1:
                    freq_code = "A"

                # Last dividend per share (actual payment, not annual rate)
                last_div = recent[recent > 0]
                if not last_div.empty:
                    result["div"] = round(float(last_div.iloc[-1]), 6)

                # Ex-div date from last payment
                result["ex_div_date"] = last_div.index[-1].strftime("%m/%d/%y") if not last_div.empty else None
        except Exception:
            # Fallback to info fields
            annual_rate = info.get("dividendRate") or 0
            if annual_rate:
                result["div"] = round(annual_rate, 6)
            ex_ts = info.get("exDividendDate")
            if ex_ts:
                try:
                    result["ex_div_date"] = _dt.utcfromtimestamp(ex_ts).strftime("%m/%d/%y")
                except Exception:
                    pass

        result["div_frequency"] = freq_code

        # Dividend pay date — try yfinance calendar, else estimate from ex-div + 2-4 weeks
        try:
            cal = tk.calendar
            if cal is not None:
                # yfinance returns calendar as dict or DataFrame depending on version
                pay_date = None
                if isinstance(cal, dict):
                    pay_date = cal.get("Dividend Date") or cal.get("Payment Date")
                elif hasattr(cal, "loc"):
                    for key in ["Dividend Date", "Payment Date"]:
                        if key in cal.index:
                            pay_date = cal.loc[key].iloc[0] if hasattr(cal.loc[key], "iloc") else cal.loc[key]
                            break
                if pay_date is not None:
                    if hasattr(pay_date, "strftime"):
                        result["div_pay_date"] = pay_date.strftime("%m/%d/%y")
                    elif isinstance(pay_date, (int, float)):
                        result["div_pay_date"] = _dt.fromtimestamp(pay_date, tz=__import__('datetime').timezone.utc).strftime("%m/%d/%y")
        except Exception:
            pass

        # Estimate: ex-div + 3 weeks if we have ex-div but no pay date
        if not result["div_pay_date"] and result["ex_div_date"]:
            try:
                ex = _dt.strptime(result["ex_div_date"], "%m/%d/%y")
                est_pay = ex + pd.Timedelta(days=21)
                result["div_pay_date"] = est_pay.strftime("%m/%d/%y")
            except Exception:
                pass

    except Exception as e:
        return jsonify({"error": f"Could not look up {ticker}: {str(e)}"}), 404

    return jsonify(result)


def _parse_timestamp_value(value):
    """Best-effort conversion of a date-like value to pandas.Timestamp."""
    if value in (None, ""):
        return None
    try:
        ts = pd.Timestamp(value)
        if ts.tzinfo is not None:
            ts = ts.tz_localize(None)
        return ts
    except Exception:
        return None


def _format_mdy_date(value):
    """Format a date-like value as MM/DD/YY, or None when unavailable."""
    ts = _parse_timestamp_value(value)
    return ts.strftime("%m/%d/%y") if ts is not None else None


def _infer_dividend_frequency_from_count(count):
    """Infer dividend frequency code from the number of positive payments in ~1 year."""
    if count >= 45:
        return "W"
    if count >= 10:
        return "M"
    if count >= 3:
        return "Q"
    if count >= 2:
        return "SA"
    if count >= 1:
        return "A"
    return None


def _fetch_refresh_dividend_snapshot(yf_ticker, preferred_freq=None):
    """Fetch dividend metadata for refresh when the batch download is incomplete."""
    from datetime import timezone as _timezone

    snapshot = {
        "known": True,
        "has_dividend": False,
        "div": 0.0,
        "ex_div_date": None,
        "div_pay_date": None,
        "freq": None,
        "history": pd.Series(dtype=float),
    }

    preferred = (preferred_freq or "").strip().upper() or None
    info = {}
    try:
        info = yf_ticker.info or {}
    except Exception:
        info = {}

    try:
        hist = yf_ticker.dividends
    except Exception:
        hist = None

    recent = pd.Series(dtype=float)
    if hist is not None and len(hist) > 0:
        try:
            if getattr(hist.index, "tz", None) is not None:
                one_year_ago = pd.Timestamp.now(tz=hist.index.tz) - pd.Timedelta(days=365)
            else:
                one_year_ago = pd.Timestamp.now() - pd.Timedelta(days=365)
            recent = hist[hist.index >= one_year_ago]
            recent = recent[recent > 0].dropna()
        except Exception:
            recent = pd.Series(dtype=float)

    if not recent.empty:
        snapshot["has_dividend"] = True
        snapshot["div"] = float(recent.iloc[-1])
        snapshot["ex_div_date"] = _format_mdy_date(recent.index[-1])
        snapshot["freq"] = _infer_dividend_frequency_from_count(len(recent))
        snapshot["history"] = recent
    else:
        annual_rate = info.get("dividendRate") or 0
        ex_ts = info.get("exDividendDate")
        if annual_rate and annual_rate > 0:
            freq = preferred or "Q"
            mult = {"W": 52, "52": 52, "M": 12, "Q": 4, "SA": 2, "A": 1}.get(freq, 4)
            snapshot["has_dividend"] = True
            snapshot["freq"] = freq
            snapshot["div"] = float(annual_rate) / mult if mult else float(annual_rate)
        if ex_ts:
            snapshot["has_dividend"] = True
            if isinstance(ex_ts, (int, float)):
                ex_ts = pd.Timestamp(ex_ts, unit="s", tz=_timezone.utc)
            snapshot["ex_div_date"] = _format_mdy_date(ex_ts)

    try:
        cal = yf_ticker.calendar
        pay_date = None
        if isinstance(cal, dict):
            pay_date = cal.get("Dividend Date") or cal.get("Payment Date")
        elif hasattr(cal, "loc"):
            for key in ["Dividend Date", "Payment Date"]:
                if key in cal.index:
                    pay_date = cal.loc[key].iloc[0] if hasattr(cal.loc[key], "iloc") else cal.loc[key]
                    break
        if isinstance(pay_date, (int, float)):
            pay_date = pd.Timestamp(pay_date, unit="s", tz=_timezone.utc)
        snapshot["div_pay_date"] = _format_mdy_date(pay_date)
    except Exception:
        pass

    if not snapshot["div_pay_date"] and snapshot["ex_div_date"]:
        ex_ts = _parse_timestamp_value(snapshot["ex_div_date"])
        if ex_ts is not None:
            snapshot["div_pay_date"] = (ex_ts + pd.Timedelta(days=21)).strftime("%m/%d/%y")

    if snapshot["has_dividend"] and not snapshot["freq"]:
        snapshot["freq"] = preferred or "Q"

    return snapshot


def _simulate_drip_refresh(div_series, close_series, base_qty, start_date=None, today=None):
    """Simulate post-start DRIP activity and actual dividend cash by dividend date."""
    if div_series is None or div_series.empty or close_series is None or close_series.empty:
        return {
            "quantity": float(base_qty or 0),
            "drip_shares": 0.0,
            "drip_total_divs": 0.0,
            "ytd_divs": 0.0,
            "current_month_income": 0.0,
            "event_count": 0,
        }

    today_ts = _parse_timestamp_value(today) or pd.Timestamp.now().normalize()
    start_ts = _parse_timestamp_value(start_date)
    jan1 = pd.Timestamp(today_ts.year, 1, 1)
    first_of_month = pd.Timestamp(today_ts.year, today_ts.month, 1)
    ytd_floor = max(jan1, start_ts) if start_ts is not None else jan1
    month_floor = max(first_of_month, ytd_floor)

    eligible = div_series[div_series.index <= today_ts]
    if start_ts is not None:
        eligible = eligible[eligible.index > start_ts]

    running_qty = float(base_qty or 0)
    drip_shares = 0.0
    drip_total_divs = 0.0
    ytd_total = 0.0
    current_total = 0.0
    event_count = 0

    for div_date, div_amt in eligible.items():
        prices_on_or_before = close_series[close_series.index <= div_date]
        if prices_on_or_before.empty:
            continue
        reinvest_price = float(prices_on_or_before.iloc[-1])
        if reinvest_price <= 0:
            continue

        div_income = float(div_amt) * running_qty
        if div_date >= ytd_floor:
            ytd_total += div_income
        if div_date >= month_floor:
            current_total += div_income

        new_shares = div_income / reinvest_price
        drip_shares += new_shares
        drip_total_divs += div_income
        running_qty += new_shares
        event_count += 1

    return {
        "quantity": running_qty,
        "drip_shares": drip_shares,
        "drip_total_divs": drip_total_divs,
        "ytd_divs": ytd_total,
        "current_month_income": current_total,
        "event_count": event_count,
    }


# ── Refresh Market Data ─────────────────────────────────────────────────────────

@app.route("/api/refresh", methods=["POST"])
def refresh_market_data():
    """Update current price, div/share, ex-div date, and frequency for all holdings from Yahoo Finance."""
    import yfinance as yf
    from datetime import datetime as _dt

    conn = get_connection()
    refresh_info = _get_refresh_target_info(conn)
    scope = refresh_info["scope"]
    source_pids = refresh_info["source_profile_ids"]
    selected_profile_id = refresh_info["selected_profile_id"]

    # Refresh only the profiles relevant to the current holdings view:
    # - single profile: just that profile
    # - Owner: source accounts that feed Owner
    # - Aggregate: aggregate member profiles
    all_rows = conn.execute(
        "SELECT profile_id, ticker, quantity, price_paid, purchase_value, purchase_date, reinvest, base_quantity, import_date FROM all_account_info WHERE profile_id IN ({})".format(
            ",".join("?" * len(source_pids))
        ), source_pids,
    ).fetchall()

    if not all_rows:
        conn.close()
        return jsonify({"updated": 0, "message": "No holdings to refresh"})

    tickers = list({r["ticker"] for r in all_rows})
    # Build per-profile holding map
    holding_map = {}
    for r in all_rows:
        holding_map[(r["profile_id"], r["ticker"])] = {
            "qty": r["quantity"] or 0,
            "price_paid": r["price_paid"] or 0,
            "purchase_value": r["purchase_value"] or 0,
            "purchase_date": r["purchase_date"] or "",
            "reinvest": (r["reinvest"] or "").upper(),
            "base_quantity": r["base_quantity"] or r["quantity"] or 0,
            "import_date": r["import_date"] or "",
        }

    # Batch download prices + dividends (one yfinance call for all tickers)
    ticker_str = " ".join(tickers)
    price_map = {}
    div_map = {}
    exdiv_map = {}
    freq_map = {}
    div_history = {}    # ticker -> pandas Series of positive dividends (index=dates)
    close_history = {}  # ticker -> pandas Series of close prices (index=dates)
    div_snapshot_map = {}
    raw = pd.DataFrame()

    try:
        raw = yf.download(ticker_str, period="1y", progress=False, auto_adjust=False, actions=True)
        if not raw.empty:
            def _col(name):
                if isinstance(raw.columns, pd.MultiIndex):
                    return raw[name] if name in raw.columns.get_level_values(0) else None
                return raw[name] if name in raw.columns else None

            # Prices
            close = _col("Close")
            if close is not None:
                if isinstance(close, pd.Series):
                    s = close.dropna()
                    if len(s):
                        price_map[tickers[0]] = float(s.iloc[-1])
                        close_history[tickers[0]] = s
                else:
                    for t in tickers:
                        if t in close.columns:
                            s = close[t].dropna()
                            if len(s):
                                price_map[t] = float(s.iloc[-1])
                                close_history[t] = s

            # Dividends
            divs = _col("Dividends")
            if divs is not None:
                if isinstance(divs, pd.Series):
                    d = divs[divs > 0].dropna()
                    if not d.empty:
                        t0 = tickers[0]
                        div_map[t0] = float(d.iloc[-1])
                        exdiv_map[t0] = d.index[-1].strftime("%m/%d/%y")
                        n = len(d)
                        freq_map[t0] = "W" if n >= 45 else "M" if n >= 10 else "Q" if n >= 3 else "SA" if n >= 2 else "A"
                        div_history[t0] = d
                else:
                    for t in tickers:
                        if t in divs.columns:
                            d = divs[t][divs[t] > 0].dropna()
                            if not d.empty:
                                div_map[t] = float(d.iloc[-1])
                                exdiv_map[t] = d.index[-1].strftime("%m/%d/%y")
                                n = len(d)
                                freq_map[t] = "W" if n >= 45 else "M" if n >= 10 else "Q" if n >= 3 else "SA" if n >= 2 else "A"
                                div_history[t] = d
    except Exception:
        pass

    # Detect renamed tickers: if a ticker got no price data from batch download,
    # check if yfinance maps it to a new symbol and pull data under that column.
    rename_map = {}  # old_ticker -> new_ticker
    for t in tickers:
        if t not in price_map:
            try:
                info = yf.Ticker(t).info or {}
                new_sym = (info.get("symbol") or "").upper()
                if new_sym and new_sym != t:
                    rename_map[t] = new_sym
                    # Try to pull data from the new symbol's column in the batch download
                    if not raw.empty and isinstance(raw.columns, pd.MultiIndex):
                        close = raw["Close"] if "Close" in raw.columns.get_level_values(0) else None
                        if close is not None and not isinstance(close, pd.Series) and new_sym in close.columns:
                            s = close[new_sym].dropna()
                            if len(s):
                                price_map[t] = float(s.iloc[-1])
                                close_history[t] = s
                        divs = raw["Dividends"] if "Dividends" in raw.columns.get_level_values(0) else None
                        if divs is not None and not isinstance(divs, pd.Series) and new_sym in divs.columns:
                            d = divs[new_sym][divs[new_sym] > 0].dropna()
                            if not d.empty:
                                div_map[t] = float(d.iloc[-1])
                                exdiv_map[t] = d.index[-1].strftime("%m/%d/%y")
                                n = len(d)
                                freq_map[t] = "W" if n >= 45 else "M" if n >= 10 else "Q" if n >= 3 else "SA" if n >= 2 else "A"
                                div_history[t] = d
                    # If still no data, try a separate download for the new symbol
                    if t not in price_map:
                        try:
                            r2 = yf.download(new_sym, period="1y", progress=False, auto_adjust=False, actions=True)
                            if not r2.empty:
                                c2 = r2["Close"] if "Close" in r2.columns else (r2["Close"][new_sym] if isinstance(r2.columns, pd.MultiIndex) else None)
                                if c2 is not None:
                                    s2 = c2.squeeze().dropna()
                                    if len(s2):
                                        price_map[t] = float(s2.iloc[-1])
                                        close_history[t] = s2
                                d2 = r2["Dividends"] if "Dividends" in r2.columns else (r2["Dividends"][new_sym] if isinstance(r2.columns, pd.MultiIndex) else None)
                                if d2 is not None:
                                    d2s = d2.squeeze()
                                    d2s = d2s[d2s > 0].dropna()
                                    if not d2s.empty:
                                        div_map[t] = float(d2s.iloc[-1])
                                        exdiv_map[t] = d2s.index[-1].strftime("%m/%d/%y")
                                        n = len(d2s)
                                        freq_map[t] = "W" if n >= 45 else "M" if n >= 10 else "Q" if n >= 3 else "SA" if n >= 2 else "A"
                                        div_history[t] = d2s
                        except Exception:
                            pass
                    # Update description if it matches the ticker (i.e., never got enriched)
                    if new_sym:
                        try:
                            new_info = yf.Ticker(new_sym).info or {}
                            new_desc = new_info.get("longName") or new_info.get("shortName")
                            if new_desc:
                                for pid in source_pids:
                                    conn.execute(
                                        "UPDATE all_account_info SET description = ? WHERE ticker = ? AND profile_id = ? AND (description = ? OR description IS NULL OR description = '')",
                                        (new_desc[:200], t, pid, t),
                                    )
                        except Exception:
                            pass
            except Exception:
                pass

    # Load known weekly tickers across all profiles
    weekly_set = set()
    try:
        wrows = conn.execute("SELECT DISTINCT ticker FROM weekly_payout_tickers").fetchall()
        weekly_set = {r["ticker"] for r in wrows}
    except Exception:
        pass

    # Load current DB frequencies across all profiles (use highest rank per ticker)
    db_freq_map = {}
    try:
        frows = conn.execute("SELECT ticker, div_frequency FROM all_account_info").fetchall()
        freq_rank = {'W': 6, '52': 6, 'M': 5, 'Q': 4, 'SA': 3, 'A': 2, None: 0}
        for r in frows:
            t, f = r["ticker"], r["div_frequency"]
            if freq_rank.get(f, 0) > freq_rank.get(db_freq_map.get(t), 0):
                db_freq_map[t] = f
    except Exception:
        pass

    # Resolve effective frequency per ticker (once, shared across profiles)
    freq_rank = {'W': 6, '52': 6, 'M': 5, 'Q': 4, 'SA': 3, 'A': 2, None: 0}
    effective_freq = {}
    for t in tickers:
        if t in div_history and div_map.get(t) is not None:
            div_snapshot_map[t] = {
                "known": True,
                "has_dividend": True,
                "div": div_map.get(t) or 0.0,
                "ex_div_date": exdiv_map.get(t),
                "div_pay_date": _format_mdy_date(
                    (_parse_timestamp_value(exdiv_map.get(t)) + pd.Timedelta(days=21))
                    if exdiv_map.get(t) else None
                ),
                "freq": freq_map.get(t),
                "history": div_history.get(t, pd.Series(dtype=float)),
            }
        elif t in close_history:
            # Batch download returned valid market data for this ticker and
            # showed no positive dividends in the period, so treat that as a
            # definitive "no current dividend" signal instead of leaving stale
            # dividend metadata behind due to a flaky per-ticker fallback call.
            div_snapshot_map[t] = {
                "known": True,
                "has_dividend": False,
                "div": 0.0,
                "ex_div_date": None,
                "div_pay_date": None,
                "freq": None,
                "history": pd.Series(dtype=float),
            }
        else:
            preferred_freq = 'W' if t in weekly_set else db_freq_map.get(t)
            lookup_symbol = rename_map.get(t, t)
            try:
                div_snapshot_map[t] = _fetch_refresh_dividend_snapshot(
                    yf.Ticker(lookup_symbol),
                    preferred_freq=preferred_freq,
                )
            except Exception:
                div_snapshot_map[t] = {
                    "known": False,
                    "has_dividend": False,
                    "div": 0.0,
                    "ex_div_date": None,
                    "div_pay_date": None,
                    "freq": None,
                    "history": pd.Series(dtype=float),
                }

        snapshot = div_snapshot_map[t]
        if snapshot.get("has_dividend"):
            nf = snapshot.get("freq")
            if t in weekly_set:
                nf = 'W'
            else:
                db_rank = freq_rank.get(db_freq_map.get(t), 0)
                new_rank = freq_rank.get(nf, 0)
                if new_rank < db_rank:
                    nf = db_freq_map.get(t)
            effective_freq[t] = nf
        else:
            effective_freq[t] = None

    updated = 0
    updated_pids = set()
    refreshed_owner_source = False
    owner_source_ids = set(_get_owner_source_profile_ids(conn))
    for pid in source_pids:
        if pid in owner_source_ids:
            refreshed_owner_source = True
        for t in tickers:
            key = (pid, t)
            if key not in holding_map:
                continue

            snapshot = div_snapshot_map.get(t, {
                "known": False,
                "has_dividend": False,
                "div": 0.0,
                "ex_div_date": None,
                "div_pay_date": None,
                "freq": None,
                "history": pd.Series(dtype=float),
            })
            snapshot_known = bool(snapshot.get("known"))
            has_dividend = bool(snapshot.get("has_dividend"))
            new_price = price_map.get(t)
            new_div = snapshot.get("div") if has_dividend else None
            new_exdiv = snapshot.get("ex_div_date") if has_dividend else None
            new_pay_date = snapshot.get("div_pay_date") if has_dividend else None
            new_freq = effective_freq.get(t)

            if not new_price and not snapshot_known:
                continue

            h = holding_map[key]
            qty = h["qty"]
            price_paid = h["price_paid"]
            purchase_value = h["purchase_value"]
            purchase_date = h["purchase_date"]
            base_qty = h["base_quantity"]
            is_drip = h["reinvest"] == "Y"
            import_date = h["import_date"]
            sets = []
            vals = []

            # Backfill purchase_value if missing
            if not purchase_value and price_paid and qty:
                purchase_value = qty * price_paid
                sets.append("purchase_value = ?")
                vals.append(round(purchase_value, 2))

            # ── DRIP simulation ──────────────────────────────────────────
            # For reinvest=Y holdings, simulate dividend reinvestment from
            # import_date forward using actual dividend history + close prices.
            # base_quantity already includes all DRIP through the import date
            # (whether imported from spreadsheet or entered via form), so we
            # only add shares for dividends received AFTER that date.
            drip_shares = 0.0
            drip_total_divs = 0.0
            ytd_divs = 0.0
            current_month_income = 0.0
            div_series = snapshot.get("history")
            close_series = close_history.get(t)

            if (
                is_drip
                and div_series is not None
                and not div_series.empty
                and close_series is not None
            ):
                drip_start = import_date or purchase_date or None
                drip_result = _simulate_drip_refresh(
                    div_series,
                    close_series,
                    base_qty,
                    start_date=drip_start,
                )
                qty = drip_result["quantity"]
                drip_shares = drip_result["drip_shares"]
                drip_total_divs = drip_result["drip_total_divs"]
                ytd_divs = drip_result["ytd_divs"]
                current_month_income = drip_result["current_month_income"]
                if drip_result["event_count"] > 0:
                    sets.extend(["quantity = ?", "base_quantity = ?",
                                 "shares_bought_from_dividend = ?",
                                 "total_cash_reinvested = ?"])
                    vals.extend([round(qty, 6), base_qty,
                                 round(drip_shares, 6),
                                 round(drip_total_divs, 2)])
                    purchase_value = (base_qty * price_paid) + drip_total_divs
                    sets.append("purchase_value = ?")
                    vals.append(round(purchase_value, 2))
                elif qty != base_qty:
                    # No new DRIP dividends since import — reset to base
                    qty = base_qty
                    sets.extend(["quantity = ?", "base_quantity = ?",
                                 "shares_bought_from_dividend = ?",
                                 "total_cash_reinvested = ?"])
                    vals.extend([base_qty, base_qty, 0, 0])
                    purchase_value = base_qty * price_paid
                    sets.append("purchase_value = ?")
                    vals.append(round(purchase_value, 2))
            elif not is_drip and qty != base_qty:
                # reinvest changed from Y to N — clear accumulated DRIP shares
                qty = base_qty
                sets.extend(["quantity = ?", "shares_bought_from_dividend = ?",
                             "total_cash_reinvested = ?"])
                vals.extend([round(base_qty, 6), 0, 0])
                purchase_value = base_qty * price_paid
                sets.append("purchase_value = ?")
                vals.append(round(purchase_value, 2))

            if new_price:
                current_value = new_price * qty
                gain = current_value - purchase_value if purchase_value else 0
                gain_pct = (gain / purchase_value) if purchase_value else 0
                sets.extend(["current_price = ?", "current_value = ?", "gain_or_loss = ?",
                             "gain_or_loss_percentage = ?", "percent_change = ?"])
                vals.extend([new_price, current_value, gain, gain_pct, gain_pct])

            if snapshot_known:
                if new_div:
                    freq_mult = {'W': 52, '52': 52, 'M': 12, 'Q': 4, 'SA': 2, 'A': 1}
                    cur_freq = (new_freq or 'Q').upper()
                    mult = freq_mult.get(cur_freq, 4)
                    annual_div = new_div * mult
                    yoc = (annual_div / price_paid) if price_paid else 0
                    cur_yield = (annual_div / new_price) if new_price else 0
                    estim_annual = new_div * qty * mult
                    estim_monthly = estim_annual / 12 if estim_annual else 0
                    sets.extend(["div = ?", "annual_yield_on_cost = ?", "current_annual_yield = ?",
                                 "estim_payment_per_year = ?", "approx_monthly_income = ?"])
                    vals.extend([new_div, yoc, cur_yield, estim_annual, estim_monthly])
                else:
                    sets.extend(["div = ?", "annual_yield_on_cost = ?", "current_annual_yield = ?",
                                 "estim_payment_per_year = ?", "approx_monthly_income = ?"])
                    vals.extend([0, 0, 0, 0, 0])

                sets.append("ex_div_date = ?")
                vals.append(new_exdiv)
                sets.append("div_pay_date = ?")
                vals.append(new_pay_date)

                sets.append("div_frequency = ?")
                vals.append(new_freq)

            # Actual YTD and current-month dividends from yfinance history,
            # filtered to only include dividends on or after purchase_date.
            # Uses DRIP-adjusted qty for accurate income calculation.
            if snapshot_known and div_series is not None and not div_series.empty and not is_drip:
                jan1 = pd.Timestamp(_dt.now().year, 1, 1)
                today = pd.Timestamp(_dt.now().date())
                first_of_month = pd.Timestamp(today.year, today.month, 1)
                start = jan1
                if purchase_date:
                    try:
                        pdt = pd.Timestamp(purchase_date)
                        if pdt > start:
                            start = pdt
                    except Exception:
                        pass
                ytd = div_series[(div_series.index >= start) & (div_series.index <= today)]
                curmo_start = max(first_of_month, start)
                curmo = div_series[(div_series.index >= curmo_start) & (div_series.index <= today)]
                ytd_divs = round(float(ytd.sum()) * qty, 2) if not ytd.empty else 0
                current_month_income = round(float(curmo.sum()) * qty, 2) if not curmo.empty else 0

            if snapshot_known:
                sets.append("ytd_divs = ?")
                vals.append(round(ytd_divs, 2))
                sets.append("current_month_income = ?")
                vals.append(round(current_month_income, 2))

            if sets:
                vals.extend([t, pid])
                conn.execute(
                    f"UPDATE all_account_info SET {', '.join(sets)} WHERE ticker = ? AND profile_id = ?",
                    vals,
                )
                updated += 1
                updated_pids.add(pid)

    conn.commit()

    # Update derived tables for all affected profiles
    for pid in updated_pids:
        populate_holdings(pid)
        populate_dividends(pid)

    conn.close()

    # Sync Owner only when the refresh touched an Owner source account,
    # or when the Owner page itself initiated the refresh.
    if scope == "owner" or refreshed_owner_source:
        _auto_reconcile_owner()

    refreshed_rows = sum(1 for key in holding_map if key[0] in source_pids)
    if scope == "owner":
        msg = f"Refreshed {refreshed_rows} holdings across {len(source_pids)} Owner source portfolios."
    elif scope == "aggregate":
        msg = f"Refreshed {refreshed_rows} holdings across {len(source_pids)} Aggregate member portfolios."
    else:
        msg = f"Refreshed {refreshed_rows} holdings in {_get_profile_name(selected_profile_id)}."
    return jsonify({"updated": updated, "message": msg})


# ── Holdings CRUD ──────────────────────────────────────────────────────────────

@app.route("/api/holdings", methods=["GET"])
def list_holdings():
    is_agg, pids = get_profile_filter()
    conn = get_connection()
    placeholders = ",".join("?" * len(pids))

    if is_agg and len(pids) > 1:
        # Check if Owner import was used — if so, prefer Owner's per-ticker
        # income/dividend data since the Owner spreadsheet is authoritative.
        # For generic-only imports (no Owner), just SUM across sub-profiles.
        _oiu = conn.execute(
            "SELECT value FROM settings WHERE key = 'owner_import_used'"
        ).fetchone()
        use_owner = _oiu and _oiu[0] == "true"

        def _own_or_sum(field):
            """COALESCE from Owner if Owner import exists, else plain SUM."""
            if use_owner:
                return (f"COALESCE((SELECT o.{field} FROM all_account_info o "
                        f"WHERE o.ticker = a.ticker AND o.profile_id = 1), SUM(a.{field}))")
            return f"SUM(a.{field})"

        # Income fields are recalculated by refresh for all profiles, so just SUM.
        # Payout history fields (ytd, total divs, current month) only exist in
        # Owner, so COALESCE from Owner when available.
        ytd = _own_or_sum("ytd_divs")
        tot_div = _own_or_sum("total_divs_received")
        cur_mo = _own_or_sum("current_month_income")

        # Aggregate: combine duplicate tickers across portfolios
        rows = conn.execute(
            f"""SELECT
                   a.ticker,
                   MAX(a.description) as description,
                   MAX(a.classification_type) as classification_type,
                   SUM(a.quantity) as quantity,
                   CASE WHEN SUM(a.quantity) > 0 THEN SUM(a.purchase_value) / SUM(a.quantity) ELSE 0 END as price_paid,
                   MAX(a.current_price) as current_price,
                   SUM(a.purchase_value) as purchase_value,
                   SUM(a.current_value) as current_value,
                   SUM(a.gain_or_loss) as gain_or_loss,
                   CASE WHEN SUM(a.purchase_value) > 0 THEN SUM(a.gain_or_loss) / SUM(a.purchase_value) ELSE 0 END as gain_or_loss_percentage,
                   CASE WHEN SUM(a.purchase_value) > 0 THEN SUM(a.gain_or_loss) / SUM(a.purchase_value) ELSE 0 END as percent_change,
                   MAX(a.div_frequency) as div_frequency,
                   MAX(a.reinvest) as reinvest,
                   MAX(a.ex_div_date) as ex_div_date,
                   MAX(a.div_pay_date) as div_pay_date,
                   CASE WHEN SUM(a.quantity) > 0 THEN SUM(a.dividend_paid) / SUM(a.quantity) ELSE MAX(a.div) END as div,
                   SUM(a.dividend_paid) as dividend_paid,
                   SUM(a.estim_payment_per_year) as estim_payment_per_year,
                   SUM(a.approx_monthly_income) as approx_monthly_income,
                   SUM(CASE WHEN a.reinvest = 'Y' THEN a.approx_monthly_income ELSE 0 END) as monthly_income_reinvested,
                   SUM(CASE WHEN a.reinvest != 'Y' OR a.reinvest IS NULL THEN a.approx_monthly_income ELSE 0 END) as monthly_income_not_reinvested,
                   SUM(a.withdraw_8pct_cost_annually) as withdraw_8pct_cost_annually,
                   SUM(a.withdraw_8pct_per_month) as withdraw_8pct_per_month,
                   SUM(a.cash_not_reinvested) as cash_not_reinvested,
                   SUM(a.total_cash_reinvested) as total_cash_reinvested,
                   CASE WHEN SUM(a.purchase_value) > 0 THEN SUM(a.estim_payment_per_year) / SUM(a.purchase_value) ELSE 0 END as annual_yield_on_cost,
                   CASE WHEN SUM(a.current_value) > 0 THEN SUM(a.estim_payment_per_year) / SUM(a.current_value) ELSE 0 END as current_annual_yield,
                   NULL as percent_of_account,
                   SUM(a.shares_bought_from_dividend) as shares_bought_from_dividend,
                   SUM(a.shares_bought_in_year) as shares_bought_in_year,
                   SUM(a.shares_in_month) as shares_in_month,
                   {ytd} as ytd_divs,
                   {tot_div} as total_divs_received,
                   CASE WHEN SUM(a.purchase_value) > 0 THEN {tot_div} / SUM(a.purchase_value) ELSE 0 END as paid_for_itself,
                   MAX(a.import_date) as import_date,
                   MIN(a.purchase_date) as purchase_date,
                   {cur_mo} as current_month_income,
                   (SELECT c2.name FROM ticker_categories tc2
                    JOIN categories c2 ON tc2.category_id = c2.id
                    WHERE tc2.ticker = a.ticker AND tc2.profile_id = (
                        SELECT a2.profile_id FROM all_account_info a2
                        WHERE a2.ticker = a.ticker AND a2.profile_id IN ({placeholders})
                        ORDER BY a2.quantity DESC LIMIT 1
                    ) LIMIT 1) as category
               FROM all_account_info a
               WHERE a.profile_id IN ({placeholders})
               GROUP BY a.ticker
               ORDER BY a.ticker""",
            pids + pids,
        ).fetchall()
    else:
        pid = pids[0]
        rows = conn.execute(
            """SELECT a.*,
                      (SELECT c.name FROM ticker_categories tc
                       JOIN categories c ON tc.category_id = c.id
                       WHERE tc.ticker = a.ticker AND tc.profile_id = a.profile_id
                       LIMIT 1) AS category
               FROM all_account_info a
               WHERE a.profile_id = ?
               ORDER BY a.ticker""",
            (pid,),
        ).fetchall()

    # Recalculate percent_of_account
    results = rows_to_dicts(rows)
    total_value = sum(r.get("current_value") or 0 for r in results)
    if total_value > 0:
        for r in results:
            r["percent_of_account"] = (r.get("current_value") or 0) / total_value

    # Use stored actuals from refresh if available, otherwise estimate.
    # IMPORTANT: use "is None" not falsy check — 0 is a valid actual value
    # (e.g. no dividends received yet this month) and should NOT trigger
    # the full-month estimate fallback.
    for r in results:
        if r.get("current_month_income") is None:
            r["current_month_income"] = _estimate_current_month_income(r)
        if r.get("ytd_divs") is None:
            r["ytd_divs"] = _estimate_ytd_income(r)
    # For single-profile queries, compute reinvested/not-reinvested splits.
    # Owner (profile_id=1) uses sub-account DRIP ratios since the Owner
    # flag may be stale.  Sub-accounts use their own flag directly.
    if not is_agg or len(pids) <= 1:
        pid = pids[0]
        if pid == 1:
            # Owner: derive split from sub-accounts flagged as part of Owner
            member_rows = conn.execute(
                "SELECT id FROM profiles WHERE include_in_owner = 1 AND id != 1"
            ).fetchall()
            member_ids = [r["id"] if isinstance(r, dict) else r[0]
                          for r in member_rows]

            if member_ids:
                mph = ",".join("?" * len(member_ids))
                drip_rows = conn.execute(
                    f"""SELECT ticker,
                           SUM(CASE WHEN reinvest = 'Y' THEN approx_monthly_income ELSE 0 END) as ri,
                           SUM(approx_monthly_income) as ti
                        FROM all_account_info
                        WHERE profile_id IN ({mph})
                        GROUP BY ticker""",
                    member_ids,
                ).fetchall()
                drip_map = {}
                for dr in drip_rows:
                    t = dr["ticker"] if isinstance(dr, dict) else dr[0]
                    ri = (dr["ri"] if isinstance(dr, dict) else dr[1]) or 0
                    ti = (dr["ti"] if isinstance(dr, dict) else dr[2]) or 0
                    drip_map[t] = ri / ti if ti > 0 else 0

                for r in results:
                    if r.get("monthly_income_reinvested") is None:
                        mi = r.get("approx_monthly_income") or 0
                        pct = drip_map.get(r["ticker"], 0)
                        r["monthly_income_reinvested"] = round(mi * pct, 2)
                        r["monthly_income_not_reinvested"] = round(mi * (1 - pct), 2)
            else:
                for r in results:
                    if r.get("monthly_income_reinvested") is None:
                        mi = r.get("approx_monthly_income") or 0
                        r["monthly_income_reinvested"] = 0
                        r["monthly_income_not_reinvested"] = mi
        else:
            # Sub-account: use its own DRIP flag directly
            for r in results:
                if r.get("monthly_income_reinvested") is None:
                    mi = r.get("approx_monthly_income") or 0
                    if r.get("reinvest") == "Y":
                        r["monthly_income_reinvested"] = mi
                        r["monthly_income_not_reinvested"] = 0
                    else:
                        r["monthly_income_reinvested"] = 0
                        r["monthly_income_not_reinvested"] = mi

    conn.close()
    return jsonify(results)


@app.route("/api/holdings", methods=["POST"])
def add_holding():
    """Add a single holding manually."""
    profile_id = get_profile_id()
    data = request.get_json()
    ticker = (data.get("ticker") or "").strip().upper()
    if not ticker:
        return jsonify({"error": "Ticker is required"}), 400

    conn = get_connection()
    # Check for duplicate
    existing = conn.execute(
        "SELECT 1 FROM all_account_info WHERE ticker = ? AND profile_id = ?",
        (ticker, profile_id),
    ).fetchone()
    if existing:
        conn.close()
        return jsonify({"error": f"{ticker} already exists for this profile"}), 409

    cols = ["ticker", "profile_id"]
    vals = [ticker, profile_id]
    allowed_fields = [
        "description", "classification_type", "price_paid", "current_price",
        "quantity", "purchase_value", "current_value", "gain_or_loss",
        "gain_or_loss_percentage", "percent_change", "div_frequency", "reinvest",
        "ex_div_date", "div_pay_date", "div", "dividend_paid", "estim_payment_per_year",
        "approx_monthly_income", "annual_yield_on_cost", "current_annual_yield",
        "purchase_date", "ytd_divs", "total_divs_received", "paid_for_itself",
    ]
    for field in allowed_fields:
        if field in data and data[field] is not None:
            cols.append(field)
            vals.append(data[field])

    # Set base_quantity and import_date for DRIP tracking
    from datetime import date as _date
    qty_val = data.get("quantity")
    if qty_val is not None:
        cols.append("base_quantity")
        vals.append(qty_val)
    cols.append("import_date")
    vals.append(_date.today().isoformat())

    placeholders = ", ".join(["?"] * len(cols))
    conn.execute(
        f"INSERT INTO all_account_info ({', '.join(cols)}) VALUES ({placeholders})",
        vals,
    )
    # Handle category assignment
    cat_name = (data.get("category") or "").strip()
    if cat_name:
        cat_row = conn.execute(
            "SELECT id FROM categories WHERE name = ? AND profile_id = ?",
            (cat_name, profile_id),
        ).fetchone()
        if cat_row:
            conn.execute(
                "INSERT OR IGNORE INTO ticker_categories (ticker, category_id, profile_id) VALUES (?, ?, ?)",
                (ticker, cat_row["id"], profile_id),
            )

    conn.commit()

    # Update derived tables
    populate_holdings(profile_id)
    populate_dividends(profile_id)
    conn.close()
    return jsonify({"ticker": ticker, "message": f"{ticker} added"}), 201


@app.route("/api/holdings/<ticker>", methods=["PUT"])
def update_holding(ticker):
    """Update a holding's fields."""
    is_agg, pids = get_profile_filter()
    data = request.get_json()
    ticker = ticker.upper()

    if is_agg:
        profile_id = _resolve_aggregate_profile(ticker, pids)
    else:
        profile_id = pids[0]

    conn = get_connection()
    existing = conn.execute(
        "SELECT 1 FROM all_account_info WHERE ticker = ? AND profile_id = ?",
        (ticker, profile_id),
    ).fetchone()
    if not existing:
        conn.close()
        return jsonify({"error": f"{ticker} not found"}), 404

    allowed_fields = [
        "description", "classification_type", "price_paid", "current_price",
        "quantity", "purchase_value", "current_value", "gain_or_loss",
        "gain_or_loss_percentage", "percent_change", "div_frequency", "reinvest",
        "ex_div_date", "div_pay_date", "div", "dividend_paid", "estim_payment_per_year",
        "approx_monthly_income", "annual_yield_on_cost", "current_annual_yield",
        "purchase_date", "ytd_divs", "total_divs_received", "paid_for_itself",
        "cash_not_reinvested", "total_cash_reinvested", "shares_bought_from_dividend",
    ]
    updates = []
    vals = []
    for field in allowed_fields:
        if field in data:
            updates.append(f"{field} = ?")
            vals.append(data[field])

    # When quantity is manually changed, reset base_quantity and import_date
    # so DRIP simulation restarts from this point (the user's new share count
    # already accounts for any DRIP up to now)
    if "quantity" in data:
        from datetime import date as _date
        updates.append("base_quantity = ?")
        vals.append(data["quantity"])
        updates.append("import_date = ?")
        vals.append(_date.today().isoformat())

    if not updates:
        conn.close()
        return jsonify({"error": "No fields to update"}), 400

    vals.extend([ticker, profile_id])
    conn.execute(
        f"UPDATE all_account_info SET {', '.join(updates)} WHERE ticker = ? AND profile_id = ?",
        vals,
    )
    # Handle category assignment
    if "category" in data:
        cat_name = (data["category"] or "").strip()
        # Remove existing assignment
        conn.execute(
            "DELETE FROM ticker_categories WHERE ticker = ? AND profile_id = ?",
            (ticker, profile_id),
        )
        if cat_name:
            cat_row = conn.execute(
                "SELECT id FROM categories WHERE name = ? AND profile_id = ?",
                (cat_name, profile_id),
            ).fetchone()
            if cat_row:
                conn.execute(
                    "INSERT INTO ticker_categories (ticker, category_id, profile_id) VALUES (?, ?, ?)",
                    (ticker, cat_row["id"], profile_id),
                )

    conn.commit()

    populate_holdings(profile_id)
    populate_dividends(profile_id)
    conn.close()
    return jsonify({"ticker": ticker, "message": f"{ticker} updated"})


@app.route("/api/holdings/<ticker>", methods=["DELETE"])
def delete_holding(ticker):
    """Delete a holding."""
    is_agg, pids = get_profile_filter()
    if is_agg:
        profile_id = _resolve_aggregate_profile(ticker, pids)
    else:
        profile_id = pids[0]
    ticker = ticker.upper()
    conn = get_connection()
    conn.execute(
        "DELETE FROM all_account_info WHERE ticker = ? AND profile_id = ?",
        (ticker, profile_id),
    )
    conn.execute("DELETE FROM holdings WHERE ticker = ? AND profile_id = ?", (ticker, profile_id))
    conn.execute("DELETE FROM dividends WHERE ticker = ? AND profile_id = ?", (ticker, profile_id))
    conn.execute("DELETE FROM transactions WHERE ticker = ? AND profile_id = ?", (ticker, profile_id))
    conn.commit()
    conn.close()
    return jsonify({"ticker": ticker, "message": f"{ticker} deleted"})


# ── Transactions ───────────────────────────────────────────────────────────────

def _rerollup_after_import(profile_id):
    """After an import, re-rollup any tickers that have transactions so the
    transaction-managed position data overrides whatever the spreadsheet had."""
    conn = get_connection()
    if _profile_is_positions_managed(profile_id, conn):
        conn.close()
        return
    rows = conn.execute(
        "SELECT DISTINCT ticker FROM transactions WHERE profile_id = ?",
        (profile_id,),
    ).fetchall()
    for r in rows:
        ticker = r["ticker"] if isinstance(r, dict) else r[0]
        _rollup_transactions(ticker, profile_id, conn)
    conn.close()


def _snapshot_positions(profile_id):
    """Capture pre-import position data + transaction status for each ticker.

    Returns {ticker: {qty, price, date, has_txns}} — call BEFORE the import
    writes to all_account_info so we know what to seed vs. what's new.
    """
    conn = get_connection()
    rows = conn.execute(
        "SELECT ticker, quantity, price_paid, purchase_date "
        "FROM all_account_info WHERE profile_id = ?",
        (profile_id,),
    ).fetchall()
    snap = {}
    for r in rows:
        ticker = r["ticker"] if isinstance(r, dict) else r[0]
        cnt_row = conn.execute(
            "SELECT COUNT(*) as cnt FROM transactions WHERE ticker = ? AND profile_id = ?",
            (ticker, profile_id),
        ).fetchone()
        has_txns = (cnt_row["cnt"] if isinstance(cnt_row, dict) else cnt_row[0]) > 0
        snap[ticker] = {
            "qty": r["quantity"] if isinstance(r, dict) else r[1],
            "price": r["price_paid"] if isinstance(r, dict) else r[2],
            "date": r["purchase_date"] if isinstance(r, dict) else r[3],
            "has_txns": has_txns,
        }
    conn.close()
    return snap


def _import_as_transactions(profile_id, pre_snapshot):
    """Convert freshly-imported rows into delta-based transactions.

    pre_snapshot is the result of _snapshot_positions() taken BEFORE the import.
    The imported quantity represents the user's NEW total position. We compare
    it against the pre-import quantity to determine the delta:

    - delta > 0 → BUY delta shares at the imported price
    - delta < 0 → SELL abs(delta) shares at the imported price
    - delta == 0 → no transaction (position unchanged)
    - New ticker (not in snapshot) → first BUY for the full imported quantity
    - Existing ticker WITHOUT transactions → seed old position first, then
      apply the delta as a BUY or SELL
    """
    from datetime import date as _date
    conn = get_connection()

    rows = conn.execute(
        "SELECT ticker, quantity, price_paid, purchase_date "
        "FROM all_account_info WHERE profile_id = ?",
        (profile_id,),
    ).fetchall()

    for r in rows:
        ticker = r["ticker"] if isinstance(r, dict) else r[0]
        imp_qty = (r["quantity"] if isinstance(r, dict) else r[1]) or 0
        imp_price = (r["price_paid"] if isinstance(r, dict) else r[2]) or 0
        imp_date = (r["purchase_date"] if isinstance(r, dict) else r[3]) or _date.today().isoformat()

        old = pre_snapshot.get(ticker)

        if not old:
            # New ticker — create first BUY from imported data
            if imp_qty > 0:
                conn.execute(
                    "INSERT INTO transactions (ticker, profile_id, transaction_type, "
                    "transaction_date, shares, price_per_share, fees, notes) "
                    "VALUES (?, ?, 'BUY', ?, ?, ?, 0, 'Imported from spreadsheet')",
                    (ticker, profile_id, imp_date, imp_qty, imp_price),
                )
                conn.commit()
                _rollup_transactions(ticker, profile_id, conn)
            continue

        old_qty = old["qty"] or 0
        delta = round(imp_qty - old_qty, 6)

        if not old["has_txns"]:
            # Seed the pre-import position as the first transaction
            if old_qty > 0:
                conn.execute(
                    "INSERT INTO transactions (ticker, profile_id, transaction_type, "
                    "transaction_date, shares, price_per_share, fees, notes) "
                    "VALUES (?, ?, 'BUY', ?, ?, ?, 0, 'Seed from pre-import position')",
                    (ticker, profile_id, old["date"] or _date.today().isoformat(),
                     old_qty, old["price"] or 0),
                )
                conn.commit()

        if delta > 0:
            conn.execute(
                "INSERT INTO transactions (ticker, profile_id, transaction_type, "
                "transaction_date, shares, price_per_share, fees, notes) "
                "VALUES (?, ?, 'BUY', ?, ?, ?, 0, 'Imported from spreadsheet')",
                (ticker, profile_id, imp_date, delta, imp_price),
            )
            conn.commit()
            _rollup_transactions(ticker, profile_id, conn)
        elif delta < 0:
            conn.execute(
                "INSERT INTO transactions (ticker, profile_id, transaction_type, "
                "transaction_date, shares, price_per_share, fees, notes) "
                "VALUES (?, ?, 'SELL', ?, ?, ?, 0, 'Imported from spreadsheet')",
                (ticker, profile_id, imp_date, abs(delta), imp_price),
            )
            conn.commit()
            _rollup_transactions(ticker, profile_id, conn)
        elif not old["has_txns"] and old_qty > 0:
            # No delta but we seeded — still need to rollup
            _rollup_transactions(ticker, profile_id, conn)

    conn.close()


def _rollup_transactions(ticker, profile_id, conn):
    """Recalculate all_account_info from transactions.

    BUY transactions add to a lot queue.
    SELL transactions consume lots using specific-lot allocations when stored
    in transaction_lot_allocations, otherwise falling back to FIFO.
    The remaining lots determine quantity, weighted-average price_paid, and
    purchase_value.  Realized gains are stored per-sell transaction and summed.
    """
    rows = conn.execute(
        "SELECT id, transaction_type, shares, price_per_share, fees, transaction_date "
        "FROM transactions WHERE ticker = ? AND profile_id = ? ORDER BY transaction_date, id",
        (ticker, profile_id),
    ).fetchall()
    if not rows:
        return

    def _val(r, key, idx):
        return r[key] if isinstance(r, dict) else r[idx]

    # Pre-load lot allocations for all sells in this ticker
    all_sell_ids = [
        _val(r, "id", 0)
        for r in rows
        if (_val(r, "transaction_type", 1) or "BUY").upper() == "SELL"
    ]
    alloc_map = _load_lot_alloc_map(conn, all_sell_ids)

    # Build lot queue and compute realized gains
    lots = []  # each lot: { id, shares, cost_per_share (incl. fees), date }
    share_deficit = 0.0
    total_realized = 0
    earliest_buy = None

    for r in rows:
        txn_id = _val(r, "id", 0)
        txn_type = (_val(r, "transaction_type", 1) or "BUY").upper()
        shares = _val(r, "shares", 2) or 0
        price = _val(r, "price_per_share", 3) or 0
        fees = _val(r, "fees", 4) or 0
        tdate = _val(r, "transaction_date", 5)

        if txn_type == "BUY":
            if share_deficit > 1e-9:
                covered = min(share_deficit, shares)
                share_deficit -= covered
                shares -= covered
            _apply_buy_to_lots(lots, shares, price, fees, txn_id=txn_id, txn_date=tdate)
            if shares > 1e-9 and tdate and (earliest_buy is None or tdate < earliest_buy):
                earliest_buy = tdate
            # Clear any realized_gain on BUY rows
            conn.execute("UPDATE transactions SET realized_gain = NULL WHERE id = ?", (txn_id,))
        else:  # SELL
            sell_proceeds = (shares * price) - fees  # fees reduce proceeds
            cost_of_sold, sell_remaining = _consume_sell_lots(
                lots,
                shares,
                alloc_map.get(txn_id),
            )
            share_deficit += sell_remaining

            realized = sell_proceeds - cost_of_sold
            total_realized += realized
            conn.execute("UPDATE transactions SET realized_gain = ? WHERE id = ?",
                         (round(realized, 2), txn_id))

    # Remaining lots = current position
    total_shares = sum(lot["shares"] for lot in lots) - share_deficit
    total_cost = sum(lot["shares"] * lot["cost_per_share"] for lot in lots)
    avg_price = total_cost / total_shares if total_shares else 0

    from datetime import date as _date

    if total_shares <= 1e-9:
        # Fully sold — remove the holding row entirely
        conn.execute(
            "DELETE FROM all_account_info WHERE ticker = ? AND profile_id = ?",
            (ticker, profile_id),
        )
    else:
        # Get current_price so we can recompute value and G/L immediately
        cp_row = conn.execute(
            "SELECT current_price FROM all_account_info WHERE ticker = ? AND profile_id = ?",
            (ticker, profile_id),
        ).fetchone()
        cp = (cp_row["current_price"] if isinstance(cp_row, dict) else cp_row[0]) if cp_row else None
        cur_val = round(total_shares * cp, 2) if cp else None
        gl = round(cur_val - total_cost, 2) if cur_val is not None else None
        gl_pct = round(gl / total_cost, 6) if gl is not None and total_cost > 0 else None

        conn.execute(
            """UPDATE all_account_info
               SET quantity = ?, price_paid = ?, purchase_value = ?,
                   purchase_date = ?, base_quantity = ?, import_date = ?,
                   realized_gains = ?,
                   current_value = ?, gain_or_loss = ?,
                   gain_or_loss_percentage = ?, percent_change = ?
               WHERE ticker = ? AND profile_id = ?""",
            (round(total_shares, 6), round(avg_price, 4), round(total_cost, 2),
             earliest_buy, round(total_shares, 6), _date.today().isoformat(),
             round(total_realized, 2),
             cur_val, gl, gl_pct, gl_pct,
             ticker, profile_id),
        )
    conn.commit()
    populate_holdings(profile_id)
    populate_dividends(profile_id)


def _seed_transaction_if_needed(ticker, profile_id, conn):
    """Auto-create a seed transaction from existing all_account_info data
    when a user first uses transactions for this ticker."""
    existing = conn.execute(
        "SELECT COUNT(*) as cnt FROM transactions WHERE ticker = ? AND profile_id = ?",
        (ticker, profile_id),
    ).fetchone()
    cnt = existing["cnt"] if isinstance(existing, dict) else existing[0]
    if cnt > 0:
        return  # already has transactions

    holding = conn.execute(
        "SELECT quantity, price_paid, purchase_date FROM all_account_info "
        "WHERE ticker = ? AND profile_id = ?",
        (ticker, profile_id),
    ).fetchone()
    if not holding:
        return
    qty = holding["quantity"] if isinstance(holding, dict) else holding[0]
    price = holding["price_paid"] if isinstance(holding, dict) else holding[1]
    pdate = holding["purchase_date"] if isinstance(holding, dict) else holding[2]
    if not qty:
        return
    conn.execute(
        "INSERT INTO transactions (ticker, profile_id, transaction_date, shares, price_per_share, fees, notes) "
        "VALUES (?, ?, ?, ?, ?, 0, 'Initial seed from existing holding')",
        (ticker, profile_id, pdate, qty, price),
    )
    conn.commit()


def _load_lot_alloc_map(conn, sell_ids):
    """Return {sell_txn_id: [{buy_txn_id, shares}, ...]} for the given sells."""
    alloc_map = {}
    if not sell_ids:
        return alloc_map

    placeholders = ",".join("?" * len(sell_ids))
    alloc_rows = conn.execute(
        f"SELECT sell_txn_id, buy_txn_id, shares FROM transaction_lot_allocations "
        f"WHERE sell_txn_id IN ({placeholders}) ORDER BY id",
        sell_ids,
    ).fetchall()
    for ar in alloc_rows:
        sid = ar["sell_txn_id"] if isinstance(ar, dict) else ar[0]
        bid = ar["buy_txn_id"] if isinstance(ar, dict) else ar[1]
        sh = ar["shares"] if isinstance(ar, dict) else ar[2]
        alloc_map.setdefault(sid, []).append({
            "buy_txn_id": bid,
            "shares": float(sh or 0),
        })
    return alloc_map


def _consume_sell_lots(lots, sell_shares, allocations=None):
    """Consume lots for a sell and return (cost_of_sold, remaining_sell_shares)."""
    cost_of_sold = 0.0
    sell_remaining = float(sell_shares or 0)

    if allocations:
        for alloc in allocations:
            buy_id = alloc["buy_txn_id"]
            alloc_shares = float(alloc["shares"] or 0)
            for lot in lots:
                if lot["id"] != buy_id:
                    continue
                take = min(alloc_shares, lot["shares"])
                cost_of_sold += take * lot["cost_per_share"]
                lot["shares"] -= take
                sell_remaining -= take
                break
    else:
        while sell_remaining > 1e-9 and lots:
            lot = lots[0]
            if lot["shares"] <= sell_remaining + 1e-9:
                cost_of_sold += lot["shares"] * lot["cost_per_share"]
                sell_remaining -= lot["shares"]
                lots.pop(0)
            else:
                cost_of_sold += sell_remaining * lot["cost_per_share"]
                lot["shares"] -= sell_remaining
                sell_remaining = 0

    lots[:] = [lot for lot in lots if lot["shares"] > 1e-9]
    return cost_of_sold, max(0.0, sell_remaining)


def _apply_buy_to_lots(lots, shares, price, fees, txn_id=None, txn_date=None):
    """Append a BUY lot and return its weighted cost per share."""
    shares = float(shares or 0)
    price = float(price or 0)
    fees = float(fees or 0)
    lot_cost = (shares * price) + fees
    cost_per = lot_cost / shares if shares else 0
    if shares > 1e-9:
        lot = {"shares": shares, "cost_per_share": cost_per}
        if txn_id is not None:
            lot["id"] = txn_id
        if txn_date is not None:
            lot["date"] = txn_date
        lots.append(lot)
    return cost_per


def _get_open_lots(conn, ticker, profile_ids, exclude_txn_id=None):
    """Return open BUY lots for a ticker after applying sells and allocations."""
    placeholders = ",".join("?" * len(profile_ids))
    rows = conn.execute(
        f"SELECT id, transaction_type, shares, price_per_share, fees, transaction_date "
        f"FROM transactions WHERE ticker = ? AND profile_id IN ({placeholders}) "
        f"ORDER BY transaction_date, id",
        [ticker] + profile_ids,
    ).fetchall()

    sell_ids = []
    for row in rows:
        txn_id = row["id"] if isinstance(row, dict) else row[0]
        txn_type = ((row["transaction_type"] if isinstance(row, dict) else row[1]) or "BUY").upper()
        if txn_id == exclude_txn_id:
            continue
        if txn_type == "SELL":
            sell_ids.append(txn_id)
    alloc_map = _load_lot_alloc_map(conn, sell_ids)

    lots = []
    share_deficit = 0.0
    for row in rows:
        txn_id = row["id"] if isinstance(row, dict) else row[0]
        if txn_id == exclude_txn_id:
            continue

        txn_type = ((row["transaction_type"] if isinstance(row, dict) else row[1]) or "BUY").upper()
        shares = float((row["shares"] if isinstance(row, dict) else row[2]) or 0)
        price = float((row["price_per_share"] if isinstance(row, dict) else row[3]) or 0)
        fees = float((row["fees"] if isinstance(row, dict) else row[4]) or 0)
        txn_date = row["transaction_date"] if isinstance(row, dict) else row[5]

        if txn_type == "BUY":
            if share_deficit > 1e-9:
                covered = min(share_deficit, shares)
                share_deficit -= covered
                shares -= covered
            cost_per = _apply_buy_to_lots(lots, shares, price, fees, txn_id=txn_id, txn_date=txn_date)
            if shares > 1e-9:
                lots[-1]["price_per_share"] = price
                lots[-1]["cost_per_share"] = cost_per
                lots[-1]["fees"] = fees
                lots[-1]["transaction_date"] = txn_date
        else:
            _, sell_remaining = _consume_sell_lots(lots, shares, alloc_map.get(txn_id))
            share_deficit += sell_remaining

    result = []
    for lot in lots:
        if lot["shares"] <= 1e-9:
            continue
        result.append({
            "id": lot["id"],
            "shares_remaining": round(lot["shares"], 6),
            "price_per_share": lot["price_per_share"],
            "cost_per_share": round(lot["cost_per_share"], 4),
            "fees": lot["fees"],
            "transaction_date": lot["transaction_date"],
        })
    return result


def _normalize_lot_allocations(conn, ticker, profile_id, shares, lot_allocations, exclude_txn_id=None):
    """Validate specific-lot allocations and return a clean list."""
    if not lot_allocations:
        return []

    try:
        shares = float(shares or 0)
    except (TypeError, ValueError):
        raise ValueError("Shares is required")

    open_lots = _get_open_lots(conn, ticker, [profile_id], exclude_txn_id=exclude_txn_id)
    available_by_id = {lot["id"]: float(lot["shares_remaining"] or 0) for lot in open_lots}
    cleaned = {}

    for alloc in lot_allocations:
        try:
            buy_id = int(alloc["buy_txn_id"])
            alloc_shares = float(alloc["shares"])
        except (KeyError, TypeError, ValueError):
            raise ValueError("Each lot allocation must include buy_txn_id and shares")

        if alloc_shares <= 0:
            continue
        if buy_id not in available_by_id:
            raise ValueError("One or more selected lots are no longer available")
        cleaned[buy_id] = cleaned.get(buy_id, 0.0) + alloc_shares

    if not cleaned:
        raise ValueError("Specific-lot mode requires at least one lot allocation")

    for buy_id, alloc_shares in cleaned.items():
        available = available_by_id[buy_id]
        if alloc_shares - available > 1e-6:
            raise ValueError(
                f"Allocated shares exceed what is available for lot {buy_id} "
                f"({alloc_shares:.6f} > {available:.6f})"
            )

    total_alloc = sum(cleaned.values())
    if abs(total_alloc - shares) > 1e-6:
        raise ValueError(
            f"Specific-lot allocations must add up to shares sold "
            f"({total_alloc:.6f} allocated vs {shares:.6f} entered)"
        )

    return [
        {"buy_txn_id": buy_id, "shares": round(alloc_shares, 6)}
        for buy_id, alloc_shares in cleaned.items()
    ]


def _annotate_transaction_rows(rows, alloc_map):
    """Attach lot metadata and running position/cost info to transaction rows."""
    annotated = []
    lots = []
    share_deficit = 0.0

    for row in rows:
        txn = dict(row) if isinstance(row, dict) else dict(zip(row.keys(), row))
        txn_id = txn["id"]
        txn_type = (txn.get("transaction_type") or "BUY").upper()
        shares = float(txn.get("shares") or 0)
        price = float(txn.get("price_per_share") or 0)
        fees = float(txn.get("fees") or 0)

        txn_allocs = alloc_map.get(txn_id, [])
        txn["lot_allocations"] = [
            {"buy_txn_id": alloc["buy_txn_id"], "shares": round(float(alloc["shares"] or 0), 6)}
            for alloc in txn_allocs
        ]
        txn["cost_basis_method"] = "SPECIFIC" if txn_allocs else "FIFO"

        if txn_type == "BUY":
            if share_deficit > 1e-9:
                covered = min(share_deficit, shares)
                share_deficit -= covered
                shares -= covered
            _apply_buy_to_lots(lots, shares, price, fees, txn_id=txn_id)
        else:
            _, sell_remaining = _consume_sell_lots(lots, shares, txn_allocs)
            share_deficit += sell_remaining

        total_shares = sum(lot["shares"] for lot in lots) - share_deficit
        total_cost = sum(lot["shares"] * lot["cost_per_share"] for lot in lots)
        avg_cost = total_cost / total_shares if total_shares else 0
        txn["position_after"] = round(total_shares, 6)
        txn["total_cost_after"] = round(total_cost, 2)
        txn["avg_cost_after"] = round(avg_cost, 4) if total_shares else 0
        annotated.append(txn)

    return annotated


@app.route("/api/holdings/<ticker>/open-lots", methods=["GET"])
def open_lots(ticker):
    """Return open (unsold) BUY lots for a ticker, computed via FIFO of existing sells."""
    is_agg, pids = get_profile_filter()
    ticker = ticker.upper()
    conn = get_connection()
    exclude_txn_id = request.args.get("exclude_txn_id", type=int)
    result = _get_open_lots(conn, ticker, pids, exclude_txn_id=exclude_txn_id)
    conn.close()
    return jsonify(result)


@app.route("/api/holdings/<ticker>/transactions", methods=["GET"])
def list_transactions(ticker):
    """List all transactions for a ticker."""
    is_agg, pids = get_profile_filter()
    ticker = ticker.upper()
    conn = get_connection()
    placeholders = ",".join("?" * len(pids))
    rows = conn.execute(
        f"SELECT * FROM transactions WHERE ticker = ? AND profile_id IN ({placeholders}) ORDER BY transaction_date, id",
        [ticker] + pids,
    ).fetchall()
    sell_ids = [
        row["id"] if isinstance(row, dict) else row[0]
        for row in rows
        if ((row["transaction_type"] if isinstance(row, dict) else row[1]) or "BUY").upper() == "SELL"
    ]
    alloc_map = _load_lot_alloc_map(conn, sell_ids)
    conn.close()
    return jsonify(_annotate_transaction_rows(rows, alloc_map))


@app.route("/api/holdings/<ticker>/transactions", methods=["POST"])
def add_transaction(ticker):
    """Add a new transaction for a ticker. Auto-seeds if first transaction."""
    is_agg, pids = get_profile_filter()
    if is_agg:
        profile_id = _resolve_aggregate_profile(ticker.upper(), pids)
    else:
        profile_id = pids[0]

    data = request.get_json()
    ticker = ticker.upper()

    shares = data.get("shares")
    if not shares:
        return jsonify({"error": "Shares is required"}), 400

    conn = get_connection()

    # Check if the holding exists; if not, create it (new ticker via transaction flow)
    existing = conn.execute(
        "SELECT 1 FROM all_account_info WHERE ticker = ? AND profile_id = ?",
        (ticker, profile_id),
    ).fetchone()
    if not existing:
        # Create a minimal holding — the rollup will fill in quantity/price/value
        from datetime import date as _date
        conn.execute(
            "INSERT INTO all_account_info (ticker, profile_id, description, import_date) VALUES (?, ?, ?, ?)",
            (ticker, profile_id, data.get("description", ""), _date.today().isoformat()),
        )
        # If a lookup was done on the frontend, apply those fields
        for field in ["description", "classification_type", "div", "div_frequency",
                      "ex_div_date", "div_pay_date", "current_price", "reinvest"]:
            if data.get(field) is not None:
                conn.execute(
                    f"UPDATE all_account_info SET {field} = ? WHERE ticker = ? AND profile_id = ?",
                    (data[field], ticker, profile_id),
                )
        # Handle category for new ticker
        cat_name = (data.get("category") or "").strip()
        if cat_name:
            cat_row = conn.execute(
                "SELECT id FROM categories WHERE name = ? AND profile_id = ?",
                (cat_name, profile_id),
            ).fetchone()
            if cat_row:
                cat_id = cat_row["id"] if isinstance(cat_row, dict) else cat_row[0]
                conn.execute(
                    "INSERT OR IGNORE INTO ticker_categories (ticker, category_id, profile_id) VALUES (?, ?, ?)",
                    (ticker, cat_id, profile_id),
                )
        conn.commit()
    else:
        # Auto-seed existing holding's current data as the first transaction
        _seed_transaction_if_needed(ticker, profile_id, conn)

    # Validate date year if provided
    txn_date = data.get("transaction_date")
    if txn_date:
        try:
            year = int(str(txn_date).split("-")[0])
            if year < 1900 or year > 2099:
                conn.close()
                return jsonify({"error": f"Invalid year {year} — must be between 1900 and 2099"}), 400
        except (ValueError, IndexError):
            conn.close()
            return jsonify({"error": "Invalid date format"}), 400

    # Insert the new transaction
    txn_type = (data.get("transaction_type") or "BUY").upper()
    if txn_type not in ("BUY", "SELL"):
        conn.close()
        return jsonify({"error": "transaction_type must be BUY or SELL"}), 400

    try:
        normalized_allocs = []
        if txn_type == "SELL":
            normalized_allocs = _normalize_lot_allocations(
                conn,
                ticker,
                profile_id,
                shares,
                data.get("lot_allocations"),
            )
    except ValueError as e:
        conn.close()
        return jsonify({"error": str(e)}), 400

    cur = conn.execute(
        "INSERT INTO transactions (ticker, profile_id, transaction_type, transaction_date, shares, price_per_share, fees, notes) "
        "VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
        (ticker, profile_id, txn_type, data.get("transaction_date"), float(shares),
         data.get("price_per_share"), data.get("fees", 0), data.get("notes")),
    )
    new_txn_id = cur.lastrowid
    conn.commit()

    # Store specific-lot allocations if provided
    if txn_type == "SELL" and normalized_allocs:
        for alloc in normalized_allocs:
            conn.execute(
                "INSERT INTO transaction_lot_allocations (sell_txn_id, buy_txn_id, shares) VALUES (?, ?, ?)",
                (new_txn_id, alloc["buy_txn_id"], float(alloc["shares"])),
            )
        conn.commit()

    # Rollup into all_account_info
    _rollup_transactions(ticker, profile_id, conn)
    conn.close()
    return jsonify({"ticker": ticker, "message": f"Transaction added for {ticker}"}), 201


@app.route("/api/holdings/<ticker>/transactions/<int:txn_id>", methods=["PUT"])
def update_transaction(ticker, txn_id):
    """Update an existing transaction."""
    is_agg, pids = get_profile_filter()
    ticker = ticker.upper()
    if is_agg:
        profile_id = _resolve_aggregate_profile(ticker, pids)
    else:
        profile_id = pids[0]

    data = request.get_json()

    # Validate date year if provided
    txn_date = data.get("transaction_date")
    if txn_date:
        try:
            year = int(str(txn_date).split("-")[0])
            if year < 1900 or year > 2099:
                return jsonify({"error": f"Invalid year {year} — must be between 1900 and 2099"}), 400
        except (ValueError, IndexError):
            return jsonify({"error": "Invalid date format"}), 400

    conn = get_connection()

    existing = conn.execute(
        "SELECT id, ticker, profile_id, transaction_type, shares FROM transactions WHERE id = ?",
        (txn_id,),
    ).fetchone()
    if not existing:
        conn.close()
        return jsonify({"error": "Transaction not found"}), 404
    existing_ticker = (existing["ticker"] if isinstance(existing, dict) else existing[1]).upper()
    existing_profile_id = existing["profile_id"] if isinstance(existing, dict) else existing[2]
    existing_type = ((existing["transaction_type"] if isinstance(existing, dict) else existing[3]) or "BUY").upper()
    existing_shares = existing["shares"] if isinstance(existing, dict) else existing[4]
    if existing_ticker != ticker or existing_profile_id != profile_id:
        conn.close()
        return jsonify({"error": "Transaction does not belong to this holding/profile"}), 404

    existing_allocs = _load_lot_alloc_map(conn, [txn_id]).get(txn_id, [])

    new_type = (data.get("transaction_type") or existing_type).upper()
    new_shares = data.get("shares", existing_shares)
    if new_type not in ("BUY", "SELL"):
        conn.close()
        return jsonify({"error": "transaction_type must be BUY or SELL"}), 400

    try:
        normalized_allocs = []
        if new_type == "SELL":
            if "lot_allocations" in data:
                normalized_allocs = _normalize_lot_allocations(
                    conn,
                    ticker,
                    profile_id,
                    new_shares,
                    data.get("lot_allocations"),
                    exclude_txn_id=txn_id,
                )
            elif existing_allocs:
                existing_alloc_total = sum(float(alloc["shares"] or 0) for alloc in existing_allocs)
                if abs(float(new_shares or 0) - existing_alloc_total) > 1e-6:
                    raise ValueError(
                        "This sell uses specific lots. Re-select the lots when changing the share count."
                    )
                normalized_allocs = [
                    {"buy_txn_id": alloc["buy_txn_id"], "shares": float(alloc["shares"] or 0)}
                    for alloc in existing_allocs
                ]
    except ValueError as e:
        conn.close()
        return jsonify({"error": str(e)}), 400

    updates = []
    vals = []
    for field in ["transaction_type", "transaction_date", "shares", "price_per_share", "fees", "notes"]:
        if field in data:
            updates.append(f"{field} = ?")
            vals.append(data[field])

    if not updates:
        conn.close()
        return jsonify({"error": "No fields to update"}), 400

    vals.append(txn_id)
    conn.execute(f"UPDATE transactions SET {', '.join(updates)} WHERE id = ?", vals)
    conn.execute("DELETE FROM transaction_lot_allocations WHERE sell_txn_id = ?", (txn_id,))
    if new_type == "SELL" and normalized_allocs:
        for alloc in normalized_allocs:
            conn.execute(
                "INSERT INTO transaction_lot_allocations (sell_txn_id, buy_txn_id, shares) VALUES (?, ?, ?)",
                (txn_id, alloc["buy_txn_id"], float(alloc["shares"])),
            )
    conn.commit()

    _rollup_transactions(ticker, profile_id, conn)
    conn.close()
    return jsonify({"ticker": ticker, "message": f"Transaction {txn_id} updated"})


@app.route("/api/holdings/<ticker>/transactions/<int:txn_id>", methods=["DELETE"])
def delete_transaction(ticker, txn_id):
    """Delete a transaction and re-rollup."""
    is_agg, pids = get_profile_filter()
    ticker = ticker.upper()
    if is_agg:
        profile_id = _resolve_aggregate_profile(ticker, pids)
    else:
        profile_id = pids[0]

    conn = get_connection()
    conn.execute("DELETE FROM transactions WHERE id = ?", (txn_id,))
    conn.commit()

    # Check if any transactions remain
    remaining = conn.execute(
        "SELECT COUNT(*) as cnt FROM transactions WHERE ticker = ? AND profile_id = ?",
        (ticker, profile_id),
    ).fetchone()
    cnt = remaining["cnt"] if isinstance(remaining, dict) else remaining[0]
    if cnt > 0:
        _rollup_transactions(ticker, profile_id, conn)
    # If no transactions left, the holding stays as-is (user can manage via edit modal)

    conn.close()
    return jsonify({"ticker": ticker, "message": f"Transaction {txn_id} deleted"})


@app.route("/api/holdings/<ticker>/has_transactions", methods=["GET"])
def has_transactions(ticker):
    """Check if a ticker has any transactions (used by frontend to toggle read-only)."""
    is_agg, pids = get_profile_filter()
    ticker = ticker.upper()
    conn = get_connection()
    placeholders = ",".join("?" * len(pids))
    row = conn.execute(
        f"SELECT COUNT(*) as cnt FROM transactions WHERE ticker = ? AND profile_id IN ({placeholders})",
        [ticker] + pids,
    ).fetchone()
    conn.close()
    cnt = row["cnt"] if isinstance(row, dict) else row[0]
    return jsonify({"has_transactions": cnt > 0, "count": cnt})


# ── Dividends ──────────────────────────────────────────────────────────────────

@app.route("/api/dividends", methods=["GET"])
def list_dividends():
    is_agg, pids = get_profile_filter()
    conn = get_connection()
    placeholders = ",".join("?" * len(pids))
    rows = conn.execute(
        f"SELECT * FROM dividends WHERE profile_id IN ({placeholders}) ORDER BY ticker", pids
    ).fetchall()
    conn.close()
    return jsonify(rows_to_dicts(rows))


@app.route("/api/dividends/<ticker>", methods=["PUT"])
def update_dividend(ticker):
    data = request.get_json()
    ticker = ticker.upper()
    conn = get_connection()

    allowed = [
        "div_frequency", "reinvest", "ex_div_date", "div_per_share",
        "dividend_paid", "estim_payment_per_year", "approx_monthly_income",
        "annual_yield_on_cost", "current_annual_yield",
        "ytd_divs", "total_divs_received", "paid_for_itself",
    ]
    updates = []
    vals = []
    for field in allowed:
        if field in data:
            updates.append(f"{field} = ?")
            vals.append(data[field])

    if not updates:
        conn.close()
        return jsonify({"error": "No fields to update"}), 400

    is_agg, pids = get_profile_filter()
    if is_agg:
        upd_pid = _resolve_aggregate_profile(ticker, pids)
    else:
        upd_pid = pids[0]
    vals.append(ticker)
    vals.append(upd_pid)
    conn.execute(f"UPDATE dividends SET {', '.join(updates)} WHERE ticker = ? AND profile_id = ?", vals)
    conn.commit()
    conn.close()
    return jsonify({"ticker": ticker, "message": f"{ticker} dividend info updated"})


# ── Income Tracking ────────────────────────────────────────────────────────────

@app.route("/api/income-tracking", methods=["GET"])
def list_income_tracking():
    profile_id = get_profile_id()
    conn = get_connection()
    rows = conn.execute(
        "SELECT * FROM income_tracking WHERE profile_id = ? ORDER BY import_date DESC, ticker",
        (profile_id,),
    ).fetchall()
    conn.close()
    return jsonify(rows_to_dicts(rows))


# ── Payouts ────────────────────────────────────────────────────────────────────

@app.route("/api/payouts/weekly", methods=["GET"])
def list_weekly_payouts():
    profile_id = get_profile_id()
    conn = get_connection()
    rows = conn.execute(
        "SELECT * FROM weekly_payouts WHERE profile_id = ? ORDER BY pay_date DESC",
        (profile_id,),
    ).fetchall()
    conn.close()
    return jsonify(rows_to_dicts(rows))


@app.route("/api/payouts/monthly", methods=["GET"])
def list_monthly_payouts():
    profile_id = get_profile_id()
    conn = get_connection()
    rows = conn.execute(
        "SELECT * FROM monthly_payouts WHERE profile_id = ? ORDER BY year DESC, month DESC",
        (profile_id,),
    ).fetchall()
    conn.close()
    return jsonify(rows_to_dicts(rows))


@app.route("/api/payouts/monthly/recalculate", methods=["POST"])
def recalculate_monthly_payouts():
    """Recalculate monthly_payouts from current holdings data for the trailing 12 months."""
    import datetime
    from collections import defaultdict
    profile_id = get_profile_id()
    conn = get_connection()
    cur = conn.cursor()
    today_d = datetime.date.today()
    month_start = today_d.replace(day=1)

    def _add_m(d, n):
        m = d.month - 1 + n
        y = d.year + m // 12
        m = m % 12 + 1
        return d.replace(year=y, month=m)

    window = [_add_m(month_start, i) for i in range(-11, 1)]

    # Load holdings
    rows = conn.execute(
        "SELECT ticker, div_frequency, estim_payment_per_year, quantity "
        "FROM all_account_info WHERE profile_id = ? AND quantity > 0",
        (profile_id,),
    ).fetchall()

    # Load pay-month schedule
    mpt_rows = conn.execute(
        "SELECT ticker, pay_month FROM monthly_payout_tickers WHERE profile_id = ?",
        (profile_id,),
    ).fetchall()
    ticker_pay_months = defaultdict(list)
    for r in mpt_rows:
        ticker_pay_months[r["ticker"]].append(int(r["pay_month"]))

    # Build estimate by calendar month (1-12)
    est_by_month = defaultdict(float)
    for r in rows:
        annual = float(r["estim_payment_per_year"] or 0)
        if annual <= 0:
            continue
        freq = str(r["div_frequency"] or "").strip()
        ticker = r["ticker"]
        pay_months = ticker_pay_months.get(ticker, [])
        if freq in ("M", "52", "W"):
            for mo in range(1, 13):
                est_by_month[mo] += annual / 12
        elif pay_months:
            n_pays = len(pay_months)
            per_pay = annual / n_pays
            for mo in pay_months:
                est_by_month[mo] += per_pay
        else:
            for mo in range(1, 13):
                est_by_month[mo] += annual / 12

    # Upsert each month in the window
    updated = 0
    for m in window:
        amount = round(est_by_month.get(m.month, 0), 2)
        if amount <= 0:
            continue
        existing = cur.execute(
            "SELECT 1 FROM monthly_payouts WHERE year = ? AND month = ? AND profile_id = ?",
            (m.year, m.month, profile_id),
        ).fetchone()
        if existing is None:
            cur.execute(
                "INSERT INTO monthly_payouts (year, month, amount, profile_id) VALUES (?, ?, ?, ?)",
                (m.year, m.month, amount, profile_id),
            )
        else:
            cur.execute(
                "UPDATE monthly_payouts SET amount = ? WHERE year = ? AND month = ? AND profile_id = ?",
                (amount, m.year, m.month, profile_id),
            )
        updated += 1

    conn.commit()
    conn.close()
    return jsonify({"updated": updated, "message": f"Recalculated {updated} months of payouts from current holdings."})


@app.route("/api/payouts/weekly-tickers", methods=["GET"])
def list_weekly_tickers():
    profile_id = get_profile_id()
    conn = get_connection()
    rows = conn.execute(
        "SELECT * FROM weekly_payout_tickers WHERE profile_id = ? ORDER BY ticker",
        (profile_id,),
    ).fetchall()
    conn.close()
    return jsonify(rows_to_dicts(rows))


@app.route("/api/payouts/monthly-tickers", methods=["GET"])
def list_monthly_tickers():
    profile_id = get_profile_id()
    conn = get_connection()
    rows = conn.execute(
        "SELECT * FROM monthly_payout_tickers WHERE profile_id = ? ORDER BY ticker, pay_month",
        (profile_id,),
    ).fetchall()
    conn.close()
    return jsonify(rows_to_dicts(rows))


# ── Template download ──────────────────────────────────────────────────────────

@app.route("/api/template/download", methods=["GET"])
def download_template():
    template_path = os.path.join(os.path.dirname(__file__), '..', 'templates', 'portfolio_upload_template.xlsx')
    if not os.path.exists(template_path):
        from create_template import create_template
        create_template()
    return send_file(
        os.path.abspath(template_path),
        as_attachment=True,
        download_name='portfolio_upload_template.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@app.route("/api/template/etrade-download", methods=["GET"])
def download_etrade_template():
    template_path = os.path.join(os.path.dirname(__file__), '..', 'templates', 'etrade_positions_template.csv')
    if not os.path.exists(template_path):
        from create_template import create_etrade_template
        create_etrade_template()
    return send_file(
        os.path.abspath(template_path),
        as_attachment=True,
        download_name='etrade_positions_template.csv',
        mimetype='text/csv',
    )


# ── Export Holdings ───────────────────────────────────────────────────────────

# Shared column definitions (matches create_template.py headers exactly)
_EXPORT_COL_MAP = [
    ("Ticker",              "ticker"),
    ("Shares",              "quantity"),
    ("Price Paid",          "price_paid"),
    ("Current Price",       "current_price"),
    ("Description",         "description"),
    ("Type",                "classification_type"),
    ("Date Purchased",      "purchase_date"),
    ("Purchase Value",      "purchase_value"),
    ("Current Value",       "current_value"),
    ("Gain/Loss",           "gain_or_loss"),
    ("Gain/Loss %",         "gain_or_loss_percentage"),
    ("% Change",            "percent_change"),
    ("Div/Share",           "div"),
    ("Frequency",           "div_frequency"),
    ("Ex-Div Date",         "ex_div_date"),
    ("Pay Date",            "div_pay_date"),
    ("DRIP",                "reinvest"),
    ("Div Paid",            "dividend_paid"),
    ("Est. Annual Pmt",     "estim_payment_per_year"),
    ("Monthly Income",      "approx_monthly_income"),
    ("Yield On Cost",       "annual_yield_on_cost"),
    ("Current Yield",       "current_annual_yield"),
    ("% of Account",        "percent_of_account"),
    ("YTD Divs",            "ytd_divs"),
    ("Total Divs Received", "total_divs_received"),
    ("Paid For Itself",     "paid_for_itself"),
    ("Cash Not Reinvest",   "cash_not_reinvested"),
    ("Cash Reinvested",     "total_cash_reinvested"),
    ("Shares from Div",     "shares_bought_from_dividend"),
    ("Shares/Year",         "shares_bought_in_year"),
    ("Shares/Month",        "shares_in_month"),
    ("8% Annual Wdraw",     "withdraw_8pct_cost_annually"),
    ("8% Monthly Wdraw",    "withdraw_8pct_per_month"),
    ("Category",            None),  # joined separately
]


def _export_profile_data(conn, profile_id):
    """Fetch export rows for a single profile. Returns (profile_name, [row_dicts])."""
    headers = [h for h, _ in _EXPORT_COL_MAP]
    sql_cols = [c for _, c in _EXPORT_COL_MAP if c is not None]

    prof = conn.execute("SELECT name FROM profiles WHERE id = ?", (profile_id,)).fetchone()
    profile_name = prof["name"] if prof else f"Portfolio {profile_id}"

    rows = conn.execute(
        f"SELECT {', '.join(sql_cols)} FROM all_account_info WHERE profile_id = ?",
        (profile_id,),
    ).fetchall()

    cat_map = {}
    cat_rows = conn.execute(
        "SELECT tc.ticker, c.name AS category_name "
        "FROM ticker_categories tc JOIN categories c ON c.id = tc.category_id "
        "WHERE tc.profile_id = ?",
        (profile_id,),
    ).fetchall()
    for cr in cat_rows:
        cat_map.setdefault(cr["ticker"], []).append(cr["category_name"])

    out_rows = []
    for row in rows:
        out = {}
        for header, sql_col in _EXPORT_COL_MAP:
            if header == "Category":
                out[header] = ", ".join(cat_map.get(row["ticker"], []))
            else:
                val = row[sql_col]
                out[header] = val if val is not None else ""
        out_rows.append(out)

    return profile_name, out_rows


@app.route("/api/export/holdings", methods=["GET"])
def export_holdings():
    """Export holdings as an Excel file compatible with both Generic and Owner reimport."""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    is_agg, profile_ids = get_profile_filter()
    conn = get_connection()
    headers = [h for h, _ in _EXPORT_COL_MAP]

    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True, color="FFFFFF", size=11)
    required_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
    optional_fill = PatternFill(start_color="37474F", end_color="37474F", fill_type="solid")
    thin_border = Border(bottom=Side(style="thin", color="90CAF9"))

    for pid in profile_ids:
        profile_name, rows = _export_profile_data(conn, pid)
        ws = wb.create_sheet(title=profile_name[:31])

        for ci, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=header)
            cell.font = header_font
            cell.fill = required_fill if ci <= 2 else optional_fill
            cell.alignment = Alignment(horizontal="center")

        for ri, row in enumerate(rows, 2):
            for ci, header in enumerate(headers, 1):
                cell = ws.cell(row=ri, column=ci, value=row[header])
                cell.border = thin_border

        for i, header in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(i)].width = max(len(header) + 4, 12)

    conn.close()

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    if len(profile_ids) == 1:
        prof_name = wb.sheetnames[0].replace(" ", "_")
        fname = f"portfolio_export_{prof_name}.xlsx"
    else:
        fname = "portfolio_export_all.xlsx"

    return send_file(buf, as_attachment=True,
                     download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/api/export/holdings/csv", methods=["GET"])
def export_holdings_csv():
    """Export holdings as a CSV file compatible with Generic reimport."""
    import csv
    from io import StringIO, BytesIO

    is_agg, profile_ids = get_profile_filter()
    conn = get_connection()
    headers = [h for h, _ in _EXPORT_COL_MAP]

    profile_name = None
    buf = StringIO()
    writer = csv.DictWriter(buf, fieldnames=headers)
    writer.writeheader()

    for pid in profile_ids:
        name, rows = _export_profile_data(conn, pid)
        if profile_name is None:
            profile_name = name
        writer.writerows(rows)

    conn.close()

    out = BytesIO(buf.getvalue().encode("utf-8"))

    if len(profile_ids) == 1:
        prof_name = (profile_name or "portfolio").replace(" ", "_")
        fname = f"portfolio_export_{prof_name}.csv"
    else:
        fname = "portfolio_export_all.csv"

    return send_file(out, as_attachment=True,
                     download_name=fname,
                     mimetype="text/csv")


# ── Dividend Comparison (Forward vs TTM) ──────────────────────────────────────

@app.route("/api/dividend-compare/holdings", methods=["GET"])
def dividend_compare_holdings():
    """Return forward and trailing-12-month dividend data for current portfolio holdings."""
    import yfinance as yf
    from datetime import datetime as _dt

    is_agg, pids = get_profile_filter()
    conn = get_connection()
    placeholders = ",".join("?" * len(pids))

    if is_agg and len(pids) > 1:
        rows = conn.execute(
            f"""SELECT ticker, MAX(description) as description, SUM(quantity) as quantity,
                   MAX(div_frequency) as div_frequency, MAX(current_price) as current_price,
                   CASE WHEN SUM(quantity) > 0 THEN SUM(purchase_value) / SUM(quantity) ELSE 0 END as price_paid
               FROM all_account_info
               WHERE profile_id IN ({placeholders}) AND quantity > 0
               GROUP BY ticker""",
            pids,
        ).fetchall()
    else:
        rows = conn.execute(
            f"""SELECT ticker, description, quantity, div_frequency, current_price, price_paid
               FROM all_account_info
               WHERE profile_id IN ({placeholders}) AND quantity > 0""",
            pids,
        ).fetchall()
    conn.close()

    if not rows:
        return jsonify([])

    tickers = [r["ticker"] for r in rows]
    holding_map = {r["ticker"]: dict(r) for r in rows}

    # Batch download dividend history (one yfinance call for TTM)
    ttm_map = {}
    try:
        raw = yf.download(" ".join(tickers), period="1y", progress=False, auto_adjust=False, actions=True)
        if not raw.empty:
            if isinstance(raw.columns, pd.MultiIndex):
                divs = raw["Dividends"] if "Dividends" in raw.columns.get_level_values(0) else None
            else:
                divs = raw[["Dividends"]] if "Dividends" in raw.columns else None

            if divs is not None:
                one_year_ago = pd.Timestamp.now() - pd.Timedelta(days=365)
                if hasattr(divs.index, 'tz') and divs.index.tz is not None:
                    one_year_ago = pd.Timestamp.now(tz=divs.index.tz) - pd.Timedelta(days=365)
                recent = divs[divs.index >= one_year_ago]

                if isinstance(recent, pd.Series):
                    # Single ticker
                    total = float(recent[recent > 0].sum())
                    if total > 0:
                        ttm_map[tickers[0]] = total
                else:
                    for t in tickers:
                        if t in recent.columns:
                            col = recent[t].dropna()
                            total = float(col[col > 0].sum())
                            if total > 0:
                                ttm_map[t] = total
    except Exception:
        pass

    # Fetch forward dividend rate via info (use threads for speed)
    from concurrent.futures import ThreadPoolExecutor, as_completed
    fwd_map = {}

    def _get_fwd(sym):
        try:
            info = yf.Ticker(sym).info or {}
            rate = info.get("dividendRate")
            return sym, rate if rate and rate > 0 else None
        except Exception:
            return sym, None

    with ThreadPoolExecutor(max_workers=8) as pool:
        futures = {pool.submit(_get_fwd, t): t for t in tickers}
        for fut in as_completed(futures):
            sym, rate = fut.result()
            if rate:
                fwd_map[sym] = rate

    # Build results
    results = []
    for tk_sym in tickers:
        h = holding_map[tk_sym]
        price = h["current_price"] or 0
        qty = h["quantity"] or 0
        fwd = fwd_map.get(tk_sym)
        ttm = ttm_map.get(tk_sym)

        entry = {
            "ticker": tk_sym,
            "description": h["description"] or tk_sym,
            "quantity": qty,
            "current_price": price,
            "price_paid": h["price_paid"] or 0,
            "div_frequency": h["div_frequency"] or "Q",
            "forward_annual_dividend": round(fwd, 4) if fwd else None,
            "forward_dividend_yield": round(fwd / price, 6) if fwd and price else None,
            "forward_income": round(fwd * qty, 2) if fwd else None,
            "ttm_dividend": round(ttm, 4) if ttm else None,
            "ttm_dividend_yield": round(ttm / price, 6) if ttm and price else None,
            "ttm_income": round(ttm * qty, 2) if ttm else None,
            "error": None if (fwd or ttm) else "Not enough information to compute",
        }
        results.append(entry)

    results.sort(key=lambda r: r["ticker"])
    return jsonify(results)


@app.route("/api/dividend-compare/lookup", methods=["POST"])
def dividend_compare_lookup():
    """Fetch forward and TTM dividend data for user-supplied tickers."""
    import yfinance as yf

    data = request.get_json()
    tickers = data.get("tickers", [])
    if not tickers:
        return jsonify([])

    results = []
    for tk_sym in tickers:
        tk_sym = tk_sym.strip().upper()
        if not tk_sym:
            continue
        entry = {
            "ticker": tk_sym,
            "description": tk_sym,
            "quantity": None,
            "current_price": 0,
            "price_paid": None,
            "div_frequency": None,
            "forward_annual_dividend": None,
            "forward_dividend_yield": None,
            "ttm_dividend": None,
            "ttm_dividend_yield": None,
            "ttm_income": None,
            "forward_income": None,
            "error": None,
        }
        try:
            tk = yf.Ticker(tk_sym)
            info = tk.info or {}
            price = info.get("regularMarketPrice") or info.get("currentPrice") or 0
            entry["current_price"] = price
            entry["description"] = (info.get("longName") or info.get("shortName") or tk_sym)[:200]

            # Infer frequency from history
            try:
                hist = tk.dividends
                if hist is not None and len(hist) > 0:
                    if hist.index.tz is not None:
                        one_year_ago = pd.Timestamp.now(tz=hist.index.tz) - pd.Timedelta(days=365)
                    else:
                        one_year_ago = pd.Timestamp.now() - pd.Timedelta(days=365)
                    recent = hist[hist.index >= one_year_ago]
                    n = len(recent[recent > 0])
                    entry["div_frequency"] = "W" if n >= 45 else "M" if n >= 10 else "Q" if n >= 3 else "SA" if n >= 2 else "A" if n >= 1 else None

                    ttm_total = float(recent[recent > 0].sum()) if n > 0 else 0
                    if ttm_total > 0:
                        entry["ttm_dividend"] = round(ttm_total, 4)
                        entry["ttm_dividend_yield"] = round(ttm_total / price, 6) if price else None
            except Exception:
                pass

            # Forward
            fwd = info.get("dividendRate")
            if fwd and fwd > 0:
                entry["forward_annual_dividend"] = round(fwd, 4)
                entry["forward_dividend_yield"] = round(fwd / price, 6) if price else None

            if entry["forward_annual_dividend"] is None and entry["ttm_dividend"] is None:
                entry["error"] = "Not enough information to compute"

        except Exception as e:
            entry["error"] = f"Lookup failed: {str(e)}"

        results.append(entry)

    return jsonify(results)


# ── Upcoming Dividends ─────────────────────────────────────────────────────────

@app.route("/api/upcoming-dividends", methods=["GET"])
def upcoming_dividends():
    """Return holdings with ex-div dates projected into the upcoming week."""
    from datetime import datetime, timedelta

    is_agg, pids = get_profile_filter()
    conn = get_connection()
    placeholders = ",".join("?" * len(pids))
    if is_agg and len(pids) > 1:
        rows = conn.execute(
            f"""SELECT ticker, MAX(description) as description, MAX(ex_div_date) as ex_div_date,
                   MAX(div) as div, MAX(div_frequency) as div_frequency,
                   SUM(quantity) as quantity, SUM(approx_monthly_income) as approx_monthly_income
               FROM all_account_info
               WHERE profile_id IN ({placeholders}) AND ex_div_date IS NOT NULL AND ex_div_date != ''
                 AND ex_div_date != '--' AND quantity > 0
               GROUP BY ticker""",
            pids,
        ).fetchall()
    else:
        rows = conn.execute(
            f"""SELECT ticker, description, ex_div_date, div, div_frequency, quantity, approx_monthly_income
               FROM all_account_info
               WHERE profile_id IN ({placeholders}) AND ex_div_date IS NOT NULL AND ex_div_date != ''
                 AND ex_div_date != '--' AND quantity > 0""",
            pids,
        ).fetchall()
    conn.close()

    if not rows:
        return jsonify([])

    today = datetime.today().date()
    week_end = today + timedelta(days=7)

    freq_days = {"W": 7, "52": 7, "M": 30, "Q": 91, "SA": 182, "A": 365}
    freq_labels = {"W": "Weekly", "52": "Weekly", "M": "Monthly", "Q": "Quarterly", "SA": "Semi-Annual", "A": "Annual"}
    freq_colors = {"M": "#00c9a7", "52": "#FFD700", "W": "#FFD700", "Q": "#7ecfff", "SA": "#f0a0ff", "A": "#f0a0ff"}

    def next_biz(d):
        if d.weekday() == 5:
            d += timedelta(days=2)
        elif d.weekday() == 6:
            d += timedelta(days=1)
        return d

    events = []
    for r in rows:
        try:
            ex = datetime.strptime(r["ex_div_date"], "%m/%d/%y").date()
        except (ValueError, TypeError):
            try:
                ex = datetime.strptime(r["ex_div_date"], "%Y-%m-%d").date()
            except (ValueError, TypeError):
                continue

        freq = (r["div_frequency"] or "Q").upper()
        step = freq_days.get(freq, 91)

        # Project forward — keep advancing until the estimated pay date >= today
        nxt = ex
        while True:
            if freq in ("W", "52"):
                pay = next_biz(nxt + timedelta(days=1))
            elif freq == "M":
                pay = next_biz(nxt + timedelta(days=2))
            else:
                pay = next_biz(nxt + timedelta(days=14))
            # Show if pay date is today or later
            if pay >= today:
                break
            nxt += timedelta(days=step)

        if nxt <= week_end:
            events.append({
                "ticker": r["ticker"],
                "description": r["description"] or r["ticker"],
                "ex_date": nxt.strftime("%Y-%m-%d"),
                "ex_weekday": nxt.strftime("%a"),
                "pay_date": pay.strftime("%Y-%m-%d"),
                "pay_weekday": pay.strftime("%a"),
                "amount": r["div"] or 0,
                "est_payment": round((r["div"] or 0) * (r["quantity"] or 0), 2),
                "frequency": freq,
                "freq_label": freq_labels.get(freq, "Quarterly"),
                "color": freq_colors.get(freq, "#7ecfff"),
            })

    events.sort(key=lambda e: e["ex_date"])
    return jsonify(events)


# ── Portfolio Summary Data (grades) ────────────────────────────────────────────

@app.route("/api/portfolio-coverage", methods=["GET"])
def portfolio_coverage():
    """Compute TTM yield-based coverage for each ticker and aggregate.

    Coverage ≈ (TTM price return % + TTM distribution yield %) / TTM distribution yield %
    This approximates  (SEC Yield + Option Yield) / 12-mo Distribution Yield.

    Interpretation:
      > 1.0  → sustainable
      0.8–1.0 → borderline
      < 0.8  → likely NAV decay
    """
    import warnings
    warnings.filterwarnings("ignore")

    pid = get_profile_id()
    conn = get_connection()
    rows = conn.execute(
        "SELECT ticker, current_price, quantity "
        "FROM all_account_info "
        "WHERE profile_id = ? AND quantity > 0",
        (pid,),
    ).fetchall()
    conn.close()

    import yfinance as yf
    from datetime import datetime as _dt, timedelta as _td

    # Deduplicate tickers, aggregate quantities and dollar values
    ticker_info = {}
    for r in rows:
        tk = r[0]
        cur_price = float(r[1] or 0)
        qty = float(r[2] or 0)
        if tk not in ticker_info:
            ticker_info[tk] = {"current_price": cur_price, "quantity": qty}
        else:
            ticker_info[tk]["quantity"] += qty
            if cur_price > 0:
                ticker_info[tk]["current_price"] = cur_price

    tickers = list(ticker_info.keys())
    one_year_ago = (_dt.now() - _td(days=365)).strftime("%Y-%m-%d")

    results = []
    total_price_return_dollars = 0.0
    total_dist_dollars = 0.0

    for tk in tickers:
        info = ticker_info[tk]
        cur_price = info["current_price"]
        qty = info["quantity"]

        if cur_price <= 0:
            results.append({"ticker": tk, "coverage_ratio": None})
            continue

        try:
            yf_tk = yf.Ticker(tk)
            hist = yf_tk.history(start=one_year_ago, interval="1d", auto_adjust=True)
            if hist.empty or len(hist) < 2:
                results.append({"ticker": tk, "coverage_ratio": None})
                continue

            price_1yr_ago = float(hist["Close"].iloc[0])
            if price_1yr_ago <= 0:
                results.append({"ticker": tk, "coverage_ratio": None})
                continue

            # TTM price return as a percentage of starting price
            ttm_price_return_pct = (cur_price - price_1yr_ago) / price_1yr_ago

            # TTM distributions per share
            divs = yf_tk.dividends
            if divs is not None and not divs.empty:
                cutoff = _dt.now() - _td(days=365)
                if divs.index.tz is not None:
                    cutoff = cutoff.astimezone(divs.index.tz)
                ttm_divs = divs[divs.index >= cutoff]
                ttm_dist_per_share = float(ttm_divs.sum())
            else:
                ttm_dist_per_share = 0.0

            # TTM distribution yield
            ttm_dist_yield = ttm_dist_per_share / cur_price if cur_price > 0 else 0.0

            if ttm_dist_yield > 0:
                coverage = round((ttm_price_return_pct + ttm_dist_yield) / ttm_dist_yield, 4)
            else:
                coverage = None

            results.append({"ticker": tk, "coverage_ratio": coverage})

            if coverage is not None:
                # Dollar-weight for aggregate
                dist_dollars = ttm_dist_per_share * qty
                price_return_dollars = (cur_price - price_1yr_ago) * qty
                total_dist_dollars += dist_dollars
                total_price_return_dollars += price_return_dollars

        except Exception:
            results.append({"ticker": tk, "coverage_ratio": None})

    if total_dist_dollars > 0:
        agg_coverage = round((total_price_return_dollars + total_dist_dollars) / total_dist_dollars, 4)
    else:
        agg_coverage = None

    return jsonify(results=results, aggregate_coverage=agg_coverage)


@app.route("/api/portfolio-summary/data", methods=["GET"])
def portfolio_summary_data():
    """Compute per-ticker grades and portfolio-level grade via yfinance."""
    import warnings
    import numpy as np
    import yfinance as yf
    from grading import ticker_score, grade_portfolio, letter_grade
    warnings.filterwarnings("ignore")

    is_agg, pids = get_profile_filter()
    conn = get_connection()
    placeholders = ",".join("?" * len(pids))
    rows = conn.execute(
        f"SELECT ticker, SUM(current_value) as current_value FROM all_account_info WHERE profile_id IN ({placeholders}) AND purchase_value > 0 AND quantity > 0 GROUP BY ticker",
        pids,
    ).fetchall()
    conn.close()

    if not rows:
        return jsonify({"error": "No data"}), 400

    tickers = [r["ticker"] for r in rows]
    all_dl = list(set(tickers + ["SPY"]))

    # Detect renamed tickers and include both old and new in download
    rename_map = {}
    for t in tickers:
        try:
            _info = yf.Ticker(t).info or {}
            _new = (_info.get("symbol") or "").upper()
            if _new and _new != t:
                rename_map[t] = _new
                if _new not in all_dl:
                    all_dl.append(_new)
        except Exception:
            pass

    try:
        raw = yf.download(" ".join(all_dl), period="1y", auto_adjust=True, progress=False)
        if raw.empty:
            return jsonify({"error": "No price data from yfinance"}), 500
    except Exception as e:
        return jsonify({"error": str(e)}), 500

    if isinstance(raw.columns, pd.MultiIndex):
        close = raw["Close"].dropna(how="all")
    else:
        close = raw[["Close"]].dropna(how="all")
        close.columns = [all_dl[0]]

    # Map renamed ticker columns back to original names
    for old_t, new_t in rename_map.items():
        if old_t not in close.columns and new_t in close.columns:
            close[old_t] = close[new_t]

    bench_close = close["SPY"] if "SPY" in close.columns else None
    bench_ret = bench_close.pct_change().dropna() if bench_close is not None else None

    ticker_grades = {}
    available = []
    for t in tickers:
        if t not in close.columns:
            ticker_grades[t] = {"grade": "N/A", "score": None}
            continue
        tc = close[t].dropna()
        if len(tc) < 30:
            ticker_grades[t] = {"grade": "N/A", "score": None}
            continue
        tr = tc.pct_change().dropna()
        score, *_ = ticker_score(tc, tr, bench_ret)
        ticker_grades[t] = {"grade": letter_grade(score), "score": score}
        available.append(t)

    import numpy as np
    portfolio_grade_info = {}
    if len(available) >= 2:
        returns_df = close[available].pct_change().fillna(0)
        val_map = {r["ticker"]: float(r["current_value"] or 0) for r in rows}
        weights_arr = np.array([val_map.get(t, 0.0) for t in available])
        pm = grade_portfolio(returns_df, weights_arr, bench_ret)
        portfolio_grade_info = pm.get("grade", {})
        portfolio_grade_info["sharpe"] = pm.get("sharpe")
        portfolio_grade_info["sortino"] = pm.get("sortino")
        portfolio_grade_info["calmar"] = pm.get("calmar")
        portfolio_grade_info["omega"] = pm.get("omega")
        portfolio_grade_info["max_drawdown"] = pm.get("max_drawdown")
        portfolio_grade_info["ulcer_index"] = pm.get("ulcer_index")

    return jsonify(ticker_grades=ticker_grades, portfolio_grade=portfolio_grade_info)


# ── Ticker Return Chart ───────────────────────────────────────────────────────

@app.route("/api/ticker-return/<ticker>", methods=["GET"])
def ticker_return_chart(ticker):
    """Return price return % and total return % data since purchase date."""
    import warnings
    import yfinance as yf
    warnings.filterwarnings("ignore")

    profile_id = get_profile_id()
    ticker = ticker.strip().upper()
    conn = get_connection()
    row = conn.execute(
        "SELECT purchase_date, price_paid, description FROM all_account_info WHERE ticker = ? AND profile_id = ? AND purchase_value > 0",
        (ticker, profile_id),
    ).fetchone()
    conn.close()

    if not row:
        return jsonify({"error": f"No data found for {ticker}"}), 404

    purchase_date = pd.to_datetime(row["purchase_date"])
    price_paid = float(row["price_paid"] or 0)
    description = row["description"] or ticker

    if pd.isna(purchase_date) or price_paid <= 0:
        return jsonify({"error": f"Missing purchase date or price for {ticker}"}), 404

    start_str = purchase_date.strftime("%Y-%m-%d")

    # Detect ticker rename (e.g. TOPW → WPAY)
    dl_ticker = ticker
    try:
        _info = yf.Ticker(ticker).info or {}
        _new_sym = (_info.get("symbol") or "").upper()
        if _new_sym and _new_sym != ticker:
            dl_ticker = _new_sym
    except Exception:
        pass

    try:
        raw = yf.download(dl_ticker, start=start_str, progress=False, auto_adjust=False, actions=True)
    except Exception as e:
        return jsonify({"error": f"Yahoo Finance error for {ticker}: {str(e)}"}), 404

    if raw.empty:
        return jsonify({"error": f"No Yahoo Finance data for {ticker}"}), 404

    try:
        if isinstance(raw.columns, pd.MultiIndex):
            close_col = raw["Close"][dl_ticker] if dl_ticker in raw["Close"].columns else raw["Close"].iloc[:, 0]
            divs_col = raw["Dividends"][dl_ticker] if dl_ticker in raw["Dividends"].columns else raw["Dividends"].iloc[:, 0]
        else:
            close_col = raw["Close"]
            divs_col = raw["Dividends"]

        # Ensure we have Series (not scalar from squeeze on single-row data)
        if not isinstance(close_col, pd.Series):
            close_col = pd.Series(close_col) if hasattr(close_col, '__iter__') else pd.Series()
        else:
            close_col = close_col.squeeze()
            if not isinstance(close_col, pd.Series):
                return jsonify({"error": f"Not enough price history for {ticker}"}), 404
        if not isinstance(divs_col, pd.Series):
            divs_col = pd.Series(0, index=close_col.index)
        else:
            divs_col = divs_col.squeeze()
            if not isinstance(divs_col, pd.Series):
                divs_col = pd.Series(0, index=close_col.index)

        if len(close_col) < 2:
            return jsonify({"error": f"Not enough price history for {ticker}"}), 404

        cum_divs = divs_col.reindex(close_col.index, fill_value=0).cumsum()

        price_return = ((close_col - price_paid) / price_paid * 100).round(2)
        total_return = ((close_col - price_paid + cum_divs) / price_paid * 100).round(2)

        return jsonify({
            "ticker": ticker,
            "description": description,
            "purchase_date": start_str,
            "price_paid": price_paid,
            "dates": close_col.index.strftime("%Y-%m-%d").tolist(),
            "price_return": price_return.tolist(),
            "total_return": total_return.tolist(),
        })
    except Exception as e:
        return jsonify({"error": f"Could not compute return data for {ticker}: {str(e)}"}), 500


@app.route("/api/ticker-return-1y/<ticker>", methods=["GET"])
def ticker_return_1y(ticker):
    """Return 1-year price return % and total return % for any ticker (no portfolio required)."""
    import warnings
    import yfinance as yf
    warnings.filterwarnings("ignore")

    ticker = ticker.strip().upper()
    start_date = (pd.Timestamp.now() - pd.DateOffset(years=1)).strftime("%Y-%m-%d")

    dl_ticker = ticker
    description = ticker
    try:
        _info = yf.Ticker(ticker).info or {}
        description = _info.get("longName") or _info.get("shortName") or ticker
        _new_sym = (_info.get("symbol") or "").upper()
        if _new_sym and _new_sym != ticker:
            dl_ticker = _new_sym
    except Exception:
        pass

    try:
        raw = yf.download(dl_ticker, start=start_date, progress=False, auto_adjust=False, actions=True)
    except Exception as e:
        return jsonify({"error": f"Yahoo Finance error for {ticker}: {str(e)}"}), 404

    if raw.empty:
        return jsonify({"error": f"No Yahoo Finance data for {ticker}"}), 404

    try:
        if isinstance(raw.columns, pd.MultiIndex):
            close_col = raw["Close"][dl_ticker] if dl_ticker in raw["Close"].columns else raw["Close"].iloc[:, 0]
            divs_col = raw["Dividends"][dl_ticker] if dl_ticker in raw["Dividends"].columns else raw["Dividends"].iloc[:, 0]
        else:
            close_col = raw["Close"]
            divs_col = raw["Dividends"]

        if not isinstance(close_col, pd.Series):
            close_col = pd.Series(close_col) if hasattr(close_col, '__iter__') else pd.Series()
        else:
            close_col = close_col.squeeze()
            if not isinstance(close_col, pd.Series):
                return jsonify({"error": f"Not enough price history for {ticker}"}), 404
        if not isinstance(divs_col, pd.Series):
            divs_col = pd.Series(0, index=close_col.index)
        else:
            divs_col = divs_col.squeeze()
            if not isinstance(divs_col, pd.Series):
                divs_col = pd.Series(0, index=close_col.index)

        if len(close_col) < 2:
            return jsonify({"error": f"Not enough price history for {ticker}"}), 404

        start_price = float(close_col.iloc[0])
        if start_price <= 0:
            return jsonify({"error": f"Invalid start price for {ticker}"}), 404

        cum_divs = divs_col.reindex(close_col.index, fill_value=0).cumsum()
        price_return = ((close_col - start_price) / start_price * 100).round(2)
        total_return = ((close_col - start_price + cum_divs) / start_price * 100).round(2)

        return jsonify({
            "ticker": ticker,
            "description": description,
            "start_price": round(start_price, 2),
            "dates": close_col.index.strftime("%Y-%m-%d").tolist(),
            "price_return": price_return.tolist(),
            "total_return": total_return.tolist(),
        })
    except Exception as e:
        return jsonify({"error": f"Could not compute return data for {ticker}: {str(e)}"}), 500


# ── Data Management ───────────────────────────────────────────────────────────

@app.route("/api/data/clear-all", methods=["POST"])
def clear_all_data():
    """Delete all holdings, dividends, and related data for the current profile."""
    profile_id = get_profile_id()
    conn = get_connection()
    tables = [
        "all_account_info", "holdings", "dividends", "income_tracking",
        "weekly_payouts", "monthly_payouts", "weekly_payout_tickers", "monthly_payout_tickers",
    ]
    counts = {}
    for t in tables:
        if t == "all_account_info":
            r = conn.execute(f"DELETE FROM {t} WHERE profile_id = ?", (profile_id,))
        else:
            r = conn.execute(f"DELETE FROM {t}")
        counts[t] = r.rowcount
    conn.commit()
    conn.close()
    return jsonify({"message": "All data cleared", "deleted": counts})


@app.route("/api/data/stats", methods=["GET"])
def data_stats():
    """Return row counts for key tables."""
    is_agg, pids = get_profile_filter()
    conn = get_connection()
    placeholders = ",".join("?" * len(pids))
    holdings = conn.execute(
        f"SELECT COUNT(DISTINCT ticker) as c FROM all_account_info WHERE profile_id IN ({placeholders})", pids
    ).fetchone()["c"]
    dividends = conn.execute(
        f"SELECT COUNT(*) as c FROM dividends WHERE profile_id IN ({placeholders})", pids
    ).fetchone()["c"]
    income = conn.execute(
        f"SELECT COUNT(*) as c FROM income_tracking WHERE profile_id IN ({placeholders})", pids
    ).fetchone()["c"]
    conn.close()
    return jsonify({"holdings": holdings, "dividends": dividends, "income_tracking": income})


# ── Income Summary ────────────────────────────────────────────────────────────

@app.route("/api/income-summary", methods=["GET"])
def income_summary():
    """Compute current month and YTD income from payout tables + estimates."""
    import datetime
    is_agg, pids = get_profile_filter()
    conn = get_connection()
    today = datetime.date.today()
    placeholders = ",".join("?" * len(pids))

    # Use stored actuals from refresh, fall back to estimates for holdings without data
    holdings = conn.execute(
        f"""SELECT quantity, div, ex_div_date, div_frequency,
                   ytd_divs, current_month_income
            FROM all_account_info
            WHERE profile_id IN ({placeholders})
              AND quantity IS NOT NULL AND quantity > 0""",
        pids,
    ).fetchall()

    total_month = 0.0
    total_ytd = 0.0
    for row in holdings:
        h = dict(row)
        stored_month = h.get("current_month_income")
        stored_ytd = h.get("ytd_divs")
        # Use "is None" not falsy — 0 means refresh ran and found no dividends
        # yet this period, which is correct (not a reason to use the full estimate)
        total_month += _estimate_current_month_income(h) if stored_month is None else stored_month
        total_ytd += _estimate_ytd_income(h) if stored_ytd is None else stored_ytd

    conn.close()
    return jsonify({
        "ytd_income": total_ytd,
        "current_month_income": total_month,
        "month_label": today.strftime("%B"),
    })


# ── Categories ────────────────────────────────────────────────────────────────

_CLASSIFICATION_NAMES = {
    "A": "Anchors", "B": "Boosters", "G": "Growth", "J": "Juicers",
    "BDC": "BDC", "HA": "Hedged Anchor", "GS": "Gold Silver",
    "ETF": "ETF", "EQUITY": "Equity", "CEF": "CEF", "REIT": "REIT",
}


@app.route("/api/categories/data", methods=["GET"])
def categories_data():
    """Return categories with assigned tickers, unallocated tickers, and total value."""
    profile_id = get_profile_id()
    conn = get_connection()

    # Auto-seed categories from classification_type if none exist
    cat_count = conn.execute(
        "SELECT COUNT(*) as c FROM categories WHERE profile_id = ?", (profile_id,)
    ).fetchone()["c"]

    if cat_count == 0:
        types = conn.execute(
            "SELECT DISTINCT classification_type FROM all_account_info WHERE profile_id = ? AND classification_type IS NOT NULL",
            (profile_id,),
        ).fetchall()
        for i, row in enumerate(types):
            ct = row["classification_type"]
            name = _CLASSIFICATION_NAMES.get(ct, ct)
            conn.execute(
                "INSERT OR IGNORE INTO categories (name, profile_id, sort_order) VALUES (?, ?, ?)",
                (name, profile_id, i),
            )
        conn.commit()
        # Auto-assign tickers to their seeded categories
        cats = conn.execute(
            "SELECT id, name FROM categories WHERE profile_id = ?", (profile_id,)
        ).fetchall()
        name_to_id = {r["name"]: r["id"] for r in cats}
        holdings = conn.execute(
            "SELECT ticker, classification_type FROM all_account_info WHERE profile_id = ?",
            (profile_id,),
        ).fetchall()
        for h in holdings:
            ct = h["classification_type"]
            cat_name = _CLASSIFICATION_NAMES.get(ct, ct)
            cat_id = name_to_id.get(cat_name)
            if cat_id:
                conn.execute(
                    "INSERT OR IGNORE INTO ticker_categories (ticker, category_id, profile_id) VALUES (?, ?, ?)",
                    (h["ticker"], cat_id, profile_id),
                )
        conn.commit()

    # Fetch categories
    cats = conn.execute(
        "SELECT id, name, target_pct, sort_order FROM categories WHERE profile_id = ? ORDER BY sort_order, name",
        (profile_id,),
    ).fetchall()

    # Fetch all holdings
    all_holdings = conn.execute(
        "SELECT ticker, description, current_value FROM all_account_info WHERE profile_id = ?",
        (profile_id,),
    ).fetchall()
    total_value = sum(float(h["current_value"] or 0) for h in all_holdings)

    # Fetch ticker-category assignments
    assignments = conn.execute(
        "SELECT ticker, category_id FROM ticker_categories WHERE profile_id = ?",
        (profile_id,),
    ).fetchall()
    # Build response
    holding_map = {h["ticker"]: h for h in all_holdings}
    categories = []
    assigned_tickers = set()
    for cat in cats:
        cat_tickers = []
        for a in assignments:
            if a["category_id"] == cat["id"]:
                t = a["ticker"]
                h = holding_map.get(t)
                if h:
                    cat_tickers.append({
                        "ticker": t,
                        "description": h["description"],
                        "current_value": float(h["current_value"] or 0),
                    })
                    assigned_tickers.add(t)
        cat_value = sum(t["current_value"] for t in cat_tickers)
        categories.append({
            "id": cat["id"],
            "name": cat["name"],
            "target_pct": cat["target_pct"],
            "sort_order": cat["sort_order"],
            "tickers": sorted(cat_tickers, key=lambda x: x["ticker"]),
            "actual_value": cat_value,
            "actual_pct": (cat_value / total_value * 100) if total_value else 0,
        })

    unallocated = []
    for h in all_holdings:
        if h["ticker"] not in assigned_tickers:
            unallocated.append({
                "ticker": h["ticker"],
                "description": h["description"],
                "current_value": float(h["current_value"] or 0),
            })
    unallocated.sort(key=lambda x: x["ticker"])

    conn.close()
    return jsonify({
        "categories": categories,
        "unallocated": unallocated,
        "total_value": total_value,
    })


@app.route("/api/categories", methods=["POST"])
def create_category():
    profile_id = get_profile_id()
    data = request.get_json()
    name = (data.get("name") or "").strip()
    if not name:
        return jsonify({"error": "Name is required"}), 400
    target_pct = data.get("target_pct")
    conn = get_connection()
    max_order = conn.execute(
        "SELECT COALESCE(MAX(sort_order), -1) + 1 as n FROM categories WHERE profile_id = ?",
        (profile_id,),
    ).fetchone()["n"]
    try:
        conn.execute(
            "INSERT INTO categories (name, target_pct, profile_id, sort_order) VALUES (?, ?, ?, ?)",
            (name, target_pct, profile_id, max_order),
        )
        conn.commit()
    except Exception:
        conn.close()
        return jsonify({"error": f"Category '{name}' already exists"}), 409
    conn.close()
    return jsonify({"message": f"Category '{name}' created"})


@app.route("/api/categories/<int:cat_id>", methods=["PUT"])
def update_category(cat_id):
    profile_id = get_profile_id()
    data = request.get_json()
    conn = get_connection()
    sets, vals = [], []
    if "name" in data:
        sets.append("name = ?")
        vals.append(data["name"].strip())
    if "target_pct" in data:
        sets.append("target_pct = ?")
        vals.append(data["target_pct"])
    if not sets:
        conn.close()
        return jsonify({"error": "Nothing to update"}), 400
    vals.extend([cat_id, profile_id])
    conn.execute(
        f"UPDATE categories SET {', '.join(sets)} WHERE id = ? AND profile_id = ?", vals
    )
    conn.commit()
    conn.close()
    return jsonify({"message": "Category updated"})


@app.route("/api/categories/<int:cat_id>", methods=["DELETE"])
def delete_category(cat_id):
    profile_id = get_profile_id()
    conn = get_connection()
    conn.execute(
        "DELETE FROM ticker_categories WHERE category_id = ? AND profile_id = ?",
        (cat_id, profile_id),
    )
    conn.execute(
        "DELETE FROM categories WHERE id = ? AND profile_id = ?",
        (cat_id, profile_id),
    )
    conn.commit()
    conn.close()
    return jsonify({"message": "Category deleted"})


@app.route("/api/categories/assign", methods=["POST"])
def assign_tickers():
    profile_id = get_profile_id()
    data = request.get_json()
    category_id = data.get("category_id")
    tickers = data.get("tickers", [])
    if not category_id or not tickers:
        return jsonify({"error": "category_id and tickers required"}), 400
    conn = get_connection()
    for t in tickers:
        conn.execute(
            "DELETE FROM ticker_categories WHERE ticker = ? AND profile_id = ?",
            (t, profile_id),
        )
        conn.execute(
            "INSERT INTO ticker_categories (ticker, category_id, profile_id) VALUES (?, ?, ?)",
            (t, category_id, profile_id),
        )
    conn.commit()
    conn.close()
    return jsonify({"message": f"Assigned {len(tickers)} ticker(s)"})


@app.route("/api/categories/unassign", methods=["POST"])
def unassign_tickers():
    profile_id = get_profile_id()
    data = request.get_json()
    tickers = data.get("tickers", [])
    if not tickers:
        return jsonify({"error": "tickers required"}), 400
    conn = get_connection()
    for t in tickers:
        conn.execute(
            "DELETE FROM ticker_categories WHERE ticker = ? AND profile_id = ?",
            (t, profile_id),
        )
    conn.commit()
    conn.close()
    return jsonify({"message": f"Unassigned {len(tickers)} ticker(s)"})


@app.route("/api/categories/reorder", methods=["POST"])
def reorder_categories():
    profile_id = get_profile_id()
    data = request.get_json()
    order = data.get("order", [])  # list of category IDs in desired order
    conn = get_connection()
    for i, cat_id in enumerate(order):
        conn.execute(
            "UPDATE categories SET sort_order = ? WHERE id = ? AND profile_id = ?",
            (i, cat_id, profile_id),
        )
    conn.commit()
    conn.close()
    return jsonify({"message": "Reordered"})


# ── Dividend Analysis ──────────────────────────────────────────────────────────

@app.route("/api/dividend-analysis/data", methods=["GET"])
def dividend_analysis_data():
    """Per-ticker dividend summary with charts, totals, grade, and category filter."""
    import math
    import json
    import datetime
    import warnings
    import numpy as np
    import yfinance as yf
    import plotly.graph_objects as go
    import plotly.utils
    from grading import grade_portfolio, letter_grade, _sharpe, _sortino
    warnings.filterwarnings("ignore")

    def _clean(v):
        if v is None:
            return None
        try:
            if math.isnan(v) or math.isinf(v):
                return None
        except (TypeError, ValueError):
            pass
        return v

    profile_id = get_profile_id()
    conn = get_connection()

    # Categories for filter dropdown
    cats = conn.execute(
        "SELECT id, name FROM categories WHERE profile_id = ? ORDER BY sort_order, name",
        (profile_id,),
    ).fetchall()
    categories = [{"id": c["id"], "name": c["name"]} for c in cats]

    # Category filter
    cat_param = request.args.get("category", "").strip()
    cat_ids = [c.strip() for c in cat_param.split(",") if c.strip()] if cat_param else []

    # Load holdings
    rows = conn.execute(
        """SELECT ticker, description, classification_type,
                  ytd_divs, total_divs_received, paid_for_itself,
                  dividend_paid, estim_payment_per_year, approx_monthly_income,
                  annual_yield_on_cost, current_annual_yield,
                  purchase_value, current_value, gain_or_loss,
                  div_frequency, ex_div_date, reinvest, div, current_price,
                  quantity
           FROM all_account_info
           WHERE purchase_value IS NOT NULL AND purchase_value > 0
             AND IFNULL(quantity, 0) > 0
             AND profile_id = ?
           ORDER BY IFNULL(total_divs_received, 0) DESC, ticker""",
        (profile_id,),
    ).fetchall()
    df = pd.DataFrame([dict(r) for r in rows])

    if df.empty:
        conn.close()
        return jsonify({"rows": [], "totals": {}, "charts": {}, "grade": {}, "categories": categories})

    # Enrich with category names
    try:
        cat_map_rows = conn.execute(
            "SELECT tc.ticker, c.name AS category_name "
            "FROM ticker_categories tc "
            "JOIN categories c ON c.id = tc.category_id "
            "WHERE tc.profile_id = ?", (profile_id,)
        ).fetchall()
        cat_map = pd.DataFrame([dict(r) for r in cat_map_rows])
        if not cat_map.empty:
            df = df.merge(cat_map, on="ticker", how="left")
        else:
            df["category_name"] = None
    except Exception:
        df["category_name"] = None

    if "classification_type" in df.columns:
        mask = df["category_name"].isna() | (df["category_name"] == "")
        df.loc[mask, "category_name"] = df.loc[mask, "classification_type"].map(
            lambda c: _CLASSIFICATION_NAMES.get(str(c).strip(), str(c).strip()) if pd.notna(c) else "Other"
        )
    df["category_name"] = df["category_name"].fillna("Other")

    # Apply category filter
    if cat_ids:
        cat_names = [c["name"] for c in categories if str(c["id"]) in cat_ids]
        if cat_names:
            df = df[df["category_name"].isin(cat_names)]

    if df.empty:
        conn.close()
        return jsonify({"rows": [], "totals": {}, "charts": {}, "grade": {}, "categories": categories})

    # Coerce numerics
    num_cols = ["ytd_divs", "total_divs_received", "estim_payment_per_year",
                "approx_monthly_income", "purchase_value", "dividend_paid",
                "current_value", "gain_or_loss", "annual_yield_on_cost",
                "current_annual_yield", "paid_for_itself", "div", "current_price", "quantity"]
    for col in num_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")
    df["description"] = df["description"].fillna("").astype(str)
    df["div_frequency"] = df["div_frequency"].fillna("").astype(str).str.strip()

    # ── Totals ──
    totals = {
        "ytd_divs": _clean(float(df["ytd_divs"].fillna(0).sum())),
        "total_divs_received": _clean(float(df["total_divs_received"].fillna(0).sum())),
        "dividend_paid": _clean(float(df["dividend_paid"].fillna(0).sum())),
        "estim_payment_per_year": _clean(float(df["estim_payment_per_year"].fillna(0).sum())),
        "approx_monthly_income": _clean(float(df["approx_monthly_income"].fillna(0).sum())),
    }

    # Actual monthly income from monthly_payouts
    today_d = datetime.date.today()
    month_label = today_d.strftime("%b %Y")
    totals["current_month_label"] = month_label
    try:
        actual_row = conn.execute(
            "SELECT amount FROM monthly_payouts WHERE year = ? AND month = ? AND profile_id = ?",
            (today_d.year, today_d.month, profile_id),
        ).fetchone()
        totals["actual_monthly_income"] = _clean(float(actual_row["amount"])) if actual_row else 0
    except Exception:
        totals["actual_monthly_income"] = 0

    # ── Build table rows ──
    table_rows = []
    for _, row in df.iterrows():
        table_rows.append({
            "ticker": row["ticker"],
            "description": row["description"],
            "category_name": row["category_name"],
            "ytd_divs": _clean(row.get("ytd_divs")),
            "total_divs_received": _clean(row.get("total_divs_received")),
            "paid_for_itself": _clean(row.get("paid_for_itself")),
            "dividend_paid": _clean(row.get("dividend_paid")),
            "estim_payment_per_year": _clean(row.get("estim_payment_per_year")),
            "approx_monthly_income": _clean(row.get("approx_monthly_income")),
            "annual_yield_on_cost": _clean(row.get("annual_yield_on_cost")),
            "current_annual_yield": _clean(row.get("current_annual_yield")),
            "gain_or_loss": _clean(row.get("gain_or_loss")),
            "current_value": _clean(row.get("current_value")),
        })

    # ── Charts ──
    charts = {}
    dark_layout = dict(template="plotly_dark", paper_bgcolor="#1a1f2e", plot_bgcolor="rgba(255,255,255,0.03)")

    # Chart 1: Est. Annual Income by Ticker (top 20)
    c1 = df[df["estim_payment_per_year"].fillna(0) > 0].nlargest(20, "estim_payment_per_year")
    if not c1.empty:
        fig1 = go.Figure(go.Bar(
            x=c1["ticker"].tolist(), y=c1["estim_payment_per_year"].tolist(),
            marker_color="#a855f7",
            text=[f"${v:,.0f}" for v in c1["estim_payment_per_year"].tolist()],
            textposition="outside", customdata=c1["description"].tolist(),
            hovertemplate="<b>%{x}</b><br>%{customdata}<br>Est. Annual: <b>$%{y:,.2f}</b><extra></extra>",
        ))
        fig1.update_layout(title="Est. Annual Income by Ticker (Top 20)", xaxis_tickangle=-45,
            yaxis=dict(title="Est. Annual Income ($)", range=[0, c1["estim_payment_per_year"].max() * 1.2]),
            margin=dict(t=50, b=100, l=60, r=20), **dark_layout)
        charts["annual_income"] = json.dumps(fig1, cls=plotly.utils.PlotlyJSONEncoder)

    # Chart 4: Total Divs Received by Ticker
    c4 = df[df["total_divs_received"].fillna(0) > 0].sort_values("total_divs_received", ascending=False)
    if not c4.empty:
        fig4 = go.Figure(go.Bar(
            x=c4["ticker"].tolist(), y=c4["total_divs_received"].tolist(),
            marker_color="#38bdf8",
            text=[f"${v:,.0f}" for v in c4["total_divs_received"].tolist()],
            textposition="outside", customdata=c4["description"].tolist(),
            hovertemplate="<b>%{x}</b><br>%{customdata}<br>Total Divs: <b>$%{y:,.2f}</b><extra></extra>",
        ))
        fig4.update_layout(title="Total Dividends Received by Ticker", xaxis_tickangle=-45,
            yaxis=dict(title="Total Divs Received ($)", range=[0, c4["total_divs_received"].max() * 1.2]),
            margin=dict(t=50, b=100, l=60, r=20), **dark_layout)
        charts["total_divs_ticker"] = json.dumps(fig4, cls=plotly.utils.PlotlyJSONEncoder)

    # Chart 5: Paid For Itself
    paid_df = df.copy()
    paid_df["paid_for_itself"] = pd.to_numeric(paid_df["paid_for_itself"], errors="coerce")
    paid_df = paid_df[paid_df["paid_for_itself"] > 0].copy()
    paid_df["paid_pct"] = (paid_df["paid_for_itself"] * 100).round(1)
    paid_df = paid_df.sort_values("paid_pct", ascending=False)
    if not paid_df.empty:
        fig5 = go.Figure()
        for ctype in paid_df["category_name"].unique().tolist():
            g = paid_df[paid_df["category_name"] == ctype]
            fig5.add_trace(go.Bar(
                x=g["ticker"].tolist(), y=g["paid_pct"].tolist(), name=ctype,
                text=[f"{v:.1f}%" for v in g["paid_pct"].tolist()],
                textposition="auto", customdata=g["description"].tolist(),
                hovertemplate="<b>%{x}</b><br>%{customdata}<br>Cost Recovered: <b>%{y:.1f}%</b><extra></extra>",
            ))
        fig5.add_hline(y=100, line_dash="dash", line_color="gold",
                       annotation_text="100% - Fully Paid For Itself", annotation_font_color="gold")
        _pfi_max = paid_df["paid_pct"].max()
        fig5.update_layout(title="Paid For Itself - % of Cost Recovered via Dividends",
            xaxis_tickangle=-45,
            yaxis=dict(title="% of Cost Recovered", range=[0, max(_pfi_max * 1.15, 110)]),
            legend_title="Category", margin=dict(t=60, b=100, l=60, r=20), **dark_layout)
        charts["paid_for_itself"] = json.dumps(fig5, cls=plotly.utils.PlotlyJSONEncoder)

    # Chart 6: Total Dividends by Category (pie)
    type_grp = df.groupby("category_name")["total_divs_received"].sum().reset_index()
    type_grp = type_grp[type_grp["total_divs_received"] > 0]
    if not type_grp.empty:
        fig6 = go.Figure(go.Pie(
            labels=type_grp["category_name"].tolist(),
            values=type_grp["total_divs_received"].tolist(),
            hovertemplate="<b>%{label}</b><br>Total Dividends: $%{value:,.2f}<br>Share: %{percent}<extra></extra>",
            textinfo="label+percent",
        ))
        fig6.update_layout(title="Total Dividends Received by Category", **dark_layout)
        charts["by_type"] = json.dumps(fig6, cls=plotly.utils.PlotlyJSONEncoder)

    # Chart 2: Projected Monthly Income - next 12 months
    def _add_m(d, n):
        mo = d.month - 1 + n
        return datetime.date(d.year + mo // 12, mo % 12 + 1, min(d.day, 28))

    def _safe_float(v):
        try:
            f = float(v)
            return f if not (math.isnan(f) or math.isinf(f)) else 0.0
        except (TypeError, ValueError):
            return 0.0

    month_start = today_d.replace(day=1)
    future_months = [_add_m(month_start, i) for i in range(12)]
    proj_key = {m: 0.0 for m in future_months}

    # Load actual pay-month assignments from monthly_payout_tickers
    mpt_rows = conn.execute(
        "SELECT ticker, pay_month FROM monthly_payout_tickers WHERE profile_id = ?",
        (profile_id,),
    ).fetchall()
    from collections import defaultdict
    ticker_pay_months = defaultdict(list)
    for r in mpt_rows:
        ticker_pay_months[r["ticker"]].append(r["pay_month"])

    for _, row in df.iterrows():
        annual = _safe_float(row.get("estim_payment_per_year"))
        if annual <= 0:
            continue
        freq = str(row.get("div_frequency") or "").strip()
        if freq in ("--", ""):
            continue
        ticker = row["ticker"]
        drip = str(row.get("reinvest") or "").strip().upper() == "Y"
        div_per_share = _safe_float(row.get("div"))
        price = _safe_float(row.get("current_price"))
        g = (div_per_share / price) if (drip and price > 0 and div_per_share > 0) else 0

        pay_months = ticker_pay_months.get(ticker, [])

        if freq in ("M", "52", "W"):
            # Monthly/weekly payers distribute to every month
            for i, m in enumerate(future_months):
                growth = (1 + g) ** (i // (1 if freq == "M" else 4))
                proj_key[m] += (annual / 12) * growth
        elif pay_months:
            # Use imported pay-month schedule for Q/SA/A payers
            n_pays = len(pay_months)
            per_pay = annual / n_pays
            pay_n = 0
            for m in future_months:
                if m.month in pay_months:
                    proj_key[m] += per_pay * ((1 + g) ** pay_n)
                    pay_n += 1
        else:
            # Fallback: spread evenly
            for m in future_months:
                proj_key[m] += annual / 12

    # Floor current month projected to actual if actual is higher
    try:
        if totals["actual_monthly_income"] and proj_key.get(month_start, 0) < totals["actual_monthly_income"]:
            proj_key[month_start] = totals["actual_monthly_income"]
    except Exception:
        pass

    proj_labels = [m.strftime("%b") for m in future_months]
    proj_vals = [round(v, 2) if not math.isnan(v) else 0.0 for v in (proj_key[m] for m in future_months)]
    total_12m = sum(proj_vals)
    fig2 = go.Figure(go.Bar(
        x=proj_labels, y=proj_vals, marker_color="#22d3ee",
        text=[f"${v:,.0f}" for v in proj_vals], textposition="outside",
        hovertemplate="<b>%{x}</b><br>Projected: <b>$%{y:,.2f}</b><extra></extra>",
    ))
    _proj_max = max(proj_vals) if proj_vals else 1
    fig2.update_layout(
        title=f"Projected Monthly Income - Next 12 Months  |  Est. Annual: ${total_12m:,.0f}  |  Monthly Avg: ${total_12m/12:,.0f}",
        xaxis=dict(type="category", categoryorder="array", categoryarray=proj_labels),
        yaxis=dict(title="Projected Income ($)", range=[0, _proj_max * 1.2]),
        margin=dict(t=60, b=60, l=60, r=20), **dark_layout)
    charts["projected_monthly"] = json.dumps(fig2, cls=plotly.utils.PlotlyJSONEncoder)

    # Chart 3: Monthly Dividends Received (past 12 months)
    # Uses actual monthly_payouts where available, fills gaps with estimates
    # from monthly_payout_tickers schedule + approx_monthly_income.
    try:
        da_window = [_add_m(month_start, i) for i in range(-11, 1)]
        win_start_key = da_window[0].year * 100 + da_window[0].month
        win_end_key = da_window[-1].year * 100 + da_window[-1].month
        mp_rows = conn.execute(
            "SELECT year, month, amount FROM monthly_payouts "
            "WHERE (year * 100 + month) >= ? AND (year * 100 + month) <= ? AND profile_id = ? "
            "ORDER BY year, month",
            (win_start_key, win_end_key, profile_id),
        ).fetchall()
        actuals = {(int(r["year"]), int(r["month"])): float(r["amount"]) for r in mp_rows}

        # Build estimated monthly income from pay-month schedule for gap-filling
        est_by_month = defaultdict(float)  # keyed by month number (1-12)
        for _, row in df.iterrows():
            annual = _safe_float(row.get("estim_payment_per_year"))
            if annual <= 0:
                continue
            freq = str(row.get("div_frequency") or "").strip()
            ticker = row["ticker"]
            pay_months = ticker_pay_months.get(ticker, [])
            if freq in ("M", "52", "W"):
                for mo in range(1, 13):
                    est_by_month[mo] += annual / 12
            elif pay_months:
                n_pays = len(pay_months)
                per_pay = annual / n_pays
                for mo in pay_months:
                    est_by_month[mo] += per_pay
            else:
                for mo in range(1, 13):
                    est_by_month[mo] += annual / 12

        rx_labels = [m.strftime("%b '%y") for m in da_window]
        rx_vals = []
        rx_is_actual = []
        for m in da_window:
            actual = actuals.get((m.year, m.month))
            if actual is not None and actual > 0:
                rx_vals.append(actual)
                rx_is_actual.append(True)
            else:
                # Use estimate for past months, 0 for future
                if m <= month_start:
                    rx_vals.append(round(est_by_month.get(m.month, 0), 2))
                    rx_is_actual.append(False)
                else:
                    rx_vals.append(0.0)
                    rx_is_actual.append(False)

        total_rx = sum(rx_vals)
        # Purple for actual, lighter/dashed for estimated
        bar_colors = ["#a855f7" if is_act else "#6b4d8a" for is_act in rx_is_actual]
        fig3 = go.Figure(go.Bar(
            x=rx_labels, y=rx_vals, marker_color=bar_colors,
            text=[f"${v:,.0f}" if v else "" for v in rx_vals], textposition="outside",
            hovertemplate="<b>%{x}</b><br>Amount: <b>$%{y:,.2f}</b><extra></extra>",
        ))
        _rx_max = max((v for v in rx_vals if v), default=1000)
        win_start, win_end = da_window[0], da_window[-1]
        fig3.update_layout(
            title=f"Monthly Dividends Received  |  {win_start.strftime('%b %Y')} - {win_end.strftime('%b %Y')}  |  Total: ${total_rx:,.0f}",
            xaxis=dict(type="category", categoryorder="array", categoryarray=rx_labels),
            yaxis=dict(title="Amount ($)", range=[0, _rx_max * 1.2]),
            margin=dict(t=60, b=60, l=60, r=20), **dark_layout)
        charts["monthly_received"] = json.dumps(fig3, cls=plotly.utils.PlotlyJSONEncoder)
    except Exception:
        pass

    # ── Portfolio Grade ──
    grade_info = {}
    tickers = df["ticker"].tolist()
    values_map = {row["ticker"]: float(row["current_value"] or 0) for _, row in df.iterrows()}
    try:
        all_dl = list(set(tickers + ["SPY"]))
        raw = yf.download(" ".join(all_dl), period="1y", auto_adjust=True, progress=False)
        if not raw.empty:
            if isinstance(raw.columns, pd.MultiIndex):
                close = raw["Close"].dropna(how="all")
            else:
                close = raw[["Close"]].dropna(how="all")
                close.columns = [all_dl[0]]
            bench_close = close["SPY"] if "SPY" in close.columns else None
            bench_ret = bench_close.pct_change().dropna() if bench_close is not None else None
            available = [t for t in tickers if t in close.columns and len(close[t].dropna()) >= 30]
            if len(available) >= 2:
                returns_df = close[available].pct_change().fillna(0)
                weights_arr = np.array([values_map.get(t, 0.0) for t in available])
                pm = grade_portfolio(returns_df, weights_arr, bench_ret)
                g = pm.get("grade", {})
                grade_info = {
                    "overall": g.get("overall", "N/A"),
                    "score": _clean(g.get("score")),
                    "sharpe": _clean(pm.get("sharpe")),
                    "sortino": _clean(pm.get("sortino")),
                }
    except Exception:
        pass

    conn.close()

    return jsonify(
        rows=table_rows,
        totals=totals,
        charts=charts,
        grade=grade_info,
        categories=categories,
    )


# ── Sync DRIP flags from sub-profiles to Owner ──────────────────────────────

@app.route("/api/sync-drip-to-owner", methods=["POST"])
def sync_drip_to_owner():
    """Copy reinvest flags from sub-profiles to Owner.

    For each ticker in Owner (profile 1):
    - Set reinvest='Y' if ANY sub-profile has reinvest='Y' (only 'N' if all are 'N').
    - Set drip_quantity = sum of shares from sub-profiles where reinvest='Y'.
      This way DRIP income calculations use only the DRIP-eligible shares,
      not the full aggregate quantity.
    """
    conn = get_connection()
    ensure_tables_exist(conn)
    owner_id = 1

    # Get sub-profile ids that are included in owner
    rows = conn.execute(
        "SELECT id FROM profiles WHERE id != ? AND include_in_owner = 1",
        (owner_id,),
    ).fetchall()
    sub_ids = [r["id"] if isinstance(r, dict) else r[0] for r in rows]

    if not sub_ids:
        return jsonify({"message": "No sub-profiles found", "updated": 0})

    # Build lookup: ticker -> (any_drip, drip_shares)
    placeholders = ",".join("?" * len(sub_ids))
    sub_rows = conn.execute(
        f"SELECT ticker, reinvest, COALESCE(quantity, 0) as qty FROM all_account_info WHERE profile_id IN ({placeholders})",
        sub_ids,
    ).fetchall()

    drip_map = {}  # ticker -> {"any_drip": bool, "all_drip": bool, "drip_qty": float, "total_qty": float}
    for r in sub_rows:
        t = r["ticker"] if isinstance(r, dict) else r[0]
        v = r["reinvest"] if isinstance(r, dict) else r[1]
        q = float(r["qty"] if isinstance(r, dict) else r[2] or 0)
        if t not in drip_map:
            drip_map[t] = {"any_drip": False, "all_drip": True, "drip_qty": 0.0, "total_qty": 0.0}
        drip_map[t]["total_qty"] += q
        if v == "Y":
            drip_map[t]["any_drip"] = True
            drip_map[t]["drip_qty"] += q
        else:
            drip_map[t]["all_drip"] = False

    # Update owner holdings
    updated = 0
    for ticker, info in drip_map.items():
        new_val = "Y" if info["any_drip"] else "N"
        # If ALL accounts have DRIP on, use None (meaning use total shares).
        # If only SOME have DRIP on, store the partial DRIP-eligible share count.
        new_drip_qty = None if info["all_drip"] else (round(info["drip_qty"], 6) if info["any_drip"] else None)
        cur = conn.execute(
            "SELECT reinvest, drip_quantity FROM all_account_info WHERE ticker = ? AND profile_id = ?",
            (ticker, owner_id),
        ).fetchone()
        if cur is None:
            continue
        old_val = cur["reinvest"] if isinstance(cur, dict) else cur[0]
        old_drip_qty = cur["drip_quantity"] if isinstance(cur, dict) else cur[1]
        if old_val != new_val or old_drip_qty != new_drip_qty:
            conn.execute(
                "UPDATE all_account_info SET reinvest = ?, drip_quantity = ? WHERE ticker = ? AND profile_id = ?",
                (new_val, new_drip_qty, ticker, owner_id),
            )
            updated += 1

    conn.commit()
    conn.close()

    # Refresh derived tables for owner
    populate_holdings(owner_id)
    populate_dividends(owner_id)

    return jsonify({"message": f"Synced DRIP flags to Owner — {updated} tickers updated", "updated": updated})


# ── DRIP Matrix ──────────────────────────────────────────────────────────────

@app.route("/api/drip-matrix", methods=["GET"])
def drip_matrix():
    """Return DRIP status for every ticker across all sub-profiles included in Owner."""
    conn = get_connection()
    ensure_tables_exist(conn)
    owner_id = 1

    # Get sub-profiles
    profiles = conn.execute(
        "SELECT id, name FROM profiles WHERE id != ? AND include_in_owner = 1 ORDER BY name",
        (owner_id,),
    ).fetchall()
    profile_list = [{"id": r["id"], "name": r["name"]} for r in profiles]
    sub_ids = [p["id"] for p in profile_list]

    if not sub_ids:
        conn.close()
        return jsonify(profiles=[], tickers=[])

    # Get all holdings across sub-profiles
    placeholders = ",".join("?" * len(sub_ids))
    rows = conn.execute(
        f"""SELECT ticker, profile_id, reinvest, COALESCE(quantity, 0) as qty
            FROM all_account_info
            WHERE profile_id IN ({placeholders})
            ORDER BY ticker""",
        sub_ids,
    ).fetchall()

    # Build ticker -> {profile_id: {reinvest, qty}}
    ticker_map = {}
    for r in rows:
        t = r["ticker"]
        pid = r["profile_id"]
        if t not in ticker_map:
            ticker_map[t] = {}
        ticker_map[t][pid] = {
            "reinvest": r["reinvest"] == "Y",
            "qty": round(float(r["qty"] or 0), 2),
        }

    # Get Owner totals
    owner_rows = conn.execute(
        "SELECT ticker, reinvest, COALESCE(quantity, 0) as qty, drip_quantity, "
        "       COALESCE(estim_payment_per_year, 0) as annual_income "
        "FROM all_account_info WHERE profile_id = ? ORDER BY ticker",
        (owner_id,),
    ).fetchall()
    conn.close()

    tickers = []
    for r in owner_rows:
        t = r["ticker"]
        if t not in ticker_map:
            continue
        total_qty = round(float(r["qty"] or 0), 2)
        drip_qty = round(float(r["drip_quantity"]), 2) if r["drip_quantity"] is not None else total_qty
        owner_drip = r["reinvest"] == "Y"
        annual_income = round(float(r["annual_income"] or 0), 2)
        accounts = {}
        for pid in sub_ids:
            if pid in ticker_map.get(t, {}):
                info = ticker_map[t][pid]
                accounts[str(pid)] = {"reinvest": info["reinvest"], "qty": info["qty"]}
        # Income proportional to DRIP shares vs total shares
        drip_income = round(annual_income * drip_qty / total_qty, 2) if (owner_drip and total_qty > 0) else 0
        tickers.append({
            "ticker": t,
            "total_qty": total_qty,
            "drip_qty": drip_qty if owner_drip else 0,
            "owner_drip": owner_drip,
            "annual_income": annual_income,
            "drip_income": drip_income,
            "accounts": accounts,
        })

    return jsonify(profiles=profile_list, tickers=tickers)


@app.route("/api/drip-matrix/toggle", methods=["POST"])
def drip_matrix_toggle():
    """Toggle DRIP for a specific ticker in a specific sub-profile."""
    data = request.get_json(force=True) or {}
    ticker = data.get("ticker")
    profile_id = data.get("profile_id")
    reinvest = data.get("reinvest")  # True/False

    if not ticker or not profile_id:
        return jsonify(error="Missing ticker or profile_id"), 400

    conn = get_connection()
    new_val = "Y" if reinvest else "N"
    conn.execute(
        "UPDATE all_account_info SET reinvest = ? WHERE ticker = ? AND profile_id = ?",
        (new_val, ticker, profile_id),
    )
    conn.commit()
    conn.close()

    return jsonify(ok=True)


# ── DRIP Settings & Projections ───────────────────────────────────────────────

@app.route("/api/drip-settings", methods=["GET", "POST"])
def drip_settings():
    """GET: load saved reinvest % per ticker. POST: save/update reinvest % for multiple tickers."""
    profile_id = get_profile_id()
    conn = get_connection()
    ensure_tables_exist(conn)

    if request.method == "GET":
        rows = conn.execute(
            "SELECT ticker, reinvest_pct FROM drip_settings WHERE profile_id = ?",
            (profile_id,),
        ).fetchall()
        conn.close()
        return jsonify({r["ticker"]: r["reinvest_pct"] for r in rows})

    # POST — save settings
    data = request.get_json(force=True, silent=True) or {}
    settings = data.get("settings", {})  # {ticker: reinvest_pct}
    cur = conn.cursor()
    for ticker, pct in settings.items():
        ticker = str(ticker).strip().upper()
        if not ticker:
            continue
        try:
            pct = max(0.0, min(100.0, float(pct)))
        except (TypeError, ValueError):
            pct = 100.0
        cur.execute(
            "INSERT INTO drip_settings (ticker, reinvest_pct, profile_id) VALUES (?, ?, ?) "
            "ON CONFLICT(ticker, profile_id) DO UPDATE SET reinvest_pct = excluded.reinvest_pct",
            (ticker, pct, profile_id),
        )
    conn.commit()
    conn.close()
    return jsonify(ok=True)


@app.route("/api/drip-contribution-settings", methods=["GET", "POST"])
def drip_contribution_settings():
    """GET/POST monthly contribution settings for DRIP projections."""
    profile_id = get_profile_id()
    conn = get_connection()
    ensure_tables_exist(conn)

    if request.method == "GET":
        row = conn.execute(
            "SELECT monthly_amount, targeted FROM drip_monthly_contribution WHERE profile_id = ?",
            (profile_id,),
        ).fetchone()
        targets = [
            r["ticker"]
            for r in conn.execute(
                "SELECT ticker FROM drip_contribution_targets WHERE profile_id = ?",
                (profile_id,),
            ).fetchall()
        ]
        conn.close()
        return jsonify(
            monthly_amount=row["monthly_amount"] if row else 0,
            targeted=bool(row["targeted"]) if row else False,
            targets=targets,
        )

    # POST
    data = request.get_json(force=True, silent=True) or {}
    monthly_amount = max(0.0, float(data.get("monthly_amount", 0)))
    targeted = 1 if data.get("targeted") else 0
    targets = data.get("targets", [])

    cur = conn.cursor()
    cur.execute(
        "INSERT INTO drip_monthly_contribution (profile_id, monthly_amount, targeted) VALUES (?, ?, ?) "
        "ON CONFLICT(profile_id) DO UPDATE SET monthly_amount = excluded.monthly_amount, targeted = excluded.targeted",
        (profile_id, monthly_amount, targeted),
    )
    cur.execute("DELETE FROM drip_contribution_targets WHERE profile_id = ?", (profile_id,))
    for t in targets:
        t = str(t).strip().upper()
        if t:
            cur.execute(
                "INSERT OR IGNORE INTO drip_contribution_targets (profile_id, ticker) VALUES (?, ?)",
                (profile_id, t),
            )
    conn.commit()
    conn.close()
    return jsonify(ok=True)


@app.route("/api/drip-redirects", methods=["GET", "POST"])
def drip_redirects():
    """GET/POST distribution redirect pairs for DRIP projections."""
    profile_id = get_profile_id()
    conn = get_connection()
    ensure_tables_exist(conn)

    if request.method == "GET":
        rows = conn.execute(
            "SELECT source_ticker, target_ticker FROM drip_redirects WHERE profile_id = ?",
            (profile_id,),
        ).fetchall()
        conn.close()
        return jsonify(redirects=[{"source": r["source_ticker"], "target": r["target_ticker"]} for r in rows])

    # POST
    data = request.get_json(force=True, silent=True) or {}
    redirects = data.get("redirects", [])

    cur = conn.cursor()
    cur.execute("DELETE FROM drip_redirects WHERE profile_id = ?", (profile_id,))
    for rd in redirects:
        src = str(rd.get("source", "")).strip().upper()
        tgt = str(rd.get("target", "")).strip().upper()
        if src and tgt and src != tgt:
            cur.execute(
                "INSERT OR IGNORE INTO drip_redirects (profile_id, source_ticker, target_ticker) VALUES (?, ?, ?)",
                (profile_id, src, tgt),
            )
    conn.commit()
    conn.close()
    return jsonify(ok=True)


@app.route("/api/analytics/drip-projection", methods=["POST"])
def drip_projection():
    """Project DRIP compounding for actual portfolio holdings."""
    import math

    def _clean(v):
        if v is None:
            return None
        try:
            if math.isnan(v) or math.isinf(v):
                return None
        except (TypeError, ValueError):
            pass
        return v

    profile_id = get_profile_id()
    conn = get_connection()
    ensure_tables_exist(conn)

    data = request.get_json(force=True, silent=True) or {}
    years = max(1, min(30, int(data.get("years", 5))))
    drip_overrides = data.get("drip_settings", {})  # {ticker: reinvest_pct}
    drip_default = data.get("drip_default", None)  # fallback pct from "Set All"
    investment_overrides = data.get("investment_overrides", {})  # {ticker: dollar_amount}
    filter_categories = data.get("categories", [])   # list of category names

    # Load saved DRIP defaults
    saved_drip = {}
    for r in conn.execute("SELECT ticker, reinvest_pct FROM drip_settings WHERE profile_id = ?", (profile_id,)).fetchall():
        saved_drip[r["ticker"]] = r["reinvest_pct"]

    # Load holdings
    rows = conn.execute(
        """SELECT ticker, description, classification_type,
                  quantity, current_price, div, div_frequency,
                  current_annual_yield, estim_payment_per_year
           FROM all_account_info
           WHERE purchase_value IS NOT NULL AND purchase_value > 0
             AND IFNULL(quantity, 0) > 0
             AND profile_id = ?
           ORDER BY ticker""",
        (profile_id,),
    ).fetchall()
    df = pd.DataFrame([dict(r) for r in rows])
    if df.empty:
        conn.close()
        return jsonify(holdings=[], yearly=[], totals={}, categories=[])

    # Enrich with category names
    try:
        cat_map_rows = conn.execute(
            "SELECT tc.ticker, c.name AS category_name "
            "FROM ticker_categories tc "
            "JOIN categories c ON c.id = tc.category_id "
            "WHERE tc.profile_id = ?", (profile_id,)
        ).fetchall()
        cat_map = pd.DataFrame([dict(r) for r in cat_map_rows])
        if not cat_map.empty:
            df = df.merge(cat_map, on="ticker", how="left")
        else:
            df["category_name"] = None
    except Exception:
        df["category_name"] = None

    if "classification_type" in df.columns:
        mask = df["category_name"].isna() | (df["category_name"] == "")
        df.loc[mask, "category_name"] = df.loc[mask, "classification_type"].map(
            lambda c: _CLASSIFICATION_NAMES.get(str(c).strip(), str(c).strip()) if pd.notna(c) else "Other"
        )
    df["category_name"] = df["category_name"].fillna("Other")

    # Categories list for filter UI
    cats = conn.execute(
        "SELECT id, name FROM categories WHERE profile_id = ? ORDER BY sort_order, name",
        (profile_id,),
    ).fetchall()
    categories_list = [{"id": c["id"], "name": c["name"]} for c in cats]
    conn.close()

    # Apply category filter
    if filter_categories:
        df = df[df["category_name"].isin(filter_categories)]
    if df.empty:
        return jsonify(holdings=[], yearly=[], totals={}, categories=categories_list)

    # Coerce numerics
    for col in ["quantity", "current_price", "div", "current_annual_yield", "estim_payment_per_year"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
    df["div_frequency"] = df["div_frequency"].fillna("").astype(str).str.strip()

    # New features: monthly contributions, targeted allocation, redirects
    monthly_contribution = max(0.0, float(data.get("monthly_contribution", 0)))
    contribution_targeted = bool(data.get("contribution_targeted", False))
    contribution_targets_list = [str(t).strip().upper() for t in data.get("contribution_targets", []) if str(t).strip()]
    redirects_list = data.get("redirects", [])
    redirect_map = {}  # source_ticker -> target_ticker
    for rd in redirects_list:
        src = str(rd.get("source", "")).strip().upper()
        tgt = str(rd.get("target", "")).strip().upper()
        if src and tgt and src != tgt:
            redirect_map[src] = tgt

    # Frequency multiplier (payments per year)
    freq_map = {
        "Monthly": 12, "M": 12, "Weekly": 52, "W": 52, "52": 52,
        "Bi-Weekly": 26, "BW": 26, "Quarterly": 4, "Q": 4,
        "Semi-Annual": 2, "SA": 2, "Annual": 1, "A": 1,
    }

    # Monthly payment schedule: which months each frequency pays in (1-indexed)
    # For weekly/bi-weekly, they pay every month with fractional payments
    month_pay_map = {
        12: list(range(1, 13)),          # Monthly: every month
        52: list(range(1, 13)),          # Weekly: every month (~4.33/month)
        26: list(range(1, 13)),          # Bi-Weekly: every month (~2.17/month)
        4:  [3, 6, 9, 12],              # Quarterly
        2:  [6, 12],                     # Semi-Annual
        1:  [12],                        # Annual
    }
    # Payments per month for each frequency
    payments_per_month = {
        12: 1, 52: 52 / 12, 26: 26 / 12, 4: 1, 2: 1, 1: 1,
    }

    # Build ticker info list
    tickers_info = []
    for _, row in df.iterrows():
        ticker = row["ticker"]
        shares = float(row["quantity"])
        price = float(row["current_price"]) if row["current_price"] > 0 else 0
        div_per_share = float(row["div"]) if row["div"] > 0 else 0
        freq_str = row["div_frequency"]
        freq = freq_map.get(freq_str, 0)
        annual_yield = float(row["current_annual_yield"]) if row["current_annual_yield"] else 0

        if div_per_share == 0 and shares > 0 and freq > 0:
            est_annual = float(row["estim_payment_per_year"])
            if est_annual > 0:
                div_per_share = est_annual / (shares * freq)

        # Apply investment override: recalculate shares from dollar amount
        if ticker in investment_overrides and price > 0:
            override_amt = max(0.0, float(investment_overrides[ticker]))
            shares = override_amt / price

        if ticker in drip_overrides:
            reinvest_pct = max(0.0, min(100.0, float(drip_overrides[ticker])))
        elif drip_default is not None:
            reinvest_pct = max(0.0, min(100.0, float(drip_default)))
        else:
            reinvest_pct = 0.0

        tickers_info.append({
            "ticker": ticker,
            "description": row.get("description", ""),
            "category": row.get("category_name", "Other"),
            "shares": shares,
            "price": price,
            "div_per_share": div_per_share,
            "freq_str": freq_str,
            "freq": freq,
            "annual_yield": annual_yield,
            "reinvest_pct": reinvest_pct,
        })

    # Add custom tickers (not in portfolio) for comparison
    custom_tickers_data = data.get("custom_tickers", [])
    existing_tickers = {t["ticker"] for t in tickers_info}
    for ct in custom_tickers_data:
        sym = str(ct.get("ticker", "")).strip().upper()
        if not sym or sym in existing_tickers:
            continue
        ct_price = float(ct.get("price", 0))
        ct_div = float(ct.get("div_per_share", 0))
        ct_freq_str = str(ct.get("freq_str", "Q"))
        ct_freq = freq_map.get(ct_freq_str, 0)
        if sym in investment_overrides and ct_price > 0:
            ct_shares = max(0.0, float(investment_overrides[sym])) / ct_price
        else:
            ct_shares = 0  # default to $0 invested for tickers not owned
        ct_reinvest = float(drip_overrides.get(sym, drip_default if drip_default is not None else 0))
        ct_annual_yield = (ct_div * ct_freq / ct_price) if ct_price > 0 and ct_freq > 0 else 0
        tickers_info.append({
            "ticker": sym,
            "description": ct.get("description", sym),
            "category": "Custom",
            "shares": ct_shares,
            "price": ct_price,
            "div_per_share": ct_div,
            "freq_str": ct_freq_str,
            "freq": ct_freq,
            "annual_yield": ct_annual_yield,
            "reinvest_pct": ct_reinvest,
        })
        existing_tickers.add(sym)

    # Build lookup maps
    ticker_list = [t["ticker"] for t in tickers_info]
    idx = {t["ticker"]: i for i, t in enumerate(tickers_info)}
    prices = {t["ticker"]: t["price"] for t in tickers_info}

    # Current state
    proj_shares = {t["ticker"]: t["shares"] for t in tickers_info}
    new_shares_drip = {t: 0.0 for t in ticker_list}
    new_shares_contrib = {t: 0.0 for t in ticker_list}
    new_shares_redirect_in = {t: 0.0 for t in ticker_list}

    # Year 0 = current state
    yearly_data = {yr: {} for yr in range(0, years + 1)}
    for t in tickers_info:
        tk = t["ticker"]
        annual_income = t["div_per_share"] * t["freq"] * t["shares"]
        yearly_data[0][tk] = {"shares": round(t["shares"], 4), "annual_income": round(annual_income, 2)}

    # Determine contribution-eligible tickers (min $0.50 price to avoid penny-stock distortion)
    # Exclude tickers with 0% reinvest unless explicitly targeted
    MIN_CONTRIB_PRICE = 0.50
    reinvest_by_ticker = {t["ticker"]: t["reinvest_pct"] for t in tickers_info}
    if contribution_targeted and contribution_targets_list:
        contrib_eligible = [t for t in contribution_targets_list if t in idx and prices.get(t, 0) >= MIN_CONTRIB_PRICE]
    else:
        contrib_eligible = [t for t in ticker_list if prices.get(t, 0) >= MIN_CONTRIB_PRICE and reinvest_by_ticker.get(t, 100) > 0]

    # Month-based simulation
    for yr in range(1, years + 1):
        for month in range(1, 13):
            # Step 1: Process dividend payments
            redirect_buys = {}  # target_ticker -> dollar amount to buy
            for ti in tickers_info:
                tk = ti["ticker"]
                freq = ti["freq"]
                if freq <= 0 or prices[tk] <= 0 or ti["div_per_share"] <= 0:
                    continue
                pay_months = month_pay_map.get(freq, [])
                if month not in pay_months:
                    continue
                ppm = payments_per_month.get(freq, 1)
                dividend = ti["div_per_share"] * proj_shares[tk] * ppm
                reinvest_amt = dividend * ti["reinvest_pct"] / 100.0

                if tk in redirect_map:
                    target = redirect_map[tk]
                    if target in idx and prices.get(target, 0) > 0:
                        redirect_buys[target] = redirect_buys.get(target, 0) + reinvest_amt
                    # Source ticker doesn't get new shares from its own dividends
                else:
                    new_sh = reinvest_amt / prices[tk]
                    proj_shares[tk] += new_sh
                    new_shares_drip[tk] += new_sh

            # Apply redirect buys
            for target, amt in redirect_buys.items():
                new_sh = amt / prices[target]
                proj_shares[target] += new_sh
                new_shares_redirect_in[target] += new_sh

            # Step 2: Monthly contribution
            if monthly_contribution > 0 and contrib_eligible:
                per_ticker = monthly_contribution / len(contrib_eligible)
                for t in contrib_eligible:
                    new_sh = per_ticker / prices[t]
                    proj_shares[t] += new_sh
                    new_shares_contrib[t] += new_sh

        # Record year-end state
        for ti in tickers_info:
            tk = ti["ticker"]
            yr_income = ti["div_per_share"] * ti["freq"] * proj_shares[tk]
            yearly_data[yr][tk] = {"shares": round(proj_shares[tk], 4), "annual_income": round(yr_income, 2)}

    # Build holdings output
    holdings_out = []
    for ti in tickers_info:
        tk = ti["ticker"]
        total_new = new_shares_drip[tk] + new_shares_contrib[tk] + new_shares_redirect_in[tk]
        proj_annual_income = ti["div_per_share"] * ti["freq"] * proj_shares[tk]
        annual_income_now = ti["div_per_share"] * ti["freq"] * ti["shares"]

        redirect_target = redirect_map.get(tk)

        holdings_out.append({
            "ticker": tk,
            "description": ti["description"],
            "category": ti["category"],
            "shares": round(ti["shares"], 4),
            "price": _clean(round(ti["price"], 2)),
            "div_per_share": _clean(round(ti["div_per_share"], 4)),
            "frequency": ti["freq_str"],
            "yield_pct": _clean(round(ti["annual_yield"] * 100, 2)),
            "reinvest_pct": round(ti["reinvest_pct"], 1),
            "projected_shares": round(proj_shares[tk], 4),
            "projected_annual_income": round(proj_annual_income, 2),
            "current_annual_income": round(annual_income_now, 2),
            "new_shares": round(total_new, 4),
            "new_shares_drip": round(new_shares_drip[tk], 4),
            "new_shares_contribution": round(new_shares_contrib[tk], 4),
            "new_shares_redirect_in": round(new_shares_redirect_in[tk], 4),
            "redirect_target": redirect_target,
        })

    # Build yearly totals for chart
    yearly_totals = []
    for yr in range(0, years + 1):
        total_income = sum(yearly_data[yr][tk]["annual_income"] for tk in ticker_list)
        total_shares_added = 0
        if yr > 0:
            for tk in ticker_list:
                total_shares_added += yearly_data[yr][tk]["shares"] - yearly_data[yr - 1][tk]["shares"]
        yearly_totals.append({
            "year": yr,
            "annual_income": round(total_income, 2),
            "new_shares_this_year": round(total_shares_added, 4),
        })

    # Per-ticker yearly series for stacked chart
    ticker_yearly = {}
    for tk in ticker_list:
        ticker_yearly[tk] = [{"year": yr, "annual_income": yearly_data[yr][tk]["annual_income"]} for yr in range(0, years + 1)]

    current_total = sum(h["current_annual_income"] for h in holdings_out)
    projected_total = sum(h["projected_annual_income"] for h in holdings_out)
    total_new = sum(h["new_shares"] for h in holdings_out)
    total_contrib_shares = sum(h["new_shares_contribution"] for h in holdings_out)

    totals = {
        "current_annual_income": round(current_total, 2),
        "projected_annual_income": round(projected_total, 2),
        "income_growth_pct": round((projected_total / current_total - 1) * 100, 2) if current_total > 0 else 0,
        "total_new_shares": round(total_new, 4),
        "total_contributions": round(monthly_contribution * 12 * years, 2),
        "contribution_new_shares": round(total_contrib_shares, 4),
        "years": years,
    }

    return jsonify(
        holdings=holdings_out,
        yearly_totals=yearly_totals,
        ticker_yearly=ticker_yearly,
        totals=totals,
        categories=categories_list,
    )


# ── Income Growth Simulator ───────────────────────────────────────────────────

@app.route("/api/analytics/income-growth-sim", methods=["POST"])
def income_growth_sim():
    """Project how portfolio income changes over time with scenario growth rates."""
    import math
    from datetime import datetime, timedelta

    def _safe(v):
        if v is None:
            return None
        try:
            f = float(v)
            return None if (math.isnan(f) or math.isinf(f)) else round(f, 2)
        except (TypeError, ValueError):
            return None

    data = request.get_json(force=True) or {}
    years = max(1, min(20, int(data.get("years", 5))))
    market_type = data.get("market_type", "neutral")
    monthly_contribution = max(0.0, float(data.get("monthly_contribution", 0)))
    reinvest_pct = max(0.0, min(100.0, float(data.get("reinvest_pct", 0))))
    use_monte_carlo = bool(data.get("monte_carlo", False))

    holdings_override = data.get("holdings_override")  # optional [{ticker, shares, price, div_per_share, freq_str, description}]

    # Scenario rates
    div_growth_map = {"bullish": 0.05, "neutral": 0.0, "bearish": -0.20}
    price_drift_map = {"bullish": 0.08, "neutral": 0.0, "bearish": -0.20}
    annual_div_growth = div_growth_map.get(market_type, 0.0)
    annual_price_drift = price_drift_map.get(market_type, 0.0)

    # Frequency maps (same as drip_projection)
    freq_map = {
        "Monthly": 12, "M": 12, "Weekly": 52, "W": 52, "52": 52,
        "Bi-Weekly": 26, "BW": 26, "Quarterly": 4, "Q": 4,
        "Semi-Annual": 2, "SA": 2, "Annual": 1, "A": 1,
    }
    month_pay_map = {
        12: list(range(1, 13)), 52: list(range(1, 13)), 26: list(range(1, 13)),
        4: [3, 6, 9, 12], 2: [6, 12], 1: [12],
    }
    payments_per_month = {12: 1, 52: 52/12, 26: 26/12, 4: 1, 2: 1, 1: 1}

    if holdings_override:
        # Use caller-provided holdings instead of DB query
        holdings = []
        total_value = 0.0
        for ho in holdings_override:
            freq = freq_map.get((ho.get("freq_str") or "").strip(), 0)
            shares = float(ho.get("shares") or 0)
            price = float(ho.get("price") or 0)
            dps = float(ho.get("div_per_share") or 0)
            if shares <= 0 or price <= 0:
                continue
            total_value += price * shares
            is_reinvest = bool(ho.get("reinvest", False))
            holdings.append({
                "ticker": (ho.get("ticker") or "").upper(),
                "description": ho.get("description") or "",
                "shares": shares,
                "drip_shares": shares if is_reinvest else 0.0,
                "price": price,
                "div_per_share": dps,
                "freq": freq,
                "freq_str": (ho.get("freq_str") or "").strip(),
                "reinvest": is_reinvest,
            })
        if not holdings:
            return jsonify(error="No valid holdings in override.")
    else:
        # Query holdings from DB
        _, pids = get_profile_filter()
        conn = get_connection()
        rows = conn.execute(
            """SELECT ticker, description, classification_type,
                      quantity, current_price, div, div_frequency,
                      current_annual_yield, estim_payment_per_year,
                      reinvest, drip_quantity
               FROM all_account_info
               WHERE purchase_value IS NOT NULL AND purchase_value > 0
                 AND IFNULL(quantity, 0) > 0
                 AND profile_id IN ({})
               ORDER BY ticker""".format(",".join("?" * len(pids))),
            pids,
        ).fetchall()
        conn.close()

        if not rows:
            return jsonify(error="No holdings found.")

        # Build holdings list (aggregate across profiles)
        holdings_map = {}
        for r in rows:
            t = r["ticker"]
            if t not in holdings_map:
                holdings_map[t] = {
                    "ticker": t,
                    "description": r["description"] or "",
                    "quantity": 0.0,
                    "current_price": float(r["current_price"] or 0),
                    "div_per_share": float(r["div"] or 0),
                    "freq_str": (r["div_frequency"] or "").strip(),
                    "estim_payment_per_year": float(r["estim_payment_per_year"] or 0),
                    "reinvest": False,
                    "drip_quantity": 0.0,
                }
            holdings_map[t]["quantity"] += float(r["quantity"] or 0)
            # Per-ticker DRIP: 'Y' if any row has it on
            if (r["reinvest"] if isinstance(r, dict) else "N") == "Y":
                holdings_map[t]["reinvest"] = True
            # drip_quantity: use stored value if available, else accumulate from rows with reinvest='Y'
            dq = float(r["drip_quantity"] or 0) if r["drip_quantity"] is not None else None
            if dq is not None and dq > 0:
                holdings_map[t]["drip_quantity"] = max(holdings_map[t]["drip_quantity"], dq)
            elif (r["reinvest"] if isinstance(r, dict) else "N") == "Y":
                holdings_map[t]["drip_quantity"] += float(r["quantity"] or 0)

        holdings = []
        total_value = 0.0
        for h in holdings_map.values():
            freq = freq_map.get(h["freq_str"], 0)
            dps = h["div_per_share"]
            if dps == 0 and h["quantity"] > 0 and freq > 0:
                est = h["estim_payment_per_year"]
                if est > 0:
                    dps = est / (h["quantity"] * freq)
            total_value += h["current_price"] * h["quantity"]
            has_drip = h["reinvest"] if reinvest_pct > 0 else False
            # drip_shares: if drip_quantity is set, use it; else use all shares (all accounts have DRIP on)
            drip_sh = h["drip_quantity"] if (has_drip and h["drip_quantity"] > 0) else (h["quantity"] if has_drip else 0.0)
            holdings.append({
                "ticker": h["ticker"],
                "description": h["description"],
                "shares": h["quantity"],
                "drip_shares": drip_sh,
                "price": h["current_price"],
                "div_per_share": dps,
                "freq": freq,
                "freq_str": h["freq_str"],
                "reinvest": has_drip,
            })

    if total_value <= 0:
        return jsonify(error="Portfolio has no value.")

    # Allocation weights for monthly contributions
    weights = {}
    eligible = [h for h in holdings if h["price"] >= 0.50 and h["freq"] > 0]
    if eligible:
        elig_value = sum(h["price"] * h["shares"] for h in eligible)
        for h in eligible:
            weights[h["ticker"]] = (h["price"] * h["shares"]) / elig_value if elig_value > 0 else 0

    total_months = years * 12
    now = datetime.now()
    start_year = now.year
    start_month = now.month

    def _month_label(m):
        """Return calendar month (1-12) and label string for simulation month m (1-based)."""
        total_m = start_month + m
        y = start_year + (total_m - 1) // 12
        cm = ((total_m - 1) % 12) + 1
        month_names = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
        return cm, f"{month_names[cm - 1]} {y}"

    # ── Deterministic simulation ──────────────────────────────────────────
    def _run_deterministic():
        monthly_series = []
        annual_series = []
        # Track per-holding state: total shares and DRIP-eligible shares separately
        sim_shares = {h["ticker"]: h["shares"] for h in holdings}
        sim_drip_shares = {h["ticker"]: h.get("drip_shares", h["shares"] if h.get("reinvest") else 0.0) for h in holdings}
        contrib_shares = {h["ticker"]: 0.0 for h in holdings}
        prev_income = None

        for m in range(1, total_months + 1):
            cal_month, label = _month_label(m)
            growth_factor = (1 + annual_div_growth) ** (m / 12)
            price_factor = (1 + annual_price_drift) ** (m / 12)

            # Monthly contribution: buy shares at drifted prices
            if monthly_contribution > 0 and weights:
                for tk, w in weights.items():
                    alloc = monthly_contribution * w
                    adj_price = next(h["price"] for h in holdings if h["ticker"] == tk) * price_factor
                    if adj_price > 0:
                        new_sh = alloc / adj_price
                        sim_shares[tk] += new_sh
                        contrib_shares[tk] += new_sh

            # Calculate income for this month and DRIP reinvest
            month_income = 0.0
            month_income_existing = 0.0
            month_income_contrib = 0.0
            drip_buys = {}  # ticker -> dividend dollars to reinvest
            for h in holdings:
                tk = h["ticker"]
                freq = h["freq"]
                if freq <= 0 or h["div_per_share"] <= 0:
                    continue
                # Spread annual income evenly across 12 months for display
                adj_dps = h["div_per_share"] * growth_factor
                annual_inc = adj_dps * sim_shares[tk] * freq
                monthly_inc = annual_inc / 12
                annual_inc_existing = adj_dps * h["shares"] * freq
                monthly_inc_existing = annual_inc_existing / 12
                month_income += monthly_inc
                month_income_existing += monthly_inc_existing
                month_income_contrib += (monthly_inc - monthly_inc_existing)
                # DRIP: only reinvest dividends from DRIP-eligible shares
                pay_months = month_pay_map.get(freq, [])
                if cal_month in pay_months and h.get("reinvest"):
                    ppm = payments_per_month.get(freq, 1)
                    drip_amt = adj_dps * sim_drip_shares[tk] * ppm
                    drip_buys[tk] = drip_amt

            # DRIP: reinvest dividends back into the same holdings
            if drip_buys:
                for tk, drip_amt in drip_buys.items():
                    h_ref = next(h for h in holdings if h["ticker"] == tk)
                    adj_price = h_ref["price"] * price_factor
                    if adj_price > 0:
                        new_sh = drip_amt / adj_price
                        sim_shares[tk] += new_sh
                        sim_drip_shares[tk] += new_sh
                        contrib_shares[tk] += new_sh

            monthly_series.append({
                "month": m,
                "label": label,
                "total_income": _safe(month_income),
                "income_from_existing": _safe(month_income_existing),
                "income_from_contributions": _safe(month_income_contrib),
                "p10": None, "p90": None,
            })

        # Compute trailing-12-month annualized income and changes
        for i, entry in enumerate(monthly_series):
            if i >= 11:
                t12 = sum(monthly_series[j]["total_income"] or 0 for j in range(i - 11, i + 1))
            else:
                # Not enough months yet — extrapolate from available
                partial = sum(monthly_series[j]["total_income"] or 0 for j in range(0, i + 1))
                t12 = partial / (i + 1) * 12
            entry["annualized_income"] = _safe(t12)

        prev_ann = None
        for entry in monthly_series:
            ann = entry["annualized_income"] or 0
            if i < 12:
                # Trailing window not yet full — suppress noisy changes
                entry["change_dollar"] = None
                entry["change_pct"] = None
            elif prev_ann is not None:
                entry["change_dollar"] = _safe(ann - prev_ann)
                entry["change_pct"] = _safe((ann - prev_ann) / prev_ann * 100 if prev_ann > 0 else 0)
            else:
                entry["change_dollar"] = _safe(0)
                entry["change_pct"] = _safe(0)
            prev_ann = ann

        # Build annual series
        for yr in range(1, years + 1):
            start_idx = (yr - 1) * 12
            end_idx = yr * 12
            yr_months = monthly_series[start_idx:end_idx]
            yr_total = sum(m["total_income"] or 0 for m in yr_months)
            yr_existing = sum(m["income_from_existing"] or 0 for m in yr_months)
            yr_contrib = sum(m["income_from_contributions"] or 0 for m in yr_months)
            if yr == 1:
                prev_yr_total = sum(m["total_income"] or 0 for m in monthly_series[:12])
                # For year 1, compare against current annual income
                current_ann = sum(
                    h["div_per_share"] * h["freq"] * h["shares"]
                    for h in holdings if h["freq"] > 0
                )
                change_d = yr_total - current_ann
                change_p = (change_d / current_ann * 100) if current_ann > 0 else 0
            else:
                prev_start = (yr - 2) * 12
                prev_end = (yr - 1) * 12
                prev_yr_total = sum(m["total_income"] or 0 for m in monthly_series[prev_start:prev_end])
                change_d = yr_total - prev_yr_total
                change_p = (change_d / prev_yr_total * 100) if prev_yr_total > 0 else 0

            annual_series.append({
                "year": yr,
                "label": f"Year {yr}",
                "total_income": _safe(yr_total),
                "income_from_existing": _safe(yr_existing),
                "income_from_contributions": _safe(yr_contrib),
                "change_dollar": _safe(change_d),
                "change_pct": _safe(change_p),
                "p10": None, "p90": None,
            })

        return monthly_series, annual_series, sim_shares, contrib_shares

    # ── Monte Carlo simulation ────────────────────────────────────────────
    def _run_monte_carlo():
        import numpy as np
        N_PATHS = 300

        # Base div growth noise: sigma ~1.5% monthly for div changes
        div_sigma = 0.015
        # Price noise: sigma ~5% monthly
        price_sigma = 0.05

        monthly_div_bias = annual_div_growth / 12
        monthly_price_bias = annual_price_drift / 12

        # Run N_PATHS simulations
        # For each path, track monthly portfolio income
        np.random.seed(None)
        path_monthly_income = np.zeros((N_PATHS, total_months))
        path_monthly_existing = np.zeros((N_PATHS, total_months))
        path_monthly_contrib = np.zeros((N_PATHS, total_months))

        for p in range(N_PATHS):
            sim_shares_p = {h["ticker"]: h["shares"] for h in holdings}
            sim_drip_shares_p = {h["ticker"]: h.get("drip_shares", h["shares"] if h.get("reinvest") else 0.0) for h in holdings}
            cum_div_factor = {h["ticker"]: 1.0 for h in holdings}
            cum_price_factor = {h["ticker"]: 1.0 for h in holdings}

            for m in range(total_months):
                cal_month, _ = _month_label(m + 1)

                # Random walk for div and price factors
                for h in holdings:
                    tk = h["ticker"]
                    div_noise = np.random.normal(monthly_div_bias, div_sigma)
                    price_noise = np.random.normal(monthly_price_bias, price_sigma)
                    cum_div_factor[tk] *= (1 + div_noise)
                    cum_div_factor[tk] = max(cum_div_factor[tk], 0.1)  # floor at 10% of original
                    cum_price_factor[tk] *= (1 + price_noise)
                    cum_price_factor[tk] = max(cum_price_factor[tk], 0.1)

                # Monthly contribution
                if monthly_contribution > 0 and weights:
                    for tk, w in weights.items():
                        alloc = monthly_contribution * w
                        h_ref = next(h for h in holdings if h["ticker"] == tk)
                        adj_price = h_ref["price"] * cum_price_factor[tk]
                        if adj_price > 0:
                            sim_shares_p[tk] += alloc / adj_price

                # Income for this month (spread evenly across 12 months)
                month_inc = 0.0
                month_existing = 0.0
                mc_drip_buys = {}
                for h in holdings:
                    tk = h["ticker"]
                    freq = h["freq"]
                    if freq <= 0 or h["div_per_share"] <= 0:
                        continue
                    adj_dps = h["div_per_share"] * cum_div_factor[tk]
                    annual_inc = adj_dps * sim_shares_p[tk] * freq
                    monthly_inc = annual_inc / 12
                    annual_inc_existing = adj_dps * h["shares"] * freq
                    monthly_inc_existing = annual_inc_existing / 12
                    month_inc += monthly_inc
                    month_existing += monthly_inc_existing
                    # DRIP: only reinvest dividends from DRIP-eligible shares
                    pay_months = month_pay_map.get(freq, [])
                    if cal_month in pay_months and h.get("reinvest"):
                        ppm = payments_per_month.get(freq, 1)
                        drip_amt = adj_dps * sim_drip_shares_p[tk] * ppm
                        mc_drip_buys[tk] = drip_amt

                # DRIP reinvest
                if mc_drip_buys:
                    for tk, drip_amt in mc_drip_buys.items():
                        h_ref = next(h for h in holdings if h["ticker"] == tk)
                        adj_price = h_ref["price"] * cum_price_factor[tk]
                        if adj_price > 0:
                            new_sh = drip_amt / adj_price
                            sim_shares_p[tk] += new_sh
                            sim_drip_shares_p[tk] += new_sh

                path_monthly_income[p, m] = month_inc
                path_monthly_existing[p, m] = month_existing
                path_monthly_contrib[p, m] = month_inc - month_existing

        # Calculate percentiles
        median_income = np.median(path_monthly_income, axis=0)
        p10_income = np.percentile(path_monthly_income, 10, axis=0)
        p90_income = np.percentile(path_monthly_income, 90, axis=0)
        median_existing = np.median(path_monthly_existing, axis=0)
        median_contrib = np.median(path_monthly_contrib, axis=0)

        # Build monthly series
        monthly_series = []
        prev_income = None
        for m in range(total_months):
            _, label = _month_label(m + 1)
            inc = float(median_income[m])
            change_dollar = (inc - prev_income) if prev_income is not None else 0.0
            change_pct = (change_dollar / prev_income * 100) if prev_income and prev_income > 0 else 0.0
            monthly_series.append({
                "month": m + 1,
                "label": label,
                "total_income": _safe(inc),
                "income_from_existing": _safe(float(median_existing[m])),
                "income_from_contributions": _safe(float(median_contrib[m])),
                "change_dollar": _safe(change_dollar),
                "change_pct": _safe(change_pct),
                "p10": _safe(float(p10_income[m])),
                "p90": _safe(float(p90_income[m])),
            })
            prev_income = inc

        # Build annual series
        annual_series = []
        current_ann = sum(h["div_per_share"] * h["freq"] * h["shares"] for h in holdings if h["freq"] > 0)
        for yr in range(1, years + 1):
            s, e = (yr - 1) * 12, yr * 12
            yr_income = float(np.sum(median_income[s:e]))
            yr_existing = float(np.sum(median_existing[s:e]))
            yr_contrib = float(np.sum(median_contrib[s:e]))
            yr_p10 = float(np.sum(p10_income[s:e]))
            yr_p90 = float(np.sum(p90_income[s:e]))
            if yr == 1:
                change_d = yr_income - current_ann
                change_p = (change_d / current_ann * 100) if current_ann > 0 else 0
            else:
                ps, pe = (yr - 2) * 12, (yr - 1) * 12
                prev_yr = float(np.sum(median_income[ps:pe]))
                change_d = yr_income - prev_yr
                change_p = (change_d / prev_yr * 100) if prev_yr > 0 else 0

            annual_series.append({
                "year": yr,
                "label": f"Year {yr}",
                "total_income": _safe(yr_income),
                "income_from_existing": _safe(yr_existing),
                "income_from_contributions": _safe(yr_contrib),
                "change_dollar": _safe(change_d),
                "change_pct": _safe(change_p),
                "p10": _safe(yr_p10),
                "p90": _safe(yr_p90),
            })

        # Use deterministic for holdings end-state (MC median is complex per-ticker)
        _, _, sim_shares_det, contrib_shares_det = _run_deterministic()
        return monthly_series, annual_series, sim_shares_det, contrib_shares_det

    # Run selected mode
    if use_monte_carlo:
        monthly_series, annual_series, final_shares, contrib_shares = _run_monte_carlo()
    else:
        monthly_series, annual_series, final_shares, contrib_shares = _run_deterministic()

    # Current income
    current_annual = sum(h["div_per_share"] * h["freq"] * h["shares"] for h in holdings if h["freq"] > 0)
    current_monthly = current_annual / 12

    # Projected income (from last year of annual series)
    projected_annual = annual_series[-1]["total_income"] if annual_series else current_annual

    # Holdings detail
    holdings_out = []
    for h in holdings:
        tk = h["ticker"]
        start_ann = h["div_per_share"] * h["freq"] * h["shares"]
        end_shares = final_shares.get(tk, h["shares"])
        growth_factor = (1 + annual_div_growth) ** years
        end_ann = h["div_per_share"] * growth_factor * h["freq"] * end_shares
        growth_pct = ((end_ann / start_ann - 1) * 100) if start_ann > 0 else 0
        holdings_out.append({
            "ticker": tk,
            "description": h["description"],
            "shares_start": _safe(h["shares"]),
            "shares_end": _safe(end_shares),
            "shares_added": _safe(contrib_shares.get(tk, 0)),
            "frequency": h["freq_str"],
            "current_annual_income": _safe(start_ann),
            "projected_annual_income": _safe(end_ann),
            "growth_pct": _safe(growth_pct),
            "drip": h.get("reinvest", False),
            "drip_shares": _safe(h.get("drip_shares", 0)),
        })

    return jsonify(
        current_monthly_income=_safe(current_monthly),
        current_annual_income=_safe(current_annual),
        monthly_contribution=monthly_contribution,
        total_contributed=_safe(monthly_contribution * total_months),
        monte_carlo=use_monte_carlo,
        monthly_series=monthly_series,
        annual_series=annual_series,
        holdings=holdings_out,
        years=years,
        market_type=market_type,
        projected_annual_income=_safe(projected_annual),
    )


# ── Total Return ──────────────────────────────────────────────────────────────

@app.route("/api/total-return/summary", methods=["GET"])
def total_return_summary():
    """DB-based summary: cards, scatter chart, table rows. No yfinance needed."""
    import math, json
    import plotly.graph_objects as go
    import plotly.utils

    profile_id = get_profile_id()
    conn = get_connection()

    cats = conn.execute(
        "SELECT id, name FROM categories WHERE profile_id = ? ORDER BY sort_order, name",
        (profile_id,),
    ).fetchall()
    categories = [{"id": c["id"], "name": c["name"]} for c in cats]

    cat_param = request.args.get("category", "").strip()
    cat_ids = [c.strip() for c in cat_param.split(",") if c.strip()] if cat_param else []

    rows = conn.execute(
        """SELECT ticker, description, classification_type,
                  price_paid, current_price, quantity,
                  purchase_value, current_value,
                  gain_or_loss, gain_or_loss_percentage,
                  total_divs_received, ytd_divs,
                  estim_payment_per_year, annual_yield_on_cost
           FROM all_account_info
           WHERE purchase_value IS NOT NULL AND purchase_value > 0
             AND profile_id = ?
           ORDER BY ticker""",
        (profile_id,),
    ).fetchall()
    df = pd.DataFrame([dict(r) for r in rows])

    if df.empty:
        conn.close()
        return jsonify({"rows": [], "totals": {}, "scatter": None, "categories": categories})

    # Enrich category names
    try:
        cat_map_rows = conn.execute(
            "SELECT tc.ticker, c.name AS category_name "
            "FROM ticker_categories tc JOIN categories c ON c.id = tc.category_id "
            "WHERE tc.profile_id = ?", (profile_id,)
        ).fetchall()
        cat_map = pd.DataFrame([dict(r) for r in cat_map_rows])
        if not cat_map.empty:
            df = df.merge(cat_map, on="ticker", how="left")
        else:
            df["category_name"] = None
    except Exception:
        df["category_name"] = None

    if "classification_type" in df.columns:
        mask = df["category_name"].isna() | (df["category_name"] == "")
        df.loc[mask, "category_name"] = df.loc[mask, "classification_type"].map(
            lambda c: _CLASSIFICATION_NAMES.get(str(c).strip(), str(c).strip()) if pd.notna(c) else "Other"
        )
    df["category_name"] = df["category_name"].fillna("Other")
    conn.close()

    # Apply category filter
    if cat_ids:
        cat_names = [c["name"] for c in categories if str(c["id"]) in cat_ids]
        if cat_names:
            df = df[df["category_name"].isin(cat_names)]

    if df.empty:
        return jsonify({"rows": [], "totals": {}, "scatter": None, "categories": categories})

    def _safe(v):
        if v is None:
            return None
        try:
            f = float(v)
            return None if (math.isnan(f) or math.isinf(f)) else f
        except (TypeError, ValueError):
            return None

    # Compute return columns
    num_cols = ["price_paid", "current_price", "quantity", "purchase_value",
                "current_value", "gain_or_loss", "total_divs_received",
                "annual_yield_on_cost"]
    for c in num_cols:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce")

    df["total_divs_received"] = df["total_divs_received"].fillna(0)
    df["gain_or_loss"] = df["gain_or_loss"].fillna(0)
    df["total_return_dollar"] = df["gain_or_loss"] + df["total_divs_received"]
    df["total_return_pct"] = (df["total_return_dollar"] / df["purchase_value"].replace(0, float("nan"))) * 100
    df["price_return_pct"] = (df["gain_or_loss"] / df["purchase_value"].replace(0, float("nan"))) * 100

    # Totals
    total_cost = _safe(df["purchase_value"].sum())
    total_curr = _safe(df["current_value"].sum())
    total_gl = _safe(df["gain_or_loss"].sum())
    total_divs = _safe(df["total_divs_received"].sum())
    total_tr = (total_gl or 0) + (total_divs or 0)
    total_tr_pct = round(total_tr / total_cost * 100, 2) if total_cost else 0

    totals = {
        "total_invested": total_cost,
        "current_value": total_curr,
        "price_gl": total_gl,
        "total_divs": total_divs,
        "total_return_dollar": total_tr,
        "total_return_pct": total_tr_pct,
    }

    # Table rows
    table_rows = []
    for _, row in df.sort_values("total_return_pct", ascending=False).iterrows():
        table_rows.append({
            "ticker": row["ticker"],
            "category_name": row.get("category_name", ""),
            "quantity": _safe(row.get("quantity")),
            "price_paid": _safe(row.get("price_paid")),
            "current_price": _safe(row.get("current_price")),
            "purchase_value": _safe(row.get("purchase_value")),
            "current_value": _safe(row.get("current_value")),
            "gain_or_loss": _safe(row.get("gain_or_loss")),
            "price_return_pct": _safe(row.get("price_return_pct")),
            "total_divs_received": _safe(row.get("total_divs_received")),
            "total_return_dollar": _safe(row.get("total_return_dollar")),
            "total_return_pct": _safe(row.get("total_return_pct")),
        })

    # Scatter: Total Return % vs Yield on Cost
    scatter_json = None
    try:
        sdf = df.copy()
        sdf["yield_on_cost_pct"] = sdf["annual_yield_on_cost"].fillna(0) * 100
        sdf["category_name"] = sdf["category_name"].fillna("Other")
        max_pv = float(sdf["purchase_value"].max())
        sdf["bubble_size"] = ((sdf["purchase_value"] / max_pv * 35) + 8).clip(8, 43)

        fig_scatter = go.Figure()
        for ctype, grp in sdf.groupby("category_name"):
            fig_scatter.add_trace(go.Scatter(
                x=grp["yield_on_cost_pct"].tolist(),
                y=grp["total_return_pct"].tolist(),
                mode="markers+text", name=str(ctype),
                text=grp["ticker"].tolist(), textposition="top center",
                textfont=dict(size=9),
                marker=dict(size=grp["bubble_size"].tolist(), opacity=0.8),
                hovertemplate="<b>%{text}</b><br>Total Ret: %{y:.2f}%<br>Yield on Cost: %{x:.2f}%<extra>" + str(ctype) + "</extra>",
            ))
        fig_scatter.add_hline(y=0, line_dash="dash", line_color="gray", opacity=0.5)
        fig_scatter.update_layout(
            title="Total Return % vs Annual Yield on Cost (Since Purchase)",
            template="plotly_dark", height=520,
            xaxis_title="Annual Yield on Cost (%)",
            yaxis_title="Total Return % (Since Purchase)",
            legend_title="Category",
            paper_bgcolor="#1a1f2e", plot_bgcolor="rgba(255,255,255,0.03)",
        )
        scatter_json = json.dumps(fig_scatter, cls=plotly.utils.PlotlyJSONEncoder)
    except Exception:
        pass

    return jsonify(rows=table_rows, totals=totals, scatter=scatter_json, categories=categories)


@app.route("/api/total-return/charts", methods=["GET"])
def total_return_charts():
    """AJAX: yfinance bar + price-history charts for a given period."""
    import math, json, warnings, traceback
    from datetime import date as date_type
    import yfinance as yf
    import plotly.graph_objects as go
    import plotly.utils
    warnings.filterwarnings("ignore")

    period = request.args.get("period", "1y")
    compare = request.args.get("compare", "").strip()

    profile_id = get_profile_id()
    conn = get_connection()

    cat_param = request.args.get("category", "").strip()
    cat_ids = [c.strip() for c in cat_param.split(",") if c.strip()] if cat_param else []

    rows = conn.execute(
        "SELECT ticker, description, classification_type, purchase_value "
        "FROM all_account_info WHERE purchase_value IS NOT NULL AND purchase_value > 0 AND profile_id = ? ORDER BY ticker",
        (profile_id,),
    ).fetchall()
    df = pd.DataFrame([dict(r) for r in rows])

    if df.empty:
        conn.close()
        return jsonify({"error": "No portfolio data"}), 404

    # Enrich with category names and filter
    try:
        cat_map_rows = conn.execute(
            "SELECT tc.ticker, c.name AS category_name "
            "FROM ticker_categories tc JOIN categories c ON c.id = tc.category_id "
            "WHERE tc.profile_id = ?", (profile_id,)
        ).fetchall()
        cat_map = pd.DataFrame([dict(r) for r in cat_map_rows])
        if not cat_map.empty:
            df = df.merge(cat_map, on="ticker", how="left")
        else:
            df["category_name"] = None
    except Exception:
        df["category_name"] = None

    if "classification_type" in df.columns:
        mask = df["category_name"].isna() | (df["category_name"] == "")
        df.loc[mask, "category_name"] = df.loc[mask, "classification_type"].map(
            lambda c: _CLASSIFICATION_NAMES.get(str(c).strip(), str(c).strip()) if pd.notna(c) else "Other"
        )
    df["category_name"] = df["category_name"].fillna("Other")

    cats = conn.execute(
        "SELECT id, name FROM categories WHERE profile_id = ? ORDER BY sort_order, name",
        (profile_id,),
    ).fetchall()
    conn.close()

    if cat_ids:
        cat_names = [c["name"] for c in cats if str(c["id"]) in cat_ids]
        if cat_names:
            df = df[df["category_name"].isin(cat_names)]

    if df.empty:
        return jsonify({"error": "No holdings in selected categories"}), 404

    period_map = {
        "1mo": (dict(period="1mo"), "1d"),
        "3mo": (dict(period="3mo"), "1d"),
        "6mo": (dict(period="6mo"), "1d"),
        "ytd": (dict(period="ytd"), "1d"),
        "1y":  (dict(period="1y"), "1wk"),
        "2y":  (dict(period="2y"), "1wk"),
        "5y":  (dict(period="5y"), "1mo"),
        "max": (dict(period="max"), "1mo"),
    }

    if period.isdigit() and len(period) == 4:
        yr = int(period)
        today = date_type.today()
        start = f"{yr}-01-01"
        end = today.strftime("%Y-%m-%d") if yr == today.year else f"{yr}-12-31"
        yf_kwargs = dict(start=start, end=end)
        yf_interval = "1d" if yr == today.year else "1wk"
    else:
        yf_range, yf_interval = period_map.get(period, (dict(period="1y"), "1wk"))
        yf_kwargs = yf_range

    period_labels = {
        "1mo": "1 Month", "3mo": "3 Months", "6mo": "6 Months",
        "ytd": "Year to Date", "1y": "1 Year", "2y": "2 Years",
        "5y": "5 Years", "max": "All Available",
    }
    if period.isdigit() and len(period) == 4:
        period_labels[period] = f"Calendar {period}"
    period_label = period_labels.get(period, period)

    try:
        tickers_list = df["ticker"].tolist()
        # Parse comparison tickers
        compare_tickers = [t.strip().upper() for t in compare.replace(",", " ").split() if t.strip()] if compare else []
        # Always include SPY as benchmark
        extra = list(set(["SPY"] + compare_tickers))
        all_dl = list(set(tickers_list + extra))

        raw = yf.download(
            " ".join(all_dl), **yf_kwargs, interval=yf_interval,
            progress=False, auto_adjust=True,
        )

        if raw.empty or "Close" not in (raw.columns.get_level_values(0) if isinstance(raw.columns, pd.MultiIndex) else raw.columns):
            return jsonify({"error": "No price data from Yahoo Finance"}), 500

        if isinstance(raw.columns, pd.MultiIndex):
            close = raw["Close"]
        else:
            close = raw[["Close"]]
            close.columns = [all_dl[0]]

        # Normalize to 100
        norm = close.copy()
        for col in norm.columns:
            first_valid = norm[col].first_valid_index()
            if first_valid is not None:
                base = norm.loc[first_valid, col]
                norm[col] = (norm[col] / base) * 100 if base and base != 0 else None

        # Period return per ticker
        returns = {}
        for col in norm.columns:
            s = norm[col].dropna()
            returns[col] = round(float(s.iloc[-1] - 100), 2) if len(s) >= 2 else None

        spy_ret = returns.get("SPY")

        # Bar chart
        ret_df = df.copy()
        ret_df["period_return"] = ret_df["ticker"].map(returns)
        ret_df = ret_df[ret_df["period_return"].notna()].sort_values("period_return", ascending=True)

        if ret_df.empty:
            return jsonify({"error": f"No return data for period: {period_label}"}), 404

        colors = ["#4dff91" if v >= 0 else "#ff6b6b" for v in ret_df["period_return"]]
        fig_bar = go.Figure(go.Bar(
            x=ret_df["period_return"], y=ret_df["ticker"], orientation="h",
            marker_color=colors,
            text=ret_df["period_return"].apply(lambda v: f"{v:.1f}%"),
            textposition="outside",
            hovertemplate="<b>%{y}</b><br>Return: %{x:.2f}%<extra></extra>",
        ))
        if spy_ret is not None:
            fig_bar.add_vline(x=spy_ret, line_dash="dash", line_color="#FFD700",
                              annotation_text=f"SPY: {spy_ret:.1f}%", annotation_position="top")
        fig_bar.update_layout(
            title=f"Total Return % by Ticker - {period_label} (dividend-adjusted)",
            template="plotly_dark", xaxis_title="Total Return (%)", yaxis_title="",
            height=max(500, len(ret_df) * 20),
            margin=dict(l=70, r=80, t=60, b=50),
            paper_bgcolor="#1a1f2e", plot_bgcolor="rgba(255,255,255,0.03)",
        )

        # Price history line chart
        dates = [str(d)[:10] for d in norm.index]
        top5 = set(ret_df.nlargest(5, "period_return")["ticker"].tolist())
        bot5 = set(ret_df.nsmallest(5, "period_return")["ticker"].tolist())
        default_visible = top5 | bot5

        fig_hist = go.Figure()

        # Add SPY + comparison tickers first
        for etk in extra:
            if etk not in norm.columns:
                continue
            is_spy = etk == "SPY"
            fig_hist.add_trace(go.Scatter(
                x=dates, y=norm[etk].tolist(),
                name=f"{etk} (Benchmark)" if is_spy else etk,
                line=dict(color="#FFD700" if is_spy else None, width=3 if is_spy else 2, dash="dash" if is_spy else None),
                visible=True,
            ))

        for tkr in tickers_list:
            if tkr in extra or tkr not in norm.columns:
                continue
            vals = norm[tkr].tolist()
            desc_vals = df.loc[df["ticker"] == tkr, "description"].values
            hover_name = desc_vals[0][:35] if len(desc_vals) and pd.notna(desc_vals[0]) else tkr
            visible = True if tkr in default_visible else "legendonly"
            fig_hist.add_trace(go.Scatter(
                x=dates, y=vals, name=tkr, visible=visible,
                hovertemplate=f"<b>{tkr}</b> - {hover_name}<br>Normalized: %{{y:.1f}}<br>Date: %{{x}}<extra></extra>",
                line=dict(width=1.5),
            ))

        fig_hist.update_layout(
            title=f"Price Performance - {period_label} (normalized to 100, dividend-adjusted)",
            template="plotly_dark", xaxis_title="Date",
            yaxis_title="Normalized Price (100 = start)", height=550,
            legend=dict(orientation="v", x=1.01, y=1, font=dict(size=10)),
            hovermode="x unified",
            paper_bgcolor="#1a1f2e", plot_bgcolor="rgba(255,255,255,0.03)",
        )

        def _clean_val(v):
            if v is None:
                return None
            try:
                if math.isnan(v) or math.isinf(v):
                    return None
            except (TypeError, ValueError):
                pass
            return v

        return jsonify({
            "bar": json.loads(json.dumps(fig_bar, cls=plotly.utils.PlotlyJSONEncoder)),
            "history": json.loads(json.dumps(fig_hist, cls=plotly.utils.PlotlyJSONEncoder)),
            "spy_ret": _clean_val(spy_ret),
            "period_label": period_label,
        })

    except Exception as e:
        return jsonify({"error": str(e), "detail": traceback.format_exc(limit=3)}), 500


@app.route("/api/total-return/compare", methods=["GET"])
def total_return_compare():
    """Comparison chart: price return and total return for selected tickers."""
    import math, warnings
    import yfinance as yf
    warnings.filterwarnings("ignore")

    tickers_param = request.args.get("tickers", "").strip()
    extra_param = request.args.get("extra", "").strip()
    period = request.args.get("period", "1y")

    selected = [t.strip().upper() for t in tickers_param.replace(",", " ").split() if t.strip()]
    extras = [t.strip().upper() for t in extra_param.replace(",", " ").split() if t.strip()]
    all_tickers = list(dict.fromkeys(selected + extras))  # dedupe, preserve order

    if not all_tickers:
        return jsonify({"error": "No tickers selected"}), 400

    period_map = {
        "3mo": dict(period="3mo"), "6mo": dict(period="6mo"),
        "9mo": dict(period="9mo"), "1y": dict(period="1y"),
        "2y": dict(period="2y"), "3y": dict(period="3y"),
        "4y": dict(period="4y"), "5y": dict(period="5y"),
    }
    yf_kwargs = period_map.get(period, dict(period="1y"))
    period_labels = {
        "3mo": "3 Months", "6mo": "6 Months", "9mo": "9 Months",
        "1y": "1 Year", "2y": "2 Years", "3y": "3 Years",
        "4y": "4 Years", "5y": "5 Years",
    }
    period_label = period_labels.get(period, period)

    try:
        raw = yf.download(
            " ".join(all_tickers), **yf_kwargs,
            auto_adjust=False, actions=True, progress=False,
        )
        if raw.empty:
            return jsonify({"error": "No data from Yahoo Finance"}), 500

        if isinstance(raw.columns, pd.MultiIndex):
            close = raw["Close"]
            divs = raw["Dividends"].fillna(0) if "Dividends" in raw.columns.get_level_values(0) else pd.DataFrame(0, index=close.index, columns=close.columns)
        else:
            close = raw[["Close"]]
            close.columns = [all_tickers[0]]
            divs = raw[["Dividends"]].fillna(0) if "Dividends" in raw.columns else pd.DataFrame(0, index=close.index, columns=[all_tickers[0]])
            divs.columns = [all_tickers[0]]

        dates = [d.strftime("%Y-%m-%d") for d in close.index]

        def _clean(v):
            if v is None:
                return None
            try:
                f = float(v)
                return None if (math.isnan(f) or math.isinf(f)) else round(f, 2)
            except (TypeError, ValueError):
                return None

        price_series = {}
        total_series = {}

        for t in all_tickers:
            if t not in close.columns:
                continue
            c = close[t].dropna()
            if len(c) < 2:
                continue
            base = float(c.iloc[0])
            if base == 0:
                continue
            # Price return: normalized to 100
            price_norm = (c / base * 100)
            price_series[t] = [_clean(v) for v in price_norm]

            # Total return: price + cumulative dividends
            d = divs[t].reindex(close.index).fillna(0) if t in divs.columns else pd.Series(0, index=close.index)
            cum_div = d.cumsum()
            total_norm = ((c + cum_div) / base * 100)
            total_series[t] = [_clean(v) for v in total_norm]

        return jsonify({
            "dates": dates,
            "price": price_series,
            "total": total_series,
            "tickers": [t for t in all_tickers if t in price_series],
            "period_label": period_label,
        })

    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "detail": traceback.format_exc(limit=3)}), 500


# ── Gains & Losses ─────────────────────────────────────────────────────────────

@app.route("/api/gains-losses/summary", methods=["GET"])
def gains_losses_summary():
    """Unified unrealized + realized gains/losses with price-only and total (price+divs) columns."""
    import math

    profile_id = get_profile_id()
    conn = get_connection()

    cats = conn.execute(
        "SELECT id, name FROM categories WHERE profile_id = ? ORDER BY sort_order, name",
        (profile_id,),
    ).fetchall()
    categories = [{"id": c["id"], "name": c["name"]} for c in cats]

    cat_param = request.args.get("category", "").strip()
    cat_ids = [c.strip() for c in cat_param.split(",") if c.strip()] if cat_param else []

    # ── Unrealized (current holdings) ──
    rows = conn.execute(
        """SELECT ticker, description, classification_type,
                  price_paid, current_price, quantity,
                  purchase_value, current_value,
                  gain_or_loss, total_divs_received, purchase_date
           FROM all_account_info
           WHERE purchase_value IS NOT NULL AND purchase_value > 0
             AND profile_id = ?
           ORDER BY ticker""",
        (profile_id,),
    ).fetchall()
    udf = pd.DataFrame([dict(r) for r in rows])

    # Enrich category names
    if not udf.empty:
        try:
            cat_map_rows = conn.execute(
                "SELECT tc.ticker, c.name AS category_name "
                "FROM ticker_categories tc JOIN categories c ON c.id = tc.category_id "
                "WHERE tc.profile_id = ?", (profile_id,)
            ).fetchall()
            cat_map = pd.DataFrame([dict(r) for r in cat_map_rows])
            if not cat_map.empty:
                cat_map = cat_map.drop_duplicates(subset="ticker", keep="first")
                udf = udf.merge(cat_map, on="ticker", how="left")
            else:
                udf["category_name"] = None
        except Exception:
            udf["category_name"] = None

        if "classification_type" in udf.columns:
            mask = udf["category_name"].isna() | (udf["category_name"] == "")
            udf.loc[mask, "category_name"] = udf.loc[mask, "classification_type"].map(
                lambda c: _CLASSIFICATION_NAMES.get(str(c).strip(), str(c).strip()) if pd.notna(c) else "Other"
            )
        udf["category_name"] = udf["category_name"].fillna("Other")

        if cat_ids:
            cat_names = [c["name"] for c in categories if str(c["id"]) in cat_ids]
            if cat_names:
                udf = udf[udf["category_name"].isin(cat_names)]

    # ── Realized (sold positions) ──
    # 1) Legacy watchlist_sold table
    sold_rows = conn.execute(
        "SELECT ticker, buy_price, sell_price, shares_sold, sell_date, divs_received, notes "
        "FROM watchlist_sold ORDER BY sell_date DESC, id DESC"
    ).fetchall()
    rdf = pd.DataFrame([dict(r) for r in sold_rows]) if sold_rows else pd.DataFrame()

    # 2) Transactions-based SELL records
    # Owner (profile 1) is the master combined view — show sells from all profiles
    if profile_id == 1:
        txn_sell_rows = conn.execute(
            """SELECT t.ticker, t.profile_id, t.price_per_share AS sell_price,
                      t.shares AS shares_sold, t.transaction_date AS sell_date,
                      t.realized_gain, t.fees, t.notes
               FROM transactions t
               WHERE t.transaction_type = 'SELL'
               ORDER BY t.transaction_date DESC, t.id DESC"""
        ).fetchall()
    else:
        txn_sell_rows = conn.execute(
            """SELECT t.ticker, t.profile_id, t.price_per_share AS sell_price,
                      t.shares AS shares_sold, t.transaction_date AS sell_date,
                      t.realized_gain, t.fees, t.notes
               FROM transactions t
               WHERE t.transaction_type = 'SELL' AND t.profile_id = ?
               ORDER BY t.transaction_date DESC, t.id DESC""",
            (profile_id,),
        ).fetchall()
    if txn_sell_rows:
        # Look up total_divs_received per ticker/profile from dividends table
        div_lookup = {}
        div_rows = conn.execute(
            "SELECT ticker, profile_id, total_divs_received FROM dividends"
        ).fetchall()
        for dr in div_rows:
            dr = dict(dr)
            div_lookup[(dr["ticker"], dr["profile_id"])] = float(dr.get("total_divs_received") or 0)

        txn_rows = []
        for tr in txn_sell_rows:
            tr = dict(tr)
            sp = float(tr.get("sell_price") or 0)
            sh = float(tr.get("shares_sold") or 0)
            rg = float(tr.get("realized_gain") or 0)
            fees = float(tr.get("fees") or 0)
            proceeds = sp * sh
            cost = proceeds - rg + fees
            bp = cost / sh if sh else 0
            divs = div_lookup.get((tr["ticker"], tr["profile_id"]), 0)
            txn_rows.append({
                "ticker": tr["ticker"],
                "buy_price": bp,
                "sell_price": sp,
                "shares_sold": sh,
                "sell_date": tr.get("sell_date", ""),
                "divs_received": divs,
                "notes": tr.get("notes") or "",
            })
        txn_df = pd.DataFrame(txn_rows)
        rdf = pd.concat([rdf, txn_df], ignore_index=True) if not rdf.empty else txn_df

    conn.close()

    def _safe(v):
        if v is None or (isinstance(v, float) and (math.isnan(v) or math.isinf(v))):
            return None
        try:
            f = float(v)
            return None if (math.isnan(f) or math.isinf(f)) else round(f, 2)
        except (TypeError, ValueError):
            return None

    def _safe_str(v):
        """Sanitize string/date values — return None for NaN/NaT."""
        if v is None:
            return None
        if isinstance(v, float) and math.isnan(v):
            return None
        if pd.isna(v):
            return None
        return str(v)

    # ── Build unrealized rows ──
    unrealized = []
    u_totals = {"invested": 0, "value": 0, "price_gl": 0, "divs": 0, "total_gl": 0}

    if not udf.empty:
        for c in ["price_paid", "current_price", "quantity", "purchase_value",
                   "current_value", "gain_or_loss", "total_divs_received"]:
            if c in udf.columns:
                udf[c] = pd.to_numeric(udf[c], errors="coerce")

        udf["gain_or_loss"] = udf["gain_or_loss"].fillna(0)
        udf["total_divs_received"] = udf["total_divs_received"].fillna(0)
        udf["total_gl"] = udf["gain_or_loss"] + udf["total_divs_received"]
        udf["price_gl_pct"] = (udf["gain_or_loss"] / udf["purchase_value"].replace(0, float("nan"))) * 100
        udf["total_gl_pct"] = (udf["total_gl"] / udf["purchase_value"].replace(0, float("nan"))) * 100

        for _, row in udf.sort_values("total_gl", ascending=False).iterrows():
            pv = float(row.get("purchase_value") or 0)
            cv = float(row.get("current_value") or 0)
            pgl = float(row.get("gain_or_loss") or 0)
            dvs = float(row.get("total_divs_received") or 0)
            tgl = pgl + dvs
            u_totals["invested"] += pv
            u_totals["value"] += cv
            u_totals["price_gl"] += pgl
            u_totals["divs"] += dvs
            u_totals["total_gl"] += tgl
            unrealized.append({
                "ticker": row["ticker"],
                "description": row.get("description", ""),
                "category_name": row.get("category_name", ""),
                "quantity": _safe(row.get("quantity")),
                "price_paid": _safe(row.get("price_paid")),
                "current_price": _safe(row.get("current_price")),
                "purchase_value": _safe(pv),
                "current_value": _safe(cv),
                "price_gl": _safe(pgl),
                "price_gl_pct": _safe(row.get("price_gl_pct")),
                "divs_received": _safe(dvs),
                "total_gl": _safe(tgl),
                "total_gl_pct": _safe(row.get("total_gl_pct")),
                "purchase_date": _safe_str(row.get("purchase_date")),
            })

    # ── Build realized rows ──
    realized = []
    r_totals = {"cost": 0, "proceeds": 0, "price_gl": 0, "divs": 0, "total_gl": 0}

    if not rdf.empty:
        for c in ["buy_price", "sell_price", "shares_sold", "divs_received"]:
            if c in rdf.columns:
                rdf[c] = pd.to_numeric(rdf[c], errors="coerce").fillna(0)

        for _, row in rdf.iterrows():
            bp = float(row.get("buy_price") or 0)
            sp = float(row.get("sell_price") or 0)
            sh = float(row.get("shares_sold") or 0)
            dv = float(row.get("divs_received") or 0)
            cost = bp * sh
            proceeds = sp * sh
            pgl = proceeds - cost
            tgl = pgl + dv
            pgl_pct = (pgl / cost * 100) if cost else 0
            tgl_pct = (tgl / cost * 100) if cost else 0
            r_totals["cost"] += cost
            r_totals["proceeds"] += proceeds
            r_totals["price_gl"] += pgl
            r_totals["divs"] += dv
            r_totals["total_gl"] += tgl
            realized.append({
                "ticker": row["ticker"],
                "buy_price": _safe(bp),
                "sell_price": _safe(sp),
                "shares_sold": _safe(sh),
                "sell_date": _safe_str(row.get("sell_date")) or "",
                "cost_basis": _safe(cost),
                "proceeds": _safe(proceeds),
                "price_gl": _safe(pgl),
                "price_gl_pct": _safe(pgl_pct),
                "divs_received": _safe(dv),
                "total_gl": _safe(tgl),
                "total_gl_pct": _safe(tgl_pct),
                "notes": row.get("notes", ""),
            })

    # ── Combined (one row per ticker across unrealized + realized) ──
    combined_map = {}
    for r in unrealized:
        t = r["ticker"]
        combined_map[t] = {
            "ticker": t, "description": r["description"],
            "unrealized_price_gl": r["price_gl"] or 0,
            "unrealized_divs": r["divs_received"] or 0,
            "unrealized_total_gl": r["total_gl"] or 0,
            "realized_price_gl": 0, "realized_divs": 0, "realized_total_gl": 0,
            "status": "Open",
        }
    for r in realized:
        t = r["ticker"]
        if t not in combined_map:
            combined_map[t] = {
                "ticker": t, "description": "",
                "unrealized_price_gl": 0, "unrealized_divs": 0, "unrealized_total_gl": 0,
                "realized_price_gl": 0, "realized_divs": 0, "realized_total_gl": 0,
                "status": "Closed",
            }
        entry = combined_map[t]
        entry["realized_price_gl"] += (r["price_gl"] or 0)
        entry["realized_divs"] += (r["divs_received"] or 0)
        entry["realized_total_gl"] += (r["total_gl"] or 0)
        if entry["unrealized_total_gl"]:
            entry["status"] = "Open + Closed"

    combined = []
    for t, entry in combined_map.items():
        entry["net_price_gl"] = _safe(entry["unrealized_price_gl"] + entry["realized_price_gl"])
        entry["net_divs"] = _safe(entry["unrealized_divs"] + entry["realized_divs"])
        entry["net_total_gl"] = _safe(entry["unrealized_total_gl"] + entry["realized_total_gl"])
        entry["unrealized_price_gl"] = _safe(entry["unrealized_price_gl"])
        entry["unrealized_total_gl"] = _safe(entry["unrealized_total_gl"])
        entry["realized_price_gl"] = _safe(entry["realized_price_gl"])
        entry["realized_total_gl"] = _safe(entry["realized_total_gl"])
        entry["unrealized_divs"] = _safe(entry["unrealized_divs"])
        entry["realized_divs"] = _safe(entry["realized_divs"])
        combined.append(entry)
    combined.sort(key=lambda x: x["net_total_gl"] or 0, reverse=True)

    totals = {
        "unrealized_invested": _safe(u_totals["invested"]),
        "unrealized_value": _safe(u_totals["value"]),
        "unrealized_price_gl": _safe(u_totals["price_gl"]),
        "unrealized_divs": _safe(u_totals["divs"]),
        "unrealized_total_gl": _safe(u_totals["total_gl"]),
        "realized_cost": _safe(r_totals["cost"]),
        "realized_proceeds": _safe(r_totals["proceeds"]),
        "realized_price_gl": _safe(r_totals["price_gl"]),
        "realized_divs": _safe(r_totals["divs"]),
        "realized_total_gl": _safe(r_totals["total_gl"]),
        "combined_price_gl": _safe(u_totals["price_gl"] + r_totals["price_gl"]),
        "combined_divs": _safe(u_totals["divs"] + r_totals["divs"]),
        "combined_total_gl": _safe(u_totals["total_gl"] + r_totals["total_gl"]),
    }

    return jsonify(
        unrealized=unrealized, realized=realized, combined=combined,
        totals=totals, categories=categories,
    )


@app.route("/api/gains-losses/chart", methods=["GET"])
def gains_losses_chart():
    """Cumulative portfolio G/L over time using yfinance, plus realized events."""
    import math, warnings
    from datetime import date as date_type
    import yfinance as yf
    warnings.filterwarnings("ignore")

    period = request.args.get("period", "1y")
    profile_id = get_profile_id()
    conn = get_connection()

    cat_param = request.args.get("category", "").strip()
    cat_ids = [c.strip() for c in cat_param.split(",") if c.strip()] if cat_param else []

    rows = conn.execute(
        """SELECT ticker, quantity, price_paid, purchase_value, total_divs_received,
                  classification_type
           FROM all_account_info
           WHERE purchase_value IS NOT NULL AND purchase_value > 0 AND profile_id = ?
           ORDER BY ticker""",
        (profile_id,),
    ).fetchall()
    hdf = pd.DataFrame([dict(r) for r in rows])

    if hdf.empty:
        conn.close()
        return jsonify({"error": "No portfolio data"}), 404

    # Category filter
    if cat_ids:
        try:
            cat_map_rows = conn.execute(
                "SELECT tc.ticker, c.name AS category_name "
                "FROM ticker_categories tc JOIN categories c ON c.id = tc.category_id "
                "WHERE tc.profile_id = ?", (profile_id,)
            ).fetchall()
            cat_map = pd.DataFrame([dict(r) for r in cat_map_rows])
            if not cat_map.empty:
                hdf = hdf.merge(cat_map, on="ticker", how="left")
            else:
                hdf["category_name"] = None
            if "classification_type" in hdf.columns:
                mask = hdf["category_name"].isna() | (hdf["category_name"] == "")
                hdf.loc[mask, "category_name"] = hdf.loc[mask, "classification_type"].map(
                    lambda c: _CLASSIFICATION_NAMES.get(str(c).strip(), str(c).strip()) if pd.notna(c) else "Other"
                )
            hdf["category_name"] = hdf["category_name"].fillna("Other")

            cats = conn.execute(
                "SELECT id, name FROM categories WHERE profile_id = ? ORDER BY sort_order, name",
                (profile_id,),
            ).fetchall()
            cat_names = [c["name"] for c in cats if str(c["id"]) in cat_ids]
            if cat_names:
                hdf = hdf[hdf["category_name"].isin(cat_names)]
        except Exception:
            pass

    # Realized events (from watchlist_sold + transactions)
    sold_rows = conn.execute(
        "SELECT ticker, buy_price, sell_price, shares_sold, sell_date, divs_received "
        "FROM watchlist_sold WHERE sell_date IS NOT NULL AND sell_date != '' ORDER BY sell_date"
    ).fetchall()

    realized_events = []
    for sr in sold_rows:
        sr = dict(sr)
        try:
            bp = float(sr.get("buy_price") or 0)
            sp = float(sr.get("sell_price") or 0)
            sh = float(sr.get("shares_sold") or 0)
            dv = float(sr.get("divs_received") or 0)
            pgl = (sp - bp) * sh
            tgl = pgl + dv
            realized_events.append({
                "date": sr["sell_date"], "ticker": sr["ticker"],
                "price_gl": round(pgl, 2), "total_gl": round(tgl, 2),
            })
        except (TypeError, ValueError):
            pass

    # Also include SELL transactions
    # Owner (profile 1) is the master combined view — show sells from all profiles
    if profile_id == 1:
        txn_sell = conn.execute(
            """SELECT t.ticker, t.profile_id, t.price_per_share, t.shares,
                      t.transaction_date, t.realized_gain
               FROM transactions t
               WHERE t.transaction_type = 'SELL'
                 AND t.transaction_date IS NOT NULL AND t.transaction_date != ''
               ORDER BY t.transaction_date"""
        ).fetchall()
    else:
        txn_sell = conn.execute(
            """SELECT t.ticker, t.profile_id, t.price_per_share, t.shares,
                      t.transaction_date, t.realized_gain
               FROM transactions t
               WHERE t.transaction_type = 'SELL' AND t.profile_id = ?
                 AND t.transaction_date IS NOT NULL AND t.transaction_date != ''
               ORDER BY t.transaction_date""",
            (profile_id,),
        ).fetchall()
    # Look up dividends for total G/L
    chart_div_lookup = {}
    chart_div_rows = conn.execute(
        "SELECT ticker, profile_id, total_divs_received FROM dividends"
    ).fetchall()
    for dr in chart_div_rows:
        dr = dict(dr)
        chart_div_lookup[(dr["ticker"], dr["profile_id"])] = float(dr.get("total_divs_received") or 0)

    for tr in txn_sell:
        tr = dict(tr)
        try:
            rg = float(tr.get("realized_gain") or 0)
            divs = chart_div_lookup.get((tr["ticker"], tr["profile_id"]), 0)
            realized_events.append({
                "date": tr["transaction_date"], "ticker": tr["ticker"],
                "price_gl": round(rg, 2), "total_gl": round(rg + divs, 2),
            })
        except (TypeError, ValueError):
            pass
    realized_events.sort(key=lambda x: x["date"])

    conn.close()

    if hdf.empty:
        return jsonify({"error": "No holdings in selected categories"}), 404

    for c in ["quantity", "price_paid", "purchase_value", "total_divs_received"]:
        hdf[c] = pd.to_numeric(hdf[c], errors="coerce").fillna(0)

    period_map = {
        "3mo": dict(period="3mo"), "6mo": dict(period="6mo"),
        "1y":  dict(period="1y"),  "2y":  dict(period="2y"),
        "3y":  dict(period="3y"),  "5y":  dict(period="5y"),
    }
    yf_kwargs = period_map.get(period, dict(period="1y"))
    period_labels = {
        "3mo": "3 Months", "6mo": "6 Months", "1y": "1 Year",
        "2y": "2 Years", "3y": "3 Years", "5y": "5 Years",
    }
    period_label = period_labels.get(period, period)

    def _safe(v):
        if v is None:
            return None
        try:
            f = float(v)
            return None if (math.isnan(f) or math.isinf(f)) else round(f, 2)
        except (TypeError, ValueError):
            return None

    try:
        tickers_list = hdf["ticker"].tolist()
        raw = yf.download(
            " ".join(tickers_list), **yf_kwargs, interval="1d",
            progress=False, auto_adjust=True,
        )

        if raw.empty or "Close" not in (raw.columns.get_level_values(0) if isinstance(raw.columns, pd.MultiIndex) else raw.columns):
            return jsonify({"error": "No price data from Yahoo Finance"}), 500

        if isinstance(raw.columns, pd.MultiIndex):
            close = raw["Close"]
        else:
            close = raw[["Close"]]
            close.columns = [tickers_list[0]]

        # Build portfolio cost basis and shares vectors
        total_cost = hdf["purchase_value"].sum()

        # Cumulative portfolio value over time
        dates = close.index
        port_values = pd.Series(0.0, index=dates)
        for _, h in hdf.iterrows():
            t = h["ticker"]
            if t in close.columns:
                port_values += close[t].ffill().fillna(0) * h["quantity"]

        price_gl = (port_values - total_cost).round(2)

        # For total G/L, prorate total_divs_received linearly across the period
        total_divs = hdf["total_divs_received"].sum()
        n_days = len(dates)
        div_series = pd.Series(
            [total_divs * (i + 1) / n_days for i in range(n_days)],
            index=dates,
        )
        total_gl = (price_gl + div_series).round(2)

        # Downsample for performance
        step = max(1, n_days // 250)
        sampled_dates = [dates[i].strftime("%Y-%m-%d") for i in range(0, n_days, step)]
        sampled_price_gl = [_safe(price_gl.iloc[i]) for i in range(0, n_days, step)]
        sampled_total_gl = [_safe(total_gl.iloc[i]) for i in range(0, n_days, step)]

        # Per-ticker G/L for bar chart
        ticker_gl = []
        for _, h in hdf.iterrows():
            t = h["ticker"]
            if t in close.columns:
                last = close[t].dropna()
                if len(last) > 0:
                    cv = float(last.iloc[-1]) * h["quantity"]
                    pv = h["purchase_value"]
                    dv = h["total_divs_received"]
                    ticker_gl.append({
                        "ticker": t,
                        "price_gl": _safe(cv - pv),
                        "total_gl": _safe(cv - pv + dv),
                    })
        ticker_gl.sort(key=lambda x: x["total_gl"] or 0, reverse=True)

        return jsonify(
            dates=sampled_dates,
            price_gl=sampled_price_gl,
            total_gl=sampled_total_gl,
            period_label=period_label,
            realized_events=realized_events,
            ticker_gl=ticker_gl,
        )

    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "detail": traceback.format_exc(limit=3)}), 500


# ── Growth ─────────────────────────────────────────────────────────────────────

@app.route("/api/growth/data", methods=["GET"])
def growth_data():
    """Portfolio growth charts, heatmap, correlation, and grading."""
    import math
    import warnings
    import numpy as np
    import yfinance as yf
    from grading import grade_portfolio, letter_grade, _sharpe, _sortino
    warnings.filterwarnings("ignore")

    def _clean(v):
        """Convert NaN/Inf to None for JSON safety."""
        if v is None:
            return None
        try:
            if math.isnan(v) or math.isinf(v):
                return None
        except (TypeError, ValueError):
            pass
        return v

    profile_id = get_profile_id()
    period = request.args.get("period", "1y")
    benchmark = request.args.get("benchmark", "SPY").upper().strip()
    category_id = request.args.get("category")

    if period not in ("1y", "5y", "max"):
        period = "1y"

    conn = get_connection()

    # Fetch holdings (optionally filtered by category)
    if category_id:
        cat_ids = [int(c) for c in category_id.split(",") if c.strip().isdigit()]
        placeholders = ",".join("?" * len(cat_ids))
        rows = conn.execute(
            f"""SELECT DISTINCT a.ticker, a.quantity, a.current_value
               FROM all_account_info a
               JOIN ticker_categories tc ON a.ticker = tc.ticker AND a.profile_id = tc.profile_id
               WHERE a.profile_id = ? AND tc.category_id IN ({placeholders})
                 AND a.purchase_value > 0 AND a.quantity > 0""",
            [profile_id] + cat_ids,
        ).fetchall()
    else:
        rows = conn.execute(
            """SELECT ticker, quantity, current_value
               FROM all_account_info
               WHERE profile_id = ? AND purchase_value > 0 AND quantity > 0""",
            (profile_id,),
        ).fetchall()

    # Fetch categories for filter dropdown
    cats = conn.execute(
        "SELECT id, name FROM categories WHERE profile_id = ? ORDER BY sort_order, name",
        (profile_id,),
    ).fetchall()
    conn.close()

    categories = [{"id": c["id"], "name": c["name"]} for c in cats]

    if not rows:
        return jsonify({"error": "No holdings found", "categories": categories}), 400

    tickers = [r["ticker"] for r in rows]
    quantities = {r["ticker"]: float(r["quantity"] or 0) for r in rows}
    values = {r["ticker"]: float(r["current_value"] or 0) for r in rows}
    all_dl = list(set(tickers + [benchmark]))

    try:
        raw = yf.download(
            " ".join(all_dl), period=period, auto_adjust=True,
            actions=True, progress=False
        )
        if raw.empty:
            return jsonify({"error": "No price data"}), 500
    except Exception as e:
        return jsonify({"error": str(e)}), 500

    # Parse close prices and dividends
    if isinstance(raw.columns, pd.MultiIndex):
        close = raw["Close"].dropna(how="all")
        divs = raw["Dividends"].fillna(0) if "Dividends" in raw.columns.get_level_values(0) else pd.DataFrame(0, index=close.index, columns=close.columns)
    else:
        close = raw[["Close"]].dropna(how="all")
        close.columns = [all_dl[0]]
        divs = raw[["Dividends"]].fillna(0) if "Dividends" in raw.columns else pd.DataFrame(0, index=close.index, columns=[all_dl[0]])
        divs.columns = [all_dl[0]]

    # Filter to tickers actually present in data
    available_tickers = [t for t in tickers if t in close.columns]
    if not available_tickers:
        return jsonify({"error": "No price data for holdings"}), 500

    # ── Portfolio value series (price-only and total return) ──
    first_valid = close[available_tickers].dropna(how="all").index[0]
    close_aligned = close[available_tickers].loc[first_valid:].ffill().bfill()
    divs_aligned = divs[available_tickers].loc[first_valid:].fillna(0) if set(available_tickers).issubset(divs.columns) else pd.DataFrame(0, index=close_aligned.index, columns=available_tickers)

    # Cumulative dividends per share
    cum_divs = divs_aligned.cumsum()

    # Portfolio price-only value
    port_price = pd.Series(0.0, index=close_aligned.index)
    port_total = pd.Series(0.0, index=close_aligned.index)
    for t in available_tickers:
        q = quantities.get(t, 0)
        port_price += close_aligned[t] * q
        port_total += (close_aligned[t] + cum_divs[t]) * q

    # Normalize to 100
    p0 = port_price.iloc[0]
    t0 = port_total.iloc[0]
    port_price_norm = (port_price / p0 * 100) if p0 > 0 else port_price
    port_total_norm = (port_total / t0 * 100) if t0 > 0 else port_total

    # ── Benchmark series ──
    bench_price_norm = pd.Series(dtype=float)
    bench_total_norm = pd.Series(dtype=float)
    if benchmark in close.columns:
        bc = close[benchmark].loc[first_valid:].ffill().bfill()
        bd = divs[benchmark].loc[first_valid:].fillna(0) if benchmark in divs.columns else pd.Series(0, index=bc.index)
        bc_cum_div = bd.cumsum()
        b0 = bc.iloc[0]
        bench_price_norm = (bc / b0 * 100) if b0 > 0 else bc
        bench_total_norm = ((bc + bc_cum_div) / b0 * 100) if b0 > 0 else bc

    dates_str = [d.strftime("%Y-%m-%d") for d in port_price_norm.index]

    # ── Per-ticker returns for bar chart ──
    def pct_return(series, n):
        if len(series) < n + 1:
            return None
        v = float((series.iloc[-1] / series.iloc[-n] - 1) * 100)
        return _clean(round(v, 2))

    def ytd_return(series):
        yr = series.index[-1].year
        yr_start = series.loc[series.index >= f"{yr}-01-01"]
        if len(yr_start) < 2:
            return None
        v = float((yr_start.iloc[-1] / yr_start.iloc[0] - 1) * 100)
        return _clean(round(v, 2))

    windows = {"1M": 21, "3M": 63, "6M": 126, "1Y": 252}
    ticker_returns = []
    for t in available_tickers:
        tc = close_aligned[t]
        row = {"ticker": t}
        for label, n in windows.items():
            row[label] = pct_return(tc, n)
        row["YTD"] = ytd_return(tc)
        ticker_returns.append(row)
    ticker_returns.sort(key=lambda r: r.get("1Y") or 0, reverse=True)

    # ── Heatmap ──
    heatmap_windows = {"1D": 1, "7D": 5, "1M": 21, "3M": 63, "6M": 126, "1Y": 252}
    heatmap_tickers = [r["ticker"] for r in ticker_returns]
    heatmap_labels = list(heatmap_windows.keys()) + ["YTD"]
    heatmap_values = []
    for t in heatmap_tickers:
        tc = close_aligned[t]
        row = []
        for label, n in heatmap_windows.items():
            row.append(pct_return(tc, n))
        row.append(ytd_return(tc))
        heatmap_values.append(row)

    # ── Correlation matrix ──
    daily_returns = close_aligned[available_tickers].pct_change().dropna()
    corr = daily_returns.corr()
    corr_tickers = list(corr.columns)
    corr_matrix = [[_clean(round(float(corr.iloc[i, j]), 3)) for j in range(len(corr_tickers))] for i in range(len(corr_tickers))]

    # ── Portfolio grade ──
    grade_info = {}
    if len(available_tickers) >= 2 and len(daily_returns) >= 30:
        weights_arr = np.array([values.get(t, 0.0) for t in available_tickers])
        bench_ret = daily_returns[benchmark].dropna() if benchmark in daily_returns.columns else None
        returns_for_grade = daily_returns[available_tickers]
        pm = grade_portfolio(returns_for_grade, weights_arr, bench_ret)
        g = pm.get("grade", {})
        grade_info = {
            "overall": g.get("overall", "N/A"),
            "score": _clean(g.get("score")),
            "sharpe": _clean(pm.get("sharpe")),
            "sortino": _clean(pm.get("sortino")),
        }

    # ── Benchmark metrics ──
    benchmark_metrics = {}
    if benchmark in close.columns:
        bc_full = close[benchmark].loc[first_valid:].dropna()
        benchmark_metrics = {
            "sharpe": _clean(_sharpe(bc_full)),
            "sortino": _clean(_sortino(bc_full)),
        }

    def _clean_series(s):
        return [_clean(round(float(v), 2)) for v in s]

    return jsonify(
        portfolio_price={"dates": dates_str, "values": _clean_series(port_price_norm)},
        portfolio_total={"dates": dates_str, "values": _clean_series(port_total_norm)},
        benchmark_price={"dates": [d.strftime("%Y-%m-%d") for d in bench_price_norm.index] if len(bench_price_norm) else [], "values": _clean_series(bench_price_norm)},
        benchmark_total={"dates": [d.strftime("%Y-%m-%d") for d in bench_total_norm.index] if len(bench_total_norm) else [], "values": _clean_series(bench_total_norm)},
        benchmark_ticker=benchmark,
        ticker_returns=ticker_returns,
        heatmap={"tickers": heatmap_tickers, "windows": heatmap_labels, "values": heatmap_values},
        correlation={"tickers": corr_tickers, "matrix": corr_matrix},
        grade=grade_info,
        benchmark_metrics=benchmark_metrics,
        categories=categories,
    )


# ── ETF Screen ─────────────────────────────────────────────────────────────────

def _blend_price_drip(close_series, divs_series, frac, track_cash=True):
    """Simulate reinvestment of dividends.

    frac=0 → price + cash dividends (no reinvestment)
    frac=1 → full DRIP (all dividends reinvested)
    Returns a pandas Series normalised to 100 at start.
    """
    shares = 1.0
    cash_divs = 0.0
    start_price = float(close_series.iloc[0])
    vals = []
    for i in range(len(close_series)):
        price = float(close_series.iloc[i])
        d = float(divs_series.iloc[i]) if i < len(divs_series) else 0.0
        if d > 0 and price > 0:
            reinvest = d * shares * frac
            if track_cash:
                cash_divs += d * shares * (1 - frac)
            shares += reinvest / price
        vals.append((shares * price + cash_divs) / start_price * 100)
    return pd.Series(vals, index=close_series.index)


@app.route("/api/etf-screen/data")
def etf_screen_data():
    """Return OHLCV + return data for one or more tickers."""
    import yfinance as yf

    ticker = request.args.get("ticker", "").strip().upper()
    extra = request.args.get("extra", "").strip()
    period = request.args.get("period", "1y")
    mode = request.args.get("mode", "ohlcv")  # ohlcv | total | price | pricediv | both | all3 | all4
    reinvest_pct = min(100, max(0, int(request.args.get("reinvest", 100))))
    interval = request.args.get("interval", "")

    if not ticker:
        return jsonify(error="ticker is required"), 400

    # Collect all symbols
    symbols = [ticker]
    if extra:
        for s in extra.split(","):
            s = s.strip().upper()
            if s and s not in symbols:
                symbols.append(s)

    # Auto-select interval
    if not interval:
        period_intervals = {
            "1mo": "1d", "3mo": "1d", "6mo": "1d",
            "ytd": "1d", "1y": "1d", "2y": "1wk",
            "5y": "1wk", "10y": "1mo", "max": "1mo",
        }
        interval = period_intervals.get(period, "1d")

    # ---------- OHLCV mode (for candlestick / technical chart) ----------
    if mode == "ohlcv":
        try:
            tk = yf.Ticker(ticker)
            df = tk.history(period=period, interval=interval, auto_adjust=False)
            if df.empty:
                return jsonify(error=f"No data found for {ticker}"), 404

            df = df.dropna(subset=["Open", "High", "Low", "Close"])
            records = []
            for dt, row in df.iterrows():
                records.append({
                    "date": dt.strftime("%Y-%m-%d %H:%M") if interval in ("1m","2m","5m","15m","30m","60m","90m","1h") else dt.strftime("%Y-%m-%d"),
                    "open": round(float(row["Open"]), 4),
                    "high": round(float(row["High"]), 4),
                    "low": round(float(row["Low"]), 4),
                    "close": round(float(row["Close"]), 4),
                    "volume": int(row["Volume"]),
                })

            info = tk.info or {}

            # Fetch ATM implied volatility from nearest options expiration
            iv_data = None
            try:
                expirations = tk.options
                if expirations:
                    chain = tk.option_chain(expirations[0])
                    last_close = float(df["Close"].iloc[-1])
                    # Find ATM call (closest strike to last close)
                    calls = chain.calls
                    if not calls.empty and "strike" in calls.columns and "impliedVolatility" in calls.columns:
                        calls = calls.dropna(subset=["impliedVolatility"])
                        if not calls.empty:
                            atm_idx = (calls["strike"] - last_close).abs().idxmin()
                            atm_call_iv = float(calls.loc[atm_idx, "impliedVolatility"]) * 100
                            # Also get ATM put
                            puts = chain.puts
                            atm_put_iv = None
                            if not puts.empty and "impliedVolatility" in puts.columns:
                                puts = puts.dropna(subset=["impliedVolatility"])
                                if not puts.empty:
                                    atm_put_idx = (puts["strike"] - last_close).abs().idxmin()
                                    atm_put_iv = float(puts.loc[atm_put_idx, "impliedVolatility"]) * 100
                            # Average call and put IV for ATM IV
                            if atm_put_iv is not None:
                                iv_data = {"atm_iv": round((atm_call_iv + atm_put_iv) / 2, 2),
                                           "call_iv": round(atm_call_iv, 2),
                                           "put_iv": round(atm_put_iv, 2),
                                           "expiration": expirations[0]}
                            else:
                                iv_data = {"atm_iv": round(atm_call_iv, 2),
                                           "call_iv": round(atm_call_iv, 2),
                                           "put_iv": None,
                                           "expiration": expirations[0]}
            except Exception:
                pass  # Options data not available for all tickers

            return jsonify(
                ticker=ticker,
                name=info.get("longName") or info.get("shortName") or ticker,
                records=records,
                iv=iv_data,
            )
        except Exception as e:
            return jsonify(error=str(e)), 500

    # ---------- Return modes ----------
    frac = reinvest_pct / 100.0
    try:
        # Download price data (unadjusted)
        price_df = yf.download(symbols, period=period, interval=interval, auto_adjust=False, progress=False)
        if price_df.empty:
            return jsonify(error="No price data found"), 404

        # Download daily data with dividends for return calcs
        div_df = yf.download(symbols, period=period, interval="1d", auto_adjust=False, actions=True, progress=False)

        result = {"mode": mode, "reinvest_pct": reinvest_pct, "series": {}, "stats": {}, "warnings": []}

        def _extract_col(df, col, sym):
            """Safely extract a Series from yfinance download (handles multi-level columns)."""
            if col not in df.columns and hasattr(df.columns, 'get_level_values'):
                # Multi-level columns
                if col in df.columns.get_level_values(0):
                    sub = df[col]
                    if isinstance(sub, pd.DataFrame):
                        if sym in sub.columns:
                            return sub[sym].dropna()
                        elif len(sub.columns) == 1:
                            return sub.iloc[:, 0].dropna()
                    return sub.dropna() if isinstance(sub, pd.Series) else pd.Series(dtype=float)
                return pd.Series(dtype=float)
            if col in df.columns:
                s = df[col]
                if isinstance(s, pd.DataFrame):
                    if sym in s.columns:
                        return s[sym].dropna()
                    elif len(s.columns) == 1:
                        return s.iloc[:, 0].dropna()
                    return pd.Series(dtype=float)
                return s.dropna()
            return pd.Series(dtype=float)

        for sym in symbols:
            close = _extract_col(price_df, "Close", sym)
            div_close = _extract_col(div_df, "Close", sym)
            if div_close.empty:
                div_close = close
            divs_raw = _extract_col(div_df, "Dividends", sym)
            divs = divs_raw.reindex(div_close.index, fill_value=0.0) if not divs_raw.empty else pd.Series(0.0, index=div_close.index)

            if close.empty and div_close.empty:
                result["warnings"].append(f"{sym}: no data available")
                continue

            # Use div_close (daily) for dates and price normalization so all
            # traces share the same x-axis.  close (from price_df) may use a
            # coarser interval (weekly/monthly) whose length differs from the
            # daily return traces, causing misaligned chart data.
            base = div_close if not div_close.empty else close
            dates = [d.strftime("%Y-%m-%d") for d in base.index]
            norm_price = [round(float(v), 4) for v in (base / float(base.iloc[0]) * 100)]

            # Compute return series based on mode
            traces = {}
            if mode == "price":
                traces["price"] = norm_price
            elif mode == "pricediv":
                pd_series = _blend_price_drip(div_close, divs, 0.0, track_cash=True)
                traces["pricediv"] = [round(v, 4) for v in pd_series.tolist()]
            elif mode == "total":
                tot = _blend_price_drip(div_close, divs, frac, track_cash=True)
                traces["total"] = [round(v, 4) for v in tot.tolist()]
            elif mode == "both":
                tot = _blend_price_drip(div_close, divs, frac, track_cash=True)
                traces["total"] = [round(v, 4) for v in tot.tolist()]
                traces["price"] = norm_price
            elif mode == "all3":
                traces["price"] = norm_price
                blend = _blend_price_drip(div_close, divs, frac, track_cash=True)
                traces["blend"] = [round(v, 4) for v in blend.tolist()]
                drip = _blend_price_drip(div_close, divs, 1.0, track_cash=True)
                traces["drip"] = [round(v, 4) for v in drip.tolist()]
            elif mode == "all4":
                traces["price"] = norm_price
                pdiv = _blend_price_drip(div_close, divs, 0.0, track_cash=True)
                traces["pricediv"] = [round(v, 4) for v in pdiv.tolist()]
                blend = _blend_price_drip(div_close, divs, frac, track_cash=True)
                traces["blend"] = [round(v, 4) for v in blend.tolist()]
                drip = _blend_price_drip(div_close, divs, 1.0, track_cash=True)
                traces["drip"] = [round(v, 4) for v in drip.tolist()]

            # Statistics — use base (daily) consistently with traces
            price_ret = round((float(base.iloc[-1]) / float(base.iloc[0]) - 1) * 100, 2)
            # Use the best available total return figure
            if "total" in traces:
                total_ret = round(traces["total"][-1] - 100, 2)
            elif "drip" in traces:
                total_ret = round(traces["drip"][-1] - 100, 2)
            elif "pricediv" in traces:
                total_ret = round(traces["pricediv"][-1] - 100, 2)
            else:
                total_ret = price_ret
            div_contrib = round(total_ret - price_ret, 2)

            # Annualized return
            n_days = (base.index[-1] - base.index[0]).days
            if n_days > 30:
                years = n_days / 365.25
                final_norm = 100 + total_ret
                ann = round(((final_norm / 100) ** (1 / years) - 1) * 100, 2)
            else:
                ann = None

            # Max drawdown (on price)
            running_max = base.cummax()
            drawdown = ((base - running_max) / running_max * 100)
            mdd = round(float(drawdown.min()), 2)

            result["series"][sym] = {"dates": dates, "traces": traces}
            result["stats"][sym] = {
                "total_ret": total_ret,
                "price_ret": price_ret,
                "div_contrib": div_contrib,
                "annualized": ann,
                "max_drawdown": mdd,
            }

        return jsonify(result)
    except Exception as e:
        return jsonify(error=str(e)), 500


@app.route("/api/etf-screen/tickers")
def etf_screen_tickers():
    """Return list of portfolio tickers for the quick-pick dropdown."""
    profile_id = get_profile_id()
    conn = get_connection()
    conn.row_factory = __import__("sqlite3").Row
    rows = conn.execute(
        """SELECT ticker, description
           FROM all_account_info WHERE profile_id = ? ORDER BY ticker""",
        (profile_id,),
    ).fetchall()
    conn.close()
    return jsonify(tickers=[dict(r) for r in rows])


# ── Dividend Calendar ──────────────────────────────────────────────────────────

def _yf_div_pay_date(ticker):
    """Fetch next dividend pay date from yfinance. Returns date or None."""
    try:
        import yfinance as yf
        cal = yf.Ticker(ticker).calendar
        if not cal or not isinstance(cal, dict):
            return None
        d = cal.get("Dividend Date")
        if d is None:
            return None
        if hasattr(d, "date"):
            d = d.date()
        elif isinstance(d, str):
            from datetime import datetime as _dtm
            d = _dtm.strptime(d[:10], "%Y-%m-%d").date()
        import datetime as _dtmod
        if not isinstance(d, _dtmod.date):
            return None
        if d < _dtmod.date.today() - _dtmod.timedelta(days=548):
            return None
        return d
    except Exception:
        return None


def _build_cal_events():
    """Build dividend calendar events from all_account_info."""
    from datetime import datetime, timedelta
    import calendar as _cal

    FREQ_LABEL = {
        "M": "monthly", "52": "weekly", "W": "weekly",
        "Q": "quarterly", "SA": "semi-annual", "A": "annual",
    }
    FREQ_COLOR = {
        "M": "#00c9a7", "52": "#FFD700", "W": "#FFD700",
        "Q": "#7ecfff", "SA": "#f0a0ff", "A": "#f0a0ff",
    }
    FREQ_OFFSET = {"M": 10, "Q": 14, "SA": 21, "A": 21}

    profile_id = get_profile_id()
    conn = get_connection()
    try:
        df = pd.read_sql("""
            SELECT ticker, description, ex_div_date, div, div_frequency
            FROM all_account_info
            WHERE ex_div_date IS NOT NULL
              AND ex_div_date NOT IN ('', '--')
              AND current_price IS NOT NULL
              AND COALESCE(quantity, 0) > 0
              AND profile_id = ?
            ORDER BY ticker
        """, conn, params=[profile_id])
    except Exception:
        conn.close()
        return []
    conn.close()

    # Pre-fetch yfinance pay dates for non-weekly tickers
    yf_pay_dates = {}
    for tkr in df.loc[~df["div_frequency"].isin(["52", "W"]), "ticker"].tolist():
        pd_date = _yf_div_pay_date(tkr)
        if pd_date:
            yf_pay_dates[tkr] = pd_date

    events = []
    today_d = datetime.today().date()

    for _, row in df.iterrows():
        raw = str(row["ex_div_date"]).strip()
        dt = None
        for fmt in ("%m/%d/%y", "%m/%d/%Y", "%Y-%m-%d"):
            try:
                dt = datetime.strptime(raw, fmt).date()
                break
            except ValueError:
                continue
        if dt is None:
            continue

        freq = str(row["div_frequency"]).strip() if pd.notna(row["div_frequency"]) else ""
        amount = float(row["div"]) if pd.notna(row["div"]) and row["div"] else None
        ticker = row["ticker"]

        # Project ex-div date forward to next upcoming occurrence
        # Owner (pid==1) non-weekly dates come from Excel import and are kept as-is
        # Weekly payers always project forward since the next ex-div is always ~1 week away
        threshold = today_d - timedelta(days=1)
        is_weekly = freq in ("52", "W")
        if (is_weekly or profile_id != 1) and dt < threshold and freq:
            def _add_months(d, n):
                m = d.month + n
                y = d.year + (m - 1) // 12
                m = (m - 1) % 12 + 1
                day = min(d.day, _cal.monthrange(y, m)[1])
                return d.replace(year=y, month=m, day=day)

            _period = {
                "M":  lambda d: _add_months(d, 1),
                "Q":  lambda d: _add_months(d, 3),
                "SA": lambda d: _add_months(d, 6),
                "A":  lambda d: _add_months(d, 12),
                "W":  lambda d: d + timedelta(weeks=1),
                "52": lambda d: d + timedelta(weeks=1),
            }.get(freq)
            if _period:
                while dt < threshold:
                    dt = _period(dt)

        # Determine pay date
        pay_estimated = True

        def _next_biz(base, days):
            d = base + timedelta(days=days)
            if d.weekday() == 5:
                d += timedelta(days=2)
            elif d.weekday() == 6:
                d += timedelta(days=1)
            return d

        if freq in ("52", "W"):
            # Weekly: pay 1 business day after ex-div
            pay_dt = _next_biz(dt, 1)
        elif ticker in yf_pay_dates:
            # Confirmed pay date from Yahoo Finance
            pay_dt = yf_pay_dates[ticker]
            pay_estimated = False
        elif freq == "M":
            # Monthly ETF distributions: pay within 2 days of ex-div
            pay_dt = _next_biz(dt, 2)
        else:
            offset = FREQ_OFFSET.get(freq, 10)
            pay_dt = dt + timedelta(days=offset)

        events.append({
            "ticker":        ticker,
            "description":   str(row["description"]) if pd.notna(row["description"]) else "",
            "date":          dt.isoformat(),
            "day":           str(dt.day),
            "month":         dt.strftime("%b"),
            "weekday":       dt.strftime("%a"),
            "amount":        round(amount, 4) if amount is not None else None,
            "freq":          freq,
            "freq_label":    FREQ_LABEL.get(freq, freq.lower() if freq else ""),
            "color":         FREQ_COLOR.get(freq, "#8899aa"),
            "pay_date":      pay_dt.isoformat(),
            "pay_month":     pay_dt.strftime("%b"),
            "pay_day":       str(pay_dt.day),
            "pay_estimated": pay_estimated,
        })

    events.sort(key=lambda e: e["pay_date"])
    return events


@app.route("/api/div-calendar")
def div_calendar():
    """Return dividend calendar events as JSON."""
    from datetime import date
    try:
        events = _build_cal_events()
        return jsonify(events=events, today=date.today().isoformat())
    except Exception as e:
        return jsonify(error=str(e)), 500


# ── Buy / Sell Signals ─────────────────────────────────────────────────────────

SIGNAL_COLOR = {"BUY": "#00c853", "SELL": "#d50000", "NEUTRAL": "#f9a825"}
SIGNAL_ORDER = {"BUY": 0, "NEUTRAL": 1, "SELL": 2}

SECTOR_WATCHLIST = [
    "XLB", "XLC", "XLE", "XLF", "XLI", "XLK", "XLP", "XLRE", "XLU", "XLV", "XLY",
]
ALT_WATCHLIST = [
    "HDG", "HFGM", "WTMF", "DBMF", "HFMF", "GMOM", "MFUT", "CFIT", "GAA",
    "RPAR", "IRVH", "WTIP", "RLY", "FTLS", "BTAL", "CSM", "WTLS", "ORR",
    "NLSI", "TAIL", "KMLM", "CTA", "MNA",
]
BSS_WATCHLIST_NAMES = {
    "XLB": "Materials Select Sector SPDR Fund",
    "XLC": "Communication Services Select Sector SPDR Fund",
    "XLE": "Energy Select Sector SPDR Fund",
    "XLF": "Financial Select Sector SPDR Fund",
    "XLI": "Industrial Select Sector SPDR Fund",
    "XLK": "Technology Select Sector SPDR Fund",
    "XLP": "Consumer Staples Select Sector SPDR Fund",
    "XLRE": "Real Estate Select Sector SPDR Fund",
    "XLU": "Utilities Select Sector SPDR Fund",
    "XLV": "Health Care Select Sector SPDR Fund",
    "XLY": "Consumer Discretionary Select Sector SPDR Fund",
    "HDG": "ProShares Hedge Replication ETF",
    "HFGM": "Unlimited HFGM Global Macro ETF",
    "WTMF": "WisdomTree Managed Futures Strategy Fund",
    "DBMF": "iMGP DBi Managed Futures Strategy ETF",
    "HFMF": "Unlimited HFMF Managed Futures ETF",
    "GMOM": "Cambria Global Momentum ETF",
    "MFUT": "Cambria Managed Futures Strategy ETF",
    "CFIT": "Cambria Fixed Income Trend ETF",
    "GAA": "Cambria Global Asset Allocation ETF",
    "RPAR": "RPAR Risk Parity ETF",
    "IRVH": "Global X Interest Rate Volatility & Inflation Hedge ETF",
    "WTIP": "WisdomTree Inflation Plus Fund",
    "RLY": "SPDR SSgA Multi-Asset Real Return ETF",
    "FTLS": "First Trust Long/Short Equity ETF",
    "BTAL": "AGFiQ US Market Neutral Anti-Beta Fund",
    "CSM": "ProShares Large Cap Core Plus",
    "WTLS": "WisdomTree Efficient Long/Short U.S. Equity Fund",
    "ORR": "Militia Long/Short Equity ETF",
    "NLSI": "NEOS Long/Short Equity Income ETF",
    "TAIL": "Cambria Tail Risk ETF",
    "KMLM": "KFA Mount Lucas Index Strategy ETF",
    "CTA": "Simplify Managed Futures Strategy ETF",
    "MNA": "IQ Merger Arbitrage ETF",
}


def _bss_ao(high, low):
    if len(high) < 34:
        return "NEUTRAL", None, ""
    mid = (high + low) / 2
    ao_v = (mid.rolling(5).mean() - mid.rolling(34).mean()).dropna()
    if len(ao_v) < 2:
        return "NEUTRAL", None, ""
    cur, prv = float(ao_v.iloc[-1]), float(ao_v.iloc[-2])
    direction = "Rising" if cur > prv else ("Falling" if cur < prv else "Flat")
    if cur > 0 and cur > prv:
        return "BUY", cur, direction
    if cur < 0 and cur < prv:
        return "SELL", cur, direction
    return "NEUTRAL", cur, direction


def _bss_rsi(close, period=14):
    if len(close) < period + 1:
        return "NEUTRAL", None
    delta = close.diff()
    avg_gain = delta.clip(lower=0).ewm(com=period - 1, adjust=False, min_periods=period).mean()
    avg_loss = (-delta).clip(lower=0).ewm(com=period - 1, adjust=False, min_periods=period).mean()
    rs = avg_gain / avg_loss
    val = float((100 - 100 / (1 + rs)).iloc[-1])
    if pd.isna(val):
        return "NEUTRAL", None
    if val < 30:
        return "BUY", val
    if val > 70:
        return "SELL", val
    return "NEUTRAL", val


def _bss_macd(close):
    if len(close) < 26:
        return "NEUTRAL"
    macd_line = close.ewm(span=12, adjust=False).mean() - close.ewm(span=26, adjust=False).mean()
    sig_line = macd_line.ewm(span=9, adjust=False).mean()
    m, s = float(macd_line.iloc[-1]), float(sig_line.iloc[-1])
    if pd.isna(m) or pd.isna(s):
        return "NEUTRAL"
    return "BUY" if m > s else "SELL"


def _bss_sma(close, period):
    if len(close) < period:
        return "NEUTRAL", None, None
    sma_val = close.rolling(period).mean().iloc[-1]
    price = close.iloc[-1]
    if pd.isna(sma_val) or pd.isna(price) or float(sma_val) == 0:
        return "NEUTRAL", None, None
    sma_f, price_f = float(sma_val), float(price)
    pct = (price_f - sma_f) / sma_f * 100
    if price_f > sma_f * 1.01:
        return "BUY", sma_f, pct
    if price_f < sma_f * 0.99:
        return "SELL", sma_f, pct
    return "NEUTRAL", sma_f, pct


def _slow_stochastic(high, low, close, k_period=14, k_smooth=3, d_period=3):
    """Slow Stochastic. Returns (slow_k, slow_d) latest values."""
    needed = k_period + k_smooth + d_period
    if len(close) < needed:
        return None, None
    lowest_low = low.rolling(k_period).min()
    highest_high = high.rolling(k_period).max()
    denom = highest_high - lowest_low
    denom = denom.replace(0, float("nan"))
    raw_k = (close - lowest_low) / denom * 100
    slow_k = raw_k.rolling(k_smooth).mean()
    slow_d = slow_k.rolling(d_period).mean()
    k_val = float(slow_k.iloc[-1]) if not pd.isna(slow_k.iloc[-1]) else None
    d_val = float(slow_d.iloc[-1]) if not pd.isna(slow_d.iloc[-1]) else None
    return k_val, d_val


def _bss_vote(signals):
    """Majority-vote: need >50% of valid (non-None) signals to agree."""
    valid = [s for s in signals if s is not None]
    if not valid:
        return "NEUTRAL"
    threshold = len(valid) / 2
    if valid.count("BUY") > threshold:
        return "BUY"
    if valid.count("SELL") > threshold:
        return "SELL"
    return "NEUTRAL"


def _bss_coverage(close, divs_series):
    """Compute TTM yield-based coverage ratio — same formula as /api/portfolio-coverage.

    coverage = (TTM_price_return_pct + TTM_dist_yield) / TTM_dist_yield
    Returns (coverage_ratio, signal, nav_erosion_label).
    """
    try:
        if close is None or len(close) < 2:
            return None, None, None
        cur_price = float(close.iloc[-1])
        price_1yr_ago = float(close.iloc[0])
        if cur_price <= 0 or price_1yr_ago <= 0:
            return None, None, None

        ttm_price_return_pct = (cur_price - price_1yr_ago) / price_1yr_ago

        ttm_dist_per_share = float(divs_series.sum()) if divs_series is not None and not divs_series.empty else 0.0
        ttm_dist_yield = ttm_dist_per_share / cur_price if cur_price > 0 else 0.0

        if ttm_dist_yield <= 0:
            return None, None, None

        coverage = round((ttm_price_return_pct + ttm_dist_yield) / ttm_dist_yield, 4)

        if coverage > 1:
            sig = "BUY"
            erosion = "Low"
        elif coverage < 1:
            sig = "SELL"
            erosion = "High"
        else:
            sig = "NEUTRAL"
            erosion = "Medium"

        return coverage, sig, erosion
    except Exception:
        return None, None, None


def _bss_sharpe(close, risk_free_annual=0.05):
    import numpy as np
    try:
        if len(close) < 30:
            return None
        daily_ret = close.pct_change().dropna()
        if len(daily_ret) < 30:
            return None
        std = float(daily_ret.std())
        if std == 0 or np.isnan(std):
            return None
        return round((float(daily_ret.mean()) - risk_free_annual / 252) / std * np.sqrt(252), 2)
    except Exception:
        return None


def _bss_sortino(close, risk_free_annual=0.05):
    import numpy as np
    try:
        if len(close) < 30:
            return None
        daily_ret = close.pct_change().dropna()
        if len(daily_ret) < 30:
            return None
        neg_ret = daily_ret[daily_ret < 0]
        if len(neg_ret) == 0:
            return None
        down_std = float(neg_ret.std())
        if down_std == 0 or np.isnan(down_std):
            return None
        return round((float(daily_ret.mean()) - risk_free_annual / 252) / down_std * np.sqrt(252), 2)
    except Exception:
        return None


@app.route("/api/buy-sell-signals")
def buy_sell_signals_data():
    """Compute 5-indicator majority-vote signals for portfolio + sector/watchlist tickers."""
    import yfinance as yf
    import json
    import math
    import warnings
    warnings.filterwarnings("ignore")

    try:
        import plotly.graph_objects as go
        import plotly.utils
        has_plotly = True
    except ImportError:
        has_plotly = False

    WATCHLIST_SIZE = 1000

    def _fmt_pct(v):
        if v is None:
            return "\u2014"
        return f"+{v:.1f}%" if v >= 0 else f"{v:.1f}%"

    profile_id = get_profile_id()
    conn = get_connection()
    try:
        port = pd.read_sql("""
            SELECT ticker, description, classification_type, purchase_value
            FROM all_account_info
            WHERE purchase_value IS NOT NULL AND purchase_value > 0
              AND profile_id = ?
            ORDER BY ticker
        """, conn, params=[profile_id])
    except Exception:
        port = pd.DataFrame(columns=["ticker", "description", "classification_type", "purchase_value"])
    conn.close()

    port["description"] = port["description"].fillna("")
    port["classification_type"] = port["classification_type"].fillna("")
    port_sizes = dict(zip(port["ticker"].tolist(), port["purchase_value"].tolist()))
    port_desc = dict(zip(port["ticker"].tolist(), port["description"].tolist()))
    port_type = dict(zip(port["ticker"].tolist(), port["classification_type"].tolist()))

    all_wl = SECTOR_WATCHLIST + ALT_WATCHLIST
    all_tickers = sorted(set(port["ticker"].tolist() + all_wl))

    error = None
    fig_json = None
    table_rows = []

    try:
        raw = yf.download(
            " ".join(all_tickers),
            period="1y",
            interval="1d",
            auto_adjust=True,
            actions=True,
            progress=False,
        )

        if raw.empty:
            error = "No price data returned from Yahoo Finance."
        else:
            try:
                high_df = raw["High"]
                low_df = raw["Low"]
                close_df = raw["Close"]
            except KeyError:
                high_df = low_df = close_df = None

            # Extract dividends for coverage ratio
            try:
                divs_df = raw["Dividends"]
            except KeyError:
                divs_df = None

            if any(df is None for df in [high_df, low_df, close_df]):
                error = "Missing OHLC price data from Yahoo Finance."
            else:
                SECTOR_SET = set(SECTOR_WATCHLIST)
                labels = ["State Street Sectors"]
                parents_list = [""]
                values_list = [0]
                colors_list = ["#1a1a3a"]
                hover_texts = [""]
                empty = pd.Series([], dtype=float)

                for ticker in all_tickers:
                    has_data = (ticker in close_df.columns and
                                ticker in high_df.columns and
                                ticker in low_df.columns)
                    if has_data:
                        close = close_df[ticker].dropna()
                        high = high_df[ticker].dropna()
                        low = low_df[ticker].dropna()
                    else:
                        close = high = low = empty

                    ao_sig, ao_val, ao_dir = _bss_ao(high, low)
                    rsi_sig, rsi_val = _bss_rsi(close)
                    macd_sig = _bss_macd(close)
                    sma50_sig, sma50_v, sma50_pct = _bss_sma(close, 50)
                    sma200_sig, sma200_v, sma200_pct = _bss_sma(close, 200)
                    sharpe_val = _bss_sharpe(close)
                    sortino_val = _bss_sortino(close)

                    # Coverage ratio (same formula as /api/portfolio-coverage)
                    tk_divs = pd.Series(dtype=float)
                    if divs_df is not None and has_data:
                        if isinstance(divs_df, pd.DataFrame) and ticker in divs_df.columns:
                            tk_divs = divs_df[ticker].dropna()
                        elif isinstance(divs_df, pd.Series):
                            tk_divs = divs_df.dropna()
                    cov_ratio, cov_sig, nav_erosion = _bss_coverage(close, tk_divs)

                    signal = _bss_vote([ao_sig, rsi_sig, macd_sig, sma50_sig, sma200_sig, cov_sig])

                    is_portfolio = ticker in port_sizes
                    is_sector = ticker in SECTOR_SET and not is_portfolio
                    size = port_sizes.get(ticker, WATCHLIST_SIZE)

                    # Treemap nodes
                    if has_plotly:
                        ao_val_str = f"{ao_val:.4f}" if ao_val is not None else "\u2014"
                        rsi_val_str = f"{rsi_val:.1f}" if rsi_val is not None else "\u2014"
                        cov_str = f"{cov_ratio:.2f}" if cov_ratio is not None else "\u2014"
                        hover_text = (
                            f"<b>{signal}</b><br>"
                            f"AO: {ao_sig} ({ao_val_str}, {ao_dir or chr(8212)})<br>"
                            f"RSI: {rsi_sig} ({rsi_val_str})<br>"
                            f"MACD: {macd_sig}<br>"
                            f"SMA 50: {sma50_sig} ({_fmt_pct(sma50_pct)})<br>"
                            f"SMA 200: {sma200_sig} ({_fmt_pct(sma200_pct)})<br>"
                            f"Coverage: {cov_str} ({nav_erosion or chr(8212)} erosion risk)"
                        )
                        if is_portfolio:
                            labels.append(ticker)
                            parents_list.append("")
                            values_list.append(float(size))
                            colors_list.append(SIGNAL_COLOR[signal])
                            hover_texts.append(hover_text)
                        elif is_sector:
                            labels.append(ticker)
                            parents_list.append("State Street Sectors")
                            values_list.append(WATCHLIST_SIZE)
                            colors_list.append(SIGNAL_COLOR[signal])
                            hover_texts.append(hover_text)

                    table_rows.append({
                        "ticker": ticker,
                        "desc": port_desc.get(ticker) or BSS_WATCHLIST_NAMES.get(ticker, ""),
                        "ctype": port_type.get(ticker, ""),
                        "source": "Portfolio" if is_portfolio else ("Sectors" if is_sector else "Watchlist"),
                        "signal": signal,
                        "sig_order": SIGNAL_ORDER[signal],
                        "ao_sig": ao_sig,
                        "ao_sig_ord": SIGNAL_ORDER[ao_sig],
                        "ao_value": f"{ao_val:.4f}" if ao_val is not None else "\u2014",
                        "ao_val_num": ao_val if ao_val is not None else "",
                        "ao_dir": ao_dir,
                        "rsi_sig": rsi_sig,
                        "rsi_sig_ord": SIGNAL_ORDER[rsi_sig],
                        "rsi_value": f"{rsi_val:.1f}" if rsi_val is not None else "\u2014",
                        "rsi_val_num": rsi_val if rsi_val is not None else "",
                        "macd_sig": macd_sig,
                        "macd_sig_ord": SIGNAL_ORDER[macd_sig],
                        "sma50_sig": sma50_sig,
                        "sma50_sig_ord": SIGNAL_ORDER[sma50_sig],
                        "sma50_pct": _fmt_pct(sma50_pct),
                        "sma200_sig": sma200_sig,
                        "sma200_sig_ord": SIGNAL_ORDER[sma200_sig],
                        "sma200_pct": _fmt_pct(sma200_pct),
                        "sharpe_val": f"{sharpe_val:.2f}" if sharpe_val is not None else "\u2014",
                        "sharpe_val_num": sharpe_val if sharpe_val is not None else "",
                        "sortino_val": f"{sortino_val:.2f}" if sortino_val is not None else "\u2014",
                        "sortino_val_num": sortino_val if sortino_val is not None else "",
                        "cov_ratio": f"{cov_ratio:.2f}" if cov_ratio is not None else "\u2014",
                        "cov_ratio_num": cov_ratio if cov_ratio is not None else "",
                        "cov_sig": cov_sig or "NEUTRAL",
                        "cov_sig_ord": SIGNAL_ORDER.get(cov_sig or "NEUTRAL", 1),
                        "nav_erosion": nav_erosion or "\u2014",
                        "pv_fmt": f"${size:,.2f}" if is_portfolio else "\u2014",
                        "pv_num": float(size) if is_portfolio else 0,
                        "src_order": 0 if is_portfolio else (1 if is_sector else 2),
                    })

                table_rows.sort(key=lambda r: (r["src_order"], r["sig_order"], -r["pv_num"]))

                if has_plotly:
                    fig = go.Figure(go.Treemap(
                        labels=labels,
                        parents=parents_list,
                        values=values_list,
                        text=hover_texts,
                        marker=dict(colors=colors_list),
                        textinfo="label",
                        hovertemplate="<b>%{label}</b><br>%{text}<extra></extra>",
                    ))
                    fig.update_layout(
                        title="Buy / Sell Signal Dashboard",
                        template="plotly_dark",
                        margin=dict(t=50, l=5, r=5, b=5),
                        height=720,
                        hoverlabel=dict(bgcolor="#111124", bordercolor="#3a3a5c",
                                        font=dict(color="#e0e0e0", size=13)),
                    )
                    fig_json = json.dumps(fig, cls=plotly.utils.PlotlyJSONEncoder)

    except Exception:
        import traceback
        error = traceback.format_exc(limit=5)

    def _scrub(obj):
        if isinstance(obj, list):
            return [_scrub(v) for v in obj]
        if isinstance(obj, dict):
            return {k: _scrub(v) for k, v in obj.items()}
        if isinstance(obj, float) and (math.isnan(obj) or math.isinf(obj)):
            return None
        if obj is pd.NaT:
            return None
        if hasattr(obj, 'isoformat'):
            return obj.isoformat()
        return obj

    table_rows = _scrub(table_rows)
    return jsonify(fig_json=fig_json, error=error, table_rows=table_rows)


# ── Watchlist ──────────────────────────────────────────────────────────────────

@app.route("/api/watchlist/watching", methods=["GET", "POST"])
def watchlist_watching_list():
    """GET: return watching rows.  POST: bulk-replace watching list."""
    conn = get_connection()

    if request.method == "POST":
        data = request.get_json(force=True)
        rows = data.get("rows", [])
        conn.execute("DELETE FROM watchlist_watching")
        for i, r in enumerate(rows):
            ticker = str(r.get("ticker", "")).strip().upper()
            if not ticker:
                continue
            notes = str(r.get("notes", ""))[:500]
            conn.execute(
                "INSERT INTO watchlist_watching (ticker, notes, sort_order) VALUES (?, ?, ?)",
                (ticker, notes, i),
            )
        conn.commit()
        conn.close()
        return jsonify(ok=True)

    # GET
    rows = conn.execute(
        "SELECT ticker, notes, added_date FROM watchlist_watching ORDER BY sort_order, id"
    ).fetchall()
    conn.close()
    result = []
    for r in rows:
        result.append({
            "ticker": r["ticker"],
            "notes": r["notes"] or "",
            "added_date": r["added_date"] or "",
        })
    return jsonify(rows=result)


@app.route("/api/watchlist/data")
def watchlist_data():
    """Analysis data for all watchlist tickers."""
    import yfinance as yf
    import numpy as np
    import warnings
    import math
    warnings.filterwarnings("ignore")

    conn = get_connection()
    try:
        watching_rows = conn.execute(
            "SELECT ticker, notes FROM watchlist_watching ORDER BY sort_order, id"
        ).fetchall()
    except Exception:
        watching_rows = []
    conn.close()

    watching_tickers = [r["ticker"] for r in watching_rows]
    if not watching_tickers:
        return jsonify(watching=[], counts={"BUY": 0, "SELL": 0, "NEUTRAL": 0})

    error = None
    result_rows = []
    counts = {"BUY": 0, "SELL": 0, "NEUTRAL": 0}

    # ── Indicator helpers ──

    def _ao(high, low):
        if len(high) < 34:
            return "NEUTRAL", None, ""
        mid = (high + low) / 2
        ao_v = (mid.rolling(5).mean() - mid.rolling(34).mean()).dropna()
        if len(ao_v) < 2:
            return "NEUTRAL", None, ""
        cur, prv = float(ao_v.iloc[-1]), float(ao_v.iloc[-2])
        direction = "Rising" if cur > prv else ("Falling" if cur < prv else "Flat")
        if cur > 0 and cur > prv:
            return "BUY", cur, direction
        if cur < 0 and cur < prv:
            return "SELL", cur, direction
        return "NEUTRAL", cur, direction

    def _rsi(close, period=14):
        if len(close) < period + 1:
            return "NEUTRAL", None
        delta = close.diff()
        avg_gain = delta.clip(lower=0).ewm(com=period - 1, adjust=False, min_periods=period).mean()
        avg_loss = (-delta).clip(lower=0).ewm(com=period - 1, adjust=False, min_periods=period).mean()
        rs = avg_gain / avg_loss
        val = float((100 - 100 / (1 + rs)).iloc[-1])
        if pd.isna(val):
            return "NEUTRAL", None
        if val < 30:
            return "BUY", val
        if val > 70:
            return "SELL", val
        return "NEUTRAL", val

    def _macd(close):
        if len(close) < 26:
            return "NEUTRAL"
        macd_line = close.ewm(span=12, adjust=False).mean() - close.ewm(span=26, adjust=False).mean()
        sig_line = macd_line.ewm(span=9, adjust=False).mean()
        m, s = float(macd_line.iloc[-1]), float(sig_line.iloc[-1])
        if pd.isna(m) or pd.isna(s):
            return "NEUTRAL"
        return "BUY" if m > s else "SELL"

    def _sma(close, period):
        if len(close) < period:
            return "NEUTRAL", None, None
        sma_val = close.rolling(period).mean().iloc[-1]
        price = close.iloc[-1]
        if pd.isna(sma_val) or pd.isna(price) or float(sma_val) == 0:
            return "NEUTRAL", None, None
        sma_f, price_f = float(sma_val), float(price)
        pct = (price_f - sma_f) / sma_f * 100
        if price_f > sma_f * 1.01:
            return "BUY", sma_f, pct
        if price_f < sma_f * 0.99:
            return "SELL", sma_f, pct
        return "NEUTRAL", sma_f, pct

    def _vote(signals):
        valid = [s for s in signals if s is not None]
        if not valid:
            return "NEUTRAL"
        threshold = len(valid) / 2
        if valid.count("BUY") > threshold:
            return "BUY"
        if valid.count("SELL") > threshold:
            return "SELL"
        return "NEUTRAL"

    def _sharpe(close, risk_free_annual=0.05):
        try:
            if len(close) < 30:
                return None
            daily_ret = close.pct_change().dropna()
            if len(daily_ret) < 30:
                return None
            std = float(daily_ret.std())
            if std == 0 or np.isnan(std):
                return None
            daily_rf = risk_free_annual / 252
            excess = float(daily_ret.mean()) - daily_rf
            return round(excess / std * np.sqrt(252), 2)
        except Exception:
            return None

    def _sortino(close, risk_free_annual=0.05):
        try:
            if len(close) < 30:
                return None
            daily_ret = close.pct_change().dropna()
            if len(daily_ret) < 30:
                return None
            daily_rf = risk_free_annual / 252
            neg_ret = daily_ret[daily_ret < 0]
            if len(neg_ret) == 0:
                return None
            down_std = float(neg_ret.std())
            if down_std == 0 or np.isnan(down_std):
                return None
            excess = float(daily_ret.mean()) - daily_rf
            return round(excess / down_std * np.sqrt(252), 2)
        except Exception:
            return None

    try:
        raw = yf.download(
            " ".join(watching_tickers),
            period="1y",
            interval="1d",
            auto_adjust=False,
            actions=True,
            progress=False,
        )

        if raw.empty:
            return jsonify(watching=[], counts=counts, error="No price data returned.")

        try:
            _top = raw.columns.get_level_values(0)
            has_multi = True
        except AttributeError:
            has_multi = False

        if has_multi:
            close_df = raw["Adj Close"] if "Adj Close" in _top else raw["Close"]
            unadj_close_df = raw["Close"]
            high_df = raw["High"]
            low_df = raw["Low"]
            divs_df = raw["Dividends"] if "Dividends" in _top else None
        else:
            close_df = raw[["Adj Close"]] if "Adj Close" in raw.columns else raw[["Close"]]
            unadj_close_df = raw[["Close"]]
            high_df = raw[["High"]]
            low_df = raw[["Low"]]
            divs_df = raw[["Dividends"]] if "Dividends" in raw.columns else None

        empty = pd.Series([], dtype=float)
        ticker_info = {}

        for ticker in watching_tickers:
            has_data = (ticker in close_df.columns and
                        ticker in high_df.columns and
                        ticker in low_df.columns)
            if has_data:
                close = close_df[ticker].dropna()
                high = high_df[ticker].dropna()
                low = low_df[ticker].dropna()
            else:
                close = high = low = empty

            ao_sig, ao_val, ao_dir = _ao(high, low)
            rsi_sig, rsi_val = _rsi(close)
            macd_sig = _macd(close)
            sma50_sig, sma50_v, sma50_pct = _sma(close, 50)
            sma200_sig, sma200_v, sma200_pct = _sma(close, 200)
            # Coverage ratio
            cov_ratio, cov_sig, cov_erosion = None, None, None
            if divs_df is not None and has_data:
                try:
                    t_divs_cov = divs_df[ticker].dropna() if ticker in divs_df.columns else pd.Series([], dtype=float)
                    t_divs_cov = t_divs_cov[t_divs_cov > 0]
                    cov_ratio, cov_sig, cov_erosion = _bss_coverage(close, t_divs_cov)
                except Exception:
                    pass

            signal = _vote([ao_sig, rsi_sig, macd_sig, sma50_sig, sma200_sig, cov_sig])

            sharpe_val = _sharpe(close)
            sortino_val = _sortino(close)

            price = float(close.iloc[-1]) if len(close) >= 1 else None
            prev_price = float(close.iloc[-2]) if len(close) >= 2 else None
            change_1d = round((price - prev_price) / prev_price * 100, 2) \
                if price is not None and prev_price else None

            div_yield = None
            if divs_df is not None and price is not None and price > 0:
                try:
                    t_divs = divs_df[ticker].dropna() if ticker in divs_df.columns else pd.Series([], dtype=float)
                    ttm_divs = t_divs[t_divs > 0].sum()
                    if ttm_divs > 0:
                        div_yield = round(ttm_divs / price * 100, 2)
                except Exception:
                    pass

            one_yr_ret = None
            if len(close) >= 2:
                first_price = float(close.iloc[0])
                if first_price > 0:
                    one_yr_ret = round((price - first_price) / first_price * 100, 2)

            nav_erosion = False
            unadj_close = unadj_close_df[ticker].dropna() if ticker in unadj_close_df.columns else pd.Series([], dtype=float)
            if len(unadj_close) >= 2:
                unadj_first = float(unadj_close.iloc[0])
                unadj_last = float(unadj_close.iloc[-1])
                if unadj_first > 0:
                    unadj_ret = (unadj_last - unadj_first) / unadj_first * 100
                    if unadj_ret < -5:
                        nav_erosion = True

            ticker_info[ticker] = {
                "price": round(price, 2) if price is not None else None,
                "change_1d": change_1d,
                "div_yield": div_yield,
                "signal": signal,
                "ao_sig": ao_sig,
                "ao_val": round(ao_val, 4) if ao_val is not None else None,
                "ao_dir": ao_dir,
                "rsi_sig": rsi_sig,
                "rsi_val": round(rsi_val, 1) if rsi_val is not None else None,
                "macd_sig": macd_sig,
                "sma50_sig": sma50_sig,
                "sma50_pct": round(sma50_pct, 1) if sma50_pct is not None else None,
                "sma200_sig": sma200_sig,
                "sma200_pct": round(sma200_pct, 1) if sma200_pct is not None else None,
                "one_yr_ret": one_yr_ret,
                "nav_erosion": nav_erosion,
                "sharpe": round(sharpe_val, 2) if sharpe_val is not None else None,
                "sortino": round(sortino_val, 2) if sortino_val is not None else None,
                "cov_ratio": round(cov_ratio, 4) if cov_ratio is not None else None,
                "cov_sig": cov_sig,
                "nav_erosion_prob": cov_erosion,
            }

        # Build watching result rows
        for wr in watching_rows:
            t = wr["ticker"]
            info = ticker_info.get(t, {})
            result_rows.append({
                "ticker": t,
                "notes": wr["notes"] or "",
                **info,
            })

    except Exception:
        import traceback
        error = traceback.format_exc(limit=5)

    # Scrub NaN/NaT
    def _scrub(obj):
        if isinstance(obj, list):
            return [_scrub(v) for v in obj]
        if isinstance(obj, dict):
            return {k: _scrub(v) for k, v in obj.items()}
        if isinstance(obj, float) and (math.isnan(obj) or math.isinf(obj)):
            return None
        if obj is pd.NaT:
            return None
        if hasattr(obj, 'isoformat'):
            return obj.isoformat()
        return obj

    result_rows = _scrub(result_rows)

    counts = {"BUY": 0, "SELL": 0, "NEUTRAL": 0}
    for row in result_rows:
        sig = row.get("signal", "NEUTRAL")
        counts[sig] = counts.get(sig, 0) + 1

    return jsonify(watching=result_rows, counts=counts, error=error)


# ── NAV Erosion Back-Tester ────────────────────────────────────────────────────

@app.route("/api/nav-erosion/data")
def nav_erosion_data():
    """Back-test a ticker for NAV erosion over a custom date range."""
    import warnings
    warnings.filterwarnings("ignore")

    sym = request.args.get("ticker", "").strip().upper()
    try:
        initial_investment = float(request.args.get("amount", 0))
    except (TypeError, ValueError):
        initial_investment = 0.0
    start_date = request.args.get("start", "")
    end_date = request.args.get("end", "")
    try:
        reinvest_pct = float(request.args.get("reinvest", 0))
    except (TypeError, ValueError):
        reinvest_pct = 0.0

    if not sym:
        return jsonify(error="Please enter a ticker symbol.")
    if initial_investment <= 0:
        return jsonify(error="Initial investment must be greater than zero.")
    if not start_date or not end_date:
        return jsonify(error="Please select both start and end dates.")

    error = None
    warning = None
    fig_json = None
    rows = []
    summary = {}

    try:
        from datetime import datetime as _dt
        import json
        import yfinance as yf
        import plotly.graph_objects as go
        import plotly.utils

        hist = yf.Ticker(sym).history(
            start=start_date, end=end_date,
            interval="1d", auto_adjust=False, actions=True,
        )

        if hist.empty:
            return jsonify(error=f"No data found for ticker {sym}. Check the symbol and date range.")

        # Detect if data starts later than requested
        requested_start = _dt.strptime(start_date, "%Y-%m-%d").date()
        actual_start = hist.index[0].date()
        if (actual_start - requested_start).days > 30:
            warning = (
                f"{sym} only has data going back to {actual_start.strftime('%B %d, %Y')}. "
                f"Results are shown from that date — your requested start "
                f"({requested_start.strftime('%B %d, %Y')}) predates when this ETF existed."
            )

        monthly_close = hist["Close"].resample("ME").last()
        monthly_divs = hist["Dividends"].resample("ME").sum()

        df = pd.DataFrame({"price": monthly_close, "div": monthly_divs}).dropna(subset=["price"])
        df["div"] = df["div"].fillna(0.0)

        if df.empty:
            return jsonify(error=f"No usable monthly data for {sym}.")

        initial_price = float(df["price"].iloc[0])
        if initial_price == 0:
            return jsonify(error=f"Initial price for {sym} is zero — cannot calculate.")

        current_shares = initial_investment / initial_price
        cumulative_dist = 0.0
        cumulative_reinvested = 0.0
        cumulative_shares_bought = 0.0
        prev_price = initial_price
        cumulative_divs_per_share = 0.0

        for dt, row in df.iterrows():
            price = float(row["price"])
            div_per_share = float(row["div"])
            total_dist = div_per_share * current_shares
            reinvest_amt = total_dist * reinvest_pct / 100.0
            shares_bought = (reinvest_amt / price) if price > 0 else 0.0
            current_shares += shares_bought
            portfolio_val = current_shares * price
            breakeven_sh = (initial_investment / price) if price > 0 else 0.0
            shares_deficit = breakeven_sh - current_shares
            price_delta_pct = (price - initial_price) / initial_price * 100 if initial_price else 0.0
            cumulative_dist += total_dist
            cumulative_reinvested += reinvest_amt
            cumulative_shares_bought += shares_bought
            cumulative_divs_per_share += div_per_share

            # Coverage ratio: (month-over-month price change + distribution) / distribution
            price_change = price - prev_price
            if div_per_share > 0:
                coverage_ratio = round((price_change + div_per_share) / div_per_share, 4)
            else:
                coverage_ratio = None
            prev_price = price

            rows.append({
                "date": dt.strftime("%b %Y"),
                "price": round(price, 4),
                "price_delta_pct": round(price_delta_pct, 2),
                "div_per_share": round(div_per_share, 4),
                "total_dist": round(total_dist, 2),
                "reinvested": round(reinvest_amt, 2),
                "shares_bought": round(shares_bought, 4),
                "total_shares": round(current_shares, 4),
                "portfolio_val": round(portfolio_val, 2),
                "breakeven_sh": round(breakeven_sh, 4),
                "shares_deficit": round(shares_deficit, 4),
                "coverage_ratio": coverage_ratio,
            })

        final_row = rows[-1]
        # Aggregate coverage: (total price change + total divs per share) / total divs per share
        total_price_change = final_row["price"] - initial_price
        if cumulative_divs_per_share > 0:
            total_coverage = round((total_price_change + cumulative_divs_per_share) / cumulative_divs_per_share, 4)
        else:
            total_coverage = None
        summary = {
            "total_dist": round(cumulative_dist, 2),
            "total_shares_bought": round(cumulative_shares_bought, 4),
            "total_reinvested": round(cumulative_reinvested, 2),
            "final_value": final_row["portfolio_val"],
            "price_chg_pct": final_row["price_delta_pct"],
            "has_erosion": final_row["shares_deficit"] > 0,
            "final_deficit": final_row["shares_deficit"],
            "total_coverage": total_coverage,
        }

        dates_list = [r["date"] for r in rows]
        prices_list = [r["price"] for r in rows]
        vals_list = [r["portfolio_val"] for r in rows]
        breakeven_list = [initial_investment] * len(rows)

        fig = go.Figure()
        fig.add_trace(go.Scatter(
            x=dates_list, y=prices_list,
            name="Share Price",
            line=dict(color="#7ecfff", width=2),
            yaxis="y1",
            hovertemplate="<b>%{x}</b><br>Price: $%{y:.2f}<extra></extra>",
        ))
        fig.add_trace(go.Scatter(
            x=dates_list, y=vals_list,
            name="Portfolio Value",
            line=dict(color="#00e89a", width=2),
            yaxis="y2",
            hovertemplate="<b>%{x}</b><br>Portfolio Value: $%{y:,.2f}<extra></extra>",
        ))
        fig.add_trace(go.Scatter(
            x=dates_list, y=breakeven_list,
            name="Initial Investment",
            line=dict(color="#888", width=1.5, dash="dash"),
            yaxis="y2",
            hovertemplate="<b>%{x}</b><br>Break-Even: $%{y:,.2f}<extra></extra>",
        ))
        fig.update_layout(
            title=f"{sym} — NAV Erosion Back-Test ({rows[0]['date']} → {rows[-1]['date']})",
            template="plotly_dark",
            margin=dict(t=50, l=60, r=60, b=50),
            height=420,
            legend=dict(orientation="h", y=1.08, x=0),
            hoverlabel=dict(
                bgcolor="#111124",
                bordercolor="#3a3a5c",
                font=dict(color="#e0e0e0", size=13),
            ),
            yaxis=dict(title="Share Price ($)", tickprefix="$", side="left"),
            yaxis2=dict(
                title="Portfolio Value ($)", tickprefix="$",
                overlaying="y", side="right", showgrid=False,
            ),
            hovermode="x unified",
        )
        fig_json = json.dumps(fig, cls=plotly.utils.PlotlyJSONEncoder)

    except Exception:
        import traceback
        error = traceback.format_exc(limit=5)

    return jsonify(fig_json=fig_json, error=error, warning=warning, rows=rows, summary=summary)


# ── NAV Erosion Portfolio Screener ─────────────────────────────────────────────

@app.route("/api/nav-erosion-portfolio/list", methods=["GET", "POST"])
def nav_erosion_portfolio_list():
    conn = get_connection()
    ensure_tables_exist(conn)
    cur = conn.cursor()

    if request.method == "GET":
        cur.execute(
            "SELECT ticker, amount, reinvest_pct FROM nav_erosion_portfolio_list ORDER BY sort_order"
        )
        rows = [{"ticker": r[0], "amount": r[1], "reinvest_pct": r[2]} for r in cur.fetchall()]
        conn.close()
        return jsonify(rows=rows)

    # POST — replace list
    data = request.get_json(force=True, silent=True) or {}
    rows = data.get("rows", [])

    if len(rows) > 80:
        conn.close()
        return jsonify(error="Maximum 80 ETFs allowed.")

    validated = []
    for i, r in enumerate(rows):
        ticker = str(r.get("ticker", "")).strip().upper()
        if not ticker:
            conn.close()
            return jsonify(error=f"Row {i+1}: ticker is required.")
        try:
            amount = float(r.get("amount", 0))
        except (TypeError, ValueError):
            conn.close()
            return jsonify(error=f"Row {i+1}: invalid amount.")
        if amount <= 0:
            conn.close()
            return jsonify(error=f"Row {i+1}: amount must be greater than 0.")
        try:
            reinvest_pct = float(r.get("reinvest_pct", 0))
        except (TypeError, ValueError):
            conn.close()
            return jsonify(error=f"Row {i+1}: invalid reinvest %.")
        if reinvest_pct < 0 or reinvest_pct > 100:
            conn.close()
            return jsonify(error=f"Row {i+1}: reinvest % must be 0–100.")
        validated.append((ticker, amount, reinvest_pct, i))

    cur.execute("DELETE FROM nav_erosion_portfolio_list")
    for ticker, amount, reinvest_pct, sort_order in validated:
        cur.execute(
            "INSERT INTO nav_erosion_portfolio_list (ticker, amount, reinvest_pct, sort_order) "
            "VALUES (?, ?, ?, ?)",
            (ticker, amount, reinvest_pct, sort_order),
        )
    conn.commit()
    conn.close()
    return jsonify(ok=True)


@app.route("/api/nav-erosion-portfolio/saved", methods=["GET", "POST"])
def nav_erosion_portfolio_saved():
    import json as _json
    conn = get_connection()
    ensure_tables_exist(conn)
    cur = conn.cursor()

    if request.method == "GET":
        cur.execute(
            "SELECT id, name, created_at, start_date, end_date "
            "FROM nav_erosion_saved_backtests ORDER BY created_at DESC"
        )
        saved = [
            {
                "id": r[0], "name": r[1],
                "created_at": r[2] or "",
                "start_date": r[3], "end_date": r[4],
            }
            for r in cur.fetchall()
        ]
        conn.close()
        return jsonify(saved=saved)

    # POST — save a new named backtest
    data = request.get_json(force=True, silent=True) or {}
    name = str(data.get("name", "")).strip()
    if not name:
        conn.close()
        return jsonify(error="Name is required.")
    if len(name) > 200:
        conn.close()
        return jsonify(error="Name must be 200 characters or less.")
    rows_input = data.get("rows", [])
    start = str(data.get("start", ""))
    end = str(data.get("end", ""))
    rows_json = _json.dumps(rows_input)
    cur.execute(
        "INSERT INTO nav_erosion_saved_backtests (name, start_date, end_date, rows_json) "
        "VALUES (?, ?, ?, ?)",
        (name, start, end, rows_json),
    )
    new_id = cur.lastrowid
    conn.commit()
    conn.close()
    return jsonify(ok=True, id=new_id)


@app.route("/api/nav-erosion-portfolio/saved/<int:saved_id>", methods=["GET", "PUT", "DELETE"])
def nav_erosion_portfolio_saved_item(saved_id):
    import json as _json
    conn = get_connection()
    cur = conn.cursor()

    if request.method == "DELETE":
        cur.execute("DELETE FROM nav_erosion_saved_backtests WHERE id = ?", (saved_id,))
        conn.commit()
        conn.close()
        return jsonify(ok=True)

    if request.method == "PUT":
        data = request.get_json(force=True, silent=True) or {}
        name = str(data.get("name", "")).strip()
        if not name:
            conn.close()
            return jsonify(error="Name is required.")
        if len(name) > 200:
            conn.close()
            return jsonify(error="Name must be 200 characters or less.")
        rows_input = data.get("rows", [])
        start = str(data.get("start", ""))
        end = str(data.get("end", ""))
        rows_json = _json.dumps(rows_input)
        cur.execute(
            "UPDATE nav_erosion_saved_backtests "
            "SET name=?, start_date=?, end_date=?, rows_json=?, created_at=CURRENT_TIMESTAMP "
            "WHERE id=?",
            (name, start, end, rows_json, saved_id),
        )
        conn.commit()
        conn.close()
        return jsonify(ok=True)

    # GET — load one saved backtest
    cur.execute(
        "SELECT name, start_date, end_date, rows_json "
        "FROM nav_erosion_saved_backtests WHERE id = ?",
        (saved_id,),
    )
    row = cur.fetchone()
    conn.close()
    if not row:
        return jsonify(error="Not found."), 404
    return jsonify(
        name=row[0], start=row[1], end=row[2],
        rows=_json.loads(row[3]),
    )


@app.route("/api/nav-erosion-portfolio/data", methods=["POST"])
def nav_erosion_portfolio_data():
    import yfinance as yf
    import warnings
    warnings.filterwarnings("ignore")

    data = request.get_json(force=True, silent=True) or {}
    start_date = data.get("start", "")
    end_date = data.get("end", "")
    rows_input = data.get("rows", [])

    if not start_date or not end_date:
        return jsonify(error="Please select both start and end dates.")
    if not rows_input:
        return jsonify(error="No ETFs provided.")
    if len(rows_input) > 80:
        return jsonify(error="Maximum 80 ETFs allowed.")

    # Validate rows
    validated = []
    for i, r in enumerate(rows_input):
        ticker = str(r.get("ticker", "")).strip().upper()
        if not ticker:
            return jsonify(error=f"Row {i+1}: ticker is required.")
        try:
            amount = float(r.get("amount", 0))
        except (TypeError, ValueError):
            return jsonify(error=f"Row {i+1}: invalid amount.")
        if amount <= 0:
            return jsonify(error=f"Row {i+1}: amount must be greater than 0.")
        try:
            reinvest_pct = float(r.get("reinvest_pct", 0))
        except (TypeError, ValueError):
            return jsonify(error=f"Row {i+1}: invalid reinvest %.")
        if reinvest_pct < 0 or reinvest_pct > 100:
            return jsonify(error=f"Row {i+1}: reinvest % must be 0–100.")
        validated.append({"ticker": ticker, "amount": amount, "reinvest_pct": reinvest_pct})

    unique_tickers = list(dict.fromkeys(r["ticker"] for r in validated))

    # Batch download all tickers
    try:
        from datetime import datetime as _dt
        raw = yf.download(
            unique_tickers,
            start=start_date, end=end_date,
            interval="1d", auto_adjust=False, actions=True,
            group_by="ticker", progress=False,
        )
    except Exception as e:
        return jsonify(error=f"Failed to fetch data: {str(e)}")

    requested_start = _dt.strptime(start_date, "%Y-%m-%d").date()

    def get_ticker_df(sym):
        try:
            if isinstance(raw.columns, pd.MultiIndex):
                top_keys = raw.columns.get_level_values(0).unique()
                if sym in top_keys:
                    sub = raw[sym]
                    close = sub["Close"] if "Close" in sub.columns else None
                    divs = sub["Dividends"] if "Dividends" in sub.columns else pd.Series(0.0, index=raw.index)
                elif "Close" in top_keys:
                    close = raw["Close"][sym] if sym in raw["Close"].columns else raw["Close"].iloc[:, 0]
                    divs_df = raw["Dividends"] if "Dividends" in top_keys else None
                    if divs_df is not None:
                        divs = divs_df[sym] if sym in divs_df.columns else pd.Series(0.0, index=raw.index)
                    else:
                        divs = pd.Series(0.0, index=raw.index)
                else:
                    return None, None
            else:
                close = raw["Close"] if "Close" in raw.columns else None
                divs = raw["Dividends"] if "Dividends" in raw.columns else pd.Series(0.0, index=raw.index)
        except Exception:
            return None, None
        if close is None:
            return None, None
        return close, divs

    results = []
    for r in validated:
        sym = r["ticker"]
        amount = r["amount"]
        reinvest_pct = r["reinvest_pct"]

        close, divs = get_ticker_df(sym)

        if close is None or close.dropna().empty:
            results.append({
                "ticker": sym, "amount": amount, "reinvest_pct": reinvest_pct,
                "error": f"No data found for {sym}.",
            })
            continue

        monthly_close = close.resample("ME").last()
        monthly_divs = divs.resample("ME").sum()

        df = pd.DataFrame({"price": monthly_close, "div": monthly_divs}).dropna(subset=["price"])
        df["div"] = df["div"].fillna(0.0)

        if df.empty:
            results.append({
                "ticker": sym, "amount": amount, "reinvest_pct": reinvest_pct,
                "error": f"No usable monthly data for {sym}.",
            })
            continue

        actual_start = df.index[0].date()
        warning = None
        if (actual_start - requested_start).days > 30:
            warning = (
                f"{sym} only has data going back to {actual_start.strftime('%B %d, %Y')}. "
                f"Results start from that date."
            )

        initial_price = float(df["price"].iloc[0])
        if initial_price == 0:
            results.append({
                "ticker": sym, "amount": amount, "reinvest_pct": reinvest_pct,
                "error": f"Initial price for {sym} is zero.",
            })
            continue

        current_shares = amount / initial_price
        cumul_dist = 0.0
        cumul_reinvested = 0.0
        cumul_divs_per_share = 0.0

        for dt, row in df.iterrows():
            price = float(row["price"])
            div_per_share = float(row["div"])
            total_dist = div_per_share * current_shares
            reinvest_amt = total_dist * reinvest_pct / 100.0
            shares_bought = (reinvest_amt / price) if price > 0 else 0.0
            current_shares += shares_bought
            cumul_dist += total_dist
            cumul_reinvested += reinvest_amt
            cumul_divs_per_share += div_per_share

        final_price = float(df["price"].iloc[-1])
        portfolio_val = current_shares * final_price
        breakeven_final = (amount / final_price) if final_price > 0 else 0.0
        final_deficit = breakeven_final - current_shares
        has_erosion_at_end = final_deficit > 0
        price_delta_pct = (final_price - initial_price) / initial_price * 100 if initial_price else 0.0
        gain_loss_dollar = portfolio_val - amount
        gain_loss_pct = gain_loss_dollar / amount * 100 if amount else 0.0
        cash_taken = cumul_dist - cumul_reinvested
        total_return_dollar = portfolio_val + cash_taken - amount
        total_return_pct = total_return_dollar / amount * 100 if amount else 0.0

        # Coverage ratio: (total price change + total divs per share) / total divs per share
        total_price_change = final_price - initial_price
        if cumul_divs_per_share > 0:
            coverage_ratio = round((total_price_change + cumul_divs_per_share) / cumul_divs_per_share, 4)
        else:
            coverage_ratio = None

        results.append({
            "ticker": sym,
            "amount": round(amount, 2),
            "reinvest_pct": round(reinvest_pct, 1),
            "start_price": round(initial_price, 4),
            "end_price": round(final_price, 4),
            "price_delta_pct": round(price_delta_pct, 2),
            "total_dist": round(cumul_dist, 2),
            "total_reinvested": round(cumul_reinvested, 2),
            "final_value": round(portfolio_val, 2),
            "gain_loss_dollar": round(gain_loss_dollar, 2),
            "gain_loss_pct": round(gain_loss_pct, 2),
            "total_return_dollar": round(total_return_dollar, 2),
            "total_return_pct": round(total_return_pct, 2),
            "has_erosion": has_erosion_at_end,
            "final_deficit": round(final_deficit, 4),
            "coverage_ratio": coverage_ratio,
            "warning": warning,
            "error": None,
        })

    return jsonify(results=results)


# ── Portfolio Income Simulator ─────────────────────────────────────────────────

@app.route("/api/pis/portfolio-tickers")
def pis_portfolio_tickers():
    """Return all live portfolio tickers for the picker modal."""
    pid = get_profile_id()
    conn = get_connection()
    rows = conn.execute(
        """SELECT ticker, description, classification_type,
                  purchase_value, reinvest, current_annual_yield
           FROM all_account_info
           WHERE current_price IS NOT NULL AND profile_id = ?
           ORDER BY ticker""",
        (pid,),
    ).fetchall()
    conn.close()
    tickers = [
        {
            "ticker": r[0],
            "description": r[1] or "",
            "type": r[2] or "",
            "amount": round(float(r[3] or 0), 2),
            "drip": (r[4] or "").upper() == "Y",
            "current_yield": round(float(r[5] or 0), 2),
        }
        for r in rows
    ]
    return jsonify(tickers=tickers)


@app.route("/api/pis/list", methods=["GET", "POST"])
def pis_list():
    conn = get_connection()
    ensure_tables_exist(conn)
    cur = conn.cursor()

    if request.method == "GET":
        cur.execute(
            "SELECT ticker, amount, reinvest_pct, yield_override "
            "FROM portfolio_income_sim_list ORDER BY sort_order"
        )
        rows = [
            {"ticker": r[0], "amount": r[1], "reinvest_pct": r[2], "yield_override": r[3]}
            for r in cur.fetchall()
        ]
        conn.close()
        return jsonify(rows=rows)

    # POST — replace list
    data = request.get_json(force=True, silent=True) or {}
    rows = data.get("rows", [])
    if len(rows) > 80:
        conn.close()
        return jsonify(error="Maximum 80 ETFs allowed.")

    validated = []
    for i, r in enumerate(rows):
        ticker = str(r.get("ticker", "")).strip().upper()
        if not ticker:
            continue  # skip empty rows
        try:
            amount = float(r.get("amount", 0))
        except (TypeError, ValueError):
            conn.close()
            return jsonify(error=f"Row {i+1}: invalid amount.")
        if amount <= 0:
            conn.close()
            return jsonify(error=f"Row {i+1}: amount must be greater than 0.")
        try:
            reinvest_pct = float(r.get("reinvest_pct", 0))
        except (TypeError, ValueError):
            conn.close()
            return jsonify(error=f"Row {i+1}: invalid reinvest %.")
        if reinvest_pct < 0 or reinvest_pct > 100:
            conn.close()
            return jsonify(error=f"Row {i+1}: reinvest % must be 0–100.")
        yield_override = r.get("yield_override")
        if yield_override is not None and yield_override != "":
            try:
                yield_override = float(yield_override)
                if yield_override <= 0:
                    yield_override = None
            except (TypeError, ValueError):
                yield_override = None
        else:
            yield_override = None
        validated.append((ticker, amount, reinvest_pct, yield_override, len(validated)))

    cur.execute("DELETE FROM portfolio_income_sim_list")
    for ticker, amount, reinvest_pct, yield_override, sort_order in validated:
        cur.execute(
            "INSERT INTO portfolio_income_sim_list (ticker, amount, reinvest_pct, yield_override, sort_order) "
            "VALUES (?, ?, ?, ?, ?)",
            (ticker, amount, reinvest_pct, yield_override, sort_order),
        )
    conn.commit()
    conn.close()
    return jsonify(ok=True)


@app.route("/api/pis/saved", methods=["GET", "POST"])
def pis_saved():
    import json as _json
    conn = get_connection()
    ensure_tables_exist(conn)
    cur = conn.cursor()

    if request.method == "GET":
        cur.execute(
            "SELECT id, name, created_at, mode, start_date, end_date, market_type, duration_months "
            "FROM portfolio_income_sim_saved ORDER BY created_at DESC"
        )
        saved = [
            {
                "id": r[0], "name": r[1], "created_at": r[2] or "",
                "mode": r[3], "start_date": r[4], "end_date": r[5],
                "market_type": r[6], "duration_months": r[7],
            }
            for r in cur.fetchall()
        ]
        conn.close()
        return jsonify(saved=saved)

    # POST — save new
    data = request.get_json(force=True, silent=True) or {}
    name = str(data.get("name", "")).strip()
    if not name:
        conn.close()
        return jsonify(error="Name is required.")
    if len(name) > 200:
        conn.close()
        return jsonify(error="Name must be 200 characters or less.")
    rows_input = data.get("rows", [])
    mode = str(data.get("mode", "historical"))
    start = str(data.get("start", ""))
    end = str(data.get("end", ""))
    market_type = str(data.get("market_type", ""))
    duration_months = data.get("duration_months")
    if duration_months is not None:
        try:
            duration_months = int(duration_months)
        except (TypeError, ValueError):
            duration_months = None
    rows_json = _json.dumps(rows_input)
    comparison_input = data.get("comparison_tickers", [])
    comparison_json = _json.dumps(comparison_input) if comparison_input else None
    cur.execute(
        "INSERT INTO portfolio_income_sim_saved "
        "(name, mode, start_date, end_date, market_type, duration_months, rows_json, comparison_json) "
        "VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
        (name, mode, start, end, market_type, duration_months, rows_json, comparison_json),
    )
    new_id = cur.lastrowid
    conn.commit()
    conn.close()
    return jsonify(ok=True, id=new_id)


@app.route("/api/pis/saved/<int:saved_id>", methods=["GET", "PUT", "DELETE"])
def pis_saved_item(saved_id):
    import json as _json
    conn = get_connection()
    cur = conn.cursor()

    if request.method == "DELETE":
        cur.execute("DELETE FROM portfolio_income_sim_saved WHERE id = ?", (saved_id,))
        conn.commit()
        conn.close()
        return jsonify(ok=True)

    if request.method == "PUT":
        data = request.get_json(force=True, silent=True) or {}
        name = str(data.get("name", "")).strip()
        if not name:
            conn.close()
            return jsonify(error="Name is required.")
        if len(name) > 200:
            conn.close()
            return jsonify(error="Name must be 200 characters or less.")
        rows_input = data.get("rows", [])
        mode = str(data.get("mode", "historical"))
        start = str(data.get("start", ""))
        end = str(data.get("end", ""))
        market_type = str(data.get("market_type", ""))
        duration_months = data.get("duration_months")
        if duration_months is not None:
            try:
                duration_months = int(duration_months)
            except (TypeError, ValueError):
                duration_months = None
        rows_json = _json.dumps(rows_input)
        comparison_input = data.get("comparison_tickers", [])
        comparison_json = _json.dumps(comparison_input) if comparison_input else None
        cur.execute(
            "UPDATE portfolio_income_sim_saved "
            "SET name=?, mode=?, start_date=?, end_date=?, market_type=?, "
            "    duration_months=?, rows_json=?, comparison_json=?, created_at=CURRENT_TIMESTAMP "
            "WHERE id=?",
            (name, mode, start, end, market_type, duration_months, rows_json, comparison_json, saved_id),
        )
        conn.commit()
        conn.close()
        return jsonify(ok=True)

    # GET
    cur.execute(
        "SELECT name, mode, start_date, end_date, market_type, duration_months, rows_json, comparison_json "
        "FROM portfolio_income_sim_saved WHERE id = ?",
        (saved_id,),
    )
    row = cur.fetchone()
    conn.close()
    if not row:
        return jsonify(error="Not found."), 404
    comparison = _json.loads(row[7]) if row[7] else []
    return jsonify(
        name=row[0], mode=row[1], start=row[2], end=row[3],
        market_type=row[4], duration_months=row[5],
        rows=_json.loads(row[6]),
        comparison_tickers=comparison,
    )


@app.route("/api/pis/run", methods=["POST"])
def pis_run():
    import traceback as _tb
    try:
        return _pis_run_inner()
    except Exception as _e:
        return jsonify(error=f"Server error: {str(_e)}", detail=_tb.format_exc())


def _pis_run_inner():
    import yfinance as yf
    import numpy as np
    import warnings
    warnings.filterwarnings("ignore")

    data = request.get_json(force=True, silent=True) or {}
    mode = data.get("mode", "historical")
    rows_input = data.get("rows", [])
    # comparison_tickers: accept either strings or {ticker, reinvest_pct, amount} objects
    raw_comp = data.get("comparison_tickers", [])
    comparison_tickers = []
    for ct in raw_comp:
        if isinstance(ct, dict):
            t = str(ct.get("ticker", "")).strip().upper()
            rp = float(ct.get("reinvest_pct", 0))
            amt = float(ct.get("amount", 10000))
            yo = ct.get("yield_override", None)
            if yo is not None:
                try:
                    yo = float(yo)
                except (ValueError, TypeError):
                    yo = None
            if t:
                comparison_tickers.append({"ticker": t, "reinvest_pct": max(0.0, min(100.0, rp)), "amount": max(1.0, amt), "yield_override": yo})
        elif isinstance(ct, str) and ct.strip():
            comparison_tickers.append({"ticker": ct.strip().upper(), "reinvest_pct": 0.0, "amount": 10000.0, "yield_override": None})

    if not rows_input and not comparison_tickers:
        return jsonify(error="No ETFs provided.")
    if len(rows_input) > 80:
        return jsonify(error="Maximum 80 ETFs allowed.")

    # Validate portfolio rows
    validated = []
    for i, r in enumerate(rows_input):
        ticker = str(r.get("ticker", "")).strip().upper()
        if not ticker:
            continue
        try:
            amount = float(r.get("amount", 0))
        except (TypeError, ValueError):
            return jsonify(error=f"Row {i+1}: invalid amount.")
        if amount <= 0:
            return jsonify(error=f"Row {i+1}: amount must be greater than 0.")
        try:
            reinvest_pct = float(r.get("reinvest_pct", 0))
        except (TypeError, ValueError):
            return jsonify(error=f"Row {i+1}: invalid reinvest %.")
        if reinvest_pct < 0 or reinvest_pct > 100:
            return jsonify(error=f"Row {i+1}: reinvest % must be 0–100.")
        yield_override = r.get("yield_override")
        if yield_override is not None and yield_override != "":
            try:
                yield_override = float(yield_override)
                if yield_override <= 0:
                    yield_override = None
            except (TypeError, ValueError):
                yield_override = None
        else:
            yield_override = None
        validated.append({
            "ticker": ticker, "amount": amount,
            "reinvest_pct": reinvest_pct, "yield_override": yield_override,
            "is_comparison": False,
        })

    # Add comparison tickers (normalized $10,000, per-ticker reinvest %)
    reinvest_compare = data.get("reinvest_compare", False)
    reinvest_compare_pct = float(data.get("reinvest_compare_pct", 50))

    for ct in comparison_tickers:
        validated.append({
            "ticker": ct["ticker"], "amount": ct["amount"],
            "reinvest_pct": ct["reinvest_pct"], "yield_override": ct["yield_override"],
            "is_comparison": True,
        })

    if not validated:
        return jsonify(error="No valid ETFs to process.")

    unique_tickers = list(dict.fromkeys(r["ticker"] for r in validated))

    # ── HISTORICAL MODE ──────────────────────────────────────────────────────
    if mode == "historical":
        start_date = data.get("start", "")
        end_date = data.get("end", "")
        if not start_date or not end_date:
            return jsonify(error="Please select both start and end dates.")

        try:
            from datetime import datetime as _dt
            raw = yf.download(
                unique_tickers,
                start=start_date, end=end_date,
                interval="1d", auto_adjust=False, actions=True,
                group_by="ticker", progress=False,
            )
        except Exception as e:
            return jsonify(error=f"Failed to fetch data: {str(e)}")

        requested_start = _dt.strptime(start_date, "%Y-%m-%d").date()

        def get_ticker_df(sym):
            try:
                if isinstance(raw.columns, pd.MultiIndex):
                    top_keys = raw.columns.get_level_values(0).unique()
                    if sym in top_keys:
                        sub = raw[sym]
                        close = sub["Close"] if "Close" in sub.columns else None
                        divs = sub["Dividends"] if "Dividends" in sub.columns else pd.Series(0.0, index=raw.index)
                    elif "Close" in top_keys:
                        close = raw["Close"][sym] if sym in raw["Close"].columns else raw["Close"].iloc[:, 0]
                        divs_df = raw["Dividends"] if "Dividends" in top_keys else None
                        if divs_df is not None:
                            divs = divs_df[sym] if sym in divs_df.columns else pd.Series(0.0, index=raw.index)
                        else:
                            divs = pd.Series(0.0, index=raw.index)
                    else:
                        return None, None
                else:
                    close = raw["Close"] if "Close" in raw.columns else None
                    divs = raw["Dividends"] if "Dividends" in raw.columns else pd.Series(0.0, index=raw.index)
            except Exception:
                return None, None
            if close is None:
                return None, None
            return close, divs

        def _run_hist_sim(df, amount, reinvest_pct):
            initial_price = float(df["price"].iloc[0])
            current_shares = amount / initial_price
            cumul_dist = 0.0
            cumul_reinvested = 0.0
            m_prices, m_vals, m_divs = [], [], []
            for _, row in df.iterrows():
                price = float(row["price"])
                div_per_share = float(row["div"])
                total_dist_m = div_per_share * current_shares
                reinvest_amt = total_dist_m * reinvest_pct / 100.0
                shares_bought = (reinvest_amt / price) if price > 0 else 0.0
                current_shares += shares_bought
                cumul_dist += total_dist_m
                cumul_reinvested += reinvest_amt
                m_prices.append(round(price, 4))
                m_vals.append(round(current_shares * price, 2))
                m_divs.append(round(total_dist_m, 2))
            final_price = float(df["price"].iloc[-1])
            portfolio_val = current_shares * final_price
            breakeven_final = (amount / final_price) if final_price > 0 else 0.0
            final_deficit = breakeven_final - current_shares
            price_delta_pct = (final_price - initial_price) / initial_price * 100 if initial_price else 0.0
            gain_loss_dollar = portfolio_val - amount
            gain_loss_pct = gain_loss_dollar / amount * 100 if amount else 0.0
            effective_yield_pct = cumul_dist / amount * 100 if amount else 0.0
            return {
                "initial_price": initial_price, "final_price": final_price,
                "price_delta_pct": price_delta_pct, "cumul_dist": cumul_dist,
                "cumul_reinvested": cumul_reinvested, "portfolio_val": portfolio_val,
                "gain_loss_dollar": gain_loss_dollar, "gain_loss_pct": gain_loss_pct,
                "effective_yield_pct": effective_yield_pct, "has_erosion": final_deficit > 0,
                "final_deficit": final_deficit,
                "m_prices": m_prices, "m_vals": m_vals, "m_divs": m_divs,
            }

        results = []
        for r in validated:
            sym = r["ticker"]
            amount = r["amount"]
            reinvest_pct = r["reinvest_pct"]

            close, divs = get_ticker_df(sym)

            if close is None or close.dropna().empty:
                results.append({"ticker": sym, "amount": amount, "reinvest_pct": reinvest_pct,
                                "is_comparison": r["is_comparison"],
                                "error": f"No data found for {sym}.",
                                "monthly_prices": [], "monthly_portfolio_vals": [], "monthly_dividends": []})
                continue

            monthly_close = close.resample("ME").last()
            monthly_divs = divs.resample("ME").sum()
            df = pd.DataFrame({"price": monthly_close, "div": monthly_divs}).dropna(subset=["price"])
            df["div"] = df["div"].fillna(0.0)

            if df.empty:
                results.append({"ticker": sym, "amount": amount, "reinvest_pct": reinvest_pct,
                                "is_comparison": r["is_comparison"],
                                "error": f"No usable monthly data for {sym}.",
                                "monthly_prices": [], "monthly_portfolio_vals": [], "monthly_dividends": []})
                continue

            actual_start = df.index[0].date()
            warning = None
            if (actual_start - requested_start).days > 30:
                warning = (
                    f"{sym} only has data going back to {actual_start.strftime('%B %d, %Y')}. "
                    f"Results start from that date."
                )

            initial_price = float(df["price"].iloc[0])
            if initial_price == 0:
                results.append({"ticker": sym, "amount": amount, "reinvest_pct": reinvest_pct,
                                "is_comparison": r["is_comparison"],
                                "error": f"Initial price for {sym} is zero.",
                                "monthly_prices": [], "monthly_portfolio_vals": [], "monthly_dividends": []})
                continue

            date_labels = [dt.strftime("%Y-%m") for dt in df.index]

            # Determine which reinvest runs to perform
            # Comparison tickers use their own per-ticker reinvest_pct (single run)
            if reinvest_compare and not r["is_comparison"]:
                runs = [
                    (0.0, "baseline"),
                    (reinvest_compare_pct, "reinvested"),
                ]
            else:
                runs = [(reinvest_pct, None)]

            for run_pct, compare_group in runs:
                s = _run_hist_sim(df, amount, run_pct)
                rec = {
                    "ticker": sym,
                    "is_comparison": r["is_comparison"],
                    "amount": round(amount, 2),
                    "reinvest_pct": round(run_pct, 1),
                    "start_price": round(s["initial_price"], 4),
                    "end_price": round(s["final_price"], 4),
                    "price_delta_pct": round(s["price_delta_pct"], 2),
                    "total_dist": round(s["cumul_dist"], 2),
                    "total_reinvested": round(s["cumul_reinvested"], 2),
                    "final_value": round(s["portfolio_val"], 2),
                    "gain_loss_dollar": round(s["gain_loss_dollar"], 2),
                    "gain_loss_pct": round(s["gain_loss_pct"], 2),
                    "effective_yield_pct": round(s["effective_yield_pct"], 2),
                    "ttm_yield_pct": None,
                    "has_erosion": s["has_erosion"],
                    "final_deficit": round(s["final_deficit"], 4),
                    "monthly_prices": s["m_prices"],
                    "monthly_portfolio_vals": s["m_vals"],
                    "monthly_dividends": s["m_divs"],
                    "date_labels": date_labels,
                    "warning": warning,
                    "error": None,
                }
                if compare_group is not None:
                    rec["compare_group"] = compare_group
                results.append(rec)

        return jsonify(results=results)

    # ── SIMULATION MODE ──────────────────────────────────────────────────────
    if mode == "simulate":
        market_type = data.get("market_type", "neutral")
        duration_months = int(data.get("duration_months", 36))
        if duration_months < 1 or duration_months > 600:
            return jsonify(error="Duration must be between 1 and 600 months.")

        bias_map = {"bullish": +0.010, "bearish": -0.015, "neutral": 0.0}
        vol_mult_map = {"bullish": 0.9, "bearish": 1.2, "neutral": 1.0}
        bias = bias_map.get(market_type, 0.0)
        vol_mult = vol_mult_map.get(market_type, 1.0)

        results = []
        for r in validated:
            sym = r["ticker"]
            amount = r["amount"]
            reinvest_pct = r["reinvest_pct"]
            yo = r["yield_override"]

            try:
                hist = yf.Ticker(sym).history(period="1y", auto_adjust=False, actions=True)
            except Exception as e:
                results.append({"ticker": sym, "amount": amount, "reinvest_pct": reinvest_pct,
                                "is_comparison": r["is_comparison"],
                                "error": f"Failed to fetch data for {sym}: {str(e)}",
                                "monthly_prices": [], "monthly_portfolio_vals": [], "monthly_dividends": []})
                continue

            if hist is None or hist.empty:
                results.append({"ticker": sym, "amount": amount, "reinvest_pct": reinvest_pct,
                                "is_comparison": r["is_comparison"],
                                "error": f"No data found for {sym}.",
                                "monthly_prices": [], "monthly_portfolio_vals": [], "monthly_dividends": []})
                continue

            close = hist["Close"]
            divs = hist["Dividends"] if "Dividends" in hist.columns else pd.Series(0.0, index=hist.index)

            if close.dropna().empty:
                results.append({"ticker": sym, "amount": amount, "reinvest_pct": reinvest_pct,
                                "is_comparison": r["is_comparison"],
                                "error": f"No data found for {sym}.",
                                "monthly_prices": [], "monthly_portfolio_vals": [], "monthly_dividends": []})
                continue

            current_price = float(close.dropna().iloc[-1])
            if current_price <= 0:
                results.append({"ticker": sym, "amount": amount, "reinvest_pct": reinvest_pct,
                                "is_comparison": r["is_comparison"],
                                "error": f"Could not determine current price for {sym}.",
                                "monthly_prices": [], "monthly_portfolio_vals": [], "monthly_dividends": []})
                continue

            # TTM yield
            if yo is not None:
                ttm_yield = yo / 100.0
            else:
                ttm_divs_sum = float(divs.sum()) if divs is not None else 0.0
                ttm_yield = ttm_divs_sum / current_price if current_price > 0 else 0.0

            # Historical monthly return statistics
            monthly_returns = close.dropna().resample("ME").last().pct_change().dropna()
            if len(monthly_returns) >= 2:
                hist_sigma = float(monthly_returns.std())
                hist_mu = float(monthly_returns.mean())
            else:
                hist_sigma = 0.05
                hist_mu = 0.0

            # Historical skewness (captures covered-call truncated upside, etc.)
            if len(monthly_returns) >= 12:
                hist_skew = float(monthly_returns.skew())
                hist_kurt = float(monthly_returns.kurtosis())  # excess kurtosis
            else:
                hist_skew = 0.0
                hist_kurt = 0.0

            # Use ticker's own historical drift + market bias overlay
            mu = hist_mu + bias
            sigma = hist_sigma * vol_mult
            SIGMA_CAP = 0.25
            sigma_capped = sigma > SIGMA_CAP
            sigma = min(sigma, SIGMA_CAP)

            # Monte-Carlo GBM — 300 paths with skew-adjusted returns
            N_PATHS = 300
            drift = mu - 0.5 * sigma ** 2
            np.random.seed(None)
            Z = np.random.normal(0.0, 1.0, (N_PATHS, duration_months))
            # Cornish-Fisher expansion: adjust for skewness and excess kurtosis
            # This transforms normal draws to approximate the historical return shape
            skew_c = max(-2.0, min(2.0, hist_skew))   # clamp to avoid extreme distortion
            kurt_c = max(-2.0, min(7.0, hist_kurt))
            Z_adj = (Z
                     + (skew_c / 6.0) * (Z ** 2 - 1)
                     + (kurt_c / 24.0) * (Z ** 3 - 3 * Z)
                     - (skew_c ** 2 / 36.0) * (2 * Z ** 3 - 5 * Z))
            log_rets = drift + sigma * Z_adj
            cum_log = np.cumsum(
                np.hstack([np.zeros((N_PATHS, 1)), log_rets]), axis=1
            )
            price_floor = max(current_price * 0.0001, 1e-10)
            price_matrix = np.maximum(current_price * np.exp(cum_log), price_floor)
            prices = [float(v) for v in np.median(price_matrix, axis=0)]

            # Helper: simulate month-by-month for a given reinvest_pct
            def _run_fwd_sim(prices_arr, amount_v, reinvest_pct_v, ttm_yield_v, cur_price):
                initial_shares = amount_v / cur_price
                cs = initial_shares
                cd = 0.0; cr = 0.0
                mp, mv, md = [], [], []
                for price in prices_arr[1:]:
                    pct_chg = (price - cur_price) / cur_price * 100
                    if pct_chg >= 10:
                        factor = min(1.0 + (pct_chg - 10) * 0.02, 1.30)
                    elif pct_chg >= -10:
                        factor = 1.0
                    elif pct_chg >= -20:
                        factor = 0.85
                    elif pct_chg >= -30:
                        factor = 0.70
                    else:
                        factor = max(0.40, 1.0 + pct_chg * 0.02)
                    dist_ps = (ttm_yield_v / 12) * cur_price * factor
                    total_d = dist_ps * cs
                    ra = total_d * (reinvest_pct_v / 100)
                    sb = ra / price if price > 0 else 0.0
                    cs += sb; cd += total_d; cr += ra
                    mp.append(round(price, 4))
                    mv.append(round(cs * price, 2))
                    md.append(round(total_d, 2))
                fp = prices_arr[-1]
                fv = cs * fp
                be = amount_v / fp if fp else 0.0
                fd = be - cs
                return {
                    "final_price": fp, "final_value": fv, "final_deficit": fd,
                    "cumul_dist": cd, "cumul_reinvested": cr,
                    "gain_loss_dollar": fv - amount_v,
                    "gain_loss_pct": (fv - amount_v) / amount_v * 100 if amount_v else 0.0,
                    "eff_yield_pct": cd / amount_v * 100 if amount_v else 0.0,
                    "has_erosion": bool(fd > 0),
                    "mp": mp, "mv": mv, "md": md,
                }

            date_labels = [f"Month {i+1}" for i in range(duration_months)]
            price_delta_pct = (prices[-1] - current_price) / current_price * 100 if current_price else 0.0
            sim_warning = (
                f"{sym} has extreme historical volatility; monthly sigma was capped "
                f"at 25% (raw: {hist_sigma * vol_mult * 100:.0f}%) to prevent "
                f"unrealistic simulation outcomes."
            ) if sigma_capped else None

            # Pass per-ticker historical stats for transparency
            sim_stats = {
                "hist_mean_monthly": round(hist_mu * 100, 2),
                "hist_sigma_monthly": round(hist_sigma * 100, 2),
                "hist_skewness": round(hist_skew, 2) if len(monthly_returns) >= 12 else None,
            }

            # Determine which reinvest runs to perform
            # Comparison tickers use their own per-ticker reinvest_pct (single run)
            if reinvest_compare and not r["is_comparison"]:
                runs = [
                    (0.0, "baseline"),
                    (reinvest_compare_pct, "reinvested"),
                ]
            else:
                runs = [(reinvest_pct, None)]

            for run_pct, compare_group in runs:
                s = _run_fwd_sim(prices, amount, run_pct, ttm_yield, current_price)
                rec = {
                    "ticker": sym,
                    "is_comparison": r["is_comparison"],
                    "amount": round(amount, 2),
                    "reinvest_pct": round(run_pct, 1),
                    "start_price": round(current_price, 4),
                    "end_price": round(s["final_price"], 4),
                    "price_delta_pct": round(price_delta_pct, 2),
                    "total_dist": round(s["cumul_dist"], 2),
                    "total_reinvested": round(s["cumul_reinvested"], 2),
                    "final_value": round(s["final_value"], 2),
                    "gain_loss_dollar": round(s["gain_loss_dollar"], 2),
                    "gain_loss_pct": round(s["gain_loss_pct"], 2),
                    "effective_yield_pct": round(s["eff_yield_pct"], 2),
                    "ttm_yield_pct": round(ttm_yield * 100, 2),
                    "has_erosion": s["has_erosion"],
                    "final_deficit": round(s["final_deficit"], 4),
                    "monthly_prices": s["mp"],
                    "monthly_portfolio_vals": s["mv"],
                    "monthly_dividends": s["md"],
                    "date_labels": date_labels,
                    "warning": sim_warning,
                    "sim_stats": sim_stats,
                    "error": None,
                }
                if compare_group is not None:
                    rec["compare_group"] = compare_group
                results.append(rec)

        return jsonify(results=results)

    return jsonify(error=f"Unknown mode: {mode}")


# ── Portfolio Analytics ─────────────────────────────────────────────────────────

def _portfolio_sharpe(weights, returns_df, risk_free_annual=0.05):
    import numpy as np
    port_ret = returns_df.dot(weights)
    daily_rf = risk_free_annual / 252
    excess = float(port_ret.mean()) - daily_rf
    std = float(port_ret.std())
    if std == 0:
        return 0.0
    return excess / std * np.sqrt(252)


def _portfolio_max_dd(weights, returns_df):
    port_ret = returns_df.dot(weights)
    cum = (1 + port_ret).cumprod()
    running_max = cum.cummax()
    dd = (cum - running_max) / running_max
    return float(dd.min())


def _portfolio_sortino(w, returns_df):
    """Annualized Sortino ratio for a portfolio with given weights."""
    import numpy as np
    r_p = returns_df.values @ w
    rf_daily = 0.05 / 252
    excess = r_p - rf_daily
    neg = r_p[r_p < 0]
    down_std = float(neg.std()) if len(neg) > 1 else 1e-6
    return (float(excess.mean()) / max(down_std, 1e-6)) * np.sqrt(252)


def _optimize_sharpe(returns_df, current_weights=None, weight_caps=None, turnover_penalty=0.15):
    """Maximize returns using 60% Sharpe + 40% Sortino blend.
    Respects current weights (partial sells), NAV caps, and turnover penalty.
    """
    import numpy as np
    from scipy.optimize import minimize
    n = returns_df.shape[1]
    if current_weights is None:
        current_weights = np.ones(n) / n
    cw = np.array(current_weights)
    if weight_caps is None:
        weight_caps = [0.40] * n
    bounds = []
    for i in range(n):
        cap = weight_caps[i]
        if cap <= 0:
            bounds.append((0.0, 0.0))
        else:
            floor = cw[i] * 0.25 if cw[i] > 0.005 else 0.0
            bounds.append((floor, min(cap, 0.40)))
    n_current = int(np.sum(cw > 0.005))
    def neg_objective(w):
        sharpe = _portfolio_sharpe(w, returns_df)
        sortino = _portfolio_sortino(w, returns_df)
        norm_sharpe = min(max(sharpe, 0), 4) / 4.0
        norm_sortino = min(max(sortino, 0), 6) / 6.0
        score = 0.6 * norm_sharpe + 0.4 * norm_sortino
        turnover = float(np.sum(np.abs(w - cw)))
        n_active = float(np.sum(w > 0.005))
        diversification_loss = max(0, n_current - n_active) * 0.01
        return -(score - turnover_penalty * turnover - diversification_loss)
    constraints = [{"type": "eq", "fun": lambda w: w.sum() - 1}]
    init = _clamp_and_normalize(cw, bounds)
    best = None
    starts = [init]
    rng = np.random.default_rng(42)
    for _ in range(4):
        d = rng.dirichlet(np.ones(n))
        starts.append(_clamp_and_normalize(d, bounds))
    for w0 in starts:
        try:
            res = minimize(neg_objective, w0, method="SLSQP", bounds=bounds,
                           constraints=constraints, options={"maxiter": 1000})
            if res.success and (best is None or res.fun < best.fun):
                best = res
        except Exception:
            pass
    if best is None:
        return init
    return _clamp_and_normalize(best.x, bounds)


def _compute_nav_erosion(close_df, tickers, trading_days=252):
    """Compute annualized price return per ticker. Negative = NAV erosion."""
    import numpy as np
    erosion = {}
    for t in tickers:
        if t not in close_df.columns:
            erosion[t] = 0.0
            continue
        s = close_df[t].dropna()
        if len(s) < 2:
            erosion[t] = 0.0
            continue
        total_ret = float(s.iloc[-1] / s.iloc[0]) - 1.0
        n_days = len(s)
        ann_ret = (1 + total_ret) ** (trading_days / max(n_days, 1)) - 1.0
        erosion[t] = ann_ret
    return erosion


def _adjust_yields_for_nav(yields, nav_returns, tickers):
    """Produce effective yields: yield + nav_return (clamped so minimum is 0).
    A fund yielding 40% but losing 30% NAV has effective yield of 10%.
    A fund yielding 5% with 10% NAV growth keeps its 5% yield (we don't boost beyond raw yield)."""
    import numpy as np
    adjusted = []
    for i, t in enumerate(tickers):
        raw_yield = yields[i]
        nav_ret = nav_returns.get(t, 0.0)
        if nav_ret < 0:
            # Penalize: effective yield = raw yield + nav return (which is negative)
            eff = max(raw_yield + nav_ret, 0.0)
        else:
            # Don't boost yield for price appreciation — that's captured by Sharpe/returns
            eff = raw_yield
        adjusted.append(eff)
    return adjusted


def _nav_weight_caps(nav_returns, tickers, severe=-0.30, moderate=-0.15):
    """Return per-ticker upper weight bounds based on NAV erosion severity.
    Severe erosion (>30% annualized loss): cap at 5%
    Moderate erosion (>15%): cap at 15%
    Otherwise: up to 100%"""
    caps = []
    for t in tickers:
        nav = nav_returns.get(t, 0.0)
        if nav <= severe:
            caps.append(0.05)
        elif nav <= moderate:
            caps.append(0.15)
        else:
            caps.append(1.0)
    return caps


def _clamp_and_normalize(w, bounds):
    """Enforce bounds and renormalize weights."""
    import numpy as np
    w = np.array(w, dtype=float)
    for i, (lo, hi) in enumerate(bounds):
        w[i] = max(lo, min(hi, w[i]))
    s = w.sum()
    if s > 0:
        w = w / s
        # Re-clamp after normalization (may shift slightly)
        for i, (lo, hi) in enumerate(bounds):
            w[i] = max(lo, min(hi, w[i]))
    return w


def _portfolio_income_score(w, returns_df_values, mu, rf_daily=0.05/252):
    """Composite quality score from Sortino, Omega, Calmar, Ulcer Index.
    All inputs are numpy arrays for speed. Returns 0-1 score (higher = better).
    """
    import numpy as np
    r_p = returns_df_values @ w
    n_days = len(r_p)
    if n_days < 30:
        return 0.0
    # Sortino
    excess_daily = r_p - rf_daily
    neg = r_p[r_p < 0]
    down_std = float(neg.std()) if len(neg) > 1 else 1e-6
    sortino = (float(excess_daily.mean()) / max(down_std, 1e-6)) * np.sqrt(252)
    norm_sortino = max(0, min(sortino, 4)) / 4.0
    # Omega
    gains = float(excess_daily[excess_daily > 0].sum())
    losses = abs(float(excess_daily[excess_daily <= 0].sum()))
    omega = gains / max(losses, 1e-8)
    norm_omega = max(0, min(omega, 3) - 0.5) / 2.5
    # Calmar
    cum = np.cumprod(1 + r_p)
    ann_ret = float(cum[-1] ** (252 / n_days) - 1)
    running_max = np.maximum.accumulate(cum)
    dd = (cum - running_max) / np.where(running_max > 0, running_max, 1)
    mdd = abs(float(dd.min()))
    calmar = ann_ret / max(mdd, 1e-6) if ann_ret > 0 else 0
    norm_calmar = max(0, min(calmar, 5)) / 5.0
    # Ulcer Index (lower = better)
    pct_dd = ((cum - running_max) / np.where(running_max > 0, running_max, 1)) * 100
    ulcer = float(np.sqrt((pct_dd ** 2).mean()))
    norm_ulcer = 1.0 - max(0, min(ulcer, 30)) / 30.0
    # Weighted composite: Ulcer 30%, Calmar 25%, Omega 25%, Sortino 20%
    return 0.30 * norm_ulcer + 0.25 * norm_calmar + 0.25 * norm_omega + 0.20 * norm_sortino


def _optimize_income(returns_df, yields, current_weights=None, weight_caps=None, turnover_penalty=0.2, nav_penalties=None):
    """Maximize income: blend yield and income-quality score (Sortino/Omega/Calmar/Ulcer).
    Uses 5 optimizer starts and a 20% per-holding cap to force diversification.
    nav_penalties: per-ticker penalty for NAV erosion.
    """
    import numpy as np
    from scipy.optimize import minimize
    n = returns_df.shape[1]
    mu = returns_df.mean().values * 252
    yields_arr = np.array(yields)
    target_yield = 0.15  # normalize against 15% target, not max ticker yield
    if current_weights is None:
        current_weights = np.ones(n) / n
    cw = np.array(current_weights)
    nav_pen = np.array(nav_penalties) if nav_penalties is not None else np.zeros(n)
    # Build bounds: partial sells (floor at 25% of current), zero-cap respected, 20% per-holding cap
    if weight_caps is None:
        weight_caps = [0.20] * n
    bounds = []
    for i in range(n):
        cap = weight_caps[i]
        if cap <= 0:
            bounds.append((0.0, 0.0))
        else:
            floor = cw[i] * 0.25 if cw[i] > 0.005 else 0.0
            bounds.append((floor, min(cap, 0.20)))
    n_current = int(np.sum(cw > 0.005))
    # Precompute returns matrix for income score
    ret_vals = returns_df.values
    ret_mu = mu  # already computed above
    # Income-focused blend: 70% yield, 30% income quality (Sortino/Omega/Calmar/Ulcer)
    import math
    blend = 0.7
    def neg_objective(w):
        port_yield = float(yields_arr @ w)
        # Log-scale: always rewards more yield but with diminishing returns
        # At 15%→0.69, 30%→1.10, 50%→1.50 — prevents extreme yields from dominating
        norm_yield = math.log(1.0 + port_yield / target_yield) if port_yield > 0 else 0.0
        quality = _portfolio_income_score(w, ret_vals, ret_mu)
        nav_cost = float(w.dot(nav_pen))
        turnover = float(np.sum(np.abs(w - cw)))
        n_active = float(np.sum(w > 0.005))
        diversification_loss = max(0, n_current - n_active) * 0.03
        hhi = float((w ** 2).sum())
        concentration_penalty = hhi * 0.15
        return -(blend * norm_yield + (1.0 - blend) * quality - turnover_penalty * turnover - diversification_loss - nav_cost - concentration_penalty)
    constraints = [{"type": "eq", "fun": lambda w: w.sum() - 1}]
    init = _clamp_and_normalize(cw, bounds)
    best = None
    starts = [init]
    rng = np.random.default_rng(42)
    for _ in range(4):
        d = rng.dirichlet(np.ones(n))
        starts.append(_clamp_and_normalize(d, bounds))
    for w0 in starts:
        try:
            res = minimize(neg_objective, w0, method="SLSQP", bounds=bounds,
                           constraints=constraints, options={"maxiter": 500, "ftol": 1e-8})
            if res.success and (best is None or res.fun < best.fun):
                best = res
        except Exception:
            pass
    if best is None:
        return init
    return _clamp_and_normalize(best.x, bounds)


def _optimize_balanced(returns_df, yields, balance=0.5, current_weights=None, min_sharpe=0.8, max_dd=-0.20, weight_caps=None, turnover_penalty=0.3, nav_penalties=None):
    """Balanced optimization: slider controls yield vs safety blend.
    balance=1.0 → pure yield focus.
    balance=0.0 → pure safety focus (income quality: Sortino/Omega/Calmar/Ulcer).
    balance=0.5 → equal blend (default).
    Enforces quality floor penalty scaled by (1-balance), heavier turnover penalty.
    nav_penalties: per-ticker penalty for NAV erosion (higher = worse, 0 = no erosion).
    """
    import numpy as np
    from scipy.optimize import minimize
    n = returns_df.shape[1]
    yields_arr = np.array(yields)
    target_yield = 0.15  # normalize against 15% target, not max ticker yield
    if current_weights is None:
        current_weights = np.ones(n) / n
    cw = np.array(current_weights)
    mu = returns_df.mean().values * 252
    ret_vals = returns_df.values
    nav_pen = np.array(nav_penalties) if nav_penalties is not None else np.zeros(n)
    import math
    # Dynamic per-holding cap: 10% at safety end (balance=0), 20% at income end (balance=1)
    max_per_holding = 0.10 + 0.10 * balance
    if weight_caps is None:
        weight_caps = [max_per_holding] * n
    bounds = []
    for i in range(n):
        cap = weight_caps[i]
        if cap <= 0:
            bounds.append((0.0, 0.0))
        else:
            floor = cw[i] * 0.25 if cw[i] > 0.005 else 0.0
            bounds.append((floor, min(cap, max_per_holding)))
    n_current = int(np.sum(cw > 0.005))
    def neg_objective(w):
        port_yield = float(w.dot(yields_arr))
        # Log-scale: always rewards more yield but with diminishing returns
        norm_yield = math.log(1.0 + port_yield / target_yield) if port_yield > 0 else 0.0
        quality = _portfolio_income_score(w, ret_vals, mu)
        # NAV erosion penalty: scaled by balance (softer when income-focused)
        nav_scale = 1.0 - 0.5 * balance
        nav_cost = float(w.dot(nav_pen)) * nav_scale
        # Slider blend: balance toward yield, (1-balance) toward income quality
        blended = balance * norm_yield + (1.0 - balance) * quality
        # Safety penalty: quality floor enforced, scaled by safety focus
        safety_penalty = max(0, 0.4 - quality) * 2.0 * (1.0 - balance)
        turnover = float(np.sum(np.abs(w - cw)))
        n_active = float(np.sum(w > 0.005))
        diversification_loss = max(0, n_current - n_active) * 0.03
        hhi = float((w ** 2).sum())
        concentration_penalty = hhi * 0.15
        return -(blended - turnover_penalty * turnover - safety_penalty - diversification_loss - nav_cost - concentration_penalty)
    constraints = [{"type": "eq", "fun": lambda w: w.sum() - 1}]
    init = _clamp_and_normalize(cw, bounds)
    best = None
    starts = [init]
    rng = np.random.default_rng(42)
    for _ in range(3):
        d = rng.dirichlet(np.ones(n))
        starts.append(_clamp_and_normalize(d, bounds))
    for w0 in starts:
        try:
            res = minimize(neg_objective, w0, method="SLSQP", bounds=bounds,
                           constraints=constraints, options={"maxiter": 500, "ftol": 1e-8})
            if res.success and (best is None or res.fun < best.fun):
                best = res
        except Exception:
            pass
    if best is None:
        return init
    return _clamp_and_normalize(best.x, bounds)


def _before_after_comparison(returns_df, opt_weights, bench_ret,
                              port_metrics_before, curr_income, opt_income,
                              current_weights=None, coverage_map=None,
                              available_tickers=None):
    """Compute before/after grade, income, coverage, and key metrics.
    Uses already-computed port_metrics and income values from the optimization branch
    so numbers match exactly what's shown in the optimization summary.
    If weights are essentially unchanged, reuses before metrics to avoid grading noise."""
    import numpy as np
    from grading import grade_portfolio
    pm = port_metrics_before or {}
    grade_before = pm.get("grade", {})

    # Compute weighted coverage for current and optimal weights
    def _weighted_coverage(weights_arr):
        if coverage_map is None or available_tickers is None:
            return None
        total_w = 0.0
        weighted_sum = 0.0
        for i, t in enumerate(available_tickers):
            cov = coverage_map.get(t)
            w = float(weights_arr[i]) if i < len(weights_arr) else 0
            if cov is not None and w > 0.001:
                weighted_sum += cov * w
                total_w += w
        return round(weighted_sum / total_w, 4) if total_w > 0 else None

    curr_cov = _weighted_coverage(current_weights) if current_weights is not None else None
    opt_cov = _weighted_coverage(opt_weights)

    before = {
        "grade": grade_before.get("overall", "—") if isinstance(grade_before, dict) else str(grade_before),
        "score": grade_before.get("score", 0) if isinstance(grade_before, dict) else 0,
        "sharpe": pm.get("sharpe"),
        "sortino": pm.get("sortino"),
        "omega": pm.get("omega"),
        "calmar": pm.get("calmar"),
        "ulcer_index": pm.get("ulcer_index"),
        "max_drawdown": pm.get("max_drawdown"),
        "annual_income": round(curr_income, 2),
        "monthly_income": round(curr_income / 12, 2),
        "coverage": curr_cov,
    }
    # If weights barely changed (all within 0.5%), reuse before metrics — no real change
    if current_weights is not None:
        max_change = float(np.max(np.abs(np.array(opt_weights) - np.array(current_weights)))) * 100
        if max_change < 0.5:
            return {"before": before, "after": dict(before)}
    pm_after = grade_portfolio(returns_df, opt_weights, bench_ret)
    return {
        "before": before,
        "after": {
            "grade": pm_after.get("grade", {}).get("overall", "—"),
            "score": pm_after.get("grade", {}).get("score", 0),
            "sharpe": pm_after.get("sharpe"),
            "sortino": pm_after.get("sortino"),
            "omega": pm_after.get("omega"),
            "calmar": pm_after.get("calmar"),
            "ulcer_index": pm_after.get("ulcer_index"),
            "max_drawdown": pm_after.get("max_drawdown"),
            "annual_income": round(opt_income, 2),
            "monthly_income": round(opt_income / 12, 2),
            "coverage": opt_cov,
        },
    }


def _enrich_weights_with_actions(weights_out, close_df, total_val, nav_returns=None, threshold=0.5):
    """Add action / dollar_change / shares_change / current_price / nav_change_pct to each weight dict."""
    total_buy = 0.0
    total_sell = 0.0
    for w in weights_out:
        ticker = w["ticker"]
        change_pct = w["optimal_pct"] - w["current_pct"]
        dollar_change = round((change_pct / 100) * total_val, 2)
        price = 0.0
        if ticker in close_df.columns:
            s = close_df[ticker].dropna()
            if len(s) > 0:
                price = round(float(s.iloc[-1]), 2)
        shares = int(dollar_change / price) if price > 0 else 0
        if nav_returns:
            w["nav_change_pct"] = round(nav_returns.get(ticker, 0.0) * 100, 1)
        if abs(change_pct) < threshold:
            action = "HOLD"
        elif change_pct > 0:
            action = "BUY"
        else:
            action = "SELL"
        w["action"] = action
        w["dollar_change"] = dollar_change
        w["shares_change"] = shares
        w["current_price"] = price
        if action == "BUY":
            total_buy += dollar_change
        elif action == "SELL":
            total_sell += abs(dollar_change)
    summary = {
        "total_buy": round(total_buy, 2),
        "total_sell": round(total_sell, 2),
        "num_buys": sum(1 for w in weights_out if w["action"] == "BUY"),
        "num_sells": sum(1 for w in weights_out if w["action"] == "SELL"),
        "num_holds": sum(1 for w in weights_out if w["action"] == "HOLD"),
    }
    return weights_out, summary


def _build_efficient_frontier(returns_df, n_points=30):
    import numpy as np
    from scipy.optimize import minimize
    n = returns_df.shape[1]
    mean_ret = returns_df.mean().values * 252
    cov = returns_df.cov().values * 252
    ret_range = np.linspace(mean_ret.min(), mean_ret.max(), n_points)
    frontier = []
    for target in ret_range:
        def portfolio_vol(w):
            return np.sqrt(w @ cov @ w)
        constraints = [
            {"type": "eq", "fun": lambda w: w.sum() - 1},
            {"type": "eq", "fun": lambda w, t=target: w @ mean_ret - t},
        ]
        bounds = [(0, 1)] * n
        init = np.ones(n) / n
        result = minimize(portfolio_vol, init, method="SLSQP", bounds=bounds,
                          constraints=constraints, options={"maxiter": 500})
        if result.success:
            vol = float(np.sqrt(result.x @ cov @ result.x))
            ret = float(result.x @ mean_ret)
            frontier.append({"vol": round(vol * 100, 2), "ret": round(ret * 100, 2)})
    return frontier


GROWTH_ETFS = ["QQQ", "VOO", "VGT", "VUG", "SCHG", "IWF", "MGK", "QQQM", "XLK", "SMH"]
INCOME_ETFS = ["SPYI", "QQQI", "MLPI", "TSPY", "TDAQ", "IWMI", "JEPQ", "CHPY", "ULTY", "O", "MAIN", "PBDC", "XQQI"]

# Single-stock ETFs — highest risk, destroy portfolio scores.
# Only recommended as BUY in optimize_income or balanced at 100%.
_DEFAULT_SINGLE_STOCK_ETFS = {
    # YieldMax single-stock option-income ETFs
    "TSLY", "NVDY", "CONY", "AMZY", "MSFO", "APLY", "GOOY", "FBY",
    "NFLY", "DISO", "PYPY", "SQY", "AMDY", "OARK", "MRNY", "SNOY",
    "JPMO", "BABO", "TSMY", "YBIT", "PLTY", "FIAT", "GMEY", "MARO",
    "CRSH", "ABNY", "SMCY", "MSTY", "METY", "FIVY", "LFGY",
    # YieldMax multi-single-stock baskets (composed entirely of single-stock ETFs)
    "ULTY", "YMAG", "YMAX",
    # Kurv single-stock
    "TSLP", "NVDP",
    # GraniteShares / REX single-stock leveraged
    "NVDL", "TSLL", "CONL", "AMDL",
}


def _get_single_stock_etfs():
    """Return built-in set merged with any user-added tickers from settings."""
    result = set(_DEFAULT_SINGLE_STOCK_ETFS)
    try:
        conn = get_connection()
        row = conn.execute("SELECT value FROM settings WHERE key = 'single_stock_etfs'").fetchone()
        conn.close()
        if row and row["value"]:
            user_tickers = {t.strip().upper() for t in row["value"].split(",") if t.strip()}
            result |= user_tickers
    except Exception:
        pass
    return result


@app.route("/api/analytics/data", methods=["POST"])
def analytics_data():
    """Compute risk metrics, portfolio grade, charts and optional optimization."""
    import warnings, math
    import numpy as np
    import yfinance as yf
    from grading import (ticker_score, grade_portfolio, letter_grade,
                         _sharpe, _sortino, _calmar, _omega,
                         _ulcer_index, _max_drawdown, _capture_ratios, _safe)
    warnings.filterwarnings("ignore")

    data = request.get_json(force=True, silent=True) or {}
    tickers = [str(t).strip().upper() for t in data.get("tickers", []) if str(t).strip()]
    benchmark = str(data.get("benchmark", "SPY")).strip().upper()
    period = data.get("period", "1y")
    mode = data.get("mode", "metrics")
    min_sharpe = float(data.get("min_sharpe", 0.8))
    max_dd = float(data.get("max_dd", -0.20))
    balance = float(data.get("balance", 0.5))

    if not tickers:
        return jsonify(error="No tickers provided."), 400

    valid_periods = {"1mo", "3mo", "6mo", "ytd", "1y", "2y", "5y", "max"}
    if period not in valid_periods:
        period = "1y"

    all_dl = list(set(tickers + [benchmark]))
    try:
        raw = yf.download(" ".join(all_dl), period=period, auto_adjust=True, progress=False)
        if raw.empty:
            return jsonify(error="No price data returned."), 500
    except Exception as e:
        return jsonify(error=f"yfinance error: {str(e)}"), 500

    if isinstance(raw.columns, pd.MultiIndex):
        close = raw["Close"].dropna(how="all")
    else:
        close = raw[["Close"]].dropna(how="all")
        close.columns = [all_dl[0]]

    # DB weights and yields
    profile_id = get_profile_id()
    conn = get_connection()
    try:
        db_rows = conn.execute(
            "SELECT ticker, current_value, estim_payment_per_year FROM all_account_info "
            "WHERE current_value IS NOT NULL AND current_value > 0 AND profile_id = ?",
            (profile_id,)
        ).fetchall()
    except Exception:
        db_rows = []
    conn.close()

    db_df = pd.DataFrame([dict(r) for r in db_rows]) if db_rows else pd.DataFrame(columns=["ticker", "current_value", "estim_payment_per_year"])
    total_val = float(db_df["current_value"].sum()) if not db_df.empty else 1
    weight_map = {}
    yield_map = {}
    income_map = {}
    for _, r in db_df.iterrows():
        t = r["ticker"]
        weight_map[t] = float(r["current_value"]) / total_val if total_val > 0 else 0
        cv = float(r["current_value"]) if r["current_value"] and float(r["current_value"]) > 0 else 1
        ep = float(r["estim_payment_per_year"]) if pd.notna(r.get("estim_payment_per_year")) else 0
        yield_map[t] = ep / cv
        income_map[t] = ep

    # Per-ticker coverage ratio: (TTM price return % + TTM dist yield %) / TTM dist yield %
    coverage_map = {}
    for t in tickers:
        if t not in close.columns:
            continue
        tc = close[t].dropna()
        if len(tc) < 2:
            continue
        price_start = float(tc.iloc[0])
        price_end = float(tc.iloc[-1])
        if price_start <= 0:
            continue
        price_return_pct = (price_end - price_start) / price_start
        dist_yield = yield_map.get(t, 0)
        if dist_yield > 0:
            coverage_map[t] = round((price_return_pct + dist_yield) / dist_yield, 4)

    bench_close = close[benchmark] if benchmark in close.columns else None
    bench_ret = bench_close.pct_change().dropna() if bench_close is not None else None

    def safe(v):
        if v is None:
            return None
        if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
            return None
        return v

    # Per-ticker metrics
    metrics = []
    available_tickers = []
    for t in tickers:
        if t not in close.columns:
            continue
        tc = close[t].dropna()
        if len(tc) < 30:
            continue
        available_tickers.append(t)
        tr = tc.pct_change().dropna()
        score, sharpe_v, sortino_v, calmar_v, omega_v, mdd_v, dc_v, ulcer_v = ticker_score(tc, tr, bench_ret)
        uc_v, _ = _capture_ratios(tr, bench_ret) if bench_ret is not None else (None, None)
        annual_ret = round(float(tr.mean() * 252) * 100, 2)
        annual_total_ret = round(annual_ret + yield_map.get(t, 0) * 100, 2)
        annual_vol = round(float(tr.std() * np.sqrt(252)) * 100, 2)
        metrics.append({
            "ticker": t,
            "weight": round(weight_map.get(t, 0) * 100, 2),
            "sharpe": safe(sharpe_v),
            "sortino": safe(sortino_v),
            "calmar": safe(calmar_v),
            "omega": safe(omega_v),
            "ulcer_index": safe(ulcer_v),
            "up_capture": safe(uc_v),
            "down_capture": safe(dc_v),
            "max_drawdown": round(mdd_v * 100, 2) if mdd_v is not None else None,
            "annual_ret": annual_ret,
            "annual_total_ret": annual_total_ret,
            "annual_vol": annual_vol,
            "score": round(score, 1),
            "grade": letter_grade(score),
            "annual_income": round(income_map.get(t, 0), 2),
        })

    # Portfolio-level metrics and grade
    port_metrics = {}
    result_corr = None
    result_dd = None
    if len(available_tickers) >= 2:
        returns_df = close[available_tickers].pct_change().fillna(0)
        weights_arr = np.array([weight_map.get(t, 0) for t in available_tickers])
        w_sum = weights_arr.sum()
        if w_sum > 0:
            weights_arr = weights_arr / w_sum
        else:
            # No DB weights available — use equal weight
            weights_arr = np.ones(len(available_tickers)) / len(available_tickers)

        pm = grade_portfolio(returns_df, weights_arr, bench_ret)
        port_metrics = {
            "sharpe": pm.get("sharpe"),
            "sortino": pm.get("sortino"),
            "calmar": pm.get("calmar"),
            "omega": pm.get("omega"),
            "ulcer_index": pm.get("ulcer_index"),
            "max_drawdown": pm.get("max_drawdown"),
            "up_capture": pm.get("up_capture"),
            "down_capture": pm.get("down_capture"),
            "top_weight": pm.get("top_weight"),
            "effective_n": pm.get("effective_n"),
            "n_holdings": len(available_tickers),
            "grade": pm.get("grade"),
            "total_value": round(total_val, 2),
            "est_annual_income": round(sum(income_map.values()), 2),
        }

        # Correlation matrix
        corr = returns_df.corr()
        result_corr = {
            "labels": available_tickers,
            "matrix": [[round(float(corr.loc[a, b]), 3) for b in available_tickers]
                       for a in available_tickers],
        }

        # Portfolio drawdown series
        port_daily = returns_df.dot(weights_arr)
        port_cum = (1 + port_daily).cumprod()
        roll_max = port_cum.cummax()
        drawdown_s = (port_cum / roll_max - 1)
        step = max(1, len(drawdown_s) // 200)
        dd_sampled = drawdown_s.iloc[::step]
        result_dd = {
            "dates": [d.strftime("%Y-%m-%d") for d in dd_sampled.index],
            "values": [round(float(v) * 100, 2) for v in dd_sampled.values],
        }

    # Include per-ticker coverage and aggregate in response
    cov_results = [{"ticker": t, "coverage_ratio": coverage_map.get(t)} for t in available_tickers]
    # Dollar-weighted aggregate coverage
    _cov_num = 0.0
    _cov_den = 0.0
    for t in available_tickers:
        cov = coverage_map.get(t)
        w = weight_map.get(t, 0)
        if cov is not None and w > 0:
            _cov_num += cov * w
            _cov_den += w
    agg_cov = round(_cov_num / _cov_den, 4) if _cov_den > 0 else None

    result = {"metrics": metrics, "portfolio_metrics": port_metrics,
              "coverage": {"results": cov_results, "aggregate_coverage": agg_cov}}
    if result_corr:
        result["correlation"] = result_corr
        result["drawdown_series"] = result_dd

    # Sector breakdown (parallel yf.Ticker.info fetch)
    try:
        from concurrent.futures import ThreadPoolExecutor
        def _fetch_sector(t):
            try:
                info = yf.Ticker(t).info
                return t, info.get("sector", info.get("category", "Other")), info.get("quoteType", "Unknown")
            except Exception:
                return t, "Unknown", "Unknown"
        with ThreadPoolExecutor(max_workers=8) as pool:
            sector_info = dict()
            type_info = dict()
            for t, sector, qtype in pool.map(lambda t: _fetch_sector(t), available_tickers):
                val = weight_map.get(t, 0) * 100 if weight_map else 100 / max(len(available_tickers), 1)
                sector_info[sector or "Unknown"] = sector_info.get(sector or "Unknown", 0) + val
                type_info[qtype or "Unknown"] = type_info.get(qtype or "Unknown", 0) + val
        result["sector_breakdown"] = {
            "by_sector": [{"label": k, "value": round(v, 2)} for k, v in sorted(sector_info.items(), key=lambda x: -x[1])],
            "by_type": [{"label": k, "value": round(v, 2)} for k, v in sorted(type_info.items(), key=lambda x: -x[1])],
        }
    except Exception:
        pass

    # Risk contribution breakdown
    if result_corr and weight_map and len(available_tickers) >= 2:
        try:
            w_arr = np.array([weight_map.get(t, 0) for t in available_tickers])
            w_sum = w_arr.sum()
            if w_sum > 0:
                w_arr = w_arr / w_sum
                cov_mat = returns_df.cov().values * 252
                port_vol = float(np.sqrt(w_arr @ cov_mat @ w_arr))
                if port_vol > 0:
                    mctr = cov_mat @ w_arr / port_vol
                    risk_contrib = w_arr * mctr
                    rc_sum = risk_contrib.sum()
                    if rc_sum > 0:
                        pct_contrib = risk_contrib / rc_sum * 100
                        result["risk_contribution"] = [{"ticker": t, "pct": round(float(p), 2)} for t, p in zip(available_tickers, pct_contrib)]
        except Exception:
            pass

    # Detect portfolio type and suggest complementary ETFs
    if weight_map:
        weighted_yield = sum(yield_map.get(t, 0) * weight_map.get(t, 0) for t in weight_map)
        tickers_upper = [t.upper() for t in tickers]
        if weighted_yield > 0.05:
            result["portfolio_type"] = "income"
            result["suggested_growth"] = [t for t in GROWTH_ETFS if t not in tickers_upper]
        elif weighted_yield < 0.02:
            result["portfolio_type"] = "growth"
            result["suggested_income"] = [t for t in INCOME_ETFS if t not in tickers_upper]
        else:
            result["portfolio_type"] = "mixed"

    # Optimization modes
    if mode in ("optimize_returns", "optimize_income", "optimize_balanced") and len(available_tickers) >= 2:
        returns_df = close[available_tickers].pct_change().dropna()
        current_weights = np.array([weight_map.get(t, 0) for t in available_tickers])
        cw_sum = current_weights.sum()
        has_portfolio = cw_sum > 0  # True when user loaded real portfolio data
        if has_portfolio:
            current_weights = current_weights / cw_sum
        else:
            current_weights = np.ones(len(available_tickers)) / len(available_tickers)

        # Compute NAV erosion for yield-based optimizers
        nav_returns = _compute_nav_erosion(close, available_tickers)
        single_stock_etfs = _get_single_stock_etfs()

        if mode == "optimize_returns":
            ret_weight_caps = _nav_weight_caps(nav_returns, available_tickers)
            # Force sell single-stock ETFs — too risky for growth optimization
            for i, t in enumerate(available_tickers):
                if t in single_stock_etfs:
                    ret_weight_caps[i] = 0.0
            opt_w = _optimize_sharpe(returns_df, current_weights=current_weights, weight_caps=ret_weight_caps)
            frontier = _build_efficient_frontier(returns_df)
            opt_sharpe_val = _portfolio_sharpe(opt_w, returns_df)
            opt_sortino_val = _portfolio_sortino(opt_w, returns_df)
            curr_sharpe_val = _portfolio_sharpe(current_weights, returns_df)
            curr_sortino_val = _portfolio_sortino(current_weights, returns_df)
            mean_ret = float(returns_df.mean().values @ opt_w) * 252
            cov = returns_df.cov().values * 252
            opt_vol = float(np.sqrt(opt_w @ cov @ opt_w))
            curr_ret = float(returns_df.mean().values @ current_weights) * 252
            curr_vol = float(np.sqrt(current_weights @ cov @ current_weights))
            weights_out = [{"ticker": t, "current_pct": round(current_weights[i] * 100, 2),
                            "optimal_pct": round(opt_w[i] * 100, 2)} for i, t in enumerate(available_tickers)]
            weights_out, rebal_summary = _enrich_weights_with_actions(weights_out, close, total_val, nav_returns)
            opt_dict = {
                "weights": weights_out, "frontier": frontier,
                "optimal_point": {"vol": round(opt_vol * 100, 2), "ret": round(mean_ret * 100, 2)},
                "current_point": {"vol": round(curr_vol * 100, 2), "ret": round(curr_ret * 100, 2)},
                "summary": {"sharpe": round(opt_sharpe_val, 2), "sortino": round(opt_sortino_val, 2),
                             "curr_sharpe": round(curr_sharpe_val, 2), "curr_sortino": round(curr_sortino_val, 2),
                             "expected_return": round(mean_ret * 100, 2),
                             "expected_vol": round(opt_vol * 100, 2)},
                "rebalance_summary": rebal_summary,
            }
            if has_portfolio:
                yields_list_ret = [yield_map.get(t, 0) for t in available_tickers]
                ret_curr_income = float(current_weights.dot(np.array(yields_list_ret))) * total_val
                ret_opt_income = float(opt_w.dot(np.array(yields_list_ret))) * total_val
                opt_dict["comparison"] = _before_after_comparison(returns_df, opt_w, bench_ret,
                                                                   port_metrics, ret_curr_income, ret_opt_income,
                                                                   current_weights=current_weights,
                                                                   coverage_map=coverage_map,
                                                                   available_tickers=available_tickers)
            result["optimization"] = opt_dict

        elif mode == "optimize_income":
            yields_list = [yield_map.get(t, 0) for t in available_tickers]
            adj_yields = _adjust_yields_for_nav(yields_list, nav_returns, available_tickers)
            weight_caps = _nav_weight_caps(nav_returns, available_tickers)
            # Per-ticker NAV erosion penalty: abs(erosion) scaled, 0 for non-eroding
            nav_pen = [max(0, -nav_returns.get(t, 0.0)) * 1.5 for t in available_tickers]
            # Zero-yield tickers get capped to 0 in income modes — no reason to buy them
            for i, t in enumerate(available_tickers):
                if yields_list[i] <= 0:
                    weight_caps[i] = 0.0
            opt_w = _optimize_income(returns_df, adj_yields, current_weights=current_weights, weight_caps=weight_caps, nav_penalties=nav_pen)
            opt_yield = float(opt_w.dot(np.array(yields_list)))
            curr_yield = float(current_weights.dot(np.array(yields_list)))
            # Use port_metrics for current (matches Impact Analysis "before"), grade_portfolio for optimized
            opt_gp = grade_portfolio(returns_df, opt_w, bench_ret)
            opt_income = opt_yield * total_val
            curr_income = curr_yield * total_val
            weights_out = [{"ticker": t, "current_pct": round(current_weights[i] * 100, 2),
                            "optimal_pct": round(opt_w[i] * 100, 2), "yield_pct": round(yields_list[i] * 100, 2)}
                           for i, t in enumerate(available_tickers)]
            weights_out, rebal_summary = _enrich_weights_with_actions(weights_out, close, total_val, nav_returns)
            scatter_data = []
            for i, t in enumerate(available_tickers):
                tc = close[t].dropna()
                tr = tc.pct_change().dropna()
                vol = float(tr.std() * np.sqrt(252)) if len(tr) > 1 else 0
                scatter_data.append({"ticker": t, "yield_pct": round(yields_list[i] * 100, 2),
                                     "vol_pct": round(vol * 100, 2), "is_optimal": float(opt_w[i]) > 0.01})
            opt_dict = {
                "weights": weights_out, "scatter": scatter_data,
                "summary": {"opt_sortino": round(safe(opt_gp.get("sortino")) or 0, 2),
                             "curr_sortino": round(safe(port_metrics.get("sortino")) or 0, 2),
                             "opt_omega": round(safe(opt_gp.get("omega")) or 0, 2),
                             "curr_omega": round(safe(port_metrics.get("omega")) or 0, 2),
                             "opt_calmar": round(safe(opt_gp.get("calmar")) or 0, 2),
                             "curr_calmar": round(safe(port_metrics.get("calmar")) or 0, 2),
                             "opt_ulcer": round(safe(opt_gp.get("ulcer_index")) or 0, 2),
                             "curr_ulcer": round(safe(port_metrics.get("ulcer_index")) or 0, 2),
                             "max_dd": round((safe(opt_gp.get("max_drawdown")) or 0) * 100, 2),
                             "opt_yield": round(opt_yield * 100, 2), "curr_yield": round(curr_yield * 100, 2),
                             "opt_income": round(opt_income, 2), "curr_income": round(curr_income, 2)},
                "rebalance_summary": rebal_summary,
            }
            if has_portfolio:
                opt_dict["comparison"] = _before_after_comparison(returns_df, opt_w, bench_ret,
                                                                   port_metrics, curr_income, opt_income,
                                                                   current_weights=current_weights,
                                                                   coverage_map=coverage_map,
                                                                   available_tickers=available_tickers)
            result["optimization"] = opt_dict

        elif mode == "optimize_balanced":
            yields_list = [yield_map.get(t, 0) for t in available_tickers]
            adj_yields = _adjust_yields_for_nav(yields_list, nav_returns, available_tickers)
            weight_caps = _nav_weight_caps(nav_returns, available_tickers)
            # Zero-yield tickers get capped to 0 in income-oriented modes
            for i, t in enumerate(available_tickers):
                if yields_list[i] <= 0:
                    weight_caps[i] = 0.0
            # Single-stock ETFs: tiered handling based on balance slider and yield
            # <55%: scale toward sell (cap from 0 at balance=0 to current weight at 55%)
            # 55-70%: can buy but reduced priority (cap at half of max_per_holding)
            # >=70%: full priority ONLY if high yield (>=10%), otherwise still reduced
            high_yield_threshold = 0.10
            max_per_holding = 0.10 + 0.10 * balance  # match the cap used inside _optimize_balanced
            if balance < 0.55:
                for i, t in enumerate(available_tickers):
                    if t in single_stock_etfs:
                        ss_cap = current_weights[i] * (balance / 0.55)
                        weight_caps[i] = min(weight_caps[i], ss_cap)
            elif balance < 0.70:
                half_cap = max_per_holding * 0.5
                for i, t in enumerate(available_tickers):
                    if t in single_stock_etfs:
                        weight_caps[i] = min(weight_caps[i], half_cap)
            else:
                # >=70%: only high-yield single-stock ETFs get full priority
                half_cap = max_per_holding * 0.5
                for i, t in enumerate(available_tickers):
                    if t in single_stock_etfs and yield_map.get(t, 0) < high_yield_threshold:
                        weight_caps[i] = min(weight_caps[i], half_cap)
            nav_pen = [max(0, -nav_returns.get(t, 0.0)) * 1.5 for t in available_tickers]
            opt_w = _optimize_balanced(returns_df, adj_yields, balance=balance, current_weights=current_weights, min_sharpe=min_sharpe, max_dd=max_dd, weight_caps=weight_caps, nav_penalties=nav_pen)
            opt_yield = float(opt_w.dot(np.array(yields_list)))
            curr_yield = float(current_weights.dot(np.array(yields_list)))
            # Use port_metrics for current (matches Impact Analysis "before"), grade_portfolio for optimized
            opt_gp = grade_portfolio(returns_df, opt_w, bench_ret)
            opt_income = opt_yield * total_val
            curr_income = curr_yield * total_val
            weights_out = [{"ticker": t, "current_pct": round(current_weights[i] * 100, 2),
                            "optimal_pct": round(opt_w[i] * 100, 2), "yield_pct": round(yields_list[i] * 100, 2)}
                           for i, t in enumerate(available_tickers)]
            scatter_data = []
            for i, t in enumerate(available_tickers):
                tc = close[t].dropna()
                tr = tc.pct_change().dropna()
                vol = float(tr.std() * np.sqrt(252)) if len(tr) > 1 else 0
                sharpe_t = safe(_sharpe(tc))
                scatter_data.append({"ticker": t, "yield_pct": round(yields_list[i] * 100, 2),
                                     "vol_pct": round(vol * 100, 2), "sharpe": sharpe_t if sharpe_t else 0,
                                     "is_optimal": float(opt_w[i]) > 0.01})
            weights_out, rebal_summary = _enrich_weights_with_actions(weights_out, close, total_val, nav_returns)
            opt_dict = {
                "weights": weights_out, "scatter": scatter_data,
                "summary": {"opt_yield": round(opt_yield * 100, 2), "curr_yield": round(curr_yield * 100, 2),
                             "opt_income": round(opt_income, 2), "curr_income": round(curr_income, 2),
                             "opt_sortino": round(safe(opt_gp.get("sortino")) or 0, 2),
                             "curr_sortino": round(safe(port_metrics.get("sortino")) or 0, 2),
                             "opt_omega": round(safe(opt_gp.get("omega")) or 0, 2),
                             "curr_omega": round(safe(port_metrics.get("omega")) or 0, 2),
                             "opt_calmar": round(safe(opt_gp.get("calmar")) or 0, 2),
                             "curr_calmar": round(safe(port_metrics.get("calmar")) or 0, 2),
                             "opt_ulcer": round(safe(opt_gp.get("ulcer_index")) or 0, 2),
                             "curr_ulcer": round(safe(port_metrics.get("ulcer_index")) or 0, 2),
                             "max_dd": round((safe(opt_gp.get("max_drawdown")) or 0) * 100, 2),
                             "balance": round(balance * 100)},
                "rebalance_summary": rebal_summary,
            }
            if has_portfolio:
                opt_dict["comparison"] = _before_after_comparison(returns_df, opt_w, bench_ret,
                                                                   port_metrics, curr_income, opt_income,
                                                                   current_weights=current_weights,
                                                                   coverage_map=coverage_map,
                                                                   available_tickers=available_tickers)
            result["optimization"] = opt_dict

    return jsonify(result)


# ── Correlation Matrix ─────────────────────────────────────────────────────────

@app.route("/api/correlation/data", methods=["POST"])
def correlation_data():
    """Compute correlation matrix for arbitrary tickers over a given period."""
    import math
    import warnings
    import numpy as np
    import yfinance as yf
    warnings.filterwarnings("ignore")

    data = request.get_json(force=True, silent=True) or {}
    tickers = [str(t).strip().upper() for t in data.get("tickers", []) if str(t).strip()]
    period = data.get("period", "1y")

    if len(tickers) < 2:
        return jsonify(error="Please enter at least 2 tickers.")
    if len(tickers) > 50:
        return jsonify(error="Maximum 50 tickers allowed.")

    valid_periods = {"3mo", "6mo", "1y", "2y", "5y", "max"}
    if period not in valid_periods:
        period = "1y"

    unique = list(dict.fromkeys(tickers))

    try:
        raw = yf.download(" ".join(unique), period=period, auto_adjust=True, progress=False)
        if raw.empty:
            return jsonify(error="No price data returned from Yahoo Finance.")
    except Exception as e:
        return jsonify(error=f"Failed to fetch data: {str(e)}")

    if isinstance(raw.columns, pd.MultiIndex):
        close = raw["Close"].dropna(how="all")
    else:
        close = raw[["Close"]].dropna(how="all")
        close.columns = [unique[0]]

    available = [t for t in unique if t in close.columns and close[t].dropna().count() >= 30]
    missing = [t for t in unique if t not in available]

    if len(available) < 2:
        return jsonify(error="Need at least 2 tickers with sufficient data.")

    daily_returns = close[available].pct_change().dropna()
    corr = daily_returns.corr()

    def _safe(v):
        if v is None:
            return None
        try:
            f = float(v)
            return None if (math.isnan(f) or math.isinf(f)) else round(f, 3)
        except (TypeError, ValueError):
            return None

    matrix = [[_safe(corr.iloc[i, j]) for j in range(len(available))] for i in range(len(available))]

    return jsonify(
        tickers=available,
        matrix=matrix,
        missing=missing,
        period=period,
        data_points=len(daily_returns),
    )


# ── Analytics: Income Calendar ─────────────────────────────────────────────────

@app.route("/api/analytics/income-calendar", methods=["POST"])
def analytics_income_calendar():
    data = request.get_json() or {}
    req_tickers = [t.upper() for t in data.get("tickers", [])]
    if not req_tickers:
        return jsonify({"months": [], "tickers": [], "monthly_totals": []})

    _, cal_pids = get_profile_filter()
    cal_ph = ",".join("?" * len(cal_pids))
    conn = get_connection()
    months_labels = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

    ticker_data = []
    for t in req_tickers:
        # Get pay months from monthly_payout_tickers
        row = conn.execute("SELECT pay_month FROM monthly_payout_tickers WHERE ticker = ?", (t,)).fetchone()
        div_row = conn.execute(
            f"SELECT estim_payment_per_year, div_frequency FROM dividends WHERE ticker = ? AND profile_id IN ({cal_ph})",
            [t] + cal_pids
        ).fetchone()
        if not div_row:
            continue
        annual = float(div_row["estim_payment_per_year"] or 0)
        freq = (div_row["div_frequency"] or "").strip().upper()

        amounts = [0.0] * 12
        if freq == "W":
            # Weekly: spread evenly
            weekly = annual / 52
            amounts = [round(weekly * 4.33, 2)] * 12
        elif freq == "M":
            monthly = annual / 12
            amounts = [round(monthly, 2)] * 12
        elif row and row["pay_month"]:
            # Quarterly/Semi-annual/Annual — use pay_month data
            pay_months = [int(m.strip()) for m in str(row["pay_month"]).split(",") if m.strip().isdigit()]
            if pay_months:
                per_payment = annual / len(pay_months)
                for m in pay_months:
                    if 1 <= m <= 12:
                        amounts[m - 1] = round(per_payment, 2)
        else:
            # Default: spread quarterly starting month 3
            if freq == "Q":
                for m in [2, 5, 8, 11]:
                    amounts[m] = round(annual / 4, 2)
            elif freq in ("SA", "S"):
                for m in [5, 11]:
                    amounts[m] = round(annual / 2, 2)
            elif freq == "A":
                amounts[11] = round(annual, 2)
            else:
                monthly = annual / 12
                amounts = [round(monthly, 2)] * 12

        ticker_data.append({"ticker": t, "amounts": amounts})

    conn.close()

    monthly_totals = [0.0] * 12
    for td in ticker_data:
        for i in range(12):
            monthly_totals[i] += td["amounts"][i]
    monthly_totals = [round(v, 2) for v in monthly_totals]

    return jsonify({"months": months_labels, "tickers": ticker_data, "monthly_totals": monthly_totals})


# ── Analytics: Backtest ───────────────────────────────────────────────────────

@app.route("/api/analytics/backtest", methods=["POST"])
def analytics_backtest():
    import yfinance as yf
    data = request.get_json() or {}
    req_tickers = [t.upper() for t in data.get("tickers", [])]
    period = data.get("period", "1y")
    if not req_tickers:
        return jsonify({"dates": [], "series": []})

    try:
        close = yf.download(req_tickers, period=period, auto_adjust=True, progress=False)
        if hasattr(close, "columns") and isinstance(close.columns, pd.MultiIndex):
            close = close["Close"]
        elif "Close" in close.columns:
            close = close[["Close"]]
            close.columns = req_tickers[:1]
    except Exception as e:
        return jsonify({"error": str(e)}), 500

    if close.empty:
        return jsonify({"dates": [], "series": []})

    series = []
    for t in req_tickers:
        if t not in close.columns:
            continue
        col = close[t].dropna()
        if len(col) < 2:
            continue
        normalized = (col / col.iloc[0]) * 10000
        step = max(1, len(normalized) // 200)
        sampled = normalized.iloc[::step]
        series.append({"ticker": t, "values": [round(float(v), 2) for v in sampled.values]})

    # Use the longest ticker's dates
    if series:
        longest = max(req_tickers, key=lambda t: len(close[t].dropna()) if t in close.columns else 0)
        col = close[longest].dropna()
        step = max(1, len(col) // 200)
        dates = [d.strftime("%Y-%m-%d") for d in col.index[::step]]
    else:
        dates = []

    return jsonify({"dates": dates, "series": series})


# ── Analytics: Yield Trend ────────────────────────────────────────────────────

@app.route("/api/analytics/yield-trend", methods=["POST"])
def analytics_yield_trend():
    import yfinance as yf
    data = request.get_json() or {}
    req_tickers = [t.upper() for t in data.get("tickers", [])]
    period = data.get("period", "2y")
    if not req_tickers:
        return jsonify({"series": []})

    result_series = []
    for t in req_tickers:
        try:
            tk = yf.Ticker(t)
            hist = tk.history(period=period)
            divs = tk.dividends
            if hist.empty or divs.empty:
                continue

            # Make both timezone-naive for comparison
            hist.index = hist.index.tz_localize(None)
            divs.index = divs.index.tz_localize(None)

            dates_out = []
            values_out = []
            close_prices = hist["Close"]
            # Compute trailing 12-month yield at sampled points
            step = max(1, len(close_prices) // 100)
            for i in range(step, len(close_prices), step):
                date = close_prices.index[i]
                price = float(close_prices.iloc[i])
                if price <= 0:
                    continue
                start = date - pd.Timedelta(days=365)
                ttm_divs = divs[(divs.index >= start) & (divs.index <= date)]
                ttm_yield = float(ttm_divs.sum()) / price * 100
                dates_out.append(date.strftime("%Y-%m-%d"))
                values_out.append(round(ttm_yield, 3))

            if dates_out:
                result_series.append({"ticker": t, "dates": dates_out, "values": values_out})
        except Exception:
            continue

    return jsonify({"series": result_series})


# ── Analytics: Rolling Metrics ────────────────────────────────────────────────

@app.route("/api/analytics/rolling-metrics", methods=["POST"])
def analytics_rolling_metrics():
    import yfinance as yf
    data = request.get_json() or {}
    req_tickers = [t.upper() for t in data.get("tickers", [])]
    period = data.get("period", "2y")
    window = int(data.get("window", 126))
    if not req_tickers:
        return jsonify({"dates": [], "sharpe": [], "sortino": []})

    try:
        close = yf.download(req_tickers, period=period, auto_adjust=True, progress=False)
        if hasattr(close, "columns") and isinstance(close.columns, pd.MultiIndex):
            close = close["Close"]
        elif "Close" in close.columns:
            close = close[["Close"]]
            close.columns = req_tickers[:1]
    except Exception as e:
        return jsonify({"error": str(e)}), 500

    if close.empty:
        return jsonify({"dates": [], "sharpe": [], "sortino": []})

    returns = close.pct_change().dropna()
    sharpe_series = []
    sortino_series = []

    for t in req_tickers:
        if t not in returns.columns:
            continue
        r = returns[t].dropna()
        if len(r) < window + 10:
            continue

        sharpe_vals = []
        sortino_vals = []
        indices = range(window, len(r), max(1, (len(r) - window) // 200))
        for i in indices:
            w_ret = r.iloc[i - window:i]
            mean_ret = float(w_ret.mean()) * 252
            std_ret = float(w_ret.std()) * (252 ** 0.5)
            downside = w_ret[w_ret < 0]
            down_std = float(downside.std()) * (252 ** 0.5) if len(downside) > 1 else 0

            sharpe_vals.append(round(mean_ret / std_ret, 3) if std_ret > 0 else 0)
            sortino_vals.append(round(mean_ret / down_std, 3) if down_std > 0 else 0)

        dates_idx = [r.index[i] for i in indices]
        sharpe_series.append({"ticker": t, "values": sharpe_vals})
        sortino_series.append({"ticker": t, "values": sortino_vals})

    if sharpe_series:
        # Use first ticker's dates as reference
        t0 = req_tickers[0] if req_tickers[0] in returns.columns else list(returns.columns)[0]
        r0 = returns[t0].dropna()
        indices = range(window, len(r0), max(1, (len(r0) - window) // 200))
        dates = [r0.index[i].strftime("%Y-%m-%d") for i in indices]
    else:
        dates = []

    return jsonify({"dates": dates, "sharpe": sharpe_series, "sortino": sortino_series})


# ── Analytics: NAV Erosion Chart ──────────────────────────────────────────────

@app.route("/api/analytics/nav-erosion-chart", methods=["POST"])
def analytics_nav_erosion_chart():
    import yfinance as yf
    data = request.get_json() or {}
    req_tickers = [t.upper() for t in data.get("tickers", [])]
    period = data.get("period", "2y")
    if not req_tickers:
        return jsonify({"series": []})

    result_series = []
    for t in req_tickers:
        try:
            tk = yf.Ticker(t)
            hist = tk.history(period=period)
            divs = tk.dividends
            if hist.empty:
                continue

            hist.index = hist.index.tz_localize(None)
            close_prices = hist["Close"]

            # Build cumulative dividends
            cum_divs = [0.0] * len(close_prices)
            if not divs.empty:
                divs.index = divs.index.tz_localize(None)
                running = 0.0
                div_idx = 0
                sorted_divs = divs.sort_index()
                for i, date in enumerate(close_prices.index):
                    while div_idx < len(sorted_divs) and sorted_divs.index[div_idx] <= date:
                        running += float(sorted_divs.iloc[div_idx])
                        div_idx += 1
                    cum_divs[i] = round(running, 4)

            # Total return line (price + cumulative dividends, normalized to starting price)
            start_price = float(close_prices.iloc[0])
            total_return = [round(float(close_prices.iloc[i]) + cum_divs[i] - start_price + start_price, 2) for i in range(len(close_prices))]

            step = max(1, len(close_prices) // 200)
            dates = [d.strftime("%Y-%m-%d") for d in close_prices.index[::step]]
            prices = [round(float(v), 2) for v in close_prices.values[::step]]
            cum_d = [cum_divs[i] for i in range(0, len(cum_divs), step)]
            total_r = [total_return[i] for i in range(0, len(total_return), step)]

            result_series.append({
                "ticker": t, "dates": dates, "prices": prices,
                "cum_dividends": cum_d, "total_return_line": total_r,
            })
        except Exception:
            continue

    return jsonify({"series": result_series})


# ── Analytics: Peer Comparison ────────────────────────────────────────────────

PEER_GROUPS = {
    "S&P 500 Covered Calls": ["SPYI", "JEPI", "JEPY", "XYLD", "PBP", "SPY"],
    "Nasdaq Covered Calls": ["QQQI", "JEPQ", "QYLD", "QYLG", "QQQ"],
    "Dividend Growth": ["SCHD", "VIG", "DGRO", "DGRW", "NOBL", "SDY", "DVY", "HDV"],
    "Broad US Market": ["VOO", "VTI", "SPY", "IVV", "QQQ", "SPLG"],
    "Growth": ["VUG", "SCHG", "IWF", "SPYG", "QQQ", "VGT", "XLK"],
    "High-Yield Bonds": ["HYG", "JNK", "HYGV", "USHY", "SHYG"],
    "REITs": ["VNQ", "O", "STAG", "NNN", "SCHH", "IYR"],
    "BDCs": ["BIZD", "ARCC", "MAIN", "HTGC", "GBDC", "BXSL"],
    "MLPs & Infrastructure": ["MLPA", "AMLP", "TPVG", "EMD"],
    "Preferred Stock": ["PFF", "PGX", "PFFD", "PSK"],
    "Senior Loans / CLOs": ["JAAA", "CLOZ", "SRLN", "FLOT"],
    "International Dividend": ["VIGI", "IDV", "SCHY", "DWX"],
    "Bonds / Aggregate": ["BND", "AGG", "VCIT", "SCHZ", "TLT"],
    "Small-Cap": ["VBK", "IJR", "IWM", "SCHA"],
    "Semiconductors": ["SMH", "SOXX", "XSD"],
    "Russell 2000 CC": ["RYLD", "RYLG", "IWMI"],
}

@app.route("/api/analytics/peers", methods=["POST"])
def analytics_peers():
    import yfinance as yf
    data = request.get_json() or {}
    ticker = (data.get("ticker", "")).upper()
    if not ticker:
        return jsonify({"ticker": "", "category": "", "peers": []})

    # Find which group(s) this ticker belongs to
    found_category = ""
    peer_tickers = []
    for cat, members in PEER_GROUPS.items():
        if ticker in members:
            found_category = cat
            peer_tickers = [t for t in members if t != ticker]
            break

    # Also check single-stock ETFs
    if not found_category and ticker in _get_single_stock_etfs():
        found_category = "Single-Stock ETFs"
        ss = _get_single_stock_etfs()
        peer_tickers = sorted([t for t in ss if t != ticker])[:10]

    if not peer_tickers:
        return jsonify({"ticker": ticker, "category": "", "peers": []})

    # Fetch quick metrics for peers
    from concurrent.futures import ThreadPoolExecutor

    def _fetch_peer(t):
        try:
            info = yf.Ticker(t).info
            name = info.get("shortName", info.get("longName", ""))
            yld = info.get("yield", info.get("dividendYield", 0))
            if yld and yld > 0:
                yld = float(yld) * 100
            else:
                yld = None
            # 1Y return from regularMarketPrice / fiftyTwoWeekLow approximate
            price = info.get("regularMarketPrice", info.get("previousClose", 0))
            year_low = info.get("fiftyTwoWeekLow", 0)
            if price and year_low and year_low > 0:
                ret_1y = (price / year_low - 1) * 100  # approximation from 52w low
            else:
                ret_1y = None
            return {"ticker": t, "name": name, "yield_pct": round(yld, 2) if yld else None, "return_1y": round(ret_1y, 1) if ret_1y else None}
        except Exception:
            return {"ticker": t, "name": "", "yield_pct": None, "return_1y": None}

    with ThreadPoolExecutor(max_workers=6) as pool:
        peers = list(pool.map(_fetch_peer, peer_tickers))

    return jsonify({"ticker": ticker, "category": found_category, "peers": peers})


# ── Portfolio Builder CRUD ────────────────────────────────────────────────────

@app.route("/api/builder/portfolios", methods=["GET"])
def builder_list_portfolios():
    pid = get_profile_id()
    conn = get_connection()
    rows = conn.execute("""
        SELECT p.id, p.name, p.notes, p.created_at, p.updated_at,
               (SELECT COUNT(*) FROM builder_holdings h WHERE h.portfolio_id = p.id) AS holding_count
        FROM builder_portfolios p WHERE p.profile_id = ?
        ORDER BY p.updated_at DESC
    """, (pid,)).fetchall()
    conn.close()
    return jsonify({"portfolios": rows_to_dicts(rows)})


@app.route("/api/builder/portfolios", methods=["POST"])
def builder_create_portfolio():
    pid = get_profile_id()
    data = request.get_json()
    name = (data.get("name") or "").strip()
    if not name:
        return jsonify({"error": "Name is required"}), 400
    notes = (data.get("notes") or "").strip()
    conn = get_connection()
    try:
        cur = conn.execute(
            "INSERT INTO builder_portfolios (profile_id, name, notes) VALUES (?, ?, ?)",
            (pid, name, notes))
        conn.commit()
        new_id = cur.lastrowid
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 400
    conn.close()
    return jsonify({"id": new_id, "name": name}), 201


@app.route("/api/builder/portfolios/<int:port_id>", methods=["PATCH"])
def builder_update_portfolio(port_id):
    pid = get_profile_id()
    data = request.get_json()
    conn = get_connection()
    fields, vals = [], []
    if "name" in data:
        fields.append("name = ?")
        vals.append(data["name"].strip())
    if "notes" in data:
        fields.append("notes = ?")
        vals.append(data["notes"].strip())
    if not fields:
        conn.close()
        return jsonify({"error": "Nothing to update"}), 400
    fields.append("updated_at = CURRENT_TIMESTAMP")
    vals.extend([port_id, pid])
    conn.execute(f"UPDATE builder_portfolios SET {', '.join(fields)} WHERE id = ? AND profile_id = ?", vals)
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/builder/portfolios/<int:port_id>", methods=["DELETE"])
def builder_delete_portfolio(port_id):
    pid = get_profile_id()
    conn = get_connection()
    conn.execute("DELETE FROM builder_holdings WHERE portfolio_id = ?", (port_id,))
    conn.execute("DELETE FROM builder_portfolios WHERE id = ? AND profile_id = ?", (port_id, pid))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/builder/portfolios/<int:port_id>/holdings", methods=["GET"])
def builder_list_holdings(port_id):
    conn = get_connection()
    rows = conn.execute(
        "SELECT id, ticker, dollar_amount, added_at FROM builder_holdings WHERE portfolio_id = ? ORDER BY added_at",
        (port_id,)).fetchall()
    conn.close()
    return jsonify({"holdings": rows_to_dicts(rows)})


@app.route("/api/builder/portfolios/<int:port_id>/holdings", methods=["POST"])
def builder_add_holding(port_id):
    data = request.get_json()
    ticker = (data.get("ticker") or "").strip().upper()
    dollar_amount = float(data.get("dollar_amount", 0))
    if not ticker:
        return jsonify({"error": "Ticker is required"}), 400
    conn = get_connection()
    conn.execute(
        "INSERT OR REPLACE INTO builder_holdings (portfolio_id, ticker, dollar_amount) VALUES (?, ?, ?)",
        (port_id, ticker, dollar_amount))
    conn.execute("UPDATE builder_portfolios SET updated_at = CURRENT_TIMESTAMP WHERE id = ?", (port_id,))
    conn.commit()
    conn.close()
    return jsonify({"ticker": ticker, "dollar_amount": dollar_amount})


@app.route("/api/builder/portfolios/<int:port_id>/holdings/<ticker>", methods=["DELETE"])
def builder_delete_holding(port_id, ticker):
    conn = get_connection()
    conn.execute("DELETE FROM builder_holdings WHERE portfolio_id = ? AND ticker = ?", (port_id, ticker.upper()))
    conn.execute("UPDATE builder_portfolios SET updated_at = CURRENT_TIMESTAMP WHERE id = ?", (port_id,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


# ── Portfolio Builder Analyze ────────────────────────────────────────────────

@app.route("/api/builder/portfolios/<int:port_id>/analyze", methods=["POST"])
def builder_analyze(port_id):
    import numpy as np
    import math
    import yfinance as yf
    from grading import (ticker_score, grade_portfolio, letter_grade,
                         _ulcer_index, _sharpe, _sortino, _calmar, _omega,
                         _max_drawdown, _capture_ratios, _safe)

    data = request.get_json() or {}
    benchmark = (data.get("benchmark") or "SPY").strip().upper()
    period = data.get("period", "1y")

    conn = get_connection()
    rows = conn.execute(
        "SELECT ticker, dollar_amount FROM builder_holdings WHERE portfolio_id = ?",
        (port_id,)).fetchall()
    pname = conn.execute("SELECT name FROM builder_portfolios WHERE id = ?", (port_id,)).fetchone()
    conn.close()

    if not rows:
        return jsonify({"error": "No holdings in this portfolio"}), 400

    holdings_raw = [{"ticker": r["ticker"].upper(), "dollar_amount": float(r["dollar_amount"])} for r in rows]
    tickers = [h["ticker"] for h in holdings_raw]
    all_tickers = list(set(tickers + [benchmark]))

    # Download price data
    try:
        df = yf.download(all_tickers, period=period, auto_adjust=True, progress=False)
    except Exception as e:
        return jsonify({"error": f"yfinance download failed: {e}"}), 500

    if df.empty:
        return jsonify({"error": "No price data returned"}), 500

    if isinstance(df.columns, pd.MultiIndex):
        close_df = df["Close"] if "Close" in df.columns.get_level_values(0) else df
    else:
        close_df = df if len(all_tickers) == 1 else df
        if len(all_tickers) == 1:
            close_df = pd.DataFrame({all_tickers[0]: df["Close"] if "Close" in df.columns else df.iloc[:, 0]})

    close_df = close_df.ffill().dropna(how="all")
    bench_close = close_df[benchmark] if benchmark in close_df.columns else None
    bench_ret = bench_close.pct_change().dropna() if bench_close is not None else None

    # NAV erosion
    nav_erosion = _compute_nav_erosion(close_df, tickers)

    # Total portfolio value
    total_value = sum(h["dollar_amount"] for h in holdings_raw)
    if total_value <= 0:
        total_value = 1.0

    # Per-ticker analysis
    result_holdings = []
    returns_cols = {}
    weights_list = []

    for h in holdings_raw:
        t = h["ticker"]
        amt = h["dollar_amount"]
        weight = amt / total_value

        row = {
            "ticker": t, "dollar_amount": amt, "weight_pct": round(weight * 100, 2),
            "shares": 0, "current_price": 0, "score": 0, "grade": "N/A",
            "ulcer_index": None, "sharpe": None, "sortino": None, "calmar": None,
            "omega": None, "up_capture": None, "down_capture": None,
            "annual_ret": None, "annual_vol": None, "max_drawdown": None,
            "week52_high": None, "week52_low": None,
            "annual_yield_pct": 0, "annual_income": 0, "monthly_income": 0,
            "nav_erosion_pct": round((nav_erosion.get(t, 0)) * 100, 2),
        }

        if t not in close_df.columns:
            result_holdings.append(row)
            weights_list.append(weight)
            continue

        tc = close_df[t].dropna()
        if len(tc) < 10:
            result_holdings.append(row)
            weights_list.append(weight)
            continue

        daily_ret = tc.pct_change().dropna()
        returns_cols[t] = daily_ret

        # Current price and shares
        cur_price = float(tc.iloc[-1])
        row["current_price"] = round(cur_price, 2)
        row["shares"] = round(amt / cur_price, 4) if cur_price > 0 else 0

        # 52-week
        row["week52_high"] = round(float(tc.max()), 2)
        row["week52_low"] = round(float(tc.min()), 2)

        # Annualized return and vol
        n_days = len(tc)
        total_ret = float(tc.iloc[-1] / tc.iloc[0]) - 1.0
        ann_ret = (1 + total_ret) ** (252 / max(n_days, 1)) - 1.0
        ann_vol = float(daily_ret.std()) * np.sqrt(252) if len(daily_ret) > 1 else None

        row["annual_ret"] = round(ann_ret * 100, 2)
        row["annual_vol"] = round(ann_vol * 100, 2) if ann_vol is not None else None
        row["annual_total_ret"] = row["annual_ret"]

        # Ticker score (includes ulcer index)
        score, sharpe, sortino, calmar, omega, mdd, dc, ulcer = ticker_score(tc, daily_ret, bench_ret)
        row["score"] = score
        row["grade"] = letter_grade(score)
        row["sharpe"] = sharpe
        row["sortino"] = sortino
        row["calmar"] = calmar
        row["omega"] = omega
        row["max_drawdown"] = round(mdd * 100, 2) if mdd is not None else None
        row["down_capture"] = dc
        row["ulcer_index"] = ulcer

        # Up capture
        if bench_ret is not None:
            uc, _ = _capture_ratios(daily_ret, bench_ret)
            row["up_capture"] = uc

        # Dividend income — try info first, fall back to dividend history
        try:
            yf_ticker = yf.Ticker(t)
            info = yf_ticker.info or {}
            div_yield = info.get("yield") or 0
            # dividendYield from yfinance can be a percentage (e.g. 11.84) or
            # decimal (e.g. 0.1184) depending on the ticker — use "yield" which
            # is consistently a decimal, and fall back to dividendYield only if
            # it looks like a decimal (< 1)
            if div_yield == 0:
                dy = info.get("dividendYield") or 0
                if dy > 0:
                    div_yield = dy / 100 if dy > 1 else dy

            # Fallback: compute yield from dividend history, annualizing if < 1 year
            if div_yield == 0 and cur_price > 0:
                try:
                    divs = yf_ticker.dividends
                    if divs is not None and len(divs) > 0:
                        cutoff = divs.index[-1] - pd.Timedelta(days=365)
                        recent_divs = divs[divs.index >= cutoff]
                        if len(recent_divs) > 0:
                            total_div = float(recent_divs.sum())
                            span_days = (recent_divs.index[-1] - recent_divs.index[0]).days
                            if span_days < 300 and len(recent_divs) >= 2:
                                # Short history: annualize using avg payment × estimated frequency
                                avg_gap = span_days / (len(recent_divs) - 1)
                                payments_per_year = 365.0 / max(avg_gap, 1)
                                avg_payment = total_div / len(recent_divs)
                                annual_div_per_share = avg_payment * payments_per_year
                            else:
                                annual_div_per_share = total_div
                            div_yield = annual_div_per_share / cur_price
                except Exception:
                    pass

            # Fallback 2: check distributions (for return-of-capital / covered-call ETFs)
            if div_yield == 0 and cur_price > 0:
                try:
                    actions = yf_ticker.actions
                    if actions is not None and "Dividends" in actions.columns:
                        div_col = actions["Dividends"]
                        div_col = div_col[div_col > 0]
                        if len(div_col) > 0:
                            cutoff = div_col.index[-1] - pd.Timedelta(days=365)
                            recent = div_col[div_col.index >= cutoff]
                            if len(recent) > 0:
                                total_div = float(recent.sum())
                                span_days = (recent.index[-1] - recent.index[0]).days
                                if span_days < 300 and len(recent) >= 2:
                                    avg_gap = span_days / (len(recent) - 1)
                                    payments_per_year = 365.0 / max(avg_gap, 1)
                                    avg_payment = total_div / len(recent)
                                    annual_div_per_share = avg_payment * payments_per_year
                                else:
                                    annual_div_per_share = total_div
                                div_yield = annual_div_per_share / cur_price
                except Exception:
                    pass

            row["annual_yield_pct"] = round(div_yield * 100, 2)
            row["annual_income"] = round(amt * div_yield, 2)
            row["monthly_income"] = round(amt * div_yield / 12, 2)
        except Exception:
            pass

        result_holdings.append(row)
        weights_list.append(weight)

    # Portfolio-level metrics
    if returns_cols:
        # Exclude tickers with < 30 data points so they don't truncate
        # the overlap window below the grading minimum
        long_cols = {t: s for t, s in returns_cols.items() if len(s) >= 30}
        if not long_cols:
            long_cols = returns_cols  # fallback: use whatever we have
        returns_df = pd.DataFrame(long_cols).dropna()
        available = [t for t in tickers if t in long_cols]
        avail_weights = np.array([weights_list[tickers.index(t)] for t in available])
        if avail_weights.sum() > 0:
            avail_weights = avail_weights / avail_weights.sum()
        port_grade = grade_portfolio(returns_df[available], avail_weights, bench_ret)
    else:
        port_grade = {"grade": {"overall": "N/A", "score": 0, "breakdown": []}}
        returns_df = pd.DataFrame()
        available = []

    # NAV Health factor — add to grade breakdown
    if returns_cols and available:
        weighted_nav = sum(
            avail_weights[i] * nav_erosion.get(available[i], 0)
            for i in range(len(available))
        )
        if weighted_nav < 0:
            nav_health_score = max(0.0, 100.0 + weighted_nav * 200.0)
        else:
            nav_health_score = 100.0

        port_grade["grade"]["breakdown"].append({
            "category": "NAV Health",
            "score": round(nav_health_score, 1),
            "weight": 10,
            "grade": letter_grade(nav_health_score),
        })
        # Recalculate overall
        total_w = sum(b["weight"] for b in port_grade["grade"]["breakdown"])
        total_s = sum(b["score"] * b["weight"] for b in port_grade["grade"]["breakdown"])
        new_score = round(total_s / total_w, 1) if total_w > 0 else 0.0
        port_grade["grade"]["score"] = new_score
        port_grade["grade"]["overall"] = letter_grade(new_score)
        port_grade["nav_erosion_avg_pct"] = round(weighted_nav * 100, 2)

    port_grade["n_holdings"] = len(tickers)
    port_grade["total_value"] = round(total_value, 2)
    ann_income = sum(h["annual_income"] for h in result_holdings)
    port_grade["est_annual_income"] = round(ann_income, 2)
    port_grade["est_monthly_income"] = round(ann_income / 12, 2)

    # Correlation matrix
    corr_data = None
    if len(returns_cols) > 1:
        corr_df = pd.DataFrame(returns_cols).dropna().corr()
        labels = list(corr_df.columns)
        matrix = []
        for i in range(len(labels)):
            row = []
            for j in range(len(labels)):
                v = corr_df.iloc[i, j]
                row.append(round(float(v), 3) if not (math.isnan(v) or math.isinf(v)) else 0)
            matrix.append(row)
        corr_data = {"labels": labels, "matrix": matrix}

    # Drawdown series
    dd_data = None
    if returns_cols and available:
        port_daily = returns_df[available].dot(avail_weights)
        port_cum = (1 + port_daily).cumprod()
        running_max = port_cum.cummax()
        drawdown = ((port_cum - running_max) / running_max) * 100
        dates = [d.strftime("%Y-%m-%d") for d in drawdown.index]
        vals = [round(float(v), 2) for v in drawdown.values]
        # Subsample if > 300 points
        if len(dates) > 300:
            step = len(dates) // 300
            dates = dates[::step]
            vals = vals[::step]
        dd_data = {"dates": dates, "values": vals}

    return jsonify({
        "portfolio_id": port_id,
        "portfolio_name": pname["name"] if pname else "",
        "portfolio_metrics": port_grade,
        "holdings": result_holdings,
        "correlation": corr_data,
        "drawdown_series": dd_data,
    })


# ── Portfolio Builder Compare ────────────────────────────────────────────────

@app.route("/api/builder/compare", methods=["POST"])
def builder_compare():
    import numpy as np
    import yfinance as yf
    from grading import grade_portfolio, letter_grade

    data = request.get_json() or {}
    port_ids = data.get("portfolio_ids", [])
    period = data.get("period", "1y")
    benchmark = (data.get("benchmark") or "SPY").strip().upper()

    if len(port_ids) < 2:
        return jsonify({"error": "Select at least 2 portfolios"}), 400

    conn = get_connection()
    results = []

    # Gather all tickers across all portfolios
    all_tickers_set = {benchmark}
    port_data = []
    for pid in port_ids:
        rows = conn.execute(
            "SELECT ticker, dollar_amount FROM builder_holdings WHERE portfolio_id = ?", (pid,)).fetchall()
        pname = conn.execute("SELECT name FROM builder_portfolios WHERE id = ?", (pid,)).fetchone()
        holdings = [{"ticker": r["ticker"].upper(), "dollar_amount": float(r["dollar_amount"])} for r in rows]
        for h in holdings:
            all_tickers_set.add(h["ticker"])
        port_data.append({"id": pid, "name": pname["name"] if pname else f"Portfolio {pid}", "holdings": holdings})
    conn.close()

    all_tickers = list(all_tickers_set)
    try:
        df = yf.download(all_tickers, period=period, auto_adjust=True, progress=False)
    except Exception as e:
        return jsonify({"error": f"yfinance download failed: {e}"}), 500

    if df.empty:
        return jsonify({"error": "No price data"}), 500

    if isinstance(df.columns, pd.MultiIndex):
        close_df = df["Close"]
    else:
        close_df = df if len(all_tickers) > 1 else pd.DataFrame({all_tickers[0]: df["Close"]})

    close_df = close_df.ffill().dropna(how="all")
    bench_ret = close_df[benchmark].pct_change().dropna() if benchmark in close_df.columns else None

    for p in port_data:
        holdings = p["holdings"]
        total_val = sum(h["dollar_amount"] for h in holdings)
        if total_val <= 0:
            continue

        avail_tickers = [h["ticker"] for h in holdings if h["ticker"] in close_df.columns]
        if not avail_tickers:
            continue

        # Build returns per ticker, then exclude short-history tickers
        # (< 30 days) so they don't truncate the overlap below grading minimum
        returns_cols = {}
        for t in avail_tickers:
            col = close_df[t].dropna()
            if len(col) >= 2:
                returns_cols[t] = col.pct_change().dropna()
        if not returns_cols:
            continue
        long_cols = {t: s for t, s in returns_cols.items() if len(s) >= 30}
        if not long_cols:
            long_cols = returns_cols  # fallback
        returns_df = pd.DataFrame(long_cols).dropna()
        if len(returns_df) < 2:
            returns_df = pd.DataFrame(long_cols).fillna(0)
        used_tickers = list(long_cols.keys())
        weights = np.array([next(h["dollar_amount"] for h in holdings if h["ticker"] == t) for t in used_tickers])
        weights = weights / weights.sum()

        pg = grade_portfolio(returns_df[used_tickers], weights, bench_ret)

        # Add NAV Health to match analyze endpoint grading
        nav_erosion = _compute_nav_erosion(close_df, used_tickers)
        weighted_nav = sum(
            weights[i] * nav_erosion.get(used_tickers[i], 0)
            for i in range(len(used_tickers))
        )
        nav_health_score = max(0.0, 100.0 + weighted_nav * 200.0) if weighted_nav < 0 else 100.0
        pg["grade"]["breakdown"].append({
            "category": "NAV Health",
            "score": round(nav_health_score, 1),
            "weight": 10,
            "grade": letter_grade(nav_health_score),
        })
        total_w = sum(b["weight"] for b in pg["grade"]["breakdown"])
        total_s = sum(b["score"] * b["weight"] for b in pg["grade"]["breakdown"])
        new_score = round(total_s / total_w, 1) if total_w > 0 else 0.0
        pg["grade"]["score"] = new_score
        pg["grade"]["overall"] = letter_grade(new_score)

        ann_income = 0
        for h in holdings:
            try:
                info = yf.Ticker(h["ticker"]).info or {}
                dy = info.get("yield") or 0
                if dy == 0:
                    raw_dy = info.get("dividendYield") or 0
                    if raw_dy > 0:
                        dy = raw_dy / 100 if raw_dy > 1 else raw_dy
                ann_income += h["dollar_amount"] * dy
            except Exception:
                pass

        results.append({
            "id": p["id"],
            "name": p["name"],
            "score": pg["grade"]["score"],
            "grade": pg["grade"]["overall"],
            "monthly_income": round(ann_income / 12, 2),
            "sharpe": pg.get("sharpe"),
            "sortino": pg.get("sortino"),
            "calmar": pg.get("calmar"),
            "omega": pg.get("omega"),
            "ulcer_index": pg.get("ulcer_index"),
            "max_drawdown": pg.get("max_drawdown"),
            "effective_n": pg.get("effective_n"),
            "breakdown": {b["category"]: b["score"] for b in pg["grade"]["breakdown"]},
        })

    return jsonify({"results": results})


# ── Portfolio Builder All Weather ────────────────────────────────────────────

ALL_WEATHER_TARGETS_INCOME = [
    {"asset_class": "US Stocks",         "target_pct": 30.0, "etfs": ["SCHD", "SPYI", "QQQI", "JEPQ", "O", "MAIN"]},
    {"asset_class": "Long-Term Bonds",   "target_pct": 40.0, "etfs": ["TLT", "TLTW", "EDV"]},
    {"asset_class": "Intermediate Bonds", "target_pct": 15.0, "etfs": ["VCIT", "BND", "AGG"]},
    {"asset_class": "Gold",              "target_pct": 5.0,  "etfs": ["IAUI", "KGLD", "GLDN", "GLDM", "GLD", "IAU"]},
    {"asset_class": "Silver",            "target_pct": 2.5,  "etfs": ["KSLV", "SVLX"]},
    {"asset_class": "Commodities",       "target_pct": 7.5,  "etfs": ["PDBC", "DJP", "GSG"]},
]

ALL_WEATHER_TARGETS_GROWTH = [
    {"asset_class": "US Stocks",         "target_pct": 30.0, "etfs": ["VTI", "VOO", "QQQ", "VUG", "SCHG"]},
    {"asset_class": "Long-Term Bonds",   "target_pct": 40.0, "etfs": ["TLT", "EDV", "ZROZ"]},
    {"asset_class": "Intermediate Bonds", "target_pct": 15.0, "etfs": ["IEF", "BND", "GOVT"]},
    {"asset_class": "Gold",              "target_pct": 7.5,  "etfs": ["GLD", "GLDM", "IAU"]},
    {"asset_class": "Commodities",       "target_pct": 7.5,  "etfs": ["DJP", "PDBC", "GSG"]},
]

# ── Income Factory Targets (Steven Bavaria style: 2/3 credit, 1/3 equity-income) ──

INCOME_FACTORY_TARGETS_INCOME = [
    {"asset_class": "Covered Call ETFs",     "target_pct": 20.0, "etfs": ["SPYI", "QQQI", "JEPQ", "JEPI", "XYLD"]},
    {"asset_class": "High-Yield Bonds",      "target_pct": 20.0, "etfs": ["HYG", "JNK", "HYGV", "USHY"]},
    {"asset_class": "Senior Loans / CLOs",   "target_pct": 15.0, "etfs": ["JAAA", "CLOZ", "SRLN", "FLOT"]},
    {"asset_class": "BDCs",                  "target_pct": 15.0, "etfs": ["BIZD", "ARCC", "MAIN", "HTGC"]},
    {"asset_class": "REITs",                 "target_pct": 10.0, "etfs": ["O", "VNQ", "STAG", "NNN"]},
    {"asset_class": "MLPs & Infrastructure", "target_pct": 10.0, "etfs": ["MLPA", "AMLP", "TPVG"]},
    {"asset_class": "Preferred Stock",       "target_pct": 10.0, "etfs": ["PFF", "PGX", "PFFD"]},
]

INCOME_FACTORY_TARGETS_GROWTH = [
    {"asset_class": "Dividend Growth",       "target_pct": 20.0, "etfs": ["SCHD", "VIG", "DGRO", "DGRW"]},
    {"asset_class": "Covered Call ETFs",     "target_pct": 15.0, "etfs": ["SPYI", "QQQI", "JEPQ", "JEPI"]},
    {"asset_class": "High-Yield Bonds",      "target_pct": 15.0, "etfs": ["HYG", "JNK", "HYGV"]},
    {"asset_class": "Senior Loans / CLOs",   "target_pct": 15.0, "etfs": ["JAAA", "CLOZ", "SRLN", "FLOT"]},
    {"asset_class": "BDCs",                  "target_pct": 10.0, "etfs": ["BIZD", "ARCC", "MAIN"]},
    {"asset_class": "REITs",                 "target_pct": 10.0, "etfs": ["O", "VNQ", "STAG"]},
    {"asset_class": "MLPs & Infrastructure", "target_pct": 10.0, "etfs": ["MLPA", "AMLP", "TPVG"]},
    {"asset_class": "Preferred Stock",       "target_pct": 5.0,  "etfs": ["PFF", "PGX", "PFFD"]},
]

# ── Covered Call Income Portfolio ──────────────────────────────────────────
# Heavy allocation to covered-call / option-income ETFs for maximum premium income.

COVERED_CALL_TARGETS_INCOME = [
    {"asset_class": "S&P 500 Covered Calls",  "target_pct": 25.0, "etfs": ["SPYI", "JEPI", "XYLD", "PBP"]},
    {"asset_class": "Nasdaq Covered Calls",    "target_pct": 20.0, "etfs": ["QQQI", "JEPQ", "QYLD", "QYLG"]},
    {"asset_class": "Single-Stock CC / Yield", "target_pct": 15.0, "etfs": ["TSLY", "NVDY", "CONY", "MSTY", "AMZY", "APLY"]},
    {"asset_class": "Russell 2000 / Broad CC", "target_pct": 10.0, "etfs": ["RYLD", "RYLG", "IWMI"]},
    {"asset_class": "Dividend Growth Anchor",  "target_pct": 15.0, "etfs": ["SCHD", "VIG", "DGRO", "DGRW"]},
    {"asset_class": "Bonds / Stability",       "target_pct": 10.0, "etfs": ["JAAA", "BND", "VCIT", "AGG"]},
    {"asset_class": "REITs / Alternatives",    "target_pct": 5.0,  "etfs": ["O", "VNQ", "STAG", "NNN"]},
]

COVERED_CALL_TARGETS_GROWTH = [
    {"asset_class": "S&P 500 Covered Calls",  "target_pct": 20.0, "etfs": ["SPYI", "JEPI", "JEPY"]},
    {"asset_class": "Nasdaq Covered Calls",    "target_pct": 20.0, "etfs": ["QQQI", "JEPQ", "QYLG"]},
    {"asset_class": "Dividend Growth Core",    "target_pct": 25.0, "etfs": ["SCHD", "VIG", "DGRO", "DGRW"]},
    {"asset_class": "Russell 2000 / Broad CC", "target_pct": 10.0, "etfs": ["RYLD", "RYLG", "IWMI"]},
    {"asset_class": "Growth Equity",           "target_pct": 10.0, "etfs": ["VUG", "QQQ", "SCHG", "VTI"]},
    {"asset_class": "Bonds / Stability",       "target_pct": 10.0, "etfs": ["JAAA", "BND", "IEF", "GOVT"]},
    {"asset_class": "REITs / Alternatives",    "target_pct": 5.0,  "etfs": ["O", "VNQ", "STAG"]},
]


# ── Dividend Growth Portfolio ──────────────────────────────────────────────
# Focus on companies/ETFs with strong dividend growth track records.

DIVIDEND_GROWTH_TARGETS_INCOME = [
    {"asset_class": "Dividend Aristocrats",    "target_pct": 25.0, "etfs": ["NOBL", "SDY", "KNG"]},
    {"asset_class": "Broad Dividend Growth",   "target_pct": 25.0, "etfs": ["SCHD", "VIG", "DGRO", "DGRW"]},
    {"asset_class": "High-Yield Dividend",     "target_pct": 15.0, "etfs": ["HDV", "VYM", "SPYD", "DVY"]},
    {"asset_class": "International Dividend",  "target_pct": 10.0, "etfs": ["VIGI", "IDV", "SCHY", "DWX"]},
    {"asset_class": "REITs",                   "target_pct": 10.0, "etfs": ["O", "VNQ", "STAG", "NNN"]},
    {"asset_class": "Utilities / Staples",     "target_pct": 10.0, "etfs": ["XLU", "VPU", "XLP", "VDC"]},
    {"asset_class": "Bonds / Stability",       "target_pct": 5.0,  "etfs": ["BND", "VCIT", "AGG", "JAAA"]},
]

DIVIDEND_GROWTH_TARGETS_GROWTH = [
    {"asset_class": "Dividend Aristocrats",    "target_pct": 20.0, "etfs": ["NOBL", "SDY", "KNG"]},
    {"asset_class": "Broad Dividend Growth",   "target_pct": 30.0, "etfs": ["SCHD", "VIG", "DGRO", "DGRW"]},
    {"asset_class": "Quality / Low Vol",       "target_pct": 15.0, "etfs": ["QUAL", "SPHQ", "USMV", "SPLV"]},
    {"asset_class": "International Dividend",  "target_pct": 10.0, "etfs": ["VIGI", "IDV", "SCHY"]},
    {"asset_class": "Growth Equity",           "target_pct": 15.0, "etfs": ["VUG", "QQQ", "SCHG", "VTI"]},
    {"asset_class": "Bonds / Stability",       "target_pct": 10.0, "etfs": ["BND", "IEF", "GOVT", "JAAA"]},
]

# ── Retirement Income (Bucket Strategy) ───────────────────────────────────
# 3-bucket approach: near-term cash, medium-term bonds, long-term growth.

RETIREMENT_INCOME_TARGETS_INCOME = [
    {"asset_class": "Cash / Ultra-Short",      "target_pct": 15.0, "etfs": ["SGOV", "BIL", "SHV", "USFR"]},
    {"asset_class": "Short-Term Bonds",        "target_pct": 15.0, "etfs": ["JAAA", "VCSH", "SCHO", "FLOT"]},
    {"asset_class": "Intermediate Bonds",      "target_pct": 15.0, "etfs": ["VCIT", "BND", "AGG", "SCHZ"]},
    {"asset_class": "Dividend Income",         "target_pct": 20.0, "etfs": ["SCHD", "VYM", "HDV", "SPYD"]},
    {"asset_class": "Covered Call Income",     "target_pct": 15.0, "etfs": ["SPYI", "JEPI", "QQQI", "JEPQ"]},
    {"asset_class": "REITs",                   "target_pct": 10.0, "etfs": ["O", "VNQ", "STAG", "NNN"]},
    {"asset_class": "TIPS / Inflation Hedge",  "target_pct": 10.0, "etfs": ["TIP", "SCHP", "VTIP", "STIP"]},
]

RETIREMENT_INCOME_TARGETS_GROWTH = [
    {"asset_class": "Cash / Ultra-Short",      "target_pct": 10.0, "etfs": ["SGOV", "BIL", "SHV", "USFR"]},
    {"asset_class": "Short-Term Bonds",        "target_pct": 10.0, "etfs": ["JAAA", "VCSH", "SCHO"]},
    {"asset_class": "Intermediate Bonds",      "target_pct": 15.0, "etfs": ["VCIT", "BND", "AGG"]},
    {"asset_class": "Dividend Growth",         "target_pct": 20.0, "etfs": ["SCHD", "VIG", "DGRO", "DGRW"]},
    {"asset_class": "US Equity Growth",        "target_pct": 20.0, "etfs": ["VTI", "VOO", "VUG", "QQQ"]},
    {"asset_class": "International Equity",    "target_pct": 10.0, "etfs": ["VXUS", "VEA", "VIGI", "SCHY"]},
    {"asset_class": "TIPS / Inflation Hedge",  "target_pct": 10.0, "etfs": ["TIP", "SCHP", "VTIP", "STIP"]},
    {"asset_class": "REITs",                   "target_pct": 5.0,  "etfs": ["VNQ", "O", "STAG"]},
]

# ── Growth / Aggressive Growth Portfolio ──────────────────────────────────
# Capital appreciation focused, heavy tech/innovation, minimal income.

GROWTH_TARGETS_INCOME = [
    {"asset_class": "US Large-Cap Growth",     "target_pct": 30.0, "etfs": ["VUG", "SCHG", "IWF", "SPYG"]},
    {"asset_class": "Nasdaq / Tech",           "target_pct": 25.0, "etfs": ["QQQ", "QQQM", "XLK", "VGT"]},
    {"asset_class": "Semiconductors",          "target_pct": 10.0, "etfs": ["SMH", "SOXX", "XSD"]},
    {"asset_class": "Mid-Cap Growth",          "target_pct": 10.0, "etfs": ["VOT", "IJK", "MDYG"]},
    {"asset_class": "Small-Cap Growth",        "target_pct": 5.0,  "etfs": ["VBK", "IJT", "SLYG"]},
    {"asset_class": "International Growth",    "target_pct": 10.0, "etfs": ["VXUS", "VEA", "EFG", "VWO"]},
    {"asset_class": "Bonds / Stability",       "target_pct": 10.0, "etfs": ["BND", "JAAA", "VCSH", "AGG"]},
]

GROWTH_TARGETS_GROWTH = [
    {"asset_class": "US Large-Cap Growth",     "target_pct": 30.0, "etfs": ["VUG", "SCHG", "IWF", "SPYG"]},
    {"asset_class": "Nasdaq / Tech",           "target_pct": 25.0, "etfs": ["QQQ", "QQQM", "XLK", "VGT"]},
    {"asset_class": "Semiconductors",          "target_pct": 15.0, "etfs": ["SMH", "SOXX", "XSD"]},
    {"asset_class": "Innovation / Thematic",   "target_pct": 10.0, "etfs": ["ARKK", "ARKW", "MOON", "KOMP"]},
    {"asset_class": "Mid-Cap Growth",          "target_pct": 10.0, "etfs": ["VOT", "IJK", "MDYG"]},
    {"asset_class": "Small-Cap Growth",        "target_pct": 5.0,  "etfs": ["VBK", "IJT", "SLYG"]},
    {"asset_class": "International Growth",    "target_pct": 5.0,  "etfs": ["EFG", "VWO", "VXUS"]},
]


def get_strategy_targets(strategy, mode):
    """Pick the right target allocation list based on strategy + mode."""
    if strategy == "income_factory":
        return INCOME_FACTORY_TARGETS_INCOME if mode == "income" else INCOME_FACTORY_TARGETS_GROWTH
    if strategy == "covered_call":
        return COVERED_CALL_TARGETS_INCOME if mode == "income" else COVERED_CALL_TARGETS_GROWTH
    if strategy == "dividend_growth":
        return DIVIDEND_GROWTH_TARGETS_INCOME if mode == "income" else DIVIDEND_GROWTH_TARGETS_GROWTH
    if strategy == "retirement_income":
        return RETIREMENT_INCOME_TARGETS_INCOME if mode == "income" else RETIREMENT_INCOME_TARGETS_GROWTH
    if strategy == "growth":
        return GROWTH_TARGETS_INCOME if mode == "income" else GROWTH_TARGETS_GROWTH
    return ALL_WEATHER_TARGETS_INCOME if mode == "income" else ALL_WEATHER_TARGETS_GROWTH


@app.route("/api/builder/all-weather", methods=["POST"])
def builder_all_weather():
    data = request.get_json() or {}
    mode = data.get("mode", "income")
    budget = float(data.get("budget", 100000))
    strategy = data.get("strategy", "all_weather")
    funds_per_class = min(int(data.get("funds_per_class", 1)), 4)
    if funds_per_class < 1:
        funds_per_class = 1

    targets = get_strategy_targets(strategy, mode)

    # Get user's existing holdings for matching
    _, h_pids = get_profile_filter()
    conn = get_connection()
    h_ph = ",".join("?" * len(h_pids))
    existing = conn.execute(
        f"SELECT ticker, SUM(current_value) as current_value FROM holdings WHERE profile_id IN ({h_ph}) GROUP BY ticker", h_pids
    ).fetchall()
    conn.close()
    existing_map = {r["ticker"].upper(): float(r["current_value"] or 0) for r in existing}

    allocations = []
    used_tickers = set()  # no duplicate tickers across classes
    for slot in targets:
        candidates = slot["etfs"]
        class_budget = budget * slot["target_pct"] / 100

        # Rank candidates: existing holdings first (by value desc), then remaining
        owned = [(etf, existing_map[etf.upper()]) for etf in candidates if etf.upper() in existing_map]
        owned.sort(key=lambda x: -x[1])
        not_owned = [etf for etf in candidates if etf.upper() not in existing_map]
        ranked = [etf for etf, _ in owned] + not_owned

        # Filter out already-used tickers
        available = [etf for etf in ranked if etf.upper() not in used_tickers]
        n_funds = min(funds_per_class, len(available))
        if n_funds < 1:
            n_funds = 1
            available = ranked[:1]  # fallback: reuse if no alternatives
        per_fund = round(class_budget / n_funds, 2)

        for i in range(n_funds):
            ticker = available[i]
            used_tickers.add(ticker.upper())
            source = "existing" if ticker.upper() in existing_map else "recommended"
            # Last fund gets any rounding remainder
            amt = per_fund if i < n_funds - 1 else round(class_budget - per_fund * (n_funds - 1), 2)
            allocations.append({
                "asset_class": slot["asset_class"],
                "target_pct": round(slot["target_pct"] / n_funds, 2),
                "ticker": ticker,
                "source": source,
                "dollar_amount": amt,
                "candidates": [c for c in candidates if c.upper() not in used_tickers or c.upper() == ticker.upper()],
            })

    return jsonify({"allocations": allocations, "mode": mode, "budget": budget})


# ── Settings API ──────────────────────────────────────────────────────────────

@app.route("/api/single-stock-etfs", methods=["GET"])
def get_single_stock_etfs():
    """Return built-in and user-added single-stock ETFs."""
    conn = get_connection()
    row = conn.execute("SELECT value FROM settings WHERE key = 'single_stock_etfs'").fetchone()
    conn.close()
    user_added = []
    if row and row["value"]:
        user_added = [t.strip().upper() for t in row["value"].split(",") if t.strip()]
    return jsonify({
        "builtin": sorted(_DEFAULT_SINGLE_STOCK_ETFS),
        "user_added": sorted(set(user_added)),
    })


@app.route("/api/single-stock-etfs", methods=["POST"])
def save_single_stock_etfs():
    """Save user-added single-stock ETF tickers."""
    data = request.get_json() or {}
    tickers = data.get("tickers", [])
    # Normalize and deduplicate, exclude any already in defaults
    cleaned = sorted({t.strip().upper() for t in tickers if t.strip()} - _DEFAULT_SINGLE_STOCK_ETFS)
    value = ",".join(cleaned)
    conn = get_connection()
    conn.execute("INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)",
                 ("single_stock_etfs", value))
    conn.commit()
    conn.close()
    return jsonify({"ok": True, "user_added": cleaned})


@app.route("/api/settings", methods=["GET"])
def get_settings():
    conn = get_connection()
    rows = conn.execute("SELECT key, value FROM settings").fetchall()
    conn.close()
    out = {}
    for r in rows:
        k = r["key"]
        v = r["value"]
        # Mask sensitive keys
        if "key" in k.lower() or "secret" in k.lower():
            out[k] = v[:4] + "..." + v[-4:] if v and len(v) > 8 else "***"
        else:
            out[k] = v
    return jsonify(out)


@app.route("/api/settings", methods=["POST"])
def save_settings():
    data = request.get_json() or {}
    conn = get_connection()
    for k, v in data.items():
        conn.execute("INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)", (k, str(v)))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


# ── Portfolio Builder Rebalance ────────────────────────────────────────────────

@app.route("/api/builder/portfolios/<int:pid>/rebalance", methods=["POST"])
def builder_rebalance(pid):
    data = request.get_json() or {}
    mode = data.get("mode", "income")
    strategy = data.get("strategy", "all_weather")

    targets = get_strategy_targets(strategy, mode)

    # Fetch portfolio holdings
    conn = get_connection()
    rows = conn.execute(
        "SELECT ticker, dollar_amount FROM builder_holdings WHERE portfolio_id = ?", (pid,)
    ).fetchall()
    if not rows:
        conn.close()
        return jsonify({"error": "Portfolio has no holdings"}), 400

    port_holdings = [{"ticker": r["ticker"].upper(), "amount": float(r["dollar_amount"] or 0)} for r in rows]
    total_value = sum(h["amount"] for h in port_holdings)
    if total_value <= 0:
        conn.close()
        return jsonify({"error": "Portfolio total value is zero"}), 400

    # Build ETF → asset class lookup
    etf_to_class = {}
    for slot in targets:
        for etf in slot["etfs"]:
            etf_to_class[etf.upper()] = slot["asset_class"]

    # Classify holdings
    class_holdings = {slot["asset_class"]: [] for slot in targets}
    unclassified = []
    for h in port_holdings:
        ac = etf_to_class.get(h["ticker"])
        if ac:
            class_holdings[ac].append(h)
        else:
            unclassified.append(h)

    # Get user's real holdings for suggesting new ETFs
    _, r_pids = get_profile_filter()
    r_ph = ",".join("?" * len(r_pids))
    existing = conn.execute(
        f"SELECT ticker, SUM(current_value) as current_value FROM holdings WHERE profile_id IN ({r_ph}) GROUP BY ticker", r_pids
    ).fetchall()
    conn.close()
    existing_map = {r["ticker"].upper(): float(r["current_value"] or 0) for r in existing}

    suggestions = []
    for slot in targets:
        ac = slot["asset_class"]
        current_amount = sum(h["amount"] for h in class_holdings[ac])
        current_pct = round(current_amount / total_value * 100, 1)
        target_pct = slot["target_pct"]
        drift_pct = round(current_pct - target_pct, 1)
        target_amount = round(total_value * target_pct / 100, 2)
        change_amount = round(target_amount - current_amount, 2)

        # Determine action
        if not class_holdings[ac]:
            action = "add_new"
        elif change_amount > 0:
            action = "buy"
        elif change_amount < 0:
            action = "reduce"
        else:
            action = "on_target"

        # Suggest ETF: use existing holding in that class, or auto-select
        if class_holdings[ac]:
            suggested = max(class_holdings[ac], key=lambda h: h["amount"])["ticker"]
        else:
            # Same auto-select logic: prefer user's real holdings
            suggested = None
            best_val = -1
            for etf in slot["etfs"]:
                if etf.upper() in existing_map and existing_map[etf.upper()] > best_val:
                    best_val = existing_map[etf.upper()]
                    suggested = etf
            if suggested is None:
                suggested = slot["etfs"][0]

        suggestions.append({
            "asset_class": ac,
            "target_pct": target_pct,
            "current_pct": current_pct,
            "drift_pct": drift_pct,
            "current_amount": round(current_amount, 2),
            "target_amount": target_amount,
            "change_amount": change_amount,
            "action": action,
            "holdings": class_holdings[ac],
            "suggested_ticker": suggested,
        })

    return jsonify({
        "total_value": round(total_value, 2),
        "mode": mode,
        "suggestions": suggestions,
        "unclassified": unclassified,
    })


# ── Distribution Comparison ────────────────────────────────────────────────────

@app.route("/api/distribution-compare/lookup")
def distribution_compare_lookup():
    import yfinance as yf
    import warnings
    warnings.filterwarnings("ignore")

    ticker = request.args.get("ticker", "").strip().upper()
    if not ticker:
        return jsonify(error="No ticker provided."), 400

    try:
        t = yf.Ticker(ticker)
        hist = t.history(period="1y", auto_adjust=False, actions=True)
    except Exception as e:
        return jsonify(error=f"Failed to fetch data for {ticker}: {str(e)}"), 400

    if hist is None or hist.empty:
        return jsonify(error=f"No data found for {ticker}."), 404

    close = hist["Close"].dropna()
    if close.empty:
        return jsonify(error=f"No price data for {ticker}."), 404

    current_price = float(close.iloc[-1])
    divs = hist["Dividends"] if "Dividends" in hist.columns else pd.Series(0.0, index=hist.index)
    ttm_divs = float(divs.sum())
    ttm_yield = ttm_divs / current_price if current_price > 0 else 0.0
    yield_found = ttm_divs > 0

    return jsonify(
        ticker=ticker,
        price=round(current_price, 4),
        ttm_yield=round(ttm_yield * 100, 4),
        yield_found=yield_found,
    )


@app.route("/api/distribution-compare/run", methods=["POST"])
def distribution_compare_run():
    import traceback as _tb
    try:
        result = _distribution_compare_compute()
        if isinstance(result, dict) and "error" in result:
            return jsonify(error=result["error"])
        return jsonify(**result)
    except Exception as _e:
        return jsonify(error=f"Server error: {str(_e)}", detail=_tb.format_exc())


@app.route("/api/distribution-compare/export", methods=["POST"])
def distribution_compare_export():
    import traceback as _tb
    try:
        return _distribution_compare_export_inner()
    except Exception as _e:
        return jsonify(error=f"Server error: {str(_e)}")


def _distribution_compare_export_inner():
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, numbers

    result_data = _distribution_compare_compute()
    if isinstance(result_data, dict) and "error" in result_data:
        return jsonify(error=result_data["error"])

    wb = Workbook()
    # Summary sheet
    ws = wb.active
    ws.title = "Summary"
    header_font = Font(bold=True)
    ws.cell(row=1, column=1, value="Distribution Comparison Summary").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"Comparison Type: {result_data.get('comparison_type', '')}")
    ws.cell(row=3, column=1, value=f"Cash Wedge Initial: ${result_data.get('cash_wedge_initial', 0):,.2f}")

    summary_row = 5
    summary_headers = ["Metric", "Fund A", "Fund B"]
    if "fund_c" in result_data:
        summary_headers.append("Fund C")
    for ci, h in enumerate(summary_headers, 1):
        ws.cell(row=summary_row, column=ci, value=h).font = header_font

    fund_keys = ["fund_a", "fund_b"]
    if "fund_c" in result_data:
        fund_keys.append("fund_c")

    metrics = [
        ("Ticker", "ticker"),
        ("Role", "role"),
        ("Investment", "investment"),
        ("Initial Shares", "initial_shares"),
        ("Final Portfolio", "final_portfolio"),
        ("Final Withdrawn", "final_withdrawn"),
        ("Final Distributions", "final_distributions"),
        ("Final Total", "final_total"),
        ("Depleted", "depleted"),
    ]
    for mi, (metric_label, metric_key) in enumerate(metrics):
        row_num = summary_row + 1 + mi
        ws.cell(row=row_num, column=1, value=metric_label)
        for fi, fk in enumerate(fund_keys):
            if fk in result_data:
                val = result_data[fk].get(metric_key, "")
                ws.cell(row=row_num, column=2 + fi, value=val)

    # Fund detail sheets
    for fund_key in fund_keys:
        if fund_key not in result_data:
            continue
        fund = result_data[fund_key]
        ws_f = wb.create_sheet(title=f"{fund['ticker']} Detail")
        headers = ["Month", "Price", "Shares", "Portfolio", "Dist/Share", "Income",
                    "Withdrawal", "Excess", "Shares +/-", "Growth", "Cum Income",
                    "ROI ($)", "ROI (%)"]
        if fund.get("has_cash_wedge"):
            headers.insert(7, "CW Drawn")
            headers.insert(8, "CW Balance")
        if fund.get("drip") is False:
            headers.append("Cash Accumulated")
        for ci, h in enumerate(headers, 1):
            ws_f.cell(row=1, column=ci, value=h).font = header_font
        months = result_data.get("months", [])
        for ri, row in enumerate(fund["monthly_rows"], 2):
            vals = [months[ri - 2] if ri - 2 < len(months) else "",
                    row["price"], row["shares"], row["portfolio"],
                    row["dist_per_share"], row["income"], row["withdrawal"]]
            if fund.get("has_cash_wedge"):
                vals.extend([row.get("wedge_drawn", 0), row.get("wedge_bal", 0)])
            vals.extend([row["excess"], row["shares_delta"], row["growth"],
                         row["cum_income"], row["roi_dollar"], row["roi_pct"]])
            if fund.get("drip") is False:
                vals.append(row.get("cash_accumulated", 0))
            for ci, v in enumerate(vals, 1):
                ws_f.cell(row=ri, column=ci, value=v)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name="distribution_comparison.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


def _distribution_compare_compute():
    """Core computation for distribution comparison. Returns a plain dict."""
    import yfinance as yf
    import numpy as np
    import warnings
    warnings.filterwarnings("ignore")
    from grading import _max_drawdown, _ulcer_index

    data = request.get_json(force=True, silent=True) or {}
    mode = data.get("mode", "historical")
    monthly_withdrawal = float(data.get("monthly_withdrawal", 0))
    if monthly_withdrawal < 0:
        return {"error": "Monthly withdrawal cannot be negative."}

    # Withdrawal strategy parameters
    withdrawal_strategy = data.get("withdrawal_strategy", "fixed")
    withdrawal_pct = float(data.get("withdrawal_pct", 4))
    inflation_rate = data.get("inflation_rate")
    if inflation_rate is not None:
        try:
            inflation_rate = float(inflation_rate)
        except (TypeError, ValueError):
            inflation_rate = None
    dynamic_reduce_pct = float(data.get("dynamic_reduce_pct", 25))
    dynamic_threshold_pct = float(data.get("dynamic_threshold_pct", 80))

    comparison_type = data.get("comparison_type", "income_vs_growth")
    cash_wedge_initial = float(data.get("cash_wedge", 0))
    if cash_wedge_initial < 0:
        cash_wedge_initial = 0.0

    growth_funds = set()
    if comparison_type == "income_vs_growth":
        growth_funds.add("fund_b")
    elif comparison_type == "growth_vs_growth":
        growth_funds.add("fund_a")
        growth_funds.add("fund_b")

    fund_a_input = data.get("fund_a", {})
    fund_b_input = data.get("fund_b", {})

    def _validate_fund(f, label):
        ticker = str(f.get("ticker", "")).strip().upper()
        if not ticker:
            return None, f"{label}: ticker is required."
        try:
            investment = float(f.get("investment", 0))
        except (TypeError, ValueError):
            return None, f"{label}: invalid investment amount."
        if investment <= 0:
            return None, f"{label}: investment must be greater than 0."
        yield_override = f.get("yield_override")
        if yield_override is not None and yield_override != "" and yield_override is not False:
            try:
                yield_override = float(yield_override)
                if yield_override <= 0:
                    yield_override = None
            except (TypeError, ValueError):
                yield_override = None
        else:
            yield_override = None
        drip = f.get("drip", True)
        return {"ticker": ticker, "investment": investment,
                "yield_override": yield_override, "drip": drip}, None

    fa, err = _validate_fund(fund_a_input, "Fund A")
    if err:
        return {"error": err}
    fb, err = _validate_fund(fund_b_input, "Fund B")
    if err:
        return {"error": err}

    # Optional Fund C
    fund_c_input = data.get("fund_c")
    fc = None
    if fund_c_input:
        fc, err = _validate_fund(fund_c_input, "Fund C")
        if err:
            return {"error": err}

    # Growth funds for fund_c
    if fc:
        if comparison_type == "income_vs_growth":
            growth_funds.add("fund_c")
        elif comparison_type == "growth_vs_growth":
            growth_funds.add("fund_c")
        # income_vs_income: fund_c stays as income (not in growth_funds)

    def _compute_effective_withdrawal(base_withdrawal, month_index, current_pv, initial_inv):
        """Compute effective withdrawal based on strategy + inflation."""
        if withdrawal_strategy == "percentage":
            eff = current_pv * (withdrawal_pct / 100.0 / 12.0)
        elif withdrawal_strategy == "dynamic":
            eff = base_withdrawal
            threshold = initial_inv * (dynamic_threshold_pct / 100.0)
            if current_pv < threshold:
                eff = eff * (1.0 - dynamic_reduce_pct / 100.0)
        else:
            eff = base_withdrawal
        # Apply inflation
        if inflation_rate is not None and inflation_rate != 0:
            eff *= (1.0 + inflation_rate / 100.0) ** (month_index / 12.0)
        return eff

    def _compute_risk_metrics(portfolio_values, monthly_rows):
        """Compute risk metrics for a fund's portfolio value series."""
        pv_series = pd.Series(portfolio_values)
        max_dd = _max_drawdown(pv_series) if len(pv_series) > 1 else None
        ulcer = _ulcer_index(pv_series)
        worst_month_val = min((r["growth"] for r in monthly_rows), default=0)
        worst_month_idx = next((i for i, r in enumerate(monthly_rows)
                                if r["growth"] == worst_month_val), None)
        recovery_months = None
        if max_dd is not None and max_dd < 0 and len(pv_series) > 1:
            running_max = pv_series.cummax()
            drawdowns = (pv_series - running_max) / running_max
            trough_idx = drawdowns.idxmin()
            peak_before = running_max.iloc[trough_idx]
            for ri in range(trough_idx + 1, len(pv_series)):
                if pv_series.iloc[ri] >= peak_before:
                    recovery_months = ri - trough_idx
                    break
        return {
            "max_drawdown_pct": round(max_dd * 100, 2) if max_dd else None,
            "ulcer_index": ulcer,
            "worst_month_value": round(worst_month_val, 2),
            "worst_month_idx": worst_month_idx,
            "recovery_months": recovery_months,
        }

    def _compute_depletion_month(depleted, monthly_rows):
        """Find the first month where depletion occurred."""
        if not depleted:
            return None
        for di, r in enumerate(monthly_rows):
            if r["shares"] == 0 and r["portfolio"] == 0:
                return di
        return None

    # ── HISTORICAL MODE ──
    if mode == "historical":
        duration_str = data.get("duration", "10y")

        all_funds = [fa, fb]
        if fc:
            all_funds.append(fc)
        tickers = [f["ticker"] for f in all_funds]
        unique_tickers = list(dict.fromkeys(tickers))

        try:
            raw = yf.download(
                unique_tickers, period="max", interval="1d",
                auto_adjust=False, actions=True, group_by="ticker", progress=False,
            )
        except Exception as e:
            return {"error": f"Failed to fetch data: {str(e)}"}

        if raw is None or raw.empty:
            return {"error": "No data returned from Yahoo Finance."}

        def _extract(sym):
            try:
                if isinstance(raw.columns, pd.MultiIndex):
                    top_keys = raw.columns.get_level_values(0).unique()
                    if sym in top_keys:
                        sub = raw[sym]
                        close = sub["Close"] if "Close" in sub.columns else None
                        divs = sub["Dividends"] if "Dividends" in sub.columns else pd.Series(0.0, index=raw.index)
                    elif "Close" in top_keys:
                        close = raw["Close"][sym] if sym in raw["Close"].columns else raw["Close"].iloc[:, 0]
                        divs_df = raw["Dividends"] if "Dividends" in top_keys else None
                        divs = divs_df[sym] if divs_df is not None and sym in divs_df.columns else pd.Series(0.0, index=raw.index)
                    else:
                        return None, None
                else:
                    close = raw["Close"] if "Close" in raw.columns else None
                    divs = raw["Dividends"] if "Dividends" in raw.columns else pd.Series(0.0, index=raw.index)
            except Exception:
                return None, None
            return close, divs

        common_start = None
        newer_ticker = None
        for fund in all_funds:
            close_tmp, _ = _extract(fund["ticker"])
            if close_tmp is None or close_tmp.dropna().empty:
                return {"error": f"No data found for {fund['ticker']}."}
            first_valid = close_tmp.dropna().index[0]
            if common_start is None or first_valid > common_start:
                common_start = first_valid
                newer_ticker = fund["ticker"]

        from datetime import datetime as _dt, timedelta
        available_months = (pd.Timestamp.now() - common_start).days / 30.44
        available_years = available_months / 12
        if duration_str != "max":
            requested_years = int(duration_str.replace("y", ""))
            if requested_years > available_years:
                max_yrs = int(available_years)
                return {"error": (
                    f"{newer_ticker} only has data from {common_start.strftime('%Y-%m-%d')} "
                    f"(~{max_yrs} year{'s' if max_yrs != 1 else ''} of overlap). "
                    f"Please select {max_yrs} year{'s' if max_yrs != 1 else ''} or less, or use Max."
                )}
            trim_start = pd.Timestamp(_dt.now() - timedelta(days=requested_years * 365))
            if trim_start > common_start:
                common_start = trim_start

        results = {}
        fund_list = [("fund_a", fa), ("fund_b", fb)]
        if fc:
            fund_list.append(("fund_c", fc))
        date_labels = []
        for label, fund in fund_list:
            sym = fund["ticker"]
            investment = fund["investment"]
            drip = fund["drip"]

            close, divs = _extract(sym)
            close = close[close.index >= common_start]
            divs = divs[divs.index >= common_start]

            monthly_close = close.resample("ME").last()
            monthly_divs = divs.resample("ME").sum()
            df_m = pd.DataFrame({"price": monthly_close, "div": monthly_divs}).dropna(subset=["price"])
            df_m["div"] = df_m["div"].fillna(0.0)

            if df_m.empty:
                return {"error": f"No usable monthly data for {sym}."}

            initial_price = float(df_m["price"].iloc[0])
            if initial_price <= 0:
                return {"error": f"Initial price for {sym} is zero."}

            is_growth = label in growth_funds
            wedge = cash_wedge_initial if is_growth else 0.0
            shares = investment / initial_price
            cum_withdrawals = 0.0
            cum_distributions = 0.0
            prev_pv = investment
            cash_accumulated = 0.0

            portfolio_values = []
            cumulative_withdrawals_list = []
            cumulative_distributions_list = []
            share_counts = []
            total_values = []
            monthly_rows = []
            depleted = False

            for month_index, (dt, row) in enumerate(df_m.iterrows()):
                price = float(row["price"])
                div_per_share = float(row["div"])

                if depleted or shares <= 0:
                    depleted = True
                    portfolio_values.append(0.0)
                    cumulative_withdrawals_list.append(round(cum_withdrawals, 2))
                    cumulative_distributions_list.append(round(cum_distributions, 2))
                    share_counts.append(0.0)
                    tv = cum_withdrawals + wedge + cash_accumulated
                    total_values.append(round(tv, 2))
                    row_data = {
                        "price": 0, "shares": 0, "portfolio": 0,
                        "dist_per_share": 0, "income": 0, "withdrawal": 0,
                        "wedge_drawn": 0, "wedge_bal": round(wedge, 2),
                        "excess": 0, "shares_delta": 0, "growth": 0,
                        "cum_income": round(cum_distributions, 2),
                        "roi_dollar": round(cum_withdrawals + wedge + cash_accumulated - investment - (cash_wedge_initial if is_growth else 0), 2),
                        "roi_pct": round(((cum_withdrawals + wedge + cash_accumulated - investment - (cash_wedge_initial if is_growth else 0)) / investment * 100), 2) if investment else 0,
                    }
                    if not drip:
                        row_data["cash_accumulated"] = round(cash_accumulated, 2)
                    monthly_rows.append(row_data)
                    continue

                shares_before = shares
                dist_this_month = div_per_share * shares
                cum_distributions += dist_this_month

                # Compute effective withdrawal for this month
                current_pv = shares * price
                effective_withdrawal = _compute_effective_withdrawal(
                    monthly_withdrawal, month_index, current_pv, investment)

                wedge_drawn = 0.0
                if is_growth:
                    need = effective_withdrawal
                    cw_draw = min(wedge, need)
                    wedge -= cw_draw
                    wedge_drawn = cw_draw
                    need -= cw_draw
                    div_used = min(dist_this_month, need)
                    need -= div_used
                    leftover_div = dist_this_month - div_used
                    if drip:
                        if leftover_div > 0 and price > 0:
                            shares += leftover_div / price
                    else:
                        cash_accumulated += leftover_div
                    if need > 0:
                        if not drip and cash_accumulated > 0:
                            cash_draw = min(cash_accumulated, need)
                            cash_accumulated -= cash_draw
                            need -= cash_draw
                        if need > 0 and price > 0:
                            shares -= need / price
                    cum_withdrawals += effective_withdrawal
                    excess = dist_this_month + cw_draw - effective_withdrawal
                else:
                    excess = dist_this_month - effective_withdrawal
                    if dist_this_month >= effective_withdrawal:
                        if drip:
                            if price > 0:
                                shares += excess / price
                        else:
                            cash_accumulated += excess
                        cum_withdrawals += effective_withdrawal
                    else:
                        shortfall = -excess
                        if not drip and cash_accumulated > 0:
                            cash_draw = min(cash_accumulated, shortfall)
                            cash_accumulated -= cash_draw
                            shortfall -= cash_draw
                        if shortfall > 0 and price > 0:
                            shares -= shortfall / price
                        cum_withdrawals += effective_withdrawal

                if shares <= 0:
                    shares = 0.0
                    depleted = True

                pv = shares * price
                growth = pv - prev_pv
                shares_delta = shares - shares_before
                total_basis = investment + (cash_wedge_initial if is_growth else 0)
                roi_dollar = pv + cum_withdrawals + wedge + cash_accumulated - total_basis
                roi_pct = roi_dollar / investment * 100 if investment else 0

                portfolio_values.append(round(pv, 2))
                cumulative_withdrawals_list.append(round(cum_withdrawals, 2))
                cumulative_distributions_list.append(round(cum_distributions, 2))
                share_counts.append(round(shares, 4))
                total_values.append(round(pv + cum_withdrawals + wedge + cash_accumulated, 2))
                row_data = {
                    "price": round(price, 2),
                    "shares": round(shares, 4),
                    "portfolio": round(pv, 2),
                    "dist_per_share": round(div_per_share, 4),
                    "income": round(dist_this_month, 2),
                    "withdrawal": round(effective_withdrawal, 2),
                    "wedge_drawn": round(wedge_drawn, 2),
                    "wedge_bal": round(wedge, 2),
                    "excess": round(excess, 2),
                    "shares_delta": round(shares_delta, 4),
                    "growth": round(growth, 2),
                    "cum_income": round(cum_distributions, 2),
                    "roi_dollar": round(roi_dollar, 2),
                    "roi_pct": round(roi_pct, 2),
                }
                if not drip:
                    row_data["cash_accumulated"] = round(cash_accumulated, 2)
                monthly_rows.append(row_data)
                prev_pv = pv

            date_labels = [dt.strftime("%Y-%m") for dt in df_m.index]
            total_bought = sum(r["shares_delta"] for r in monthly_rows if r["shares_delta"] > 0)
            total_sold = sum(-r["shares_delta"] for r in monthly_rows if r["shares_delta"] < 0)
            initial_shares = investment / initial_price
            role = "Growth" if is_growth else "Income"

            risk_metrics = _compute_risk_metrics(portfolio_values, monthly_rows)
            depletion_month = _compute_depletion_month(depleted, monthly_rows)

            fund_result = {
                "ticker": sym,
                "role": role,
                "investment": round(investment, 2),
                "initial_shares": round(initial_shares, 4),
                "has_cash_wedge": is_growth and cash_wedge_initial > 0,
                "cash_wedge_remaining": round(wedge, 2) if is_growth else None,
                "portfolio_values": portfolio_values,
                "cumulative_withdrawals": cumulative_withdrawals_list,
                "cumulative_distributions": cumulative_distributions_list,
                "total_values": total_values,
                "shares": share_counts,
                "monthly_rows": monthly_rows,
                "total_shares_bought": round(total_bought, 4),
                "total_shares_sold": round(total_sold, 4),
                "final_portfolio": portfolio_values[-1] if portfolio_values else 0.0,
                "final_withdrawn": round(cum_withdrawals, 2),
                "final_distributions": round(cum_distributions, 2),
                "final_total": total_values[-1] if total_values else 0.0,
                "depleted": depleted,
                "depletion_month": depletion_month,
                "risk_metrics": risk_metrics,
                "drip": drip,
            }
            if not drip:
                fund_result["cash_accumulated"] = round(cash_accumulated, 2)
            results[label] = fund_result

        results["months"] = date_labels
        results["data_start"] = common_start.strftime("%Y-%m-%d")
        results["cash_wedge_initial"] = cash_wedge_initial
        results["comparison_type"] = comparison_type
        return results

    # ── SIMULATION MODE ──
    if mode == "simulate":
        market_type = data.get("market", "neutral")
        duration_months = int(data.get("duration_months", 120))
        if duration_months < 1 or duration_months > 240:
            return {"error": "Simulation duration must be between 1 and 240 months (20 years max)."}

        bias_map = {"bullish": +0.010, "bearish": -0.015, "neutral": 0.0}
        vol_mult_map = {"bullish": 0.9, "bearish": 1.2, "neutral": 1.0}
        bias = bias_map.get(market_type, 0.0)
        vol_mult = vol_mult_map.get(market_type, 1.0)

        results = {}
        fund_list = [("fund_a", fa), ("fund_b", fb)]
        if fc:
            fund_list.append(("fund_c", fc))
        date_labels = []
        for label, fund in fund_list:
            sym = fund["ticker"]
            investment = fund["investment"]
            yo = fund["yield_override"]
            drip = fund["drip"]

            try:
                hist = yf.Ticker(sym).history(period="1y", auto_adjust=False, actions=True)
            except Exception as e:
                return {"error": f"Failed to fetch data for {sym}: {str(e)}"}

            if hist is None or hist.empty:
                return {"error": f"No data found for {sym}."}

            close = hist["Close"].dropna()
            if close.empty:
                return {"error": f"No price data for {sym}."}

            current_price = float(close.iloc[-1])
            if current_price <= 0:
                return {"error": f"Could not determine current price for {sym}."}

            divs = hist["Dividends"] if "Dividends" in hist.columns else pd.Series(0.0, index=hist.index)
            if yo is not None:
                ttm_yield = yo / 100.0
            else:
                ttm_divs_sum = float(divs.sum()) if divs is not None else 0.0
                ttm_yield = ttm_divs_sum / current_price if current_price > 0 else 0.0

            monthly_returns = close.resample("ME").last().pct_change().dropna()
            hist_sigma = float(monthly_returns.std()) if len(monthly_returns) >= 2 else 0.05

            SIGMA_CAP = 0.25
            REGIME_MONTHS = 42
            FADE_MONTHS = 6
            RALLY_PROB = 0.18
            RALLY_LEN_LO = 2
            RALLY_LEN_HI = 4

            neutral_bias = 0.0
            neutral_vol = 1.0

            mu_arr = np.empty(duration_months)
            vmul_arr = np.empty(duration_months)
            rally_remaining = 0

            for m in range(duration_months):
                if m < REGIME_MONTHS:
                    regime_w = 1.0
                elif m < REGIME_MONTHS + FADE_MONTHS:
                    regime_w = 1.0 - (m - REGIME_MONTHS) / FADE_MONTHS
                else:
                    regime_w = 0.0

                m_bias = bias * regime_w + neutral_bias * (1.0 - regime_w)
                m_vmul = vol_mult * regime_w + neutral_vol * (1.0 - regime_w)

                if market_type == "bearish" and regime_w > 0:
                    if rally_remaining > 0:
                        m_bias = abs(bias) * 0.6 * regime_w
                        m_vmul = 0.9 * regime_w + neutral_vol * (1.0 - regime_w)
                        rally_remaining -= 1
                    elif np.random.random() < RALLY_PROB:
                        rally_remaining = np.random.randint(RALLY_LEN_LO, RALLY_LEN_HI + 1)
                        m_bias = abs(bias) * 0.6 * regime_w
                        m_vmul = 0.9 * regime_w + neutral_vol * (1.0 - regime_w)
                        rally_remaining -= 1

                mu_arr[m] = m_bias
                vmul_arr[m] = m_vmul

            sigma_arr = np.minimum(hist_sigma * vmul_arr, SIGMA_CAP)

            N_PATHS = 300
            drift_arr = mu_arr - 0.5 * sigma_arr ** 2
            np.random.seed(None)
            Z = np.random.normal(0.0, 1.0, (N_PATHS, duration_months))
            log_rets = drift_arr[np.newaxis, :] + sigma_arr[np.newaxis, :] * Z
            cum_log = np.cumsum(np.hstack([np.zeros((N_PATHS, 1)), log_rets]), axis=1)
            price_floor = max(current_price * 0.0001, 1e-10)
            price_matrix = np.maximum(current_price * np.exp(cum_log), price_floor)
            prices = [float(v) for v in np.median(price_matrix, axis=0)]

            is_growth = label in growth_funds
            wedge = cash_wedge_initial if is_growth else 0.0
            shares = investment / current_price
            cum_withdrawals = 0.0
            cum_distributions = 0.0
            prev_pv = investment
            cash_accumulated = 0.0

            portfolio_values = []
            cumulative_withdrawals_list = []
            cumulative_distributions_list = []
            share_counts = []
            total_values = []
            monthly_rows = []
            depleted = False

            for month_index, price in enumerate(prices[1:]):
                if depleted or shares <= 0:
                    depleted = True
                    portfolio_values.append(0.0)
                    cumulative_withdrawals_list.append(round(cum_withdrawals, 2))
                    cumulative_distributions_list.append(round(cum_distributions, 2))
                    share_counts.append(0.0)
                    tv = cum_withdrawals + wedge + cash_accumulated
                    total_values.append(round(tv, 2))
                    row_data = {
                        "price": 0, "shares": 0, "portfolio": 0,
                        "dist_per_share": 0, "income": 0, "withdrawal": 0,
                        "wedge_drawn": 0, "wedge_bal": round(wedge, 2),
                        "excess": 0, "shares_delta": 0, "growth": 0,
                        "cum_income": round(cum_distributions, 2),
                        "roi_dollar": round(cum_withdrawals + wedge + cash_accumulated - investment - (cash_wedge_initial if is_growth else 0), 2),
                        "roi_pct": round(((cum_withdrawals + wedge + cash_accumulated - investment - (cash_wedge_initial if is_growth else 0)) / investment * 100), 2) if investment else 0,
                    }
                    if not drip:
                        row_data["cash_accumulated"] = round(cash_accumulated, 2)
                    monthly_rows.append(row_data)
                    continue

                pct_chg = (price - current_price) / current_price * 100
                if pct_chg >= 10:
                    factor = min(1.0 + (pct_chg - 10) * 0.02, 1.30)
                elif pct_chg >= -10:
                    factor = 1.0
                elif pct_chg >= -20:
                    factor = 0.85
                elif pct_chg >= -30:
                    factor = 0.70
                else:
                    factor = max(0.40, 1.0 + pct_chg * 0.02)

                shares_before = shares
                dist_per_share = (ttm_yield / 12) * price * factor
                dist_this_month = dist_per_share * shares
                cum_distributions += dist_this_month

                # Compute effective withdrawal for this month
                current_pv = shares * price
                effective_withdrawal = _compute_effective_withdrawal(
                    monthly_withdrawal, month_index, current_pv, investment)

                wedge_drawn = 0.0
                if is_growth:
                    need = effective_withdrawal
                    cw_draw = min(wedge, need)
                    wedge -= cw_draw
                    wedge_drawn = cw_draw
                    need -= cw_draw
                    div_used = min(dist_this_month, need)
                    need -= div_used
                    leftover_div = dist_this_month - div_used
                    if drip:
                        if leftover_div > 0 and price > 0:
                            shares += leftover_div / price
                    else:
                        cash_accumulated += leftover_div
                    if need > 0:
                        if not drip and cash_accumulated > 0:
                            cash_draw = min(cash_accumulated, need)
                            cash_accumulated -= cash_draw
                            need -= cash_draw
                        if need > 0 and price > 0:
                            shares -= need / price
                    cum_withdrawals += effective_withdrawal
                    excess = dist_this_month + cw_draw - effective_withdrawal
                else:
                    excess = dist_this_month - effective_withdrawal
                    if dist_this_month >= effective_withdrawal:
                        if drip:
                            if price > 0:
                                shares += excess / price
                        else:
                            cash_accumulated += excess
                        cum_withdrawals += effective_withdrawal
                    else:
                        shortfall = -excess
                        if not drip and cash_accumulated > 0:
                            cash_draw = min(cash_accumulated, shortfall)
                            cash_accumulated -= cash_draw
                            shortfall -= cash_draw
                        if shortfall > 0 and price > 0:
                            shares -= shortfall / price
                        cum_withdrawals += effective_withdrawal

                if shares <= 0:
                    shares = 0.0
                    depleted = True

                pv = shares * price
                growth = pv - prev_pv
                shares_delta = shares - shares_before
                total_basis = investment + (cash_wedge_initial if is_growth else 0)
                roi_dollar = pv + cum_withdrawals + wedge + cash_accumulated - total_basis
                roi_pct = roi_dollar / investment * 100 if investment else 0

                portfolio_values.append(round(pv, 2))
                cumulative_withdrawals_list.append(round(cum_withdrawals, 2))
                cumulative_distributions_list.append(round(cum_distributions, 2))
                share_counts.append(round(shares, 4))
                total_values.append(round(pv + cum_withdrawals + wedge + cash_accumulated, 2))
                row_data = {
                    "price": round(price, 2),
                    "shares": round(shares, 4),
                    "portfolio": round(pv, 2),
                    "dist_per_share": round(dist_per_share, 4),
                    "income": round(dist_this_month, 2),
                    "withdrawal": round(effective_withdrawal, 2),
                    "wedge_drawn": round(wedge_drawn, 2),
                    "wedge_bal": round(wedge, 2),
                    "excess": round(excess, 2),
                    "shares_delta": round(shares_delta, 4),
                    "growth": round(growth, 2),
                    "cum_income": round(cum_distributions, 2),
                    "roi_dollar": round(roi_dollar, 2),
                    "roi_pct": round(roi_pct, 2),
                }
                if not drip:
                    row_data["cash_accumulated"] = round(cash_accumulated, 2)
                monthly_rows.append(row_data)
                prev_pv = pv

            date_labels = [f"Month {i+1}" for i in range(duration_months)]
            total_bought = sum(r["shares_delta"] for r in monthly_rows if r["shares_delta"] > 0)
            total_sold = sum(-r["shares_delta"] for r in monthly_rows if r["shares_delta"] < 0)
            initial_shares = investment / current_price
            role = "Growth" if is_growth else "Income"

            risk_metrics = _compute_risk_metrics(portfolio_values, monthly_rows)
            depletion_month = _compute_depletion_month(depleted, monthly_rows)

            fund_result = {
                "ticker": sym,
                "role": role,
                "investment": round(investment, 2),
                "initial_shares": round(initial_shares, 4),
                "has_cash_wedge": is_growth and cash_wedge_initial > 0,
                "cash_wedge_remaining": round(wedge, 2) if is_growth else None,
                "portfolio_values": portfolio_values,
                "cumulative_withdrawals": cumulative_withdrawals_list,
                "cumulative_distributions": cumulative_distributions_list,
                "total_values": total_values,
                "shares": share_counts,
                "monthly_rows": monthly_rows,
                "total_shares_bought": round(total_bought, 4),
                "total_shares_sold": round(total_sold, 4),
                "final_portfolio": portfolio_values[-1] if portfolio_values else 0.0,
                "final_withdrawn": round(cum_withdrawals, 2),
                "final_distributions": round(cum_distributions, 2),
                "final_total": total_values[-1] if total_values else 0.0,
                "depleted": depleted,
                "depletion_month": depletion_month,
                "risk_metrics": risk_metrics,
                "drip": drip,
            }
            if not drip:
                fund_result["cash_accumulated"] = round(cash_accumulated, 2)
            results[label] = fund_result

        results["months"] = date_labels
        results["cash_wedge_initial"] = cash_wedge_initial
        results["comparison_type"] = comparison_type
        return results

    return {"error": f"Unknown mode: {mode}"}


# ── Consolidation Analysis ─────────────────────────────────────────────────────

@app.route("/api/consolidation/clusters", methods=["POST"])
def consolidation_clusters():
    """Auto-group holdings by underlying exposure using correlation analysis."""
    import math
    import warnings
    import numpy as np
    import yfinance as yf
    from scipy.cluster.hierarchy import linkage, fcluster
    from scipy.spatial.distance import squareform
    from collections import Counter
    warnings.filterwarnings("ignore")

    def _safe(v):
        if v is None:
            return None
        try:
            f = float(v)
            return None if (math.isnan(f) or math.isinf(f)) else round(f, 4)
        except (TypeError, ValueError):
            return None

    # Get threshold from request (default 0.80)
    body = request.get_json(silent=True) or {}
    corr_threshold = float(body.get("threshold", 0.80))
    corr_threshold = max(0.50, min(0.95, corr_threshold))  # clamp
    dist_threshold = 1.0 - corr_threshold

    _, pids = get_profile_filter()
    placeholders = ",".join("?" * len(pids))
    conn = get_connection()
    rows = conn.execute(
        f"SELECT ticker, description, quantity, current_value, current_price, "
        f"approx_monthly_income, current_annual_yield "
        f"FROM all_account_info WHERE profile_id IN ({placeholders}) "
        f"AND quantity > 0 AND current_value > 0",
        pids,
    ).fetchall()
    conn.close()

    if not rows:
        return jsonify(error="No holdings found for this profile.")

    # Aggregate rows by ticker (same ticker may appear for multiple owners)
    holdings = {}
    for r in rows:
        t = r["ticker"]
        if t in holdings:
            holdings[t]["quantity"] = (holdings[t]["quantity"] or 0) + (r["quantity"] or 0)
            holdings[t]["current_value"] = (holdings[t]["current_value"] or 0) + (r["current_value"] or 0)
            holdings[t]["approx_monthly_income"] = (holdings[t]["approx_monthly_income"] or 0) + (r["approx_monthly_income"] or 0)
            # Keep current_price and current_annual_yield from existing entry (same ticker, same price)
        else:
            holdings[t] = dict(r)
    # Recalculate yield from aggregated values
    for t, h in holdings.items():
        cv = h.get("current_value") or 0
        mi = h.get("approx_monthly_income") or 0
        h["current_annual_yield"] = (mi * 12 / cv * 100) if cv > 0 else 0
    tickers = list(holdings.keys())

    if len(tickers) < 2:
        # Not enough tickers to cluster
        unclustered = []
        for t in tickers:
            h = holdings[t]
            unclustered.append({
                "ticker": t,
                "description": h.get("description") or "",
                "quantity": _safe(h.get("quantity")),
                "current_value": _safe(h.get("current_value")),
                "monthly_income": _safe(h.get("approx_monthly_income")),
                "current_yield": _safe(h.get("current_annual_yield")),
            })
        return jsonify(clusters=[], unclustered=unclustered, correlation_matrix={})

    # Download 1 year of daily close prices
    try:
        raw = yf.download(" ".join(tickers), period="1y", auto_adjust=True, progress=False)
        if raw.empty:
            return jsonify(error="No price data returned from Yahoo Finance.")
    except Exception as e:
        return jsonify(error=f"Failed to fetch data: {str(e)}")

    if isinstance(raw.columns, pd.MultiIndex):
        close = raw["Close"].dropna(how="all")
    else:
        close = raw[["Close"]].dropna(how="all")
        close.columns = [tickers[0]]

    available = [t for t in tickers if t in close.columns and close[t].dropna().count() >= 30]
    unavailable = [t for t in tickers if t not in available]

    if len(available) < 2:
        unclustered = []
        for t in tickers:
            h = holdings[t]
            unclustered.append({
                "ticker": t,
                "description": h.get("description") or "",
                "quantity": _safe(h.get("quantity")),
                "current_value": _safe(h.get("current_value")),
                "monthly_income": _safe(h.get("approx_monthly_income")),
                "current_yield": _safe(h.get("current_annual_yield")),
            })
        return jsonify(clusters=[], unclustered=unclustered, correlation_matrix={})

    daily_returns = close[available].pct_change().dropna()
    corr = daily_returns.corr()

    # Build correlation matrix response
    corr_matrix = {}
    for t1 in available:
        corr_matrix[t1] = {}
        for t2 in available:
            corr_matrix[t1][t2] = _safe(corr.loc[t1, t2])

    # Hierarchical clustering
    # Convert correlation to distance: distance = 1 - correlation
    dist_matrix = 1 - corr.values
    np.fill_diagonal(dist_matrix, 0)
    # Ensure symmetry and no negative values
    dist_matrix = np.clip((dist_matrix + dist_matrix.T) / 2, 0, 2)

    try:
        condensed = squareform(dist_matrix)
        Z = linkage(condensed, method="average")
        labels = fcluster(Z, t=dist_threshold, criterion="distance")
    except Exception:
        # Fallback: treat every ticker as unclustered
        labels = list(range(1, len(available) + 1))

    # Group tickers by cluster label
    cluster_groups = {}
    for i, t in enumerate(available):
        lbl = int(labels[i])
        cluster_groups.setdefault(lbl, []).append(t)

    clusters = []
    unclustered = []
    cluster_id = 0

    for lbl, group_tickers in sorted(cluster_groups.items()):
        if len(group_tickers) < 2:
            # Single ticker = unclustered
            h = holdings[group_tickers[0]]
            unclustered.append({
                "ticker": group_tickers[0],
                "description": h.get("description") or "",
                "quantity": _safe(h.get("quantity")),
                "current_value": _safe(h.get("current_value")),
                "monthly_income": _safe(h.get("approx_monthly_income")),
                "current_yield": _safe(h.get("current_annual_yield")),
            })
            continue

        cluster_id += 1

        # Calculate avg correlation within group
        pair_corrs = []
        for i, t1 in enumerate(group_tickers):
            for t2 in group_tickers[i + 1:]:
                val = corr.loc[t1, t2]
                if not math.isnan(val):
                    pair_corrs.append(val)
        avg_corr = float(np.mean(pair_corrs)) if pair_corrs else 0.0

        # Identify underlying: most common meaningful words in descriptions
        all_words = []
        for t in group_tickers:
            desc = (holdings[t].get("description") or "").strip()
            if desc:
                # Filter out common filler words
                stop = {"etf", "fund", "inc", "corp", "the", "and", "of", "trust",
                        "a", "an", "for", "in", "on", "to", "with", "&", "-", ""}
                words = [w for w in desc.lower().split() if w not in stop and len(w) > 1]
                all_words.extend(words)

        if all_words:
            word_counts = Counter(all_words)
            # Take the top 2-3 most common words to form the underlying name
            common = word_counts.most_common(3)
            underlying = " ".join(w.title() for w, _ in common if _ > 1)
            if not underlying:
                underlying = common[0][0].title() if common else group_tickers[0]
        else:
            # Fallback: ticker with most history
            max_count = 0
            best = group_tickers[0]
            for t in group_tickers:
                cnt = close[t].dropna().count() if t in close.columns else 0
                if cnt > max_count:
                    max_count = cnt
                    best = t
            underlying = best

        # Build ticker detail list
        ticker_details = []
        total_value = 0.0
        total_income = 0.0
        for t in group_tickers:
            h = holdings[t]
            cv = float(h.get("current_value") or 0)
            mi = float(h.get("approx_monthly_income") or 0)
            total_value += cv
            total_income += mi

            # Correlation to group average
            group_others = [t2 for t2 in group_tickers if t2 != t]
            if group_others:
                corr_to_group = float(np.mean([corr.loc[t, t2] for t2 in group_others
                                                if not math.isnan(corr.loc[t, t2])]))
            else:
                corr_to_group = 1.0

            ticker_details.append({
                "ticker": t,
                "description": h.get("description") or "",
                "quantity": _safe(h.get("quantity")),
                "current_value": _safe(cv),
                "monthly_income": _safe(mi),
                "current_yield": _safe(h.get("current_annual_yield")),
                "correlation_to_group": _safe(corr_to_group),
            })

        clusters.append({
            "cluster_id": cluster_id,
            "underlying": underlying,
            "tickers": ticker_details,
            "avg_correlation": _safe(avg_corr),
            "total_value": _safe(total_value),
            "total_monthly_income": _safe(total_income),
        })

    # Add unavailable tickers to unclustered
    for t in unavailable:
        h = holdings[t]
        unclustered.append({
            "ticker": t,
            "description": h.get("description") or "",
            "quantity": _safe(h.get("quantity")),
            "current_value": _safe(h.get("current_value")),
            "monthly_income": _safe(h.get("approx_monthly_income")),
            "current_yield": _safe(h.get("current_annual_yield")),
        })

    # Add "nearest cluster" info for each unclustered ticker
    for item in unclustered:
        t = item["ticker"]
        if t not in corr_matrix:
            continue
        best_cluster = None
        best_corr = -1.0
        for c in clusters:
            cluster_tickers = [ct["ticker"] for ct in c["tickers"]]
            vals = []
            for ct in cluster_tickers:
                if ct in corr_matrix.get(t, {}):
                    v = corr_matrix[t][ct]
                    if v is not None:
                        vals.append(v)
            if vals:
                avg = sum(vals) / len(vals)
                if avg > best_corr:
                    best_corr = avg
                    best_cluster = c["underlying"]
        if best_cluster and best_corr > 0:
            item["nearest_cluster"] = best_cluster
            item["nearest_correlation"] = _safe(best_corr)

    return jsonify(
        clusters=clusters,
        unclustered=unclustered,
        correlation_matrix=corr_matrix,
    )


@app.route("/api/consolidation/simulate", methods=["POST"])
def consolidation_simulate():
    """Simulate selling one ticker and redistributing to another."""
    import math
    import warnings
    import numpy as np
    import yfinance as yf
    warnings.filterwarnings("ignore")

    def _safe(v):
        if v is None:
            return None
        try:
            f = float(v)
            return None if (math.isnan(f) or math.isinf(f)) else round(f, 4)
        except (TypeError, ValueError):
            return None

    data = request.get_json(force=True, silent=True) or {}
    sell_ticker = str(data.get("sell_ticker", "")).strip().upper()
    buy_ticker = str(data.get("buy_ticker", "")).strip().upper()
    period = data.get("period", "1y")

    if not sell_ticker or not buy_ticker:
        return jsonify(error="Both sell_ticker and buy_ticker are required."), 400
    if sell_ticker == buy_ticker:
        return jsonify(error="sell_ticker and buy_ticker must be different."), 400

    valid_periods = {"3mo", "6mo", "1y", "2y", "5y"}
    if period not in valid_periods:
        period = "1y"

    # Get current holdings data (aggregate across all owners in profile)
    _, pids = get_profile_filter()
    placeholders = ",".join("?" * len(pids))
    conn = get_connection()
    sell_rows = conn.execute(
        f"SELECT ticker, description, quantity, current_value, current_price, "
        f"approx_monthly_income, current_annual_yield "
        f"FROM all_account_info WHERE ticker = ? AND profile_id IN ({placeholders}) "
        f"AND quantity > 0",
        [sell_ticker] + pids,
    ).fetchall()
    buy_rows = conn.execute(
        f"SELECT ticker, description, quantity, current_value, current_price, "
        f"approx_monthly_income, current_annual_yield "
        f"FROM all_account_info WHERE ticker = ? AND profile_id IN ({placeholders}) "
        f"AND quantity > 0",
        [buy_ticker] + pids,
    ).fetchall()
    conn.close()

    if not sell_rows:
        return jsonify(error=f"No holding found for {sell_ticker}."), 404
    if not buy_rows:
        return jsonify(error=f"No holding found for {buy_ticker}."), 404

    # Aggregate across owners
    def _aggregate_rows(rows):
        agg = dict(rows[0])
        for r in rows[1:]:
            agg["quantity"] = (agg["quantity"] or 0) + (r["quantity"] or 0)
            agg["current_value"] = (agg["current_value"] or 0) + (r["current_value"] or 0)
            agg["approx_monthly_income"] = (agg["approx_monthly_income"] or 0) + (r["approx_monthly_income"] or 0)
        cv = agg.get("current_value") or 0
        mi = agg.get("approx_monthly_income") or 0
        agg["current_annual_yield"] = (mi * 12 / cv * 100) if cv > 0 else 0
        return agg

    sell_data = _aggregate_rows(sell_rows)
    buy_data = _aggregate_rows(buy_rows)

    sell_value = float(sell_data.get("current_value") or 0)
    sell_income = float(sell_data.get("approx_monthly_income") or 0)
    sell_yield = float(sell_data.get("current_annual_yield") or 0)
    buy_value = float(buy_data.get("current_value") or 0)
    buy_income = float(buy_data.get("approx_monthly_income") or 0)
    buy_yield = float(buy_data.get("current_annual_yield") or 0)
    buy_price = float(buy_data.get("current_price") or 0)

    # _aggregate_rows always returns yield as percentage (e.g. 8.59 = 8.59%)
    # Convert to decimal for income calculations
    sell_yield_dec = sell_yield / 100
    buy_yield_dec = buy_yield / 100

    if buy_price <= 0:
        return jsonify(error=f"Cannot simulate: {buy_ticker} has no valid price."), 400

    # Calculate consolidation impact
    new_shares_added = sell_value / buy_price
    new_total_value = sell_value + buy_value  # value doesn't change at moment of swap
    # New income from added shares: sell_value * (buy annual yield as decimal) / 12
    new_monthly_from_added = (sell_value * buy_yield_dec) / 12 if buy_yield_dec else 0
    new_monthly_income = buy_income + new_monthly_from_added
    old_combined_income = sell_income + buy_income
    income_change = new_monthly_income - old_combined_income
    income_change_pct = (income_change / old_combined_income * 100) if old_combined_income > 0 else 0
    new_yield = (new_monthly_income * 12 / new_total_value * 100) if new_total_value > 0 else 0

    # Download historical data for performance comparison
    both_tickers = list(set([sell_ticker, buy_ticker]))
    try:
        raw = yf.download(" ".join(both_tickers), period=period, auto_adjust=True,
                          actions=True, progress=False)
        if raw.empty:
            return jsonify(error="No historical data returned from Yahoo Finance.")
    except Exception as e:
        return jsonify(error=f"Failed to fetch data: {str(e)}")

    if isinstance(raw.columns, pd.MultiIndex):
        close = raw["Close"].dropna(how="all")
        # Try to get dividends
        try:
            divs = raw["Dividends"].fillna(0)
        except KeyError:
            divs = pd.DataFrame(0, index=close.index, columns=close.columns)
    else:
        close = raw[["Close"]].dropna(how="all")
        close.columns = [both_tickers[0]]
        try:
            divs = raw[["Dividends"]].fillna(0)
            divs.columns = [both_tickers[0]]
        except KeyError:
            divs = pd.DataFrame(0, index=close.index, columns=[both_tickers[0]])

    # Ensure both tickers present
    for t in [sell_ticker, buy_ticker]:
        if t not in close.columns:
            return jsonify(error=f"No price data available for {t}.")
        if t not in divs.columns:
            divs[t] = 0

    # Calculate performance metrics for each ticker
    perf = {}
    history = {"dates": [], "sell_total_return": [], "buy_total_return": [],
               "sell_price_return": [], "buy_price_return": []}

    for label, t in [("sell_ticker", sell_ticker), ("buy_ticker", buy_ticker)]:
        tc = close[t].dropna()
        if len(tc) < 2:
            perf[label] = {"total_return": None, "price_return": None,
                           "volatility": None, "max_drawdown": None, "sharpe": None}
            continue

        prices = tc.values
        start_price = float(prices[0])

        # Price return series
        price_ret_series = (prices / start_price - 1) * 100

        # Total return (including dividends reinvested)
        td = divs[t].reindex(tc.index).fillna(0)
        cum_shares = 1.0
        total_ret_series = np.zeros(len(tc))
        for i in range(len(tc)):
            d = float(td.iloc[i])
            p = float(prices[i])
            if p > 0 and d > 0:
                cum_shares += d * cum_shares / p
            total_ret_series[i] = (cum_shares * p / start_price - 1) * 100

        # Daily returns for volatility/sharpe
        daily_ret = np.diff(prices) / prices[:-1]
        vol = float(np.std(daily_ret) * np.sqrt(252) * 100) if len(daily_ret) > 1 else 0

        # Max drawdown
        running_max = np.maximum.accumulate(prices)
        drawdowns = (prices - running_max) / running_max * 100
        max_dd = float(np.min(drawdowns))

        # Sharpe (annualized, assuming 0 risk-free rate)
        mean_daily = float(np.mean(daily_ret)) if len(daily_ret) > 0 else 0
        std_daily = float(np.std(daily_ret)) if len(daily_ret) > 0 else 1
        sharpe = (mean_daily / std_daily * np.sqrt(252)) if std_daily > 0 else 0

        perf[label] = {
            "total_return": _safe(total_ret_series[-1]),
            "price_return": _safe(price_ret_series[-1]),
            "volatility": _safe(vol),
            "max_drawdown": _safe(max_dd),
            "sharpe": _safe(sharpe),
        }

        # Build history arrays
        dates_str = [d.strftime("%Y-%m-%d") for d in tc.index]
        if label == "sell_ticker":
            history["dates"] = dates_str
            history["sell_total_return"] = [_safe(v) for v in total_ret_series]
            history["sell_price_return"] = [_safe(v) for v in price_ret_series]
        else:
            # Align buy to same dates if needed (use its own dates)
            history["buy_total_return"] = [_safe(v) for v in total_ret_series]
            history["buy_price_return"] = [_safe(v) for v in price_ret_series]

    # If dates differ between tickers, use the intersection
    # For simplicity, use sell_ticker dates as base (already set)

    return jsonify(
        sell={
            "ticker": sell_ticker,
            "current_value": _safe(sell_value),
            "monthly_income": _safe(sell_income),
            "current_yield": _safe(sell_yield),
        },
        buy={
            "ticker": buy_ticker,
            "current_value": _safe(buy_value),
            "monthly_income": _safe(buy_income),
            "current_yield": _safe(buy_yield),
        },
        after_consolidation={
            "new_shares_added": _safe(new_shares_added),
            "new_total_value": _safe(new_total_value),
            "new_monthly_income": _safe(new_monthly_income),
            "income_change": _safe(income_change),
            "income_change_pct": _safe(income_change_pct),
            "new_yield": _safe(new_yield),
        },
        performance_comparison=perf,
        history=history,
    )


@app.route("/api/consolidation/regimes", methods=["POST"])
def consolidation_regimes():
    """Analyze performance by market regime (bull/bear/sideways + high vol overlay)."""
    import math
    import warnings
    import numpy as np
    import yfinance as yf
    warnings.filterwarnings("ignore")

    def _safe(v):
        if v is None:
            return None
        try:
            f = float(v)
            return None if (math.isnan(f) or math.isinf(f)) else round(f, 4)
        except (TypeError, ValueError):
            return None

    data = request.get_json(force=True, silent=True) or {}
    tickers = [str(t).strip().upper() for t in data.get("tickers", []) if str(t).strip()]
    period = data.get("period", "2y")

    if not tickers:
        return jsonify(error="No tickers provided."), 400
    if len(tickers) > 20:
        return jsonify(error="Maximum 20 tickers allowed."), 400

    valid_periods = {"6mo", "1y", "2y", "5y", "max"}
    if period not in valid_periods:
        period = "2y"

    # Download SPY, VIX, and all requested tickers
    all_dl = list(set(tickers + ["SPY", "^VIX"]))
    try:
        raw = yf.download(" ".join(all_dl), period=period, auto_adjust=True,
                          actions=True, progress=False)
        if raw.empty:
            return jsonify(error="No price data returned from Yahoo Finance.")
    except Exception as e:
        return jsonify(error=f"Failed to fetch data: {str(e)}")

    if isinstance(raw.columns, pd.MultiIndex):
        close = raw["Close"].dropna(how="all")
        try:
            divs = raw["Dividends"].fillna(0)
        except KeyError:
            divs = pd.DataFrame(0, index=close.index, columns=close.columns)
    else:
        close = raw[["Close"]].dropna(how="all")
        close.columns = [all_dl[0]]
        try:
            divs = raw[["Dividends"]].fillna(0)
            divs.columns = [all_dl[0]]
        except KeyError:
            divs = pd.DataFrame(0, index=close.index, columns=[all_dl[0]])

    if "SPY" not in close.columns:
        return jsonify(error="Could not fetch SPY benchmark data.")

    # VIX data
    has_vix = "^VIX" in close.columns
    if has_vix:
        vix_series = close["^VIX"].ffill()
    else:
        vix_series = pd.Series(20.0, index=close.index)

    spy = close["SPY"].dropna()
    if len(spy) < 63:
        return jsonify(error="Not enough SPY data to calculate regimes (need 63+ trading days).")

    # Calculate SPY 63-day rolling return (percentage)
    spy_rolling = spy.pct_change(63) * 100
    spy_rolling = spy_rolling.reindex(close.index)

    # Classify regimes
    regime_series = pd.Series("sideways", index=close.index)
    regime_series[spy_rolling > 5] = "bull"
    regime_series[spy_rolling < -5] = "bear"
    # First 63 days have no rolling return; label as sideways
    regime_series.iloc[:63] = "sideways"

    # High vol overlay
    high_vol = vix_series > 25

    # Count regime days
    total_days = len(regime_series)
    regime_counts = regime_series.value_counts()
    regimes_summary = {}
    for r in ["bull", "bear", "sideways"]:
        days = int(regime_counts.get(r, 0))
        regimes_summary[r] = {
            "days": days,
            "pct_of_period": _safe(days / total_days * 100) if total_days > 0 else 0,
        }

    # Analyze each ticker per regime
    data_warnings = {}
    ticker_performance = {}

    for t in tickers:
        if t not in close.columns:
            data_warnings[t] = "No price data available"
            continue

        tc = close[t].dropna()
        if len(tc) < 30:
            data_warnings[t] = f"Only {len(tc)} days of history available"
            continue

        # Note if limited history
        expected_days = len(close.index)
        actual_days = len(tc)
        if actual_days < expected_days * 0.8:
            months_avail = round(actual_days / 21)
            data_warnings[t] = f"Only {months_avail} months of history available"

        td = divs[t].reindex(tc.index).fillna(0) if t in divs.columns else pd.Series(0, index=tc.index)

        ticker_perf = {}

        for regime_name in ["bull", "bear", "sideways", "high_vol"]:
            if regime_name == "high_vol":
                mask = high_vol.reindex(tc.index).fillna(False)
            else:
                mask = (regime_series.reindex(tc.index) == regime_name)

            regime_prices = tc[mask]
            regime_divs = td[mask]

            if len(regime_prices) < 2:
                ticker_perf[regime_name] = {
                    "price_return": None, "income_return": None,
                    "total_return": None, "max_drawdown": None, "volatility": None,
                }
                continue

            # Daily returns within regime days
            prices_arr = regime_prices.values
            daily_rets = np.diff(prices_arr) / prices_arr[:-1]
            daily_rets = daily_rets[np.isfinite(daily_rets)]

            if len(daily_rets) < 1:
                ticker_perf[regime_name] = {
                    "price_return": None, "income_return": None,
                    "total_return": None, "max_drawdown": None, "volatility": None,
                }
                continue

            # Annualized price return
            n_days = len(regime_prices)
            cumulative_price_ret = (float(prices_arr[-1]) / float(prices_arr[0]) - 1)
            ann_factor = 252 / n_days if n_days > 0 else 1
            ann_price_ret = cumulative_price_ret * ann_factor * 100

            # Income return: sum of dividends / average price, annualized
            total_div = float(regime_divs.sum())
            avg_price = float(regime_prices.mean())
            if avg_price > 0 and n_days > 0:
                income_ret_period = total_div / avg_price
                ann_income_ret = income_ret_period * ann_factor * 100
            else:
                ann_income_ret = 0

            ann_total_ret = ann_price_ret + ann_income_ret

            # Volatility (annualized)
            vol = float(np.std(daily_rets) * np.sqrt(252) * 100) if len(daily_rets) > 1 else 0

            # Max drawdown within regime periods
            running_max = np.maximum.accumulate(prices_arr)
            drawdowns = (prices_arr - running_max) / running_max * 100
            max_dd = float(np.min(drawdowns))

            ticker_perf[regime_name] = {
                "price_return": _safe(ann_price_ret),
                "income_return": _safe(ann_income_ret),
                "total_return": _safe(ann_total_ret),
                "max_drawdown": _safe(max_dd),
                "volatility": _safe(vol),
            }

        ticker_performance[t] = ticker_perf

    # Build timeline for visualization
    timeline = {
        "dates": [d.strftime("%Y-%m-%d") for d in close.index],
        "regime": [regime_series.iloc[i] for i in range(len(regime_series))],
        "spy_price": [_safe(float(spy.reindex(close.index).iloc[i]))
                      if i < len(spy.reindex(close.index)) and not math.isnan(spy.reindex(close.index).iloc[i])
                      else None
                      for i in range(len(close.index))],
        "vix": [_safe(float(vix_series.iloc[i]))
                if not math.isnan(vix_series.iloc[i])
                else None
                for i in range(len(close.index))],
    }

    return jsonify(
        regimes=regimes_summary,
        ticker_performance=ticker_performance,
        timeline=timeline,
        data_warnings=data_warnings,
    )


# ── Macro Regime Dashboard ─────────────────────────────────────────────────────

# Macro proxy tickers fetched from yfinance
MACRO_TICKERS = {
    "tip":        "TIP",       # iShares TIPS ETF (inflation-linked)
    "ief":        "IEF",       # 7-10yr Treasury ETF (nominal bonds)
    "oil":        "CL=F",      # WTI Crude futures
    "rates_10y":  "^TNX",      # 10-Year Treasury Yield
    "rates_short":"^IRX",      # 13-Week T-Bill yield
    "usd":        "DX-Y.NYB",  # US Dollar Index
    "gold":       "GC=F",      # Gold futures
    "vix":        "^VIX",      # Volatility Index
    "spy":        "SPY",       # S&P 500 ETF
    "xli":        "XLI",       # Industrial Select Sector SPDR (growth proxy)
}

# ── 4-Quadrant regime model ──────────────────────────────────────────────────
QUADRANT_NAMES = {1: "Goldilocks", 2: "Reflation", 3: "Stagflation", 4: "Deflation"}
QUADRANT_DESCRIPTIONS = {
    1: "Growth UP + Inflation DOWN: Favors equities, tech, growth stocks",
    2: "Growth UP + Inflation UP: Favors commodities, energy, equities",
    3: "Growth DOWN + Inflation UP: Favors gold, TIPS, utilities",
    4: "Growth DOWN + Inflation DOWN: Favors long-term bonds, cash, defensives",
}
QUADRANT_ASSET_TILTS = {
    1: {"Tech/Growth": "Best", "Commodities": "Avoid", "Gold": "Neutral",
        "Long-Treasuries": "Neutral", "Healthcare/Staples": "Underperform"},
    2: {"Tech/Growth": "Neutral", "Commodities": "Best", "Gold": "Good",
        "Long-Treasuries": "Avoid", "Healthcare/Staples": "Neutral"},
    3: {"Tech/Growth": "Avoid", "Commodities": "Good", "Gold": "Best",
        "Long-Treasuries": "Avoid", "Healthcare/Staples": "Good"},
    4: {"Tech/Growth": "Avoid", "Commodities": "Avoid", "Gold": "Neutral",
        "Long-Treasuries": "Best", "Healthcare/Staples": "Good"},
}

# Pillar classification -> macro sensitivity tags
MACRO_SENSITIVITY_BY_PILLAR = {
    "HA":  ["rate_sensitive_negative", "inflation_neutral"],
    "A":   ["rate_sensitive_negative", "inflation_negative"],
    "GS":  ["inflation_benefiting", "commodity_linked", "safe_haven"],
    "B":   ["rate_sensitive_mild", "inflation_neutral"],
    "J":   ["rate_sensitive_negative", "inflation_negative"],
    "BDC": ["rate_sensitive_positive", "inflation_neutral"],
    "G":   ["growth_equity", "rate_sensitive_negative"],
}

# yfinance sector -> macro sensitivity tags (Tier 2 fallback)
MACRO_SENSITIVITY_BY_SECTOR = {
    "Energy":                 ["commodity_linked", "inflation_benefiting"],
    "Basic Materials":        ["commodity_linked", "inflation_benefiting"],
    "Real Estate":            ["rate_sensitive_negative", "inflation_neutral"],
    "Utilities":              ["rate_sensitive_negative", "inflation_negative"],
    "Financial Services":     ["rate_sensitive_positive", "inflation_neutral"],
    "Financials":             ["rate_sensitive_positive", "inflation_neutral"],
    "Technology":             ["growth_equity", "rate_sensitive_negative"],
    "Communication Services": ["growth_equity", "rate_sensitive_negative"],
    "Consumer Cyclical":      ["growth_equity", "inflation_neutral"],
    "Consumer Defensive":     ["inflation_neutral", "rate_sensitive_mild"],
    "Healthcare":             ["inflation_neutral", "rate_sensitive_mild"],
    "Industrials":            ["inflation_neutral", "rate_sensitive_mild"],
}

# yfinance fund category -> macro sensitivity tags (Tier 2 for ETFs/funds)
MACRO_SENSITIVITY_BY_CATEGORY = {
    "Inflation-Protected Bond":  ["inflation_benefiting", "rate_sensitive_mild"],
    "Commodities Broad Basket":  ["commodity_linked", "inflation_benefiting"],
    "Commodities Focused":       ["commodity_linked", "inflation_benefiting"],
    "Equity Energy":             ["commodity_linked", "inflation_benefiting"],
    "Equity Precious Metals":    ["safe_haven", "inflation_benefiting"],
    "High Yield Bond":           ["rate_sensitive_negative", "inflation_negative"],
    "Bank Loan":                 ["rate_sensitive_positive", "inflation_neutral"],
    "Floating Rate":             ["rate_sensitive_positive", "inflation_neutral"],
    "Long Government":           ["rate_sensitive_negative", "inflation_negative"],
    "Long-Term Bond":            ["rate_sensitive_negative", "inflation_negative"],
    "Short Government":          ["rate_sensitive_mild", "inflation_neutral"],
    "Short-Term Bond":           ["rate_sensitive_mild", "inflation_neutral"],
    "Ultrashort Bond":           ["rate_sensitive_mild", "inflation_neutral"],
    "Real Estate":               ["rate_sensitive_negative", "inflation_neutral"],
    "Large Growth":              ["growth_equity", "rate_sensitive_negative"],
    "Large Blend":               ["growth_equity", "rate_sensitive_mild"],
    "Large Value":               ["inflation_neutral", "rate_sensitive_mild"],
    "Mid-Cap Growth":            ["growth_equity", "rate_sensitive_negative"],
    "Small Growth":              ["growth_equity", "rate_sensitive_negative"],
    "Diversified Emerging Mkts": ["growth_equity", "commodity_linked"],
}

# Macro score weights per sensitivity tag, keyed by regime component
# Positive = favorable in that condition, negative = unfavorable
MACRO_SCORE_WEIGHTS = {
    "inflation_rising": {
        "inflation_benefiting": 1.0, "inflation_negative": -1.0,
        "inflation_neutral": 0.0, "commodity_linked": 0.5,
        "safe_haven": 0.3, "rate_sensitive_positive": 0.3,
        "rate_sensitive_negative": -0.3, "rate_sensitive_mild": -0.1,
        "growth_equity": -0.5, "unclassified": 0.0,
    },
    "inflation_falling": {
        "inflation_benefiting": -0.3, "inflation_negative": 0.5,
        "inflation_neutral": 0.0, "commodity_linked": -0.3,
        "safe_haven": -0.1, "rate_sensitive_positive": -0.3,
        "rate_sensitive_negative": 0.5, "rate_sensitive_mild": 0.1,
        "growth_equity": 0.7, "unclassified": 0.0,
    },
    "rates_rising": {
        "rate_sensitive_positive": 1.0, "rate_sensitive_negative": -1.0,
        "rate_sensitive_mild": -0.3, "inflation_benefiting": 0.2,
        "inflation_negative": -0.3, "inflation_neutral": 0.0,
        "commodity_linked": 0.1, "safe_haven": -0.2,
        "growth_equity": -0.7, "unclassified": 0.0,
    },
    "rates_falling": {
        "rate_sensitive_positive": -0.5, "rate_sensitive_negative": 0.8,
        "rate_sensitive_mild": 0.2, "inflation_benefiting": 0.0,
        "inflation_negative": 0.3, "inflation_neutral": 0.0,
        "commodity_linked": 0.0, "safe_haven": 0.3,
        "growth_equity": 0.8, "unclassified": 0.0,
    },
    "oil_rising": {
        "commodity_linked": 1.0, "inflation_benefiting": 0.5,
        "inflation_negative": -0.3, "inflation_neutral": 0.0,
        "rate_sensitive_positive": 0.1, "rate_sensitive_negative": -0.2,
        "rate_sensitive_mild": 0.0, "safe_haven": 0.2,
        "growth_equity": -0.3, "unclassified": 0.0,
    },
    "oil_falling": {
        "commodity_linked": -0.5, "inflation_benefiting": -0.2,
        "inflation_negative": 0.3, "inflation_neutral": 0.0,
        "rate_sensitive_positive": 0.0, "rate_sensitive_negative": 0.1,
        "rate_sensitive_mild": 0.0, "safe_haven": -0.1,
        "growth_equity": 0.5, "unclassified": 0.0,
    },
}

# In-memory cache for macro conditions (TTL-based)
# Candidate ETFs to suggest when user lacks exposure in a sensitivity category
CANDIDATE_ETFS = {
    "inflation_benefiting": [
        {"ticker": "TIP",  "name": "iShares TIPS Bond ETF"},
        {"ticker": "SCHP", "name": "Schwab U.S. TIPS ETF"},
        {"ticker": "VTIP", "name": "Vanguard Short-Term Inflation-Protected Securities"},
        {"ticker": "DJP",  "name": "iPath Bloomberg Commodity Index"},
        {"ticker": "PDBC", "name": "Invesco Optimum Yield Diversified Commodity Strategy"},
    ],
    "commodity_linked": [
        {"ticker": "DBC",  "name": "Invesco DB Commodity Index Tracking Fund"},
        {"ticker": "GSG",  "name": "iShares S&P GSCI Commodity-Indexed Trust"},
        {"ticker": "XLE",  "name": "State Street Energy Select Sector SPDR Fund"},
        {"ticker": "XLEI", "name": "YieldMax XLE Option Income Strategy ETF"},
        {"ticker": "MLPI", "name": "NEOS MLP & Energy Infrastructure High Income ETF"},
        {"ticker": "USO",  "name": "United States Oil Fund"},
        {"ticker": "PDBC", "name": "Invesco Optimum Yield Diversified Commodity Strategy"},
    ],
    "safe_haven": [
        {"ticker": "GLD",  "name": "SPDR Gold Shares"},
        {"ticker": "IAU",  "name": "iShares Gold Trust"},
        {"ticker": "SLV",  "name": "iShares Silver Trust"},
        {"ticker": "GLDM", "name": "SPDR Gold MiniShares Trust"},
        {"ticker": "SGOL", "name": "Aberdeen Standard Physical Gold Shares ETF"},
    ],
    "rate_sensitive_positive": [
        {"ticker": "SRLN", "name": "SPDR Blackstone Senior Loan ETF"},
        {"ticker": "BKLN", "name": "Invesco Senior Loan ETF"},
        {"ticker": "FLOT", "name": "iShares Floating Rate Bond ETF"},
        {"ticker": "ARCC", "name": "Ares Capital Corporation"},
        {"ticker": "MAIN", "name": "Main Street Capital"},
        {"ticker": "BIZD", "name": "VanEck BDC Income ETF"},
    ],
    "rate_sensitive_negative": [
        {"ticker": "TLT",  "name": "iShares 20+ Year Treasury Bond ETF"},
        {"ticker": "IEF",  "name": "iShares 7-10 Year Treasury Bond ETF"},
        {"ticker": "VGLT", "name": "Vanguard Long-Term Treasury ETF"},
        {"ticker": "AGG",  "name": "iShares Core U.S. Aggregate Bond ETF"},
        {"ticker": "BND",  "name": "Vanguard Total Bond Market ETF"},
    ],
    "growth_equity": [
        {"ticker": "QQQ",  "name": "Invesco QQQ Trust"},
        {"ticker": "SCHG", "name": "Schwab U.S. Large-Cap Growth ETF"},
        {"ticker": "VUG",  "name": "Vanguard Growth ETF"},
        {"ticker": "IWF",  "name": "iShares Russell 1000 Growth ETF"},
        {"ticker": "MGK",  "name": "Vanguard Mega Cap Growth ETF"},
    ],
    "rate_sensitive_mild": [
        {"ticker": "SCHD", "name": "Schwab U.S. Dividend Equity ETF"},
        {"ticker": "VYM",  "name": "Vanguard High Dividend Yield ETF"},
        {"ticker": "HDV",  "name": "iShares Core High Dividend ETF"},
        {"ticker": "JEPI", "name": "JPMorgan Equity Premium Income ETF"},
        {"ticker": "DIVO", "name": "Amplify CWP Enhanced Dividend Income ETF"},
    ],
    "inflation_neutral": [
        {"ticker": "JEPI", "name": "JPMorgan Equity Premium Income ETF"},
        {"ticker": "SCHD", "name": "Schwab U.S. Dividend Equity ETF"},
        {"ticker": "VIG",  "name": "Vanguard Dividend Appreciation ETF"},
    ],
}

_macro_cache = {"data": None, "timestamp": 0, "ttl": 1800}  # 30 min TTL
_quadrant_cache = {"data": None, "timestamp": 0, "ttl": 1800}  # 30 min TTL

# In-memory cache for yfinance ticker info (sector/category) — 24hr TTL
_ticker_info_cache = {}


def _get_ticker_sensitivity(ticker, classification_type, description="", overrides=None):
    """Determine macro sensitivity tags for a ticker using tiered fallback.
    Tier 0: User override (macro_overrides table)
    Tier 1: Pillar classification
    Tier 2: yfinance sector/category
    Tier 2.5: Name/description heuristics (yfinance name + DB description)
    Tier 3: Neutral default
    Returns (tags_list, source_label).
    """
    import yfinance as yf

    # Tier 0: User override
    if overrides and ticker in overrides:
        return overrides[ticker], "Override"

    # Tier 1: Pillar classification (skip generic values like "ETF", "Stock")
    if classification_type:
        ct = classification_type.strip().upper()
        skip_values = {"ETF", "STOCK", "EQUITY", "FUND", "CEF", "BOND", ""}
        if ct not in skip_values and ct in MACRO_SENSITIVITY_BY_PILLAR:
            return MACRO_SENSITIVITY_BY_PILLAR[ct], "Pillar"

    # Tier 2: yfinance metadata (cached)
    cache_entry = _ticker_info_cache.get(ticker)
    now = time.time()
    if cache_entry and (now - cache_entry["ts"]) < 86400:
        info = cache_entry["info"]
    else:
        try:
            info = yf.Ticker(ticker).info or {}
            _ticker_info_cache[ticker] = {"info": info, "ts": now}
        except Exception:
            info = {}
            _ticker_info_cache[ticker] = {"info": info, "ts": now}

    # Check fund category first (more specific for ETFs)
    category = info.get("category") or ""
    for cat_key, tags in MACRO_SENSITIVITY_BY_CATEGORY.items():
        if cat_key.lower() in category.lower():
            return tags, "Category"

    # Check sector
    sector = info.get("sector") or ""
    if sector in MACRO_SENSITIVITY_BY_SECTOR:
        return MACRO_SENSITIVITY_BY_SECTOR[sector], "Sector"

    # Tier 2.5: Name heuristics — check BOTH yfinance name AND the DB description
    long_name = (info.get("longName") or info.get("shortName") or "").lower()
    db_desc = (description or "").lower()
    # Combine both for broader matching
    combined = long_name + " " + db_desc + " " + ticker.lower()

    # Gold / Silver / Precious metals
    if any(w in combined for w in ["gold", "silver", "precious", "mining"]):
        return ["safe_haven", "inflation_benefiting", "commodity_linked"], "Name"
    # Energy / Oil / MLP / Pipeline
    if any(w in combined for w in ["oil", "energy", "petroleum", "crude", "mlp",
                                    "pipeline", "infrastructure"]):
        return ["commodity_linked", "inflation_benefiting"], "Name"
    # Inflation-protected
    if any(w in combined for w in ["treasury inflation", "tips", "inflation-protected",
                                    "inflation protected"]):
        return ["inflation_benefiting", "rate_sensitive_mild"], "Name"
    # Floating rate / BDC / Senior loans
    if any(w in combined for w in ["floating rate", "bank loan", "senior loan", "bdc",
                                    "business development"]):
        return ["rate_sensitive_positive", "inflation_neutral"], "Name"
    # Credit / High yield
    if any(w in combined for w in ["high yield", "credit opportunit", "junk bond"]):
        return ["rate_sensitive_negative", "inflation_negative"], "Name"
    # Preferred stock
    if any(w in combined for w in ["preferred", "pffd", "pff"]):
        return ["rate_sensitive_negative", "inflation_negative"], "Name"
    # Real estate / REIT
    if any(w in combined for w in ["real estate", "reit", "realty"]):
        return ["rate_sensitive_negative", "inflation_neutral"], "Name"
    # Utilities
    if any(w in combined for w in ["utilit"]):
        return ["rate_sensitive_negative", "inflation_negative"], "Name"
    # Bitcoin / Crypto
    if any(w in combined for w in ["bitcoin", "btc", "crypto", "ethereum"]):
        return ["growth_equity", "inflation_neutral"], "Name"
    # Nuclear
    if any(w in combined for w in ["nuclear", "uranium"]):
        return ["commodity_linked", "inflation_neutral"], "Name"
    # Covered call / Option income on broad indices (S&P, Nasdaq, Russell)
    if any(w in combined for w in ["s&p", "s & p", "spy", "500"]) and \
       any(w in combined for w in ["income", "inc ", "option", "covered", "premium",
                                    "yield", "dividend", "enhanced", "lift"]):
        return ["rate_sensitive_negative", "inflation_negative"], "Name"
    if any(w in combined for w in ["nasdaq", "qqq", "100", "tech", "innov"]) and \
       any(w in combined for w in ["income", "inc ", "option", "covered", "premium",
                                    "yield", "enhanced", "lift"]):
        return ["rate_sensitive_negative", "inflation_negative"], "Name"
    if any(w in combined for w in ["russell", "2000", "small"]) and \
       any(w in combined for w in ["income", "inc ", "option", "covered", "premium",
                                    "yield", "lift"]):
        return ["rate_sensitive_negative", "inflation_negative"], "Name"
    # Tesla / NVDA / single-stock option ETFs (juicer-like)
    if any(w in combined for w in ["tesla", "tsla", "nvd", "amazon", "amzn", "google",
                                    "goog", "netflix", "nflx", "apple", "aapl", "meta"]) and \
       any(w in combined for w in ["income", "inc ", "option", "covered", "premium",
                                    "yield", "lift"]):
        return ["rate_sensitive_negative", "inflation_negative"], "Name"
    # Generic option/income/dividend/yield ETFs (broad match — anchor-like)
    if any(w in combined for w in ["option income", "high income", "premium income",
                                    "covered call", "enhanced dividend", "enhanced div",
                                    "equity premium", "yieldmax", "yield prem",
                                    "tappalpha", "tapp alpha"]):
        return ["rate_sensitive_negative", "inflation_negative"], "Name"
    # Global equity funds
    if any(w in combined for w in ["global equity", "world equity", "international equity",
                                    "global fund", "eafe"]):
        return ["growth_equity", "rate_sensitive_mild"], "Name"
    # Generic dividend / equity income
    if any(w in combined for w in ["dividend", "div income", "equity income"]):
        return ["inflation_neutral", "rate_sensitive_mild"], "Name"
    # Growth / large cap growth
    if any(w in combined for w in ["growth", "large cap", "large-cap", "mega cap"]):
        return ["growth_equity", "rate_sensitive_negative"], "Name"

    # Tier 3: Unclassified
    return ["unclassified"], "Unclassified"


@app.route("/api/macro/conditions", methods=["POST"])
def macro_conditions():
    """Fetch current macro indicators and determine the macro regime."""
    import math
    import warnings
    import numpy as np
    import yfinance as yf
    warnings.filterwarnings("ignore")

    now = time.time()
    if _macro_cache["data"] and (now - _macro_cache["timestamp"]) < _macro_cache["ttl"]:
        return jsonify(_macro_cache["data"])

    def _safe(v):
        if v is None:
            return None
        try:
            f = float(v)
            return None if (math.isnan(f) or math.isinf(f)) else round(f, 4)
        except (TypeError, ValueError):
            return None

    # Download all macro proxy tickers
    all_symbols = list(MACRO_TICKERS.values())
    try:
        raw = yf.download(" ".join(all_symbols), period="1y", auto_adjust=True, progress=False)
        if raw.empty:
            return jsonify(error="No macro data returned from Yahoo Finance.")
    except Exception as e:
        return jsonify(error=f"Failed to fetch macro data: {str(e)}")

    if isinstance(raw.columns, pd.MultiIndex):
        close = raw["Close"].dropna(how="all")
    else:
        close = raw[["Close"]].dropna(how="all")
        close.columns = [all_symbols[0]]

    # Compute TIP/IEF spread as inflation proxy
    has_tip = "TIP" in close.columns and close["TIP"].dropna().count() > 30
    has_ief = "IEF" in close.columns and close["IEF"].dropna().count() > 30
    if has_tip and has_ief:
        close["INFLATION_SPREAD"] = close["TIP"] / close["IEF"]
    else:
        close["INFLATION_SPREAD"] = pd.Series(dtype=float)

    # Compute 63-day (3-month) rolling percentage change for direction
    LOOKBACK = 63
    RISING_THRESHOLD = 3.0    # % change to qualify as "rising"
    FALLING_THRESHOLD = -3.0  # % change to qualify as "falling"
    # For rates/VIX use absolute point change thresholds
    RATE_RISE_THRESH = 0.3    # 30 bps
    RATE_FALL_THRESH = -0.3

    def direction_pct(series, rise_thresh=RISING_THRESHOLD, fall_thresh=FALLING_THRESHOLD):
        """Compute 3-month % change and return (current_value, change_pct, direction)."""
        clean = series.dropna()
        if len(clean) < LOOKBACK:
            if len(clean) >= 5:
                curr = float(clean.iloc[-1])
                old = float(clean.iloc[0])
                chg = ((curr - old) / abs(old) * 100) if old != 0 else 0
                d = "rising" if chg > rise_thresh else ("falling" if chg < fall_thresh else "flat")
                return _safe(curr), _safe(chg), d
            return None, None, "flat"
        curr = float(clean.iloc[-1])
        old = float(clean.iloc[-LOOKBACK])
        chg = ((curr - old) / abs(old) * 100) if old != 0 else 0
        d = "rising" if chg > rise_thresh else ("falling" if chg < fall_thresh else "flat")
        return _safe(curr), _safe(chg), d

    def direction_abs(series, rise_thresh=RATE_RISE_THRESH, fall_thresh=RATE_FALL_THRESH):
        """For yield-type data: use absolute change instead of percentage."""
        clean = series.dropna()
        if len(clean) < LOOKBACK:
            if len(clean) >= 5:
                curr = float(clean.iloc[-1])
                old = float(clean.iloc[0])
                chg = curr - old
                d = "rising" if chg > rise_thresh else ("falling" if chg < fall_thresh else "flat")
                return _safe(curr), _safe(chg), d
            return None, None, "flat"
        curr = float(clean.iloc[-1])
        old = float(clean.iloc[-LOOKBACK])
        chg = curr - old
        d = "rising" if chg > rise_thresh else ("falling" if chg < fall_thresh else "flat")
        return _safe(curr), _safe(chg), d

    # Build indicator results
    indicators = {}

    # Inflation proxy (TIP/IEF spread)
    val, chg, d = direction_pct(close.get("INFLATION_SPREAD", pd.Series(dtype=float)))
    indicators["inflation_proxy"] = {"value": val, "direction": d, "change_3m": chg, "label": "Inflation Expectations"}

    # Oil
    sym = MACRO_TICKERS["oil"]
    val, chg, d = direction_pct(close.get(sym, pd.Series(dtype=float)), rise_thresh=5.0, fall_thresh=-5.0)
    indicators["oil"] = {"value": val, "direction": d, "change_3m": chg, "label": "Oil (WTI)"}

    # 10-Year Rate
    sym = MACRO_TICKERS["rates_10y"]
    val, chg, d = direction_abs(close.get(sym, pd.Series(dtype=float)))
    indicators["rates_10y"] = {"value": val, "direction": d, "change_3m": chg, "label": "10-Year Yield"}

    # Short-term Rate
    sym = MACRO_TICKERS["rates_short"]
    val, chg, d = direction_abs(close.get(sym, pd.Series(dtype=float)))
    indicators["rates_short"] = {"value": val, "direction": d, "change_3m": chg, "label": "Short-Term Rate"}

    # USD
    sym = MACRO_TICKERS["usd"]
    val, chg, d = direction_pct(close.get(sym, pd.Series(dtype=float)))
    indicators["usd"] = {"value": val, "direction": d, "change_3m": chg, "label": "US Dollar Index"}

    # Gold
    sym = MACRO_TICKERS["gold"]
    val, chg, d = direction_pct(close.get(sym, pd.Series(dtype=float)))
    indicators["gold"] = {"value": val, "direction": d, "change_3m": chg, "label": "Gold"}

    # VIX
    sym = MACRO_TICKERS["vix"]
    val, chg, d = direction_abs(close.get(sym, pd.Series(dtype=float)), rise_thresh=5, fall_thresh=-5)
    indicators["vix"] = {"value": val, "direction": d, "change_3m": chg, "label": "VIX"}

    # SPY
    sym = MACRO_TICKERS["spy"]
    val, chg, d = direction_pct(close.get(sym, pd.Series(dtype=float)))
    indicators["spy"] = {"value": val, "direction": d, "change_3m": chg, "label": "S&P 500"}

    # Determine composite regime
    infl_dir = indicators["inflation_proxy"]["direction"]
    rate_dir = indicators["rates_10y"]["direction"]
    oil_dir = indicators["oil"]["direction"]
    vix_val = indicators["vix"]["value"]
    vix_dir = indicators["vix"]["direction"]

    regime_parts = []
    if infl_dir == "rising":
        regime_parts.append("Rising Inflation")
    elif infl_dir == "falling":
        regime_parts.append("Falling Inflation")
    else:
        regime_parts.append("Stable Inflation")

    if rate_dir == "rising":
        regime_parts.append("Rising Rates")
    elif rate_dir == "falling":
        regime_parts.append("Falling Rates")
    else:
        regime_parts.append("Stable Rates")

    current_regime = " + ".join(regime_parts)

    # Overlay flags
    overlays = []
    if oil_dir == "rising":
        overlays.append("Oil Rising")
    elif oil_dir == "falling":
        overlays.append("Oil Falling")
    if vix_val is not None and vix_val > 25:
        overlays.append("High Volatility")
    elif vix_dir == "rising":
        overlays.append("Rising Volatility")

    # Regime description
    descriptions = {
        "Rising Inflation + Rising Rates": "Inflation expectations are climbing and rates are moving higher. Floating-rate and commodity-linked assets tend to outperform. Fixed-rate bonds and growth stocks face headwinds.",
        "Rising Inflation + Falling Rates": "A stagflationary environment — inflation is rising while rates are being cut. Commodities and real assets tend to shine. Growth may benefit from lower rates but inflation erodes returns.",
        "Rising Inflation + Stable Rates": "Inflation is heating up but rates haven't moved yet. Inflation beneficiaries have an edge while rate-sensitive assets are in a holding pattern.",
        "Falling Inflation + Rising Rates": "A tightening environment — inflation is easing but rates are still climbing. Defensive and high-quality income assets tend to hold up best.",
        "Falling Inflation + Falling Rates": "An easing environment — both inflation and rates are declining. Growth equities and long-duration bonds typically rally. Risk-on positioning tends to be rewarded.",
        "Falling Inflation + Stable Rates": "Inflation is cooling with rates holding steady. A benign environment for most asset classes. Quality growth and income tend to perform well.",
        "Stable Inflation + Rising Rates": "Inflation is contained but rates are climbing. Floating-rate assets benefit while long-duration holdings face pressure.",
        "Stable Inflation + Falling Rates": "Goldilocks territory — stable inflation with easing rates. Broad equity and bond markets tend to do well.",
        "Stable Inflation + Stable Rates": "A calm macro environment. Focus on individual security selection and yield optimization rather than macro positioning.",
    }
    regime_description = descriptions.get(current_regime, "Mixed macro signals — diversification across asset classes remains important.")

    # Build history arrays for sparklines
    dates = [d.strftime("%Y-%m-%d") for d in close.index]
    def _series_list(col):
        if col not in close.columns:
            return []
        return [_safe(float(v)) if not math.isnan(v) else None for v in close[col].values]

    indicator_history = {
        "dates": dates,
        "inflation_proxy": _series_list("INFLATION_SPREAD"),
        "oil": _series_list(MACRO_TICKERS["oil"]),
        "rates_10y": _series_list(MACRO_TICKERS["rates_10y"]),
        "rates_short": _series_list(MACRO_TICKERS["rates_short"]),
        "usd": _series_list(MACRO_TICKERS["usd"]),
        "gold": _series_list(MACRO_TICKERS["gold"]),
        "vix": _series_list(MACRO_TICKERS["vix"]),
        "spy": _series_list(MACRO_TICKERS["spy"]),
    }

    # Determine active regime components for scoring
    active_components = []
    if infl_dir == "rising":
        active_components.append("inflation_rising")
    elif infl_dir == "falling":
        active_components.append("inflation_falling")
    if rate_dir == "rising":
        active_components.append("rates_rising")
    elif rate_dir == "falling":
        active_components.append("rates_falling")
    if oil_dir == "rising":
        active_components.append("oil_rising")
    elif oil_dir == "falling":
        active_components.append("oil_falling")

    result = {
        "current_regime": current_regime,
        "regime_description": regime_description,
        "overlays": overlays,
        "active_components": active_components,
        "indicators": indicators,
        "indicator_history": indicator_history,
    }

    # Cache it
    _macro_cache["data"] = result
    _macro_cache["timestamp"] = time.time()

    return jsonify(result)


PILLAR_DISPLAY_NAMES = {
    "HA": "Hedged Anchor", "A": "Anchor", "GS": "Gold/Silver",
    "B": "Booster", "J": "Juicer", "BDC": "BDC", "G": "Growth",
}

SENSITIVITY_DISPLAY_NAMES = {
    "inflation_benefiting":     "Inflation Benefiting",
    "inflation_negative":       "Inflation Negative",
    "inflation_neutral":        "Inflation Neutral",
    "rate_sensitive_positive":  "Rate Sensitive (Positive)",
    "rate_sensitive_negative":  "Rate Sensitive (Negative)",
    "rate_sensitive_mild":      "Rate Sensitive (Mild)",
    "commodity_linked":         "Commodity Linked",
    "safe_haven":               "Safe Haven",
    "growth_equity":            "Growth Equity",
    "excluded":                 "Excluded",
    "unclassified":             "Unclassified",
}


@app.route("/api/macro/exposure", methods=["POST"])
def macro_exposure():
    """Analyze portfolio exposure relative to current macro conditions."""
    import json
    import math
    import warnings
    warnings.filterwarnings("ignore")

    def _safe(v):
        if v is None:
            return None
        try:
            f = float(v)
            return None if (math.isnan(f) or math.isinf(f)) else round(f, 4)
        except (TypeError, ValueError):
            return None

    # Get current macro conditions (use cached if available)
    now = time.time()
    if _macro_cache["data"] and (now - _macro_cache["timestamp"]) < _macro_cache["ttl"]:
        macro = _macro_cache["data"]
    else:
        # Trigger a fresh fetch
        with app.test_request_context(method="POST", json={}):
            resp = macro_conditions()
            if hasattr(resp, "get_json"):
                macro = resp.get_json()
            else:
                macro = resp

    current_regime = macro.get("current_regime", "Unknown")
    active_components = macro.get("active_components", [])

    # Get portfolio holdings (aggregate by ticker across owners)
    _, pids = get_profile_filter()
    placeholders = ",".join("?" * len(pids))
    conn = get_connection()
    rows = conn.execute(
        f"""SELECT ticker, description, classification_type, quantity,
               current_value, approx_monthly_income, current_annual_yield
           FROM all_account_info
           WHERE profile_id IN ({placeholders}) AND quantity > 0 AND current_value > 0""",
        pids,
    ).fetchall()

    # Load user overrides for this profile
    override_rows = conn.execute(
        f"SELECT ticker, sensitivity_tags FROM macro_overrides WHERE profile_id IN ({placeholders})",
        pids,
    ).fetchall()
    conn.close()
    overrides = {}
    for ovr in override_rows:
        overrides[ovr["ticker"]] = json.loads(ovr["sensitivity_tags"])

    if not rows:
        return jsonify(error="No holdings found for this profile.")

    # Aggregate by ticker
    holdings = {}
    for r in rows:
        t = r["ticker"]
        if t in holdings:
            holdings[t]["quantity"] = (holdings[t]["quantity"] or 0) + (r["quantity"] or 0)
            holdings[t]["current_value"] = (holdings[t]["current_value"] or 0) + (r["current_value"] or 0)
            holdings[t]["approx_monthly_income"] = (holdings[t]["approx_monthly_income"] or 0) + (r["approx_monthly_income"] or 0)
        else:
            holdings[t] = dict(r)

    total_value = sum(float(h.get("current_value") or 0) for h in holdings.values())
    if total_value <= 0:
        return jsonify(error="Portfolio has no value.")

    # Classify each holding and compute macro scores
    holdings_detail = []
    by_sensitivity = {}
    unclassified_value = 0.0

    excluded_value = 0.0
    for ticker, h in holdings.items():
        cv = float(h.get("current_value") or 0)
        mi = float(h.get("approx_monthly_income") or 0)
        ct = (h.get("classification_type") or "").strip()
        tags, source = _get_ticker_sensitivity(ticker, ct, h.get("description") or "", overrides=overrides)

        is_excluded = "excluded" in tags

        if is_excluded:
            excluded_value += cv
            score = None
            macro_label = "Excluded"
        else:
            # Compute macro score for this holding
            score = 0.0
            n_components = 0
            for component in active_components:
                weights = MACRO_SCORE_WEIGHTS.get(component, {})
                for tag in tags:
                    score += weights.get(tag, 0.0)
                n_components += 1
            if n_components > 0:
                score /= n_components

            # Clamp to [-1, 1]
            score = max(-1.0, min(1.0, score))

            if score > 0.3:
                macro_label = "Favorable"
            elif score > -0.3:
                macro_label = "Neutral"
            else:
                macro_label = "Unfavorable"

        if "unclassified" in tags:
            unclassified_value += cv

        # Track by sensitivity (skip excluded from breakdown)
        if not is_excluded:
            for tag in tags:
                if tag not in by_sensitivity:
                    by_sensitivity[tag] = {"value": 0.0, "tickers": [], "label": SENSITIVITY_DISPLAY_NAMES.get(tag, tag)}
                by_sensitivity[tag]["value"] += cv
                by_sensitivity[tag]["tickers"].append(ticker)

        holdings_detail.append({
            "ticker": ticker,
            "description": h.get("description") or "",
            "classification_type": ct,
            "pillar_name": PILLAR_DISPLAY_NAMES.get(ct, ct or "—"),
            "current_value": _safe(cv),
            "monthly_income": _safe(mi),
            "pct_of_portfolio": _safe(cv / total_value * 100),
            "sensitivity_tags": tags,
            "sensitivity_source": source,
            "macro_score": _safe(score),
            "macro_label": macro_label,
        })

    # Add percentages to by_sensitivity
    for tag, data in by_sensitivity.items():
        data["pct"] = _safe(data["value"] / total_value * 100)
        data["value"] = _safe(data["value"])

    # Compute portfolio alignment score (value-weighted, excluding excluded holdings)
    alignment_score = 0.0
    included_value = total_value - excluded_value
    for h in holdings_detail:
        if h["macro_label"] == "Excluded":
            continue
        cv = float(h.get("current_value") or 0)
        s = float(h.get("macro_score") or 0)
        alignment_score += cv * s
    alignment_score = alignment_score / included_value if included_value > 0 else 0
    alignment_score = max(-1.0, min(1.0, alignment_score))

    if alignment_score > 0.3:
        alignment_label = "Well Positioned"
    elif alignment_score > 0.1:
        alignment_label = "Slightly Favorable"
    elif alignment_score > -0.1:
        alignment_label = "Neutral"
    elif alignment_score > -0.3:
        alignment_label = "Slightly Unfavorable"
    else:
        alignment_label = "Poorly Positioned"

    # Sort holdings: unfavorable first, then neutral, then favorable, excluded last
    order = {"Unfavorable": 0, "Neutral": 1, "Favorable": 2, "Excluded": 3}
    holdings_detail.sort(key=lambda h: (order.get(h["macro_label"], 1), -(h.get("current_value") or 0)))

    favorable = [h for h in holdings_detail if h["macro_label"] == "Favorable"]
    unfavorable = [h for h in holdings_detail if h["macro_label"] == "Unfavorable"]

    unclassified_pct = _safe(unclassified_value / total_value * 100) if total_value > 0 else 0

    return jsonify(
        current_regime=current_regime,
        active_components=active_components,
        portfolio_alignment_score=_safe(alignment_score),
        alignment_label=alignment_label,
        total_value=_safe(total_value),
        by_sensitivity=by_sensitivity,
        holdings_detail=holdings_detail,
        favorable_holdings=favorable,
        unfavorable_holdings=unfavorable,
        unclassified_pct=unclassified_pct,
        unclassified_warning=f"{unclassified_pct}% of your portfolio couldn't be classified. Assign pillar types in Manage Holdings for more accurate results." if (unclassified_pct or 0) > 20 else None,
    )


@app.route("/api/macro/rebalance-suggestions", methods=["POST"])
def macro_rebalance_suggestions():
    """Generate rebalancing tilt suggestions based on macro conditions and portfolio exposure."""
    import math
    import warnings
    warnings.filterwarnings("ignore")

    def _safe(v):
        if v is None:
            return None
        try:
            f = float(v)
            return None if (math.isnan(f) or math.isinf(f)) else round(f, 4)
        except (TypeError, ValueError):
            return None

    # Get exposure data by calling the exposure endpoint directly
    # (we're already in a request context with the same query string)
    resp = macro_exposure()
    if hasattr(resp, "get_json"):
        exposure = resp.get_json()
    else:
        exposure = resp

    if "error" in exposure:
        return jsonify(exposure)

    current_regime = exposure.get("current_regime", "Unknown")
    active_components = exposure.get("active_components", [])
    by_sensitivity = exposure.get("by_sensitivity", {})
    holdings_detail = exposure.get("holdings_detail", [])
    alignment_score = float(exposure.get("portfolio_alignment_score") or 0)
    total_value = float(exposure.get("total_value") or 0)

    # Collect ALL portfolio tickers so candidate ETFs exclude anything already owned
    all_portfolio_tickers = set()
    for data in by_sensitivity.values():
        for t in data.get("tickers", []):
            all_portfolio_tickers.add(t.upper())

    # Build a human-readable conditions string that includes active overlays
    component_labels = {
        "inflation_rising": "rising inflation", "inflation_falling": "falling inflation",
        "rates_rising": "rising rates", "rates_falling": "falling rates",
        "oil_rising": "rising oil prices", "oil_falling": "falling oil prices",
    }
    active_labels = [component_labels.get(c, c) for c in active_components]
    if active_labels:
        conditions_str = ", ".join(active_labels)
    else:
        conditions_str = "stable macro conditions"

    # Determine which sensitivity tags are favorable / unfavorable in current regime
    tag_scores = {}
    for component in active_components:
        weights = MACRO_SCORE_WEIGHTS.get(component, {})
        for tag, weight in weights.items():
            if tag not in tag_scores:
                tag_scores[tag] = 0.0
            tag_scores[tag] += weight
    # Average across active components
    n = len(active_components) or 1
    for tag in tag_scores:
        tag_scores[tag] /= n

    favorable_tags = sorted([(t, s) for t, s in tag_scores.items() if s > 0.2], key=lambda x: -x[1])
    unfavorable_tags = sorted([(t, s) for t, s in tag_scores.items() if s < -0.2], key=lambda x: x[1])

    suggestions = []

    # Generate "increase" suggestions for favorable tags that are underweight
    for tag, score in favorable_tags:
        if tag == "unclassified":
            continue
        current_data = by_sensitivity.get(tag, {})
        current_pct = float(current_data.get("pct") or 0)
        current_tickers = current_data.get("tickers", [])

        # Suggest increasing if below a reasonable threshold
        suggested_increase = min(current_pct + 10, 40)  # don't suggest more than 40%
        if suggested_increase > current_pct + 3:  # only suggest if meaningful increase
            # Build candidate ETFs list (exclude ones already owned in ANY category)
            candidates = [
                c for c in CANDIDATE_ETFS.get(tag, [])
                if c["ticker"].upper() not in all_portfolio_tickers
            ]
            suggestions.append({
                "action": "increase",
                "target_sensitivity": tag,
                "target_label": SENSITIVITY_DISPLAY_NAMES.get(tag, tag),
                "reason": f"Portfolio is underweight {SENSITIVITY_DISPLAY_NAMES.get(tag, tag).lower()} assets, which tend to outperform during {conditions_str}.",
                "current_pct": _safe(current_pct),
                "suggested_pct": _safe(suggested_increase),
                "tickers_in_portfolio": current_tickers,
                "candidate_etfs": candidates,
                "macro_favorability": _safe(score),
            })

    # Generate "reduce" suggestions for unfavorable tags that are overweight
    for tag, score in unfavorable_tags:
        if tag == "unclassified":
            continue
        current_data = by_sensitivity.get(tag, {})
        current_pct = float(current_data.get("pct") or 0)
        current_tickers = current_data.get("tickers", [])

        if current_pct > 15:  # only flag if meaningful allocation
            suggested_pct = max(current_pct - 10, 5)  # don't suggest going below 5%
            # Find the biggest unfavorable holdings for consolidation linking
            unfavorable_in_tag = [
                h for h in holdings_detail
                if tag in h.get("sensitivity_tags", []) and h.get("macro_label") == "Unfavorable"
            ]
            unfavorable_in_tag.sort(key=lambda h: -(h.get("current_value") or 0))
            reduce_tickers = [h["ticker"] for h in unfavorable_in_tag[:10]]

            suggestions.append({
                "action": "reduce",
                "target_sensitivity": tag,
                "target_label": SENSITIVITY_DISPLAY_NAMES.get(tag, tag),
                "reason": f"Heavy {SENSITIVITY_DISPLAY_NAMES.get(tag, tag).lower()} exposure ({current_pct:.0f}%) faces headwinds during {conditions_str}.",
                "current_pct": _safe(current_pct),
                "suggested_pct": _safe(suggested_pct),
                "tickers_to_consider_reducing": reduce_tickers,
                "all_tickers": current_tickers,
                "consolidation_link": len(reduce_tickers) > 0,
                "macro_favorability": _safe(score),
            })

    # Build "next dollar" allocation
    next_dollar = {}
    if favorable_tags:
        total_fav_score = sum(s for _, s in favorable_tags if s > 0)
        for tag, score in favorable_tags:
            if tag == "unclassified" or score <= 0:
                continue
            pct = round(score / total_fav_score * 100) if total_fav_score > 0 else 0
            next_dollar[SENSITIVITY_DISPLAY_NAMES.get(tag, tag)] = pct
        # Normalize to 100
        total_alloc = sum(next_dollar.values())
        if total_alloc > 0 and total_alloc != 100:
            factor = 100 / total_alloc
            for k in next_dollar:
                next_dollar[k] = round(next_dollar[k] * factor)
            # Fix rounding to exactly 100
            diff = 100 - sum(next_dollar.values())
            if diff != 0 and next_dollar:
                top_key = max(next_dollar, key=next_dollar.get)
                next_dollar[top_key] += diff

    # If no active components (everything flat), suggest balanced allocation
    if not active_components:
        suggestions.append({
            "action": "hold",
            "target_sensitivity": None,
            "target_label": "Balanced",
            "reason": "Macro conditions are stable across the board. No strong tilts are needed — continue with your current allocation strategy and focus on yield optimization.",
            "current_pct": None,
            "suggested_pct": None,
            "tickers_in_portfolio": [],
            "macro_favorability": 0,
        })

    # ── Breakeven analysis: calculate how much needs to shift ──
    # alignment_score is value-weighted avg of per-holding macro_scores in [-1, 1].
    # To reach breakeven (score ≥ 0), we need to shift $ from unfavorable → favorable.
    breakeven_target = {}
    if alignment_score < 0 and total_value > 0:
        # Current favorability breakdown by tag
        fav_value = sum(float(h.get("current_value") or 0) for h in holdings_detail if h.get("macro_label") == "Favorable")
        unfav_value = sum(float(h.get("current_value") or 0) for h in holdings_detail if h.get("macro_label") == "Unfavorable")
        neutral_value = total_value - fav_value - unfav_value

        # Avg macro_score for favorable and unfavorable holdings
        fav_avg_score = 0
        unfav_avg_score = 0
        if fav_value > 0:
            fav_avg_score = sum(float(h.get("current_value") or 0) * float(h.get("macro_score") or 0)
                               for h in holdings_detail if h.get("macro_label") == "Favorable") / fav_value
        if unfav_value > 0:
            unfav_avg_score = sum(float(h.get("current_value") or 0) * float(h.get("macro_score") or 0)
                                 for h in holdings_detail if h.get("macro_label") == "Unfavorable") / unfav_value

        # To reach alignment 0: need shift_amount from unfav→fav such that
        # (alignment_score * total_value + shift * (fav_avg - unfav_avg)) / total_value = 0
        score_swing = (fav_avg_score - unfav_avg_score) if (fav_avg_score - unfav_avg_score) > 0 else 0.5
        shift_needed = abs(alignment_score * total_value) / score_swing if score_swing > 0 else 0
        shift_needed = min(shift_needed, unfav_value * 0.5)  # cap at 50% of unfavorable

        # Distribute shift across favorable tags proportional to their score
        shift_by_tag = []
        for tag, score in favorable_tags:
            if tag == "unclassified":
                continue
            current_data = by_sensitivity.get(tag, {})
            current_pct = float(current_data.get("pct") or 0)
            current_val = float(current_data.get("value") or 0)
            tag_share = score / sum(s for _, s in favorable_tags if s > 0) if sum(s for _, s in favorable_tags if s > 0) > 0 else 0
            tag_shift = shift_needed * tag_share
            target_val = current_val + tag_shift
            target_pct = target_val / total_value * 100 if total_value > 0 else 0

            shift_by_tag.append({
                "tag": tag,
                "label": SENSITIVITY_DISPLAY_NAMES.get(tag, tag),
                "current_pct": _safe(current_pct),
                "target_pct": _safe(target_pct),
                "gap_pct": _safe(target_pct - current_pct),
                "gap_dollars": _safe(tag_shift),
                "favorability_score": _safe(score),
            })

        breakeven_target = {
            "total_shift_needed": _safe(shift_needed),
            "total_shift_pct": _safe(shift_needed / total_value * 100) if total_value > 0 else 0,
            "current_favorable_pct": _safe(fav_value / total_value * 100),
            "current_unfavorable_pct": _safe(unfav_value / total_value * 100),
            "current_neutral_pct": _safe(neutral_value / total_value * 100),
            "tags": shift_by_tag,
        }

    return jsonify(
        current_regime=current_regime,
        conditions=conditions_str,
        alignment_score=_safe(alignment_score),
        suggestions=suggestions,
        next_dollar_allocation=next_dollar,
        breakeven_target=breakeven_target,
    )


# ── Income Benchmark ──────────────────────────────────────────────────────────

INCOME_BENCHMARK_DEFAULTS = {
    "Covered Call / Options Income": 15,
    "BDCs": 8,
    "CEFs": 10,
    "REITs / Real Estate": 15,
    "Preferred Stock / Credit": 12,
    "Dividend Growth": 20,
    "Commodities / Gold & Silver": 5,
    "Bonds / Fixed Income": 15,
}

# Kept for backward compat — points to defaults
INCOME_BENCHMARK_TARGETS = INCOME_BENCHMARK_DEFAULTS


def _load_income_targets(pids):
    """Load custom income benchmark targets for profile, falling back to defaults."""
    conn = get_connection()
    placeholders = ",".join("?" * len(pids))
    rows = conn.execute(
        f"SELECT bucket, target_pct FROM income_benchmark_targets WHERE profile_id IN ({placeholders})",
        pids,
    ).fetchall()
    conn.close()
    if rows:
        return {r["bucket"]: r["target_pct"] for r in rows}
    return dict(INCOME_BENCHMARK_DEFAULTS)

INCOME_BUCKET_BY_PILLAR = {
    "BDC": "BDCs",
    "GS": "Commodities / Gold & Silver",
    "CEF": "CEFs",
    "REIT": "REITs / Real Estate",
    "HA": "Covered Call / Options Income",
    "J": "Covered Call / Options Income",
    "B": "Covered Call / Options Income",
    "A": "Dividend Growth",
    "G": "Dividend Growth",
}

INCOME_BUCKET_KEYWORDS = {
    "Covered Call / Options Income": [
        "covered call", "option income", "premium income", "buy-write",
        "buywrite", "options-based", "nasdaq premium", "s&p 500 premium",
        "high income etf", "high income fund",
        "yieldmax", "tappalpha", "defiance", "kurv", "roundhill",
        "neos", "boosted", "growth and da", "growth & daily",
        "equity premium income", "enhanced div",
        "lift etf", "leveraged",
    ],
    "BDCs": [
        "business development", "bdc", "direct lending",
    ],
    "CEFs": [
        "closed-end", "closed end", "adams diversified", "adams natural",
        "blackrock science", "reaves utility", "saba closed",
        "cohen steers", "aberdeen", "abrdn",
        " cf", " cef",
    ],
    "REITs / Real Estate": [
        "real estate", "reit", "mortgage", "rlty", "iyri",
    ],
    "Preferred Stock / Credit": [
        "preferred", "pfd", "pffa", "pff ",
        "senior secured", "high yield bond",
        "floating rate", "bank loan", "leveraged loan",
    ],
    "Dividend Growth": [
        "dividend growth", "dividend appreciation", "aristocrat",
        "dividend achiever", "quality dividend",
    ],
    "Commodities / Gold & Silver": [
        "gold", "silver", "commodity", "mining", "precious metal",
        "energy", "oil", "natural gas", "copper", "mlp",
    ],
    "Bonds / Fixed Income": [
        "treasury", "bond", "fixed income", "aggregate bond",
        "investment grade", "tips", "municipal", "clo",
        "tbll", "t-bill", "credit",
    ],
}


def _get_income_bucket(ticker, classification_type, description="", overrides=None):
    """Classify a holding into one of the income benchmark buckets.
    Tier 0: User override (income_overrides table)
    Tier 1: Specific pillar codes that map 1:1 to income buckets (GS, BDC)
    Tier 2: Name/description heuristics
    Tier 3: Pillar code mapping (for remaining codes like CEF, HA, J, B, A, G)
    Tier 4: yfinance sector fallback
    Tier 5: Unclassified
    """
    # Tier 0: User override
    if overrides and ticker in overrides:
        return overrides[ticker]

    # Tier 1: Specific pillar codes — these take priority over keyword guessing
    # so that income benchmark percentages stay consistent with the dashboard.
    SPECIFIC_PILLAR_BUCKETS = {
        "GS": "Commodities / Gold & Silver",
        "BDC": "BDCs",
    }
    ct = classification_type.strip().upper() if classification_type else ""
    if ct in SPECIFIC_PILLAR_BUCKETS:
        return SPECIFIC_PILLAR_BUCKETS[ct]

    # Tier 2: Name-based heuristics
    cache_entry = _ticker_info_cache.get(ticker)
    info = cache_entry.get("info", {}) if cache_entry else {}
    yf_name = (info.get("longName") or info.get("shortName") or "").lower()
    yf_category = (info.get("category") or "").lower()
    yf_sector = (info.get("sector") or "").lower()
    combined = f"{yf_name} {yf_category} {description.lower()} {ticker.lower()}"

    # Check specific buckets (order matters — more specific before generic)
    check_order = [
        "REITs / Real Estate",
        "Commodities / Gold & Silver",
        "Bonds / Fixed Income",
        "BDCs",
        "CEFs",
        "Preferred Stock / Credit",
        "Dividend Growth",
        "Covered Call / Options Income",  # catch-all last
    ]
    for bucket in check_order:
        # Skip Commodities for holdings the user explicitly classified into
        # a different pillar — prevents broad keywords like "energy"/"mlp"
        # from pulling non-GS holdings into Commodities.
        if bucket == "Commodities / Gold & Silver" and ct and ct != "GS":
            continue
        # Same for BDCs — trust the user's pillar classification.
        if bucket == "BDCs" and ct and ct != "BDC":
            continue
        keywords = INCOME_BUCKET_KEYWORDS.get(bucket, [])
        for kw in keywords:
            if kw in combined:
                return bucket

    # Tier 3: Pillar classification (remaining codes)
    if ct:
        skip_values = {"ETF", "STOCK", "EQUITY", "FUND", ""}
        if ct not in skip_values and ct in INCOME_BUCKET_BY_PILLAR:
            return INCOME_BUCKET_BY_PILLAR[ct]

    # Tier 4: Sector-based fallback
    if "real estate" in yf_sector:
        return "REITs / Real Estate"

    return "Unclassified"


@app.route("/api/macro/income-benchmark", methods=["POST"])
def macro_income_benchmark():
    """Compare portfolio allocation against an income-focused benchmark."""
    import math

    def _safe(v):
        if v is None:
            return None
        try:
            f = float(v)
            return None if (math.isnan(f) or math.isinf(f)) else round(f, 4)
        except (TypeError, ValueError):
            return None

    _, pids = get_profile_filter()
    placeholders = ",".join("?" * len(pids))
    conn = get_connection()
    rows = conn.execute(
        f"SELECT ticker, description, classification_type, "
        f"current_value, approx_monthly_income, current_annual_yield, quantity "
        f"FROM all_account_info "
        f"WHERE profile_id IN ({placeholders}) AND IFNULL(quantity, 0) > 0",
        pids,
    ).fetchall()

    # Load income bucket overrides
    inc_ovr_rows = conn.execute(
        f"SELECT ticker, bucket FROM income_overrides WHERE profile_id IN ({placeholders})",
        pids,
    ).fetchall()
    conn.close()
    income_overrides = {r["ticker"]: r["bucket"] for r in inc_ovr_rows}

    if not rows:
        return jsonify(error="No holdings found.")

    # Aggregate by ticker across profiles
    holdings = {}
    for r in rows:
        t = r["ticker"]
        if t not in holdings:
            holdings[t] = {
                "ticker": t,
                "description": r["description"] or "",
                "classification_type": r["classification_type"] or "",
                "current_value": 0, "monthly_income": 0, "quantity": 0,
            }
        holdings[t]["current_value"] += float(r["current_value"] or 0)
        holdings[t]["monthly_income"] += float(r["approx_monthly_income"] or 0)
        holdings[t]["quantity"] += float(r["quantity"] or 0)

    total_value = sum(h["current_value"] for h in holdings.values())
    total_monthly = sum(h["monthly_income"] for h in holdings.values())
    if total_value <= 0:
        return jsonify(error="Portfolio has no value.")

    # Load targets (custom or defaults)
    targets = _load_income_targets(pids)

    # Classify each holding into a bucket
    buckets = {name: {"value": 0, "monthly_income": 0, "quantity": 0, "tickers": []}
               for name in list(targets.keys()) + ["Unclassified"]}

    holdings_detail = []
    excluded_value = 0.0
    for h in holdings.values():
        bucket = _get_income_bucket(h["ticker"], h["classification_type"], h["description"], overrides=income_overrides)
        cv = h["current_value"]
        mi = h["monthly_income"]
        qty = h["quantity"]
        yld = (mi * 12 / cv * 100) if cv > 0 else 0
        is_overridden = h["ticker"] in income_overrides

        if bucket == "Excluded":
            excluded_value += cv
        elif bucket not in buckets:
            bucket = "Unclassified"
        if bucket != "Excluded":
            buckets[bucket]["value"] += cv
            buckets[bucket]["monthly_income"] += mi
            buckets[bucket]["quantity"] += qty
            buckets[bucket]["tickers"].append(h["ticker"])

        holdings_detail.append({
            "ticker": h["ticker"],
            "description": h["description"],
            "bucket": bucket,
            "current_value": _safe(cv),
            "pct_of_portfolio": _safe(cv / total_value * 100),
            "monthly_income": _safe(mi),
            "annual_yield": _safe(yld),
            "quantity": _safe(qty),
            "is_overridden": is_overridden,
        })

    # Build comparison table
    comparison = []
    for bucket_name, target_pct in targets.items():
        data = buckets.get(bucket_name, {"value": 0, "monthly_income": 0, "tickers": []})
        actual_pct = data["value"] / total_value * 100 if total_value > 0 else 0
        diff_pct = actual_pct - target_pct
        target_value = total_value * target_pct / 100
        gap_dollars = target_value - data["value"]
        bucket_yield = (data["monthly_income"] * 12 / data["value"] * 100) if data["value"] > 0 else 0

        comparison.append({
            "bucket": bucket_name,
            "target_pct": target_pct,
            "actual_pct": _safe(actual_pct),
            "diff_pct": _safe(diff_pct),
            "actual_value": _safe(data["value"]),
            "target_value": _safe(target_value),
            "gap_dollars": _safe(gap_dollars),
            "monthly_income": _safe(data["monthly_income"]),
            "bucket_yield": _safe(bucket_yield),
            "quantity": _safe(data["quantity"]),
            "tickers": data["tickers"],
        })

    # Unclassified
    unclass = buckets.get("Unclassified", {"value": 0, "monthly_income": 0, "tickers": []})
    unclassified_pct = unclass["value"] / total_value * 100 if total_value > 0 else 0

    # Portfolio-level metrics
    blended_yield = (total_monthly * 12 / total_value * 100) if total_value > 0 else 0
    # Diversification score: 1 - HHI (Herfindahl index). Higher = more diversified.
    bucket_shares = [(b["actual_pct"] or 0) / 100 for b in comparison if (b["actual_pct"] or 0) > 0]
    hhi = sum(s ** 2 for s in bucket_shares)
    diversification_score = round((1 - hhi) * 100)

    return jsonify(
        comparison=comparison,
        holdings_detail=holdings_detail,
        summary={
            "total_value": _safe(total_value),
            "total_monthly_income": _safe(total_monthly),
            "total_annual_income": _safe(total_monthly * 12),
            "blended_yield": _safe(blended_yield),
            "diversification_score": diversification_score,
        },
        unclassified_pct=_safe(unclassified_pct),
        unclassified_tickers=unclass["tickers"],
    )


# ── Macro Classification Overrides ─────────────────────────────────────────────

@app.route("/api/macro/overrides", methods=["GET"])
def macro_overrides_list():
    """Return all macro sensitivity overrides for the current profile."""
    import json
    _, pids = get_profile_filter()
    conn = get_connection()
    placeholders = ",".join("?" * len(pids))
    rows = conn.execute(
        f"SELECT ticker, sensitivity_tags, updated_at FROM macro_overrides WHERE profile_id IN ({placeholders})",
        pids,
    ).fetchall()
    conn.close()
    valid_tags = {k: v for k, v in SENSITIVITY_DISPLAY_NAMES.items() if k not in ("unclassified",)}
    return jsonify(
        overrides={r["ticker"]: json.loads(r["sensitivity_tags"]) for r in rows},
        sensitivity_options=valid_tags,
    )


@app.route("/api/macro/overrides", methods=["PUT"])
def macro_overrides_save():
    """Save a macro sensitivity override for a ticker."""
    import json
    data = request.get_json(force=True) or {}
    ticker = (data.get("ticker") or "").strip().upper()
    tags = data.get("sensitivity_tags") or []

    if not ticker:
        return jsonify(error="Ticker is required."), 400
    valid_tags = set(SENSITIVITY_DISPLAY_NAMES.keys()) - {"unclassified"}
    if not tags or not all(t in valid_tags for t in tags):
        return jsonify(error=f"Invalid tags. Valid: {sorted(valid_tags)}"), 400

    _, pids = get_profile_filter()
    pid = pids[0]
    conn = get_connection()
    conn.execute(
        """INSERT INTO macro_overrides (ticker, profile_id, sensitivity_tags, updated_at)
           VALUES (?, ?, ?, CURRENT_TIMESTAMP)
           ON CONFLICT(ticker, profile_id) DO UPDATE SET sensitivity_tags = excluded.sensitivity_tags,
                                                          updated_at = excluded.updated_at""",
        (ticker, pid, json.dumps(tags)),
    )
    conn.commit()
    conn.close()
    return jsonify(ok=True, ticker=ticker, sensitivity_tags=tags)


@app.route("/api/macro/overrides", methods=["DELETE"])
def macro_overrides_delete():
    """Remove a macro sensitivity override (revert to auto-classification)."""
    data = request.get_json(force=True) or {}
    ticker = (data.get("ticker") or "").strip().upper()
    if not ticker:
        return jsonify(error="Ticker is required."), 400

    _, pids = get_profile_filter()
    conn = get_connection()
    placeholders = ",".join("?" * len(pids))
    conn.execute(
        f"DELETE FROM macro_overrides WHERE ticker = ? AND profile_id IN ({placeholders})",
        [ticker] + pids,
    )
    conn.commit()
    conn.close()
    return jsonify(ok=True, ticker=ticker)


# ── Income Bucket Overrides ────────────────────────────────────────────────────

@app.route("/api/income/overrides", methods=["GET"])
def income_overrides_list():
    """Return all income bucket overrides for the current profile."""
    _, pids = get_profile_filter()
    conn = get_connection()
    placeholders = ",".join("?" * len(pids))
    rows = conn.execute(
        f"SELECT ticker, bucket FROM income_overrides WHERE profile_id IN ({placeholders})",
        pids,
    ).fetchall()
    conn.close()
    valid_buckets = list(INCOME_BENCHMARK_TARGETS.keys()) + ["Excluded"]
    return jsonify(
        overrides={r["ticker"]: r["bucket"] for r in rows},
        bucket_options=valid_buckets,
    )


@app.route("/api/income/overrides", methods=["PUT"])
def income_overrides_save():
    """Save an income bucket override for a ticker."""
    data = request.get_json(force=True) or {}
    ticker = (data.get("ticker") or "").strip().upper()
    bucket = (data.get("bucket") or "").strip()

    if not ticker:
        return jsonify(error="Ticker is required."), 400
    valid = set(INCOME_BENCHMARK_TARGETS.keys()) | {"Excluded"}
    if bucket not in valid:
        return jsonify(error=f"Invalid bucket. Valid: {sorted(valid)}"), 400

    _, pids = get_profile_filter()
    pid = pids[0]
    conn = get_connection()
    conn.execute(
        """INSERT INTO income_overrides (ticker, profile_id, bucket, updated_at)
           VALUES (?, ?, ?, CURRENT_TIMESTAMP)
           ON CONFLICT(ticker, profile_id) DO UPDATE SET bucket = excluded.bucket,
                                                          updated_at = excluded.updated_at""",
        (ticker, pid, bucket),
    )
    conn.commit()
    conn.close()
    return jsonify(ok=True, ticker=ticker, bucket=bucket)


@app.route("/api/income/overrides", methods=["DELETE"])
def income_overrides_delete():
    """Remove an income bucket override (revert to auto-classification)."""
    data = request.get_json(force=True) or {}
    ticker = (data.get("ticker") or "").strip().upper()
    if not ticker:
        return jsonify(error="Ticker is required."), 400

    _, pids = get_profile_filter()
    conn = get_connection()
    placeholders = ",".join("?" * len(pids))
    conn.execute(
        f"DELETE FROM income_overrides WHERE ticker = ? AND profile_id IN ({placeholders})",
        [ticker] + pids,
    )
    conn.commit()
    conn.close()
    return jsonify(ok=True, ticker=ticker)


# ── Income Benchmark Targets ──────────────────────────────────────────────────

@app.route("/api/income/targets", methods=["GET"])
def income_targets_list():
    """Return current income benchmark targets (custom or defaults)."""
    _, pids = get_profile_filter()
    targets = _load_income_targets(pids)
    # Check if custom targets exist
    conn = get_connection()
    placeholders = ",".join("?" * len(pids))
    count = conn.execute(
        f"SELECT COUNT(*) as cnt FROM income_benchmark_targets WHERE profile_id IN ({placeholders})",
        pids,
    ).fetchone()["cnt"]
    conn.close()
    return jsonify(
        targets=targets,
        is_custom=count > 0,
        defaults=INCOME_BENCHMARK_DEFAULTS,
    )


@app.route("/api/income/targets", methods=["PUT"])
def income_targets_save():
    """Save custom income benchmark targets."""
    data = request.get_json(force=True) or {}
    targets = data.get("targets")
    if not targets or not isinstance(targets, dict):
        return jsonify(error="targets dict is required."), 400

    # Validate all values are numbers and sum to ~100
    total = 0
    for bucket, pct in targets.items():
        try:
            pct = float(pct)
        except (TypeError, ValueError):
            return jsonify(error=f"Invalid percentage for {bucket}"), 400
        if pct < 0:
            return jsonify(error=f"Percentage for {bucket} cannot be negative"), 400
        total += pct
    if abs(total - 100) > 0.5:
        return jsonify(error=f"Targets must sum to 100% (currently {total:.1f}%)"), 400

    _, pids = get_profile_filter()
    pid = pids[0]
    conn = get_connection()
    # Clear existing and insert fresh
    conn.execute("DELETE FROM income_benchmark_targets WHERE profile_id = ?", (pid,))
    for bucket, pct in targets.items():
        conn.execute(
            """INSERT INTO income_benchmark_targets (bucket, profile_id, target_pct, updated_at)
               VALUES (?, ?, ?, CURRENT_TIMESTAMP)""",
            (bucket, pid, float(pct)),
        )
    conn.commit()
    conn.close()
    return jsonify(ok=True, targets=targets)


@app.route("/api/income/targets", methods=["DELETE"])
def income_targets_reset():
    """Reset income benchmark targets to defaults."""
    _, pids = get_profile_filter()
    conn = get_connection()
    placeholders = ",".join("?" * len(pids))
    conn.execute(
        f"DELETE FROM income_benchmark_targets WHERE profile_id IN ({placeholders})",
        pids,
    )
    conn.commit()
    conn.close()
    return jsonify(ok=True, targets=INCOME_BENCHMARK_DEFAULTS)


# ── Dividend History ───────────────────────────────────────────────────────────

@app.route("/api/dividend-history/data", methods=["GET"])
def dividend_history_data():
    """Return dividend income time-series in yearly, monthly, or weekly granularity."""
    import datetime
    from collections import defaultdict

    view = request.args.get("view", "monthly")  # yearly | monthly | weekly
    months_back = int(request.args.get("months_back", 60 if view == "monthly" else 12))
    cat_param = request.args.get("category", "").strip()
    cat_ids = [c.strip() for c in cat_param.split(",") if c.strip()] if cat_param else []

    is_agg, pids = get_profile_filter()
    profile_id = pids[0]
    conn = get_connection()

    # Categories for filter dropdown
    cats = conn.execute(
        "SELECT id, name FROM categories WHERE profile_id = ? ORDER BY sort_order, name",
        (profile_id,),
    ).fetchall()
    categories = [{"id": c["id"], "name": c["name"]} for c in cats]

    # Resolve category filter to ticker set
    filtered_tickers = None
    if cat_ids:
        cat_names = [c["name"] for c in categories if str(c["id"]) in cat_ids]
        if cat_names:
            placeholders = ",".join("?" * len(pids))
            cat_rows = conn.execute(
                f"SELECT DISTINCT tc.ticker FROM ticker_categories tc "
                f"JOIN categories c ON c.id = tc.category_id "
                f"WHERE tc.profile_id IN ({placeholders}) AND c.name IN ({','.join('?' * len(cat_names))})",
                pids + cat_names,
            ).fetchall()
            filtered_tickers = {r["ticker"] for r in cat_rows}
            # Also include tickers mapped via classification_type
            _CLASSIFICATION_NAMES_REV = {}
            for ct_val, ct_name in _CLASSIFICATION_NAMES.items():
                _CLASSIFICATION_NAMES_REV.setdefault(ct_name, []).append(ct_val)
            ct_vals = []
            for cn in cat_names:
                ct_vals.extend(_CLASSIFICATION_NAMES_REV.get(cn, []))
            if ct_vals:
                ct_ph = ",".join("?" * len(ct_vals))
                ct_rows = conn.execute(
                    f"SELECT DISTINCT ticker FROM all_account_info "
                    f"WHERE classification_type IN ({ct_ph}) AND profile_id IN ({placeholders})",
                    ct_vals + pids,
                ).fetchall()
                filtered_tickers |= {r["ticker"] for r in ct_rows}

    # Compute category ratio for proportional filtering on aggregate tables
    cat_ratio = 1.0
    if filtered_tickers is not None:
        total_annual = conn.execute(
            f"SELECT IFNULL(SUM(estim_payment_per_year), 0) as total FROM all_account_info "
            f"WHERE profile_id IN ({','.join('?' * len(pids))}) AND IFNULL(quantity, 0) > 0",
            pids,
        ).fetchone()["total"]
        if total_annual > 0:
            ticker_ph = ",".join("?" * len(filtered_tickers))
            cat_annual = conn.execute(
                f"SELECT IFNULL(SUM(estim_payment_per_year), 0) as total FROM all_account_info "
                f"WHERE profile_id IN ({','.join('?' * len(pids))}) AND IFNULL(quantity, 0) > 0 "
                f"AND ticker IN ({ticker_ph})",
                pids + list(filtered_tickers),
            ).fetchone()["total"]
            cat_ratio = cat_annual / total_annual
        else:
            cat_ratio = 0.0

    today = datetime.date.today()
    labels = []
    values = []
    placeholders = ",".join("?" * len(pids))

    if view == "yearly":
        rows = conn.execute(
            f"SELECT year, SUM(amount) as total FROM monthly_payouts "
            f"WHERE profile_id IN ({placeholders}) GROUP BY year ORDER BY year",
            pids,
        ).fetchall()
        for r in rows:
            labels.append(str(r["year"]))
            values.append(round(float(r["total"]) * cat_ratio, 2))

    elif view == "monthly":
        # Compute start date
        start_y = today.year
        start_m = today.month - months_back
        while start_m <= 0:
            start_m += 12
            start_y -= 1
        start_key = start_y * 100 + start_m
        end_key = today.year * 100 + today.month

        rows = conn.execute(
            f"SELECT year, month, amount FROM monthly_payouts "
            f"WHERE profile_id IN ({placeholders}) AND (year * 100 + month) >= ? AND (year * 100 + month) <= ? "
            f"ORDER BY year, month",
            pids + [start_key, end_key],
        ).fetchall()
        for r in rows:
            dt = datetime.date(int(r["year"]), int(r["month"]), 1)
            labels.append(dt.strftime("%b '%y"))
            values.append(round(float(r["amount"]) * cat_ratio, 2))

    elif view == "weekly":
        start_date = today - datetime.timedelta(days=months_back * 30)
        rows = conn.execute(
            f"SELECT pay_date, amount FROM weekly_payouts "
            f"WHERE profile_id IN ({placeholders}) AND pay_date >= ? "
            f"ORDER BY pay_date",
            pids + [start_date.isoformat()],
        ).fetchall()
        for r in rows:
            labels.append(r["pay_date"])
            values.append(round(float(r["amount"]) * cat_ratio, 2))

    conn.close()

    # Summary stats
    total = sum(values) if values else 0
    avg = total / len(values) if values else 0
    mn = min(values) if values else 0
    mx = max(values) if values else 0
    # Growth: compare last value to first
    trend_pct = 0.0
    if len(values) >= 2 and values[0] > 0:
        trend_pct = round((values[-1] - values[0]) / values[0] * 100, 1)

    # Cumulative
    cumulative = []
    running = 0
    for v in values:
        running += v
        cumulative.append(round(running, 2))

    return jsonify({
        "categories": categories,
        "view": view,
        "series": {
            "labels": labels,
            "values": values,
            "cumulative": cumulative,
        },
        "summary": {
            "total": round(total, 2),
            "average": round(avg, 2),
            "min": round(mn, 2),
            "max": round(mx, 2),
            "trend_pct": trend_pct,
        },
    })


# ── 4-Quadrant Regime & Markov Chain ──────────────────────────────────────────

# FRED series configuration for regime classification
# transform: "roc_3m" = 3-month rate-of-change then Z-score; "level" = Z-score raw value
# invert: True = negate Z-score (higher raw value = worse for that category)
FRED_SERIES_CONFIG = {
    # Growth indicators
    "Industrial Production": {"id": "INDPRO",        "category": "growth",    "transform": "roc_3m"},
    "Housing Starts":        {"id": "HOUST",         "category": "growth",    "transform": "roc_3m"},
    "Nonfarm Payrolls":      {"id": "PAYEMS",        "category": "growth",    "transform": "roc_3m"},
    "Jobless Claims":        {"id": "ICSA",          "category": "growth",    "transform": "level", "invert": True},
    "Unemployment Rate":     {"id": "UNRATE",        "category": "growth",    "transform": "level", "invert": True},
    # Inflation indicators
    "CPI":                   {"id": "CPIAUCSL",      "category": "inflation", "transform": "roc_3m"},
    "Core CPI":              {"id": "CPILFESL",      "category": "inflation", "transform": "roc_3m"},
    "Breakeven Inflation":   {"id": "T10YIE",        "category": "inflation", "transform": "level"},
    # Financial conditions
    "HY Credit Spread":     {"id": "BAMLH0A0HYM2",  "category": "financial", "transform": "level", "invert": True},
    "Yield Curve (10Y-2Y)":  {"id": "T10Y2Y",       "category": "financial", "transform": "level"},
    # Sentiment
    "Consumer Sentiment":    {"id": "UMCSENT",       "category": "sentiment", "transform": "level"},
}


def _fetch_fred_series(series_id, start="2000-01-01"):
    """Fetch a FRED series as a DataFrame via the official FRED JSON API."""
    import requests as _req
    url = "https://api.stlouisfed.org/fred/series/observations"
    params = {
        "series_id": series_id,
        "api_key": FRED_API_KEY,
        "file_type": "json",
        "observation_start": start,
        "sort_order": "asc",
    }
    r = _req.get(url, params=params, timeout=15)
    r.raise_for_status()
    data = r.json()
    obs = data.get("observations", [])
    records = [(o["date"], o["value"]) for o in obs if o["value"] != "."]
    df = pd.DataFrame(records, columns=["date", "value"])
    df["date"] = pd.to_datetime(df["date"])
    df["value"] = df["value"].astype(float)
    return df.set_index("date")


@app.route("/api/macro/quadrant", methods=["POST"])
def macro_quadrant():
    """4-quadrant regime classification with Markov transition matrix.

    Uses FRED economic data (CPI, Industrial Production, Housing Starts)
    for current-state Z-score classification, and market proxies
    (SPY, XLI, TIP, IEF) for the historical weekly transition matrix.
    """
    import numpy as np
    import yfinance as yf

    now = time.time()
    if _quadrant_cache["data"] and (now - _quadrant_cache["timestamp"]) < _quadrant_cache["ttl"]:
        return jsonify(_quadrant_cache["data"])

    try:
        # ── A. FRED economic data for current classification ──────────────
        from concurrent.futures import ThreadPoolExecutor, as_completed
        fred_ok = True
        fred_indicators = {}
        try:
            # Fetch all series in parallel for speed
            fred_dfs = {}
            with ThreadPoolExecutor(max_workers=6) as pool:
                futures = {
                    pool.submit(_fetch_fred_series, cfg["id"]): name
                    for name, cfg in FRED_SERIES_CONFIG.items()
                }
                for future in as_completed(futures):
                    name = futures[future]
                    try:
                        fred_dfs[name] = future.result()
                    except Exception:
                        pass  # Skip failed series, continue with the rest

            if len(fred_dfs) < 3:
                raise ValueError("Too few FRED series returned")

            for name, df in fred_dfs.items():
                cfg = FRED_SERIES_CONFIG[name]
                if cfg["transform"] == "roc_3m":
                    df["roc_3m"] = df["value"].pct_change(3) * 100
                    series = df["roc_3m"].dropna()
                    latest_val = float(series.iloc[-1])
                    display_val = f"{latest_val:.2f}% (3m ROC)"
                else:
                    series = df["value"].dropna()
                    latest_val = float(series.iloc[-1])
                    # Format display based on series type
                    if name == "Jobless Claims":
                        display_val = f"{latest_val:,.0f}"
                    elif name == "Nonfarm Payrolls":
                        display_val = f"{latest_val:,.0f}K"
                    else:
                        display_val = f"{latest_val:.2f}"

                mean_val = float(series.mean())
                std_val = float(series.std())
                z = (latest_val - mean_val) / std_val if std_val > 0 else 0.0

                # Previous value and Z-score for trend comparison
                prev_val = float(series.iloc[-2]) if len(series) >= 2 else latest_val
                prev_z = (prev_val - mean_val) / std_val if std_val > 0 else 0.0

                # Invert Z-score for indicators where higher = worse
                if cfg.get("invert"):
                    z = -z
                    prev_z = -prev_z

                extremity = ("Extreme" if abs(z) > 2 else
                             "Elevated" if abs(z) > 1 else "Normal")

                # Format previous display value
                if cfg["transform"] == "roc_3m":
                    prev_display = f"{prev_val:.2f}%"
                elif name == "Jobless Claims":
                    prev_display = f"{prev_val:,.0f}"
                elif name == "Nonfarm Payrolls":
                    prev_display = f"{prev_val:,.0f}K"
                else:
                    prev_display = f"{prev_val:.2f}"

                # Direction logic
                raw_change = latest_val - prev_val
                if cfg["transform"] == "roc_3m":
                    direction = "Rising" if latest_val > 0 else "Falling"
                else:
                    if cfg.get("invert"):
                        direction = "Improving" if raw_change < 0 else ("Worsening" if raw_change > 0 else "Stable")
                    else:
                        direction = "Rising" if raw_change > 0 else ("Falling" if raw_change < 0 else "Stable")

                fred_indicators[name] = {
                    "current_value": display_val,
                    "previous_value": prev_display,
                    "previous_date": series.index[-2].strftime("%Y-%m-%d") if len(series) >= 2 else None,
                    "z_score": round(z, 2),
                    "previous_z": round(prev_z, 2),
                    "z_change": round(z - prev_z, 2),
                    "extremity": extremity,
                    "direction": direction,
                    "latest_date": df.index[-1].strftime("%Y-%m-%d"),
                    "history_mean": round(mean_val, 2),
                    "history_std": round(std_val, 2),
                    "category": cfg["category"],
                }

            # Composite Z-scores — average all available indicators per category
            growth_zs = [v["z_score"] for v in fred_indicators.values() if v["category"] == "growth"]
            inflation_zs = [v["z_score"] for v in fred_indicators.values() if v["category"] == "inflation"]
            fred_growth_z = round(sum(growth_zs) / len(growth_zs), 2) if growth_zs else None
            fred_inflation_z = round(sum(inflation_zs) / len(inflation_zs), 2) if inflation_zs else None

        except Exception:
            fred_ok = False
            fred_growth_z = None
            fred_inflation_z = None

        # ── B. Market proxy data for historical transition matrix ─────────
        tickers = ["SPY", "XLI", "TIP", "IEF"]
        raw = yf.download(tickers, period="5y", auto_adjust=True, progress=False)
        close = raw["Close"].dropna(how="all")

        # Resample to weekly (Friday close) for noise reduction
        weekly = close.resample("W-FRI").last().dropna()

        # Market-proxy axes (13-week lookback = ~1 quarter)
        spy_mom = weekly["SPY"].pct_change(13) * 100
        xli_mom = weekly["XLI"].pct_change(13) * 100
        mkt_growth_score = (spy_mom + xli_mom) / 2

        inflation_spread = weekly["TIP"] / weekly["IEF"]
        mkt_inflation_score = inflation_spread.pct_change(13) * 100

        # ── C. Classify each historical week (market proxies for history) ──
        def classify(g, inf):
            if g > 0 and inf <= 0:
                return 1  # Goldilocks
            if g > 0 and inf > 0:
                return 2  # Reflation
            if g <= 0 and inf > 0:
                return 3  # Stagflation
            return 4      # Deflation

        valid = mkt_growth_score.notna() & mkt_inflation_score.notna()
        dates = mkt_growth_score[valid].index
        g_vals = mkt_growth_score[valid].values
        i_vals = mkt_inflation_score[valid].values
        regimes = [classify(g, i) for g, i in zip(g_vals, i_vals)]

        # ── D. Current quadrant: keep FRED as the macro "now" view, but anchor
        # the Markov engine to the market-proxy state it was trained on.
        market_current_quad = regimes[-1] if regimes else 2
        market_current_growth = round(float(g_vals[-1]), 4) if len(g_vals) else 0
        market_current_inflation = round(float(i_vals[-1]), 4) if len(i_vals) else 0

        if fred_ok and fred_growth_z is not None:
            current_quad = classify(fred_growth_z, fred_inflation_z)
            current_growth = fred_growth_z
            current_inflation = fred_inflation_z
            classification_source = "FRED"
        else:
            current_quad = market_current_quad
            current_growth = market_current_growth
            current_inflation = market_current_inflation
            classification_source = "Market Proxy"

        transition_anchor_quad = market_current_quad
        states_aligned = current_quad == transition_anchor_quad

        if len(regimes) < 2:
            return jsonify({"error": "Insufficient data for transition matrix"}), 502

        # Step 4: Persist to regime_history
        conn = get_connection()
        for idx in range(len(dates)):
            d_str = dates[idx].strftime("%Y-%m-%d")
            g_dir = "up" if g_vals[idx] > 0 else "down"
            i_dir = "up" if i_vals[idx] > 0 else "down"
            conn.execute(
                """INSERT OR REPLACE INTO regime_history
                   (date, quadrant, growth_score, inflation_score,
                    growth_direction, inflation_direction)
                   VALUES (?, ?, ?, ?, ?, ?)""",
                (d_str, regimes[idx],
                 round(float(g_vals[idx]), 4),
                 round(float(i_vals[idx]), 4),
                 g_dir, i_dir),
            )
        conn.commit()
        conn.close()

        # Step 5: Compute 4x4 transition matrix
        transition_counts = np.zeros((4, 4), dtype=int)
        for idx in range(len(regimes) - 1):
            fr = regimes[idx] - 1
            to = regimes[idx + 1] - 1
            transition_counts[fr][to] += 1

        smoothing_alpha = 0.35
        smoothed_counts = transition_counts.astype(float) + smoothing_alpha
        row_sums = smoothed_counts.sum(axis=1, keepdims=True)
        row_sums[row_sums == 0] = 1
        transition_matrix = (smoothed_counts / row_sums).tolist()

        # Step 5b: Compute CONDITIONAL transition matrix
        # The static matrix treats all weeks in a quadrant equally.
        # We improve it by computing transitions only from weeks whose
        # momentum conditions resemble the current state.
        #
        # For the current row, we filter historical transitions to those
        # where growth and inflation momentum were trending in the same
        # direction as now (both decelerating, both near boundary, etc.)

        g_now_mkt = float(g_vals[-1])
        i_now_mkt = float(i_vals[-1])
        g_4w_mkt = float(g_vals[-5]) if len(g_vals) >= 5 else g_now_mkt
        i_4w_mkt = float(i_vals[-5]) if len(i_vals) >= 5 else i_now_mkt
        g_mom_dir = "decelerating" if (g_now_mkt - g_4w_mkt) < -0.3 else (
            "accelerating" if (g_now_mkt - g_4w_mkt) > 0.3 else "stable")
        i_mom_dir = "decelerating" if (i_now_mkt - i_4w_mkt) < -0.3 else (
            "accelerating" if (i_now_mkt - i_4w_mkt) > 0.3 else "stable")

        # Filter: find historical weeks in the transition anchor state with
        # similar momentum.
        conditional_counts = np.zeros(4, dtype=int)
        lookback = 4  # compare 4-week delta
        for idx in range(lookback, len(regimes) - 1):
            if regimes[idx] != transition_anchor_quad:
                continue
            hist_g_delta = float(g_vals[idx]) - float(g_vals[idx - lookback])
            hist_i_delta = float(i_vals[idx]) - float(i_vals[idx - lookback])
            hist_g_dir = "decelerating" if hist_g_delta < -0.3 else (
                "accelerating" if hist_g_delta > 0.3 else "stable")
            hist_i_dir = "decelerating" if hist_i_delta < -0.3 else (
                "accelerating" if hist_i_delta > 0.3 else "stable")

            # Match: same growth direction OR same inflation direction
            # (relaxed to get enough observations)
            if hist_g_dir == g_mom_dir or hist_i_dir == i_mom_dir:
                next_q = regimes[idx + 1] - 1
                conditional_counts[next_q] += 1

        cond_total = int(conditional_counts.sum())
        static_anchor_row = np.array(transition_matrix[transition_anchor_quad - 1], dtype=float)
        conditional_weight = 0.0
        if cond_total > 0:
            conditional_empirical = conditional_counts / cond_total
            conditional_weight = min(cond_total / 24.0, 1.0)
            conditional_row = (
                conditional_empirical * conditional_weight
                + static_anchor_row * (1.0 - conditional_weight)
            ).tolist()
            use_conditional = True
        else:
            conditional_row = static_anchor_row.tolist()
            use_conditional = False

        # Further adjust conditional row using FRED Z-score mean-reversion
        # If FRED growth Z is high (>1.0), increase probability of moving to
        # quadrants where growth is DOWN (Q3, Q4) — mean reversion pressure
        adjusted_row = list(conditional_row)
        if fred_ok and fred_growth_z is not None:
            # Mean-reversion adjustment factor
            # Higher Z → stronger pull toward opposite-growth quadrants
            if fred_growth_z > 1.0 and transition_anchor_quad in (1, 2):
                # Growth is elevated in a growth-UP quad → risk of Q3/Q4
                reversion_strength = min((fred_growth_z - 1.0) * 0.15, 0.25)
                # Move probability from self → growth-DOWN quads
                if transition_anchor_quad == 2:  # Reflation
                    # Risk: Q3 (Stagflation) if inflation stays, Q4 if both fall
                    shift_to_q3 = reversion_strength * 0.75
                    shift_to_q4 = reversion_strength * 0.25
                    adjusted_row[transition_anchor_quad - 1] -= reversion_strength
                    adjusted_row[2] += shift_to_q3  # Q3
                    adjusted_row[3] += shift_to_q4  # Q4
                elif transition_anchor_quad == 1:  # Goldilocks
                    shift_to_q4 = reversion_strength * 0.75
                    shift_to_q2 = reversion_strength * 0.25
                    adjusted_row[transition_anchor_quad - 1] -= reversion_strength
                    adjusted_row[3] += shift_to_q4
                    adjusted_row[1] += shift_to_q2

            if fred_inflation_z is not None and fred_inflation_z > 1.0 and transition_anchor_quad in (1, 4):
                # Inflation rising in a low-inflation quad → risk of Q2/Q3
                reversion_strength = min((fred_inflation_z - 1.0) * 0.15, 0.25)
                if transition_anchor_quad == 1:
                    adjusted_row[transition_anchor_quad - 1] -= reversion_strength
                    adjusted_row[1] += reversion_strength  # Q2
                elif transition_anchor_quad == 4:
                    adjusted_row[transition_anchor_quad - 1] -= reversion_strength
                    adjusted_row[2] += reversion_strength  # Q3

            if fred_growth_z < -1.0 and transition_anchor_quad in (3, 4):
                # Growth depressed → risk of recovery to Q1/Q2
                reversion_strength = min((abs(fred_growth_z) - 1.0) * 0.15, 0.25)
                if transition_anchor_quad == 3:
                    adjusted_row[transition_anchor_quad - 1] -= reversion_strength
                    adjusted_row[1] += reversion_strength  # Q2
                elif transition_anchor_quad == 4:
                    adjusted_row[transition_anchor_quad - 1] -= reversion_strength
                    adjusted_row[0] += reversion_strength  # Q1

        # ── Financial conditions adjustments ──
        # HY credit spreads widening → stress building → favor downside quadrants
        hy_z = fred_indicators.get("HY Credit Spread", {}).get("z_score")
        if hy_z is not None and hy_z < -1.0:
            # Inverted Z: negative = spreads widening = stress
            stress_strength = min((abs(hy_z) - 1.0) * 0.12, 0.20)
            if transition_anchor_quad in (1, 2):
                # Growth-UP quads → risk of moving to growth-DOWN
                if fred_inflation_z is not None and fred_inflation_z > 0.5:
                    # Inflation still elevated → Stagflation risk
                    adjusted_row[transition_anchor_quad - 1] -= stress_strength
                    adjusted_row[2] += stress_strength * 0.7   # Q3 Stagflation
                    adjusted_row[3] += stress_strength * 0.3   # Q4 Deflation
                else:
                    # Inflation low → Deflation risk
                    adjusted_row[transition_anchor_quad - 1] -= stress_strength
                    adjusted_row[3] += stress_strength * 0.7   # Q4 Deflation
                    adjusted_row[2] += stress_strength * 0.3   # Q3 Stagflation

        # Yield curve inversion → recession signal → favor deflation/stagflation
        yc_z = fred_indicators.get("Yield Curve (10Y-2Y)", {}).get("z_score")
        if yc_z is not None and yc_z < -1.0:
            # Deeply inverted/flat curve
            yc_strength = min((abs(yc_z) - 1.0) * 0.10, 0.15)
            if transition_anchor_quad in (1, 2):
                adjusted_row[transition_anchor_quad - 1] -= yc_strength
                adjusted_row[3] += yc_strength * 0.6   # Q4 Deflation
                adjusted_row[2] += yc_strength * 0.4   # Q3 Stagflation

        # ── Sentiment adjustments ──
        # Consumer sentiment deeply depressed → economic weakness ahead
        sent_z = fred_indicators.get("Consumer Sentiment", {}).get("z_score")
        if sent_z is not None and sent_z < -1.0:
            # Depressed sentiment → headwind for growth-UP quadrants
            sent_strength = min((abs(sent_z) - 1.0) * 0.10, 0.20)
            if transition_anchor_quad in (1, 2):
                adjusted_row[transition_anchor_quad - 1] -= sent_strength
                if fred_inflation_z is not None and fred_inflation_z > 0.5:
                    # Weak sentiment + elevated inflation → Stagflation
                    adjusted_row[2] += sent_strength * 0.65  # Q3 Stagflation
                    adjusted_row[3] += sent_strength * 0.35  # Q4 Deflation
                else:
                    adjusted_row[3] += sent_strength * 0.65  # Q4 Deflation
                    adjusted_row[2] += sent_strength * 0.35  # Q3 Stagflation
            elif transition_anchor_quad == 3:
                # Already in Stagflation with bad sentiment → more sticky
                adjusted_row[transition_anchor_quad - 1] += sent_strength * 0.3
                adjusted_row[0] -= sent_strength * 0.15  # Less likely to escape to Q1
                adjusted_row[1] -= sent_strength * 0.15  # Less likely to escape to Q2

        # ── Combined stress signal ──
        # When multiple stress indicators align, amplify the shift
        stress_count = sum(1 for z_val in [hy_z, yc_z, sent_z]
                           if z_val is not None and z_val < -1.0)
        if stress_count >= 2 and transition_anchor_quad in (1, 2):
            # Multiple stress signals → additional shift away from growth-UP
            combo_strength = 0.05 * stress_count
            adjusted_row[transition_anchor_quad - 1] -= combo_strength
            adjusted_row[2] += combo_strength * 0.5   # Q3
            adjusted_row[3] += combo_strength * 0.5   # Q4

        # Clamp to [0, 1] and renormalize
        adjusted_row = [max(0.0, v) for v in adjusted_row]
        adj_sum = sum(adjusted_row)
        if adj_sum > 0:
            adjusted_row = [v / adj_sum for v in adjusted_row]

        # Build the adjusted transition matrix (only current row is modified)
        adjusted_matrix = [list(r) for r in transition_matrix]
        adjusted_matrix[transition_anchor_quad - 1] = adjusted_row

        # Step 6: Forward projections via matrix exponentiation
        # Use adjusted matrix for projections
        P_static = np.array(transition_matrix)
        P_adj = np.array(adjusted_matrix)
        current_vec = np.zeros(4)
        current_vec[transition_anchor_quad - 1] = 1.0

        projections = {}
        for n_weeks, label in [(1, "1_week"), (2, "2_week"),
                                (4, "4_week"), (8, "8_week"),
                                (13, "13_week")]:
            Pn = np.linalg.matrix_power(P_adj, n_weeks)
            probs = current_vec @ Pn
            projections[label] = {
                "Q1": round(float(probs[0]), 4),
                "Q2": round(float(probs[1]), 4),
                "Q3": round(float(probs[2]), 4),
                "Q4": round(float(probs[3]), 4),
            }

        # Step 7: Confidence = adjusted self-transition probability
        confidence = round(adjusted_row[transition_anchor_quad - 1] * 100, 1)

        # Step 8: Regime duration stats
        regime_counts = {1: 0, 2: 0, 3: 0, 4: 0}
        for r in regimes:
            regime_counts[r] += 1
        total_obs = len(regimes)

        # Step 9: Momentum trend analysis & narrative interpretation
        g_now = float(g_vals[-1])
        i_now = float(i_vals[-1])
        # Short-term momentum (4-week change in the market proxy scores)
        g_4w_ago = float(g_vals[-5]) if len(g_vals) >= 5 else g_now
        i_4w_ago = float(i_vals[-5]) if len(i_vals) >= 5 else i_now
        g_delta = g_now - g_4w_ago
        i_delta = i_now - i_4w_ago
        g_accel = "accelerating" if g_delta > 0.3 else "decelerating" if g_delta < -0.3 else "stable"
        i_accel = "accelerating" if i_delta > 0.3 else "decelerating" if i_delta < -0.3 else "stable"

        # Find primary risk transition (highest prob excluding self)
        # Use the adjusted (conditional) row, not the static one
        row = adjusted_row
        risk_probs = [(q + 1, row[q]) for q in range(4) if q != transition_anchor_quad - 1]
        risk_probs.sort(key=lambda x: -x[1])
        primary_risk_quad = risk_probs[0][0]
        primary_risk_pct = round(risk_probs[0][1] * 100, 1)
        primary_risk_name = QUADRANT_NAMES[primary_risk_quad]

        # Determine regime change flag — factor in FRED Z-score extremity
        self_prob = confidence
        # If FRED data shows extreme readings, downgrade the flag
        fred_extreme_count = 0
        if fred_ok:
            for fi in fred_indicators.values():
                if fi["extremity"] == "Extreme":
                    fred_extreme_count += 1

        if fred_extreme_count >= 2:
            # Multiple extreme Z-scores override market-based confidence
            regime_flag = "RED"
            regime_flag_text = (
                f"Regime Change IMMINENT. {fred_extreme_count} FRED indicators "
                f"at Extreme Z-scores — historical mean reversion is likely."
            )
        elif fred_extreme_count == 1 or self_prob < 40:
            regime_flag = "YELLOW" if self_prob >= 40 else "RED"
            extreme_name = next(
                (n for n, fi in fred_indicators.items() if fi["extremity"] == "Extreme"),
                None,
            ) if fred_ok else None
            if extreme_name and self_prob >= 40:
                regime_flag_text = (
                    f"Regime Change POSSIBLE. {extreme_name} is at Extreme "
                    f"Z-score ({fred_indicators[extreme_name]['z_score']:.2f}) "
                    f"— mean reversion pressure building."
                )
            elif self_prob < 40:
                regime_flag_text = "Regime Change IMMINENT. Current quadrant hold probability is low."
            else:
                regime_flag_text = "Regime Change POSSIBLE. Transition probabilities are elevated."
        elif self_prob < 60:
            regime_flag = "YELLOW"
            regime_flag_text = "Regime Change POSSIBLE. Transition probabilities are elevated."
        else:
            regime_flag = "GREEN"
            regime_flag_text = "Regime is STABLE. High probability of remaining in current quadrant."

        # Build narrative about likely direction
        direction_parts = []

        if classification_source == "FRED" and not states_aligned:
            direction_parts.append(
                f"Current FRED macro classification is Q{current_quad} "
                f"({QUADRANT_NAMES[current_quad]}), while the transition engine is "
                f"anchored to market proxy Q{transition_anchor_quad} "
                f"({QUADRANT_NAMES[transition_anchor_quad]}) for consistency with "
                f"the historical Markov training data."
            )

        # Lead with FRED Z-score analysis when available
        if fred_ok:
            # Report elevated/extreme Z-scores
            z_parts = []
            for name, info in fred_indicators.items():
                z = info["z_score"]
                ext = info["extremity"]
                if ext in ("Extreme", "Elevated"):
                    z_parts.append(f"{name} Z-score is {z:.2f} ({ext})")
            if z_parts:
                direction_parts.append(
                    "FRED economic data: " + "; ".join(z_parts) + "."
                )

            # Growth composite analysis
            n_growth = len([v for v in fred_indicators.values() if v["category"] == "growth"])
            if fred_growth_z is not None and fred_growth_z > 1.5:
                direction_parts.append(
                    f"Growth composite Z-score ({fred_growth_z:.2f}, avg of {n_growth} indicators) is elevated — "
                    f"historical mean reversion suggests growth may decelerate."
                )
            elif fred_growth_z is not None and fred_growth_z < -1.0:
                direction_parts.append(
                    f"Growth composite Z-score ({fred_growth_z:.2f}, avg of {n_growth} indicators) is depressed — "
                    f"potential for recovery/bounce."
                )

            # Inflation composite analysis
            n_inflation = len([v for v in fred_indicators.values() if v["category"] == "inflation"])
            if fred_inflation_z is not None and fred_inflation_z > 1.5:
                direction_parts.append(
                    f"Inflation composite Z-score ({fred_inflation_z:.2f}, avg of {n_inflation} indicators) indicates persistent "
                    f"inflation pressure well above historical norms."
                )

            # Financial conditions warning
            hy_info = fred_indicators.get("HY Credit Spread", {})
            yc_info = fred_indicators.get("Yield Curve (10Y-2Y)", {})
            if hy_info.get("extremity") in ("Extreme", "Elevated"):
                direction_parts.append(
                    f"High-yield credit spreads are {hy_info['extremity'].lower()} (Z: {hy_info['z_score']:.2f}) — "
                    f"financial stress is building."
                )
            if yc_info and yc_info.get("z_score", 0) < -1.0:
                direction_parts.append(
                    f"Yield curve (10Y-2Y) Z-score is {yc_info['z_score']:.2f} — "
                    f"flat/inverted curve signals recession risk."
                )

            # Consumer sentiment warning
            sent_info = fred_indicators.get("Consumer Sentiment", {})
            if sent_info.get("z_score", 0) < -1.5:
                direction_parts.append(
                    f"Consumer Sentiment is very depressed (Z: {sent_info['z_score']:.2f}) — "
                    f"historically precedes or accompanies economic weakness."
                )

        # Quad-specific market proxy analysis
        if transition_anchor_quad == 1:  # Goldilocks
            if g_accel == "decelerating":
                direction_parts.append(f"Market growth momentum is decelerating ({g_delta:+.2f}% over 4 weeks), which could push toward Q4 (Deflation) if it turns negative.")
            if i_accel == "accelerating":
                direction_parts.append(f"Inflation momentum is accelerating ({i_delta:+.2f}%), risking a shift to Q2 (Reflation).")
        elif transition_anchor_quad == 2:  # Reflation
            if g_accel == "decelerating":
                direction_parts.append(f"Market growth momentum is decelerating ({g_delta:+.2f}% over 4 weeks). If growth turns negative while inflation persists, the market shifts to Q3 (Stagflation).")
            if i_accel == "decelerating":
                direction_parts.append(f"Inflation momentum is cooling ({i_delta:+.2f}%), which could mean a favorable shift to Q1 (Goldilocks) if growth holds.")
            if fred_ok and fred_growth_z is not None and fred_growth_z > 0 and fred_growth_z < 0.5 and fred_inflation_z > 1.0:
                direction_parts.append(
                    f"Growth Z-score ({fred_growth_z:.2f}) is weakening while "
                    f"inflation Z-score ({fred_inflation_z:.2f}) remains elevated — "
                    f"high probability of a shift toward Q3 (Stagflation) if "
                    f"growth continues to mean-revert."
                )
        elif transition_anchor_quad == 3:  # Stagflation
            if g_accel == "accelerating":
                direction_parts.append(f"Growth is recovering ({g_delta:+.2f}% over 4 weeks), which could push toward Q2 (Reflation) if inflation stays elevated.")
            if i_accel == "decelerating":
                direction_parts.append(f"Inflation momentum is fading ({i_delta:+.2f}%), opening a path to Q4 (Deflation) if growth remains weak.")
        else:  # Q4 Deflation
            if g_accel == "accelerating":
                direction_parts.append(f"Growth is recovering ({g_delta:+.2f}% over 4 weeks), which could push toward Q1 (Goldilocks).")
            if i_accel == "accelerating":
                direction_parts.append(f"Inflation momentum is rising ({i_delta:+.2f}%), which could shift to Q3 (Stagflation) if growth stays negative.")

        # Always add the primary risk from the transition matrix
        direction_parts.append(
            f"Historical transition data suggests the primary risk is a shift to Q{primary_risk_quad} "
            f"({primary_risk_name}) with {primary_risk_pct}% weekly probability."
        )

        # 4-week projection summary
        p4 = projections.get("4_week", {})
        p4_sorted = sorted(p4.items(), key=lambda x: -x[1])
        top2 = p4_sorted[:2]
        direction_parts.append(
            f"At the 4-week horizon, the most likely states are "
            f"{top2[0][0]} {QUADRANT_NAMES[int(top2[0][0][1])]} ({top2[0][1]*100:.1f}%) and "
            f"{top2[1][0]} {QUADRANT_NAMES[int(top2[1][0][1])]} ({top2[1][1]*100:.1f}%)."
        )

        interpretation = {
            "regime_flag": regime_flag,
            "regime_flag_text": regime_flag_text,
            "growth_trend": g_accel,
            "growth_delta_4w": round(g_delta, 4),
            "inflation_trend": i_accel,
            "inflation_delta_4w": round(i_delta, 4),
            "states_aligned": states_aligned,
            "transition_anchor_quad": transition_anchor_quad,
            "transition_anchor_name": QUADRANT_NAMES[transition_anchor_quad],
            "primary_risk_quad": primary_risk_quad,
            "primary_risk_name": primary_risk_name,
            "primary_risk_pct": primary_risk_pct,
            "direction_narrative": " ".join(direction_parts),
        }

        # Step 10: Log predictions & compute Brier score
        from datetime import timedelta as _td
        _today_str = dates[-1].strftime("%Y-%m-%d")
        _pred_conn = get_connection()

        # Save predictions for each horizon
        for n_weeks, label in [(1, "1_week"), (4, "4_week"), (8, "8_week")]:
            target_d = (dates[-1] + _td(weeks=n_weeks)).strftime("%Y-%m-%d")
            p = projections[label]
            _pred_conn.execute(
                """INSERT OR REPLACE INTO regime_predictions
                   (prediction_date, horizon, target_date,
                    prob_q1, prob_q2, prob_q3, prob_q4)
                   VALUES (?, ?, ?, ?, ?, ?, ?)""",
                (_today_str, label, target_d,
                 p["Q1"], p["Q2"], p["Q3"], p["Q4"]),
            )

        # Back-fill actuals: find past predictions whose target_date has arrived
        _pred_conn.execute(
            """UPDATE regime_predictions SET actual_quadrant = (
                   SELECT rh.quadrant FROM regime_history rh
                   WHERE rh.date = regime_predictions.target_date
               )
               WHERE actual_quadrant IS NULL
                 AND target_date <= ?""",
            (_today_str,),
        )
        _pred_conn.commit()

        # Compute Brier scores per horizon
        brier_scores = {}
        for label in ["1_week", "4_week", "8_week"]:
            rows = _pred_conn.execute(
                """SELECT prob_q1, prob_q2, prob_q3, prob_q4, actual_quadrant
                   FROM regime_predictions
                   WHERE horizon = ? AND actual_quadrant IS NOT NULL""",
                (label,),
            ).fetchall()
            if len(rows) >= 2:
                total_bs = 0.0
                for row in rows:
                    probs = [row["prob_q1"], row["prob_q2"],
                             row["prob_q3"], row["prob_q4"]]
                    actual = row["actual_quadrant"]
                    # Brier score: sum of (predicted - actual)^2 across classes
                    for qi in range(4):
                        outcome = 1.0 if (qi + 1) == actual else 0.0
                        total_bs += (probs[qi] - outcome) ** 2
                total_bs /= len(rows)
                brier_scores[label] = {
                    "score": round(total_bs, 4),
                    "n_predictions": len(rows),
                    "rating": ("Excellent" if total_bs < 0.1 else
                               "Good" if total_bs < 0.25 else
                               "Fair" if total_bs < 0.5 else "Poor"),
                }
        _pred_conn.close()

        result = {
            "current_quadrant": current_quad,
            "current_quadrant_name": QUADRANT_NAMES[current_quad],
            "current_quadrant_description": QUADRANT_DESCRIPTIONS[current_quad],
            "confidence_pct": confidence,
            "classification_source": classification_source,
            "growth_score": round(float(current_growth), 4),
            "inflation_score": round(float(current_inflation), 4),
            "market_growth_score": round(float(g_vals[-1]), 4),
            "market_inflation_score": round(float(i_vals[-1]), 4),
            "transition_anchor_quadrant": transition_anchor_quad,
            "transition_anchor_name": QUADRANT_NAMES[transition_anchor_quad],
            "states_aligned": states_aligned,
            "fred_indicators": fred_indicators if fred_ok else None,
            "fred_growth_z": fred_growth_z,
            "fred_inflation_z": fred_inflation_z,
            "transition_matrix": adjusted_matrix,
            "static_transition_matrix": transition_matrix,
            "transition_counts": transition_counts.tolist(),
            "conditional_observations": int(cond_total) if use_conditional else None,
            "conditional_weight": round(float(conditional_weight), 3),
            "transition_smoothing_alpha": smoothing_alpha,
            "projections": projections,
            "brier_scores": brier_scores if brier_scores else None,
            "asset_tilts": QUADRANT_ASSET_TILTS[current_quad],
            "all_asset_tilts": QUADRANT_ASSET_TILTS,
            "interpretation": interpretation,
            "regime_distribution": {
                f"Q{k}": {"count": v, "pct": round(v / total_obs * 100, 1)}
                for k, v in regime_counts.items()
            },
            "history": {
                "dates": [d.strftime("%Y-%m-%d") for d in dates],
                "quadrants": regimes,
                "growth_scores": [round(float(v), 4) for v in g_vals],
                "inflation_scores": [round(float(v), 4) for v in i_vals],
            },
            "total_observations": total_obs,
        }

        _quadrant_cache["data"] = result
        _quadrant_cache["timestamp"] = time.time()
        return jsonify(result)

    except Exception as e:
        return jsonify({"error": str(e)}), 502


# ── S&P 500 Performance ───────────────────────────────────────────────────────

@app.route("/api/sp500-performance")
def sp500_performance():
    """Return S&P 500 YTD and 1-day percent change from Yahoo Finance."""
    import yfinance as yf
    from datetime import datetime, timedelta
    try:
        spy = yf.Ticker("^GSPC")
        today = datetime.now()
        year_start = datetime(today.year, 1, 1)
        hist = spy.history(start=year_start - timedelta(days=5), end=today + timedelta(days=1))
        if hist.empty:
            return jsonify({"error": "No S&P 500 data available"}), 502

        # YTD: compare latest close to last close of prior year
        first_of_year = hist.loc[hist.index >= str(year_start)]
        if first_of_year.empty:
            return jsonify({"error": "No YTD data"}), 502

        # Use the close before Jan 1 as the baseline
        prior = hist.loc[hist.index < str(year_start)]
        if not prior.empty:
            baseline = float(prior["Close"].iloc[-1])
        else:
            baseline = float(first_of_year["Close"].iloc[0])

        latest = float(hist["Close"].iloc[-1])
        ytd_pct = ((latest - baseline) / baseline) * 100

        # 1-day change
        if len(hist) >= 2:
            prev_close = float(hist["Close"].iloc[-2])
            day_pct = ((latest - prev_close) / prev_close) * 100
        else:
            day_pct = 0.0

        return jsonify({
            "price": round(latest, 2),
            "ytd_pct": round(ytd_pct, 2),
            "day_pct": round(day_pct, 2),
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 502


# ── Technical Scanner ─────────────────────────────────────────────────────────

@app.route("/api/scanner/tickers", methods=["GET", "POST"])
def scanner_tickers_list():
    """GET: return saved scanner tickers.  POST: bulk-replace list."""
    conn = get_connection()

    if request.method == "POST":
        data = request.get_json(force=True)
        rows = data.get("rows", [])
        conn.execute("DELETE FROM scanner_tickers")
        for i, r in enumerate(rows):
            ticker = str(r.get("ticker", "")).strip().upper()
            if not ticker:
                continue
            conn.execute(
                "INSERT OR IGNORE INTO scanner_tickers (ticker, sort_order) VALUES (?, ?)",
                (ticker, i),
            )
        conn.commit()
        conn.close()
        return jsonify(ok=True)

    # GET
    rows = conn.execute(
        "SELECT ticker, added_date FROM scanner_tickers ORDER BY sort_order, id"
    ).fetchall()
    conn.close()
    return jsonify(rows=[{"ticker": r["ticker"], "added_date": r["added_date"] or ""} for r in rows])


@app.route("/api/scanner/scan")
def scanner_scan():
    """Run technical scan on saved tickers."""
    import yfinance as yf
    import warnings
    warnings.filterwarnings("ignore")

    timeframe = request.args.get("timeframe", "daily")
    period = request.args.get("period", "1y" if timeframe == "daily" else "5y")
    interval = "1d" if timeframe == "daily" else "1wk"
    sma_pct = float(request.args.get("sma_pct", 5)) / 100
    stoch_min = float(request.args.get("stoch_min", 19))
    stoch_max = float(request.args.get("stoch_max", 21))

    conn = get_connection()
    ticker_rows = conn.execute("SELECT ticker FROM scanner_tickers ORDER BY sort_order, id").fetchall()
    conn.close()
    tickers = [r["ticker"] for r in ticker_rows]
    if not tickers:
        return jsonify(rows=[], error=None)

    try:
        df = yf.download(tickers, period=period, interval=interval,
                         auto_adjust=True, progress=False)
    except Exception as e:
        return jsonify(rows=[], error=str(e))

    multi = isinstance(df.columns, pd.MultiIndex)
    results = []
    for t in tickers:
        try:
            if multi:
                close = df["Close"][t].dropna()
                high = df["High"][t].dropna()
                low = df["Low"][t].dropna()
            else:
                close = df["Close"].dropna()
                high = df["High"].dropna()
                low = df["Low"].dropna()

            if len(close) < 20:
                results.append({"ticker": t, "price": None, "sma_50": None,
                                "sma_175": None, "slow_k": None, "slow_d": None,
                                "buy_signal": False, "error": "Insufficient data"})
                continue

            price = float(close.iloc[-1])

            sma_50_raw = close.rolling(50).mean().iloc[-1] if len(close) >= 50 else None
            sma_50_val = float(sma_50_raw) if sma_50_raw is not None and not pd.isna(sma_50_raw) else None

            sma_175_raw = close.rolling(175).mean().iloc[-1] if len(close) >= 175 else None
            sma_175_val = float(sma_175_raw) if sma_175_raw is not None and not pd.isna(sma_175_raw) else None

            slow_k, slow_d = _slow_stochastic(high, low, close)

            # Evaluate conditions (all must be computable for BUY)
            cond_sma_cross = sma_50_val is not None and sma_175_val is not None and sma_50_val >= sma_175_val
            cond_near_175 = sma_175_val is not None and abs(price - sma_175_val) / sma_175_val <= sma_pct
            cond_stoch = slow_k is not None and stoch_min <= slow_k <= stoch_max
            buy_signal = cond_sma_cross and cond_near_175 and cond_stoch

            results.append({
                "ticker": t,
                "price": round(price, 2),
                "sma_50": round(sma_50_val, 2) if sma_50_val is not None else None,
                "sma_175": round(sma_175_val, 2) if sma_175_val is not None else None,
                "slow_k": round(slow_k, 2) if slow_k is not None else None,
                "slow_d": round(slow_d, 2) if slow_d is not None else None,
                "buy_signal": buy_signal,
                "error": None,
            })
        except Exception:
            results.append({"ticker": t, "price": None, "sma_50": None,
                            "sma_175": None, "slow_k": None, "slow_d": None,
                            "buy_signal": False, "error": "Failed"})

    return jsonify(rows=results, error=None)


@app.route("/api/scanner/chart/<ticker>")
def scanner_chart(ticker):
    """Return Plotly chart JSON for a single ticker with scanner indicators."""
    import yfinance as yf
    import warnings
    warnings.filterwarnings("ignore")
    from plotly.subplots import make_subplots
    import plotly.graph_objects as go
    import json, plotly

    timeframe = request.args.get("timeframe", "daily")
    period = request.args.get("period", "1y" if timeframe == "daily" else "5y")
    interval = "1d" if timeframe == "daily" else "1wk"

    # Fetch extra history to warm up the 175 SMA so it spans the full display range
    warmup_map = {"1mo": "1y", "3mo": "2y", "6mo": "2y", "1y": "2y",
                  "2y": "5y", "3y": "5y", "5y": "10y", "10y": "max", "max": "max"}
    fetch_period = warmup_map.get(period, "max")

    try:
        df = yf.download(ticker, period=fetch_period, interval=interval,
                         auto_adjust=True, progress=False)
    except Exception as e:
        return jsonify(error=str(e)), 502

    if df.empty or len(df) < 50:
        return jsonify(error="Insufficient data for " + ticker), 400

    # Flatten MultiIndex columns if present
    if hasattr(df.columns, "levels"):
        df.columns = [c[0] if isinstance(c, tuple) else c for c in df.columns]

    # Compute indicators on full dataset (including warmup)
    close_full = df["Close"]
    high_full = df["High"]
    low_full = df["Low"]

    sma_50_full = close_full.rolling(50).mean()
    sma_175_full = close_full.rolling(175).mean()

    k_period, k_smooth, d_period = 14, 3, 3
    lowest_low = low_full.rolling(k_period).min()
    highest_high = high_full.rolling(k_period).max()
    denom = highest_high - lowest_low
    denom = denom.replace(0, float("nan"))
    raw_k = (close_full - lowest_low) / denom * 100
    slow_k_full = raw_k.rolling(k_smooth).mean()
    slow_d_full = slow_k_full.rolling(d_period).mean()

    # Trim to display period — find the cutoff date
    from dateutil.relativedelta import relativedelta
    import re
    display_start = None
    if period != "max":
        m = re.match(r"(\d+)(mo|y)", period)
        if m:
            n, unit = int(m.group(1)), m.group(2)
            if unit == "y":
                display_start = df.index[-1] - relativedelta(years=n)
            else:
                display_start = df.index[-1] - relativedelta(months=n)

    if display_start is not None:
        mask = df.index >= display_start
        df = df[mask]
        sma_50_full = sma_50_full[mask]
        sma_175_full = sma_175_full[mask]
        slow_k_full = slow_k_full[mask]
        slow_d_full = slow_d_full[mask]

    dates = list(df.index)
    close = df["Close"].tolist()
    high = df["High"].tolist()
    low = df["Low"].tolist()
    op = df["Open"].tolist()
    sma_50_list = sma_50_full.tolist()
    sma_175_list = sma_175_full.tolist()
    slow_k_list = slow_k_full.tolist()
    slow_d_list = slow_d_full.tolist()

    tf_label = "Daily" if timeframe == "daily" else "Weekly"
    fig = make_subplots(rows=2, cols=1, shared_xaxes=True,
                        row_heights=[0.7, 0.3], vertical_spacing=0.04,
                        subplot_titles=[f"{ticker} ({tf_label} {period})", "Slow Stochastic (14,3)"])

    fig.add_trace(go.Candlestick(x=dates, open=op, high=high, low=low, close=close,
                                  name="Price", increasing_line_color="#4dff91",
                                  decreasing_line_color="#ff6b6b"), row=1, col=1)
    fig.add_trace(go.Scatter(x=dates, y=sma_50_list, mode="lines",
                              name="50 SMA", line=dict(color="#7ecfff", width=1.5)), row=1, col=1)
    fig.add_trace(go.Scatter(x=dates, y=sma_175_list, mode="lines",
                              name="175 SMA", line=dict(color="#ff9800", width=1.5)), row=1, col=1)

    fig.add_trace(go.Scatter(x=dates, y=slow_k_list, mode="lines",
                              name="%K", line=dict(color="#7ecfff", width=1.2)), row=2, col=1)
    fig.add_trace(go.Scatter(x=dates, y=slow_d_list, mode="lines",
                              name="%D", line=dict(color="#ff9800", width=1.2)), row=2, col=1)

    # Overbought/oversold reference lines
    fig.add_hline(y=80, line_dash="dot", line_color="#8899aa", line_width=0.8, row=2, col=1)
    fig.add_hline(y=20, line_dash="dot", line_color="#8899aa", line_width=0.8, row=2, col=1)

    fig.update_layout(
        paper_bgcolor="#0e1117", plot_bgcolor="#0e1117",
        font=dict(color="#e0e8f5", size=12),
        xaxis_rangeslider_visible=False,
        showlegend=True,
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1,
                    font=dict(size=10)),
        margin=dict(l=50, r=20, t=50, b=30),
        height=560,
    )
    for ax in ["xaxis", "xaxis2", "yaxis", "yaxis2"]:
        fig.update_layout(**{ax: dict(gridcolor="#1a2233", zerolinecolor="#1a2233")})

    fig_json = json.loads(json.dumps(fig, cls=plotly.utils.PlotlyJSONEncoder))

    return jsonify(fig_data=fig_json["data"], fig_layout=fig_json["layout"], error=None)


@app.route("/api/general-scanner/chart/<ticker>")
def general_scanner_chart(ticker):
    """Return Plotly chart JSON with SMA 20/50/200, MACD, RSI, and Slow Stochastic."""
    import yfinance as yf
    import warnings
    warnings.filterwarnings("ignore")
    from plotly.subplots import make_subplots
    import plotly.graph_objects as go
    import json, plotly
    import numpy as np

    period = request.args.get("period", "1y")

    # Extra history for SMA warm-up
    warmup_map = {"1mo": "1y", "3mo": "2y", "6mo": "2y", "1y": "2y",
                  "2y": "5y", "3y": "5y", "5y": "10y", "10y": "max", "max": "max"}
    fetch_period = warmup_map.get(period, "max")

    try:
        df = yf.download(ticker, period=fetch_period, interval="1d",
                         auto_adjust=True, progress=False)
    except Exception as e:
        return jsonify(error=str(e)), 502

    if df.empty or len(df) < 30:
        return jsonify(error="Insufficient data for " + ticker), 400

    if hasattr(df.columns, "levels"):
        df.columns = [c[0] if isinstance(c, tuple) else c for c in df.columns]

    close_full = df["Close"]
    high_full = df["High"]
    low_full = df["Low"]

    # SMAs
    sma_20_full = close_full.rolling(20).mean()
    sma_50_full = close_full.rolling(50).mean()
    sma_200_full = close_full.rolling(200).mean()

    # MACD (12, 26, 9)
    ema_12 = close_full.ewm(span=12, adjust=False).mean()
    ema_26 = close_full.ewm(span=26, adjust=False).mean()
    macd_line = ema_12 - ema_26
    macd_signal = macd_line.ewm(span=9, adjust=False).mean()
    macd_hist = macd_line - macd_signal

    # RSI (14)
    delta = close_full.diff()
    gain = delta.where(delta > 0, 0.0).rolling(14).mean()
    loss = (-delta.where(delta < 0, 0.0)).rolling(14).mean()
    rs = gain / loss.replace(0, float("nan"))
    rsi_full = 100 - (100 / (1 + rs))

    # Slow Stochastic (14, 3, 3)
    lowest_low = low_full.rolling(14).min()
    highest_high = high_full.rolling(14).max()
    denom = (highest_high - lowest_low).replace(0, float("nan"))
    raw_k = (close_full - lowest_low) / denom * 100
    slow_k_full = raw_k.rolling(3).mean()
    slow_d_full = slow_k_full.rolling(3).mean()

    # Trim to display period
    from dateutil.relativedelta import relativedelta
    import re
    display_start = None
    if period != "max":
        m = re.match(r"(\d+)(mo|y)", period)
        if m:
            n, unit = int(m.group(1)), m.group(2)
            if unit == "y":
                display_start = df.index[-1] - relativedelta(years=n)
            else:
                display_start = df.index[-1] - relativedelta(months=n)

    if display_start is not None:
        mask = df.index >= display_start
        df = df[mask]
        sma_20_full = sma_20_full[mask]
        sma_50_full = sma_50_full[mask]
        sma_200_full = sma_200_full[mask]
        macd_line = macd_line[mask]
        macd_signal = macd_signal[mask]
        macd_hist = macd_hist[mask]
        rsi_full = rsi_full[mask]
        slow_k_full = slow_k_full[mask]
        slow_d_full = slow_d_full[mask]

    dates = list(df.index)

    fig = make_subplots(
        rows=4, cols=1, shared_xaxes=True,
        row_heights=[0.45, 0.18, 0.18, 0.19], vertical_spacing=0.03,
        subplot_titles=[f"{ticker} (Daily {period})", "MACD (12,26,9)",
                        "RSI (14)", "Slow Stochastic (14,3,3)"],
    )

    # Row 1: Candlestick + SMAs
    fig.add_trace(go.Candlestick(
        x=dates, open=df["Open"], high=df["High"], low=df["Low"], close=df["Close"],
        name="Price", increasing_line_color="#4dff91", decreasing_line_color="#ff6b6b",
    ), row=1, col=1)
    fig.add_trace(go.Scatter(x=dates, y=sma_20_full, mode="lines",
        name="SMA 20", line=dict(color="#ffeb3b", width=1.2)), row=1, col=1)
    fig.add_trace(go.Scatter(x=dates, y=sma_50_full, mode="lines",
        name="SMA 50", line=dict(color="#7ecfff", width=1.2)), row=1, col=1)
    fig.add_trace(go.Scatter(x=dates, y=sma_200_full, mode="lines",
        name="SMA 200", line=dict(color="#ff9800", width=1.5)), row=1, col=1)

    # Row 2: MACD
    hist_colors = ["#4dff91" if v >= 0 else "#ff6b6b" for v in macd_hist.fillna(0)]
    fig.add_trace(go.Bar(x=dates, y=macd_hist, name="Histogram",
        marker_color=hist_colors, showlegend=False), row=2, col=1)
    fig.add_trace(go.Scatter(x=dates, y=macd_line, mode="lines",
        name="MACD", line=dict(color="#7ecfff", width=1.2)), row=2, col=1)
    fig.add_trace(go.Scatter(x=dates, y=macd_signal, mode="lines",
        name="Signal", line=dict(color="#ff9800", width=1.2)), row=2, col=1)
    fig.add_hline(y=0, line_dash="dot", line_color="#555", line_width=0.6, row=2, col=1)

    # Row 3: RSI
    fig.add_trace(go.Scatter(x=dates, y=rsi_full, mode="lines",
        name="RSI 14", line=dict(color="#b388ff", width=1.2)), row=3, col=1)
    fig.add_hline(y=70, line_dash="dot", line_color="#ff6b6b", line_width=0.7, row=3, col=1)
    fig.add_hline(y=30, line_dash="dot", line_color="#4dff91", line_width=0.7, row=3, col=1)
    fig.add_hline(y=50, line_dash="dot", line_color="#555", line_width=0.5, row=3, col=1)

    # Row 4: Slow Stochastic
    fig.add_trace(go.Scatter(x=dates, y=slow_k_full, mode="lines",
        name="%K", line=dict(color="#7ecfff", width=1.2)), row=4, col=1)
    fig.add_trace(go.Scatter(x=dates, y=slow_d_full, mode="lines",
        name="%D", line=dict(color="#ff9800", width=1.2)), row=4, col=1)
    fig.add_hline(y=80, line_dash="dot", line_color="#ff6b6b", line_width=0.7, row=4, col=1)
    fig.add_hline(y=20, line_dash="dot", line_color="#4dff91", line_width=0.7, row=4, col=1)

    fig.update_layout(
        paper_bgcolor="#0e1117", plot_bgcolor="#0e1117",
        font=dict(color="#e0e8f5", size=11),
        xaxis_rangeslider_visible=False,
        showlegend=True,
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1,
                    font=dict(size=10)),
        margin=dict(l=50, r=20, t=50, b=30),
        height=780,
    )
    for ax in ["xaxis", "xaxis2", "xaxis3", "xaxis4",
               "yaxis", "yaxis2", "yaxis3", "yaxis4"]:
        fig.update_layout(**{ax: dict(gridcolor="#1a2233", zerolinecolor="#1a2233")})

    fig_json = json.loads(json.dumps(fig, cls=plotly.utils.PlotlyJSONEncoder))
    return jsonify(fig_data=fig_json["data"], fig_layout=fig_json["layout"], error=None)


# ── General Scanner ────────────────────────────────────────────────────────────

@app.route("/api/general-scanner/universe", methods=["GET", "POST", "DELETE"])
def general_scanner_universe():
    """Manage the ticker universe for the general scanner."""
    conn = get_connection()

    if request.method == "POST":
        data = request.get_json(force=True)
        tickers = data.get("tickers", [])
        asset_type = data.get("asset_type", "Stock")
        added = 0
        for t in tickers:
            t = str(t).strip().upper()
            if not t:
                continue
            conn.execute(
                "INSERT OR IGNORE INTO general_scanner_universe (ticker, asset_type) VALUES (?, ?)",
                (t, asset_type),
            )
            added += 1
        conn.commit()
        conn.close()
        return jsonify(ok=True, added=added)

    if request.method == "DELETE":
        data = request.get_json(force=True)
        tickers = data.get("tickers", [])
        if tickers:
            placeholders = ",".join("?" for _ in tickers)
            conn.execute(f"DELETE FROM general_scanner_universe WHERE ticker IN ({placeholders})",
                         [t.upper() for t in tickers])
            conn.execute(f"DELETE FROM general_scanner_cache WHERE ticker IN ({placeholders})",
                         [t.upper() for t in tickers])
        conn.commit()
        conn.close()
        return jsonify(ok=True)

    # GET
    rows = conn.execute(
        "SELECT ticker, asset_type, added_date FROM general_scanner_universe ORDER BY ticker"
    ).fetchall()
    conn.close()
    return jsonify(rows=[{"ticker": r["ticker"], "asset_type": r["asset_type"]} for r in rows])


@app.route("/api/general-scanner/refresh", methods=["POST"])
def general_scanner_refresh():
    """Fetch/refresh data from yfinance for all universe tickers. Uses ThreadPoolExecutor."""
    import yfinance as yf
    import warnings
    warnings.filterwarnings("ignore")
    from concurrent.futures import ThreadPoolExecutor, as_completed
    import numpy as np

    conn = get_connection()
    rows = conn.execute("SELECT ticker, asset_type FROM general_scanner_universe ORDER BY ticker").fetchall()
    # Find tickers that already have cached info (name is not null/empty)
    cached = conn.execute("SELECT ticker FROM general_scanner_cache WHERE name IS NOT NULL AND name != ''").fetchall()
    cached_set = {r["ticker"] for r in cached}
    conn.close()
    tickers = [r["ticker"] for r in rows]
    type_map = {r["ticker"]: r["asset_type"] for r in rows}
    # Only fetch .info for tickers not yet cached (or use force param to refetch all)
    force_info = request.args.get("force", "false").lower() == "true"
    info_tickers = tickers if force_info else [t for t in tickers if t not in cached_set]

    if not tickers:
        return jsonify(ok=True, refreshed=0, errors=[])

    import time as _time

    # ── 1. Bulk download price history in batches to avoid rate limits ────────
    BATCH_SIZE = 50
    all_hist = {}  # ticker -> {Close: Series, Volume: Series}

    for i in range(0, len(tickers), BATCH_SIZE):
        batch = tickers[i:i + BATCH_SIZE]
        try:
            hist = yf.download(batch, period="1y", interval="1d",
                               auto_adjust=True, progress=False, threads=True)
            multi = len(batch) > 1
            for t in batch:
                try:
                    if multi:
                        close = hist["Close"][t].dropna()
                        vol = hist["Volume"][t].dropna()
                        high = hist["High"][t].dropna()
                        low = hist["Low"][t].dropna()
                    else:
                        close = hist["Close"].dropna()
                        vol = hist["Volume"].dropna()
                        high = hist["High"].dropna()
                        low = hist["Low"].dropna()
                    all_hist[t] = {"Close": close, "Volume": vol, "High": high, "Low": low}
                except Exception:
                    pass
        except Exception:
            pass
        if i + BATCH_SIZE < len(tickers):
            _time.sleep(2)  # pause between batches to avoid rate limits

    def _compute_technicals(t):
        """Compute SMA-20/50/200, RSI-14, MACD, Slow Stochastic, change% from downloaded price data."""
        try:
            data = all_hist.get(t)
            if not data:
                return t, {}
            close = data["Close"]
            vol = data["Volume"]

            if len(close) < 2:
                return t, {}

            price = float(close.iloc[-1])
            prev = float(close.iloc[-2])
            change_pct = round((price - prev) / prev * 100, 2) if prev else None

            sma_20 = round(float(close.rolling(20).mean().iloc[-1]), 2) if len(close) >= 20 else None
            sma_50 = round(float(close.rolling(50).mean().iloc[-1]), 2) if len(close) >= 50 else None
            sma_200 = round(float(close.rolling(200).mean().iloc[-1]), 2) if len(close) >= 200 else None

            # Previous-day SMAs (for cross detection)
            prev_sma_20 = round(float(close.rolling(20).mean().iloc[-2]), 2) if len(close) >= 21 else None
            prev_sma_50 = round(float(close.rolling(50).mean().iloc[-2]), 2) if len(close) >= 51 else None
            prev_sma_200 = round(float(close.rolling(200).mean().iloc[-2]), 2) if len(close) >= 201 else None

            # RSI 14
            rsi_14 = None
            if len(close) >= 15:
                delta = close.diff()
                gain = delta.where(delta > 0, 0.0).rolling(14).mean().iloc[-1]
                loss = (-delta.where(delta < 0, 0.0)).rolling(14).mean().iloc[-1]
                if loss != 0:
                    rs = gain / loss
                    rsi_14 = round(100 - (100 / (1 + rs)), 2)

            # MACD (12, 26, 9)
            macd_line = macd_signal = macd_hist = None
            if len(close) >= 35:
                ema_12 = close.ewm(span=12, adjust=False).mean()
                ema_26 = close.ewm(span=26, adjust=False).mean()
                macd_series = ema_12 - ema_26
                signal_series = macd_series.ewm(span=9, adjust=False).mean()
                hist_series = macd_series - signal_series
                macd_line = round(float(macd_series.iloc[-1]), 4)
                macd_signal = round(float(signal_series.iloc[-1]), 4)
                macd_hist = round(float(hist_series.iloc[-1]), 4)
                # Previous day for cross detection
                prev_macd = round(float(macd_series.iloc[-2]), 4)
                prev_signal = round(float(signal_series.iloc[-2]), 4)

            # Slow Stochastic (14, 3, 3)
            stoch_k = stoch_d = None
            if len(close) >= 17 and "High" in data and "Low" in data:
                high = data["High"]
                low = data["Low"]
                low_14 = low.rolling(14).min()
                high_14 = high.rolling(14).max()
                fast_k = 100 * (close - low_14) / (high_14 - low_14)
                slow_k = fast_k.rolling(3).mean()  # %K (smoothed)
                slow_d = slow_k.rolling(3).mean()  # %D
                stoch_k = round(float(slow_k.iloc[-1]), 2) if not np.isnan(slow_k.iloc[-1]) else None
                stoch_d = round(float(slow_d.iloc[-1]), 2) if not np.isnan(slow_d.iloc[-1]) else None

            avg_volume = round(float(vol.tail(50).mean()), 0) if len(vol) >= 5 else None
            volume = float(vol.iloc[-1]) if len(vol) > 0 else None
            week52_high = round(float(close.tail(252).max()), 2) if len(close) >= 20 else None
            week52_low = round(float(close.tail(252).min()), 2) if len(close) >= 20 else None

            return t, {
                "price": round(price, 2), "change_pct": change_pct,
                "sma_20": sma_20, "sma_50": sma_50, "sma_200": sma_200,
                "prev_sma_20": prev_sma_20, "prev_sma_50": prev_sma_50, "prev_sma_200": prev_sma_200,
                "rsi_14": rsi_14,
                "macd_line": macd_line, "macd_signal": macd_signal, "macd_hist": macd_hist,
                "stoch_k": stoch_k, "stoch_d": stoch_d,
                "avg_volume": avg_volume, "volume": volume,
                "week52_high": week52_high, "week52_low": week52_low,
            }
        except Exception:
            return t, {}

    tech_data = {}
    for t in tickers:
        ticker, vals = _compute_technicals(t)
        tech_data[ticker] = vals

    # ── 2. Fetch fundamentals via Ticker.info (batched + threaded) ───────────
    errors = []
    fund_data = {}

    # Ticker-based overrides for ETFs that yfinance doesn't classify well
    _TICKER_STRATEGY_OVERRIDES = {
        # Options / Covered Call Income
        **{t: "Options Income" for t in [
            "JEPI", "JEPQ", "QYLD", "XYLD", "RYLD", "DJIA", "SPYI", "QQQI",
            "IWMI", "GPIQ", "GPIX", "FTQI", "SVOL", "TLTW", "KLIP", "YMAX", "YMAG",
            "CONY", "TSLY", "NVDY", "AMZY", "APLY", "MSFO", "GOOY", "FBY", "OARK",
            "JEPY", "AIPI", "FEPI", "GIAX", "USOI", "QQQY", "XDTE", "QDTE", "RDTE",
            "WDTE", "ULTY", "BALI", "ISPY", "JEPX", "SPXX", "QQXX", "IWMW",
            "DIVO", "NUSI", "PUTW",
        ]},
        # CEF Income
        **{t: "CEF" for t in [
            "PCEF", "CEFS", "YYY", "XMPT", "FCEF", "ADX", "ASGI", "BST",
        ]},
        # BDC
        **{t: "BDC" for t in [
            "BIZD", "PBDC",
            "MAIN", "ARCC", "HTGC", "GBDC", "BXSL", "TPVG", "GSBD", "OBDC", "ORCC", "FSK",
            "NEWT", "GAIN", "GLAD", "PSEC", "SLRC", "OCSL", "CGBD", "FDUS", "MFIC", "CSWC",
        ]},
        # Leveraged Loans / Senior Loans
        **{t: "Loans" for t in [
            "BKLN", "SRLN", "FLBL", "FTSL",
        ]},
        # Preferred Stock
        **{t: "Preferred" for t in [
            "PFF", "PFFD", "PGX", "PSK", "FPE", "PFFV", "VRP", "FPEI",
            "SPFF", "PREF", "EPRF",
        ]},
    }

    def _classify_etf(category, ticker=""):
        """Derive strategy and cap size from yfinance ETF category string."""
        # Check ticker-based override first
        override = _TICKER_STRATEGY_OVERRIDES.get(ticker, "")

        if not category and not override:
            return "", override or "", ""
        cat = (category or "").lower()

        # ── Strategy ──
        strategy = ""
        if override:
            strategy = override
        elif any(w in cat for w in ["derivative income", "covered call", "option"]):
            strategy = "Options Income"
        elif "preferred" in cat:
            strategy = "Preferred"
        elif any(w in cat for w in ["income", "dividend", "yield", "high yield equity"]):
            strategy = "Dividend"
        elif "bond" in cat or "government" in cat or "treasury" in cat or "fixed" in cat or "inflation" in cat or "mortgage" in cat:
            strategy = "Bonds"
        elif "growth" in cat:
            strategy = "Growth"
        elif "value" in cat:
            strategy = "Value"
        elif "blend" in cat:
            strategy = "Blend"
        elif any(w in cat for w in ["commodit", "precious", "gold", "silver", "metal"]):
            strategy = "Commodity"
        elif "real estate" in cat or "reit" in cat:
            strategy = "Real Estate"
        elif any(w in cat for w in ["emerging", "international", "foreign", "world", "global", "china", "europe", "japan", "pacific"]):
            strategy = "International"
        elif any(w in cat for w in ["technology", "health", "energy", "financial", "utilities", "consumer", "industrial", "communication"]):
            strategy = "Sector"
        elif "target" in cat:
            strategy = "Target Date"
        else:
            strategy = "Other"

        # ── Cap size ──
        cap_size = ""
        if "large" in cat:
            cap_size = "Large Cap"
        elif "mid" in cat:
            cap_size = "Mid Cap"
        elif "small" in cat:
            cap_size = "Small Cap"
        elif "micro" in cat:
            cap_size = "Micro Cap"
        elif any(w in cat for w in ["total market", "all cap", "blend"]) and strategy in ("Growth", "Value", "Blend"):
            cap_size = "All Cap"

        return category or "", strategy, cap_size

    def _fetch_info(t):
        try:
            info = yf.Ticker(t).info or {}
            etf_cat = info.get("category", "")
            etf_category, etf_strategy, etf_cap_size = _classify_etf(etf_cat, t)
            return t, {
                "name": info.get("shortName") or info.get("longName", ""),
                "sector": info.get("sector", ""),
                "industry": info.get("industry", ""),
                "country": info.get("country", ""),
                "market_cap": info.get("marketCap"),
                "pe_ratio": info.get("trailingPE"),
                "forward_pe": info.get("forwardPE"),
                "peg_ratio": info.get("trailingPegRatio"),
                "ps_ratio": info.get("priceToSalesTrailing12Months"),
                "pb_ratio": info.get("priceToBook"),
                "dividend_yield": round(info.get("dividendYield", 0) * 100, 2) if info.get("dividendYield") else None,
                "eps": info.get("trailingEps"),
                "revenue": info.get("totalRevenue"),
                "profit_margin": round(info.get("profitMargins", 0) * 100, 2) if info.get("profitMargins") else None,
                "roe": round(info.get("returnOnEquity", 0) * 100, 2) if info.get("returnOnEquity") else None,
                "debt_to_equity": info.get("debtToEquity"),
                "current_ratio": info.get("currentRatio"),
                "beta": info.get("beta"),
                "expense_ratio": round(info.get("annualReportExpenseRatio", 0) * 100, 4) if info.get("annualReportExpenseRatio") else None,
                "aum": info.get("totalAssets"),
                "etf_category": etf_category,
                "etf_strategy": etf_strategy,
                "etf_cap_size": etf_cap_size,
            }
        except Exception as exc:
            return t, {"_error": str(exc)}

    INFO_BATCH = 10
    for i in range(0, len(info_tickers), INFO_BATCH):
        batch = info_tickers[i:i + INFO_BATCH]
        with ThreadPoolExecutor(max_workers=2) as pool:
            futures = {pool.submit(_fetch_info, t): t for t in batch}
            for fut in as_completed(futures):
                t, result = fut.result()
                if "_error" in result:
                    errors.append({"ticker": t, "error": result["_error"]})
                else:
                    fund_data[t] = result
        if i + INFO_BATCH < len(info_tickers):
            _time.sleep(1)  # pause between info batches to avoid rate limits

    # ── 3. Merge and upsert into cache ───────────────────────────────────────
    conn = get_connection()
    refreshed = 0
    for t in tickers:
        tech = tech_data.get(t, {})
        fund = fund_data.get(t, {})
        if not tech and not fund:
            continue
        # Apply ticker-based strategy override if info wasn't fetched
        if "etf_strategy" not in fund and t in _TICKER_STRATEGY_OVERRIDES:
            fund["etf_strategy"] = _TICKER_STRATEGY_OVERRIDES[t]
        merged = {**tech, **fund, "asset_type": type_map.get(t, "Stock")}
        cols = ["ticker"] + list(merged.keys()) + ["updated_at"]
        vals = [t] + list(merged.values()) + [pd.Timestamp.now().isoformat()]
        placeholders = ",".join("?" for _ in cols)
        upsert_cols = ",".join(cols)
        update_set = ",".join(f"{c}=excluded.{c}" for c in cols if c != "ticker")
        conn.execute(
            f"INSERT INTO general_scanner_cache ({upsert_cols}) VALUES ({placeholders}) "
            f"ON CONFLICT(ticker) DO UPDATE SET {update_set}",
            vals,
        )
        refreshed += 1

    conn.commit()
    conn.close()
    return jsonify(ok=True, refreshed=refreshed, errors=errors,
                   info_fetched=len(fund_data), info_skipped=len(cached_set) if not force_info else 0)


@app.route("/api/general-scanner/scan")
def general_scanner_scan():
    """Return cached scanner data with server-side filtering, sorting, and pagination."""
    conn = get_connection()

    # Filters
    asset_type = request.args.get("asset_type", "")
    sector = request.args.get("sector", "")
    industry = request.args.get("industry", "")
    country = request.args.get("country", "")
    etf_strategy = request.args.get("etf_strategy", "")
    etf_cap_size = request.args.get("etf_cap_size", "")

    # Range filters: param_min / param_max
    range_cols = [
        "market_cap", "price", "pe_ratio", "forward_pe", "peg_ratio", "ps_ratio",
        "pb_ratio", "dividend_yield", "eps", "profit_margin", "roe",
        "debt_to_equity", "current_ratio", "beta", "rsi_14", "change_pct",
        "volume", "avg_volume", "expense_ratio", "sma_20", "sma_50", "sma_200",
    ]

    where = []
    params = []

    if asset_type:
        where.append("c.asset_type = ?")
        params.append(asset_type)
    if sector:
        where.append("c.sector = ?")
        params.append(sector)
    if industry:
        where.append("c.industry = ?")
        params.append(industry)
    if country:
        where.append("c.country = ?")
        params.append(country)
    if etf_strategy:
        where.append("c.etf_strategy = ?")
        params.append(etf_strategy)
    if etf_cap_size:
        where.append("c.etf_cap_size = ?")
        params.append(etf_cap_size)

    for col in range_cols:
        lo = request.args.get(f"{col}_min")
        hi = request.args.get(f"{col}_max")
        if lo is not None and lo != "":
            where.append(f"c.{col} >= ?")
            params.append(float(lo))
        if hi is not None and hi != "":
            where.append(f"c.{col} <= ?")
            params.append(float(hi))

    # ── 20-Day SMA filter ──
    sma20_f = request.args.get("sma20_filter", "")
    _sma_map_20 = {
        "price_above":       "c.price > c.sma_20 AND c.sma_20 IS NOT NULL",
        "price_below":       "c.price < c.sma_20 AND c.sma_20 IS NOT NULL",
        "price_crossed_above": "c.price > c.sma_20 AND c.prev_sma_20 IS NOT NULL AND c.price - c.sma_20 >= 0 AND (SELECT c2.price FROM general_scanner_cache c2 WHERE c2.ticker = c.ticker) IS NOT NULL AND c.prev_sma_20 IS NOT NULL",
        "price_10pct_above": "c.sma_20 IS NOT NULL AND c.price > c.sma_20 * 1.10",
        "price_10pct_below": "c.sma_20 IS NOT NULL AND c.price < c.sma_20 * 0.90",
        "price_20pct_above": "c.sma_20 IS NOT NULL AND c.price > c.sma_20 * 1.20",
        "price_20pct_below": "c.sma_20 IS NOT NULL AND c.price < c.sma_20 * 0.80",
        "price_30pct_below": "c.sma_20 IS NOT NULL AND c.price < c.sma_20 * 0.70",
        "price_50pct_above": "c.sma_20 IS NOT NULL AND c.price > c.sma_20 * 1.50",
        "price_50pct_below": "c.sma_20 IS NOT NULL AND c.price < c.sma_20 * 0.50",
    }
    if sma20_f in _sma_map_20:
        where.append(_sma_map_20[sma20_f])

    # ── 50-Day SMA filter ──
    sma50_f = request.args.get("sma50_filter", "")
    _sma_map_50 = {
        "price_above":       "c.price > c.sma_50 AND c.sma_50 IS NOT NULL",
        "price_below":       "c.price < c.sma_50 AND c.sma_50 IS NOT NULL",
        "price_10pct_above": "c.sma_50 IS NOT NULL AND c.price > c.sma_50 * 1.10",
        "price_10pct_below": "c.sma_50 IS NOT NULL AND c.price < c.sma_50 * 0.90",
        "price_20pct_above": "c.sma_50 IS NOT NULL AND c.price > c.sma_50 * 1.20",
        "price_20pct_below": "c.sma_50 IS NOT NULL AND c.price < c.sma_50 * 0.80",
        "price_30pct_below": "c.sma_50 IS NOT NULL AND c.price < c.sma_50 * 0.70",
        "price_50pct_above": "c.sma_50 IS NOT NULL AND c.price > c.sma_50 * 1.50",
        "price_50pct_below": "c.sma_50 IS NOT NULL AND c.price < c.sma_50 * 0.50",
        "sma20_above":       "c.sma_20 IS NOT NULL AND c.sma_50 IS NOT NULL AND c.sma_20 > c.sma_50",
        "sma20_below":       "c.sma_20 IS NOT NULL AND c.sma_50 IS NOT NULL AND c.sma_20 < c.sma_50",
        "sma20_cross_above": "c.sma_20 IS NOT NULL AND c.sma_50 IS NOT NULL AND c.prev_sma_20 IS NOT NULL AND c.prev_sma_50 IS NOT NULL AND c.sma_20 > c.sma_50 AND c.prev_sma_20 <= c.prev_sma_50",
        "sma20_cross_below": "c.sma_20 IS NOT NULL AND c.sma_50 IS NOT NULL AND c.prev_sma_20 IS NOT NULL AND c.prev_sma_50 IS NOT NULL AND c.sma_20 < c.sma_50 AND c.prev_sma_20 >= c.prev_sma_50",
    }
    if sma50_f in _sma_map_50:
        where.append(_sma_map_50[sma50_f])

    # ── 200-Day SMA filter ──
    sma200_f = request.args.get("sma200_filter", "")
    _sma_map_200 = {
        "price_above":       "c.price > c.sma_200 AND c.sma_200 IS NOT NULL",
        "price_below":       "c.price < c.sma_200 AND c.sma_200 IS NOT NULL",
        "price_10pct_above": "c.sma_200 IS NOT NULL AND c.price > c.sma_200 * 1.10",
        "price_10pct_below": "c.sma_200 IS NOT NULL AND c.price < c.sma_200 * 0.90",
        "price_20pct_above": "c.sma_200 IS NOT NULL AND c.price > c.sma_200 * 1.20",
        "price_20pct_below": "c.sma_200 IS NOT NULL AND c.price < c.sma_200 * 0.80",
        "price_30pct_below": "c.sma_200 IS NOT NULL AND c.price < c.sma_200 * 0.70",
        "price_50pct_above": "c.sma_200 IS NOT NULL AND c.price > c.sma_200 * 1.50",
        "price_50pct_below": "c.sma_200 IS NOT NULL AND c.price < c.sma_200 * 0.50",
        "sma50_above":       "c.sma_50 IS NOT NULL AND c.sma_200 IS NOT NULL AND c.sma_50 > c.sma_200",
        "sma50_below":       "c.sma_50 IS NOT NULL AND c.sma_200 IS NOT NULL AND c.sma_50 < c.sma_200",
        "sma50_cross_above": "c.sma_50 IS NOT NULL AND c.sma_200 IS NOT NULL AND c.prev_sma_50 IS NOT NULL AND c.prev_sma_200 IS NOT NULL AND c.sma_50 > c.sma_200 AND c.prev_sma_50 <= c.prev_sma_200",
        "sma50_cross_below": "c.sma_50 IS NOT NULL AND c.sma_200 IS NOT NULL AND c.prev_sma_50 IS NOT NULL AND c.prev_sma_200 IS NOT NULL AND c.sma_50 < c.sma_200 AND c.prev_sma_50 >= c.prev_sma_200",
        "sma20_above":       "c.sma_20 IS NOT NULL AND c.sma_200 IS NOT NULL AND c.sma_20 > c.sma_200",
        "sma20_below":       "c.sma_20 IS NOT NULL AND c.sma_200 IS NOT NULL AND c.sma_20 < c.sma_200",
    }
    if sma200_f in _sma_map_200:
        where.append(_sma_map_200[sma200_f])

    # ── Combined SMA alignment filters ──
    sma_align = request.args.get("sma_alignment", "")
    if sma_align == "20_above_50_above_200":
        where.append("c.sma_20 IS NOT NULL AND c.sma_50 IS NOT NULL AND c.sma_200 IS NOT NULL AND c.sma_20 > c.sma_50 AND c.sma_50 > c.sma_200")
    elif sma_align == "20_below_50_below_200":
        where.append("c.sma_20 IS NOT NULL AND c.sma_50 IS NOT NULL AND c.sma_200 IS NOT NULL AND c.sma_20 < c.sma_50 AND c.sma_50 < c.sma_200")
    elif sma_align == "price_above_all":
        where.append("c.sma_20 IS NOT NULL AND c.sma_50 IS NOT NULL AND c.sma_200 IS NOT NULL AND c.price > c.sma_20 AND c.price > c.sma_50 AND c.price > c.sma_200")
    elif sma_align == "price_below_all":
        where.append("c.sma_20 IS NOT NULL AND c.sma_50 IS NOT NULL AND c.sma_200 IS NOT NULL AND c.price < c.sma_20 AND c.price < c.sma_50 AND c.price < c.sma_200")

    # ── MACD filter ──
    macd_f = request.args.get("macd_filter", "")
    _macd_map = {
        "bullish":           "c.macd_line IS NOT NULL AND c.macd_line > c.macd_signal",
        "bearish":           "c.macd_line IS NOT NULL AND c.macd_line < c.macd_signal",
        "bullish_cross":     "c.macd_line IS NOT NULL AND c.macd_hist > 0 AND c.macd_hist IS NOT NULL",
        "bearish_cross":     "c.macd_line IS NOT NULL AND c.macd_hist < 0 AND c.macd_hist IS NOT NULL",
        "positive":          "c.macd_line IS NOT NULL AND c.macd_line > 0",
        "negative":          "c.macd_line IS NOT NULL AND c.macd_line < 0",
    }
    if macd_f in _macd_map:
        where.append(_macd_map[macd_f])

    # ── Stochastic filter ──
    stoch_f = request.args.get("stoch_filter", "")
    _stoch_map = {
        "overbought":    "c.stoch_k IS NOT NULL AND c.stoch_k > 80",
        "oversold":      "c.stoch_k IS NOT NULL AND c.stoch_k < 20",
        "bullish":       "c.stoch_k IS NOT NULL AND c.stoch_d IS NOT NULL AND c.stoch_k > c.stoch_d",
        "bearish":       "c.stoch_k IS NOT NULL AND c.stoch_d IS NOT NULL AND c.stoch_k < c.stoch_d",
        "neutral":       "c.stoch_k IS NOT NULL AND c.stoch_k >= 20 AND c.stoch_k <= 80",
    }
    if stoch_f in _stoch_map:
        where.append(_stoch_map[stoch_f])

    # Keep backwards compatibility with old checkbox filters
    if request.args.get("above_sma_20") == "true":
        where.append("c.price > c.sma_20 AND c.sma_20 IS NOT NULL")
    if request.args.get("below_sma_20") == "true":
        where.append("c.price < c.sma_20 AND c.sma_20 IS NOT NULL")
    if request.args.get("above_sma_50") == "true":
        where.append("c.price > c.sma_50 AND c.sma_50 IS NOT NULL")
    if request.args.get("below_sma_50") == "true":
        where.append("c.price < c.sma_50 AND c.sma_50 IS NOT NULL")
    if request.args.get("above_sma_200") == "true":
        where.append("c.price > c.sma_200 AND c.sma_200 IS NOT NULL")
    if request.args.get("below_sma_200") == "true":
        where.append("c.price < c.sma_200 AND c.sma_200 IS NOT NULL")

    # 52-week filters
    pct_52hi = request.args.get("pct_from_52high")
    if pct_52hi:
        where.append("c.week52_high > 0 AND ((c.week52_high - c.price) / c.week52_high * 100) <= ?")
        params.append(float(pct_52hi))
    pct_52lo = request.args.get("pct_from_52low")
    if pct_52lo:
        where.append("c.week52_low > 0 AND ((c.price - c.week52_low) / c.week52_low * 100) <= ?")
        params.append(float(pct_52lo))

    # Sorting
    sort_col = request.args.get("sort", "ticker")
    sort_dir = "DESC" if request.args.get("dir", "asc").lower() == "desc" else "ASC"
    allowed_sort = {
        "ticker", "name", "sector", "industry", "country", "market_cap", "price",
        "pe_ratio", "forward_pe", "peg_ratio", "ps_ratio", "pb_ratio",
        "dividend_yield", "eps", "revenue", "profit_margin", "roe",
        "debt_to_equity", "current_ratio", "beta", "week52_high", "week52_low",
        "avg_volume", "sma_20", "sma_50", "sma_200", "rsi_14", "change_pct",
        "volume", "asset_type", "expense_ratio", "aum",
        "etf_category", "etf_strategy", "etf_cap_size",
        "macd_line", "macd_signal", "macd_hist", "stoch_k", "stoch_d",
    }
    if sort_col not in allowed_sort:
        sort_col = "ticker"

    # Pagination
    page = max(1, int(request.args.get("page", 1)))
    per_page = min(200, max(10, int(request.args.get("per_page", 50))))
    offset = (page - 1) * per_page

    where_clause = (" WHERE " + " AND ".join(where)) if where else ""

    count_sql = f"SELECT COUNT(*) FROM general_scanner_cache c{where_clause}"
    total = conn.execute(count_sql, params).fetchone()[0]

    data_sql = (
        f"SELECT * FROM general_scanner_cache c{where_clause} "
        f"ORDER BY c.{sort_col} {sort_dir} LIMIT ? OFFSET ?"
    )
    rows = conn.execute(data_sql, params + [per_page, offset]).fetchall()

    # Also get distinct values for dropdown filters
    sectors = [r[0] for r in conn.execute(
        "SELECT DISTINCT sector FROM general_scanner_cache WHERE sector != '' ORDER BY sector"
    ).fetchall()]
    industries = [r[0] for r in conn.execute(
        "SELECT DISTINCT industry FROM general_scanner_cache WHERE industry != '' ORDER BY industry"
    ).fetchall()]
    countries = [r[0] for r in conn.execute(
        "SELECT DISTINCT country FROM general_scanner_cache WHERE country != '' ORDER BY country"
    ).fetchall()]
    etf_strategies = [r[0] for r in conn.execute(
        "SELECT DISTINCT etf_strategy FROM general_scanner_cache WHERE etf_strategy IS NOT NULL AND etf_strategy != '' ORDER BY etf_strategy"
    ).fetchall()]
    etf_cap_sizes = [r[0] for r in conn.execute(
        "SELECT DISTINCT etf_cap_size FROM general_scanner_cache WHERE etf_cap_size IS NOT NULL AND etf_cap_size != '' ORDER BY etf_cap_size"
    ).fetchall()]

    conn.close()

    results = [dict(r) for r in rows]

    return jsonify(
        rows=results, total=total, page=page, per_page=per_page,
        pages=(total + per_page - 1) // per_page,
        filters={
            "sectors": sectors, "industries": industries, "countries": countries,
            "etf_strategies": etf_strategies, "etf_cap_sizes": etf_cap_sizes,
        },
    )


@app.route("/api/general-scanner/presets")
def general_scanner_presets():
    """Return preset ticker lists the user can load into the universe."""
    presets = {
        "sp500_top50": [
            "AAPL", "MSFT", "AMZN", "NVDA", "GOOGL", "META", "BRK-B", "LLY", "AVGO", "JPM",
            "TSLA", "UNH", "V", "XOM", "MA", "PG", "COST", "JNJ", "HD", "ABBV",
            "MRK", "WMT", "BAC", "KO", "PEP", "NFLX", "ORCL", "CRM", "AMD", "TMO",
            "CVX", "LIN", "ADBE", "MCD", "ACN", "CSCO", "ABT", "WFC", "DHR", "TXN",
            "NEE", "PM", "CMCSA", "AMGN", "IBM", "RTX", "INTC", "HON", "UNP", "QCOM",
        ],
        "popular_etfs": [
            "SPY", "QQQ", "VTI", "VOO", "IWM", "DIA", "ARKK", "XLF", "XLE", "XLK",
            "XLV", "XLI", "XLP", "XLU", "XLY", "XLB", "XLRE", "VNQ", "GLD", "SLV",
            "TLT", "HYG", "LQD", "BND", "EMB", "VWO", "EFA", "IEMG", "VEA", "SCHD",
            "JEPI", "JEPQ", "VIG", "DVY", "HDV", "DGRO", "VYM", "SPHD", "PFF", "QYLD",
        ],
        "dividend_kings": [
            "ABT", "ABBV", "ADM", "AFL", "APD", "AWR", "BDX", "BKH", "CAH", "CBU",
            "CL", "CINF", "CTAS", "CWT", "DOV", "ED", "EMR", "FRT", "GD", "GPC",
            "HRL", "ITW", "JNJ", "KMB", "KO", "LANC", "LEG", "LOW", "MCD", "MMM",
            "MKC", "NDSN", "NFG", "NWN", "PEP", "PG", "PH", "PPG", "SEIC", "SJW",
            "SWK", "SYY", "TGT", "TR", "UVV", "VFC", "WAB",
        ],
        "growth_stocks": [
            "NVDA", "AMD", "AVGO", "TSLA", "META", "NFLX", "AMZN", "CRM", "NOW", "SNOW",
            "PANW", "CRWD", "DDOG", "NET", "SHOP", "SQ", "MELI", "SE", "UBER", "ABNB",
            "COIN", "PLTR", "ENPH", "SEDG", "RIVN", "LCID", "SOFI", "ARM", "SMCI", "MSTR",
        ],
    }
    return jsonify(presets=presets)


# ── Default universe: ~400 tickers across all sectors + ETFs ─────────────────
_DEFAULT_STOCKS = [
    # Technology
    "AAPL", "MSFT", "NVDA", "AVGO", "AMD", "INTC", "CRM", "ORCL", "ADBE", "NOW",
    "TXN", "QCOM", "AMAT", "MU", "LRCX", "KLAC", "SNPS", "CDNS", "MRVL", "ADI",
    "PANW", "CRWD", "FTNT", "PLTR", "INTU", "PYPL", "IBM", "ACN", "CSCO", "DELL",
    # Healthcare
    "UNH", "JNJ", "LLY", "ABBV", "MRK", "PFE", "TMO", "ABT", "DHR", "BMY",
    "AMGN", "GILD", "VRTX", "REGN", "ISRG", "MDT", "SYK", "BDX", "ZTS", "CI",
    # Financial Services
    "JPM", "BAC", "WFC", "GS", "MS", "C", "BLK", "SCHW", "SPGI", "ICE",
    "CME", "AON", "PNC", "USB", "AXP", "COF", "MET", "PRU", "AFL", "PGR",
    # Consumer Cyclical
    "AMZN", "TSLA", "HD", "MCD", "NKE", "LOW", "SBUX", "TJX", "BKNG", "MAR",
    "CMG", "ORLY", "ROST", "DHI", "LEN", "F", "GM", "ABNB", "UBER", "YUM",
    # Consumer Defensive
    "PG", "KO", "PEP", "COST", "WMT", "PM", "MO", "CL", "KMB", "GIS",
    "HSY", "KHC", "MDLZ", "STZ", "TSN", "ADM", "SYY", "KR", "TGT", "CLX",
    # Energy
    "XOM", "CVX", "COP", "EOG", "SLB", "MPC", "PSX", "VLO", "OXY",
    "DVN", "FANG", "HES", "BKR", "KMI", "WMB", "OKE", "TRGP", "ET", "EPD",
    # Industrials
    "HON", "UNP", "UPS", "RTX", "CAT", "DE", "GE", "BA", "LMT", "NOC",
    "GD", "EMR", "ITW", "ETN", "PH", "CMI", "FDX", "CSX", "WM", "RSG",
    # Utilities
    "NEE", "DUK", "SO", "D", "AEP", "SRE", "EXC", "XEL", "ED", "WEC",
    # Real Estate
    "PLD", "AMT", "CCI", "EQIX", "PSA", "SPG", "O", "WELL", "DLR", "AVB",
    "STAG", "NNN", "OHI", "MPW",
    # Communication Services
    "GOOGL", "META", "NFLX", "DIS", "CMCSA", "T", "VZ", "TMUS", "CHTR", "EA",
    # Basic Materials
    "LIN", "APD", "SHW", "ECL", "DD", "NEM", "FCX", "NUE", "STLD", "CF",
]

_DEFAULT_ETFS = [
    # Broad Market
    "SPY", "QQQ", "VTI", "VOO", "IWM", "DIA", "MDY", "RSP", "VTV", "VUG",
    # Mid & Small Cap
    "IJH", "VO", "IJR", "VB", "SCHA", "VBK", "VBR",
    # Sector ETFs
    "XLF", "XLE", "XLK", "XLV", "XLI", "XLP", "XLU", "XLY", "XLB", "XLRE", "XLC",
    "KRE", "XBI", "IBB", "IYR", "SOXX", "SMH",
    # Dividend / Income - Equity
    "SCHD", "VYM", "HDV", "DVY", "DGRO", "VIG", "SPHD", "SPYD", "NOBL", "SDY",
    "FDVV", "DGRW", "FDL", "DHS", "FVD", "OVL", "OUSA", "LVHD", "SDOG", "GCOW",
    # Covered Call / Options Income
    "JEPI", "JEPQ", "QYLD", "XYLD", "DIVO", "NUSI", "RYLD", "SPYI", "QQQI",
    "SVOL", "TLTW", "YMAX", "YMAG", "CONY", "TSLY", "NVDY",
    "JEPY", "AIPI", "FEPI", "QQQY", "XDTE", "QDTE",
    # Preferred Stock
    "PFF", "PFFD", "PGX", "FPE", "VRP",
    # BDC
    "MAIN", "ARCC", "HTGC", "GBDC", "BXSL", "TPVG", "GSBD", "OBDC", "FSK", "BIZD",
    "NEWT", "GAIN", "GLAD", "CSWC", "PBDC",
    # CEF Income
    "PCEF", "CEFS", "YYY", "XMPT", "FCEF", "ADX", "ASGI", "BST",
    # Loans
    "BKLN", "SRLN",
    # International Dividend
    "IDV", "VYMI", "DWX", "SCHY",
    # Bond
    "TLT", "BND", "AGG", "SHY", "IEF", "TIP", "BNDX", "LQD", "HYG", "JNK",
    "VCSH", "VCIT", "MUB", "EMB", "GOVT", "SGOV",
    # International Equity
    "VWO", "EFA", "IEMG", "VEA", "EEM", "FXI", "VXUS",
    # Commodity
    "GLD", "SLV", "GDX", "USO", "PDBC", "GLDM", "IAU",
    # Real Estate
    "VNQ", "VNQI", "SCHH", "REM",
    # Thematic
    "ARKK", "TAN", "ICLN", "LIT",
    # Leveraged (popular)
    "TQQQ", "SQQQ", "SOXL",
    # Crypto
    "BITO", "IBIT",
]


@app.route("/api/general-scanner/auto-load", methods=["POST"])
def general_scanner_auto_load():
    """Load the default universe. If force=true, clears and reloads. Otherwise only loads if empty."""
    force = request.args.get("force", "false").lower() == "true"
    conn = get_connection()

    if not force:
        count = conn.execute("SELECT COUNT(*) FROM general_scanner_universe").fetchone()[0]
        if count > 0:
            conn.close()
            return jsonify(ok=True, loaded=0, message="Universe already populated")

    if force:
        conn.execute("DELETE FROM general_scanner_universe")
        conn.execute("DELETE FROM general_scanner_cache")

    loaded = 0
    for t in _DEFAULT_STOCKS:
        conn.execute("INSERT OR IGNORE INTO general_scanner_universe (ticker, asset_type) VALUES (?, 'Stock')", (t,))
        loaded += 1
    for t in _DEFAULT_ETFS:
        conn.execute("INSERT OR IGNORE INTO general_scanner_universe (ticker, asset_type) VALUES (?, 'ETF')", (t,))
        loaded += 1
    # Also load from saved defaults JSON if it exists
    import json as _json
    defaults_path = os.path.join(os.path.dirname(__file__), "scanner_defaults.json")
    if os.path.exists(defaults_path):
        with open(defaults_path, "r") as f:
            saved = _json.load(f)
        for item in saved:
            conn.execute("INSERT OR IGNORE INTO general_scanner_universe (ticker, asset_type) VALUES (?, ?)",
                         (item["ticker"], item["asset_type"]))
            loaded += 1

    conn.commit()
    conn.close()
    return jsonify(ok=True, loaded=loaded)


@app.route("/api/general-scanner/save-defaults", methods=["POST"])
def general_scanner_save_defaults():
    """Save current universe as defaults JSON file so all users get them."""
    import json as _json
    conn = get_connection()
    rows = conn.execute("SELECT ticker, asset_type FROM general_scanner_universe ORDER BY ticker").fetchall()
    conn.close()
    data = [{"ticker": r["ticker"], "asset_type": r["asset_type"]} for r in rows]
    defaults_path = os.path.join(os.path.dirname(__file__), "scanner_defaults.json")
    with open(defaults_path, "w") as f:
        _json.dump(data, f, indent=2)
    return jsonify(ok=True, count=len(data))


# ── Run ────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    is_packaged = getattr(sys, "frozen", False) or os.environ.get("ELECTRON_RUN_AS_NODE")
    app.run(debug=not is_packaged, port=5001, use_reloader=False)
