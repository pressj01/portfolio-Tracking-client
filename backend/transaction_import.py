"""
Parsers for importing transaction history from external sources.

Each parser accepts a file path + filename and returns a normalised result dict:
    {
        "transactions": [ ... ],       # list of normalised dicts
        "summary": {
            "buys": int,
            "sells": int,
            "dividends": int,
            "filtered": int,
            "drip_detected": int,
            "splits_applied": int,
        },
    }

Normalised transaction dict keys:
    type            "BUY" | "SELL" | "DIVIDEND"
    ticker          str
    date            str  (YYYY-MM-DD)
    shares          float | None  (None for DIVIDEND)
    price_per_share float | None  (None for DIVIDEND)
    fees            float
    dividend_amount float | None  (only for DIVIDEND)
    notes           str
"""

import csv
import re
from collections import defaultdict
from datetime import datetime

from snowball_assign import parse_snowball_category_label

TICKER_RE = re.compile(r"^[A-Z][A-Z0-9.\-/]{0,10}$")
ADJUSTMENT_NOTE = "Automatically generated transaction to adjust"


def clean_security_description(value):
    """Turn broker/Excel multiline security names into a single clean label."""
    text = str(value or "")
    # Some Shear Excel exports expose OOXML carriage-return/newline escapes as
    # literal text, followed by a real line break (for example `_x000d_\n`).
    text = re.sub(r"_x(?:000d|000a)_", " ", text, flags=re.IGNORECASE)
    return " ".join(text.replace("\r", " ").replace("\n", " ").split())


def _is_csv_file(file_path, filename=None):
    name = str(filename or file_path or "").lower()
    return name.endswith(".csv")


def _read_table_rows(file_path, filename=None, sheet_name=None):
    """Return a list of row lists from either CSV or Excel."""
    if _is_csv_file(file_path, filename):
        with open(file_path, "r", encoding="utf-8-sig", newline="") as fh:
            return [row for row in csv.reader(fh)]

    import openpyxl

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
        return [list(row) for row in ws.iter_rows(values_only=True)]
    finally:
        wb.close()


def _row_has_values(row):
    return any(v is not None and str(v).strip() != "" for v in row)


def _row_cell(row, index, default=""):
    if index >= len(row):
        return default
    value = row[index]
    return default if value is None else value


def _rows_to_dicts(rows, header_idx=0):
    if len(rows) <= header_idx:
        return [], []
    header = [str(c or "").strip() for c in rows[header_idx]]
    data_rows = []
    for row in rows[header_idx + 1:]:
        if not _row_has_values(row):
            continue
        values = list(row) + [None] * max(0, len(header) - len(row))
        data_rows.append(dict(zip(header, values)))
    return header, data_rows


def _header_key(value):
    """Normalize broker header labels so small punctuation/casing changes do not matter."""
    return re.sub(r"[^a-z0-9]+", "", str(value or "").strip().lower())


def _canonical_header(header, aliases):
    alias_map = {}
    for canonical, alias_list in aliases.items():
        for alias in [canonical, *alias_list]:
            key = _header_key(alias)
            if key and key not in alias_map:
                alias_map[key] = canonical
    return [alias_map.get(_header_key(col), str(col or "").strip()) for col in header]


def _find_header_row(rows, aliases, required, max_scan=40):
    """Find the best header row by matching required canonical columns."""
    required = set(required)
    best_idx = None
    best_header = None
    best_score = -1

    for idx, row in enumerate(rows[:max_scan]):
        if not _row_has_values(row):
            continue
        header = _canonical_header([str(c or "").strip() for c in row], aliases)
        present = {col for col in header if col}
        if not required.issubset(present):
            continue
        score = (len(required & present) * 10) + len(present & set(aliases.keys()))
        if score > best_score:
            best_idx = idx
            best_header = header
            best_score = score

    return best_idx, best_header


def _row_record(header, row):
    values = list(row) + [None] * max(0, len(header) - len(row))
    record = {}
    for key, value in zip(header, values):
        if not key:
            continue
        existing = record.get(key)
        existing_blank = existing is None or str(existing).strip() == ""
        value_blank = value is None or str(value).strip() == ""
        if key not in record or (existing_blank and not value_blank):
            record[key] = value
    return record


def _rows_to_flexible_dicts(rows, aliases, required, max_scan=40):
    header_idx, header = _find_header_row(rows, aliases, required, max_scan=max_scan)
    if header_idx is None:
        return None, [], []
    data_rows = []
    for row in rows[header_idx + 1:]:
        if not _row_has_values(row):
            continue
        data_rows.append(_row_record(header, row))
    return header_idx, header, data_rows


def _parse_reinvest_bool(value):
    """Return True/False for common broker DRIP flag values, or None if blank."""
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    text = str(value).strip().casefold()
    if text in ("", "nan", "none", "null", "n/a", "-"):
        return None
    if text in ("y", "yes", "true", "t", "1", "on", "checked", "x", "drip", "reinvest", "reinvested"):
        return True
    if text in ("n", "no", "false", "f", "0", "off", "unchecked", "cash", "not reinvested"):
        return False
    return None


# â”€â”€ Snowball Analytics â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

def parse_snowball_csv(file_path, filename):
    """Parse a per-account Snowball Analytics CSV export.

    Rejects combined (multi-account) exports based on filename and content
    heuristics.
    """

    # â”€â”€ Check 1: filename â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    if "combined" in filename.lower():
        raise ValueError(
            "This appears to be a combined portfolio export. "
            "Please export each account individually from Snowball."
        )

    # â”€â”€ Read all rows â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    _, rows = _rows_to_dicts(_read_table_rows(file_path, filename))

    if not rows:
        raise ValueError("The file is empty or has no data rows.")

    # â”€â”€ Check 2: multiple CASH_IN on earliest date â†’ combined export â”€â”€â”€â”€â”€â”€â”€â”€â”€
    earliest_date = None
    cash_in_dates = []
    for row in rows:
        event = (row.get("Event") or "").strip()
        raw_date = (row.get("Date") or "").strip()
        if not raw_date:
            continue
        d = _parse_date(raw_date)
        if d and (earliest_date is None or d < earliest_date):
            earliest_date = d
        if event == "CASH_IN" and d:
            cash_in_dates.append(d)

    if earliest_date:
        earliest_cash_ins = sum(1 for d in cash_in_dates if d == earliest_date)
        if earliest_cash_ins > 1:
            raise ValueError(
                "This appears to be a combined portfolio export "
                f"({earliest_cash_ins} accounts detected). "
                "Please export each account individually from Snowball."
            )

    # â”€â”€ Parse and filter rows â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    kept = []
    split_events = []
    filtered_count = 0
    importable_events = {"BUY", "SELL", "DIVIDEND"}

    for idx, row in enumerate(rows):
        event = (row.get("Event") or "").strip().upper()
        note = row.get("Note") or ""

        if event == "SPLIT":
            ticker = (row.get("Symbol") or "").strip().upper()
            raw_date = (row.get("Date") or "").strip()
            date_str = _parse_date_str(raw_date)
            ratio = _safe_float(row.get("Price"))
            if ticker and TICKER_RE.match(ticker) and date_str and ratio and ratio > 0:
                split_events.append({
                    "ticker": ticker,
                    "date": date_str,
                    "row_idx": idx,
                    "ratio": ratio,
                })
            else:
                filtered_count += 1
            continue

        # Skip non-importable event types
        if event not in importable_events:
            filtered_count += 1
            continue

        ticker = (row.get("Symbol") or "").strip().upper()
        if not ticker or not TICKER_RE.match(ticker):
            filtered_count += 1
            continue

        raw_date = (row.get("Date") or "").strip()
        date_str = _parse_date_str(raw_date)
        if not date_str:
            filtered_count += 1
            continue

        qty = _safe_float(row.get("Quantity"))
        price = _safe_float(row.get("Price"))
        fees = _safe_float(row.get("FeeTax") or row.get("Fee") or row.get("Fees") or row.get("Commission") or row.get("Commissions"))
        is_adjustment = ADJUSTMENT_NOTE in note

        if event == "DIVIDEND":
            # Skip dividend adjustments (cash balance corrections, not real payments)
            if is_adjustment:
                filtered_count += 1
                continue
            kept.append({
                "type": "DIVIDEND",
                "ticker": ticker,
                "date": date_str,
                "shares": None,
                "price_per_share": None,
                "fees": 0.0,
                "dividend_amount": qty,  # Quantity is dollar amount for dividends
                "notes": note.strip(),
            })
        else:
            if qty is None or qty == 0:
                filtered_count += 1
                continue

            # Handle Snowball balance adjustments â€” these are opening position
            # corrections that make the transaction math add up.
            # Positive qty â†’ BUY, negative qty â†’ SELL (convert accordingly).
            if is_adjustment:
                if qty > 0:
                    txn_type = "BUY"
                else:
                    txn_type = "SELL"
                adj_note = f"[Opening balance] {note.strip()}"
            else:
                # Regular transaction â€” negative BUY qty shouldn't happen
                # outside adjustments, but handle gracefully
                if event == "BUY" and qty < 0:
                    filtered_count += 1
                    continue
                txn_type = event
                adj_note = note.strip()

            kept.append({
                "type": txn_type,
                "ticker": ticker,
                "date": date_str,
                "shares": abs(qty),
                "price_per_share": price,
                "fees": fees or 0.0,
                "dividend_amount": None,
                "notes": adj_note,
                "_row_idx": idx,
            })

    splits_applied = 0
    for split in sorted(split_events, key=lambda s: (s["date"], s["row_idx"])):
        adjusted_any = False
        for txn in kept:
            if txn["type"] == "DIVIDEND" or txn["ticker"] != split["ticker"]:
                continue
            if (txn["date"], txn.get("_row_idx", -1)) >= (split["date"], split["row_idx"]):
                continue
            txn["shares"] = round((txn["shares"] or 0) * split["ratio"], 10)
            if txn["price_per_share"]:
                txn["price_per_share"] = txn["price_per_share"] / split["ratio"]
            adjusted_any = True
        if adjusted_any:
            splits_applied += 1

    for txn in kept:
        txn.pop("_row_idx", None)

    # â”€â”€ DRIP detection â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    drip_count = _detect_drip(kept)

    # â”€â”€ Build summary â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    buys = sum(1 for t in kept if t["type"] == "BUY")
    sells = sum(1 for t in kept if t["type"] == "SELL")
    divs = sum(1 for t in kept if t["type"] == "DIVIDEND")

    return {
        "transactions": kept,
        "summary": {
            "buys": buys,
            "sells": sells,
            "dividends": divs,
            "filtered": filtered_count,
            "drip_detected": drip_count,
            "splits_applied": splits_applied,
        },
    }


def _infer_snowball_asset_type(sector, category):
    sector = (sector or "").strip().lower()
    category = (category or "").strip().lower()
    if sector == "funds" or "fund" in category:
        return "Fund"
    if sector == "bonds":
        return "Bond"
    if sector:
        return sector.title()
    return "Security"


def parse_snowball_holdings_csv(file_path, filename):
    """Parse a Snowball holdings CSV/XLSX as a migration-style positions snapshot.

    Keeps only the fields this app can store and use meaningfully:
    ticker, description, shares, cost basis, current price/value, dividend
    metadata, dividends received, and category.
    """
    headers, rows = _rows_to_dicts(_read_table_rows(file_path, filename))

    required = {"Holding", "Holdings' name", "Shares", "Cost basis", "Current value", "Share price"}
    missing = sorted(required - set(headers))
    if missing:
        raise ValueError(
            "This does not look like a Snowball Holdings export. "
            f"Missing required columns: {', '.join(missing)}"
        )
    if not rows:
        raise ValueError("The file is empty or has no data rows.")

    positions = []
    filtered_count = 0

    for row in rows:
        ticker = (row.get("Holding") or "").strip().upper()
        if not ticker or not TICKER_RE.match(ticker):
            filtered_count += 1
            continue

        quantity = _safe_float(row.get("Shares"))
        purchase_value = _safe_float(row.get("Cost basis"))
        current_value = _safe_float(row.get("Current value"))
        current_price = _safe_float(row.get("Share price"))

        if quantity is None or quantity <= 0:
            filtered_count += 1
            continue

        if current_value is None and current_price is not None:
            current_value = quantity * current_price
        if purchase_value is None:
            filtered_count += 1
            continue

        current_value = current_value if current_value is not None else 0.0
        cost_per_share = (purchase_value / quantity) if quantity else 0.0
        annual_payment = _safe_float(row.get("Dividends")) or 0.0
        div_per_share = _safe_float(row.get("Dividends per share")) or 0.0
        total_divs_received = _safe_float(row.get("Div. received")) or 0.0
        dividend_yield = _safe_float(row.get("Dividend yield"))
        ex_div_date = _parse_date_str(row.get("Ex-dividend date"))
        div_pay_date = _parse_date_str(row.get("Date of the next payment"))
        category = (row.get("Category") or "").strip()
        sector = (row.get("Sector") or "").strip()

        positions.append({
            "ticker": ticker,
            "description": (row.get("Holdings' name") or "").strip(),
            "quantity": round(quantity, 6),
            "cost_per_share": round(cost_per_share, 6),
            "current_price": round(current_price or 0.0, 6),
            "purchase_value": round(purchase_value, 2),
            "current_value": round(current_value, 2),
            "gain_or_loss": round(current_value - purchase_value, 2),
            "div": round(div_per_share, 6),
            "dividend_yield": dividend_yield,
            "ex_div_date": ex_div_date,
            "div_pay_date": div_pay_date,
            "estim_payment_per_year": round(annual_payment, 2),
            "approx_monthly_income": round(annual_payment / 12.0, 2) if annual_payment > 0 else 0.0,
            "total_divs_received": round(total_divs_received, 2),
            "category": category,
            "classification_type": _infer_snowball_asset_type(sector, category),
            "asset_type": _infer_snowball_asset_type(sector, category),
        })

    if not positions:
        raise ValueError("No valid holdings rows were found in the Snowball holdings file.")

    return {
        "positions": positions,
        "summary": {
            "holdings": len(positions),
            "filtered": filtered_count,
            "options": 0,
        },
        "format_type": "positions",
        "source_format": "snowball_holdings",
    }


_parse_snowball_category_label = parse_snowball_category_label


def parse_snowball_categories_csv(file_path, filename):
    """Parse the category tree from a Snowball Holdings export.

    Snowball exports labels such as ``GROWTH / Growth-Stocks``. The leading
    value is the top-level category; the value after the slash is a
    subcategory. Labels without a slash, such as ``CASH``, are top-level
    categories with no children. Ticker/category assignments are retained,
    but position values, prices, dividends, and transactions are not parsed.
    """
    headers, rows = _rows_to_dicts(_read_table_rows(file_path, filename))

    required = {"Category"}
    missing = sorted(required - set(headers))
    if missing:
        raise ValueError(
            "This does not look like a Snowball category export. "
            f"Missing required columns: {', '.join(missing)}"
        )
    if not rows:
        raise ValueError("The file is empty or has no data rows.")

    categories = []
    category_index = {}
    subcategory_keys = set()
    assignments = []
    filtered_count = 0
    duplicates_skipped = 0
    for row in rows:
        category_name, subcategory_name = _parse_snowball_category_label(row.get("Category"))
        if not category_name:
            filtered_count += 1
            continue
        category_key = category_name.casefold()
        if category_key not in category_index:
            category_index[category_key] = len(categories)
            categories.append({"name": category_name, "subcategories": []})
        ticker = (row.get("Holding") or "").strip().upper()
        if ticker and TICKER_RE.match(ticker):
            assignments.append({
                "ticker": ticker,
                "category": category_name,
                "subcategory": subcategory_name,
            })
        if not subcategory_name:
            continue
        subcategory_key = (category_key, subcategory_name.casefold())
        if subcategory_key in subcategory_keys:
            duplicates_skipped += 1
            continue
        subcategory_keys.add(subcategory_key)
        categories[category_index[category_key]]["subcategories"].append(subcategory_name)

    if not categories:
        raise ValueError("No valid categories were found in the Snowball file.")

    subcategory_count = sum(len(item["subcategories"]) for item in categories)
    return {
        "categories": categories,
        "assignments": assignments,
        "summary": {
            "categories": len(categories),
            "subcategories": subcategory_count,
            "assignments": len(assignments),
            "filtered": filtered_count,
            "duplicates_skipped": duplicates_skipped,
        },
        "format_type": "categories",
        "source_format": "snowball_categories",
    }


# â”€â”€ DRIP detection â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

def _detect_drip(transactions):
    """Flag BUY transactions that look like DRIP reinvestments.

    Heuristic: a BUY on the same date and ticker as a DIVIDEND, where the
    BUY total cost is within 20% of the dividend amount.
    """
    # Build lookup: (ticker, date) â†’ list of DIVIDEND amounts
    div_lookup = defaultdict(list)
    for t in transactions:
        if t["type"] == "DIVIDEND":
            div_lookup[(t["ticker"], t["date"])].append(t["dividend_amount"] or 0)

    drip_count = 0
    for t in transactions:
        if t["type"] != "BUY":
            continue
        key = (t["ticker"], t["date"])
        if key not in div_lookup:
            continue
        buy_cost = (t["shares"] or 0) * (t["price_per_share"] or 0)
        for div_amt in div_lookup[key]:
            if div_amt > 0 and buy_cost > 0:
                ratio = buy_cost / div_amt
                if 0.8 <= ratio <= 1.2:
                    if "[DRIP]" not in (t["notes"] or ""):
                        t["notes"] = f"[DRIP] {t['notes']}".strip()
                        drip_count += 1
                    break

    return drip_count


# â”€â”€ Charles Schwab (Positions file) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

# Rows to skip in the positions CSV (not actual holdings)
_SCHWAB_SKIP_SYMBOLS = {
    "Futures Cash", "Futures Positions Market Value",
    "Cash & Cash Investments", "Positions Total",
}

_SCHWAB_POSITION_ALIASES = {
    "Symbol": ["Ticker"],
    "Description": ["Security Description", "Security Name", "Name"],
    "Asset Type": ["Type", "Security Type", "Holding Type"],
    "Qty (Quantity)": ["Qty", "Quantity", "Shares", "Qty #", "Quantity #"],
    "Cost Basis": ["Cost Basis Total", "Total Cost", "Purchase Value"],
    "Cost/Share": ["Cost Share", "Cost Basis/Share", "Average Cost", "Average Cost Basis", "Price Paid $"],
    "Price": ["Last Price", "Current Price", "Market Price", "Price $"],
    "Mkt Val (Market Value)": ["Mkt Val", "Market Value", "Current Value", "Value", "Value $"],
    "Gain $ (Gain/Loss $)": ["Gain $", "Gain/Loss $", "Total Gain $", "Unrealized Gain/Loss"],
    "Div Yld (Dividend Yield)": ["Div Yld", "Dividend Yield", "Dividend Yield %", "Yield"],
    "Reinvest?": ["Reinvest", "Reinvest Dividends", "DRIP"],
}


def _schwab_positions_from_records(records):
    """Turn Schwab position rows into holdings, cash, and skip counts.

    Shared by the single-account positions file and by each account block of
    the All-Accounts export, which repeats the very same columns.
    """
    positions = []
    filtered_count = 0
    options_count = 0
    options_value = 0.0
    cash_value = 0.0
    cash_seen = False
    reported_total = None

    for row in records:
        sym = str(row.get("Symbol") or "").strip()
        desc = str(row.get("Description") or "").strip()
        asset_type = str(row.get("Asset Type") or "").strip()

        if sym == "Positions Total":
            # Schwab's own account total. It counts open options, which this
            # importer does not carry as holdings, so keep it for reconciling.
            reported_total = _safe_float(row.get("Mkt Val (Market Value)"))
            filtered_count += 1
            continue

        if sym == "Cash & Cash Investments":
            cash_value = _safe_float(row.get("Mkt Val (Market Value)")) or 0.0
            cash_seen = True
            filtered_count += 1
            continue

        # Skip summary rows
        if sym in _SCHWAB_SKIP_SYMBOLS or not sym:
            filtered_count += 1
            continue

        # Skip options
        if asset_type == "Option" or " " in sym:
            options_count += 1
            options_value += _safe_float(row.get("Mkt Val (Market Value)")) or 0.0
            continue

        ticker = sym.upper()
        if not TICKER_RE.match(ticker):
            filtered_count += 1
            continue

        qty = _safe_float(row.get("Qty (Quantity)"))
        cost_basis = _safe_float(row.get("Cost Basis"))
        cost_per_share = _safe_float(row.get("Cost/Share"))
        price = _safe_float(row.get("Price"))
        mkt_val = _safe_float(row.get("Mkt Val (Market Value)"))
        gain = _safe_float(row.get("Gain $ (Gain/Loss $)"))
        div_yield = str(row.get("Div Yld (Dividend Yield)") or "").replace("%", "").strip()
        div_yield = _safe_float(div_yield)
        reinvest = _parse_reinvest_bool(row.get("Reinvest?"))

        if qty is None or qty <= 0:
            filtered_count += 1
            continue

        purchase_value = cost_basis if cost_basis is not None else qty * (cost_per_share or 0)
        if qty > 0 and purchase_value > 0:
            cost_per_share = purchase_value / qty
        current_value = mkt_val or (qty * (price or 0))

        positions.append({
            "ticker": ticker,
            "description": desc,
            "quantity": qty,
            "cost_per_share": cost_per_share or 0,
            "current_price": price or 0,
            "purchase_value": round(purchase_value, 2),
            "current_value": round(current_value, 2),
            "gain_or_loss": gain or round(current_value - purchase_value, 2),
            "dividend_yield": div_yield,
            "reinvest_dividends": reinvest,
            "asset_type": asset_type,
        })

    # Schwab can emit the same ticker more than once when lots are split across
    # sub-positions. Merge them so the import updates one consolidated holding.
    merged = {}
    for pos in positions:
        ticker = pos["ticker"]
        if ticker not in merged:
            merged[ticker] = dict(pos)
            continue

        existing = merged[ticker]
        total_qty = existing["quantity"] + pos["quantity"]
        total_purchase = existing["purchase_value"] + pos["purchase_value"]
        total_current = existing["current_value"] + pos["current_value"]
        total_gain = existing["gain_or_loss"] + pos["gain_or_loss"]

        existing["quantity"] = total_qty
        existing["purchase_value"] = round(total_purchase, 2)
        existing["current_value"] = round(total_current, 2)
        existing["gain_or_loss"] = round(total_gain, 2)
        existing["cost_per_share"] = round(total_purchase / total_qty, 4) if total_qty else 0
        existing["current_price"] = pos["current_price"] or existing["current_price"]
        existing["dividend_yield"] = pos["dividend_yield"] or existing["dividend_yield"]
        existing["reinvest_dividends"] = existing["reinvest_dividends"] or pos["reinvest_dividends"]
        if not existing["description"]:
            existing["description"] = pos["description"]
        if not existing["asset_type"]:
            existing["asset_type"] = pos["asset_type"]

    return {
        "positions": list(merged.values()),
        "filtered": filtered_count,
        "options": options_count,
        "options_value": round(options_value, 2),
        "cash": round(cash_value, 2),
        "cash_seen": cash_seen,
        "reported_total": round(reported_total, 2) if reported_total is not None else None,
    }


def parse_schwab_csv(file_path, filename):
    """Parse a Schwab Positions CSV/XLSX export.

    The positions file is the source of truth for current holdings -
    it includes shares from inter-account transfers, reverse splits,
    and all corporate actions.

    Returns a positions-based result dict (not transactions):
        {
            "positions": [ ... ],
            "summary": { "holdings": int, "filtered": int, "options": int },
            "format_type": "positions",
        }
    """

    # First line is a header like "Positions for account ..." - skip it
    # Second line is blank, then the CSV header follows
    rows = _read_table_rows(file_path, filename)
    if _schwab_looks_like_all_accounts(rows):
        raise ValueError(
            "This is a Schwab All-Accounts positions export (every account in one file). "
            "Import it with 'Charles Schwab (All Accounts Positions)' so each account "
            "can be routed to its own portfolio."
        )
    _, _, reader = _rows_to_flexible_dicts(
        rows,
        _SCHWAB_POSITION_ALIASES,
        required={"Symbol", "Qty (Quantity)"},
    )

    if not reader:
        raise ValueError(
            "Could not find the positions header row. "
            "Make sure this is a Schwab Positions export or a table with Symbol and Quantity columns."
        )

    parsed = _schwab_positions_from_records(reader)
    positions = parsed["positions"]

    if positions and all(p["purchase_value"] == 0 for p in positions):
        raise ValueError(
            "No cost basis data found - every position has a $0 cost. "
            "This usually means a Transactions file was selected with the Positions format. "
            "Please use 'Charles Schwab (Transactions)' for transaction history files."
        )

    summary = {
        "holdings": len(positions),
        "filtered": parsed["filtered"],
        "options": parsed["options"],
    }
    if parsed["cash_seen"]:
        summary["cash"] = parsed["cash"]
        summary["account_value"] = round(
            sum(position["current_value"] for position in positions) + parsed["cash"],
            2,
        )

    return {
        "positions": positions,
        "summary": summary,
        "format_type": "positions",
        "source_format": "schwab",
    }


# The All-Accounts export stacks the accounts one after another. Each block
# opens with a bare label row ("Roth_IRA ...995") on an otherwise empty line,
# then repeats the same column header, so each block parses like its own
# single-account file.
_SCHWAB_FILE_TITLE_RE = re.compile(r"^positions\s+for\b", re.I)
_SCHWAB_ACCOUNT_LABEL_RE = re.compile(
    r"^(?P<name>.*?)\s*(?:\.{2,}|\u2026|[Xx*]{3,}-?)\s*(?P<number>[A-Za-z0-9]{3,})$"
)
_SCHWAB_AS_OF_RE = re.compile(r"(\d{1,2})/(\d{1,2})/(\d{4})")


def _schwab_account_section_label(row):
    """Return the account label when this row opens an account block.

    Account labels sit on a short line of their own, unlike the column header
    and the position rows. A name holding a comma ("Doe Family Trust, Jane Doe
    ...426") is split into cells by the CSV reader, so a short row counts too
    as long as it still reads as a masked account.
    """
    if not row:
        return ""
    cells = [str(cell or "").strip() for cell in row]
    while cells and not cells[-1]:
        cells.pop()
    if not cells or not cells[0]:
        return ""
    if len(cells) > 3:
        return ""

    text = ", ".join(cell for cell in cells if cell)
    # The file's own title line ("Positions for All-Accounts as of ...") sits
    # on a short line too, but it names the export, not an account.
    if _SCHWAB_FILE_TITLE_RE.match(text):
        return ""
    if len(cells) > 1 and not _SCHWAB_ACCOUNT_LABEL_RE.match(text):
        return ""
    return text


def _split_schwab_account_label(label):
    """Split "Roth_IRA ...995" into its name and masked account number."""
    text = str(label or "").strip()
    match = _SCHWAB_ACCOUNT_LABEL_RE.match(text)
    if not match:
        return text, ""
    return match.group("name").strip(), match.group("number").strip()


def _split_schwab_account_sections(rows):
    """Group the export's rows under the account label that opens each block."""
    sections = []
    current = None
    for row in rows:
        label = _schwab_account_section_label(row)
        if label:
            current = {"label": label, "rows": []}
            sections.append(current)
            continue
        if current is not None:
            current["rows"].append(row)
    return sections


def _schwab_all_accounts_as_of(rows):
    """Read the as-of date out of the export's title line, when it carries one."""
    for row in rows[:5]:
        if not row:
            continue
        first = str(row[0] or "").strip()
        if not _SCHWAB_FILE_TITLE_RE.match(first):
            continue
        match = _SCHWAB_AS_OF_RE.search(first)
        if match:
            month, day, year = match.groups()
            return f"{year}-{int(month):02d}-{int(day):02d}"
    return ""


def _schwab_account_blocks(rows):
    """Return the account blocks that actually carry a positions table.

    A label row on its own proves nothing - export footers and note lines sit
    on short lines too. Only a block with the position columns under it counts.
    """
    blocks = []
    for section in _split_schwab_account_sections(rows):
        _, _, records = _rows_to_flexible_dicts(
            section["rows"],
            _SCHWAB_POSITION_ALIASES,
            required={"Symbol", "Qty (Quantity)"},
        )
        if records:
            blocks.append((section["label"], records))
    return blocks


def _schwab_position_table_count(rows):
    """Count distinct Schwab position headers, even if an account label is odd."""
    count = 0
    required = {"Symbol", "Qty (Quantity)"}
    for row in rows:
        if not _row_has_values(row):
            continue
        header = _canonical_header(row, _SCHWAB_POSITION_ALIASES)
        if required.issubset({column for column in header if column}):
            count += 1
    return count


def _schwab_looks_like_all_accounts(rows):
    """True only when the export actually contains multiple account tables.

    Schwab names an export ``All-Accounts-Positions-...`` and gives it an
    All-Accounts title even when the customer has only one account. That file
    is safe for the single-account importer; rejecting by title alone prevents
    a valid import.
    """
    return (
        len(_schwab_account_blocks(rows)) > 1
        or _schwab_position_table_count(rows) > 1
    )


def parse_schwab_all_accounts_csv(file_path, filename):
    """Parse a Schwab All-Accounts Positions export - every account at once.

    Schwab's All-Accounts download stacks each account in its own block: a bare
    label row ("Roth_IRA ...995"), the usual column header, that account's
    holdings, its cash row, and a Positions Total row. Each block is parsed like
    a single-account positions file and returned on its own, so one upload can
    refresh every portfolio.

    Returns:
        {
            "accounts": [ {"account_label", "account_name", "account_number",
                           "positions": [...], "summary": {...}}, ... ],
            "summary": { ... file-wide totals ... },
            "format_type": "positions_multi",
        }
    """
    rows = _read_table_rows(file_path, filename)
    if not rows:
        raise ValueError("The file is empty or has no data rows.")

    accounts = []
    for label, records in _schwab_account_blocks(rows):
        parsed = _schwab_positions_from_records(records)
        positions = parsed["positions"]
        account_name, account_number = _split_schwab_account_label(label)
        accounts.append({
            "account_label": label,
            "account_name": account_name,
            "account_number": account_number,
            "positions": positions,
            "summary": {
                "holdings": len(positions),
                "filtered": parsed["filtered"],
                "options": parsed["options"],
                "options_value": parsed["options_value"],
                "cash": parsed["cash"],
                "account_value": round(
                    sum(p["current_value"] for p in positions) + parsed["cash"], 2
                ),
                "reported_total": parsed["reported_total"],
            },
        })

    if not accounts:
        raise ValueError(
            "No account sections were found. Schwab's All-Accounts export opens each account "
            "with its own label row (for example 'Roth_IRA ...995'). If this is a single-account "
            "download, import it with 'Charles Schwab (Positions)'."
        )

    all_positions = [pos for account in accounts for pos in account["positions"]]
    total_cash = sum(account["summary"]["cash"] for account in accounts)
    if not all_positions and not total_cash:
        raise ValueError("No holdings or cash rows were found in the Schwab All-Accounts file.")
    if all_positions and all(pos["purchase_value"] == 0 for pos in all_positions):
        raise ValueError(
            "No cost basis data found - every position has a $0 cost. "
            "This usually means a Transactions file was selected with the "
            "All Accounts Positions format."
        )

    return {
        "accounts": accounts,
        "summary": {
            "accounts": len(accounts),
            "holdings": len(all_positions),
            "filtered": sum(a["summary"]["filtered"] for a in accounts),
            "options": sum(a["summary"]["options"] for a in accounts),
            "options_value": round(sum(a["summary"]["options_value"] for a in accounts), 2),
            "cash": round(sum(a["summary"]["cash"] for a in accounts), 2),
            "account_value": round(sum(a["summary"]["account_value"] for a in accounts), 2),
        },
        "as_of": _schwab_all_accounts_as_of(rows),
        "format_type": "positions_multi",
        "source_format": "schwab_all_accounts",
    }


_ETRADE_POSITION_ALIASES = {
    "Symbol": ["Ticker"],
    "Description": ["Security Description", "Security Name", "Name"],
    "Qty #": ["Qty", "Quantity", "Shares", "Quantity #", "Qty (Quantity)"],
    "Price Paid $": ["Price Paid", "Cost/Share", "Average Cost", "Average Cost Basis", "Cost Basis/Share"],
    "Last Price $": ["Last Price", "Current Price", "Market Price", "Price", "Price $"],
    "Value $": ["Value", "Market Value", "Current Value", "Mkt Val", "Mkt Val (Market Value)"],
    "Total Cost": ["Cost Basis", "Cost Basis Total", "Total Cost $", "Purchase Value"],
    "Total Gain $": ["Total Gain", "Gain $", "Gain/Loss $", "Total Gain/Loss $"],
    "Dividend Yield %": ["Dividend Yield", "Div Yld", "Div Yld (Dividend Yield)", "Yield"],
}


def parse_etrade_csv(file_path, filename):
    """Parse an E*TRADE portfolio download CSV/XLSX as current positions.

    Returns a positions-based result dict similar to the Schwab parser.
    """
    rows = _read_table_rows(file_path, filename)

    if not rows:
        raise ValueError("The file is empty.")

    account_name = ""
    account_value = 0.0
    cash_value = 0.0
    header_idx = None
    header = None

    for idx, row in enumerate(rows):
        first = str(_row_cell(row, 0)).strip()
        second = str(_row_cell(row, 1)).strip()

        if first == "Account" and second == "Net Account Value":
            data_idx = idx + 1
            while data_idx < len(rows) and not _row_has_values(rows[data_idx]):
                data_idx += 1
            if data_idx < len(rows):
                account_row = rows[data_idx]
                account_name = str(_row_cell(account_row, 0)).strip()
                account_value = _safe_float(account_row[1] if len(account_row) > 1 else None) or 0.0
                cash_value = _safe_float(account_row[7] if len(account_row) > 7 else None) or 0.0

    header_idx, header = _find_header_row(
        rows,
        _ETRADE_POSITION_ALIASES,
        required={"Symbol", "Qty #"},
    )

    if header_idx is None:
        raise ValueError(
            "Could not find the E*TRADE holdings table. "
            "Make sure this is an E*TRADE portfolio download export or a table with Symbol and Quantity columns."
        )

    positions = []
    filtered_count = 0

    for row in rows[header_idx + 1:]:
        if not _row_has_values(row):
            continue

        first = str(_row_cell(row, 0)).strip()
        if first.startswith("Generated at"):
            break

        record = _row_record(header, row)

        ticker = str(record.get("Symbol") or "").strip().upper()
        if not ticker or ticker in {"CASH", "TOTAL"}:
            filtered_count += 1
            if ticker == "CASH":
                cash_value = _safe_float(record.get("Value $")) or cash_value
            if ticker == "TOTAL":
                account_value = _safe_float(record.get("Value $")) or account_value
            continue

        if not TICKER_RE.match(ticker):
            filtered_count += 1
            continue

        qty = _safe_float(record.get("Qty #"))
        cost_per_share = _safe_float(record.get("Price Paid $"))
        current_price = _safe_float(record.get("Last Price $"))
        current_value = _safe_float(record.get("Value $"))
        purchase_value = _safe_float(record.get("Total Cost"))
        gain_or_loss = _safe_float(record.get("Total Gain $"))

        if qty is None or qty <= 0:
            filtered_count += 1
            continue

        purchase_value = purchase_value if purchase_value is not None else qty * (cost_per_share or 0)
        current_value = current_value if current_value is not None else qty * (current_price or 0)
        if gain_or_loss is None:
            gain_or_loss = round(current_value - purchase_value, 2)

        positions.append({
            "ticker": ticker,
            "description": str(record.get("Description") or "").strip(),
            "quantity": qty,
            "cost_per_share": cost_per_share or 0.0,
            "current_price": current_price or 0.0,
            "purchase_value": round(purchase_value, 2),
            "current_value": round(current_value, 2),
            "gain_or_loss": round(gain_or_loss, 2),
            "dividend_yield": _safe_float(record.get("Dividend Yield %")),
            "reinvest_dividends": None,
            "asset_type": "Security",
        })

    if not positions:
        raise ValueError("No holdings rows were found in the E*TRADE file.")

    if all(p["cost_per_share"] == 0 for p in positions):
        raise ValueError(
            "No cost basis data found â€” every position has a $0 cost. "
            "This usually means a Transactions file was selected with the Positions format. "
            "Please use the correct Transactions format for transaction history files."
        )

    return {
        "positions": positions,
        "summary": {
            "holdings": len(positions),
            "filtered": filtered_count,
            "cash": round(cash_value, 2),
            "account_value": round(account_value, 2),
        },
        "format_type": "positions",
        "account_name": account_name,
    }


def _fidelity_read_xlsx(file_path, sheet_name=None):
    import openpyxl

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
        return [list(row) for row in ws.iter_rows(values_only=True)]
    finally:
        wb.close()


def _fidelity_read_csv(file_path):
    with open(file_path, "r", encoding="utf-8-sig", newline="") as fh:
        return [row for row in csv.reader(fh)]


def _fidelity_read_rows(file_path, filename, sheet_name=None):
    if str(filename or file_path).lower().endswith(".csv"):
        return _fidelity_read_csv(file_path)
    return _fidelity_read_xlsx(file_path, sheet_name)


def _fidelity_row_record(header, row):
    return _row_record(header, row)


def _fidelity_clean_position_rows(rows, header_idx, header):
    """Keep only actual Fidelity position/cash rows after the header.

    Fidelity CSV exports append legal disclaimers and a download timestamp in
    the first column.  Because that column is normally the account number,
    those footer rows must be removed before account validation.  Clean
    exports pass through unchanged.
    """
    cleaned_rows = []
    filtered_count = 0

    for row in rows[header_idx + 1:]:
        if not _row_has_values(row):
            continue

        record = _fidelity_row_record(header, row)
        ticker = str(record.get("Symbol") or "").strip().upper()
        holding_type = str(record.get("Type") or "").strip().lower()
        quantity = _safe_float(record.get("Quantity"))
        is_cash = holding_type == "cash" or ticker.endswith("**")
        is_position = bool(
            ticker
            and TICKER_RE.match(ticker)
            and quantity is not None
            and quantity > 0
        )

        if is_cash or is_position:
            cleaned_rows.append(row)
        else:
            filtered_count += 1

    return cleaned_rows, filtered_count


def _fidelity_parse_percent_fraction(raw):
    if isinstance(raw, str) and raw.strip().endswith("%"):
        return _safe_float(raw.strip().rstrip("%"))
    val = _safe_float(raw)
    if val is None:
        return None
    return val * 100.0


_FIDELITY_POSITION_ALIASES = {
    "Account Number": ["Account #", "Acct #", "Account No", "Account Number"],
    "Account Name": ["Account", "Account Description", "Portfolio"],
    "Symbol": ["Ticker"],
    "Description": ["Security Description", "Security name", "Name"],
    "Last Price": ["Last price", "Price", "Current Price", "Market Price"],
    "Current value": ["Current Value", "Market Value", "Market value", "Value", "Value $"],
    "Cost basis total": ["Cost Basis Total", "Cost basis", "Cost Basis", "Total Cost", "Purchase Value"],
    "Average cost basis": ["Average Cost Basis", "Average cost", "Cost/Share", "Price Paid $"],
    "Total gain/loss $": ["Total Gain/Loss Dollar", "Total gain/loss dollar", "Gain/Loss $", "Total Gain $"],
    "Dist. yield": [
        "Distribution Yield", "Distribution yield", "Dist yield", "Dividend Yield", "Yield",
        "Dist. rate", "Dist rate", "Distribution Rate", "Distribution rate",
    ],
    "Amount per share": ["Amount Per Share", "Dividend Amount", "Dividend Per Share"],
    "Ex-date": ["Ex Date", "Ex-dividend Date", "Ex Dividend Date"],
    "Pay date": ["Pay Date", "Payment Date"],
    "Est. annual income": ["Estimated Annual Income", "Est Annual Income", "Annual Income"],
    "Type": ["Asset Type", "Security Type", "Holding Type"],
    "Quantity": ["Qty", "Shares", "Qty #", "Quantity #"],
}


_FIDELITY_TRANSACTION_ALIASES = {
    "Run Date": ["Date", "Settlement Date", "Trade Date", "Activity Date"],
    "Action": ["Transaction Type", "Type", "Activity Type"],
    "Symbol": ["Ticker"],
    "Quantity": ["Qty", "Shares", "Qty #", "Quantity #"],
    "Price ($)": ["Price", "Price $", "Price Paid $"],
    "Amount ($)": ["Amount", "Amount $", "Net Amount"],
    "Commission ($)": ["Commission", "Commissions", "Commission $"],
    "Fees ($)": ["Fees", "Fee", "Fees $"],
}


_GENERIC_TRANSACTION_ALIASES = {
    "Date": ["Transaction Date", "Trade Date", "Activity Date", "Payment Date"],
    "Type": ["Transaction Type", "Action", "Activity", "Event"],
    "Ticker": ["Symbol", "Stock", "ETF"],
    "Shares": ["Quantity", "Qty", "Units"],
    "Price Per Share": ["Price", "Share Price", "Price/Share", "Price Per Share ($)"],
    "Fees": ["Fee", "Commission", "Commissions", "Fees & Commissions"],
    "Dividend Amount": ["Dividend", "Amount", "Cash Amount", "Income Amount"],
    "Notes": ["Note", "Description", "Memo"],
}


def parse_generic_transactions(file_path, filename):
    """Parse the app's generic transaction XLSX/CSV format."""
    rows = _read_table_rows(file_path, filename, "Transactions")
    if not rows:
        raise ValueError("The generic transactions file is empty.")

    header_idx, header = _find_header_row(
        rows,
        _GENERIC_TRANSACTION_ALIASES,
        required={"Date", "Type", "Ticker"},
    )
    if header_idx is None:
        raise ValueError(
            "Could not find the generic transaction columns. "
            "Use the Generic Transactions template with Date, Type, and Ticker columns."
        )

    buy_actions = {"buy", "bought", "purchase", "purchased"}
    sell_actions = {"sell", "sold", "sale"}
    dividend_actions = {
        "dividend", "cash dividend", "distribution",
        "capital gain", "capital gain distribution", "cap gain",
    }
    drip_actions = {
        "drip", "reinvest", "reinvestment",
        "dividend reinvestment", "reinvest dividend",
    }

    kept = []
    filtered_count = 0

    for row in rows[header_idx + 1:]:
        if not _row_has_values(row):
            continue

        record = _row_record(header, row)
        action = str(record.get("Type") or "").strip()
        action_key = re.sub(r"[^a-z0-9]+", " ", action.casefold()).strip()
        ticker = str(record.get("Ticker") or "").strip().upper()
        date_str = _parse_date_str(record.get("Date"))
        shares = _safe_float(record.get("Shares"))
        price = _safe_float(record.get("Price Per Share"))
        fees = abs(_safe_float(record.get("Fees")) or 0.0)
        dividend_amount = _safe_float(record.get("Dividend Amount"))
        notes = str(record.get("Notes") or "").strip()

        if not ticker or not TICKER_RE.match(ticker) or not date_str:
            filtered_count += 1
            continue

        if action_key in buy_actions or action_key in sell_actions or action_key in drip_actions:
            if shares is None or shares == 0 or price is None or price <= 0:
                filtered_count += 1
                continue

            is_drip = action_key in drip_actions
            if is_drip:
                notes = f"[DRIP] {notes or 'Dividend Reinvestment'}"
            kept.append({
                "type": "SELL" if action_key in sell_actions else "BUY",
                "ticker": ticker,
                "date": date_str,
                "shares": abs(shares),
                "price_per_share": abs(price),
                "fees": round(fees, 2),
                "dividend_amount": None,
                "notes": notes,
            })
            continue

        if action_key in dividend_actions:
            if dividend_amount is None or dividend_amount == 0:
                filtered_count += 1
                continue
            kept.append({
                "type": "DIVIDEND",
                "ticker": ticker,
                "date": date_str,
                "shares": None,
                "price_per_share": None,
                "fees": 0.0,
                "dividend_amount": round(abs(dividend_amount), 2),
                "notes": notes or action.title(),
            })
            continue

        filtered_count += 1

    if not kept:
        raise ValueError(
            "No valid transactions were found. BUY, SELL, and DRIP rows require "
            "Date, Ticker, Shares, and Price Per Share; DIVIDEND rows require "
            "Date, Ticker, and Dividend Amount."
        )

    buys = sum(1 for txn in kept if txn["type"] == "BUY")
    sells = sum(1 for txn in kept if txn["type"] == "SELL")
    dividends = sum(1 for txn in kept if txn["type"] == "DIVIDEND")
    drip_count = sum(
        1 for txn in kept
        if txn["type"] == "BUY" and "[DRIP]" in (txn["notes"] or "")
    )

    return {
        "transactions": kept,
        "summary": {
            "buys": buys,
            "sells": sells,
            "dividends": dividends,
            "filtered": filtered_count,
            "drip_detected": drip_count,
            "splits_applied": 0,
        },
        "source_format": "generic_transactions",
    }


def _fidelity_positions_from_records(records):
    """Normalize one Fidelity account's already-cleaned position records."""
    positions = []
    cash_value = 0.0
    cash_count = 0

    for record in records:
        ticker = (str(record.get("Symbol") or "")).strip().upper()
        description = (str(record.get("Description") or "")).strip()
        holding_type = (str(record.get("Type") or "")).strip()
        current_value = _safe_float(record.get("Current value"))
        quantity = _safe_float(record.get("Quantity"))

        # SPAXX** and unlabeled cash stay cash. A purchased money-market fund
        # such as FZDXX is labeled Cash in some Fidelity exports, but it pays
        # a monthly dividend and belongs in the holdings table / calendar.
        if ticker.endswith("**") or (holding_type.lower() == "cash" and not ticker):
            cash_value += current_value or 0.0
            cash_count += 1
            continue

        current_price = _safe_float(record.get("Last Price")) or 0.0
        if current_value is None:
            current_value = quantity * current_price if quantity is not None else 0.0
        purchase_value = _safe_float(record.get("Cost basis total"))
        cost_per_share = _safe_float(record.get("Average cost basis"))
        gain_or_loss = _safe_float(record.get("Total gain/loss $"))
        purchase_value = purchase_value if purchase_value is not None else quantity * (cost_per_share or 0.0)
        gain_or_loss = gain_or_loss if gain_or_loss is not None else current_value - purchase_value
        dist_yield = _fidelity_parse_percent_fraction(record.get("Dist. yield"))
        est_annual_income = _safe_float(record.get("Est. annual income")) or 0.0

        positions.append({
            "ticker": ticker,
            "description": description,
            "quantity": quantity,
            "cost_per_share": cost_per_share or (purchase_value / quantity if quantity else 0.0),
            "current_price": current_price,
            "purchase_value": round(purchase_value, 2),
            "current_value": round(current_value, 2),
            "gain_or_loss": round(gain_or_loss, 2),
            "dividend_yield": dist_yield,
            "div": _safe_float(record.get("Amount per share")),
            "ex_div_date": _parse_date_str(record.get("Ex-date")),
            "div_pay_date": _parse_date_str(record.get("Pay date")),
            "estim_payment_per_year": round(est_annual_income, 2) if est_annual_income > 0 else None,
            "approx_monthly_income": round(est_annual_income / 12.0, 2) if est_annual_income > 0 else None,
            "reinvest_dividends": None,
            "asset_type": "Security",
        })

    positions_value = round(sum(p["current_value"] for p in positions), 2)
    return {
        "positions": positions,
        "summary": {
            "holdings": len(positions),
            "filtered": cash_count,
            "options": 0,
            "cash": round(cash_value, 2),
            "account_value": round(positions_value + cash_value, 2),
        },
    }


def _fidelity_position_account_groups(rows, header_idx, header):
    """Group cleaned Fidelity position rows by the account columns on each row."""
    position_rows, filtered_count = _fidelity_clean_position_rows(rows, header_idx, header)
    groups = {}

    for row in position_rows:
        record = _fidelity_row_record(header, row)
        account_name = str(record.get("Account Name") or "").strip()
        account_number = str(record.get("Account Number") or "").strip()
        number_key = re.sub(r"\W+", "", account_number).casefold()
        name_key = re.sub(r"[^a-z0-9]+", "", account_name.casefold())
        # Use both fields while parsing so a redacted sample that repeats the
        # same placeholder number still cannot merge two named accounts.
        group_key = (number_key, name_key)
        if group_key not in groups:
            groups[group_key] = {
                "account_name": account_name,
                "account_number": account_number,
                "records": [],
            }
        groups[group_key]["records"].append(record)

    return list(groups.values()), filtered_count


def _fidelity_account_label(account_name, account_number):
    """Build a readable label without displaying more than the final 4 ID characters."""
    name = str(account_name or "").strip() or "Fidelity Account"
    number = re.sub(r"\s+", "", str(account_number or "").strip())
    if not number:
        return name
    return f"{name} ...{number[-4:]}"


def _fidelity_all_accounts_as_of(rows):
    """Read Fidelity's download date footer when it is present."""
    date_re = re.compile(r"\b([A-Za-z]{3}-\d{1,2}-\d{4})\b")
    for row in reversed(rows):
        for value in row:
            text = str(value or "").strip()
            if not text.lower().startswith("date downloaded"):
                continue
            match = date_re.search(text)
            if not match:
                continue
            try:
                return datetime.strptime(match.group(1), "%b-%d-%Y").date().isoformat()
            except ValueError:
                continue
    return ""


def _read_fidelity_position_groups(file_path, filename):
    rows = _fidelity_read_rows(file_path, filename, "Position")
    if not rows:
        raise ValueError("The Fidelity positions file is empty.")

    header_idx, header = _find_header_row(
        rows,
        _FIDELITY_POSITION_ALIASES,
        required={"Symbol", "Quantity"},
    )
    if header_idx is None:
        raise ValueError(
            "Could not find the Fidelity positions columns. "
            "Make sure this is a Fidelity positions export or a table with Symbol and Quantity columns."
        )

    groups, filtered_count = _fidelity_position_account_groups(rows, header_idx, header)
    return rows, groups, filtered_count


def parse_fidelity_positions_xlsx(file_path, filename):
    """Parse a single-account Fidelity positions XLSX/XLS/CSV export."""
    _, groups, filtered_count = _read_fidelity_position_groups(file_path, filename)

    if len(groups) > 1:
        raise ValueError(
            "This appears to include more than one Fidelity account. "
            "Import it with 'Fidelity (All Accounts Positions)' so each account "
            "can be routed to its own portfolio."
        )
    if not groups:
        raise ValueError("No holdings rows were found in the Fidelity positions file.")

    group = groups[0]
    parsed = _fidelity_positions_from_records(group["records"])
    positions = parsed["positions"]

    if not positions:
        raise ValueError("No holdings rows were found in the Fidelity positions file.")

    if all(p["cost_per_share"] == 0 for p in positions):
        raise ValueError(
            "No cost basis data found â€” every position has a $0 cost. "
            "This usually means a Transactions file was selected with the Positions format. "
            "Please use 'Fidelity (Transactions)' for transaction history files."
        )

    summary = dict(parsed["summary"])
    summary["filtered"] += filtered_count
    return {
        "positions": positions,
        "summary": summary,
        "format_type": "positions",
        "source_format": "fidelity",
        "account_name": group["account_name"],
        "account_number": group["account_number"],
    }


def parse_fidelity_all_accounts_positions(file_path, filename):
    """Parse a combined Fidelity Positions export and keep every account separate."""
    rows, groups, filtered_count = _read_fidelity_position_groups(file_path, filename)
    accounts = []

    for group in groups:
        parsed = _fidelity_positions_from_records(group["records"])
        account_name = group["account_name"]
        account_number = group["account_number"]
        accounts.append({
            "account_label": _fidelity_account_label(account_name, account_number),
            "account_name": account_name,
            "account_number": account_number,
            "positions": parsed["positions"],
            "summary": parsed["summary"],
        })

    if not accounts:
        raise ValueError(
            "No populated Fidelity account rows were found. The All Accounts Positions "
            "file must include Account number, Account name, Symbol, and Quantity columns."
        )

    all_positions = [position for account in accounts for position in account["positions"]]
    total_cash = sum(account["summary"]["cash"] for account in accounts)
    if not all_positions and not total_cash:
        raise ValueError("No holdings or cash rows were found in the Fidelity All Accounts file.")
    if all_positions and all(position["purchase_value"] == 0 for position in all_positions):
        raise ValueError(
            "No cost basis data found - every position has a $0 cost. "
            "This usually means a Transactions file was selected with the "
            "All Accounts Positions format."
        )

    return {
        "accounts": accounts,
        "summary": {
            "accounts": len(accounts),
            "holdings": len(all_positions),
            "filtered": filtered_count + sum(
                account["summary"]["filtered"] for account in accounts
            ),
            "options": 0,
            "options_value": 0.0,
            "cash": round(total_cash, 2),
            "account_value": round(
                sum(account["summary"]["account_value"] for account in accounts), 2
            ),
        },
        "as_of": _fidelity_all_accounts_as_of(rows),
        "format_type": "positions_multi",
        "source_format": "fidelity_all_accounts",
    }


def parse_fidelity_transactions_xlsx(file_path, filename):
    """Parse a Fidelity transactions XLSX/XLS/CSV export."""
    rows = _fidelity_read_rows(file_path, filename, "Transactions")
    if not rows:
        raise ValueError("The Fidelity transactions file is empty.")

    header_idx, header = _find_header_row(
        rows,
        _FIDELITY_TRANSACTION_ALIASES,
        required={"Run Date", "Action", "Symbol"},
    )
    if header_idx is None:
        raise ValueError(
            "Could not find the Fidelity transaction header row. "
            "Make sure this is a Fidelity transactions export or a table with Date, Action, and Symbol columns."
        )

    kept = []
    filtered_count = 0

    for row in rows[header_idx + 1:]:
        if not any(v is not None and str(v).strip() != "" for v in row):
            continue
        record = _fidelity_row_record(header, row)
        action = (str(record.get("Action") or "")).strip()
        symbol = (str(record.get("Symbol") or "")).strip().upper()
        if not symbol or not TICKER_RE.match(symbol):
            filtered_count += 1
            continue

        date_str = _parse_date_str(record.get("Run Date"))
        if not date_str:
            filtered_count += 1
            continue

        qty_val = _safe_float(record.get("Quantity"))
        price_val = _safe_float(record.get("Price ($)"))
        amount_val = _safe_float(record.get("Amount ($)"))
        commission = _safe_float(record.get("Commission ($)") or record.get("Commission") or record.get("Commissions")) or 0.0
        fees = _safe_float(record.get("Fees ($)") or record.get("Fees") or record.get("Fee")) or 0.0
        total_fees = round(commission + fees, 2)
        action_upper = action.upper()

        if "DIVIDEND RECEIVED" in action_upper or action_upper in {"DIVIDEND", "DIVIDENDS", "CASH DIVIDEND"}:
            if amount_val is None:
                filtered_count += 1
                continue
            kept.append({
                "type": "DIVIDEND",
                "ticker": symbol,
                "date": date_str,
                "shares": None,
                "price_per_share": None,
                "fees": 0.0,
                "dividend_amount": round(abs(amount_val), 2),
                "notes": "Dividend Received",
            })
            continue

        if "REINVESTMENT" in action_upper or "REINVEST" in action_upper:
            if qty_val is None or qty_val == 0:
                filtered_count += 1
                continue
            kept.append({
                "type": "BUY",
                "ticker": symbol,
                "date": date_str,
                "shares": abs(qty_val),
                "price_per_share": price_val,
                "fees": total_fees,
                "dividend_amount": None,
                "notes": "[DRIP] Reinvestment",
            })
            continue

        if "YOU BOUGHT" in action_upper or action_upper in {"BUY", "BOUGHT"}:
            if qty_val is None or qty_val == 0:
                filtered_count += 1
                continue
            kept.append({
                "type": "BUY",
                "ticker": symbol,
                "date": date_str,
                "shares": abs(qty_val),
                "price_per_share": price_val,
                "fees": total_fees,
                "dividend_amount": None,
                "notes": "",
            })
            continue

        if "YOU SOLD" in action_upper or action_upper in {"SELL", "SOLD"}:
            if qty_val is None or qty_val == 0:
                filtered_count += 1
                continue
            kept.append({
                "type": "SELL",
                "ticker": symbol,
                "date": date_str,
                "shares": abs(qty_val),
                "price_per_share": price_val,
                "fees": total_fees,
                "dividend_amount": None,
                "notes": "",
            })
            continue

        filtered_count += 1

    drip_count = sum(1 for t in kept if t["type"] == "BUY" and "[DRIP]" in (t["notes"] or ""))
    buys = sum(1 for t in kept if t["type"] == "BUY")
    sells = sum(1 for t in kept if t["type"] == "SELL")
    divs = sum(1 for t in kept if t["type"] == "DIVIDEND")

    return {
        "transactions": kept,
        "summary": {
            "buys": buys,
            "sells": sells,
            "dividends": divs,
            "filtered": filtered_count,
            "drip_detected": drip_count,
            "splits_applied": 0,
        },
    }


# â”€â”€ Charles Schwab (Transactions file) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

# Action types that represent dividend/distribution income
_SCHWAB_DIVIDEND_ACTIONS = {
    "Cash Dividend", "Non-Qualified Div", "Reinvest Dividend",
    "Div Adjustment", "Long Term Cap Gain", "Long Term Cap Gain Reinvest",
    "Short Term Cap Gain Reinvest", "Ret Cap Reinvest",
    "Pr Yr Cash Div", "Pr Yr Non Qual Div", "Pr Yr Div Reinvest",
}

# Action types that represent share purchases from reinvested dividends/distributions
_SCHWAB_DRIP_ACTIONS = {"Reinvest Shares"}

# Action types that are adjustments to reinvestments (share corrections)
_SCHWAB_REINVEST_ADJ_ACTIONS = {"Reinvestment Adj"}

_SCHWAB_TRANSACTION_ALIASES = {
    "Date": ["Run Date", "Trade Date", "Settlement Date", "Activity Date"],
    "Action": ["Transaction Type", "Type", "Activity Type"],
    "Symbol": ["Ticker"],
    "Description": ["Security Description", "Name"],
    "Quantity": ["Qty", "Shares", "Qty #", "Quantity #"],
    "Price": ["Price $", "Price ($)", "Share Price"],
    "Fees & Commissions": ["Fees & Comm", "Fees", "Fee", "Commission", "Commissions"],
    "Amount": ["Amount $", "Amount ($)", "Net Amount"],
}


def _schwab_parse_amount(raw):
    """Parse a Schwab Amount field like '$1,234.56 ' or '($33.02)' â†’ float."""
    if not raw:
        return None
    s = str(raw).strip()
    negative = s.startswith("(") and s.endswith(")")
    s = s.strip("()").replace("$", "").replace(",", "").strip()
    val = _safe_float(s)
    if val is not None and negative:
        val = -val
    return val


def _schwab_parse_date_field(raw):
    """Parse Schwab date like '4/14/2026' or '04/02/2026 as of 03/25/2026'.

    For 'as of' dates, use the first date (the posting/settlement date).
    Returns YYYY-MM-DD string or None.
    """
    if not raw:
        return None
    # Take the first date when "as of" is present
    date_part = str(raw).split(" as of ")[0].strip()
    return _parse_date_str(date_part)


def parse_schwab_transactions_csv(file_path, filename):
    """Parse a Schwab Transactions CSV/XLSX export.

    Columns: Date, Action, Symbol, Description, Quantity, Price, Fees & Comm, Amount

    Returns a normalised transaction result dict with BUY, SELL, and DIVIDEND entries.
    Reinvest Shares are imported as BUY with [DRIP] tag.
    All dividend-like distributions (cash, reinvested, cap gains, return of capital)
    are imported as DIVIDEND.
    """

    rows = _read_table_rows(file_path, filename)

    if not rows:
        raise ValueError("The file is empty.")

    # Schwab transaction CSVs start directly with the header row
    # (Date,Action,Symbol,...) â€” no preamble lines like the positions file.
    # Find the header row to be safe.
    _, _, reader = _rows_to_flexible_dicts(
        rows,
        _SCHWAB_TRANSACTION_ALIASES,
        required={"Date", "Action", "Symbol"},
    )
    if not reader:
        raise ValueError(
            "Could not find the transaction header row (Date,Action,Symbol,...). "
            "Make sure this is a Schwab Transactions export or a table with Date, Action, and Symbol columns."
        )

    kept = []
    filtered_count = 0

    for row in reader:
        action = str(row.get("Action") or "").strip()
        action_key = action.lower()
        symbol = str(row.get("Symbol") or "").strip().upper()
        raw_date = row.get("Date")
        raw_qty = row.get("Quantity")
        raw_price = row.get("Price")
        raw_fees = row.get("Fees & Commissions") or row.get("Fees & Comm")
        raw_amount = row.get("Amount")

        # Skip rows without a valid ticker
        if not symbol or not TICKER_RE.match(symbol):
            filtered_count += 1
            continue

        date_str = _schwab_parse_date_field(raw_date)
        if not date_str:
            filtered_count += 1
            continue

        amount = _schwab_parse_amount(raw_amount)
        qty = _safe_float(raw_qty)
        price = _safe_float(raw_price)
        fees = _safe_float(raw_fees) or 0.0

        # â”€â”€ Dividend / distribution â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        if action in _SCHWAB_DIVIDEND_ACTIONS or action_key in {"dividend", "dividends", "cash dividend"}:
            div_amount = abs(amount) if amount is not None else 0.0
            # Div Adjustment can be negative (reversal)
            if action == "Div Adjustment" and amount is not None and amount < 0:
                div_amount = amount  # keep negative
            kept.append({
                "type": "DIVIDEND",
                "ticker": symbol,
                "date": date_str,
                "shares": None,
                "price_per_share": None,
                "fees": 0.0,
                "dividend_amount": round(div_amount, 2),
                "notes": action,
            })
            continue

        # â”€â”€ DRIP reinvestment shares â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        if action in _SCHWAB_DRIP_ACTIONS or "reinvest" in action_key and "adj" not in action_key:
            if qty is None or qty == 0:
                filtered_count += 1
                continue
            kept.append({
                "type": "BUY",
                "ticker": symbol,
                "date": date_str,
                "shares": abs(qty),
                "price_per_share": price,
                "fees": fees,
                "dividend_amount": None,
                "notes": f"[DRIP] {action}",
            })
            continue

        # â”€â”€ Reinvestment adjustment (share correction) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        if action in _SCHWAB_REINVEST_ADJ_ACTIONS or "reinvestment adj" in action_key:
            if qty is None or qty == 0:
                filtered_count += 1
                continue
            # Positive qty = shares added back, negative = shares removed
            txn_type = "BUY" if qty > 0 else "SELL"
            kept.append({
                "type": txn_type,
                "ticker": symbol,
                "date": date_str,
                "shares": abs(qty),
                "price_per_share": price,
                "fees": fees,
                "dividend_amount": None,
                "notes": f"[Adjustment] {action}",
            })
            continue

        # â”€â”€ Buy â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        if action_key in {"buy", "bought", "you bought"}:
            if qty is None or qty == 0:
                filtered_count += 1
                continue
            kept.append({
                "type": "BUY",
                "ticker": symbol,
                "date": date_str,
                "shares": abs(qty),
                "price_per_share": price,
                "fees": fees,
                "dividend_amount": None,
                "notes": "",
            })
            continue

        # â”€â”€ Sell â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        if action_key in {"sell", "sold", "you sold"}:
            if qty is None or qty == 0:
                filtered_count += 1
                continue
            kept.append({
                "type": "SELL",
                "ticker": symbol,
                "date": date_str,
                "shares": abs(qty),
                "price_per_share": price,
                "fees": fees,
                "dividend_amount": None,
                "notes": "",
            })
            continue

        # â”€â”€ Unknown action â€” skip â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        filtered_count += 1

    # â”€â”€ DRIP detection (for BUYs not already tagged) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    drip_count = _detect_drip(kept)
    # Don't double-count DRIP tags we already set
    already_drip = sum(1 for t in kept if t["type"] == "BUY" and "[DRIP]" in (t["notes"] or "") and t["notes"].startswith("[DRIP]"))
    drip_count = max(drip_count, already_drip)

    buys = sum(1 for t in kept if t["type"] == "BUY")
    sells = sum(1 for t in kept if t["type"] == "SELL")
    divs = sum(1 for t in kept if t["type"] == "DIVIDEND")

    return {
        "transactions": kept,
        "summary": {
            "buys": buys,
            "sells": sells,
            "dividends": divs,
            "filtered": filtered_count,
            "drip_detected": drip_count,
            "splits_applied": 0,
        },
    }


# â”€â”€ E*Trade (Transaction History â€” Buys & Sells) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

_ETRADE_TRANSACTION_ALIASES = {
    "Activity/Trade Date": ["Date", "Trade Date", "Activity Date", "Run Date"],
    "Activity Type": ["Action", "Transaction Type", "Type"],
    "Symbol": ["Ticker"],
    "Description": ["Security Description", "Name"],
    "Quantity #": ["Quantity", "Qty", "Shares", "Qty #"],
    "Price $": ["Price", "Price ($)", "Share Price"],
    "Amount $": ["Amount", "Amount ($)", "Net Amount"],
    "Commission": ["Commissions", "Commission $", "Fees", "Fee", "Fees & Commissions"],
}


def _etrade_read_rows(file_path, filename):
    """Read an E*Trade transaction history CSV/XLSX and return (account_info, data_rows).

    The common E*TRADE export layout is:
        Row 1: Title (e.g. "All Transactions Activity Types")
        Row 3: Account info (e.g. "Account Activity for IRA -2797 from ...")
        Row 5: Total line
        Row 7: Column headers
        Row 8+: Data rows (until empty/disclaimer text)
    The actual header row is detected by content so renamed files and CSV exports
    with the same fields work too.
    """
    rows = _read_table_rows(file_path, filename)

    if not rows:
        raise ValueError("The E*TRADE transaction file is empty.")

    account_info = ""
    for row in rows[:20]:
        first = str(_row_cell(row, 0)).strip()
        if first.lower().startswith("account activity for"):
            account_info = first
            break

    header_idx, header = _find_header_row(
        rows,
        _ETRADE_TRANSACTION_ALIASES,
        required={"Activity/Trade Date", "Activity Type", "Symbol"},
    )
    if header_idx is None:
        raise ValueError(
            "Could not find the E*TRADE transaction header row. "
            "Make sure this is an E*TRADE transaction export or a table with Date, Activity Type, and Symbol columns."
        )

    data_rows = []
    for row in rows[header_idx + 1:]:
        # Stop at empty rows or disclaimer text
        if not _row_has_values(row):
            continue
        val = str(_row_cell(row, 0)).strip()
        # Skip disclaimer/legal text at the bottom
        if val.startswith("The information") or val.startswith("E*TRADE") or val.startswith("Under the"):
            continue
        # Must have an activity type
        record = _row_record(header, row)
        activity = record.get("Activity Type")
        if not activity:
            continue
        data_rows.append(record)

    return account_info, data_rows


def _etrade_extract_account_name(account_info):
    """Extract the account name from an E*TRADE activity header line."""
    text = (account_info or "").strip()
    if not text:
        return ""

    match = re.search(r"Account Activity for\s+(.+?)\s+from\s+", text, re.IGNORECASE)
    if match:
        return match.group(1).strip()

    return text


def _etrade_parse_date_str(raw):
    """Parse E*Trade date like '03/06/26' (MM/DD/YY) or '2026-03-06'."""
    if raw is None:
        return None
    s = str(raw).strip()
    if not s:
        return None
    # Try MM/DD/YY (2-digit year)
    for fmt in ("%m/%d/%y", "%m/%d/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except (ValueError, AttributeError):
            continue
    return _parse_date_str(s)

# Robinhood (Positions PDF + Transactions CSV)

_ROBINHOOD_DIVIDEND_CODES = {"CDIV", "MDIV", "LCAP", "SCAP"}


def _parse_money(raw):
    if raw is None:
        return None
    text = str(raw).strip()
    if not text:
        return None
    negative = text.startswith("(") and text.endswith(")")
    text = text.strip("()").replace("$", "").replace(",", "").strip()
    val = _safe_float(text)
    if val is None:
        return None
    return -val if negative else val


def _robinhood_clean_description(description):
    lines = []
    for line in str(description or "").splitlines():
        line = line.strip()
        if not line or line.startswith("CUSIP:") or line == "Recurring":
            continue
        lines.append(line)
    return " ".join(lines)


def parse_robinhood_positions_pdf(file_path, filename):
    """Parse a Robinhood holdings PDF as a current positions snapshot.

    Robinhood's holdings PDF does not include cost basis. To keep this as a
    positions source of truth, cost basis is initialized to current value and
    transaction history can be imported separately for recordkeeping.
    """
    try:
        from pypdf import PdfReader
    except ImportError as exc:
        raise ValueError(
            "Robinhood PDF import requires the pypdf package. "
            "Install backend requirements, then try again."
        ) from exc

    reader = PdfReader(file_path)
    row_re = re.compile(
        r"^Estimated Yield:\s*(?P<yield>[0-9.]+)%\s+"
        r"(?P<ticker>[A-Z][A-Z0-9.\-/]{0,10})\s+"
        r"(?P<acct>\S+)\s+"
        r"(?P<qty>[0-9.,]+)\s+"
        r"\$(?P<price>[0-9.,]+)\s+"
        r"\$(?P<value>[0-9.,]+)\s+"
        r"\$(?P<annual>[0-9.,]+)\s+"
        r"(?P<pct>[0-9.]+)%$"
    )

    positions = []
    filtered_count = 0
    in_holdings = False
    desc_lines = []
    account_value = 0.0

    for page in reader.pages:
        text = page.extract_text() or ""
        if "Securities Held in Account" not in text and not in_holdings:
            continue

        for raw_line in text.splitlines():
            line = raw_line.strip()
            if not line:
                continue
            if "Securities Held in Account" in line:
                in_holdings = True
                desc_lines = []
                continue
            if line.startswith("Total Priced Portfolio"):
                account_value = _parse_money(line.replace("Total Priced Portfolio", "").strip()) or account_value
                in_holdings = False
                desc_lines = []
                break
            if not in_holdings:
                continue
            if line.startswith("Page ") or line == "Portfolio Summary":
                continue

            match = row_re.match(line)
            if not match:
                desc_lines.append(line)
                continue

            ticker = match.group("ticker").upper()
            quantity = _safe_float(match.group("qty"))
            current_price = _safe_float(match.group("price"))
            current_value = _safe_float(match.group("value"))
            dividend_yield = _safe_float(match.group("yield"))
            annual_payment = _safe_float(match.group("annual")) or 0.0
            description = " ".join(desc_lines).strip()
            desc_lines = []

            if not ticker or not TICKER_RE.match(ticker) or quantity is None or quantity <= 0:
                filtered_count += 1
                continue

            current_price = current_price or 0.0
            current_value = current_value if current_value is not None else quantity * current_price
            positions.append({
                "ticker": ticker,
                "description": description,
                "quantity": quantity,
                "cost_per_share": current_price,
                "current_price": current_price,
                "purchase_value": round(current_value, 2),
                "current_value": round(current_value, 2),
                "gain_or_loss": 0.0,
                "dividend_yield": dividend_yield,
                "estim_payment_per_year": round(annual_payment, 2) if annual_payment > 0 else None,
                "approx_monthly_income": round(annual_payment / 12.0, 2) if annual_payment > 0 else None,
                "reinvest_dividends": None,
                "asset_type": "Security",
            })

    if not positions:
        raise ValueError(
            "No holdings rows were found in the Robinhood PDF. "
            "Make sure this is the Robinhood holdings/positions PDF."
        )

    return {
        "positions": positions,
        "summary": {
            "holdings": len(positions),
            "filtered": filtered_count,
            "options": 0,
            "account_value": round(account_value or sum(p["current_value"] for p in positions), 2),
            "cost_basis_missing": True,
        },
        "format_type": "positions",
        "source_format": "robinhood_positions",
    }


def _robinhood_parse_quantity(raw):
    text = str(raw or "").strip().upper()
    if text.endswith("S"):
        text = text[:-1]
    return _safe_float(text)


def parse_robinhood_transactions_csv(file_path, filename):
    """Parse a Robinhood transaction/activity CSV/XLSX export."""
    headers, rows = _rows_to_dicts(_read_table_rows(file_path, filename))

    required = {"Activity Date", "Instrument", "Description", "Trans Code", "Quantity", "Price", "Amount"}
    missing = sorted(required - set(headers))
    if missing:
        raise ValueError(
            "This does not look like a Robinhood transaction CSV. "
            f"Missing required columns: {', '.join(missing)}"
        )
    if not rows:
        raise ValueError("The file is empty or has no data rows.")

    kept = []
    filtered_count = 0

    for row in rows:
        code = (row.get("Trans Code") or "").strip().upper()
        ticker = (row.get("Instrument") or "").strip().upper()
        description = row.get("Description") or ""

        if not ticker or not TICKER_RE.match(ticker):
            filtered_count += 1
            continue

        date_str = _parse_date_str(row.get("Activity Date"))
        if not date_str:
            filtered_count += 1
            continue

        qty = _robinhood_parse_quantity(row.get("Quantity"))
        price = _parse_money(row.get("Price"))
        amount = _parse_money(row.get("Amount"))
        note = _robinhood_clean_description(description) or code

        if code == "BUY":
            if qty is None or qty <= 0:
                filtered_count += 1
                continue
            if price is None and amount is not None:
                price = abs(amount) / qty
            kept.append({
                "type": "BUY",
                "ticker": ticker,
                "date": date_str,
                "shares": abs(qty),
                "price_per_share": price or 0.0,
                "fees": 0.0,
                "dividend_amount": None,
                "notes": note,
            })
        elif code == "SELL":
            if qty is None or qty <= 0:
                filtered_count += 1
                continue
            if price is None and amount is not None:
                price = abs(amount) / qty
            kept.append({
                "type": "SELL",
                "ticker": ticker,
                "date": date_str,
                "shares": abs(qty),
                "price_per_share": price or 0.0,
                "fees": 0.0,
                "dividend_amount": None,
                "notes": note,
            })
        elif code in _ROBINHOOD_DIVIDEND_CODES:
            if amount is None:
                filtered_count += 1
                continue
            if str(description).lstrip().upper().startswith("REVERT:"):
                amount = -abs(amount)
            kept.append({
                "type": "DIVIDEND",
                "ticker": ticker,
                "date": date_str,
                "shares": None,
                "price_per_share": None,
                "fees": 0.0,
                "dividend_amount": round(amount, 2),
                "notes": note,
            })
        elif code == "ACATI":
            if qty is None or qty <= 0:
                filtered_count += 1
                continue
            kept.append({
                "type": "BUY",
                "ticker": ticker,
                "date": date_str,
                "shares": abs(qty),
                "price_per_share": 0.0,
                "fees": 0.0,
                "dividend_amount": None,
                "notes": "[Transfer in] ACAT",
            })
        elif code == "ACATO":
            if qty is None or qty <= 0:
                filtered_count += 1
                continue
            kept.append({
                "type": "SELL",
                "ticker": ticker,
                "date": date_str,
                "shares": abs(qty),
                "price_per_share": 0.0,
                "fees": 0.0,
                "dividend_amount": None,
                "notes": "[Transfer out] ACAT",
            })
        else:
            filtered_count += 1

    buys = sum(1 for t in kept if t["type"] == "BUY")
    sells = sum(1 for t in kept if t["type"] == "SELL")
    divs = sum(1 for t in kept if t["type"] == "DIVIDEND")

    return {
        "transactions": kept,
        "summary": {
            "buys": buys,
            "sells": sells,
            "dividends": divs,
            "filtered": filtered_count,
            "drip_detected": 0,
            "splits_applied": 0,
        },
    }


# â”€â”€ Helpers â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

_SHEAR_GROUP_POSITION_ALIASES = {
    "Account Number": ["Account #", "Acct #", "Account No"],
    "Account Name": ["Account", "Account Description"],
    "Account Nick Name": ["Account Nickname", "Account Nick"],
    "Symbol/CUSIP": ["Symbol", "Ticker", "CUSIP"],
    "Description": ["Security Description", "Security Name", "Name"],
    "Quantity": ["Qty", "Shares", "Qty #", "Quantity #"],
    "Price ($)": ["Price", "Current Price", "Market Price", "Last Price"],
    "Value ($)": ["Value", "Market Value", "Current Value", "Mkt Val"],
    "Unit Cost": ["Average Cost", "Average Cost Basis", "Cost/Share", "Cost Basis/Share"],
    "Cost Basis ($)": ["Cost Basis", "Cost Basis Total", "Total Cost", "Purchase Value"],
    "Unrealized G/L ($)": ["Unrealized Gain/Loss", "Gain/Loss $", "Total Gain $", "Gain $"],
    "Security Type Description": ["Security Type", "Asset Type", "Type"],
}

_SHEAR_GROUP_ACTIVITY_ALIASES = {
    "Date": ["Activity Date", "Trade Date", "Run Date"],
    "Activity": ["Action", "Transaction Type", "Type", "Activity Type"],
    "Symbol": ["Ticker", "Symbol/CUSIP"],
    "Description": ["Security Description", "Name"],
    "Quantity": ["Qty", "Shares", "Qty #", "Quantity #"],
    "Unit Price": ["Price", "Price ($)", "Share Price"],
    "Value": ["Amount", "Amount ($)", "Net Amount", "Value ($)"],
    "Account Nickname": ["Account Nick Name", "Account Name", "Account"],
    "Account Number": ["Account #", "Acct #", "Account No"],
}


def _shear_group_account_parts(record):
    nickname = str(record.get("Account Nick Name") or record.get("Account Nickname") or "").strip()
    account_name = str(record.get("Account Name") or "").strip()
    account_number = str(record.get("Account Number") or "").strip()
    label = nickname or account_name
    if label and account_number:
        label = f"{label}, {account_number}"
    else:
        label = label or account_number
    return label, nickname or account_name, account_number


def _shear_group_account_label(record):
    label, _, _ = _shear_group_account_parts(record)
    return label


def _merge_positions_by_ticker(positions):
    merged = {}
    for pos in positions:
        ticker = pos["ticker"]
        if ticker not in merged:
            merged[ticker] = dict(pos)
            continue

        existing = merged[ticker]
        total_qty = (existing.get("quantity") or 0) + (pos.get("quantity") or 0)
        total_purchase = (existing.get("purchase_value") or 0) + (pos.get("purchase_value") or 0)
        total_current = (existing.get("current_value") or 0) + (pos.get("current_value") or 0)
        total_gain = (existing.get("gain_or_loss") or 0) + (pos.get("gain_or_loss") or 0)

        existing["quantity"] = total_qty
        existing["purchase_value"] = round(total_purchase, 2)
        existing["current_value"] = round(total_current, 2)
        existing["gain_or_loss"] = round(total_gain, 2)
        existing["cost_per_share"] = round(total_purchase / total_qty, 4) if total_qty else 0
        existing["current_price"] = pos.get("current_price") or existing.get("current_price") or 0
        if not existing.get("description"):
            existing["description"] = pos.get("description") or ""
        if not existing.get("asset_type"):
            existing["asset_type"] = pos.get("asset_type") or ""

    return list(merged.values())


def parse_shear_group_positions(file_path, filename):
    """Parse a Shear Group positions CSV/XLSX export as current holdings."""
    rows = _read_table_rows(file_path, filename)
    header_idx, header = _find_header_row(
        rows,
        _SHEAR_GROUP_POSITION_ALIASES,
        required={"Symbol/CUSIP", "Quantity"},
    )
    if header_idx is None:
        raise ValueError(
            "Could not find the Shear Group positions columns. "
            "Make sure this is a Positions CSV or Excel export with Symbol/CUSIP and Quantity columns."
        )

    positions = []
    filtered_count = 0
    cash_value = 0.0
    account_labels = set()
    cash_by_account = {}

    for row in rows[header_idx + 1:]:
        if not _row_has_values(row):
            continue
        record = _row_record(header, row)
        account_label, account_name, account_number = _shear_group_account_parts(record)
        if account_label:
            account_labels.add(account_label)

        ticker = str(record.get("Symbol/CUSIP") or "").strip().upper()
        asset_type = str(record.get("Security Type Description") or "").strip()
        current_value = _safe_float(record.get("Value ($)")) or 0.0
        quantity = _safe_float(record.get("Quantity"))

        if asset_type.lower() in {"cash", "money market"} or ticker in {"CASH", "USD"}:
            cash_value += current_value
            if account_label:
                cash_by_account[account_label] = cash_by_account.get(account_label, 0.0) + current_value
            filtered_count += 1
            continue

        if not ticker or not TICKER_RE.match(ticker) or quantity is None or quantity <= 0:
            cash_value += current_value if ticker and ticker.isdigit() else 0.0
            if account_label and ticker and ticker.isdigit():
                cash_by_account[account_label] = cash_by_account.get(account_label, 0.0) + current_value
            filtered_count += 1
            continue

        current_price = _safe_float(record.get("Price ($)")) or 0.0
        cost_per_share = _safe_float(record.get("Unit Cost"))
        purchase_value = _safe_float(record.get("Cost Basis ($)"))
        current_value = current_value or (quantity * current_price)
        purchase_value = purchase_value if purchase_value is not None else quantity * (cost_per_share or 0.0)
        if quantity and purchase_value:
            cost_per_share = purchase_value / quantity
        gain_or_loss = _safe_float(record.get("Unrealized G/L ($)"))
        if gain_or_loss is None:
            gain_or_loss = current_value - purchase_value

        positions.append({
            "ticker": ticker,
            "description": clean_security_description(record.get("Description")),
            "quantity": quantity,
            "cost_per_share": cost_per_share or 0.0,
            "current_price": current_price,
            "purchase_value": round(purchase_value or 0.0, 2),
            "current_value": round(current_value or 0.0, 2),
            "gain_or_loss": round(gain_or_loss or 0.0, 2),
            "dividend_yield": None,
            "reinvest_dividends": None,
            "asset_type": asset_type,
            "_account_label": account_label,
            "_account_name": account_name,
            "_account_number": account_number,
        })

    raw_positions = positions
    positions = _merge_positions_by_ticker(positions)
    if not positions:
        raise ValueError("No holdings rows were found in the Shear Group positions file.")
    if all(p["purchase_value"] == 0 for p in positions):
        raise ValueError(
            "No cost basis data found - every position has a $0 cost. "
            "This usually means an Activity file was selected with the Positions format. "
            "Please use 'Shear Group (Activity)' for activity history files."
        )

    result = {
        "positions": positions,
        "summary": {
            "holdings": len(positions),
            "filtered": filtered_count,
            "options": 0,
            "cash": round(cash_value, 2),
            "account_count": len(account_labels),
            "account_value": round(sum(p["current_value"] for p in positions) + cash_value, 2),
        },
        "format_type": "positions",
        "source_format": "shear_group",
        "_raw_positions": raw_positions,
        "_cash_by_account": {label: round(value, 2) for label, value in cash_by_account.items()},
    }
    return result


def _shear_group_multi_account_identity(account_name, account_number, account_label=""):
    """Return a stable, display-safe identity shared by Positions and Activity.

    Shear Positions exports carry the full account number while Activity exports
    carry only its last four digits. Routing on the last four lets a mapping
    confirmed for Positions also apply to Activity without returning the full
    account number to the browser.
    """
    name = str(account_name or "").strip()
    digits = re.sub(r"\D+", "", str(account_number or ""))
    suffix = digits[-4:] if digits else ""
    label = name or str(account_label or "").strip()
    if suffix:
        label = f"{label}, ending {suffix}" if label else f"Account ending {suffix}"
    return label, name, suffix


def parse_shear_group_all_accounts_positions(file_path, filename):
    """Group one Shear Positions export into independently routed accounts."""
    parsed = parse_shear_group_positions(file_path, filename)
    raw_positions = parsed.pop("_raw_positions", [])
    cash_by_account = parsed.pop("_cash_by_account", {})
    grouped = {}

    def ensure_account(account_name, account_number, raw_label):
        safe_label, safe_name, suffix = _shear_group_multi_account_identity(
            account_name, account_number, raw_label
        )
        key = f"num:{suffix}" if suffix else f"label:{safe_label.casefold()}"
        account = grouped.setdefault(key, {
            "account_label": safe_label,
            "account_name": safe_name,
            "account_number": suffix,
            "positions": [],
            "cash": 0.0,
        })
        return account

    for position in raw_positions:
        account = ensure_account(
            position.get("_account_name"),
            position.get("_account_number"),
            position.get("_account_label"),
        )
        clean_position = dict(position)
        clean_position.pop("_account_label", None)
        clean_position.pop("_account_name", None)
        clean_position.pop("_account_number", None)
        account["positions"].append(clean_position)

    # Include cash-only accounts as well as cash attached to holding accounts.
    for raw_label, cash in cash_by_account.items():
        raw_name, _, raw_number = str(raw_label or "").rpartition(",")
        account = ensure_account(raw_name.strip(), raw_number.strip(), raw_label)
        account["cash"] += float(cash or 0)

    accounts = []
    for account in grouped.values():
        positions = _merge_positions_by_ticker(account.pop("positions"))
        cash = round(account.pop("cash"), 2)
        account_value = round(sum(pos.get("current_value") or 0 for pos in positions) + cash, 2)
        account["positions"] = positions
        account["summary"] = {
            "holdings": len(positions),
            "filtered": 0,
            "options": 0,
            "cash": cash,
            "account_value": account_value,
        }
        accounts.append(account)

    accounts.sort(key=lambda account: (account.get("account_name") or "", account.get("account_number") or ""))
    summary = dict(parsed.get("summary") or {})
    summary.update({
        "accounts": len(accounts),
        "holdings": sum(account["summary"]["holdings"] for account in accounts),
        "cash": round(sum(account["summary"]["cash"] for account in accounts), 2),
        "account_value": round(sum(account["summary"]["account_value"] for account in accounts), 2),
    })
    return {
        "accounts": accounts,
        "summary": summary,
        "format_type": "positions_multi",
        "source_format": "shear_group_all_accounts",
    }


def _shear_group_price_from_amount(price, amount, quantity):
    price_val = _safe_float(price)
    amount_val = _safe_float(amount)
    qty_val = _safe_float(quantity)
    if (price_val is None or price_val == 0) and amount_val is not None and qty_val:
        return abs(amount_val) / abs(qty_val)
    return price_val


def parse_shear_group_activity(file_path, filename):
    """Parse a Shear Group activity CSV/XLSX export as transactions and dividends."""
    rows = _read_table_rows(file_path, filename)
    _, _, data_rows = _rows_to_flexible_dicts(
        rows,
        _SHEAR_GROUP_ACTIVITY_ALIASES,
        required={"Date", "Activity", "Symbol"},
    )
    if not data_rows:
        raise ValueError(
            "Could not find the Shear Group activity columns. "
            "Make sure this is an Activity CSV or Excel export with Date, Activity, and Symbol columns."
        )

    kept = []
    filtered_count = 0
    account_names = set()
    dividend_actions = {"cash dividend", "interest", "long term cap gain", "short term cap gain"}
    drip_actions = {"dividend reinvest", "lt cap gain reinvest", "st cap gain reinvest", "reinvest interest"}

    for row in data_rows:
        account_label, account_name, account_number = _shear_group_account_parts(row)
        if account_label:
            account_names.add(account_label)

        activity = str(row.get("Activity") or "").strip()
        activity_key = activity.lower()
        ticker = str(row.get("Symbol") or "").strip().upper()
        if not ticker or not TICKER_RE.match(ticker):
            filtered_count += 1
            continue

        date_str = _parse_date_str(row.get("Date"))
        if not date_str:
            filtered_count += 1
            continue

        quantity = _safe_float(row.get("Quantity"))
        amount = _safe_float(row.get("Value"))
        price = _shear_group_price_from_amount(row.get("Unit Price"), row.get("Value"), row.get("Quantity"))

        if activity_key == "buy":
            if quantity is None or quantity == 0:
                filtered_count += 1
                continue
            kept.append({
                "type": "BUY",
                "ticker": ticker,
                "date": date_str,
                "shares": abs(quantity),
                "price_per_share": price or 0.0,
                "fees": 0.0,
                "dividend_amount": None,
                "notes": activity,
                "_account_label": account_label,
                "_account_name": account_name,
                "_account_number": account_number,
            })
        elif activity_key == "sell":
            if quantity is None or quantity == 0:
                filtered_count += 1
                continue
            kept.append({
                "type": "SELL",
                "ticker": ticker,
                "date": date_str,
                "shares": abs(quantity),
                "price_per_share": price or 0.0,
                "fees": 0.0,
                "dividend_amount": None,
                "notes": activity,
                "_account_label": account_label,
                "_account_name": account_name,
                "_account_number": account_number,
            })
        elif activity_key in dividend_actions:
            if amount is None:
                filtered_count += 1
                continue
            kept.append({
                "type": "DIVIDEND",
                "ticker": ticker,
                "date": date_str,
                "shares": None,
                "price_per_share": None,
                "fees": 0.0,
                "dividend_amount": round(amount, 2),
                "notes": activity,
                "_account_label": account_label,
                "_account_name": account_name,
                "_account_number": account_number,
            })
        elif activity_key in drip_actions:
            if quantity is None or quantity == 0:
                filtered_count += 1
                continue
            kept.append({
                "type": "BUY",
                "ticker": ticker,
                "date": date_str,
                "shares": abs(quantity),
                "price_per_share": price or 0.0,
                "fees": 0.0,
                "dividend_amount": None,
                "notes": f"[DRIP] {activity}",
                "_account_label": account_label,
                "_account_name": account_name,
                "_account_number": account_number,
            })
        else:
            filtered_count += 1

    buys = sum(1 for t in kept if t["type"] == "BUY")
    sells = sum(1 for t in kept if t["type"] == "SELL")
    divs = sum(1 for t in kept if t["type"] == "DIVIDEND")
    drip_count = sum(1 for t in kept if t["type"] == "BUY" and "[DRIP]" in (t["notes"] or ""))

    result = {
        "transactions": kept,
        "summary": {
            "buys": buys,
            "sells": sells,
            "dividends": divs,
            "filtered": filtered_count,
            "drip_detected": drip_count,
            "splits_applied": 0,
        },
        "source_format": "shear_group_activity",
    }
    if account_names:
        result["summary"]["account_count"] = len(account_names)
    return result


def parse_shear_group_all_accounts_activity(file_path, filename):
    """Group one Shear Activity export into independently routed accounts."""
    parsed = parse_shear_group_activity(file_path, filename)
    grouped = {}
    for transaction in parsed.get("transactions") or []:
        safe_label, safe_name, suffix = _shear_group_multi_account_identity(
            transaction.get("_account_name"),
            transaction.get("_account_number"),
            transaction.get("_account_label"),
        )
        key = f"num:{suffix}" if suffix else f"label:{safe_label.casefold()}"
        account = grouped.setdefault(key, {
            "account_label": safe_label,
            "account_name": safe_name,
            "account_number": suffix,
            "transactions": [],
        })
        clean_transaction = dict(transaction)
        clean_transaction.pop("_account_label", None)
        clean_transaction.pop("_account_name", None)
        clean_transaction.pop("_account_number", None)
        account["transactions"].append(clean_transaction)

    accounts = []
    for account in grouped.values():
        transactions = account["transactions"]
        buys = sum(1 for txn in transactions if txn.get("type") == "BUY")
        sells = sum(1 for txn in transactions if txn.get("type") == "SELL")
        dividends = sum(1 for txn in transactions if txn.get("type") == "DIVIDEND")
        account["summary"] = {
            "transactions": len(transactions),
            "buys": buys,
            "sells": sells,
            "dividends": dividends,
            "drip_detected": sum(
                1 for txn in transactions
                if txn.get("type") == "BUY" and "[DRIP]" in (txn.get("notes") or "")
            ),
        }
        accounts.append(account)

    accounts.sort(key=lambda account: (account.get("account_name") or "", account.get("account_number") or ""))
    summary = dict(parsed.get("summary") or {})
    summary.update({
        "accounts": len(accounts),
        "transactions": sum(account["summary"]["transactions"] for account in accounts),
    })
    return {
        "accounts": accounts,
        "summary": summary,
        "format_type": "transactions_multi",
        "source_format": "shear_group_all_accounts_activity",
    }


def _parse_date(raw):
    """Parse a date string to a datetime.date, or None."""
    if raw is None:
        return None

    if isinstance(raw, datetime):
        return raw.date()

    raw = str(raw).strip()
    if not raw:
        return None

    snowball_match = re.match(r"^[A-Za-z]{3}\s+[A-Za-z]{3}\s+\d{1,2}\s+\d{4}", raw)
    if snowball_match:
        try:
            return datetime.strptime(snowball_match.group(0), "%a %b %d %Y").date()
        except ValueError:
            pass

    for fmt in (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d",
        "%m/%d/%Y",
        "%m/%d/%Y %H:%M:%S",
        "%m/%d/%Y %H:%M",
        "%b-%d-%Y",
    ):
        try:
            return datetime.strptime(raw, fmt).date()
        except (ValueError, AttributeError):
            continue
    return None


def _parse_date_str(raw):
    """Parse a date string and return YYYY-MM-DD, or None."""
    d = _parse_date(raw)
    return d.isoformat() if d else None


def _safe_float(val):
    """Convert a value to float, stripping quotes and whitespace. Returns None on failure."""
    if val is None:
        return None
    try:
        text = str(val).strip().strip('"')
        if not text or text in {"-", "--"}:
            return None
        negative = text.startswith("(") and text.endswith(")")
        text = text.strip("()").replace("$", "").replace(",", "").strip()
        if text.endswith("%"):
            text = text[:-1].strip()
        parsed = float(text)
        return -parsed if negative else parsed
    except (ValueError, TypeError):
        return None


# â”€â”€ Parser registry â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

def parse_etrade_transactions_xlsx(file_path, filename):
    """Parse one E*Trade All Transactions history CSV/XLSX export.

    The file is recognized from its E*TRADE transaction headers and row content,
    so renaming the workbook or CSV does not matter. Trades, dividend payments,
    and DRIP reinvestment buys are imported from the same export.
    """
    account_info, data_rows = _etrade_read_rows(file_path, filename)
    account_name = _etrade_extract_account_name(account_info)

    if not data_rows:
        raise ValueError("No transaction rows found in the E*Trade transactions file.")

    kept = []
    filtered_count = 0

    for row in data_rows:
        activity = (str(row.get("Activity Type") or "")).strip()
        symbol = (str(row.get("Symbol") or "")).strip().upper()
        raw_date = row.get("Activity/Trade Date")
        qty = row.get("Quantity #")
        price = row.get("Price $")
        amount = row.get("Amount $")
        commission = row.get("Commission") or row.get("Commissions") or row.get("Fee") or row.get("Fees")
        description = str(row.get("Description") or "").strip()

        if not symbol or not TICKER_RE.match(symbol):
            filtered_count += 1
            continue

        date_str = _etrade_parse_date_str(raw_date)
        if not date_str:
            filtered_count += 1
            continue

        qty_val = _safe_float(qty)
        price_val = _safe_float(price)
        amount_val = _safe_float(amount)
        fees = _safe_float(commission) or 0.0

        activity_key = activity.lower()
        activity_lower = activity.lower()
        description_lower = description.lower()
        looks_like_dividend = (
            "dividend" in activity_lower
            or "capital gain" in activity_lower
            or "dividend" in description_lower
            or "capital gain" in description_lower
        )
        looks_like_reinvestment = (
            "reinvestment" in activity_lower
            or "reinvestment" in description_lower
        )

        if activity_key in {"bought", "buy", "you bought"}:
            if qty_val is None or qty_val == 0:
                filtered_count += 1
                continue
            kept.append({
                "type": "BUY",
                "ticker": symbol,
                "date": date_str,
                "shares": abs(qty_val),
                "price_per_share": price_val,
                "fees": fees,
                "dividend_amount": None,
                "notes": "[DRIP] Dividend Reinvestment" if looks_like_reinvestment or looks_like_dividend else "",
            })
        elif activity_key in {"sold", "sell", "you sold"}:
            if qty_val is None or qty_val == 0:
                filtered_count += 1
                continue
            kept.append({
                "type": "SELL",
                "ticker": symbol,
                "date": date_str,
                "shares": abs(qty_val),
                "price_per_share": price_val,
                "fees": fees,
                "dividend_amount": None,
                "notes": "",
            })
        elif amount_val is None:
            filtered_count += 1
        elif amount_val < 0 and qty_val is not None and qty_val > 0 and looks_like_dividend:
            kept.append({
                "type": "BUY",
                "ticker": symbol,
                "date": date_str,
                "shares": abs(qty_val),
                "price_per_share": price_val,
                "fees": fees,
                "dividend_amount": None,
                "notes": "[DRIP] Dividend Reinvestment",
            })
        elif amount_val > 0 and looks_like_dividend:
            kept.append({
                "type": "DIVIDEND",
                "ticker": symbol,
                "date": date_str,
                "shares": None,
                "price_per_share": None,
                "fees": fees,
                "dividend_amount": round(amount_val, 2),
                "notes": activity or description or "Dividend",
            })
        else:
            filtered_count += 1

    drip_count = sum(1 for t in kept if t["type"] == "BUY" and "[DRIP]" in (t["notes"] or ""))
    buys = sum(1 for t in kept if t["type"] == "BUY")
    sells = sum(1 for t in kept if t["type"] == "SELL")
    divs = sum(1 for t in kept if t["type"] == "DIVIDEND")

    return {
        "account_name": account_name,
        "transactions": kept,
        "summary": {
            "buys": buys,
            "sells": sells,
            "dividends": divs,
            "filtered": filtered_count,
            "drip_detected": drip_count,
            "splits_applied": 0,
        },
    }


# Interactive Brokers (Activity Statement positions + Transaction History)

_IB_SECTION_KINDS_DATA = {"Data"}
_IB_OPTION_DESC_RE = re.compile(
    r"^\S+\s+\d{1,2}[A-Z]{3}\d{2}\s+\d+(?:\.\d+)?\s+[CP]$",
    re.I,
)
_IB_OCC_RE = re.compile(r"^[A-Z][A-Z0-9.]*\d{6}[CP]\d{8}$", re.I)
_IB_PREFERRED_RE = re.compile(r"^([A-Z][A-Z0-9.]{0,9})\s+PR([A-Z])$", re.I)
_IB_CLASS_SHARE_RE = re.compile(r"^([A-Z][A-Z0-9.]{0,9})\s+([A-Z])$", re.I)
_IB_RIGHTS_RE = re.compile(r"\s+(RTWI|RTS|RIGHTS?|WTS|WARRANT)\b", re.I)
_IB_DESC_TICKER_RE = re.compile(r"^([A-Z0-9.]+(?:\s+[A-Z0-9.]+)?)\s*\(")
_IB_BUY_TYPES = {"buy", "bought"}
_IB_SELL_TYPES = {"sell", "sold"}
_IB_DIRECTIONAL_TYPES = {"assignment", "exercise"}
_IB_DIVIDEND_TYPES = {
    "dividend",
    "dividends",
    "payment in lieu",
    "payment in lieu of dividends",
    "payment in lieu of dividend",
}


def _ib_has_section(rows, section_name):
    return any(str(row[0] or "").strip() == section_name for row in rows if row)


def _ib_section_records(rows, section_name, kinds=_IB_SECTION_KINDS_DATA):
    """Yield dicts for one IB Activity/Flex CSV section.

    IB files prefix every row with ``Section,Kind`` and then the section's
    own columns. Kind is Header, Data, Total, or SubTotal.
    """
    header = None
    for row in rows:
        if not row or str(row[0] or "").strip() != section_name:
            continue
        kind = str(row[1] if len(row) > 1 else "").strip()
        if kind == "Header":
            header = [str(cell or "").strip() for cell in row[2:]]
            continue
        if header is None or kind not in kinds:
            continue
        values = list(row[2:]) + [None] * max(0, len(header) - max(0, len(row) - 2))
        record = {"_kind": kind}
        for key, value in zip(header, values):
            if key and key not in record:
                record[key] = value
        yield record


def _ib_kv_section(rows, section_name):
    mapping = {}
    for rec in _ib_section_records(rows, section_name):
        key = str(rec.get("Field Name") or "").strip()
        if key:
            mapping[key] = rec.get("Field Value")
    return mapping


def _ib_is_ib_file(rows):
    broker = str(_ib_kv_section(rows, "Statement").get("BrokerName") or "").lower()
    if "interactive broker" in broker:
        return True
    for row in rows[:80]:
        if not row:
            continue
        section = str(row[0] or "").strip()
        kind = str(row[1] if len(row) > 1 else "").strip()
        if kind in {"Header", "Data"} and section in {
            "Open Positions",
            "Transaction History",
            "Account Information",
            "Net Asset Value",
            "Financial Instrument Information",
        }:
            return True
    return False


def _ib_normalize_ticker(symbol):
    """Turn IB symbols into the app's ticker spelling, or None to skip.

    Preferred shares such as ``CIM PRB`` become ``CIM-PRB``. Class shares
    such as ``PBR A`` become ``PBR-A``. Option contracts and rights are
    dropped so they are not stored as equity holdings.
    """
    text = str(symbol or "").strip().upper()
    if not text or text in {"-", "--"}:
        return None
    compact = re.sub(r"\s+", "", text)
    if _IB_OCC_RE.match(compact) or _IB_OPTION_DESC_RE.match(text):
        return None
    if _IB_RIGHTS_RE.search(text):
        return None
    preferred = _IB_PREFERRED_RE.match(text)
    if preferred:
        ticker = f"{preferred.group(1)}-PR{preferred.group(2)}"
        return ticker if TICKER_RE.match(ticker) else None
    class_share = _IB_CLASS_SHARE_RE.match(text)
    if class_share:
        ticker = f"{class_share.group(1)}-{class_share.group(2)}"
        return ticker if TICKER_RE.match(ticker) else None
    if " " in text:
        return None
    if TICKER_RE.match(text):
        return text
    return None


def _ib_ticker_from_description(description):
    text = str(description or "").strip()
    match = _IB_DESC_TICKER_RE.match(text)
    if not match:
        return None
    return _ib_normalize_ticker(match.group(1))


def _ib_is_option_asset(category):
    return "option" in str(category or "").strip().lower()


def _ib_is_stock_asset(category):
    text = str(category or "").strip().lower()
    return text in {"stocks", "stock", "etf", "adr", "fund", "equity"}


def _ib_is_summary_row(record):
    disc = str(record.get("DataDiscriminator") or "").strip().lower()
    return disc in {"", "summary", "data"}


def _ib_parse_date(raw):
    text = str(raw or "").strip()
    if not text or text in {"-", "--"}:
        return None
    return _parse_date_str(text.split(",")[0].strip())


def _ib_usd_price(qty, price, gross_amount):
    if qty and gross_amount is not None and abs(qty) > 0:
        return abs(gross_amount) / abs(qty)
    return price


def _ib_fx_rates(rows):
    """USD-per-unit close prices from Forex Balances, with CAD total fallback."""
    rates = {"USD": 1.0}
    for rec in _ib_section_records(rows, "Forex Balances"):
        quote = str(rec.get("Description") or "").strip().upper()
        close = _safe_float(rec.get("Close Price"))
        if quote and close:
            rates[quote] = close

    pending_local = None
    pending_ccy = None
    header = None
    for row in rows:
        if not row or str(row[0] or "").strip() != "Open Positions":
            continue
        kind = str(row[1] if len(row) > 1 else "").strip()
        if kind == "Header":
            header = [str(cell or "").strip() for cell in row[2:]]
            continue
        if kind != "Total" or not header:
            continue
        rec = _row_record(header, row[2:])
        ccy = str(rec.get("Currency") or "").strip().upper()
        value = _safe_float(rec.get("Value"))
        if ccy and ccy != "USD" and value:
            pending_local = value
            pending_ccy = ccy
        elif ccy == "USD" and pending_local and pending_ccy and value:
            rates.setdefault(pending_ccy, value / pending_local)
            pending_local = None
            pending_ccy = None
    return rates


def _ib_ending_cash_usd(rows):
    for rec in _ib_section_records(rows, "Cash Report"):
        name = str(rec.get("Currency Summary") or "").strip().lower()
        currency = str(rec.get("Currency") or "").strip().lower()
        if name == "ending cash" and currency in {
            "base currency summary", "base currency", "usd",
        }:
            value = _safe_float(rec.get("Total"))
            if value is not None:
                return value
    for rec in _ib_section_records(rows, "Net Asset Value"):
        asset = str(rec.get("Asset Class") or "").strip().lower()
        if asset.startswith("cash"):
            value = _safe_float(rec.get("Current Total"))
            if value is not None:
                return value
    return 0.0


def _ib_instrument_descriptions(rows):
    descriptions = {}
    for rec in _ib_section_records(rows, "Financial Instrument Information"):
        symbol = str(rec.get("Symbol") or "").strip()
        desc = clean_security_description(rec.get("Description"))
        if not symbol or not desc:
            continue
        descriptions[symbol] = desc
        normalized = _ib_normalize_ticker(symbol)
        if normalized:
            descriptions.setdefault(normalized, desc)
    return descriptions


def _ib_account_fields(rows):
    info = _ib_kv_section(rows, "Account Information")
    account = str(info.get("Account") or "").strip()
    name = str(info.get("Name") or "").strip()
    number = account.split()[0] if account else ""
    account_name = account or name or number
    return account_name, name, number


def _ib_statement_as_of(rows):
    """Return the ending date of the IB statement period, when present."""
    period = str(_ib_kv_section(rows, "Statement").get("Period") or "").strip()
    if not period:
        return None
    candidates = [part.strip() for part in re.split(r"\s+-\s+", period) if part.strip()]
    for candidate in reversed(candidates):
        parsed = _parse_date_str(candidate)
        if parsed:
            return parsed
        for fmt in ("%B %d, %Y", "%b %d, %Y"):
            try:
                return datetime.strptime(candidate, fmt).date().isoformat()
            except ValueError:
                continue
    return None


def _ib_merge_positions(positions):
    merged = {}
    for pos in positions:
        ticker = pos["ticker"]
        if ticker not in merged:
            merged[ticker] = dict(pos)
            continue
        existing = merged[ticker]
        total_qty = existing["quantity"] + pos["quantity"]
        total_purchase = existing["purchase_value"] + pos["purchase_value"]
        total_current = existing["current_value"] + pos["current_value"]
        total_gain = existing["gain_or_loss"] + pos["gain_or_loss"]
        existing["quantity"] = total_qty
        existing["purchase_value"] = round(total_purchase, 2)
        existing["current_value"] = round(total_current, 2)
        existing["gain_or_loss"] = round(total_gain, 2)
        existing["cost_per_share"] = round(total_purchase / total_qty, 4) if total_qty else 0
        existing["current_price"] = pos["current_price"] or existing["current_price"]
        if not existing["description"]:
            existing["description"] = pos["description"]
        if not existing["asset_type"]:
            existing["asset_type"] = pos["asset_type"]
    return list(merged.values())


def _ib_txn_row(txn_type, ticker, date_str, shares, price, fees, dividend_amount, notes):
    return {
        "type": txn_type,
        "ticker": ticker,
        "date": date_str,
        "shares": shares,
        "price_per_share": price,
        "fees": fees,
        "dividend_amount": dividend_amount,
        "notes": notes,
    }


def parse_interactive_brokers_positions(file_path, filename):
    """Parse an Interactive Brokers Activity Statement CSV as current positions.

    IB's positions export is a multi-section Activity Statement, not a flat
    table. Open Positions is the source of truth for shares and cost basis.
    Options, rights, and non-stock rows are skipped. Non-USD stock rows are
    converted with the statement's FX close.
    """
    rows = _read_table_rows(file_path, filename)
    if not rows:
        raise ValueError("The file is empty.")
    if _ib_has_section(rows, "Transaction History") and not _ib_has_section(rows, "Open Positions"):
        raise ValueError(
            "This is an Interactive Brokers Transaction History export. "
            "Import it with 'Interactive Brokers (Transactions)'."
        )
    if not _ib_has_section(rows, "Open Positions"):
        raise ValueError(
            "Could not find an Open Positions section. In Interactive Brokers, "
            "download an Activity Statement CSV (Performance / Reports > Statements) "
            "and import it as Interactive Brokers (Positions)."
        )

    descriptions = _ib_instrument_descriptions(rows)
    fx_rates = _ib_fx_rates(rows)
    positions = []
    filtered_count = 0
    options_count = 0
    options_value = 0.0

    for rec in _ib_section_records(rows, "Open Positions"):
        if not _ib_is_summary_row(rec):
            filtered_count += 1
            continue
        category = rec.get("Asset Category")
        symbol = str(rec.get("Symbol") or "").strip()
        if _ib_is_option_asset(category):
            options_count += 1
            currency = str(rec.get("Currency") or "USD").strip().upper() or "USD"
            fx = fx_rates.get(currency)
            if fx is not None:
                options_value += (_safe_float(rec.get("Value")) or 0.0) * fx
            continue
        if not _ib_is_stock_asset(category):
            filtered_count += 1
            continue
        ticker = _ib_normalize_ticker(symbol)
        if not ticker:
            filtered_count += 1
            continue
        qty = _safe_float(rec.get("Quantity"))
        if qty is None or qty <= 0:
            filtered_count += 1
            continue

        currency = str(rec.get("Currency") or "USD").strip().upper() or "USD"
        fx = fx_rates.get(currency)
        if fx is None:
            filtered_count += 1
            continue

        cost_price = (_safe_float(rec.get("Cost Price")) or 0.0) * fx
        cost_basis = (_safe_float(rec.get("Cost Basis")) or 0.0) * fx
        close_price = (_safe_float(rec.get("Close Price")) or 0.0) * fx
        value = (_safe_float(rec.get("Value")) or 0.0) * fx
        unrealized = _safe_float(rec.get("Unrealized P/L"))
        if unrealized is not None:
            unrealized *= fx

        purchase_value = cost_basis if cost_basis else qty * cost_price
        current_value = value if value else qty * close_price
        if qty > 0 and purchase_value:
            cost_price = purchase_value / qty
        gain = unrealized if unrealized is not None else round(current_value - purchase_value, 2)

        positions.append({
            "ticker": ticker,
            "description": descriptions.get(symbol) or descriptions.get(ticker) or "",
            "quantity": qty,
            "cost_per_share": cost_price or 0,
            "current_price": close_price or 0,
            "purchase_value": round(purchase_value, 2),
            "current_value": round(current_value, 2),
            "gain_or_loss": round(gain, 2) if gain is not None else round(current_value - purchase_value, 2),
            "dividend_yield": None,
            "reinvest_dividends": None,
            "asset_type": str(category or "Stocks").strip(),
        })

    positions = _ib_merge_positions(positions)
    if positions and all(p["purchase_value"] == 0 for p in positions):
        raise ValueError(
            "No cost basis data found — every position has a $0 cost. "
            "This usually means a Transaction History file was selected with the "
            "Positions format. Use 'Interactive Brokers (Transactions)' for that file."
        )

    cash = _ib_ending_cash_usd(rows)
    account_name, _, account_number = _ib_account_fields(rows)
    return {
        "positions": positions,
        "account_name": account_name,
        "account_number": account_number,
        "summary": {
            "holdings": len(positions),
            "filtered": filtered_count,
            "options": options_count,
            "options_value": round(options_value, 2),
            "cash": round(cash, 2),
            "account_value": round(
                sum(position["current_value"] for position in positions) + cash,
                2,
            ),
        },
        "as_of": _ib_statement_as_of(rows),
        "format_type": "positions",
        "source_format": "interactive_brokers",
    }


def _ib_append_trade(kept, filtered_count, ticker, date_str, qty, price, fees, notes, explicit_type=None):
    if not ticker or not date_str:
        return filtered_count + 1
    if qty is None or qty == 0:
        return filtered_count + 1
    txn_type = explicit_type or ("BUY" if qty > 0 else "SELL")
    kept.append(_ib_txn_row(
        txn_type,
        ticker,
        date_str,
        abs(qty),
        price,
        abs(fees or 0),
        None,
        notes,
    ))
    return filtered_count


def _ib_append_dividend(kept, filtered_count, ticker, date_str, amount, notes):
    if not ticker or not date_str or amount is None:
        return filtered_count + 1
    kept.append(_ib_txn_row(
        "DIVIDEND",
        ticker,
        date_str,
        None,
        None,
        0.0,
        round(amount, 2),
        notes,
    ))
    return filtered_count


def _ib_parse_transaction_history(rows):
    kept = []
    filtered_count = 0
    for rec in _ib_section_records(rows, "Transaction History"):
        ttype = str(rec.get("Transaction Type") or "").strip()
        ttype_key = ttype.lower()
        symbol = str(rec.get("Symbol") or "").strip()
        description = str(rec.get("Description") or "").strip()
        ticker = _ib_normalize_ticker(symbol) or _ib_ticker_from_description(description)
        date_str = _ib_parse_date(rec.get("Date"))
        qty = _safe_float(rec.get("Quantity"))
        price = _safe_float(rec.get("Price"))
        gross = _safe_float(rec.get("Gross Amount"))
        net = _safe_float(rec.get("Net Amount"))
        fees = abs(_safe_float(rec.get("Commission")) or 0.0)

        if (
            ttype_key in _IB_BUY_TYPES
            or ttype_key in _IB_SELL_TYPES
            or ttype_key in _IB_DIRECTIONAL_TYPES
        ):
            usd_price = _ib_usd_price(qty, price, gross)
            notes = "Assignment" if ttype_key == "assignment" else (
                "Exercise" if ttype_key == "exercise" else ""
            )
            if ttype_key in _IB_BUY_TYPES:
                explicit = "BUY"
            elif ttype_key in _IB_SELL_TYPES:
                explicit = "SELL"
            else:
                description_key = description.lower()
                if description_key.startswith(("sell ", "sold ")):
                    explicit = "SELL"
                elif description_key.startswith(("buy ", "bought ")):
                    explicit = "BUY"
                else:
                    explicit = "BUY" if (qty or 0) > 0 else "SELL"
            filtered_count = _ib_append_trade(
                kept, filtered_count, ticker, date_str, qty, usd_price, fees, notes, explicit,
            )
            continue

        if ttype_key in _IB_DIVIDEND_TYPES:
            amount = gross if gross is not None else net
            notes = ttype or "Dividend"
            filtered_count = _ib_append_dividend(
                kept, filtered_count, ticker, date_str, amount, notes,
            )
            continue

        filtered_count += 1
    return kept, filtered_count


def _ib_parse_statement_activity(rows):
    """Read buys, sells, and dividends out of an Activity Statement."""
    kept = []
    filtered_count = 0
    fx_rates = _ib_fx_rates(rows)
    for rec in _ib_section_records(rows, "Trades"):
        if str(rec.get("DataDiscriminator") or "").strip().lower() not in {"", "order", "trade"}:
            continue
        if _ib_is_option_asset(rec.get("Asset Category")):
            filtered_count += 1
            continue
        if rec.get("Asset Category") and not _ib_is_stock_asset(rec.get("Asset Category")):
            filtered_count += 1
            continue
        ticker = _ib_normalize_ticker(rec.get("Symbol"))
        date_str = _ib_parse_date(rec.get("Date/Time") or rec.get("Date"))
        qty = _safe_float(rec.get("Quantity"))
        currency = str(rec.get("Currency") or "USD").strip().upper() or "USD"
        fx = fx_rates.get(currency)
        if fx is None:
            filtered_count += 1
            continue
        proceeds = _safe_float(rec.get("Proceeds"))
        local_price = _ib_usd_price(qty, _safe_float(rec.get("T. Price")), proceeds)
        price = local_price * fx if local_price is not None else None
        fees = abs(_safe_float(rec.get("Comm/Fee")) or 0.0) * fx
        filtered_count = _ib_append_trade(
            kept, filtered_count, ticker, date_str, qty, price, fees, "",
        )

    for section, default_note in (
        ("Dividends", "Dividend"),
        ("Payment in Lieu of Dividends", "Payment in Lieu"),
    ):
        for rec in _ib_section_records(rows, section):
            description = str(rec.get("Description") or "").strip()
            ticker = _ib_normalize_ticker(rec.get("Symbol")) or _ib_ticker_from_description(description)
            date_str = _ib_parse_date(rec.get("Date"))
            amount = _safe_float(rec.get("Amount"))
            note = default_note
            if "payment in lieu" in description.lower():
                note = "Payment in Lieu"
            filtered_count = _ib_append_dividend(
                kept, filtered_count, ticker, date_str, amount, note,
            )
    return kept, filtered_count


def parse_interactive_brokers_transactions(file_path, filename):
    """Parse Interactive Brokers Transaction History, or an Activity Statement.

    The dedicated Transaction History CSV is preferred. An Activity Statement
    is accepted too: Trades become BUY/SELL and Dividends / payment-in-lieu
    become DIVIDEND. Options, interest, fees, withholding, and cash movements
    are skipped.
    """
    rows = _read_table_rows(file_path, filename)
    if not rows:
        raise ValueError("The file is empty.")

    if _ib_has_section(rows, "Transaction History"):
        kept, filtered_count = _ib_parse_transaction_history(rows)
    elif _ib_has_section(rows, "Trades") or _ib_has_section(rows, "Dividends"):
        kept, filtered_count = _ib_parse_statement_activity(rows)
    else:
        raise ValueError(
            "Could not find a Transaction History, Trades, or Dividends section. "
            "In Interactive Brokers, export Transaction History CSV, or import an "
            "Activity Statement CSV with 'Interactive Brokers (Transactions)'."
        )

    drip_count = _detect_drip(kept)
    account_name, _, account_number = _ib_account_fields(rows)
    if not account_name:
        for rec in _ib_section_records(rows, "Transaction History"):
            acct = str(rec.get("Account") or "").strip()
            if acct and acct not in {"-", "--"}:
                account_name = acct
                account_number = acct
                break
    if not account_name:
        title = str(_ib_kv_section(rows, "Statement").get("Title") or "")
        if title:
            account_name = title
    buys = sum(1 for t in kept if t["type"] == "BUY")
    sells = sum(1 for t in kept if t["type"] == "SELL")
    divs = sum(1 for t in kept if t["type"] == "DIVIDEND")
    return {
        "account_name": account_name,
        "account_number": account_number,
        "transactions": kept,
        "summary": {
            "buys": buys,
            "sells": sells,
            "dividends": divs,
            "filtered": filtered_count,
            "drip_detected": drip_count,
            "splits_applied": 0,
        },
        "source_format": "interactive_brokers_transactions",
    }


PARSERS = {
    "generic_transactions": parse_generic_transactions,
    "snowball": parse_snowball_csv,
    "snowball_holdings": parse_snowball_holdings_csv,
    "snowball_categories": parse_snowball_categories_csv,
    "schwab": parse_schwab_csv,
    "schwab_all_accounts": parse_schwab_all_accounts_csv,
    "schwab_transactions": parse_schwab_transactions_csv,
    "etrade": parse_etrade_csv,
    "etrade_transactions": parse_etrade_transactions_xlsx,
    "fidelity": parse_fidelity_positions_xlsx,
    "fidelity_all_accounts": parse_fidelity_all_accounts_positions,
    "fidelity_transactions": parse_fidelity_transactions_xlsx,
    "robinhood": parse_robinhood_positions_pdf,
    "robinhood_transactions": parse_robinhood_transactions_csv,
    "shear_group": parse_shear_group_positions,
    "shear_group_activity": parse_shear_group_activity,
    "shear_group_all_accounts": parse_shear_group_all_accounts_positions,
    "shear_group_all_accounts_activity": parse_shear_group_all_accounts_activity,
    "interactive_brokers": parse_interactive_brokers_positions,
    "interactive_brokers_transactions": parse_interactive_brokers_transactions,
}

# Labels shown in the UI format dropdown
PARSER_LABELS = {
    "generic_transactions": "Generic Transactions",
    "snowball": "Snowball Analytics",
    "snowball_holdings": "Snowball Holdings (Migration)",
    "snowball_categories": "Snowball Categories",
    "schwab": "Charles Schwab (Positions)",
    "schwab_all_accounts": "Charles Schwab (All Accounts Positions)",
    "schwab_transactions": "Charles Schwab (Transactions)",
    "etrade": "E*Trade (Positions)",
    "etrade_transactions": "E*Trade (Transactions)",
    "fidelity": "Fidelity (Positions)",
    "fidelity_all_accounts": "Fidelity (All Accounts Positions)",
    "fidelity_transactions": "Fidelity (Transactions)",
    "robinhood": "Robinhood (Positions PDF)",
    "robinhood_transactions": "Robinhood (Transactions)",
    "shear_group": "Shear Group (Positions)",
    "shear_group_activity": "Shear Group (Activity)",
    "shear_group_all_accounts": "Shear Group (All Accounts Positions)",
    "shear_group_all_accounts_activity": "Shear Group (All Accounts Activity)",
    "interactive_brokers": "Interactive Brokers (Positions)",
    "interactive_brokers_transactions": "Interactive Brokers (Transactions)",
}
