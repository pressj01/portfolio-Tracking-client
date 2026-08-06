"""Normalize generic and broker option-transaction exports.

The existing portfolio import intentionally ignores option contracts.  This
module reads the same kind of broker activity file through a separate path and
turns each recognized row into an option execution.  It never writes to the
equity ``transactions`` or holdings tables.
"""

from __future__ import annotations

import csv
import hashlib
import re
from collections import defaultdict
from datetime import date, datetime


SUPPORTED_FORMATS = {
    "generic": "Generic Options Transactions",
    "schwab": "Charles Schwab Transactions",
    "etrade": "E*TRADE Transactions",
    "fidelity": "Fidelity Transactions",
    "robinhood": "Robinhood Transactions",
    "shear_group": "Shear Group Activity",
}


HEADER_ALIASES = {
    "Date": ["Transaction Date", "Run Date", "Activity/Trade Date", "Activity Date", "Process Date", "Trade Date", "Executed At"],
    "Action": ["Type", "Transaction Type", "Transaction Code", "Trans Code", "Activity", "Activity Type", "Side"],
    "Underlying": ["Ticker", "Underlying Symbol", "Root Symbol"],
    "Symbol": ["Option Symbol", "OCC Symbol", "Instrument", "Security", "Symbol/CUSIP"],
    "Description": ["Security Description", "Instrument Description", "Name"],
    "Option Type": ["Call/Put", "Put/Call", "C/P", "Contract Type"],
    "Expiration": ["Expiration Date", "Expiry", "Exp Date"],
    "Strike": ["Strike Price"],
    "Contracts": ["Quantity", "Qty", "Quantity #"],
    "Price": ["Fill Price", "Price ($)", "Price $", "Unit Price", "Price/Contract"],
    "Fees": ["Fee", "Fees ($)", "Fees & Comm", "Fees & Commissions", "Regulatory Fees"],
    "Commission": ["Commission ($)", "Commissions"],
    "Trade ID": ["Group ID", "Strategy ID", "Position ID"],
    "Order ID": ["Transaction ID", "Activity ID", "Reference", "Reference #", "Confirmation Number"],
    "Strategy": ["Strategy Type", "Trade Type"],
    "Purpose": ["Trade Purpose", "Income Classification"],
    "Account": ["Account Name", "Account", "Account Number", "Account Nickname"],
    "Notes": ["Note", "Memo", "Comments"],
}


def _header_key(value):
    return re.sub(r"[^a-z0-9]+", "", str(value or "").strip().lower())


def _alias_map():
    result = {}
    for canonical, aliases in HEADER_ALIASES.items():
        for alias in [canonical, *aliases]:
            result.setdefault(_header_key(alias), canonical)
    return result


_ALIASES = _alias_map()


def _read_rows(file_path, filename):
    if str(filename or file_path).lower().endswith(".csv"):
        with open(file_path, "r", encoding="utf-8-sig", newline="") as handle:
            return list(csv.reader(handle))

    import openpyxl

    workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    try:
        sheet = workbook["Transactions"] if "Transactions" in workbook.sheetnames else workbook.active
        return [list(row) for row in sheet.iter_rows(values_only=True)]
    finally:
        workbook.close()


def _find_records(rows):
    best = None
    for index, row in enumerate(rows[:50]):
        header = [_ALIASES.get(_header_key(cell), str(cell or "").strip()) for cell in row]
        present = set(header)
        if "Date" not in present or "Action" not in present:
            continue
        score = len(present & set(HEADER_ALIASES))
        if best is None or score > best[0]:
            best = (score, index, header)
    if best is None:
        raise ValueError(
            "Could not find option transaction columns. The file needs at least Date and Action columns."
        )

    _, header_index, header = best
    records = []
    for row_number, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
        if not any(value is not None and str(value).strip() for value in row):
            continue
        values = list(row) + [None] * max(0, len(header) - len(row))
        record = {"_row": row_number}
        for key, value in zip(header, values):
            if not key:
                continue
            if key not in record or record[key] in (None, ""):
                record[key] = value
        records.append(record)
    return records


def _safe_float(value):
    if value is None or str(value).strip() in {"", "-", "--"}:
        return None
    text = str(value).strip().replace("$", "").replace(",", "")
    negative = text.startswith("(") and text.endswith(")")
    try:
        number = float(text.strip("()"))
        return -number if negative else number
    except (TypeError, ValueError):
        return None


def _date_string(value):
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value or "").strip()
    for fmt in (
        "%Y-%m-%d",
        "%Y-%m-%d %H:%M:%S",
        "%m/%d/%Y",
        "%m/%d/%y",
        "%m/%d/%Y %H:%M:%S",
        "%b %d, %Y",
        "%b-%d-%Y",
    ):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def parse_occ_symbol(value):
    """Return OCC contract fields from padded or compact OSI symbols."""
    raw = str(value or "").strip().upper()
    compact = re.sub(r"\s+", "", raw.removeprefix("O:").lstrip("-"))
    match = re.fullmatch(r"([A-Z][A-Z0-9.]{0,5})(\d{6})([CP])(\d{8})", compact)
    if not match:
        return None
    underlying, raw_date, option_type, raw_strike = match.groups()
    try:
        expiration = datetime.strptime(raw_date, "%y%m%d").date().isoformat()
    except ValueError:
        return None
    return {
        "underlying": underlying,
        "expiration": expiration,
        "option_type": "CALL" if option_type == "C" else "PUT",
        "strike": int(raw_strike) / 1000,
        "occ_symbol": compact,
    }


def parse_option_descriptor(value):
    """Read common broker descriptions such as ``SPY 08/21/26 600 C``."""
    text = " ".join(str(value or "").upper().replace("$", " ").split())
    occ_match = re.search(r"[A-Z][A-Z0-9.]{0,5}\s*\d{6}[CP]\d{8}", text)
    if occ_match:
        parsed = parse_occ_symbol(occ_match.group(0))
        if parsed:
            return parsed

    cp = r"(?:C|CALL|P|PUT)"
    patterns = [
        rf"(?P<underlying>[A-Z][A-Z0-9.]{{0,9}})\s+(?P<date>\d{{1,2}}/\d{{1,2}}/\d{{2,4}})\s+(?P<strike>\d+(?:\.\d+)?)\s*(?P<cp>{cp})\b",
        rf"(?P<underlying>[A-Z][A-Z0-9.]{{0,9}})\s+(?P<date>\d{{1,2}}/\d{{1,2}}/\d{{2,4}})\s+(?P<cp>{cp})\s*(?P<strike>\d+(?:\.\d+)?)\b",
        rf"(?P<cp>{cp})\s+(?P<underlying>[A-Z][A-Z0-9.]{{0,9}})\s+(?P<date>\d{{1,2}}/\d{{1,2}}/\d{{2,4}})\s+(?P<strike>\d+(?:\.\d+)?)\b",
    ]
    for pattern in patterns:
        match = re.search(pattern, text)
        if not match:
            continue
        expiration = _date_string(match.group("date"))
        if not expiration:
            continue
        option_type = match.group("cp")
        return {
            "underlying": match.group("underlying"),
            "expiration": expiration,
            "option_type": "CALL" if option_type in {"C", "CALL"} else "PUT",
            "strike": float(match.group("strike")),
            "occ_symbol": None,
        }
    return None


def _normalize_option_type(value):
    text = str(value or "").strip().upper()
    if text in {"C", "CALL", "CALLS"}:
        return "CALL"
    if text in {"P", "PUT", "PUTS"}:
        return "PUT"
    return None


def _normalize_action(value):
    text = " ".join(str(value or "").upper().replace("_", " ").replace("-", " ").split())
    compact = text.replace(" ", "")
    direct = {"BTO": "BTO", "STO": "STO", "BTC": "BTC", "STC": "STC"}
    if compact in direct:
        return direct[compact], None
    if "EXPIR" in text:
        return "EXPIRE", None
    if "ASSIGN" in text:
        return "ASSIGN", None
    if "EXERCI" in text:
        return "EXERCISE", None

    is_buy = "BUY" in text or "BOUGHT" in text
    is_sell = "SELL" in text or "SOLD" in text
    is_open = "OPEN" in text
    is_close = "CLOS" in text
    if is_buy and is_open:
        return "BTO", None
    if is_sell and is_open:
        return "STO", None
    if is_buy and is_close:
        return "BTC", None
    if is_sell and is_close:
        return "STC", None
    if is_buy:
        return "BTO", "Action did not identify opening or closing; assumed Buy to Open."
    if is_sell:
        return "STO", "Action did not identify opening or closing; assumed Sell to Open."
    return None, "Unrecognized option action."


def _strategy_for_legs(legs):
    positions = [(row["option_type"], row["position_side"], float(row["strike"])) for row in legs]
    if len(positions) == 1:
        option_type, side, _ = positions[0]
        if side == "SHORT":
            return "Short Call" if option_type == "CALL" else "Short Put"
        return "Long Call" if option_type == "CALL" else "Long Put"

    if len(positions) == 2:
        first, second = positions
        types = {first[0], second[0]}
        sides = {first[1], second[1]}
        if len(types) == 1 and sides == {"LONG", "SHORT"}:
            option_type = first[0]
            long_strike = next(strike for kind, side, strike in positions if side == "LONG")
            short_strike = next(strike for kind, side, strike in positions if side == "SHORT")
            if option_type == "CALL":
                return "Bull Call Spread" if long_strike < short_strike else "Bear Call Spread"
            return "Bear Put Spread" if long_strike > short_strike else "Bull Put Spread"
        if types == {"CALL", "PUT"} and len(sides) == 1:
            same_strike = first[2] == second[2]
            prefix = "Long" if first[1] == "LONG" else "Short"
            return f"{prefix} {'Straddle' if same_strike else 'Strangle'}"

    if len(positions) == 4:
        calls = sorted([item for item in positions if item[0] == "CALL"], key=lambda item: item[2])
        puts = sorted([item for item in positions if item[0] == "PUT"], key=lambda item: item[2])
        if len(calls) == 2 and len(puts) == 2 and {item[1] for item in calls} == {"LONG", "SHORT"} and {item[1] for item in puts} == {"LONG", "SHORT"}:
            short_strikes = [item[2] for item in positions if item[1] == "SHORT"]
            return "Iron Butterfly" if len(set(short_strikes)) == 1 else "Iron Condor"
    if len(positions) == 3:
        return "Butterfly / Custom"
    return "Custom"


def _default_purpose(strategy):
    income_strategies = {
        "Short Call", "Short Put", "Bull Put Spread", "Bear Call Spread",
        "Iron Condor", "Iron Butterfly", "Short Straddle", "Short Strangle",
    }
    return "Income" if strategy in income_strategies else "Directional"


def _dedupe_hash(row, source_format):
    parts = [
        source_format,
        row.get("account"),
        row.get("external_id"),
        row.get("executed_at"),
        row.get("action"),
        row.get("underlying"),
        row.get("expiration"),
        row.get("option_type"),
        row.get("strike"),
        row.get("contracts"),
        row.get("price"),
        row.get("fees"),
    ]
    return hashlib.sha256("|".join(str(value or "") for value in parts).encode("utf-8")).hexdigest()


def parse_option_transactions(file_path, filename, source_format="generic"):
    """Return normalized executions, review warnings, and import counts."""
    source_format = str(source_format or "generic").strip().lower()
    if source_format not in SUPPORTED_FORMATS:
        raise ValueError(f"Unsupported option transaction format: {source_format}")

    records = _find_records(_read_rows(file_path, filename))
    executions = []
    filtered = []
    for record in records:
        action, action_warning = _normalize_action(record.get("Action"))
        executed_at = _date_string(record.get("Date"))
        parsed = None
        for candidate in (record.get("Symbol"), record.get("Description")):
            parsed = parse_occ_symbol(candidate) or parse_option_descriptor(candidate)
            if parsed:
                break

        separate_type = _normalize_option_type(record.get("Option Type"))
        separate_expiration = _date_string(record.get("Expiration"))
        separate_strike = _safe_float(record.get("Strike"))
        underlying = str(record.get("Underlying") or (parsed or {}).get("underlying") or "").strip().upper()
        option_type = separate_type or (parsed or {}).get("option_type")
        expiration = separate_expiration or (parsed or {}).get("expiration")
        strike = separate_strike if separate_strike is not None else (parsed or {}).get("strike")
        occ_symbol = (parsed or {}).get("occ_symbol")
        contracts = _safe_float(record.get("Contracts"))
        price = _safe_float(record.get("Price"))
        fees = (_safe_float(record.get("Fees")) or 0) + (_safe_float(record.get("Commission")) or 0)

        missing = []
        if not action:
            missing.append("action")
        if not executed_at:
            missing.append("date")
        if not underlying:
            missing.append("underlying")
        if not option_type:
            missing.append("call/put")
        if not expiration:
            missing.append("expiration")
        if strike is None:
            missing.append("strike")
        if contracts is None or contracts == 0:
            missing.append("contracts")
        if price is None and action not in {"EXPIRE", "ASSIGN", "EXERCISE"}:
            missing.append("price")
        if missing:
            filtered.append({"row": record["_row"], "reason": f"Missing or invalid {', '.join(missing)}"})
            continue

        trade_id = str(record.get("Trade ID") or "").strip() or None
        external_id = str(record.get("Order ID") or "").strip() or None
        phase = "OPEN" if action in {"BTO", "STO"} else "CLOSE"
        group_key = (
            f"trade:{trade_id}" if trade_id else
            f"order:{external_id}" if external_id else
            f"auto:{underlying}:{executed_at}:{phase}"
        )
        position_side = "LONG" if action in {"BTO", "STC"} else "SHORT"
        warnings = [warning for warning in [action_warning] if warning]
        normalized = {
            "source_row": record["_row"],
            "account": str(record.get("Account") or "").strip() or None,
            "executed_at": executed_at,
            "action": action,
            "underlying": underlying,
            "option_type": option_type,
            "expiration": expiration,
            "strike": float(strike),
            "contracts": int(abs(contracts)),
            "price": abs(float(price or 0)),
            "fees": round(abs(float(fees)), 2),
            "occ_symbol": occ_symbol,
            "position_side": position_side,
            "trade_group_id": trade_id,
            "group_key": group_key,
            "external_id": external_id,
            "strategy_type": str(record.get("Strategy") or "").strip() or None,
            "purpose": str(record.get("Purpose") or "").strip().title() or None,
            "notes": str(record.get("Notes") or record.get("Description") or "").strip(),
            "warnings": warnings,
        }
        normalized["dedupe_hash"] = _dedupe_hash(normalized, source_format)
        executions.append(normalized)

    opening_groups = defaultdict(list)
    for execution in executions:
        if execution["action"] in {"BTO", "STO"}:
            opening_groups[execution["group_key"]].append(execution)
    for group in opening_groups.values():
        specified = next((row["strategy_type"] for row in group if row["strategy_type"]), None)
        strategy = specified or _strategy_for_legs(group)
        purpose = next((row["purpose"] for row in group if row["purpose"]), None) or _default_purpose(strategy)
        for row in group:
            row["strategy_type"] = strategy
            row["purpose"] = purpose

    executions.sort(key=lambda row: (row["executed_at"], row["source_row"]))
    warning_count = sum(bool(row["warnings"]) for row in executions)
    return {
        "source_format": source_format,
        "source_label": SUPPORTED_FORMATS[source_format],
        "executions": executions,
        "filtered_rows": filtered,
        "summary": {
            "recognized": len(executions),
            "opening": sum(row["action"] in {"BTO", "STO"} for row in executions),
            "closing": sum(row["action"] not in {"BTO", "STO"} for row in executions),
            "needs_review": warning_count,
            "filtered": len(filtered),
            "groups": len(opening_groups),
        },
    }
