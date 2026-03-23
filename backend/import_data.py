import re
import pandas as pd
from datetime import date
from config import get_connection
from database import ensure_tables_exist

# ── Column mapping for the owner's Excel spreadsheet (All Accounts sheet) ─────
COLUMN_MAP = {
    "Ticker": "ticker",
    "Description": "description",
    "Type": "classification_type",
    "Price Paid": "price_paid",
    "Current Price": "current_price",
    "% Change": "percent_change",
    "Qty": "quantity",
    "Purchase Value": "purchase_value",
    "Current Value": "current_value",
    "Gain/Loss": "gain_or_loss",
    "Gain/Loss %": "gain_or_loss_percentage",
    "Div Freq": "div_frequency",
    "DRIP": "reinvest",
    "Ex-Div Date": "ex_div_date",
    "Div/Share": "div",
    "Div Paid": "dividend_paid",
    "Est. Annual Pmt": "estim_payment_per_year",
    "Monthly Income": "approx_monthly_income",
    "8% Annual Wdraw": "withdraw_8pct_cost_annually",
    "8% Monthly Wdraw": "withdraw_8pct_per_month",
    "Cash Not Reinvest": "cash_not_reinvested",
    "Cash Reinvested": "total_cash_reinvested",
    "Yield On Cost": "annual_yield_on_cost",
    "Current Yield": "current_annual_yield",
    "% of Account": "percent_of_account",
    "Shares from Div": "shares_bought_from_dividend",
    "Shares/Year": "shares_bought_in_year",
    "Shares/Month": "shares_in_month",
    "YTD Divs": "ytd_divs",
    "Total Divs Received": "total_divs_received",
    "Paid For Itself": "paid_for_itself",
    "Date Purchased": "purchase_date",
    "Purchase Date": "purchase_date",
}

SQL_COLUMNS = list(COLUMN_MAP.values()) + ["import_date", "current_month_income"]

_VALID_TICKER = re.compile(r'^[A-Z][A-Z0-9.\-/]{0,10}$')
_EXCLUDED_TICKERS = {"TOTALS", "TOTAL", "GRAND", "SUMMARY"}


# ── Owner Excel import ─────────────────────────────────────────────────────────

def import_from_excel(file_path, sheet_name="All Accounts", profile_id=1):
    """Read the owner's Excel file and import into all_account_info.
    Returns (row_count, message).
    """
    # Auto-detect sheet if the requested name doesn't exist
    import openpyxl as _xl
    _wb_check = _xl.load_workbook(file_path, read_only=True, data_only=True)
    _available = _wb_check.sheetnames
    _wb_check.close()

    if sheet_name not in _available:
        # Try case-insensitive match first
        _lower_map = {s.lower().strip(): s for s in _available}
        _ci_match = _lower_map.get(sheet_name.lower().strip())
        if _ci_match:
            sheet_name = _ci_match
        elif sheet_name == "All Accounts":
            # Only auto-detect for the default sheet name (user didn't specify one)
            _found = None
            for sn in _available:
                try:
                    _test = pd.read_excel(file_path, sheet_name=sn, nrows=0, engine="openpyxl")
                    if "Ticker" in _test.columns:
                        _found = sn
                        break
                except Exception:
                    continue
            if _found:
                sheet_name = _found
            else:
                raise ValueError(
                    f"Sheet '{sheet_name}' not found. Available sheets: {', '.join(_available)}. "
                    f"Please enter the correct sheet name."
                )
        else:
            # User specified a sheet name that doesn't exist — don't guess, show available sheets
            raise ValueError(
                f"Sheet '{sheet_name}' not found. Available sheets: {', '.join(_available)}. "
                f"Please enter the correct sheet name."
            )

    df = pd.read_excel(file_path, sheet_name=sheet_name, engine="openpyxl")

    # ── Detect "[Month] Income" columns and capture monthly totals ──────────
    _month_pattern = re.compile(r'^([A-Za-z]+)\s+Income$')
    _current_year = date.today().year
    _monthly_income_updates = []

    _current_month = date.today().month
    _current_month_col = None

    for col_name in df.columns:
        if not isinstance(col_name, str):
            continue
        _m = _month_pattern.match(col_name)
        if not _m:
            continue
        try:
            _month_num = pd.to_datetime(_m.group(1), format='%B').month
        except Exception:
            continue
        if _month_num == _current_month:
            _current_month_col = col_name
        _mask = df['Ticker'].apply(
            lambda t: isinstance(t, str) and t.strip() != 'TOTALS' and bool(_VALID_TICKER.match(t.strip()))
        )
        _total = pd.to_numeric(df.loc[_mask, col_name], errors='coerce').fillna(0).sum()
        if _total > 0:
            _monthly_income_updates.append((_current_year, _month_num, round(float(_total), 2)))

    # Capture per-ticker current month income before columns are filtered
    _current_month_income_series = None
    if _current_month_col is not None:
        _current_month_income_series = pd.to_numeric(df[_current_month_col], errors='coerce').fillna(0)

    # Keep only mapped columns that actually exist in this sheet
    excel_cols = [c for c in COLUMN_MAP.keys() if c in df.columns]
    df = df[excel_cols].copy()
    df.rename(columns=COLUMN_MAP, inplace=True)

    # Attach per-ticker current month income
    if _current_month_income_series is not None:
        df['current_month_income'] = _current_month_income_series.reindex(df.index).fillna(0)

    # Drop rows where ticker is null or blank
    df = df.dropna(subset=["ticker"])
    df = df[df["ticker"].astype(str).str.strip() != ""]

    # Keep only valid stock tickers (exclude summary rows like TOTALS)
    df = df[df["ticker"].astype(str).str.strip().apply(
        lambda t: bool(_VALID_TICKER.match(t)) and t not in _EXCLUDED_TICKERS
    )]

    # Coerce float columns to numeric
    float_sql_cols = [
        "price_paid", "current_price", "percent_change", "quantity",
        "purchase_value", "current_value", "gain_or_loss", "gain_or_loss_percentage",
        "div", "dividend_paid", "estim_payment_per_year", "approx_monthly_income",
        "withdraw_8pct_cost_annually", "withdraw_8pct_per_month",
        "cash_not_reinvested", "total_cash_reinvested",
        "annual_yield_on_cost", "current_annual_yield", "percent_of_account",
        "shares_bought_from_dividend", "shares_bought_in_year", "shares_in_month",
        "ytd_divs", "total_divs_received", "paid_for_itself",
    ]
    for col in float_sql_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    # Coerce purchase_date
    if "purchase_date" in df.columns:
        df["purchase_date"] = pd.to_datetime(df["purchase_date"], errors="coerce")
        df["purchase_date"] = df["purchase_date"].where(df["purchase_date"].notna(), other=None)

    # Fill missing current_price with 0 instead of dropping rows
    if "current_price" in df.columns:
        df["current_price"] = df["current_price"].fillna(0)

    # ── Recompute paid_for_itself from yfinance dividend history ────────────
    if "purchase_date" in df.columns:
        import yfinance as yf
        _pd_mask = df["purchase_date"].notna()
        if _pd_mask.any():
            _dated = df.loc[_pd_mask].copy()
            _earliest = pd.Timestamp(_dated["purchase_date"].min())
            _tickers = _dated["ticker"].unique().tolist()
            _ticker_str = " ".join(_tickers)
            try:
                _raw = yf.download(
                    _ticker_str, start=_earliest.strftime("%Y-%m-%d"),
                    progress=False, auto_adjust=False, actions=True
                )
                if not _raw.empty:
                    _divs = None
                    if isinstance(_raw.columns, pd.MultiIndex):
                        if "Dividends" in _raw.columns.get_level_values(0):
                            _divs = _raw["Dividends"]
                    elif "Dividends" in _raw.columns:
                        _divs = _raw["Dividends"]

                    if _divs is not None:
                        for idx in _dated.index:
                            t = df.at[idx, "ticker"]
                            _pdate = pd.Timestamp(df.at[idx, "purchase_date"])
                            _qty = float(df.at[idx, "quantity"] or 0)
                            _pv = float(df.at[idx, "purchase_value"] or 0)

                            if isinstance(_divs, pd.DataFrame):
                                if t not in _divs.columns:
                                    continue
                                _tseries = _divs[t]
                            else:
                                _tseries = _divs

                            _since = _tseries[_tseries.index >= _pdate]
                            _total_dps = float(_since[_since > 0].sum())
                            _total_divs = _total_dps * _qty
                            _pfi = _total_divs / _pv if _pv > 0 else 0

                            df.at[idx, "total_divs_received"] = round(_total_divs, 2)
                            df.at[idx, "paid_for_itself"] = round(_pfi, 6)
            except Exception:
                pass

    # Add import_date and profile_id
    df["import_date"] = date.today().isoformat()
    df["profile_id"] = profile_id

    conn = get_connection()
    ensure_tables_exist(conn)
    cur = conn.cursor()

    # Check if data already exists for this profile (merge mode)
    existing = cur.execute(
        "SELECT ticker FROM all_account_info WHERE profile_id = ?", (profile_id,)
    ).fetchall()
    existing_tickers = {r[0] for r in existing}

    cols_to_insert = [c for c in SQL_COLUMNS if c in df.columns] + ["profile_id"]
    placeholders = ", ".join(["?"] * len(cols_to_insert))
    insert_sql = f"INSERT INTO all_account_info ({', '.join(cols_to_insert)}) VALUES ({placeholders})"

    def _row_values(row):
        values = []
        for col in cols_to_insert:
            val = row.get(col)
            if pd.isna(val):
                values.append(None)
            elif isinstance(val, pd.Timestamp):
                values.append(val.isoformat()[:10])
            else:
                values.append(val)
        return values

    if not existing_tickers:
        # First import — full load
        row_count = 0
        for _, row in df.iterrows():
            cur.execute(insert_sql, _row_values(row))
            row_count += 1
        conn.commit()
    else:
        # Merge mode — update existing, insert new
        merge_fields = [c for c in cols_to_insert if c not in ('ticker', 'profile_id')]
        inserted = 0
        updated = 0
        for _, row in df.iterrows():
            ticker = row.get('ticker')
            if not isinstance(ticker, str):
                continue
            ticker = ticker.strip().upper()

            if ticker in existing_tickers:
                sets = []
                vals = []
                for field in merge_fields:
                    val = row.get(field)
                    if pd.notna(val):
                        if isinstance(val, pd.Timestamp):
                            val = val.isoformat()[:10]
                        sets.append(f"{field} = ?")
                        vals.append(val)
                    elif field in float_sql_cols:
                        # Numeric fields: overwrite stale DB value with 0
                        sets.append(f"{field} = ?")
                        vals.append(0)
                if sets:
                    sets.append("import_date = ?")
                    vals.append(date.today().isoformat())
                    vals.extend([ticker, profile_id])
                    cur.execute(
                        f"UPDATE all_account_info SET {', '.join(sets)} WHERE ticker = ? AND profile_id = ?",
                        vals,
                    )
                    updated += 1
            else:
                cur.execute(insert_sql, _row_values(row))
                inserted += 1

        # Remove tickers no longer in the spreadsheet
        imported_tickers = set(
            row.get('ticker').strip().upper()
            for _, row in df.iterrows()
            if isinstance(row.get('ticker'), str)
        )
        removed_tickers = existing_tickers - imported_tickers
        removed = 0
        for t in removed_tickers:
            cur.execute(
                "DELETE FROM all_account_info WHERE ticker = ? AND profile_id = ?",
                (t, profile_id),
            )
            removed += 1

        row_count = updated + inserted
        conn.commit()

    # ── Upsert monthly income totals ────────────────────────────────────────
    for _yr, _mo, _amt in _monthly_income_updates:
        existing = cur.execute(
            "SELECT 1 FROM monthly_payouts WHERE year = ? AND month = ? AND profile_id = ?",
            (_yr, _mo, profile_id),
        ).fetchone()
        if existing is None:
            cur.execute(
                "INSERT INTO monthly_payouts (year, month, amount, profile_id) VALUES (?, ?, ?, ?)",
                (_yr, _mo, _amt, profile_id),
            )
        else:
            cur.execute(
                "UPDATE monthly_payouts SET amount = ? WHERE year = ? AND month = ? AND profile_id = ?",
                (_amt, _yr, _mo, profile_id),
            )
    conn.commit()

    # ── Create category assignments from classification_type ──────────────
    _CLASSIFICATION_NAMES = {
        "A": "Anchors", "B": "Boosters", "G": "Growth", "J": "Juicers",
        "BDC": "BDC", "HA": "Hedged Anchor", "GS": "Gold Silver",
        "ETF": "ETF", "EQUITY": "Equity", "CEF": "CEF", "REIT": "REIT",
    }
    if 'classification_type' in df.columns:
        for _, row in df.iterrows():
            ticker = row.get('ticker')
            ct = row.get('classification_type')
            if not isinstance(ticker, str) or not isinstance(ct, str):
                continue
            ticker = ticker.strip().upper()
            ct = ct.strip()
            cat_name = _CLASSIFICATION_NAMES.get(ct, ct)
            if not cat_name:
                continue

            # Find or create the category
            existing_cat = cur.execute(
                "SELECT id FROM categories WHERE name = ? AND profile_id = ?",
                (cat_name, profile_id),
            ).fetchone()
            if existing_cat:
                cat_id = existing_cat[0]
            else:
                max_pos = cur.execute(
                    "SELECT COALESCE(MAX(sort_order), 0) FROM categories WHERE profile_id = ?",
                    (profile_id,),
                ).fetchone()[0]
                cur.execute(
                    "INSERT INTO categories (name, target_pct, sort_order, profile_id) VALUES (?, 0, ?, ?)",
                    (cat_name, max_pos + 1, profile_id),
                )
                cat_id = cur.lastrowid

            # Assign ticker if not already assigned
            already = cur.execute(
                "SELECT 1 FROM ticker_categories WHERE ticker = ? AND category_id = ? AND profile_id = ?",
                (ticker, cat_id, profile_id),
            ).fetchone()
            if not already:
                cur.execute(
                    "INSERT INTO ticker_categories (ticker, category_id, profile_id) VALUES (?, ?, ?)",
                    (ticker, cat_id, profile_id),
                )
        conn.commit()

    conn.close()
    if not existing_tickers:
        msg = f"Imported {row_count} holdings from Excel (sheet: '{sheet_name}')."
    else:
        parts = []
        if updated: parts.append(f"updated {updated} existing")
        if inserted: parts.append(f"added {inserted} new")
        msg = f"Merge complete: {', '.join(parts)} holdings (sheet: '{sheet_name}')."
    return row_count, msg


# ── Weekly payouts import ──────────────────────────────────────────────────────

def import_weekly_payouts(file_path, profile_id=1):
    """Import weekly payout data from Weekly_Payers sheet."""
    import openpyxl
    from datetime import datetime

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True, keep_vba=True)
    ws = wb["Weekly_Payers"]

    ticker_rows = []
    for row in ws.iter_rows(min_row=2, max_row=20):
        ticker = row[0].value
        if ticker is None or not isinstance(ticker, str):
            continue
        shares = row[1].value
        dist = row[8].value
        total_div = row[9].value
        ticker_rows.append((ticker, shares, dist, total_div))

    weekly_rows = []
    for row in ws.iter_rows(min_row=22):
        pay_date_val = row[0].value
        week_val = row[1].value
        amount_val = row[2].value
        if not isinstance(pay_date_val, datetime):
            continue
        if amount_val is None:
            continue
        weekly_rows.append((
            pay_date_val.date().isoformat(),
            int(week_val) if week_val is not None else None,
            float(amount_val),
        ))
    wb.close()

    conn = get_connection()
    ensure_tables_exist(conn)
    cur = conn.cursor()

    cur.execute("DELETE FROM weekly_payout_tickers WHERE profile_id = ?", (profile_id,))
    for ticker, shares, dist, total_div in ticker_rows:
        cur.execute(
            "INSERT INTO weekly_payout_tickers (ticker, shares, distribution, total_dividend, profile_id) "
            "VALUES (?, ?, ?, ?, ?)",
            (
                ticker,
                float(shares) if shares is not None else None,
                float(dist) if dist is not None else None,
                float(total_div) if total_div is not None else None,
                profile_id,
            ),
        )

    inserted = 0
    for pay_date, week, amount in weekly_rows:
        existing = cur.execute(
            "SELECT 1 FROM weekly_payouts WHERE pay_date = ? AND profile_id = ?",
            (pay_date, profile_id),
        ).fetchone()
        if existing is None:
            cur.execute(
                "INSERT INTO weekly_payouts (pay_date, week_of_month, amount, profile_id) VALUES (?, ?, ?, ?)",
                (pay_date, week, amount, profile_id),
            )
            inserted += 1

    conn.commit()
    conn.close()
    return inserted, f"Weekly payouts: {inserted} new rows imported, {len(ticker_rows)} tickers updated."


# ── Monthly payouts import ─────────────────────────────────────────────────────

def import_monthly_payouts(file_path, profile_id=1):
    """Import monthly payout data from Monthly Tracking sheet."""
    import openpyxl

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True, keep_vba=True)
    # Support alternate sheet names
    sheet_name = None
    for candidate in ["Monthly Tracking", "Monthly Income History"]:
        if candidate in wb.sheetnames:
            sheet_name = candidate
            break
    if sheet_name is None:
        wb.close()
        raise ValueError("No monthly tracking sheet found (expected 'Monthly Tracking' or 'Monthly Income History')")
    ws = wb[sheet_name]

    rows_to_insert = []

    # Detect format from first data row
    first_val = None
    for row in ws.iter_rows(min_row=2, max_row=2):
        first_val = row[0].value

    if isinstance(first_val, str) and '-' in first_val:
        # "Monthly Income History" format: "YYYY-MM" in col 0, amount in col 1
        for row in ws.iter_rows(min_row=2):
            month_str = row[0].value
            amount_val = row[1].value
            if not isinstance(month_str, str) or '-' not in month_str:
                continue
            if amount_val is None or isinstance(amount_val, str):
                continue
            try:
                parts = month_str.split('-')
                year = int(parts[0])
                month = int(parts[1])
                rows_to_insert.append((year, month, float(amount_val)))
            except (ValueError, IndexError):
                continue
    else:
        # "Monthly Tracking" format: year in col 0, Jan-Dec in cols 1-12
        for row in ws.iter_rows(min_row=2):
            year_val = row[0].value
            if year_val is None or not isinstance(year_val, (int, float)):
                continue
            year = int(year_val)
            for month_idx in range(1, 13):
                amount = row[month_idx].value
                if amount is not None:
                    rows_to_insert.append((year, month_idx, float(amount)))
    wb.close()

    conn = get_connection()
    ensure_tables_exist(conn)
    cur = conn.cursor()

    inserted = 0
    for year, month, amount in rows_to_insert:
        existing = cur.execute(
            "SELECT 1 FROM monthly_payouts WHERE year = ? AND month = ? AND profile_id = ?",
            (year, month, profile_id),
        ).fetchone()
        if existing is None:
            cur.execute(
                "INSERT INTO monthly_payouts (year, month, amount, profile_id) VALUES (?, ?, ?, ?)",
                (year, month, amount, profile_id),
            )
            inserted += 1

    conn.commit()
    conn.close()
    return inserted, f"Monthly payouts: {inserted} new rows imported."


# ── Monthly payout tickers import ─────────────────────────────────────────────

def import_monthly_payout_tickers(file_path, profile_id=1):
    """Import ticker-to-month mapping from DivMonths sheet."""
    import openpyxl

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True, keep_vba=True)
    ws = wb["DivMonths"]

    tickers = [cell.value for cell in ws[1] if cell.value is not None]
    ticker_months = {t: set() for t in tickers}
    for row in ws.iter_rows(min_row=2):
        for i, ticker in enumerate(tickers):
            val = row[i].value
            if val is not None and isinstance(val, (int, float)):
                ticker_months[ticker].add(int(val))
    wb.close()

    conn = get_connection()
    ensure_tables_exist(conn)
    cur = conn.cursor()

    cur.execute("DELETE FROM monthly_payout_tickers WHERE profile_id = ?", (profile_id,))
    inserted = 0
    for ticker, months in ticker_months.items():
        for month in sorted(months):
            cur.execute(
                "INSERT INTO monthly_payout_tickers (ticker, pay_month, profile_id) VALUES (?, ?, ?)",
                (ticker, month, profile_id),
            )
            inserted += 1

    conn.commit()
    conn.close()
    return inserted, f"Monthly ticker mappings: {inserted} rows imported ({len(tickers)} tickers)."


# ── Generic user upload ────────────────────────────────────────────────────────

UPLOAD_COL_MAP = {
    'ticker':                     ['ticker', 'symbol', 'stock', 'etf'],
    'description':                ['description', 'name', 'security name'],
    'classification_type':        ['type', 'classification', 'asset type'],
    'quantity':                    ['shares', 'quantity', 'qty', 'units', 'amount'],
    'price_paid':                  ['price paid', 'cost basis', 'avg cost', 'buy price', 'price_paid'],
    'current_price':               ['current price', 'market price', 'last price', 'current_price'],
    'purchase_value':              ['purchase value', 'cost value', 'purchase_value'],
    'current_value':               ['current value', 'market value', 'current_value'],
    'gain_or_loss':                ['gain/loss', 'gain loss', 'gain_or_loss', 'p&l'],
    'gain_or_loss_percentage':     ['gain/loss %', 'gain loss %', 'gain_or_loss_percentage', 'p&l %'],
    'div':                         ['div/share', 'dividend', 'div', 'distribution'],
    'div_frequency':               ['frequency', 'div frequency', 'div freq', 'div_frequency'],
    'ex_div_date':                 ['ex-div date', 'ex div date', 'ex_div_date', 'exdivdate'],
    'div_pay_date':                ['pay date', 'div pay date', 'payment date', 'div_pay_date'],
    'reinvest':                    ['drip', 'reinvest', 'reinvestment'],
    'dividend_paid':               ['div paid', 'dividend paid', 'dividend_paid'],
    'estim_payment_per_year':      ['est. annual pmt', 'annual payment', 'estim_payment_per_year', 'est annual'],
    'approx_monthly_income':       ['monthly income', 'monthly div', 'approx_monthly_income'],
    'annual_yield_on_cost':        ['yield on cost', 'yoc', 'annual_yield_on_cost'],
    'current_annual_yield':        ['current yield', 'yield', 'current_annual_yield'],
    'percent_of_account':          ['% of account', 'weight', 'allocation', 'percent_of_account'],
    'shares_bought_from_dividend': ['shares from div', 'div shares', 'shares_bought_from_dividend'],
    'shares_bought_in_year':       ['shares/year', 'shares year', 'shares_bought_in_year'],
    'shares_in_month':             ['shares/month', 'shares month', 'shares_in_month'],
    'ytd_divs':                    ['ytd divs', 'ytd dividends', 'ytd_divs'],
    'total_divs_received':         ['total divs received', 'total dividends', 'total_divs_received'],
    'paid_for_itself':             ['paid for itself', 'pfi', 'paid_for_itself'],
    'purchase_date':               ['date purchased', 'purchase date', 'buy date', 'purchase_date'],
    'cash_not_reinvested':         ['cash not reinvest', 'cash not reinvested', 'cash_not_reinvested'],
    'total_cash_reinvested':       ['cash reinvested', 'total cash reinvested', 'total_cash_reinvested'],
    'withdraw_8pct_cost_annually': ['8% annual wdraw', 'annual withdrawal', 'withdraw_8pct_cost_annually'],
    'withdraw_8pct_per_month':     ['8% monthly wdraw', 'monthly withdrawal', 'withdraw_8pct_per_month'],
    'percent_change':              ['% change', 'price change', 'percent_change'],
    'category':                    ['category', 'cat', 'group'],
}


def _normalize_upload_columns(df):
    """Rename upload DataFrame columns to canonical SQL names using UPLOAD_COL_MAP."""
    lower_map = {col.lower().strip(): col for col in df.columns}
    rename = {}
    for canonical, aliases in UPLOAD_COL_MAP.items():
        for alias in aliases:
            if alias in lower_map:
                rename[lower_map[alias]] = canonical
                break
    return df.rename(columns=rename)


def import_from_upload(df, profile_id):
    """
    Import a user-uploaded DataFrame into all_account_info for the given profile.
    df must have at minimum: ticker, quantity (or shares).
    Optional: price_paid, div, div_frequency, ex_div_date, reinvest.
    Returns (row_count, message).
    """
    import yfinance as yf
    from datetime import datetime as _dt, date as _date

    df = _normalize_upload_columns(df.copy())

    if 'ticker' not in df.columns:
        raise ValueError("Upload file must have a 'Ticker' or 'Symbol' column.")
    if 'quantity' not in df.columns:
        raise ValueError("Upload file must have a 'Shares' or 'Quantity' column.")

    # Clean tickers
    df['ticker'] = df['ticker'].astype(str).str.strip().str.upper()
    df = df[df['ticker'].apply(lambda t: bool(_VALID_TICKER.match(t)))]
    df['quantity'] = pd.to_numeric(df['quantity'], errors='coerce')
    df = df[df['quantity'].notna() & (df['quantity'] > 0)]

    if df.empty:
        raise ValueError("No valid ticker rows found in upload.")

    freq_map = {1: 'A', 2: 'SA', 4: 'Q', 12: 'M', 52: 'W'}
    tickers_list = df['ticker'].tolist()
    ticker_str = ' '.join(tickers_list)

    # Load known weekly tickers from weekly_payout_tickers table
    _weekly_set = set()
    try:
        _wconn = get_connection()
        _wrows = _wconn.execute("SELECT ticker FROM weekly_payout_tickers WHERE profile_id = ?", (profile_id,)).fetchall()
        _weekly_set = {r[0] if isinstance(r, tuple) else r['ticker'] for r in _wrows}
        _wconn.close()
    except Exception:
        pass

    def _freq_from_count(n):
        if n >= 45: return 'W'
        if n >= 10: return 'M'
        if n >= 3:  return 'Q'
        if n >= 2:  return 'SA'
        return 'A'

    # Batch download from yfinance
    price_map = {}
    div_map = {}
    exdiv_map = {}
    freq_hist = {}
    try:
        raw = yf.download(ticker_str, period='1y', progress=False, auto_adjust=False, actions=True)
        if not raw.empty:
            def _col(name):
                if isinstance(raw.columns, pd.MultiIndex):
                    return raw[name] if name in raw.columns.get_level_values(0) else None
                return raw[name] if name in raw.columns else None

            close = _col('Close')
            if close is not None:
                if isinstance(close, pd.Series):
                    s = close.dropna()
                    if len(s): price_map[tickers_list[0]] = float(s.iloc[-1])
                else:
                    for t in tickers_list:
                        if t in close.columns:
                            s = close[t].dropna()
                            if len(s): price_map[t] = float(s.iloc[-1])

            divs = _col('Dividends')
            if divs is not None:
                if isinstance(divs, pd.Series):
                    d = divs[divs > 0].dropna()
                    if not d.empty:
                        t0 = tickers_list[0]
                        div_map[t0] = float(d.iloc[-1])
                        exdiv_map[t0] = d.index[-1].strftime('%m/%d/%y')
                        freq_hist[t0] = _freq_from_count(len(d))
                else:
                    for t in tickers_list:
                        if t in divs.columns:
                            d = divs[t][divs[t] > 0].dropna()
                            if not d.empty:
                                div_map[t] = float(d.iloc[-1])
                                exdiv_map[t] = d.index[-1].strftime('%m/%d/%y')
                                freq_hist[t] = _freq_from_count(len(d))
    except Exception:
        pass

    # Per-ticker info for description / quoteType
    info_map = {}
    for t in tickers_list:
        try:
            info_map[t] = yf.Ticker(t).info or {}
        except Exception:
            info_map[t] = {}

    def _val(row, col):
        """Return a user-supplied value or None if missing/blank."""
        v = row.get(col, None)
        if v is None:
            return None
        s = str(v).strip()
        if s in ('', 'nan', 'None', 'NaN', 'none'):
            return None
        return v

    def _fval(row, col):
        """Return a float from user-supplied column, or None."""
        v = _val(row, col)
        if v is None:
            return None
        try:
            return float(v)
        except (ValueError, TypeError):
            return None

    # Extract category assignments before building enriched rows
    _category_assignments = {}  # ticker -> category name
    for _, row in df.iterrows():
        t = row['ticker']
        cat_val = _val(row, 'category') if 'category' in df.columns else None
        if cat_val:
            _category_assignments[t] = str(cat_val).strip()

    enriched = []
    for _, row in df.iterrows():
        t = row['ticker']
        info = info_map.get(t, {})

        # ── Core pricing: user-supplied → yfinance → 0 ──────────────────
        current_price = (
            _fval(row, 'current_price') or
            price_map.get(t) or
            info.get('regularMarketPrice') or
            info.get('currentPrice') or 0.0
        )

        div = (
            _fval(row, 'div') or
            div_map.get(t) or
            info.get('dividendRate') or 0.0
        )

        # Ex-div date
        ex_div_raw = _val(row, 'ex_div_date')
        if ex_div_raw:
            ex_div_date = str(ex_div_raw).strip()
        else:
            ex_div_date = exdiv_map.get(t)
            if not ex_div_date:
                ex_ts = info.get('exDividendDate')
                if ex_ts:
                    try:
                        ex_div_date = _dt.utcfromtimestamp(ex_ts).strftime('%m/%d/%y')
                    except Exception:
                        ex_div_date = None

        # Frequency: user-supplied > weekly_payout_tickers table > yfinance history > yfinance info > default Q
        freq_raw = _val(row, 'div_frequency')
        if freq_raw and str(freq_raw).strip().upper() in ('A', 'SA', 'Q', 'M', 'W', '52'):
            div_frequency = str(freq_raw).strip().upper()
            if div_frequency == '52':
                div_frequency = 'W'
        elif t in _weekly_set:
            div_frequency = 'W'
        elif t in freq_hist:
            div_frequency = freq_hist[t]
        else:
            div_frequency = freq_map.get(info.get('payoutFrequency'), 'Q')

        qty = float(row['quantity'])
        price_paid = _fval(row, 'price_paid') or current_price

        # ── Computed values: use user-supplied if present, else compute ──
        purchase_value = _fval(row, 'purchase_value') or (price_paid * qty)
        current_value = _fval(row, 'current_value') or (current_price * qty)
        gain_or_loss = _fval(row, 'gain_or_loss')
        if gain_or_loss is None:
            gain_or_loss = current_value - purchase_value
        gain_or_loss_pct = _fval(row, 'gain_or_loss_percentage')
        if gain_or_loss_pct is None:
            gain_or_loss_pct = (gain_or_loss / purchase_value) if purchase_value else 0
        percent_change = _fval(row, 'percent_change')
        if percent_change is None:
            percent_change = gain_or_loss_pct

        freq_mult = {'W': 52, '52': 52, 'M': 12, 'Q': 4, 'SA': 2, 'A': 1}
        mult = freq_mult.get(div_frequency.upper(), 4) if div_frequency else 4
        estim = _fval(row, 'estim_payment_per_year') or (div * qty * mult)
        monthly_income = _fval(row, 'approx_monthly_income') or (estim / 12 if estim else 0)

        reinvest = str(_val(row, 'reinvest') or 'N').strip().upper()
        if reinvest not in ('Y', 'N'):
            reinvest = 'N'

        description = _val(row, 'description')
        if not description:
            description = info.get('longName', t)[:200] if info.get('longName') else t

        classification_type = _val(row, 'classification_type')
        if not classification_type:
            classification_type = info.get('quoteType', 'ETF')[:20]

        # Purchase date
        purchase_date = _val(row, 'purchase_date')
        if purchase_date:
            try:
                purchase_date = pd.to_datetime(purchase_date).strftime('%Y-%m-%d')
            except Exception:
                purchase_date = None

        enriched.append({
            'ticker':                     t,
            'description':                description,
            'classification_type':        str(classification_type)[:20],
            'price_paid':                 price_paid,
            'current_price':              current_price,
            'percent_change':             percent_change,
            'quantity':                    qty,
            'purchase_value':             purchase_value,
            'current_value':              current_value,
            'gain_or_loss':               gain_or_loss,
            'gain_or_loss_percentage':    gain_or_loss_pct,
            'div_frequency':              div_frequency,
            'reinvest':                   reinvest,
            'ex_div_date':                ex_div_date,
            'div_pay_date':               _val(row, 'div_pay_date') or ((_dt.strptime(ex_div_date, "%m/%d/%y") + pd.Timedelta(days=21)).strftime("%m/%d/%y") if ex_div_date else None),
            'div':                        div,
            'dividend_paid':              _fval(row, 'dividend_paid'),
            'estim_payment_per_year':     estim,
            'approx_monthly_income':      monthly_income,
            'withdraw_8pct_cost_annually': _fval(row, 'withdraw_8pct_cost_annually'),
            'withdraw_8pct_per_month':    _fval(row, 'withdraw_8pct_per_month'),
            'cash_not_reinvested':        _fval(row, 'cash_not_reinvested'),
            'total_cash_reinvested':      _fval(row, 'total_cash_reinvested'),
            'annual_yield_on_cost':       _fval(row, 'annual_yield_on_cost') or ((div * mult / price_paid) if price_paid else 0),
            'current_annual_yield':       _fval(row, 'current_annual_yield') or ((div * mult / current_price) if current_price else 0),
            'percent_of_account':         _fval(row, 'percent_of_account'),
            'shares_bought_from_dividend': _fval(row, 'shares_bought_from_dividend') or ((estim / current_price) if reinvest == 'Y' and estim and current_price else None),
            'shares_bought_in_year':      _fval(row, 'shares_bought_in_year'),
            'shares_in_month':            _fval(row, 'shares_in_month'),
            'ytd_divs':                   _fval(row, 'ytd_divs'),
            'total_divs_received':        _fval(row, 'total_divs_received'),
            'paid_for_itself':            _fval(row, 'paid_for_itself'),
            'purchase_date':              purchase_date,
            'import_date':                _date.today().isoformat(),
            'profile_id':                 profile_id,
        })

    out = pd.DataFrame(enriched)
    insert_cols = list(out.columns)
    placeholders = ', '.join(['?'] * len(insert_cols))
    insert_sql = f"INSERT INTO all_account_info ({', '.join(insert_cols)}) VALUES ({placeholders})"

    conn = get_connection()
    ensure_tables_exist(conn)
    cur = conn.cursor()

    # Check if data already exists for this profile (merge mode)
    existing = cur.execute(
        "SELECT ticker FROM all_account_info WHERE profile_id = ?", (profile_id,)
    ).fetchall()
    existing_tickers = {r[0] for r in existing}

    if not existing_tickers:
        # First import — full load
        row_count = 0
        for _, row in out.iterrows():
            values = [None if pd.isna(v) else v for v in row.tolist()]
            cur.execute(insert_sql, values)
            row_count += 1
        conn.commit()
        msg = f"Imported {row_count} holdings for profile {profile_id}."
    else:
        # Merge mode — update existing, insert new
        merge_fields = [
            'quantity', 'price_paid', 'purchase_value', 'description',
            'classification_type', 'div_frequency', 'reinvest', 'ex_div_date',
            'div_pay_date', 'div', 'dividend_paid', 'estim_payment_per_year',
            'approx_monthly_income', 'ytd_divs', 'total_divs_received',
            'paid_for_itself', 'cash_not_reinvested', 'total_cash_reinvested',
            'shares_bought_from_dividend', 'purchase_date',
            'annual_yield_on_cost', 'current_annual_yield',
            'current_price', 'current_value', 'gain_or_loss',
            'gain_or_loss_percentage', 'percent_change',
        ]
        inserted = 0
        updated = 0
        for _, row in out.iterrows():
            ticker = row.get('ticker')
            if ticker in existing_tickers:
                sets = []
                vals = []
                _numeric_merge = {
                    'quantity', 'price_paid', 'purchase_value', 'div',
                    'dividend_paid', 'estim_payment_per_year', 'approx_monthly_income',
                    'ytd_divs', 'total_divs_received', 'paid_for_itself',
                    'cash_not_reinvested', 'total_cash_reinvested',
                    'shares_bought_from_dividend', 'annual_yield_on_cost',
                    'current_annual_yield', 'current_price', 'current_value',
                    'gain_or_loss', 'gain_or_loss_percentage', 'percent_change',
                }
                for field in merge_fields:
                    if field in row.index:
                        val = row[field]
                        if pd.notna(val):
                            sets.append(f"{field} = ?")
                            vals.append(val)
                        elif field in _numeric_merge:
                            sets.append(f"{field} = ?")
                            vals.append(0)
                if sets:
                    sets.append("import_date = ?")
                    vals.append(_date.today().isoformat())
                    vals.extend([ticker, profile_id])
                    cur.execute(
                        f"UPDATE all_account_info SET {', '.join(sets)} WHERE ticker = ? AND profile_id = ?",
                        vals,
                    )
                    updated += 1
            else:
                values = [None if pd.isna(v) else v for v in row.tolist()]
                cur.execute(insert_sql, values)
                inserted += 1

        conn.commit()
        row_count = updated + inserted
        parts = []
        if updated:
            parts.append(f"updated {updated} existing")
        if inserted:
            parts.append(f"added {inserted} new")
        msg = f"Merge complete: {', '.join(parts)} holdings."

    # ── Process category assignments from spreadsheet ─────────────────────
    if _category_assignments:
        for ticker, cat_name in _category_assignments.items():
            # Find or create the category
            existing_cat = cur.execute(
                "SELECT id FROM categories WHERE name = ? AND profile_id = ?",
                (cat_name, profile_id),
            ).fetchone()
            if existing_cat:
                cat_id = existing_cat[0]
            else:
                max_pos = cur.execute(
                    "SELECT COALESCE(MAX(sort_order), 0) FROM categories WHERE profile_id = ?",
                    (profile_id,),
                ).fetchone()[0]
                cur.execute(
                    "INSERT INTO categories (name, target_pct, sort_order, profile_id) VALUES (?, 0, ?, ?)",
                    (cat_name, max_pos + 1, profile_id),
                )
                cat_id = cur.lastrowid

            # Assign ticker if not already assigned to this category
            already = cur.execute(
                "SELECT 1 FROM ticker_categories WHERE ticker = ? AND category_id = ? AND profile_id = ?",
                (ticker, cat_id, profile_id),
            ).fetchone()
            if not already:
                cur.execute(
                    "INSERT INTO ticker_categories (ticker, category_id, profile_id) VALUES (?, ?, ?)",
                    (ticker, cat_id, profile_id),
                )
        conn.commit()

    conn.close()
    return row_count, msg


# ── Multi-sheet import helpers ────────────────────────────────────────────────

def _get_or_create_profile(name, conn):
    """Find a profile by name or create it. Returns profile_id."""
    row = conn.execute("SELECT id FROM profiles WHERE name = ?", (name,)).fetchone()
    if row:
        return row[0] if isinstance(row, tuple) else row["id"]
    cur = conn.execute("INSERT INTO profiles (name) VALUES (?)", (name,))
    conn.commit()
    return cur.lastrowid


def import_multi_excel(file_path, default_profile_id=1):
    """Import multiple sheets from the owner's Excel format.

    Each sheet (except special ones) is treated as a separate portfolio.
    Returns list of {profile_id, profile_name, rows, message}.
    """
    import openpyxl
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    wb.close()

    skip_sheets = {"Weekly_Payers", "Monthly Tracking", "DivMonths", "Instructions"}
    results = []

    for sheet in sheet_names:
        if sheet in skip_sheets:
            continue
        # "All Accounts" maps to the default profile
        if sheet == "All Accounts":
            pid = default_profile_id
        else:
            conn = get_connection()
            pid = _get_or_create_profile(sheet, conn)
            conn.close()

        try:
            count, msg = import_from_excel(file_path, sheet_name=sheet, profile_id=pid)
            results.append({"profile_id": pid, "profile_name": sheet, "rows": count, "message": msg})
        except Exception as e:
            results.append({"profile_id": pid, "profile_name": sheet, "rows": 0, "message": f"Error: {str(e)}"})

    return results


def import_multi_upload(file_path):
    """Import a multi-tab generic upload file.

    Each non-empty sheet becomes a separate portfolio (auto-created from sheet name).
    Returns list of {profile_id, profile_name, rows, message}.
    """
    import openpyxl
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    sheet_names = [s for s in wb.sheetnames if s != "Instructions"]
    wb.close()

    conn = get_connection()
    results = []

    for sheet in sheet_names:
        try:
            df = pd.read_excel(file_path, sheet_name=sheet, engine="openpyxl")
        except Exception:
            continue

        # Skip empty sheets
        if df.empty or len(df) == 0:
            continue
        # Check if it has any data rows (not just headers)
        if df.dropna(how="all").empty:
            continue

        pid = _get_or_create_profile(sheet, conn)

        try:
            count, msg = import_from_upload(df, pid)
            results.append({"profile_id": pid, "profile_name": sheet, "rows": count, "message": msg})
        except Exception as e:
            results.append({"profile_id": pid, "profile_name": sheet, "rows": 0, "message": f"Error: {str(e)}"})

    conn.close()
    return results
