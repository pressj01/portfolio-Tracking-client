"""Generate downloadable import templates."""
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), '..', 'templates', 'portfolio_upload_template.xlsx')
ETRADE_TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), '..', 'templates', 'etrade_positions_template.csv')
SCHWAB_TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), '..', 'templates', 'schwab_positions_template.csv')
SCHWAB_TRANSACTIONS_TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), '..', 'templates', 'schwab_transactions_template.csv')
ETRADE_BUYS_SELLS_TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), '..', 'templates', 'etrade_buys_sells_template.xlsx')
ETRADE_DIVIDENDS_TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), '..', 'templates', 'etrade_dividends_template.xlsx')
SNOWBALL_HOLDINGS_TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), '..', 'templates', 'snowball_holdings_template.csv')
FIDELITY_TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), '..', 'templates', 'fidelity_positions_template.xlsx')
FIDELITY_TRANSACTIONS_TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), '..', 'templates', 'fidelity_transactions_template.xlsx')


def _write_csv_template(path, rows):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        writer.writerows(rows)
    return path


def _create_etrade_transaction_template(path, title, total_label, sample_rows, guidance_lines):
    os.makedirs(os.path.dirname(path), exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transactions"

    ws["A1"] = title
    ws["A3"] = "Account Activity for <E*TRADE Account Name> from 03/01/2026 to 04/13/2026"
    ws["A5"] = total_label

    headers = [
        "Run Date",
        "Settlement Date",
        "Activity/Trade Date",
        "Activity Type",
        "Symbol",
        "Description",
        "Quantity #",
        "Price $",
        "Amount $",
        "Commission",
    ]

    header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(sample_rows, 8):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    widths = {
        "A": 14, "B": 16, "C": 18, "D": 22, "E": 12,
        "F": 34, "G": 14, "H": 12, "I": 14, "J": 12,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ins = wb.create_sheet("Instructions")
    ins["A1"] = "Field"
    ins["B1"] = "How the importer uses it"
    ins["C1"] = "Notes"
    for cell in ins[1]:
        cell.fill = header_fill
        cell.font = header_font

    instruction_rows = [
        ("Sheet layout", "The importer reads the first sheet only.", "Keep the title row, account row, total row, and row 7 headers in place."),
        ("Activity Type", "Required", "Must stay in column D because the parser uses it to detect valid rows."),
        ("Symbol", "Required", "Ticker symbol. Invalid or blank tickers are skipped."),
        ("Activity/Trade Date", "Required", "Accepts MM/DD/YY, MM/DD/YYYY, or YYYY-MM-DD."),
        ("Quantity #", "Used for buys, sells, and DRIP reinvestments", "Leave blank for cash-only dividend rows."),
        ("Price $", "Used for buys, sells, and DRIP reinvestments", "Leave blank for cash-only dividend rows."),
        ("Amount $", "Used for dividend cash amounts and for matching export totals", "Negative amounts usually indicate cash outflow / reinvestment."),
        ("Commission", "Imported as fees for buys and sells", "Leave 0.00 if no commission was charged."),
    ]
    for row_idx, row in enumerate(instruction_rows, 2):
        for col_idx, value in enumerate(row, 1):
            ins.cell(row=row_idx, column=col_idx, value=value)

    extra_start = len(instruction_rows) + 3
    ins.cell(row=extra_start, column=1, value="Template guidance")
    ins.cell(row=extra_start, column=1).font = Font(bold=True)
    for offset, line in enumerate(guidance_lines, 1):
        ins.cell(row=extra_start + offset, column=1, value=f"- {line}")

    ins.column_dimensions["A"].width = 24
    ins.column_dimensions["B"].width = 34
    ins.column_dimensions["C"].width = 74

    wb.save(path)
    return path


def create_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Portfolio"

    # All columns the app supports — required columns first, then optional
    headers = [
        # Required
        "Ticker", "Shares",
        # Core pricing
        "Price Paid", "Current Price",
        # Holdings info
        "Description", "Type", "Date Purchased",
        # Gain/loss
        "Purchase Value", "Current Value", "Gain/Loss", "Gain/Loss %", "% Change",
        # Dividend info
        "Div/Share", "Frequency", "Ex-Div Date", "Pay Date", "DRIP",
        "Div Paid", "Est. Annual Pmt", "Monthly Income",
        # Yield
        "Yield On Cost", "Current Yield", "% of Account",
        # Dividend tracking
        "YTD Divs", "Total Divs Received", "Paid For Itself",
        # Reinvestment
        "Cash Not Reinvest", "Cash Reinvested",
        "Shares from Div", "Shares/Year", "Shares/Month",
        # Withdrawal
        "8% Annual Wdraw", "8% Monthly Wdraw",
        # Category
        "Category",
    ]

    header_font = Font(bold=True, color="FFFFFF", size=11)
    required_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
    optional_fill = PatternFill(start_color="37474F", end_color="37474F", fill_type="solid")
    thin_border = Border(bottom=Side(style="thin", color="90CAF9"))

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = required_fill if col <= 2 else optional_fill
        cell.alignment = Alignment(horizontal="center")

    # Example rows (only filling key columns, leaving optional ones blank)
    examples = [
        {"Ticker": "JEPI", "Shares": 100, "Price Paid": 55.50, "Current Price": 57.20,
         "Description": "JPMorgan Equity Premium Income ETF", "Type": "ETF",
         "Div/Share": 0.45, "Frequency": "M", "Ex-Div Date": "03/01/25", "DRIP": "Y",
         "Date Purchased": "2024-01-15", "YTD Divs": 135.00, "Total Divs Received": 540.00,
         "Category": "Anchors"},
        {"Ticker": "SCHD", "Shares": 50, "Price Paid": 78.20, "Current Price": 82.50,
         "Description": "Schwab US Dividend Equity ETF", "Type": "ETF",
         "Div/Share": 0.62, "Frequency": "Q", "Ex-Div Date": "03/15/25", "DRIP": "N",
         "Date Purchased": "2023-06-01", "YTD Divs": 31.00, "Total Divs Received": 186.00,
         "Category": "Growth"},
        {"Ticker": "O", "Shares": 25, "Price Paid": 52.00, "Current Price": 55.80,
         "Description": "Realty Income Corp", "Type": "REIT",
         "Div/Share": 0.26, "Frequency": "M", "Ex-Div Date": "02/28/25", "DRIP": "Y",
         "Date Purchased": "2024-03-10", "YTD Divs": 19.50, "Total Divs Received": 78.00,
         "Category": "Boosters"},
    ]
    for row_idx, example in enumerate(examples, 2):
        for col_idx, header in enumerate(headers, 1):
            val = example.get(header)
            if val is not None:
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = thin_border

    # Column widths
    for i, header in enumerate(headers, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = max(len(header) + 4, 12)

    # ── Additional portfolio sheets (2-12) ─────────────────────────────────────
    for sheet_num in range(2, 13):
        ps = wb.create_sheet(f"Portfolio {sheet_num}")
        for col, header in enumerate(headers, 1):
            cell = ps.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = required_fill if col <= 2 else optional_fill
            cell.alignment = Alignment(horizontal="center")
        for i, header in enumerate(headers, 1):
            ps.column_dimensions[openpyxl.utils.get_column_letter(i)].width = max(len(header) + 4, 12)

    # ── Instructions sheet ─────────────────────────────────────────────────────
    ins = wb.create_sheet("Instructions")
    instructions = [
        ("Column", "Required", "Description"),
        ("Ticker", "YES", "Stock/ETF ticker symbol (e.g., JEPI, SCHD, O)"),
        ("Shares", "YES", "Number of shares owned"),
        ("Price Paid", "No", "Average cost basis per share (defaults to current market price if blank)"),
        ("Current Price", "No", "Current market price per share (auto-fetched from Yahoo Finance if blank)"),
        ("Description", "No", "Security name/description (auto-fetched if blank)"),
        ("Type", "No", "Asset type: ETF, EQUITY, CEF, BDC, REIT (auto-detected if blank)"),
        ("Date Purchased", "No", "Date the position was purchased (YYYY-MM-DD format)"),
        ("Purchase Value", "No", "Total cost basis (auto-calculated as Price Paid x Shares if blank)"),
        ("Current Value", "No", "Total current value (auto-calculated as Current Price x Shares if blank)"),
        ("Gain/Loss", "No", "Dollar gain or loss (auto-calculated if blank)"),
        ("Gain/Loss %", "No", "Percentage gain or loss (auto-calculated if blank)"),
        ("% Change", "No", "Price change percentage (auto-calculated if blank)"),
        ("Div/Share", "No", "Dividend per share amount (auto-fetched from Yahoo Finance if blank)"),
        ("Frequency", "No", "Dividend frequency: W=Weekly, M=Monthly, Q=Quarterly, SA=Semi-Annual, A=Annual"),
        ("Ex-Div Date", "No", "Last ex-dividend date (auto-fetched if blank)"),
        ("Pay Date", "No", "Dividend payment date (auto-estimated as ex-div + 3 weeks if blank)"),
        ("DRIP", "No", "Dividend reinvestment: Y=Yes, N=No (defaults to N)"),
        ("Div Paid", "No", "Cash amount of one dividend payment for your current share count"),
        ("Est. Annual Pmt", "No", "Estimated annual dividend payment (auto-calculated as Div/Share x Shares if blank)"),
        ("Monthly Income", "No", "Approximate monthly dividend income (auto-calculated if blank)"),
        ("Yield On Cost", "No", "Annual yield based on price paid (auto-calculated if blank)"),
        ("Current Yield", "No", "Annual yield based on current price (auto-calculated if blank)"),
        ("% of Account", "No", "Position weight as percentage of total portfolio"),
        ("YTD Divs", "No", "Year-to-date dividends received"),
        ("Total Divs Received", "No", "Total lifetime dividends received from this position"),
        ("Paid For Itself", "No", "Ratio of total dividends received to purchase value"),
        ("Cash Not Reinvest", "No", "Cash dividends not reinvested"),
        ("Cash Reinvested", "No", "Total cash dividends that were reinvested"),
        ("Shares from Div", "No", "Number of shares acquired through DRIP"),
        ("Shares/Year", "No", "Shares bought per year from dividends"),
        ("Shares/Month", "No", "Shares bought per month from dividends"),
        ("8% Annual Wdraw", "No", "Annual withdrawal amount at 8% rate"),
        ("8% Monthly Wdraw", "No", "Monthly withdrawal amount at 8% rate"),
        ("Category", "No", "Category name to assign this ticker to (e.g., Anchors, Boosters, Growth). Creates category if it doesn't exist."),
    ]
    for row_idx, (a, b, c) in enumerate(instructions, 1):
        cell_a = ins.cell(row=row_idx, column=1, value=a)
        cell_b = ins.cell(row=row_idx, column=2, value=b)
        cell_c = ins.cell(row=row_idx, column=3, value=c)
        if row_idx == 1:
            for cell in (cell_a, cell_b, cell_c):
                cell.font = header_font
                cell.fill = required_fill
        elif b == "YES":
            cell_b.font = Font(bold=True, color="FF6B6B")

    ins.column_dimensions['A'].width = 22
    ins.column_dimensions['B'].width = 10
    ins.column_dimensions['C'].width = 80

    os.makedirs(os.path.dirname(TEMPLATE_PATH), exist_ok=True)
    wb.save(TEMPLATE_PATH)
    return TEMPLATE_PATH


def create_etrade_template():
    """Create a CSV template that matches the exact E*TRADE positions import shape."""
    os.makedirs(os.path.dirname(ETRADE_TEMPLATE_PATH), exist_ok=True)

    rows = [
        ["Account Summary"],
        ["Account", "Net Account Value", "Day's Gain", "Day's Gain %", "Market Value", "Total Gain", "Total Gain %", "Cash Purchasing Power"],
        ["<E*TRADE Account Name>", "37796.82", "0.00", "0.00%", "36963.37", "-3000.98", "-7.51%", "833.45"],
        [],
        ["View Summary - Dividends"],
        [
            "Symbol",
            "Price Paid $",
            "Last Price $",
            "Change %",
            "Day's Gain $",
            "Qty #",
            "Total Gain $",
            "Total Gain %",
            "Value $",
            "Ex-Div Date",
            "Dividend Pay Date",
            "Dividend Yield %",
            "Dividend",
            "Annual Dividend",
            "Total Cost",
        ],
        [
            "<SYMBOL>",
            "55.5000",
            "57.2000",
            "0.45%",
            "25.00",
            "100.0000",
            "170.00",
            "3.06%",
            "5720.00",
            "04/01/2026",
            "04/05/2026",
            "7.50%",
            "0.55",
            "6.60",
            "5550.00",
        ],
        ["CASH", "", "", "", "", "", "", "", "833.45", "", "", "", "", "", ""],
        ["TOTAL", "", "", "", "", "", "", "", "37796.82", "", "", "", "", "", ""],
    ]

    with open(ETRADE_TEMPLATE_PATH, "w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        writer.writerows(rows)

    return ETRADE_TEMPLATE_PATH


def create_schwab_template():
    """Create a CSV template that matches the exact Schwab positions import shape."""
    os.makedirs(os.path.dirname(SCHWAB_TEMPLATE_PATH), exist_ok=True)

    rows = [
        ['Positions for account <Charles Schwab Account Name>'],
        [],
        [
            "Symbol",
            "Description",
            "Qty (Quantity)",
            "Price",
            "Price Change $",
            "Price Change %",
            "Mkt Val (Market Value)",
            "Day Change $",
            "Day Change %",
            "Cost Basis",
            "Gain $ (Gain/Loss $)",
            "Gain % (Gain/Loss %)",
            "Reinvest?",
            "Capital Gains?",
            "% Of Account",
            "Div Yld (Dividend Yield)",
            "Last Dividend",
            "Ex-Div Date",
            "P/E Ratio",
            "52 Week Low",
            "52 Week High",
            "Volume",
            "Intrinsic Value",
            "In The Money",
            "Security Type",
            "Asset Type",
            "Cost/Share",
        ],
        [
            "SCHD",
            "Schwab U.S. Dividend Equity ETF",
            "50",
            "$82.50",
            "$0.35",
            "0.43%",
            "$4,125.00",
            "$17.50",
            "0.43%",
            "$3,910.00",
            "$215.00",
            "5.50%",
            "No",
            "No",
            "10.92%",
            "3.01%",
            "$0.62",
            "03/27/2026",
            "",
            "$68.53",
            "$87.00",
            "1234567",
            "",
            "",
            "ETF",
            "ETF",
            "$78.20",
        ],
        [
            "JEPI",
            "JPMorgan Equity Premium Income ETF",
            "100",
            "$57.20",
            "$0.18",
            "0.32%",
            "$5,720.00",
            "$18.00",
            "0.32%",
            "$5,550.00",
            "$170.00",
            "3.06%",
            "Yes",
            "No",
            "15.15%",
            "7.50%",
            "$0.55",
            "04/01/2026",
            "",
            "$49.00",
            "$60.88",
            "2345678",
            "",
            "",
            "ETF",
            "ETF",
            "$55.50",
        ],
        [
            "Cash & Cash Investments",
            "",
            "",
            "",
            "",
            "",
            "$833.45",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "Cash",
            "Cash",
            "",
        ],
        [
            "Positions Total",
            "",
            "",
            "",
            "",
            "",
            "$10,678.45",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ],
    ]

    return _write_csv_template(SCHWAB_TEMPLATE_PATH, rows)


def create_schwab_transactions_template():
    """Create a CSV template that matches the Schwab transactions import shape."""
    rows = [
        ["Date", "Action", "Symbol", "Description", "Quantity", "Price", "Fees & Comm", "Amount"],
        ["03/06/2026", "Buy", "SCHD", "Schwab U.S. Dividend Equity ETF", "25", "$82.15", "$0.00", "($2,053.75)"],
        ["03/20/2026", "Cash Dividend", "SCHD", "Schwab U.S. Dividend Equity ETF", "", "", "$0.00", "$31.00"],
        ["03/20/2026", "Reinvest Shares", "JEPI", "JPMorgan Equity Premium Income ETF", "0.9621", "$57.20", "$0.00", "($55.05)"],
        ["03/27/2026", "Sell", "DIVO", "Amplify CWP Enhanced Dividend Income ETF", "10", "$41.25", "$0.00", "$412.50"],
        ["03/31/2026", "Reinvestment Adj", "JEPI", "JPMorgan Equity Premium Income ETF", "-0.0100", "$57.20", "$0.00", "$0.00"],
    ]
    return _write_csv_template(SCHWAB_TRANSACTIONS_TEMPLATE_PATH, rows)


def create_fidelity_template():
    """Create an XLSX template that matches the Fidelity positions import shape."""
    os.makedirs(os.path.dirname(FIDELITY_TEMPLATE_PATH), exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Position"

    headers = [
        "Account Number", "Account Name", "Symbol", "Description", "Last Price",
        " Change $", "Today's gain/loss $", " Today's gain/loss %", "Total gain/loss $",
        " Total gain/loss %", "Current value", "Cost basis total", " Average cost basis",
        "% of account", "Quantity", "Ex-date", "Amount per share", "Pay date", "Dist. yield",
        " Distribution yield as of", "SEC yield", " SEC yield as of", "Est. annual income",
        "Morningstar overall rating", "52-week high date",
    ]
    sample_rows = [
        ["XXXXXXXXX", "<Fidelity Account Name>", "SPAXX**", "HELD IN MONEY MARKET", "", "", "", "", "", "", 386.87, "", "", 0.0031, "", "", "", "", "--", "--", 0.0326, "Apr-14-2026", "--", "", ""],
        ["XXXXXXXXX", "<Fidelity Account Name>", "AAPL", "APPLE INC", 258.83, -0.37, -10.69, -0.0015, 6796.28, 10.0247, 7474.23, 677.95, 23.48, 0.0608, 28.877, "Feb-09-2026", 0.26, "Feb-12-2026", 0.004, "Apr-13-2026", "--", "--", 30.03, "", "Dec-03-2025"],
        ["XXXXXXXXX", "<Fidelity Account Name>", "AVGO", "BROADCOM INC COM", 380.78, 1.03, 11.87, 0.0027, 457.02, 0.1162, 4388.48, 3931.46, 341.12, 0.0357, 11.525, "Mar-23-2026", 0.65, "Mar-31-2026", 0.0068, "Apr-13-2026", "--", "--", 29.96, "", "Dec-10-2025"],
    ]

    header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(sample_rows, 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    widths = {
        "A": 16, "B": 18, "C": 12, "D": 42, "E": 12, "F": 11, "G": 17, "H": 18, "I": 17,
        "J": 17, "K": 14, "L": 15, "M": 17, "N": 12, "O": 12, "P": 14, "Q": 16, "R": 14,
        "S": 12, "T": 18, "U": 12, "V": 16, "W": 18, "X": 18, "Y": 16,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ins = wb.create_sheet("Instructions")
    ins["A1"] = "Field"
    ins["B1"] = "How the importer uses it"
    ins["C1"] = "Notes"
    for cell in ins[1]:
        cell.fill = header_fill
        cell.font = header_font

    instruction_rows = [
        ("Account Name", "Used for broker-specific portfolio validation", "Fidelity exports often show a generic account type such as IRA."),
        ("Symbol", "Required", "Ticker symbol. Cash rows like SPAXX** are filtered out."),
        ("Description", "Imported", "Security description shown in holdings."),
        ("Quantity", "Required for holdings rows", "Rows without a positive quantity are treated as cash/filtered."),
        ("Last Price", "Imported", "Current price per share."),
        ("Current value", "Required", "Current market value for the holding."),
        ("Cost basis total", "Imported", "Used as purchase value / cost basis."),
        ("Average cost basis", "Imported", "Used as cost per share when available."),
        ("Amount per share", "Imported when present", "Stored as the current dividend per share."),
        ("Dist. yield", "Imported when present", "Fidelity exports it as a fraction, which the importer converts to a percent."),
        ("Ex-date / Pay date / Est. annual income", "Imported when present", "Optional dividend metadata used only if the workbook includes it."),
    ]
    for row_idx, row in enumerate(instruction_rows, 2):
        for col_idx, value in enumerate(row, 1):
            ins.cell(row=row_idx, column=col_idx, value=value)

    ins.column_dimensions["A"].width = 24
    ins.column_dimensions["B"].width = 34
    ins.column_dimensions["C"].width = 84

    wb.save(FIDELITY_TEMPLATE_PATH)
    return FIDELITY_TEMPLATE_PATH


def create_fidelity_transactions_template():
    """Create an XLSX template that matches the Fidelity transactions import shape."""
    os.makedirs(os.path.dirname(FIDELITY_TRANSACTIONS_TEMPLATE_PATH), exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transactions"

    headers = [
        "Run Date", "Action", "Symbol", "Description", "Type", "Price ($)", "Quantity",
        "Commission ($)", "Fees ($)", "Accrued Interest ($)", "Amount ($)",
        "Cash Balance ($)", "Settlement Date",
    ]
    sample_rows = [
        ["", "", "", "", "", "", "", "", "", "", "", "", ""],
        ["", "", "", "", "", "", "", "", "", "", "", "", ""],
        headers,
        ["04/13/2026", "YOU BOUGHT APPLE INC (AAPL) (Cash)", "AAPL", "APPLE INC", "Cash", 180.25, 5, 0.00, 0.00, "", -901.25, 1250.00, "04/14/2026"],
        ["04/10/2026", "DIVIDEND RECEIVED as of Apr-09-2026 APPLE INC (AAPL) (Cash)", "AAPL", "APPLE INC", "Cash", "", 0, 0.00, 0.00, "", 15.40, 2151.25, ""],
        ["04/10/2026", "REINVESTMENT as of Apr-09-2026 JPMORGAN EQUITY PREMIUM INCOME ETF (JEPI) (Cash)", "JEPI", "JPMORGAN EQUITY PREMIUM INCOME ETF", "Cash", 57.20, 0.2692, 0.00, 0.00, "", -15.40, 2166.65, ""],
    ]

    header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for row_idx, row in enumerate(sample_rows, 1):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 3:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

    widths = {
        "A": 14, "B": 58, "C": 12, "D": 42, "E": 12, "F": 12, "G": 12,
        "H": 15, "I": 12, "J": 18, "K": 14, "L": 16, "M": 16,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ins = wb.create_sheet("Instructions")
    ins["A1"] = "Field"
    ins["B1"] = "How the importer uses it"
    ins["C1"] = "Notes"
    for cell in ins[1]:
        cell.fill = header_fill
        cell.font = header_font

    instruction_rows = [
        ("Worksheet layout", "Required", "Keep the first two blank rows and the header row on row 3."),
        ("Action", "Required", "The importer reads YOU BOUGHT, YOU SOLD, DIVIDEND RECEIVED, and REINVESTMENT rows."),
        ("Symbol", "Required", "Invalid or blank symbols are skipped."),
        ("Run Date", "Required", "Imported as the transaction date."),
        ("Quantity / Price ($)", "Required for buys, sells, and DRIP", "Dividend cash rows can leave Price blank and Quantity at 0."),
        ("Commission ($) + Fees ($)", "Imported", "Combined into transaction fees."),
        ("Amount ($)", "Imported for dividends", "Positive amounts become cash DIVIDEND history entries."),
    ]
    for row_idx, row in enumerate(instruction_rows, 2):
        for col_idx, value in enumerate(row, 1):
            ins.cell(row=row_idx, column=col_idx, value=value)

    ins.column_dimensions["A"].width = 24
    ins.column_dimensions["B"].width = 34
    ins.column_dimensions["C"].width = 84

    wb.save(FIDELITY_TRANSACTIONS_TEMPLATE_PATH)
    return FIDELITY_TRANSACTIONS_TEMPLATE_PATH


def create_snowball_holdings_template():
    """Create a CSV template matching the Snowball holdings migration parser."""
    rows = [
        [
            "Holding",
            "Holdings' name",
            "Shares",
            "Cost basis",
            "Current value",
            "Share price",
            "Dividends",
            "Dividends per share",
            "Ex-dividend date",
            "Date of the next payment",
            "Div. received",
            "Category",
            "Country",
            "Sector",
            "ISIN",
            "Currency",
        ],
        [
            "SCHD",
            "Schwab U.S. Dividend Equity ETF",
            "50",
            "3910.00",
            "4125.00",
            "82.50",
            "124.00",
            "2.48",
            "Thu Mar 27 2026 00:00:00 GMT-0700 (Mountain Standard Time)",
            "Mon Mar 31 2026 00:00:00 GMT-0700 (Mountain Standard Time)",
            "620.00",
            "Growth",
            "United States of America",
            "Funds",
            "US8085247976",
            "USD",
        ],
        [
            "O",
            "Realty Income Corp",
            "25",
            "1300.00",
            "1395.00",
            "55.80",
            "78.00",
            "3.12",
            "Fri Feb 28 2026 00:00:00 GMT-0700 (Mountain Standard Time)",
            "Fri Mar 14 2026 00:00:00 GMT-0700 (Mountain Standard Time)",
            "156.00",
            "Anchors",
            "United States of America",
            "Real Estate",
            "US7561091049",
            "USD",
        ],
    ]
    return _write_csv_template(SNOWBALL_HOLDINGS_TEMPLATE_PATH, rows)


def create_etrade_buys_sells_template():
    """Create an XLSX template matching the E*TRADE buys/sells transaction parser."""
    sample_rows = [
        ["04/13/2026", "03/07/2026", "03/06/2026", "Bought", "SCHD", "Schwab U.S. Dividend Equity ETF", 25, 82.15, -2053.75, 0.00],
        ["04/13/2026", "03/28/2026", "03/27/2026", "Sold", "DIVO", "Amplify CWP Enhanced Dividend Income ETF", 10, 41.25, 412.50, 0.00],
    ]
    guidance_lines = [
        "Use the Buys & Sells export from E*TRADE Transaction History.",
        "The importer reads Bought rows as BUY and Sold rows as SELL.",
        "Keep Activity Type in column D and leave the header row on row 7.",
    ]
    return _create_etrade_transaction_template(
        ETRADE_BUYS_SELLS_TEMPLATE_PATH,
        "Buys & Sells Activity Types",
        "Total for all securities: <$1,641.25>",
        sample_rows,
        guidance_lines,
    )


def create_etrade_dividends_template():
    """Create an XLSX template matching the E*TRADE dividends transaction parser."""
    sample_rows = [
        ["04/13/2026", "04/04/2026", "04/03/2026", "Cash Dividend", "SCHD", "Ordinary Dividend", "", "", 31.00, 0.00],
        ["04/13/2026", "04/04/2026", "04/03/2026", "Dividend Reinvestment", "JEPI", "Dividend Reinvestment", 0.9621, 57.20, -55.05, 0.00],
    ]
    guidance_lines = [
        "Use the Dividends export from E*TRADE Transaction History.",
        "Positive Amount rows import as cash DIVIDEND entries.",
        "Negative Amount rows with Quantity and Price import as [DRIP] BUY entries.",
    ]
    return _create_etrade_transaction_template(
        ETRADE_DIVIDENDS_TEMPLATE_PATH,
        "Dividends Activity Types",
        "Total for all securities: <$86.05>",
        sample_rows,
        guidance_lines,
    )


if __name__ == "__main__":
    path = create_template()
    print(f"Template created at: {path}")
    etrade_path = create_etrade_template()
    print(f"E*TRADE template created at: {etrade_path}")
    schwab_path = create_schwab_template()
    print(f"Schwab template created at: {schwab_path}")
    schwab_txn_path = create_schwab_transactions_template()
    print(f"Schwab transactions template created at: {schwab_txn_path}")
    snowball_holdings_path = create_snowball_holdings_template()
    print(f"Snowball holdings template created at: {snowball_holdings_path}")
    etrade_bs_path = create_etrade_buys_sells_template()
    print(f"E*TRADE buys/sells template created at: {etrade_bs_path}")
    etrade_div_path = create_etrade_dividends_template()
    print(f"E*TRADE dividends template created at: {etrade_div_path}")
    fidelity_path = create_fidelity_template()
    print(f"Fidelity positions template created at: {fidelity_path}")
    fidelity_txn_path = create_fidelity_transactions_template()
    print(f"Fidelity transactions template created at: {fidelity_txn_path}")
